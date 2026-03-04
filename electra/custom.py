from __future__ import unicode_literals
import frappe,erpnext
from frappe.utils import cint
from frappe.utils import flt, fmt_money
import json
from frappe.utils import date_diff, add_months,today,add_days,add_years,nowdate,flt
from frappe.utils import cstr, cint, getdate,get_first_day, get_last_day, today, time_diff_in_hours

from frappe.model.mapper import get_mapped_doc
from frappe.utils.file_manager import get_file
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
import datetime
from datetime import date,datetime,timedelta
from electra.utils import get_dn_return_series
import openpyxl
from collections import defaultdict
from openpyxl import Workbook
import openpyxl
import xlrd
import re
from frappe.utils import date_diff,today,cstr
from datetime import datetime
from dateutil import relativedelta
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
import pandas as pd
from frappe.utils import formatdate


from erpnext.setup.utils import get_exchange_rate
from frappe import throw,_, db, get_doc, throw, whitelist


from frappe.utils import (
	add_days,
	cint,
	cstr,
	date_diff,
	flt,
	formatdate,
	get_first_day,
	get_link_to_form,
	getdate,
	money_in_words,
	rounded,
)


#Show the  Leave details for Leave Application 
@frappe.whitelist()
def get_leave_details(name,from_date,to_date,leave,lea_ty,balance):
	doc = frappe.get_doc('Leave Application',name)
	leave_type = doc.leave_type
	leave_days = frappe.db.get_value('Leave Type',{'name':doc.leave_type},'max_leaves_allowed')
	current_year = datetime.now().year
	
	
	entries = frappe.get_all(
		"Leave Ledger Entry",
		filters={
			"employee": doc.employee,
			"leave_type": doc.leave_type,
			"transaction_type": "Leave Application",
			"from_date": [">=", f"{current_year}-01-01"],
			"from_date": ["<=", f"{current_year}-12-31"],
			"from_date":["between",[f"{current_year}-01-01",add_days(from_date,-1)]]
		},
		fields=["leaves"]
	)
	expired = sum(entry.get("leaves", 0) for entry in entries)
	
	if doc.leave_type=='Leave Without Pay':
		
		total_leaves=0
	else:
		total_leaves = leave_days + expired 
	ticket=  frappe.db.get_value('Leave Type',{'name':doc.leave_type},'leave_ticket')
	if ticket == True:
		ticket="Yes"
	else:
		ticket="No"

	salary= frappe.db.get_value('Leave Type',{'name':doc.leave_type},'leave_salary')
	if salary == True:
		salary="Yes"
	else:
		salary="No"

	leaving_date = from_date
	rdoj = add_days(to_date ,1)
	leave_taken = leave
	if doc.leave_type=='Leave Without Pay':
		leave_status=0
	else:
		leave_status =  flt(balance)-flt(leave_taken)
	data = """<tr>
		<td>Leave Type</td>
		<td>Leave Days</td>
		<td>Leave Ticket</td>
		<td>Leave Salary</td>
		<td>Leaving Date</td>
		<td>Re-Joining Date</td>
		<td>Leave Taken</td>
		<td>Leave Status</td></tr>
		<tr>
		<td>%s</td>
		<td>%s</td>
		<td>%s</td>
		<td>%s</td>
		<td>%s</td>
		<td>%s</td>
		<td>%s</td>
		<td>%s</td>
		</tr>"""%(leave_type,balance,ticket,salary,leaving_date,rdoj,leave_taken,leave_status)
	return data

#Enable the hod checkbox in Leave application , Leave salary and final exit request,Rejoining form , additional vacation,HR Request form based on below condition
@frappe.whitelist()
def get_role(employee):
	user_id = frappe.get_value('Employee',{'employee':employee},['user_id'])
	hod = frappe.get_value('User',{'email':user_id},['name'])
	role = "HOD"
	hod = frappe.get_value('Has Role',{'role':role,'parent':hod})
	if hod:
		return  "HI"
	else:
		return "HII"


#Check the role of Need MD Approval with Passing the employee
@frappe.whitelist()
def get_md_role(employee):
	user_id = frappe.get_value('Employee',{'employee':employee},['user_id'])
	md = frappe.get_value('User',{'email':user_id},['name'])
	role = "Need MD Approval"
	md = frappe.get_value('Has Role',{'role':role,'parent':md})
	if md:
		return  "HI"
	else:
		return "HII"

#Set the count of Visa used and balance in Visa approval monitor after insert of employee
def visa_creation(doc,method):
	if doc.visa_application_num:
		vam = frappe.get_doc('Visa Approval Monitor',doc.visa_application_num)
		used_visa = cint(vam.used_visa)
		balance = cint(vam.balance)
		used_visa += 1
		if vam.balance:
			balance -= 1
		vam.used_visa = used_visa
		vam.balance = balance
		vam.save(ignore_permissions=True)
		frappe.db.commit()

# Get the balance qty from Bin and set it to IC Material Transfer Request and Stock Request
@frappe.whitelist()
def get_stock_balance_from_wh(item_code,warehouse):
	actual_qty = 0
	try:
		actual_qty = frappe.db.sql("""select (actual_qty - reserved_stock) as qty from tabBin
			where item_code = '%s' and warehouse = '%s' """%(item_code,warehouse),as_dict=True)
	except:
		pass
	return actual_qty

#Get the below value from bin to match with item code and set it to IC Material Transfer Request and Stock Request
@frappe.whitelist()
def get_stock_balance(item_table):
	item_table = json.loads(item_table)
	data = []
	for item in item_table:
		try:
			item_name = frappe.get_value('Item',{'name':item['item_code']},"item_name")
			stocks = frappe.db.sql("""select actual_qty,warehouse,stock_uom,stock_value from tabBin
				where item_code = '%s' """%(item["item_code"]),as_dict=True)
			for stock in stocks:
				data.append([item['item_code'],item_name,stock.warehouse,stock.actual_qty,stock.stock_uom,stock.stock_value])
		except:
			pass
	return data

#Get the previous purchase values and set it to Quotation , Sales Order and Opprtunity
@frappe.whitelist()
def get_previous_po(item_table):
	item_table = json.loads(item_table)
	data = []
	for item in item_table:
		if not item["item_code"] == "001":
			try:
				item_name = frappe.get_value('Item',{'name':item['item_code']},"item_name")
				pos = frappe.db.sql("""select `tabPurchase Order Item`.item_code as item_code,`tabPurchase Order Item`.item_name as item_name,`tabPurchase Order`.supplier as supplier,`tabPurchase Order Item`.qty as qty,`tabPurchase Order Item`.amount as amount,`tabPurchase Order`.transaction_date as date,`tabPurchase Order`.name as po from `tabPurchase Order`
				left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
				where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus != 2 """%(item["item_code"]),as_dict=True)

				for po in pos:
					data.append([item['item_code'],item_name,po.supplier,po.qty,po.date,po.amount,po.po])
			except:
				pass
	return data


# Get the Previous qty from Purchase Order set it to New Purchase Order
@frappe.whitelist()
def set_previous_po(item_table):
	item_table = json.loads(item_table)
	data = []
	for item in item_table:
		try:
			item_name = frappe.get_value('Item',{'name':item['item_code']},"item_name")
			pos = frappe.db.sql("""select `tabPurchase Order Item`.item_code as item_code,`tabPurchase Order Item`.item_name as item_name,sum(`tabPurchase Order Item`.qty) as qty from `tabPurchase Order`
			left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
			where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus != 2 """%(item["item_code"]),as_dict=True)
			for po in pos:
				data.append([item['item_code'],item_name,po.qty])
		except:
			pass
	return data


#Get the value from Previous document of the Purchase Order and Show the table while enter the Item code of Purchase Order Item
@frappe.whitelist()
def previous_po_html(item_code):
	data = ""
	item_name = frappe.get_value('Item',{'item_code':item_code},"item_name")
	pos = frappe.db.sql("""select `tabPurchase Order Item`.item_code as item_code,`tabPurchase Order Item`.item_name as item_name,`tabPurchase Order`.supplier as supplier,`tabPurchase Order Item`.qty as qty,`tabPurchase Order Item`.rate as rate,`tabPurchase Order Item`.amount as amount,`tabPurchase Order`.transaction_date as date,`tabPurchase Order`.name as po from `tabPurchase Order`
	left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
	where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus != 2 order by date"""%(item_code),as_dict=True)


	data += '<table class="table table-bordered"><tr><th style="padding:1px;border: 1px solid black;color:white;background-color:orange" colspan=6><center>Previous Purchase Order</center></th></tr>'
	data += '''
	<tr><td colspan =2 style="padding:1px;border: 1px solid black;width:300px" ><b>Item Code</b></td>
	<td style="padding:1px;border: 1px solid black;width:200px" colspan =4>%s</td></tr>
	<tr><td colspan =2 style="padding:1px;border: 1px solid black" ><b>Item Name</b></td>
	<td style="padding:1px;border: 1px solid black" colspan =4>%s</td></tr>

	<tr><td style="padding:1px;border: 1px solid black" colspan =1><b>Supplier Name</b></td>
	<td style="padding:1px;border: 1px solid black" colspan=1><b>Previous Purchase Order</b></td>
	<td style="padding:1px;border: 1px solid black" colspan=1><b>PO Date</b></td>
	<td style="padding:1px;border: 1px solid black" colspan=1><b>PO Rate</b></td>
	<td style="padding:1px;border: 1px solid black" colspan=1><b>PO Quantity</b></td>
	<td style="padding:1px;border: 1px solid black" colspan=1><b>PO Amount</b>
	</td></tr>'''%(item_code,item_name)
	for po in pos:
		data += '''<tr>
			<td style="padding:1px;border: 1px solid black" colspan =1>%s</td>
			<td style="padding:1px;border: 1px solid black" colspan=1>%s</td>
			<td style="padding:1px;border: 1px solid black" colspan=1>%s</td>
			<td style="padding:1px;border: 1px solid black" colspan=1>%s</td>
			<td style="padding:1px;border: 1px solid black" colspan=1>%s</td>
			<td style="padding:1px;border: 1px solid black" colspan=1>%s</td></tr>'''%(po.supplier,po.po,po.date,po.rate,po.qty,po.amount)

	data += '</table>'
	return data

# @frappe.whitelist()
# def get_sales_invoice(item_table):
#     item_table = json.loads(item_table)
#     data = []
#     for item in item_table:
#         try:
#             item_name = frappe.get_value('Item',{'name':item['item_code']},"item_name")
#             pos = frappe.db.sql("""select `tabPurchase Order Item`.item_code as item_code,`tabPurchase Order Item`.item_name as item_name,`tabPurchase Order`.supplier as supplier,`tabPurchase Order Item`.qty as qty,`tabPurchase Order Item`.amount as amount,`tabPurchase Order`.transaction_date as date,`tabPurchase Order`.name as po from `tabPurchase Order`
#             left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
#             where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus != 2 """%(item["item_code"]),as_dict=True)
#             for po in pos:
#                 data.append([item['item_code'],item_name,po.supplier,po.qty,po.date,po.amount,po.po])
#         except:
#             pass
#     return data


#Get the below values from Stock Ledger Entry and set the values to Stock out Qty table for Sales Order
@frappe.whitelist()
def get_out_qty(item_table):
	item_table = json.loads(item_table)
	data = []
	for item in item_table:
		try:
			sles = frappe.db.sql("""select * from `tabStock Ledger Entry`
			left join `tabStock Entry` on `tabStock Ledger Entry`.voucher_no = `tabStock Entry`.name where `tabStock Ledger Entry`.posting_date between '%s' and '%s' and `tabStock Ledger Entry`.item_code = '%s' and `tabStock Ledger Entry`.actual_qty < 0 and `tabStock Ledger Entry`.voucher_type = 'Stock Entry' and `tabStock Entry`.stock_entry_type = 'Material Issue' """%(add_months(today(),-6),today(),item['item_code']),as_dict=True)
			for sl in sles:
				data.append([item['item_code'],sl.warehouse,abs(sl.actual_qty),sl.posting_date,sl.voucher_type])
		except:
			pass
	return data


#Get the below values from Sales Order and set it to Pending sales order in Opportunity
@frappe.whitelist()
def get_under_so(item_table):
	item_table = json.loads(item_table)
	data = []
	for item in item_table:
		try:
			item_name = frappe.get_value('Item',{'name':item['item_code']},"item_name")
			sos = frappe.db.sql("""select `tabSales Order Item`.item_code as item_code,`tabSales Order Item`.item_name as item_name, `tabSales Order Item`.qty as qty,`tabSales Order`.transaction_date as date from `tabSales Order`
			left join `tabSales Order Item` on `tabSales Order`.name = `tabSales Order Item`.parent
			where `tabSales Order Item`.item_code = '%s' and `tabSales Order`.docstatus != 2 """%(item["item_code"]),as_dict=True)
			for so in sos:
				data.append([item['item_code'],item_name,so.qty,so.date])
		except:
			pass
	return data

# @frappe.whitelist()
# def get_out_qty(item_table):
#     item_table = json.loads(item_table)
#     data = []
#     for item in item_table:
#         sles = frappe.db.sql("""select * from `tabStock Ledger Entry`
#         left join `tabStock Entry` on `tabStock Ledger Entry`.voucher_no = `tabStock Entry`.name where `tabStock Ledger Entry`.posting_date between '%s' and '%s' and `tabStock Ledger Entry`.item_code = '%s' and `tabStock Ledger Entry`.actual_qty < 0 and `tabStock Ledger Entry`.voucher_type = 'Stock Entry' and `tabStock Entry`.stock_entry_type = 'Material Issue' """%(add_months(today(),-6),today(),item['item_code']),as_dict=True)
#         for sl in sles:
#             data.append([item['item_code'],sl.warehouse,abs(sl.actual_qty),sl.posting_date,sl.voucher_type])
#     return data


#Create Journal Entry from On Submission of Landed Cost Voucher
@frappe.whitelist()
def create_lcv_je(doc,method):
	from electra.utils import get_series
	tnc = doc.taxes
	for tn in tnc:
		if tn.supplier:
			jv = frappe.new_doc("Journal Entry")
			jv.naming_series = get_series(doc.company,"Journal Entry")
			jv.voucher_type = "Journal Entry"
			jv.company = doc.company
			jv.posting_date = doc.posting_date
			jv.bill_no = tn.bill_no
			jv.bill_date = tn.invoice_date
			jv.custom_description = tn.description
			# jv.custom_purchase_invoice = doc.custom_purchase_invoice
			jv.append("accounts", {
				"account": tn.expense_account,
				"debit": tn.base_amount,
				"cost_center": erpnext.get_default_cost_center(doc.company),
				"debit_in_account_currency": tn.amount
			})

			jv.append("accounts", {
				"account": frappe.get_cached_value('Company', doc.company, 'default_payable_account'),
				"party_type": "Supplier",
				"party":tn.supplier,
				"cost_center": erpnext.get_default_cost_center(doc.company),
				"credit": tn.base_amount,
				"credit_in_account_currency": tn.amount
			})
			jv.landed_cost_voucher = doc.name
			jv.insert()
			jv.submit()

#Cancel Journal Entry from On cancel of Landed Cost Voucher
@frappe.whitelist()
def cancel_lcv_je(doc,method):
	jes = frappe.get_all("Journal Entry",{'landed_cost_voucher':doc.name})
	for je in jes:
		jv = frappe.get_doc("Journal Entry",je.name)
		if jv.docstatus == 1:
			jv.cancel()
			frappe.db.commit()


#Create the Cost Estimation from Opportunity while Click the Cost Estimation Button
@frappe.whitelist()
def make_cost_estimation(source_name, target_doc=None):
	doc = frappe.get_doc("Opportunity",source_name)
	if doc.party_name and doc.opportunity_from == 'Customer':
		customer_name = frappe.db.get_value("Customer", doc.party_name, "customer_name")
	elif doc.party_name and doc.opportunity_from == 'Lead':
		lead_name = frappe.db.get_value("Lead", doc.party_name, "lead_name")
		customer_name = lead_name
	doclist = get_mapped_doc("Opportunity", source_name, {
		"Opportunity": {
			"doctype": "Cost Estimation",
			"field_map": {
				"name": "opportunity",
				"opportunity_from":"cost_estimation_for",
				"customer_name" : "lead_customer_name",
				"party_name":"lead_customer",
				"title_of_project":"project_name"
			}
		},
		"Opportunity Item": {
			"doctype": "CE MATERIALS",
			"field_map": {
				"item_code": "item",
				"uom": "unit",
				"qty": "qty"
			}
		},
		"Opportunity Scope of Work": {
			"doctype": "CE Master Scope of Work",
			"field_map": {
				"msow": "msow",
				"msow_desc": "msow_desc"
			}
		}
	}, target_doc)
	return doclist

#Create the Quotation from Cost Estimation while Click the Quotation Button
@frappe.whitelist()
def make_quotation(source_name, target_doc=None):
	doc = frappe.get_doc("Cost Estimation",source_name)
	doclist = get_mapped_doc("Cost Estimation", source_name, {
		"Cost Estimation": {
			"doctype": "Quotation",
			"field_map": {
				"cost_estimation_for": "quotation_to",
				"name": "cost_estimation",
				"project_name":"title_of_project",
				"lead_customer": "party_name",
				"prepared_by":"user_id"
			}
		},
		# "CE MATERIALS": {
		#     "doctype": "Quotation Item",
		#     "field_map": {
		#         "item": "item_code",
		#         "unit": "uom",
		#         "qty": "qty"
		#     }
		# },
		"Cost Estimation Scope Of Work": {
			"doctype": "Quotation Scope of Work",
			"field_map": {
				"msow": "msow",
				"ssow": "ssow",
				"msow_desc": "msow_desc",
				"ssow_desc" : "ssow_desc",
				"qty": "qty",
				"unit":"unit",
				"unit_price":"unit_price"
			}
		}
	}, target_doc)
	return doclist

#Create Project Budget
@frappe.whitelist()
def make_project_so(source_name, target_doc=None):
	doc = frappe.get_doc("Project Budget",source_name)
	order_type = "Project"
	doclist = get_mapped_doc("Project Budget", source_name, {
		"Project Budget": {
			"doctype": "Sales Order",
			"field_map": {
				"name": "project_budget",
				"customer":"customer",
				"order_type":"Project",
				"user_id":"prepared_by"
			}
		},
		# "Quotation Scope of Work": {
		#     "doctype": "SO Scope of Work",
		#     "field_map": {
		#         "msow": "msow",
		#         "ssow": "ssow",
		#         "msow_desc": "msow_desc",
		#         "ssow_desc" : "ssow_desc"
		#     }
		# }
	}, target_doc)
	return doclist

#Create Sales Order
@frappe.whitelist()
def make_so(source_name, target_doc=None):
	doc = frappe.get_doc("Quotation",source_name)
	doclist = get_mapped_doc("Quotation", source_name, {
		"Quotation": {
			"doctype": "Sales Order",
			"field_map": {
				"name": "quotation",
				"customer":"customer",
				"user_id":"prepared_by"
			}
		},
		"Quotation Scope of Work": {
			"doctype": "SO Scope of Work",
			"field_map": {
				"msow": "msow",
				"ssow": "ssow",
				"msow_desc": "msow_desc",
				"ssow_desc" : "ssow_desc"
			}
		}
		# "CE MATERIALS": {
		#     "doctype": "Quotation Item",
		#     "field_map": {
		#         "item": "item_code",
		#         "unit": "uom",
		#         "qty": "qty"
		#     }
		# }
	}, target_doc)
	return doclist

#Create Project Budget
@frappe.whitelist()
def make_project_budget(source_name, target_doc=None):
	doc = frappe.get_doc("Sales Order",source_name)
	cost_estimation = doc.cost_estimation
	doclist = get_mapped_doc("Sales Order", source_name, {
		"Sales Order": {
			"doctype": "Project Budget",
			"field_map": {
				"title_of_project": "title_of_project",
				"name": "sales_order",
			}
		}
	}, target_doc)
	return doclist

#Create Project
@frappe.whitelist()
def make_project(source_name, target_doc=None):
	doc = frappe.get_doc("Budgeting",source_name)
	doclist = get_mapped_doc("Budgeting", source_name, {
		"Budgeting": {
			"doctype": "Project",
			"field_map": {
				"title_of_project": "project_name",
				"sales_order": "sales_order",
				"name" : "budgeting"
							}
		}
	}, target_doc)
	return doclist

# @frappe.whitelist()
# def stock_popup(item_code):
#     item = frappe.get_value('Item',{'item_code':item_code},'item_code')
#     # item_price= frappe.get_value('Item Price',{item:'item_code'},'price_list_rate')
#     data = ''
#     stocks = frappe.db.sql("""select actual_qty,warehouse,stock_uom,stock_value from tabBin
#         where item_code = '%s' """%(item),as_dict=True)
#     # pos = frappe.db.sql("""select `tabPurchase Order Item`.item_code as item_code,`tabPurchase Order Item`.item_name as item_name,`tabPurchase Order`.supplier as supplier,`tabPurchase Order Item`.qty as qty,`tabPurchase Order Item`.amount as amount,`tabPurchase Order`.transaction_date as date,`tabPurchase Order`.company as company,`tabPurchase Order`.name as po from `tabPurchase Order`
#     # left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
#     # where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus != 1   order by amount desc limit 1"""%(item),as_dict=True)



#     pos = frappe.db.sql("""select `tabPurchase Order Item`.item_code as item_code,`tabPurchase Order Item`.item_name as item_name,`tabPurchase Order`.supplier as supplier,`tabPurchase Order Item`.qty as qty,`tabPurchase Order Item`.rate as rate,`tabPurchase Order`.transaction_date as date,`tabPurchase Order`.name as po from `tabPurchase Order`
#     left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
#     where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus != 2 order by rate asc limit 1""" % (item), as_dict=True)
#     data += '<table class="table table-bordered"><tr><th style="padding:1px;border: 1px solid black;color:white;background-color:orange" colspan=4><center>Stock Availability</center></th></tr>'
#     data += '''
#     <tr><td style="padding:1px;border: 1px solid black;width:300px" ><b>Item Code</b></td>
#     <td style="padding:1px;border: 1px solid black;width:200px" colspan =3>%s</td></tr>
#     <tr><td style="padding:1px;border: 1px solid black" ><b>Item Name</b></td>
#     <td style="padding:1px;border: 1px solid black" colspan =3>%s</td></tr>
#     <tr>
#     <td style="padding:1px;border: 1px solid black;"  ><b>Warehouse</b></td>
#     <td style="padding:1px;border: 1px solid black;max-width:50px"  ><b>QTY</b></td>
#     <td style="padding:1px;border: 1px solid black;max-width:50px;" colspan ='1' ><b>Previous Purchase Order</b></td>
#     <td style="padding:1px;border: 1px solid black;max-width:50px;" colspan ='1' ><b>PPQ</b></td></tr>'''%(item,frappe.db.get_value('Item',item,'item_name'))


#     i = 0
#     if pos:
#         for po in pos:
#             for stock in stocks:
#                 if stock.actual_qty > 0:
#                     data += '''<tr>
#                     <td style="padding:1px;border: 1px solid black" colspan =1>%s</td><td style="padding:1px;border: 1px solid black" colspan=1>%s</td>
#                     <td style="padding:1px;border: 1px solid black" colspan=1>-</td><td style="padding:1px;border: 1px solid black" colspan=1>-</td></tr>'''%(stock.warehouse,stock.actual_qty)
#     else:
#         for stock in stocks:
#             if stock.actual_qty > 0:
#                 data += '''<tr>
#                 <td style="padding:1px;border: 1px solid black" colspan =1>%s</td><td style="padding:1px;border: 1px solid black" colspan=1>%s</td>
#                 <td style="padding:1px;border: 1px solid black" colspan=1>-</td><td style="padding:1px;border: 1px solid black" colspan=1>-</td></tr>'''%(stock.warehouse,stock.actual_qty)
#     i += 1

#     data += '</table>'

#     return data

#Show the Stock availability with Warehouse while enter and Click the item in all item table
@frappe.whitelist()
def stock_popup(item_code,company):
	w_house = frappe.db.get_value("Warehouse",{'company':company,'default_for_stock_transfer':1},['name'])
	item = frappe.get_value('Item',{'item_code':item_code},'item_code')
	data = ''
	stocks = frappe.db.sql("""select actual_qty,warehouse,stock_uom,stock_value from tabBin
		where item_code = '%s' order by warehouse """%(item),as_dict=True)
	pos = frappe.db.sql("""select `tabPurchase Order Item`.item_code as item_code,`tabPurchase Order Item`.item_name as item_name,`tabPurchase Order`.supplier as supplier,`tabPurchase Order Item`.qty as qty,`tabPurchase Order Item`.rate as rate,`tabPurchase Order`.transaction_date as date,`tabPurchase Order`.name as po from `tabPurchase Order`
	left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
	where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus != 2 """ % (item_code), as_dict=True)
	data += '<table class="table table-bordered"><tr><th style="padding:1px;border: 1px solid black;color:white;background-color:orange" colspan=6><center>Stock Availability</center></th></tr>'
	data += '''
	<tr><td style="padding:1px;border: 1px solid black;width:300px" ><b>Item Code</b></td>
	<td style="padding:1px;border: 1px solid black;width:200px" colspan =5>%s</td></tr>
	<tr><td style="padding:1px;border: 1px solid black" ><b>Item Name</b></td>
	<td style="padding:1px;border: 1px solid black" colspan =5>%s</td></tr>'''%(item,frappe.db.get_value('Item',item,'item_name'))
	data += '''
	<td style="padding:1px;border: 1px solid black;background-color:orange;color:white;"  ><b>Company</b></td>
	<td style="padding:1px;border: 1px solid black;background-color:orange;color:white;"  ><b>Warehouse</b></td>
	<td style="padding:1px;border: 1px solid black;background-color:orange;color:white;width:13%"  ><b>Stock Qty</b></td>
	<td style="padding:1px;border: 1px solid black;background-color:orange;color:white;width:13%"  ><b>Res Qty</b></td>
	<td style="padding:1px;border: 1px solid black;background-color:orange;color:white;width:13%;" colspan ='1' ><b>PENDING TO RECEIVE</b></td>
	<td style="padding:1px;border: 1px solid black;background-color:orange;color:white;width:13%;" colspan ='1' ><b>PENDING TO SELL</b></td></tr>'''
	for stock in stocks:
		if stock.warehouse == w_house:
			comp = frappe.get_value("Warehouse",stock.warehouse,['company'])
			entries = frappe.db.get_all("Stock Reservation Entry",{"item_code": item, "warehouse":stock.warehouse,"status":("Not In",("Delivered")),"docstatus":('!=',2)},["reserved_qty"])
			total_reserved_qty = sum(flt(entry["reserved_qty"]) for entry in entries)
			value = stock.actual_qty - total_reserved_qty		
			new_po = frappe.db.sql("""select sum(`tabPurchase Order Item`.qty) as qty,sum(`tabPurchase Order Item`.received_qty) as d_qty from `tabPurchase Order`
			left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
			where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus = 1 and `tabPurchase Order`.company = '%s' and `tabPurchase Order`.status != 'Closed' """ % (item_code,comp), as_dict=True)[0]
			if not new_po['qty']:
				new_po['qty'] = 0
			if not new_po['d_qty']:
				new_po['d_qty'] = 0
			ppoc_total = new_po['qty'] - new_po['d_qty']	

		
			new_so = frappe.db.sql("""select sum(`tabSales Order Item`.qty) as qty,sum(`tabSales Order Item`.delivered_qty) as d_qty from `tabSales Order`
			left join `tabSales Order Item` on `tabSales Order`.name = `tabSales Order Item`.parent
			where `tabSales Order Item`.item_code = '%s' and `tabSales Order`.docstatus = 1 and `tabSales Order`.company = '%s' and `tabSales Order`.status != "Closed" """ % (item_code,comp), as_dict=True)[0]
			if not new_so['qty']:
				new_so['qty'] = 0
			if not new_so['d_qty']:
				new_so['d_qty'] = 0
			del_total = new_so['qty'] - new_so['d_qty']
			data +=''' <tr><td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan =1>%s</td><td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>%s</td><td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>%s</td>
				<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>%s</td><td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>%s</td><td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>%s</td></tr>'''%(comp,stock.warehouse,value,total_reserved_qty,ppoc_total,del_total)


	
	i = 0
	for stock in stocks:
		if stock.warehouse != w_house:
			# if stock.actual_qty > 0:
			comp = frappe.get_value("Warehouse",stock.warehouse,['company'])
			entries = frappe.db.get_all("Stock Reservation Entry",{"item_code": item, "warehouse":stock.warehouse,"status":("Not In",("Delivered")),"docstatus":('!=',2)},["reserved_qty"])
			total_reserved_qty = sum(flt(entry["reserved_qty"]) for entry in entries)
			value = stock.actual_qty - total_reserved_qty		
			new_po = frappe.db.sql("""select sum(`tabPurchase Order Item`.qty) as qty,sum(`tabPurchase Order Item`.received_qty) as d_qty from `tabPurchase Order`
			left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
			where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order Item`.warehouse = '%s' and `tabPurchase Order`.docstatus = 1 and `tabPurchase Order`.company = '%s' and `tabPurchase Order`.status != 'Closed'""" % (item_code,stock.warehouse,comp), as_dict=True)[0]
			if not new_po['qty']:
				new_po['qty'] = 0
			if not new_po['d_qty']:
				new_po['d_qty'] = 0
			ppoc_total = new_po['qty'] - new_po['d_qty']
		
			
			new_so = frappe.db.sql("""select sum(`tabSales Order Item`.qty) as qty,sum(`tabSales Order Item`.delivered_qty) as d_qty from `tabSales Order`
			left join `tabSales Order Item` on `tabSales Order`.name = `tabSales Order Item`.parent
			where `tabSales Order Item`.item_code = '%s' and `tabSales Order Item`.warehouse = '%s' and `tabSales Order`.docstatus = 1 and `tabSales Order`.company = '%s' and `tabSales Order`.status != "Closed" """ % (item_code,stock.warehouse,comp), as_dict=True)[0]
			if not new_so['qty']:
				new_so['qty'] = 0
			if not new_so['d_qty']:
				new_so['d_qty'] = 0
			del_total = new_so['qty'] - new_so['d_qty']

			data += '''<tr>
			<td style="padding:1px;border: 1px solid black" colspan =1>%s</td><td style="padding:1px;border: 1px solid black" colspan=1>%s</td><td style="padding:1px;border: 1px solid black" colspan=1>%s</td>
			<td style="padding:1px;border: 1px solid black" colspan=1>%s</td><td style="padding:1px;border: 1px solid black" colspan=1>%s</td><td style="padding:1px;border: 1px solid black" colspan=1>%s</td></tr>'''%(comp,stock.warehouse,value,total_reserved_qty,ppoc_total,del_total)
			
	
	i += 1
	p_po_total = 0
	p_so_total = 0
	stock_qty = 0
	res_qty = 0
	for stock in stocks:

		comp = frappe.get_value("Warehouse",stock.warehouse,['company'])
		entries = frappe.db.get_all("Stock Reservation Entry",{"item_code": item, "warehouse":stock.warehouse,"status":("Not In",("Delivered")),"docstatus":('!=',2)},["reserved_qty"])
		total_reserved_qty = sum(flt(entry["reserved_qty"]) for entry in entries)
		value = stock.actual_qty - total_reserved_qty		
		new_po = frappe.db.sql("""select sum(`tabPurchase Order Item`.qty) as qty,sum(`tabPurchase Order Item`.received_qty) as d_qty from `tabPurchase Order`
		left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
		where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order Item`.warehouse = '%s'  and `tabPurchase Order`.docstatus = 1 and `tabPurchase Order`.status != 'Closed' and `tabPurchase Order`.company = '%s' """ % (item_code,stock.warehouse,comp), as_dict=True)[0]
		if not new_po['qty']:
			new_po['qty'] = 0
		if not new_po['d_qty']:
			new_po['d_qty'] = 0
		ppoc_total = new_po['qty'] - new_po['d_qty']
	
		
		new_so = frappe.db.sql("""select sum(`tabSales Order Item`.qty) as qty,sum(`tabSales Order Item`.delivered_qty) as d_qty from `tabSales Order`
		left join `tabSales Order Item` on `tabSales Order`.name = `tabSales Order Item`.parent
		where `tabSales Order Item`.item_code = '%s' and `tabSales Order Item`.warehouse = '%s' and `tabSales Order`.docstatus = 1 and `tabSales Order`.company = '%s' and `tabSales Order`.status != "Closed" """ % (item_code,stock.warehouse,comp), as_dict=True)[0]
		if not new_so['qty']:
			new_so['qty'] = 0
		if not new_so['d_qty']:
			new_so['d_qty'] = 0
		del_total = new_so['qty'] - new_so['d_qty']


		p_po_total += ppoc_total
		p_so_total += del_total
		stock_qty += value
		res_qty += total_reserved_qty
	data += '''<tr>
	<td style="background-color:#ffe9ad;padding:1px;border: 1px solid black;text-align:right;font-weight:bold" colspan =2>%s</td><td style="background-color:#ffe9ad;padding:1px;border: 1px solid black;font-weight:bold" colspan=1>%s</td>
	<td style="background-color:#ffe9ad;padding:1px;border: 1px solid black;font-weight:bold" colspan=1>%s</td><td style="background-color:#ffe9ad;padding:1px;border: 1px solid black;font-weight:bold" colspan=1>%s</td><td style="background-color:#ffe9ad;padding:1px;border: 1px solid black;font-weight:bold" colspan=1>%s</td></tr>'''%("Total     ",stock_qty,res_qty,p_po_total,p_so_total)
			
	data += '</table>'

	return data




# @frappe.whitelist()
# def get_stock_price():
#     data =''
#     price = frappe.db.sql("""select `tabQuotation Item`.item_code as item_code,`tabQuotation Item`.item_name as item_name,`tabItem Price`.item_code as item_code,`tabItem Price`.price_list_rate as price_list_rate
#     left join `tabQuotation Item` on `tabItem Price`.name =`tabQuotation Item`.parent
#     where `tabItem Price`.item_code = '%s' and `tabItem Price`
#     """)

@frappe.whitelist()
def po_popup(item_code):
	item = frappe.get_value('Item',{'item_name':item_code},'item_code')

	data = ''
	pos = frappe.db.sql("""select `tabPurchase Order Item`.item_code as item_code,`tabPurchase Order Item`.item_name as item_name,`tabPurchase Order`.supplier as supplier,`tabPurchase Order Item`.qty as qty,`tabPurchase Order Item`.amount as amount,`tabPurchase Order`.transaction_date as date,`tabPurchase Order`.name as po from `tabPurchase Order`
	left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
	where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus != 2 """%(item),as_dict=True)
	data += '<table class="table table-bordered"><tr><th style="padding:1px;border: 1px solid black" colspan=6><center>Previous Purchase Order</center></th></tr>'
	data += '<tr><td style="padding:1px;border: 1px solid black"><b>Item Code</b></td><td style="padding:1px;border: 1px solid black"><b>Item Name</b></td><td style="padding:1px;border: 1px solid black"><b>Supplier</b></td><td style="padding:1px;border: 1px solid black"><b>QTY</b></td><td style="padding:1px;border: 1px solid black"><b>PO Date</b></td><td style="padding:1px;border: 1px solid black"><b>Amount</b></td></tr>'
	for po in pos:
		data += '<tr><td style="padding:1px;border: 1px solid black">%s</td><td style="padding:1px;border: 1px solid black">%s</td><td style="padding:1px;border: 1px solid black">%s</td><td style="padding:1px;border: 1px solid black">%s</td><td style="padding:1px;border: 1px solid black">%s</td><td style="padding:1px;border: 1px solid black">%s</td></tr>'%(po.item_code,po.item_name,po.supplier,po.qty,po.date,po.amount)
	data += '</table>'
	if pos:
		return data



@frappe.whitelist()
def out_qty_popup(item):
	data = ''
	sles = frappe.db.sql("""select * from `tabStock Ledger Entry`
	left join `tabStock Entry` on `tabStock Ledger Entry`.voucher_no = `tabStock Entry`.name where `tabStock Ledger Entry`.posting_date between '%s' and '%s' and `tabStock Ledger Entry`.item_code = '%s' and `tabStock Ledger Entry`.actual_qty < 0 and `tabStock Ledger Entry`.voucher_type = 'Stock Entry' and `tabStock Entry`.stock_entry_type = 'Material Issue' """%(add_months(today(),-6),today(),item),as_dict=True)
	data += '<table class="table table-bordered"><tr><th style="padding:1px;border: 1px solid black" colspan=6><center>6 Months Stock Out Qty</center></th></tr>'
	data += '<tr><td style="padding:1px;border: 1px solid black"><b>Item Code</b></td><td style="padding:1px;border: 1px solid black"><b>Warehouse</b></td><td style="padding:1px;border: 1px solid black"><b>QTY</b></td><td style="padding:1px;border: 1px solid black"><b>Date</b></td><td style="padding:1px;border: 1px solid black"><b>Out Type</b></td></tr>'
	for sl in sles:
		data += '<tr><td style="padding:1px;border: 1px solid black">%s</td><td style="padding:1px;border: 1px solid black">%s</td><td style="padding:1px;border: 1px solid black">%s</td><td style="padding:1px;border: 1px solid black">%s</td><td style="padding:1px;border: 1px solid black">%s</td></tr>'%(sl.item_code,sl.warehouse,abs(sl.actual_qty),sl.posting_date,sl.voucher_type)
	data += '</table>'
	if sles:
		return data

#Create Task from Project if sales order value there in Project
@frappe.whitelist()
def create_tasks(doc,method):
	if doc.sales_order:
		so = frappe.get_doc("Sales Order",doc.sales_order)
		for so_installation in so.installation:
			task = frappe.new_doc("Task")
			task.update({
				"project": doc.name,
				"msow":so_installation.msow,
				"item_code":so_installation.item,
				"subject": so_installation.item_name,
				"description":so_installation.description,
				"uom":so_installation.unit,
				"qty": so_installation.qty,
				"pending_qty": so_installation.qty,
				"cost":so_installation.cost,
				"cost_amount":so_installation.cost_amount,
				"budgeted_cost":so_installation.unit_price,
				"budgeted_amount":so_installation.amount,
				"selling_price":so_installation.rate_with_overheads,
				"selling_amount":so_installation.amount_with_overheads,
				"is_group": 0
			})
			task.save(ignore_permissions=True)

	# for item in so.items:
	#     if item.work_title:
	#         parent_item = frappe.get_value("Task",{'subject':item.work_title},'name')
	#         task_name = item.item_name
	#         task = frappe.new_doc("Task")
	#         task.update({
	#             "subject": task_name,
	#             "project": doc.name,
	#             "parent_task": parent_item
	#         })
	#         task.save(ignore_permissions=True)

# @frappe.whitelist()
# def employee_conversion():
# 	employee=frappe.get_all('Employee',{'employment_type':'Probation'},['date_of_joining','grade','employee_name','employee_number'])
# 	day = datetime.strptime(str(today()),"%Y-%m-%d").date()
# 	for emp in employee:
# 		date = emp.date_of_joining
# 		worker = emp.grade
# 		staff = emp.grade
# 		if worker:
# 			six = add_months(date,5)
# 			if six== day:
# 				frappe.sendmail(
# 					recipients = ["jagadeesan@groupteampro.com"],
# 					subject = 'probation period of employee ',
# 					message = 'Dear Sir / Mam <br> Probation period end for %s (%s) in %s'%(emp.employee_name,emp.employee_number,add_months(date,6))
# 				)

# 			fifteen_days = add_days(six,15)
# 			if fifteen_days== day:
# 				frappe.sendmail(
# 					recipients = ["jagadeesan.a@groupteampro.com"],
# 					subject = 'probation period ends ',
# 					message = 'Dear Sir / Mam <br> Probation period end for %s (%s) in %s'%(emp.employee_name,emp.employee_number,add_months(date,6))
# 				)

# 			seven_days = add_days(fifteen_days,8)
# 			if seven_days== day:
# 				frappe.sendmail(
# 					recipients = ["jagadeesan.a@groupteampro.com"],
# 					subject = 'probation period ends ',
# 					message = 'Dear Sir / Mam <br> Probation period end for %s (%s) in %s'%(emp.employee_name,emp.employee_number,add_months(date,6))
# 				)
# 		if staff:
# 			three = add_months(date,2)
# 			if three == day:
# 				print(day)
# 				frappe.sendmail(
# 					recipients = ["jagadeesan.a@groupteampro.com"],
# 					subject = 'probation period ends ',
# 					message = 'Dear Sir / Mam <br> Probation period end for %s (%s) in %s'%(emp.employee_name,emp.employee_number,add_months(date,3))
# 				)

# 			fifteen_days_ = add_days(three,15)
# 			if fifteen_days_== day:
# 				frappe.sendmail(
# 					recipients = ["jagadeesan.a@groupteampro.com"],
# 					sender = sender['email'],
# 					subject = 'probation period of employee ',
# 					message = 'Dear Sir / Mam <br> Probation period end for %s (%s) in %s'%(emp.employee_name,emp.employee_number,add_months(date,3))
# 				)

# 			seven_days_ = add_days(fifteen_days_,8)
# 			if seven_days_== day:
# 				frappe.sendmail(
# 					recipients = ["jagadeesan.a@groupteampro.com"],
# 					sender = sender['email'],
# 					subject = 'probation period  ',
# 					message = 'Dear Sir / Mam <br> Probation period end for %s (%s) in %s'%(emp.employee_name,emp.employee_number,add_months(date,3))
# 				)


# @frappe.whitelist()
# def check_discount_percent(user,discount):
#     user_roles = frappe.get_roles(user)
#     max_discount = 0
#     for role in user_roles:
#         d = frappe.db.get_value("Quotation Discount",{'role':role,'parent':'Sales Settings'},['role','max_dis'])
#         if d:
#             if max_discount < d[1]:
#                 max_discount = d[1]
#     if max_discount == 0:
#         return 'invalid'
#     elif float(discount) > max_discount:
#         return max_discount


@frappe.whitelist()
def get_query(doctype, txt, searchfield, start, page_len,filters):
	return frappe.db.sql(""" select follow_up_for from `tabSales Follow UP` where follow_up_for = '%s' and account_manager = '%s' """ %(filters.quotation_to,filters.user), as_dict=1)

# @frappe.whitelist()
# def alert_to_substitute(doc,method):
# 	frappe.sendmail(
# 		recipients='janisha.g@groupteampro.com',
# 		subject=('ASSIGNED AS SUBSTITUTE'),
# 		header=(''),
# 		message="""
# 			Dear %s,<br>
# 			%s %s %s is going for vacation %s to %s so you assigned as substitute<br>
# 			and getting into trainee to probation.<br>
# 			"""%(doc.sub_employee_id,doc.employee_name,doc.employee,doc.department,doc.from_date,doc.to_date)
# 	)

#Monthly run the below code - If expired document there in system between 30 days , these documents are sent to below emails 
@frappe.whitelist()
def isthimara_exp_mail(doc,method):
	vehicle_main = frappe.get_value('Vehicle Maintenance Check List',{'register_no':doc.register_no},['expiry_date'])
	Previous_Date = vehicle_main - timedelta(days=30)
	if Previous_Date:
		frappe.sendmail(
		recipients='suvarna@electraqatar.com',
		subject=('Isthimara Expiry'),
		message="""
			Dear Sir/Madam,<br>
			Vehicle No %s is going for Expire.Kindly Renew before Due date
			"""%(doc.register_no)

		)

@frappe.whitelist()
def grade(name,grade):
	doc = frappe.get_doc('Employee Grade',name)
	ticket = doc.air_ticket_allowance_
	print(ticket)

#New contact create from Quotation while create click the create contact buton
@frappe.whitelist()
def create_contact(name,email,designation,customer_name,mobile_no):
	contact = frappe.new_doc("Contact")
	contact.first_name = name
	contact.designation = designation
	contact.mobile_no = mobile_no
	contact.contact_person_email_id_ = email
	contact.append('email_ids',{
		'email_id':email,
		'is_primary':True,
	})
	# contact.append('phone_nos',{
	# 'phone': contact_no,
	# 'is_primary_phone':True,
	# 'is_primary_mobile_no':True,
	# })
	contact.append('links',{
		'link_doctype':"Customer",
		'link_name':customer_name,
	})
	contact.save(ignore_permissions=True)
	# contact = frappe.get_doc("Contact",{"first_name":name)
	# contact.contact_person_name = name
	# contact.email_id = email
	# contact.contact_number = contact_no


# @frappe.whitelist()
# def get_due_date(doc,method):
#     due = doc.payment_schedule
#     due_date = []
#     for dd in due:
#         due.append(dd.due_date)
#     sales_invoice = frappe.get_doc("Sales Invoice",doc.name)
#     sales_invoice.due_date = due_date
#     sales_invoice.save(ignore_permissions=True)
#     frappe.db.commit()


@frappe.whitelist()
def additional_salary(import_file):
	filepath = get_file(import_file)
	pps = read_csv_content(filepath[1])
	for pp in pps:
		if frappe.db.exists('Salary Structure Assignment',{'employee':pp[0]}):
			if pp[0] != "Employee":
				doj = frappe.db.get_value('Employee',{'employee':pp[0]},['date_of_joining'])
				if doj:
					if pd.to_datetime(pp[3]).date() > doj:
						company = frappe.db.get_value("Employee",{'employee':pp[0]},['company'])
						print(company)
						if company:
							doc = frappe.new_doc("Additional Salary")
							doc.employee = pp[0]
							doc.company = company
							doc.salary_component = pp[1]
							doc.amount = int(str(pp[2]).replace(',',''))
							doc.payroll_date = '2022-02-16'
							doc.save(ignore_permissions = True)
							doc.submit


@frappe.whitelist()
def gratuity_calc(employee):
	total_gratuity=0
	basic= frappe.db.get_value('Employee', employee, ['basic'])
	doj = frappe.db.get_value('Employee', employee, ['date_of_joining'])
	date_2 = datetime.now()
	diff = relativedelta.relativedelta(date_2, doj)
	current_yoe = cstr(diff.years) + ' years, ' + cstr(diff.months) + ' months and ' + cstr(diff.days) + ' days'
	exp_years = diff.years
	exp_month = diff.months
	exp_days = diff.days
	total_days = (exp_years * 365) + (exp_month * 30) + exp_days
	basic_salary = basic
	if basic_salary is not None:
		per_day_basic = (basic_salary / 30 * 21) / 365
		total_gratuity = per_day_basic * total_days

	if current_yoe <= "5 years, 0 months and 0 days":
		if diff.years > 1:
			return total_gratuity
	else:
		if diff.years < 1:
			return total_gratuity

	return total_gratuity

#Calculate the gratuity Value based on service years and set the value to gratuity field in Leave salary document
@frappe.whitelist()
def gratuity_calc_sr(employee,basic):
	doj = frappe.db.get_value('Employee',employee,['date_of_joining'])
	current_yoe_days = date_diff(nowdate(),doj)
	current_yoe = round((current_yoe_days / 365),1)
	# yoe = add_years(doc[1],5)
	if current_yoe < 5:
		gratuity_per_year = (int(basic)/30) * 21
		total_gratuity = current_yoe * gratuity_per_year
		return total_gratuity
	if current_yoe >= 5:
		total_yoe = current_yoe
		first_five_yr = ((int(basic)/30) * 21) * 5
		total_yoe -= 5
		rest_of_yrs = ((int(basic)/30) * 30) * total_yoe
		total_gratuity = first_five_yr + rest_of_yrs
		return total_gratuity


#Get the discount value from quotation discount table and check the below conditions and throw message based on return values
@frappe.whitelist()
def check_discount_percent(discount):
	max_discount = 0
	d = frappe.db.get_value("Quotation Discount",{'user':frappe.session.user,'parent':'Sales settings'},['user','max_dis'])
	if d:
		if max_discount < d[1]:
			max_discount = d[1]
	if max_discount == 0:
		return ''
	elif float(discount) > max_discount:
		return max_discount

# @frappe.whitelist()
# def get_dn_list_sales_invoice(doc,method):
#     sii = doc.items
#     dn_list = []
#     dummy = []
#     for i in sii:
#         if  i.delivery_note:
#             if i.delivery_note not in dn_list:
#                 dn_list.append(i.delivery_note or "")
#             else:
#                 dummy.append(i.delivery_note or "")
#     si = frappe.get_doc("Sales Invoice",doc.name)
#     si.delivery_note_list = str(dn_list)
#     si.save(ignore_permissions=True)
#     frappe.db.commit()

@frappe.whitelist()
def item_price():
	sellin_price = 200
	percentage = 10
	value = sellin_price * (1 + (percentage/100))
	print(round(value))

#Create new address while enter the Create button of Quotation
@frappe.whitelist()
def create_address(name,address_1,address_2,city):
	doc = frappe.new_doc("Address")
	doc.address_title = name
	doc.address_line1 = address_1
	doc.address_line1 = address_2
	doc.city = city
	doc.append('links',{
		"link_doctype":"Customer",
		"link_name" : name,
		"link_title" : name,
	})
	doc.save(ignore_permissions=True)

#Create new sales person while enter the sales person in Employee
@frappe.whitelist()
def create_sales_person(name,department,employee):
	doc = frappe.new_doc("Sales Person")
	doc.sales_person_name = employee
	doc.parent_sales_person = "Sales Team"
	doc.employee = name
	doc.department = department
	doc.enabled = True
	doc.save(ignore_permissions=True)

@frappe.whitelist()
def calculate_commission_for_item():
	items = frappe.get_all('Item',['*'])
	for item in items:
		item_group = frappe.get_all('Item Group',['*'])
		for item_g in item_group:
			if item.item_group == item_g.name:
				item.commission_ = item_g.commission

@frappe.whitelist()
def mail_trigger():
	day = datetime.strptime(str(today()),"%Y-%m-%d").date()
	past_two_day = add_days(day,-2)


	day_plan = frappe.db.get_all('Day Plan Timesheet',{'docstatus':'Draft','worked_date':past_two_day},['*'])
	for dp in day_plan:

		project = frappe.get_all('Project',{'status':'Open'},['*'])
		for p in project:
			# print(p.name)
			if dp.project == p.name:
				print(dp.project == p.name)
				frappe.sendmail(
					recipients =['mohamedyousuf.e@groupteampro.com'],
					subject = 'NO Day Plan Timesheet',
					message = 'Dear Sir <br> For the project %s Day Plan TimeSheet is not submitted'%(dp.project)
			)
				print(dp.project)


#Get the company default warehouse and set it to Quotation,DN and SO
@frappe.whitelist()
def get_default_warehouse(company):
	w_house = frappe.db.get_value("Warehouse",{'company':company,'default_for_stock_transfer':1},['name'])
	return w_house

@frappe.whitelist()
def product_bundle(item_code):
	bundle = frappe.db.exists("Product Bundle",{'name':item_code})
	if bundle:
		bundle_item = frappe.get_doc("Product Bundle",item_code)
		return bundle_item.items


#Create new payment entry submission of Sales Invoice
@frappe.whitelist()
def create_payment_entry(doc, method):
	if doc.is_paid == 1:
		doc.status = "Paid"
		pe = frappe.new_doc("Payment Entry")
		pe.posting_date = doc.posting_date
		pe.payment_type = "Pay"
		pe.company = doc.company
		pe.paid_amount = doc.total
		pe.received_amount = doc.total
		pe.paid_to_account_currency = doc.currency
		pe.paid_from_account_currency = doc.currency
		pe.paid_from = doc.debit_to
		pe.party_type = "Supplier"
		pe.party = "INDUSTRIAL EQUIPMENT & SERVICES CO."
		pe.save(ignore_permissions=True)


#Cancel the Quotation linked with Cost estimation ,Linked cost estimation also cancelled
@frappe.whitelist()
def cancel_ce(cost_estimation,quotation):
	if cost_estimation:
		cost_estimation = frappe.get_doc("Cost Estimation",cost_estimation)
		cost_estimation.cancel()

# @frappe.whitelist()
# def cancel_pb(doc,method):
#     project_bud = frappe.get_doc("Project Budget",{'sales_order':doc.name})
#     project_bud.cancel()

#Cancel and Amend the cost estimation , linked CESOW the cancel cost estimation value has been changed to new cost estimation name
@frappe.whitelist()
def amend_ce(doc,method):
	if doc.amended_from:
		ce_sow = frappe.get_all("CE SOW",{'cost_estimation':doc.amended_from},['name','cost_estimation'])
		for ce in ce_sow:
			ces = frappe.get_doc("CE SOW",{'name':ce.name,'cost_estimation':ce.cost_estimation})
			ces.cost_estimation = doc.name
			ces.save(ignore_permissions=True)


#Cancel and Amend the Project Budget , linked CESOW the cancel Project Budget value has been changed to new cost estimation name
@frappe.whitelist()
def amend_pb(doc,method):
	if doc.amended_from:
		pro = frappe.db.exists("Project",{'budgeting':doc.amended_from})
		if pro:
			project = frappe.get_doc("Project",{'budgeting':doc.amended_from})
			project.budgeting = doc.name
			project.save(ignore_permissions = True)
		pb_sow = frappe.get_all("PB SOW",{'project_budget':doc.amended_from},['name','project_budget'])
		for pb in pb_sow:
			pbs = frappe.get_doc("PB SOW",{'name':pb.name,'project_budget':pb.project_budget})
			pbs.project_budget = doc.name
			pbs.save(ignore_permissions=True)
		# update PB in MR
		mrs = frappe.get_all("Material Request", 
					 filters={"project_budget": doc.amended_from, "docstatus": 1}, 
					 pluck="name")
		for mr_name in mrs:
			mr = frappe.get_doc("Material Request", mr_name)
			mr.project_budget = doc.name
			mr.save(ignore_permissions=True)

#Check the count only
@frappe.whitelist()
def find_item():
	items = frappe.db.sql(""" select name from `tabItem` where owner = 'rizni@electraqatar.com' """, as_dict=1)
	for item in items:
		po = frappe.db.sql(""" select count(*) as count from `tabPurchase Order Item` where item_code = '%s' """%(item.name),as_dict = 1)[0]
		if po['count'] == 0:
			pi = frappe.db.sql(""" select count(*) as count from `tabPurchase Invoice Item` where item_code = '%s' """%(item.name),as_dict = 1)[0]
			if pi['count'] == 0:
				pr = frappe.db.sql(""" select count(*) as count from `tabPurchase Receipt Item` where item_code = '%s' """%(item.name),as_dict = 1)[0]
				if pr['count'] == 0:
					si = frappe.db.sql(""" select count(*) as count from `tabSales Invoice Item` where item_code = '%s' """%(item.name),as_dict = 1)[0]
					if si['count'] == 0:
						so = frappe.db.sql(""" select count(*) as count from `tabSales Order Item` where item_code = '%s' """%(item.name),as_dict = 1)[0]
						if so['count'] == 0:
							qo = frappe.db.sql(""" select count(*) as count from `tabQuotation Item` where item_code = '%s' """%(item.name),as_dict = 1)[0]
							if qo['count'] == 1:
								print(item.name)



		# quotation = frappe.db.sql(""" select `tabQuotation`.name from `tabQuotation` left join `tabQuotation Item` on `tabQuotation`.name = `tabQuotation Item`.parent where `tabQuotation Item`.item_code ='%s' """%(it.name),  as_dict=1)
		# if not quotation:
			# print(it.name)
			# sales_order = frappe.db.sql(""" select `tabSales Invoice`.name from `tabSales Invoice` left join `tabSales Invoice Item` on `tabSales Invoice`.name = `tabSales Invoice Item`.parent where `tabSales Invoice Item`.item_code ='%s' """%(it.name),  as_dict=1)

#Expired Visa details has been sent to SMJ mail
@frappe.whitelist()
def visa_expire():
	visa = frappe.db.sql(""" select name,visa_expiry_date from `tabVisa Approval Monitor`""",as_dict = 1)
	data = ''
	data += '<table class = table table - bordered><tr><td colspan = 5>Expired Visa</td></tr>'
	for date in visa:
		expire = date.visa_expiry_date
		if expire:
			str_date = datetime.strptime(str(today()),'%Y-%m-%d').date()
			expiry = add_days(today(),30)
			expiry_date = datetime.strptime(str(expiry),'%Y-%m-%d').date()
			if expire < expiry_date:
				print(expire)
				data += '<tr><td>%s</td><td>%s</td></tr>'%(expire,date.name)
	data += '</table>'
	frappe.sendmail(
			recipients=['mohamedshajith.j@groupteampro.com'],
			subject=('Visa Expiry'),
			header=('Visa Expiry List'),
			message="""
					Dear Sir,<br><br>
					%s
					""" % (data)
		)


#Execute the below method has been changed to all active employee last name
@frappe.whitelist()
def change_name():
	emp = frappe.db.get_all('Employee',{'status':'Active'},['employee'])
	for em in emp:
		name = frappe.get_doc('Employee',em.employee)
		print(name)
		fname = name.first_name
		f_name = fname.title()
		name.first_name = f_name
		if name.last_name:
			lname = name.last_name
			l_name = lname.title()
			name.last_name = l_name
			name.save(ignore_permissions=True)
			frappe.db.commit()



@frappe.whitelist()
def employee_number(doc,method):
	emp = frappe.db.sql(""" select employee_number from `tabEmployee` order by name """,as_dict = 1)[-1]
	last = emp.employee_number
	las = int(last)
	new = las+1
	# doc.employee_number = new
	# doc.name = new
	# frappe.db.set_value('Job Offer',doc.name,'employee_number',new)


#Get the below values and set it to onload of the Vehicle document
@frappe.whitelist()
def record(name):
	maintenance = frappe.db.exists("Vehicle Maintenance Check List",{'register_no':name})
	if maintenance:
		main = frappe.db.get_all("Vehicle Maintenance Check List",{'register_no':name},['complaint','employee_id','vehicle_handover_date','garage_name'])[0]

	accident = frappe.db.exists("Vehicle Accident Report",{'plate_no':name})
	if accident:
		acc = frappe.db.get_all("Vehicle Accident Report",{'plate_no':name},['name','emp_id','date_of_accident','remarks'])[0]

		return main,acc

#Enter the item code for material request , this item match with linked sales order , if not match throw the error
@frappe.whitelist()
def return_so_child(sales_order):
	if sales_order:
		list = []
		so = frappe.get_doc("Sales Order",sales_order)
		for i in so.items:
			list.append(i.item_code)
		return list


#Get the SO items and append the SI
@frappe.whitelist()
def get_so_items(parent):
	so = frappe.get_doc("Sales Order",parent)
	return so.items

@frappe.whitelist()
def get_payment_schedule(parent):
	so = frappe.get_doc("Sales Order",parent)
	return so.payment_schedule_2s

@frappe.whitelist()
def get_rb_report():
	context = {'po_no':'TRD-PO-2022-00281'}
	rb_report = frappe.get_doc('RB Report', 'RB Purchase Order')
	pdf = rb_report.get_pdf(context=context)

#Get the Sales order to match with project so value and set the project name into so
@frappe.whitelist()
def get_po_name(doc,method):
	so = frappe.get_doc("Sales Order",doc.sales_order)
	so.project = doc.name
	so.save(ignore_permissions=True)

#Get the Project Budget to match with sales order pb value and set the project budget name into so
@frappe.whitelist()
def projectbudget_name(doc,method):
	if doc.project_budget:
		pb = frappe.get_doc("Project Budget",doc.project_budget)
		pb.sales_order = doc.name
		pb.save(ignore_permissions=True)

# Create POS Opening Entry is On submission of POS Invoice 
@frappe.whitelist()
def automate_pos(doc,method):
	# pos_profile = frappe.get_list("POS Profile",['name','company'])
	today = date.today()
	# yesterday = today - timedelta(days = 9)
	# print(yesterday)
	poprof = frappe.get_doc("POS Profile",doc.pos_profile)
	pos_open = frappe.new_doc("POS Opening Entry")
	pos_open.period_start_date = today
	pos_open.posting_date = now()
	for ps in poprof.applicable_for_users:
		pos_open.user = ps.user
	pos_open.company = doc.company
	pos_open.pos_profile = poprof.name
	for ps in poprof.payments:
		pos_open.append('balance_details',{
			'mode_of_payment':ps.mode_of_payment,
			'opening_amount':0
		})
	pos_open.save(ignore_permissions=True)
	pos_open.submit()
	frappe.db.commit()

	# pos_invoice = frappe.get_list("POS Invoice",{'status':"Paid",'company':i.company},['name','posting_date','grand_total','customer'])
	pos_invoice = frappe.db.sql(""" select name,posting_date,grand_total,customer from `tabPOS Invoice` where status = "Paid" and docstatus != '2' and company = '%s' and posting_date='%s'"""%(doc.company,today),as_dict = 1)
	pos_close = frappe.new_doc("POS Closing Entry")
	pos_close.period_start_date = today
	pos_close.period_end_date = now()
	pos_close.posting_date = now()
	pos_close.posting_time = now()
	pos_close.pos_opening_entry = pos_open.name
	for ps in poprof.applicable_for_users:
		pos_close.user = ps.user
	pos_close.company = doc.company
	pos_close.pos_profile = poprof.name
	for i in pos_invoice:
		pos_close.append('pos_transactions',{
				'pos_invoice':i.name,
				'posting_date':i.posting_date,
				'grand_total':i.grand_total,
				'customer':i.customer,
				'is_return': i.is_return
			})
	pos_close.save(ignore_permissions=True)
	pos_close.submit()
	frappe.db.commit()


# @frappe.whitelist()
# def bulk_item_cancel(filename):
#     from frappe.utils.file_manager import get_file
#     _file = frappe.get_doc("File", {"file_name": filename})
#     filepath = get_file(filename)
#     pps = read_csv_content(filepath[1])
#     for pp in pps:
#         if pp[0] is not None:
#             print(pp[0])
#             item = frappe.db.sql(""" update `tabItem` set brand = '' where name = '%s'  """%(pp[0]))
#             print(item)


#Append the values to match with given supplier if not creation of payment against PO and PI
@frappe.whitelist()
def return_purchase_invoice(supplier):
	# existing_invoices = frappe.db.sql("""
	# 	SELECT `tabPurchase Invoice List`.name1
	# 	FROM `tabPurchase Invoice List`
	# 	LEFT JOIN `tabConsolidated Payment Request` 
	# 		ON `tabConsolidated Payment Request`.name = `tabPurchase Invoice List`.parent
	# 	WHERE `tabConsolidated Payment Request`.supplier = %s
	# 	AND `tabConsolidated Payment Request`.docstatus != 1
		
	# """, (supplier, today()), as_dict=False)
	if supplier:
		existing_invoices = frappe.db.sql("""
			SELECT pil.name1
			FROM `tabPurchase Invoice List` pil
			LEFT JOIN `tabConsolidated Payment Request` cpr 
				ON pil.parent = cpr.name
			WHERE cpr.supplier = %s
			AND cpr.docstatus = 1
		""", (supplier,), as_dict=False)

		# Convert to a set for easy comparison
		existing_invoice_names = {inv[0] for inv in existing_invoices} if existing_invoices else set()

		# Fetch purchase invoices not in the existing invoice list
		purchase_invoice = frappe.get_list(
			"Purchase Invoice",
			filters={
				"status": ["not in", ["Paid", "Cancelled"]],
				"supplier": supplier,
				"name": ["not in", existing_invoice_names],
			},
			fields=[
				"name", "posting_date", "bill_no", "supplier",
				"currency", "conversion_rate", "grand_total",
				"outstanding_amount", "company"
			],
		)


		pur_order = []
		po_list = frappe.db.get_list("Purchase Order", 
			filters={"supplier": supplier, "docstatus": 1},
			fields=['name', 'grand_total', 'currency', 'conversion_rate', 
					'transaction_date', 'advance_paid', 'supplier', 
					'company', 'project']
		)
		
		for po in po_list:
			total_invoiced = frappe.db.sql("""
				SELECT SUM(`tabPurchase Invoice Item`.amount)
				FROM `tabPurchase Invoice Item`
				WHERE `tabPurchase Invoice Item`.purchase_order = %s
			""", (po['name'],))[0][0] or 0
			if total_invoiced:
				total_invoiced =total_invoiced
			else:
				total_invoiced=0
			remaining_po_value = po['grand_total'] - total_invoiced
			if remaining_po_value > 0:
				pur_order.append({
					'name': po['name'],
					'grand_total': remaining_po_value,
					'total_invoiced': total_invoiced,
					'currency': po['currency'],
					'conversion_rate': po['conversion_rate'],
					'transaction_date': po['transaction_date'],
					'advance_paid': po['advance_paid'],
					'supplier': po['supplier'],
					'company': po['company'],
					'project': po['project']
				})

		# jv = frappe.db.get_list("Journal Entry",{'voucher_type':"Journal Entry","title":supplier,"docstatus":1},['name','total_amount_currency','posting_date','total_credit','total_debit','company','title','bill_no'])
		jv = frappe.db.sql("""select `tabJournal Entry`.name,`tabJournal Entry`.company,`tabJournal Entry`.title,`tabJournal Entry`.bill_no,`tabJournal Entry`.posting_date,`tabJournal Entry Account`.credit_in_account_currency as total_credit from `tabJournal Entry`
					left join `tabJournal Entry Account` on `tabJournal Entry`.name = `tabJournal Entry Account`.parent
					where `tabJournal Entry Account`.party = '%s' and `tabJournal Entry Account`.party_type = 'Supplier'  and `tabJournal Entry`.docstatus != 2 and `tabJournal Entry Account`.credit_in_account_currency > 0  ORDER BY
				`tabJournal Entry`.posting_date ASC"""%(supplier),as_dict=True)
		return purchase_invoice,pur_order,jv



@frappe.whitelist()
def return_list():
	jv = frappe.db.sql("""select `tabJournal Entry`.name,`tabJournal Entry`.company,`tabJournal Entry`.title,`tabJournal Entry`.bill_no,`tabJournal Entry`.posting_date,`tabJournal Entry Account`.credit_in_account_currency as total_credit from `tabJournal Entry`
				left join `tabJournal Entry Account` on `tabJournal Entry`.name = `tabJournal Entry Account`.parent
			where `tabJournal Entry Account`.party = '%s' and `tabJournal Entry Account`.party_type = 'Supplier' and `tabJournal Entry`.voucher_type = 'Journal Entry' and `tabJournal Entry`.title = '%s' and `tabJournal Entry`.docstatus != 2 """%("GLOBEL LINK WESTSTAR SHIPPING WLL","GLOBEL LINK WESTSTAR SHIPPING WLL"),as_dict=True)
	
	# jv = frappe.db.get_list("Journal Entry",{'party_type':"Supplier",'name':"TRD-JV-2023-00079","party":"GLOBEL LINK WESTSTAR SHIPPING WLL"},['name','total_amount_currency','posting_date','total_credit','debit','company','title','bill_no'])
	print(jv)

@frappe.whitelist()
def payment_summary(supplier):
	summary = frappe.db.sql("""select company,sum(outstanding_amount) as total from `tabPurchase Invoice` where status != 'Paid' and docstatus != 2 and supplier = '%s' group by company """%(supplier),as_dict=True)
	data = ''
	data += '<table class="table table-bordered"><tr><th style="padding:1px;border: 1px solid black;color:white;background-color:orange" colspan=4><center>Payment Summary</center></th></tr>'
	data += '''
	<tr><td style="padding:1px;border: 1px solid black" ><b>Company</b></td><td style="padding:1px;border: 1px solid black" ><b>Amount</b></td></tr>'''
	i = 0
	for ps in summary:
		data += '''
		<tr><td style="padding:1px;border: 1px solid black;"  >%s</td>
		<td style="padding:1px;border: 1px solid black;" >%s</td></tr>'''%(ps.company,ps.total)
	i += 1



	data += '</table>'

	return data

#Get the Sales Order and Sales Invoice values and set to the consolidated receivable
@frappe.whitelist()
def return_order_invoice(customer):
	sales_invoice = frappe.get_list("Sales Invoice",{"status": ('not in',("Paid","Cancelled")),'customer':customer},['conversion_rate','currency','name','grand_total','posting_date','outstanding_amount','customer','company'])
	sal_order = []
	so = frappe.db.get_list("Sales Order",{"customer":customer,"docstatus":1},['name'])
	for i in so:
		si_item = frappe.get_value("Sales Invoice Item",{"sales_order":i.name},['parent'])
		if not si_item:
			sal = frappe.db.get_list("Sales Order",{"docstatus":1,"name":i.name},['conversion_rate','currency','name','transaction_date','grand_total','advance_paid','customer','company','project'])
			if sal not in sal_order:
				sal_order.append(sal)

	jv = frappe.get_list("Journal Entry",{'voucher_type':"Journal Entry",'party_type':"Customer","title":customer},['name','total_amount_currency','posting_date','total_credit','total_debit','company','title','bill_no'])
	return sales_invoice,sal_order,jv
	



	data += '</table>'

	return data

#Get the all values from DN to match with POS Invoice DN and set the vlues for onload of the POS Invoice
@frappe.whitelist()
def get_all_dn_details(delivery_note):
	dn = frappe.get_all('Delivery Note',{'name':delivery_note},['*'])
	return dn

#Get the all values from DN items to match with POS Invoice DN and set the vlues for onload of the POS Invoice table
@frappe.whitelist()
def get_all_dn_table(delivery_note):
	dn = frappe.get_doc('Delivery Note',delivery_note)
	return dn.items

#After save of the SI and set the remarks value like below message
@frappe.whitelist()
def sales_invoice_remarks(doc,method):
	if doc.delivery_note:
		if doc.po_no and doc.po_date:
			doc.remarks = ("Against Customer Order {0} dated {1} and Delivery Note {2}").format(
				doc.po_no, formatdate(doc.po_date),doc.delivery_note
			)


#check the customer credit limit ,If not limit throw the message in SO
@frappe.whitelist()
def get_cred_lim(customer):
	empty = "empty"
	cust = frappe.get_doc("Customer",customer)
	if cust.credit_limits:
		return cust.credit_limits
	else:
		return empty


#Get the below etails to match with item group and set it CE SOW Values
@frappe.whitelist()
def get_trans_perc(item_group):
	if item_group:
		item_group = frappe.get_doc("Item Group",item_group)
		for ig_def in item_group.item_group_defaults:
			return ig_def.transfer_percentage

#Upload the credit limit for company wise in Customer
@frappe.whitelist()
def cred_limit_upload(doc,method):
	companies = [
			{"company":"ELECTRA  - NAJMA SHOWROOM","credit_limit" : 100000,"bypass_credit_limit_check":1},
			{"company":"STEEL DIVISION - ELECTRA","credit_limit" : 100000,"bypass_credit_limit_check":1},
			{"company":"MARAZEEM SECURITY SERVICES - HO","credit_limit" : 100000,"bypass_credit_limit_check":1},
			{"company":"MEP DIVISION - ELECTRA","credit_limit" : 100000,"bypass_credit_limit_check":1},
			{"company":"ELECTRA - BINOMRAN SHOWROOM","credit_limit" : 100000,"bypass_credit_limit_check":1},
			{"company":"KINGFISHER TRADING AND CONTRACTING COMPANY","credit_limit" : 100000,"bypass_credit_limit_check":1},
			{"company":"KINGFISHER - SHOWROOM","credit_limit" : 100000,"bypass_credit_limit_check":1},
			{"company":"KINGFISHER - TRANSPORTATION","credit_limit" : 100000,"bypass_credit_limit_check":1},
			{"company":"MARAZEEM SECURITY SERVICES - SHOWROOM","credit_limit" : 100000,"bypass_credit_limit_check":1},
			{"company":"MARAZEEM SECURITY SERVICES","credit_limit" : 100000,"bypass_credit_limit_check":1},
			{"company":"ELECTRA - BARWA SHOWROOM","credit_limit" : 100000,"bypass_credit_limit_check":1},
			{"company":"ELECTRICAL DIVISION - ELECTRA","credit_limit" : 100000,"bypass_credit_limit_check":1},
			{"company":"ENGINEERING DIVISION - ELECTRA","credit_limit" : 100000,"bypass_credit_limit_check":1},
			{"company":"INTERIOR DIVISION - ELECTRA","credit_limit" : 100000,"bypass_credit_limit_check":1},
			{"company":"INDUSTRIAL TOOLS DIVISION","credit_limit" : 100000,"bypass_credit_limit_check":1},
			{"company":"ELECTRA - ALKHOR SHOWROOM","credit_limit" : 100000,"bypass_credit_limit_check":1},
			{"company":"TRADING DIVISION - ELECTRA","credit_limit" : 100000,"bypass_credit_limit_check":1},
			{"company":"Al - Shaghairi Trading and Contracting Company W.L.L (ELECTRA)","credit_limit" : 100000,"bypass_credit_limit_check":1},
					]


	if doc.customer_type == "Company":
		if not doc.credit_limits:
			for company in companies:
				doc.append("credit_limits",{
					"company":company['company'],
					"credit_limit":company['credit_limit'],
					"bypass_credit_limit_check":company['bypass_credit_limit_check']
				})
			doc.save(ignore_permissions=True)


#Get the available qty to match with item code and throw the message in DN WIP
@frappe.whitelist()
def stock_request(item_code):
	stocking = frappe.db.sql("""select actual_qty,warehouse from tabBin where item_code = '%s' """%(item_code),as_dict=True)
	return stocking

@frappe.whitelist()
def warehouse_stock_check(item_code,warehouse):
	stocking = frappe.db.sql("""select actual_qty from tabBin where item_code = '%s' and warehouse = '%s' """%(item_code,warehouse),as_dict=True)
	return stocking


#While using the merge row in Material Request table , If there 2 rows for similar to merge with single row
@frappe.whitelist()
def bom_duplicate(material_request):
	childtab = frappe.db.sql(""" select `tabMaterial Request Item`.item_code,`tabMaterial Request Item`.item_name,`tabMaterial Request Item`.description,`tabMaterial Request Item`.schedule_date,sum(`tabMaterial Request Item`.qty) as qty,
	`tabMaterial Request Item`.uom,`tabMaterial Request Item`.conversion_factor,`tabMaterial Request Item`.warehouse
	from `tabMaterial Request`
	left join `tabMaterial Request Item` on `tabMaterial Request`.name = `tabMaterial Request Item`.parent where `tabMaterial Request`.name = '%s' group by `tabMaterial Request Item`.item_code"""%(material_request),as_dict = 1)
	return childtab

#Get the DN name inside the Sales Invoice Item and set the all names in one field
@frappe.whitelist()
def si_duplicate(sales_invoice,item_details):
	item_details = json.loads(item_details)
	dn_list = []
	if item_details:
		for i in item_details:
			if i not in dn_list:
				dn_list.append(i or "")
			dnlist = str(dn_list)

	sales_invoice = frappe.db.sql(""" select `tabSales Invoice Item`.item_code,`tabSales Invoice Item`.income_account,`tabSales Invoice Item`.expense_account,`tabSales Invoice Item`.item_name,`tabSales Invoice Item`.description,sum(`tabSales Invoice Item`.qty) as qty,
	`tabSales Invoice Item`.uom,`tabSales Invoice Item`.conversion_factor,`tabSales Invoice Item`.warehouse,`tabSales Invoice Item`.rate,
	`tabSales Invoice Item`.base_rate,
	`tabSales Invoice Item`.amount,
	`tabSales Invoice Item`.base_amount,
	`tabSales Invoice Item`.net_rate,`tabSales Invoice Item`.base_net_rate,
	`tabSales Invoice Item`.net_amount,`tabSales Invoice Item`.base_net_amount
	from `tabSales Invoice`
	left join `tabSales Invoice Item` on `tabSales Invoice`.name = `tabSales Invoice Item`.parent where `tabSales Invoice`.name = '%s' group by `tabSales Invoice Item`.item_code"""%(sales_invoice),as_dict = 1)
	return sales_invoice,dnlist or 'null' ,item_details

#Get the Material Request No from the PO Child table and set it to one field
@frappe.whitelist()
def mr_duplicate(item_details):
	item_details = json.loads(item_details)
	dn_list = []
	if item_details:
		for i in item_details:
			if i not in dn_list:
				dn_list.append(i or "")
			dnlist = str(dn_list)

	return dnlist or 'null'

#Get the Project from the PO Child table and set it to one field
@frappe.whitelist()
def project_duplicate(item_details):
	item_details = json.loads(item_details)
	dn_list = []
	if item_details:
		for i in item_details:
			if i not in dn_list:
				dn_list.append(i or "")
			dnlist = str(dn_list)

	return dnlist or 'null'

#Get the below value and set it to Project Time Sheet table
@frappe.whitelist()
def proj_day_plan_emp(multi_project_day_plan):
	proj = frappe.get_doc("Multi Project Day Plan",multi_project_day_plan)
	return proj.project_day_plan_employee


#Avoid the duplicate document
@frappe.whitelist()
def prob_eval(employee_name):
	prob = frappe.db.exists("Staff Probation Evaluation Form",employee_name)
	if prob:
		return True
	else:
		return False


#Get the below value and set it to Project field in Single Project Day Plan
@frappe.whitelist()
def single_proj_day_plan(day_schedule):
	ds = frappe.get_doc("Day Schedule",day_schedule)
	return ds.schedule

#Check the below condition before submission of DN
@frappe.whitelist()
def is_dummy_dn_approved(doc,method):
	if doc.dummy_dn:
		if not doc.is_approved == 1:
			frappe.throw("Delivery Note Not Approved")

#Check the below condition before submission of SO
@frappe.whitelist()
def is_si_approved(doc,method):
	if doc.is_third_party_commission_applicable:
		if not doc.is_approved == 1:
			frappe.throw("Sales Order Not Approved")


@frappe.whitelist()
def work_order_bom(item_code,company,qty,sales_order,project):
	boms = frappe.db.get_all("BOM",{'item':item_code},'name')
	if boms:
		for i in boms:
			bom = frappe.get_doc("BOM",i.name)
			work = frappe.new_doc("Work Order")
			work.qty = int(qty)
			work.company = company
			work.sales_order = sales_order
			work.project = project
			work.status = "In Process"
			work.bom_no = i.name
			work.production_item = bom.item
			work.wip_warehouse = frappe.db.get_value('Warehouse', filters={'name': ['like', '%Work In Progress%'],'company':company})
			work.fg_warehouse = frappe.db.get_value('Warehouse', filters={'name': ['like', "%" + project + "%"]})
			work.source_warehouse = frappe.db.get_value('Warehouse', {'company':company,'default_for_stock_transfer':1}, 'name')
			work.planned_start_date = now()
			for b in bom.items:
				work.append("required_items",{
					"item_code": b.item_code,
					"source_warehouse": frappe.db.get_value('Warehouse', {'company':company,'default_for_stock_transfer':1}, 'name'),
					"include_item_in_manufacturing":1,
					"item_name": b.item_name,
					"qty": b.qty * int(qty),
					"rate": b.rate,
					"amount": b.rate *(b.qty * int(qty)),
				})
			work.save(ignore_permissions=True)
			work.submit()

@frappe.whitelist()
def set_title():
	title = frappe.db.get_list("Sales Invoice",{'posting_date':"31-01-2023"},['name'])
	for i in title:
		doc = frappe.get_doc("Sales Invoice",i.name)
		doc.title = doc.customer
		doc.save(ignore_permissions=True)
		print(i.name)


#Create contact while enter the create button Quotation
@frappe.whitelist()
def contact_details(name,email,mobile,customer,designation):
	# values = json.loads(values)
	contact = frappe.new_doc("Contact")
	contact.first_name = name
	contact.contact_person_email_id_ = email
	contact.mobile_no = mobile
	contact.designation = designation
	if email:
		contact.append('email_ids',{
			'email_id': email
		})
	if mobile:
		contact.append('phone_nos',{
			'phone':mobile,
		})
	if customer:
		contact.append('links',{
			'link_doctype':'Customer',
			"link_name": customer
		})
	contact.save(ignore_permissions=True)
	frappe.msgprint("Contact Created")
	return contact.name

#The below method run by everyday 1AM (Revised the employee salary based on below condition)
@frappe.whitelist()
def update_employee_salary():
	sal = frappe.db.sql("""select * from `tabSalary Revision` where docstatus = 1""", as_dict=1)
	if sal:
		for ss in sal:
			if ss.salary_revision_status == "Permanent":
				if ss.effective_date == datetime.strptime((today()), '%Y-%m-%d').date():
					emp = frappe.get_doc("Employee",{'employee':ss.employee,'status': 'Active' })
					emp.salary_mode = ss.salary_mode
					emp.salary_currency = ss.salary_currency
					emp.basic = ss.basic
					emp.hra = ss.hra
					emp.accommodation = ss.accommodation
					emp.mobile_allowance = ss.mobile_allowance
					emp.transportation = ss.transport_allowance
					emp._other_allowance = ss.other_allowance
					emp.medical_allowance_ = ss.medical_allowance
					emp.leave_salary = ss.leave_salary
					emp.air_ticket_allowance_ = ss.air_ticket_allowance
					emp.qid_cost = ss.qid_cost
					emp.medical_renewal = ss.medical_renewal
					emp.visa_cost_ = ss.visa_cost
					emp.gross_salary = ss.gross_salary
					emp.gratuity = ss.gratuity
					emp.ctc = ss.cost_to_company
					emp.per_hour_cost = ss.per_hour_cost
					emp.append('history',{
					"date":ss.effective_date,
					"basic":ss.basic,
					"hra":ss.hra,
					"other_allowance":ss.other_allowance,
					"transport_allowance":ss.transport_allowance,
					"medical_allowance":ss.medical_allowance,
					"air_ticket_allowance":ss.air_ticket_allowance,
					"mobile_allowance":ss.mobile_allowance,
					"visa_cost":ss.visa_cost,
					"accommodation":ss.accommodation,
					"leave_salary":ss.leave_salary,
					"qid_cost":ss.qid_cost,
					"medical_renewal":ss.medical_renewal,
					"remark":ss.remarks
				})
					emp.save(ignore_permissions=True)
					frappe.db.commit()
			else:
				if ss.effective__to_date == datetime.strptime((today()), '%Y-%m-%d').date():
					emp = frappe.get_doc("Employee",{'employee':ss.employee,'status': 'Active' })
					emp.salary_mode = ss.salary_mode1
					emp.salary_currency = ss.salary_currency1
					emp.basic = ss.basic1
					emp.hra = ss.hra1
					emp.accommodation = ss.accommodation1
					emp.mobile_allowance = ss.mobile_allowance1
					emp.transportation = ss.transport_allowance1
					emp._other_allowance = ss.other_allowance1
					emp.medical_allowance_ = ss.medical_allowance1
					emp.leave_salary = ss.leave_salary1
					emp.air_ticket_allowance_ = ss.air_ticket_allowance1
					emp.qid_cost = ss.qid_cost1
					emp.medical_renewal = ss.medical_renewal1
					emp.visa_cost_ = ss.visa_cost1
					emp.gross_salary = ss.gross_salary1
					emp.gratuity = ss.gratuity1
					emp.ctc = ss.cost_to_company1
					emp.per_hour_cost = ss.per_hour_cost1
					emp.save(ignore_permissions=True)
					frappe.db.commit()
				if ss.effective__from_date == datetime.strptime((today()), '%Y-%m-%d').date():
					emp = frappe.get_doc("Employee",{'employee':ss.employee,'status': 'Active' })
					emp.salary_mode = ss.salary_mode1
					emp.salary_currency = ss.salary_currency1
					emp.basic = ss.basic2
					emp.hra = ss.hra2
					emp.accommodation = ss.accommodation2
					emp.mobile_allowance = ss.mobile_allowance2
					emp.transportation = ss.transport_allowance2
					emp._other_allowance = ss.other_allowance2
					emp.medical_allowance_ = ss.medical_allowance2
					emp.leave_salary = ss.leave_salary2
					emp.air_ticket_allowance_ = ss.air_ticket_allowance2
					emp.qid_cost = ss.qid_cost2
					emp.medical_renewal = ss.medical_renewal2
					emp.visa_cost_ = ss.visa_cost2
					emp.gross_salary = ss.gross_salary2
					emp.gratuity = ss.gratuity2
					emp.ctc = ss.cost_to_company2
					emp.per_hour_cost = ss.per_hour_cost2
					emp.save(ignore_permissions=True)
					frappe.db.commit()

#Show the below details while open the quotation document if there in profit percentage
@frappe.whitelist()
def margin_html(mar_dis_per,mar_tot_cost,mar_gross_tot,mar_dis_amt,total_selling,prof_per,margin_profit):
	data = ''
	data = "<table style='width:85%'>"
	data += "<tr><td style ='background-color:#dedede;font-weight:bold;text-align:center;border:1px solid black'>TOTAL COST</td><td style ='font-weight:bold;text-align:center;border:1px solid black'>%s</td><td style ='background-color:#dedede;font-weight:bold;text-align:center;border:1px solid black'>GROSS TOTAL SELLING</td><td style ='font-weight:bold;text-align:center;border:1px solid black'>%s</td></tr>"%(round(float(mar_tot_cost),2),round(float(mar_gross_tot),2))
	data += "<tr><td colspan  = 2 style = 'border:1px solid black;border-right-color:#dedede;background-color:#dedede;'><td style ='border-left-color:#dedede;background-color:#dedede;font-weight:bold;text-align:center;border:1px solid black'>DISCOUNT PERCENTAGE</td><td style ='background-color:#74cee7;font-weight:bold;text-align:center;border:1px solid black'>%s</td></tr>"%(round(float(mar_dis_per),2))
	data += "<tr><td style ='background-color:#dedede;font-weight:bold;text-align:center;border:1px solid black'>PROFIT PERCENTAGE</td><td style ='background-color:#74cee7;font-weight:bold;text-align:center;border:1px solid black'>%s</td><td style ='background-color:#dedede;font-weight:bold;text-align:center;border:1px solid black'>DISCOUNT AMOUNT</td><td style ='background-color:#74cee7;font-weight:bold;text-align:center;border:1px solid black'>%s</td></tr>"%(round(float(prof_per),2),round(float(mar_dis_amt),2))
	data += "<tr><td style ='background-color:#dedede;font-weight:bold;text-align:center;border:1px solid black'>PROFIT AMOUNT</td><td style ='background-color:#74cee7;font-weight:bold;text-align:center;border:1px solid black'>%s</td><td style ='background-color:#dedede;font-weight:bold;text-align:center;border:1px solid black'>TOTAL SELLING</td><td style ='font-weight:bold;text-align:center;border:1px solid black'>%s</td></tr>"%(round(float(margin_profit),2),round(float(total_selling),2))
	data += "</table>"
	return data



# @frappe.whitelist()
# def create_leave_salary(name,emp_id,grade,department,designation,status,date_of_joining):
# 	# employee = frappe.get_value("Employee",{'employee':employee,'status': 'Active' },["hra","gross_salary","air_ticket_allowance_","leave_salary"])
#         leave_sal = frappe.new_doc('Leave Salary')
#         leave_sal.employee_name = name
#         leave_sal.employee_number = emp_id
#         leave_sal.grade = grade
#         leave_sal.department = department
#         leave_sal.designation = designation
#         leave_sal.status = status
#         leave_sal.joining_date = date_of_joining
#         leave_sal.save(ignore_permissions=True)
#         frappe.msgprint("Leave Salary Created Sucessfully")

@frappe.whitelist()
def update_status():
	status = frappe.db.sql("""update `tabAdditional Salary` set docstatus = 0 where payroll_date between '2022-12-16' and '2023-01-15' and company = "TRADING DIVISION - ELECTRA" and salary_component = 'Abs Deduction' """)
	print(status)

@frappe.whitelist()
def update_si_status():
	status = frappe.db.sql("""update `tabSales Invoice` set docstatus = 1 where name = "ITD-CRD-2023-00012" """)
	print(status)

# @frappe.whitelist()
# def gratuity_calc_sr(employee,basic):
# 	doj = frappe.db.get_value('Employee',employee,['date_of_joining'])
# 	current_yoe_days = date_diff(nowdate(),doj)
# 	current_yoe = round((current_yoe_days / 365),1)
# 	# yoe = add_years(doc[1],5)
# 	if current_yoe < 5:
# 		gratuity_per_year = (int(basic)/30) * 21
# 		total_gratuity = current_yoe * gratuity_per_year
# 		return total_gratuity
# 	if current_yoe >= 5:
# 		total_yoe = current_yoe
# 		first_five_yr = ((int(basic)/30) * 21) * 5
# 		total_yoe -= 5
# 		rest_of_yrs = ((int(basic)/30) * 30) * total_yoe
# 		total_gratuity = first_five_yr + rest_of_yrs
# 		return total_gratuity


#While enter the employee in Salary Revision get the below values and set it to this document
@frappe.whitelist()
def get_data(employee):
	emp = frappe.get_value("Employee",{'employee':employee,'status': 'Active' },["basic","hra","_other_allowance","accommodation","transportation","air_ticket_allowance_","mobile_allowance","medical_renewal","visa_cost_","medical_allowance_","leave_salary","gratuity","qid_cost","gross_salary","ctc","per_hour_cost",'compensationemployee_insurence'])
	return emp

#get the below values and set it to Sales revision 
@frappe.whitelist()
def get_values(employee,basic,hra,other_allowance,accommodation,transport_allowance,air_ticket_allowance,mobile_allowance,medical_renewal,visa_cost,medical_allowance,leave_salary,qid_cost,comp):
	gratuity = 0
	doj = frappe.db.get_value('Employee',employee,['date_of_joining'])
	current_yoe_days = date_diff(nowdate(),doj)
	current_yoe = round((current_yoe_days / 365),1)
	if current_yoe < 5:
		gratuity_per_year = (int(basic)/30) * 21
		gratuity = current_yoe * gratuity_per_year
	if current_yoe >= 5:
		total_yoe = current_yoe
		first_five_yr = ((int(basic)/30) * 21) * 5
		total_yoe -= 5
		rest_of_yrs = ((int(basic)/30) * 30) * total_yoe
		gratuity = first_five_yr + rest_of_yrs
	gross = int(basic) + int(hra) + int(other_allowance) + int(accommodation) + int(transport_allowance) + int(mobile_allowance)
	ctc = int(air_ticket_allowance) + int(gross) + int(medical_renewal) + int(visa_cost) + int(medical_allowance) + int(leave_salary) + int(gratuity) + int(qid_cost) + int(comp)
	phc = (int(ctc)/(30*8))
	phc = float(basic)/30/8
	return gratuity,gross,ctc,phc


#Get the Purchase rate of the given item and set this value in Landed cost voucher
@frappe.whitelist()
def get_purchase_rate(purchase_receipt):
	pr = frappe.db.sql("""select `tabPurchase Invoice Item`.base_net_rate,`tabPurchase Invoice Item`.parent,`tabPurchase Invoice Item`.item_code from `tabPurchase Invoice Item`
	where `tabPurchase Invoice Item`.purchase_receipt = '%s' """%(purchase_receipt),as_dict=True)
	if pr:
		tax = frappe.get_doc("Purchase Invoice",pr[0].parent)
		return pr[0].base_net_rate,pr[0].item_code,tax.taxes,tax.posting_date,tax.name

#While enter the supplier in Consolidated payment request get the supplier currency and set it
@frappe.whitelist()
def get_def_currency(supplier):
	supp = frappe.db.get_value("Supplier",supplier,['default_currency'])
	conversion = get_exchange_rate(supp, "QAR")
	return supp,conversion 

@frappe.whitelist()
def get_clear():
	status = frappe.db.sql("""Update `tabSalary Structure` set docstatus = 0 where company = 'INTERIOR DIVISION - ELECTRA' """)
	print(status)
	# status = frappe.db.sql("""delete from `tabSalary Structure Assignment` """)
	# print(status)
	# status = frappe.db.sql("""delete from `tabPayroll Entry` """)
	# print(status)

#Get the norden item details to match with given item in the product search and show the details in html view
@frappe.whitelist()
def get_norden_item(item):
	url = "https://erp.nordencommunication.com/api/method/norden.custom.get_electra_details?item=%s" % (item)
	headers = { 'Content-Type': 'application/json','Authorization': 'token 28a1f5da5dffd46:812f4d4d2671af2'}
	# params = {"limit_start": 0,"limit_page_length": 20000}

	response = requests.request('GET',url,headers=headers)
	res = json.loads(response.text)
	return res

@frappe.whitelist()
def update_all_desc():
	item = frappe.db.get_all("Item",{"modified":('between',("2023-03-19","2023-03-22"))},["name"])
	for i in item:
		itm = frappe.get_doc("Item",i.name)
		itm.description = itm.item_name
		itm.save(ignore_permissions=True)
		print()

@frappe.whitelist()
def get_again_so():
	sales_order = frappe.db.sql("""select `tabSales Invoice Item`.sales_order from `tabSales Invoice` left join `tabSales Invoice Item` on `tabSales Invoice`.name = `tabSales Invoice Item`.parent where `tabSales Invoice`.name = '%s' """%("SHBW-SR-2023-00032"),as_dict=True)[0]
	print(sales_order)

@frappe.whitelist()
def copy_doc():
	sn = frappe.get_doc("Sales Invoice","SHMS-STI-2023-00118")
	dn = frappe.copy_doc(sn)
	# dn.set("items",[])
	dn.flags.ignore_permissions = True
	dn.flags.ignore_mandatory = True
	dn.insert()

# @frappe.whitelist()
# def update_stock_after_return(doc,method):
# 	# if doc.order_type == 'Project' and doc.is_return == 1:
# 	if doc.is_return == 1:
# 		if doc.update_stock == 0:
# 			if doc.delivery_note:
# 				if not doc.status == "Return":
# 					frappe.errprint(doc.status)
# 					sn = frappe.get_doc("Delivery Note",doc.delivery_note)
# 					dn = frappe.copy_doc(sn)
# 					dn.is_return = 1
# 					dn.delivery_return = doc.delivery_note_list
# 					dn.posting_date = doc.posting_date
# 					s = doc.delivery_note_list
# 					dn.naming_series = get_dn_return_series(doc.company,"Delivery Note")
# 					dn.set("items",[])
# 					stat = s.strip("[']'")
# 					dn_lst = stat.split(",")
# 					for s in dn_lst:
# 						dn.append("return_delivery_note",{
# 							"delivery_note":s.strip().lstrip("'").rstrip("'")
# 						})
# 						dn.return_against = s.strip().lstrip("'").rstrip("'")
# 					for i in doc.items:
# 						for si in sn.items:
# 							if si.item_code == i.item_code:
# 								si.returned_qty += -i.qty
# 						dn.append("items",{
# 							"item_code":i.item_code,
# 							"item_name":i.item_name,
# 							"description":i.description,
# 							"qty":i.qty,
# 							"uom":i.uom,
# 							"stock_uom":i.uom,
# 							"rate":i.rate,
# 							"conversion_factor":1,
# 							"base_rate":i.base_rate,
# 							"amount":i.amount,
# 							"warehouse":i.warehouse,
# 							"base_amount":i.base_amount,
# 						})
# 					dn.save(ignore_permissions = True)
# 					ret_qty = sum(i.returned_qty for i in sn.items)
# 					sn.per_returned = (ret_qty / sn.total_qty)*100
# 					sn.flags.ignore_validate_update_after_submit = True
# 					sn.save(ignore_permissions = True)
# 					dn.submit()
# 					sn.update_prevdoc_status()
# 					sn.update_billing_status()
# 					frappe.db.set_value("Sales Invoice",doc.name,'custom_delivery_note_return',dn.name)
					
			
#get the so items and check with material request if not match throw the message
@frappe.whitelist()
def get_so_child(sales_order):
	if sales_order:
		so = frappe.get_doc("Sales Order",sales_order)
		return so.items


@frappe.whitelist()
# def enqueue_checkin_bulk_upload_csv(filename):
# 	frappe.enqueue(
# 		checkin_bulk_upload_csv, # python function or a module path as string
# 		queue="long", # one of short, default, long
# 		timeout=36000, # pass timeout manually
# 		is_async=True, # if this is True, method is run in worker
# 		now=False, # if this is True, method is run directly (not in a worker) 
# 		job_name='Salary Component Updated', # specify a job name
# 		enqueue_after_commit=False, # enqueue the job after the database commit is done at the end of the request
# 		filename=filename, # kwargs are passed to the method as arguments
# 	)    
def checkin_bulk_upload_csv():
	# filename = "/files/Salary_Component.csv"
	from frappe.utils.file_manager import get_file
	_file = frappe.get_doc("File", {"file_url": "/files/Salary_Component.csv"})
	filepath = get_file("/files/Salary_Component.csv")
	pps = read_csv_content(filepath[1])
	for pp in pps:
		print(pp[0])
		if frappe.db.exists('Salary Component',{'name':pp[0]}):
			sc = frappe.db.exists('Salary Component',{'name':pp[0]},['name'])
			ac = frappe.db.sql("""select * from `tabSalary Component Account` where `tabSalary Component Account`.parent = '%s' """%(sc),as_dict=True)
			for i in ac:
				print("HI")
				


@frappe.whitelist()
def update():
	le = frappe.db.sql("""update `tabJob Offer` set company = "KINGFISHER - TRANSPORTATION" where name = 'HR-OFF-2023-00012' """)

# @frappe.whitelist()
# def validate_loan_amount(doc, method):
#     applicant = doc.applicant
#     loan_type = doc.loan_type
#     loan_amount = doc.loan_amount

#     # Calculate the sum of loan amounts for the employee and loan type
#     total_loan_amount = frappe.db.sql("""
#         SELECT SUM(loan_amount)
#         FROM `tabLoan Application`
#         WHERE applicant = %s
#             AND loan_type = %s
#             AND docstatus = 1
#     """, (applicant, loan_type))[0][0] or 0

#     loan_type_doc = frappe.get_doc("Loan Type", loan_type)

#     if total_loan_amount + loan_amount > loan_type_doc.maximum_loan_amount:
#         frappe.throw(_("Total Loan Amount for the employee cannot exceed Maximum Loan Amount for the Loan Type."))




# import frappe
# from datetime import datetime
# from dateutil import relativedelta

# @frappe.whitelist()
# def check_loan_amount(employee, loan_amount, interest_rate,loan_type):
# 	# Get the employee's basic amount
# 	basic_amount = frappe.db.get_value('Employee', employee, 'basic')
# 	print(basic_amount)
# 	employees = frappe.get_all('Employee', {'status': 'Active'}, ['name', 'employee_name', 'date_of_joining'])
# 	for emp in employees:
# 		date_2 = datetime.now()
# 		# Get the interval between two dates
# 		diff = relativedelta.relativedelta(date_2, emp.date_of_joining)
# 		yos = str(diff.years) + ' years, ' + str(diff.months) + ' months and ' + str(diff.days) + ' days'
# 		exp_years = diff.years
# 		exp_month = diff.months
# 		exp_days = diff.days
# 		total_days = (exp_years * 365) + (exp_month * 30) + exp_days
# 		print(total_days)
# 		per_day_basic = (float(basic_amount) / 30 * 21) / 365
# 		total_gratuity = per_day_basic * total_days
# 		gratuity = (total_gratuity * 80 /100)
# 	# Convert loan_amount and interest_rate to appropriate numeric types
# 	loan_amount = float(loan_amount)
# 	interest_rate = int(interest_rate)

# 	# Calculate the loan amount plus the calculated interest
# 	calculated_amount = loan_amount

# 	# Check if the calculated amount exceeds the basic amount
	
# 	if calculated_amount > gratuity :
# 		if loan_type == "Loan Against Gratuity" or "Loan Against Gratuity_2" or "Loan Against Gratuity_3" :
# 			frappe.throw(_('Loan amount must be Greater than Employee Gratuity'))

# 	if loan_type == "Loan Against Leave  Settlement_1" or "Loan Against Leave  Settlement" or "Loan Against Leave  Settlement_2" :
# 		gross = frappe.db.get_value('Employee', employee, 'gross_salary')
# 		loan_amount = float(loan_amount)
# 		interest_rate = int(interest_rate)

# 	# Calculate the loan amount plus the calculated interest
# 		calculated_amount = loan_amount
# 		if calculated_amount > gross :
# 			frappe.throw(_('Loan amount must be Greater than Employee Gratuity'))
		
# 	# Calculate the loan amount plus the calculated interest
		

# @frappe.whitelist()
# def get_sales_person_invoice(name):
# 	data = "<table  width = 100% border= 1px solid black><tr style= background-color:#4682B4 ; ><td colspan = 1  width = 10 ><b style = color:white ; text-align:center;>SL.NO</b></td><td colspan = 1 width=10 style = color:white><b style = color:white ; text-align:center;>Date</b></td><td colspan = 1 width=10 style = color:white ><b style = color:white ; text-align:center;>Invoice Number</b></td><td colspan = 1 width=10 style = color:white><b style = color:white ; text-align:center;>Customer Name</b></td><td colspan = 1 width= 10 style = color:white><b style = color:white ; text-align:center;>LPO NO</b><td colspan = 1 width=10 style = color:white><b style = color:white ; text-align:center;>Gross Amount</b></td><td colspan = 1 width=10 style = color:white><b style = color:white ; text-align:center;>Discount</b></td><td colspan = 1 width=10 style = color:white><b style = color:white ; text-align:center;> Ret.Amount</b></td><td width=10 style = color:white><b style = color:white ; text-align:center;>Net</b></td><td colspan = 1 width=5 style = color:white><b style = color:white ; text-align:center;> Collected </b></td><td colspan = 1 width=5 style = color:white><b style = color:white ; text-align:center;>Balance</b></td></tr>"
# 	sales_person = frappe.get_all("Sales Person",fields=["name"])
# 	for s in sales_person:
# 		# print(s.name)
# 		data += '<tr><td colspan = 11 style="border: 1px solid black; background-color:#EBF4FA" ><b style= color:"red">%s</b></td></tr>'%(s.name)
# 		sales_invoice = frappe.get_all("Sales Invoice",{"sales_person_user":s.name,"is_return":0},["name","customer","total","discount_amount","posting_date","outstanding_amount","po_no"])
# 		j=1
# 		for i in sales_invoice:
# 			net = (i.total - i.discount_amount)
# 			net_int = int(net)
# 			net_amount = (i.total - net)
# 			convert_float = int(net_amount)
# 			print(net)

# 			collected = (net - i.outstanding_amount)
# 			collected_int = int(collected)
# 			# data += "<tr><td rowspan = '%s' >%s</td><td>%s</td></tr>"%(emp_name_len,i['project_name'],i['employee_name'][0])
# 			data += '<tr><td colspan = 1 style="border: 1px solid black">%s</td><td colspan = 1 style="border: 1px solid black">%s</td><td colspan = 1 style="border: 1px solid black">%s</td><td colspan = 1 style="border: 1px solid black">%s</td><td colspan=1 style="border:1px solid black">%s</td><td colspan = 1 style="border: 1px solid black">%s</td><td colspan = 1 style="border: 1px solid black">%s</td><td colspan = 1 style="border: 1px solid black">%s</td><td style="border: 1px solid black">%s</td><td colspan = 1 style="border: 1px solid black">%s</td><td colspan = 1 style="border: 1px solid black">%s</td></tr>'%(j,i.posting_date,i.name,i.customer,i.po_no,i.total,i.discount_amount,convert_float,net_int,collected_int,i.outstanding_amount)
# 			# print(i.name,i.customer,i.grand_total,i.discount_amount)

# 			j += 1
# 	data += '</table>'
# 	return data
# @frappe.whitelist()
# def get_accounts_ledger(doc):
#     total_amount = 0
#     total_paid = 0
#     total_credit_note = 0
#     total_outstanding = 0
#     total_0_30 = 0
#     total_31_60 = 0
#     total_61_90 = 0
#     total_91_above = 0
#     sales_invoices = frappe.get_all("Sales Invoice", {'company': doc.company, 'customer': doc.customer}, ['posting_date', 'name', 'total', 'outstanding_amount'],order_by="posting_date")
#     if sales_invoices:
#         data = "<table width=100% border=1px solid black><tr style=background-color:#e35310;font-size:8px><td colspan=1 ><b style=color:white; text-align:center;width:320px>Invoice No</b></td><td colspan=2  style=color:white><b style=color:white; text-align:center;>Date</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Age</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Amount QR</b></td><td colspan=1 style=color:white><b style=color:white; text-align:center;>Paid QR</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Credit Note QR</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Outstanding QR</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>0-30</b></td><td  style=color:white><b style=color:white; text-align:center;>31-60</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>61-90</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>91-Above</b></td></tr>"
#         for i in sales_invoices:
#             days = date_diff(doc.from_date, i.posting_date)
#             if 0 <= days <= 120:
#                 total_amount += i.total
#                 total_outstanding += i.outstanding_amount
#                 data += f'<tr><td colspan=1 style="border: 1px solid black; font-size:8px" >{i.name}</td><td colspan=2 style="border: 1px solid black; font-size:8px" >{i.posting_date.strftime("%d-%m-%Y")}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{days}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{i.total}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{i.outstanding_amount}</td>'
#                 if 0 <= days <= 30:
#                     total_0_30 += i.total
#                     data += f'<td colspan=1 style="border: 1px solid black; font-size:8px" >{i.total}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td>'
#                 if 31 <= days <= 60:
#                     total_31_60 += i.total
#                     data += f'<td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{i.total}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td>'
#                 if 61 <= days <= 90:
#                     total_61_90 += i.total
#                     data += f'<td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{i.total}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td>'
#                 if 91 <= days <= 120:
#                     total_91_above += i.total
#                     data += f'<td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{i.total}</td>'					
#         data += '</tr>'
#         data += f'<tr><td colspan=4 style="border: 1px solid black; font-size:8px" >Total</td><td style="border: 1px solid black; font-size:8px" >{total_amount}</td><td style="border: 1px solid black; font-size:8px" >-</td><td style="border: 1px solid black; font-size:8px" >-</td><td style="border: 1px solid black; font-size:8px" >{total_outstanding}</td><td style="border: 1px solid black; font-size:8px" >{total_0_30}</td><td style="border: 1px solid black; font-size:8px" >{total_31_60}</td><td style="border: 1px solid black; font-size:8px" >{total_61_90}</td><td style="border: 1px solid black; font-size:8px" >{total_91_above}</td></tr>'
#         data += '</table>'
#     else:
#         data = "<table width=100% border=1px solid black><tr style=background-color:#e35310;font-size:10px;text-align:center><td colspan=1 ><b style=color:white; text-align:center;width:320px>Sales Invoice Documents Not Avaliable</b></td></tr>"

#     return data


# @frappe.whitelist()
# def get_sales_person(doc, name, from_date, to_date, company, sales_person_user):
#     data = "<table width=100% border=1px solid black><tr style=background-color:#4682B4;font-size:8px><td colspan=1 ><b style=color:white; text-align:center;width:320px>SL.NO</b></td><td colspan=2  style=color:white><b style=color:white; text-align:center;>Date</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Invoice Number</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Customer Name</b></td><td colspan=1 style=color:white><b style=color:white; text-align:center;>LPO NO</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Gross Amount</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Discount</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Ret. Amount</b></td><td  style=color:white><b style=color:white; text-align:center;>Net</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Collected</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Balance</b></td></tr>"

#     if not doc.company_multiselect and not sales_person_user:
#         sales_person = frappe.get_all("Sales Person", fields=["name"])
#         j = 1
#         for s in sales_person:
#             company_set = set()
#             prev_company_name = None
#             sales = frappe.get_all("Sales Invoice", {'posting_date': ('between', (from_date, to_date)), 'sales_person_user': s.name,'docstatus':("not in",[0,2])}, ["name", "customer", "total", "discount_amount", "posting_date", "outstanding_amount", "po_no", "company"],order_by="posting_date")
#             company_totals = {}  
#             salesperson_totals = {} 

#             if sales:
#                 data += '<tr><td colspan=12 style="border: 1px solid black; background-color:#EBF4FA; font:size:8px" ><b style=color:"red">%s</b></td></tr>' % (s.name)

#             for i in sales:
#                 company_name = i["company"]
#                 if company_name not in company_set:
#                     if prev_company_name:
#                         data += f'<tr style = font-size:8px><td colspan=6 style="border: 1px solid black"><b>Total for {prev_company_name}</b></td><td colspan=1 style="border: 1px solid black"><b>{round(company_totals[prev_company_name]["total"],2)}</b></td><td colspan=1 style="border: 1px solid black"><b>{round(company_totals[prev_company_name]["discount_amount"],2)}</b></td><td colspan=1 style="border: 1px solid black"><b>{round(company_totals[prev_company_name]["convert_float"],2)}</b></td><td style="border: 1px solid black"><b>{round(company_totals[prev_company_name]["net_total"],2)}</b></td><td colspan=1 style="border: 1px solid black"><b>{round(company_totals[prev_company_name]["collected_total"],2)}</b></td><td colspan=1 style="border: 1px solid black"><b>{round(company_totals[prev_company_name]["balance"],2)}</b></td></tr>'
#                     data += f'<tr><td colspan=12 style="border: 1px solid black; background-color:#EBF4FA;font-size:8px" ><b style=color:"red">{company_name}</b></td></tr>'
#                     prev_company_name = company_name
#                     company_set.add(company_name)

#                 net = (i.total - i.discount_amount)
#                 net_int = int(net)
#                 net_amount = (i.total - net)
#                 convert_float = int(net_amount)
#                 collected = (net - i.outstanding_amount)
#                 collected_int = int(collected)

#                 data += f'<tr style=font-size:8px width:100 height:50><td colspan=1 style="border: 1px solid black">{j}</td>'
#                 data += f'<td colspan=2 style="border: 1px solid black" nowrap>{i.posting_date.strftime("%d-%m-%Y")}</td>'
#                 data += f'<td colspan=1 style="border: 1px solid black" nowrap>{i.name}</td>'
#                 data += f'<td colspan=1 style="border: 1px solid black">{i.customer}</td>'
#                 data += f'<td colspan=1 style="border: 1px solid black">{i.po_no}</td>'
#                 data += f'<td colspan=1 style="border: 1px solid black">{i.total}</td>'
#                 data += f'<td colspan=1 style="border: 1px solid black">{i.discount_amount}</td>'
#                 data += f'<td colspan=1 style="border: 1px solid black">{convert_float}</td>'
#                 data += f'<td style="border: 1px solid black">{net_int}</td>'
#                 data += f'<td colspan=1 style="border: 1px solid black">{collected_int}</td>'
#                 data += f'<td colspan=1 style="border: 1px solid black">{i.outstanding_amount}</td></tr>'

	
#                 if company_name not in company_totals:
#                     company_totals[company_name] = {
#                         "total": i.total,
#                         "discount_amount": i.discount_amount,
#                         "net_total": net_int,
#                         "convert_float": convert_float,
#                         "collected_total": collected_int,
#                         "balance": i.outstanding_amount
#                     }
#                 else:
#                     company_totals[company_name]["total"] += i.total
#                     company_totals[company_name]["discount_amount"] += i.discount_amount
#                     company_totals[company_name]["net_total"] += net_int
#                     company_totals[company_name]["convert_float"] += convert_float
#                     company_totals[company_name]["collected_total"] += collected_int
#                     company_totals[company_name]["balance"] += i.outstanding_amount

				
#                 if s.name not in salesperson_totals:
#                     salesperson_totals[s.name] = {
#                         "total": i.total,
#                         "discount_amount": i.discount_amount,
#                         "net_total": net_int,
#                         "convert_float": convert_float,
#                         "collected_total": collected_int,
#                         "balance": i.outstanding_amount
#                     }
#                 else:
#                     salesperson_totals[s.name]["total"] += i.total
#                     salesperson_totals[s.name]["discount_amount"] += i.discount_amount
#                     salesperson_totals[s.name]["net_total"] += net_int
#                     salesperson_totals[s.name]["convert_float"] += convert_float
#                     salesperson_totals[s.name]["collected_total"] += collected_int
#                     salesperson_totals[s.name]["balance"] += i.outstanding_amount

#                 j += 1

#             if prev_company_name:
				
#                 data += f'<tr style = font-size:8px><td colspan=6 style="border: 1px solid black"><b>Total for {prev_company_name}</b></td><td colspan=1 style="border: 1px solid black"><b>{round(company_totals[prev_company_name]["total"],2)}</b></td><td colspan=1 style="border: 1px solid black"><b>{round(company_totals[prev_company_name]["discount_amount"],2)}</b></td><td colspan=1 style="border: 1px solid black"><b>{round(company_totals[prev_company_name]["convert_float"],2)}</b></td><td style="border: 1px solid black"><b>{round(company_totals[prev_company_name]["net_total"],2)}</b></td><td colspan=1 style="border: 1px solid black"><b>{round(company_totals[prev_company_name]["collected_total"],2)}</b></td><td colspan=1 style="border: 1px solid black"><b>{round(company_totals[prev_company_name]["balance"],2)}</b></td></tr>'

#             prev_company_name = None

			
#             if s.name in salesperson_totals:
#                 data += f'<tr style= font-size:8px><td colspan=6 style="border: 1px solid black"><b>Total for {s.name}</b></td><td colspan=1 style="border: 1px solid black"><b>{salesperson_totals[s.name]["total"]}</b></td><td colspan=1 style="border: 1px solid black"><b>{salesperson_totals[s.name]["discount_amount"]}</b></td><td colspan=1 style="border: 1px solid black"><b>{salesperson_totals[s.name]["convert_float"]}</b></td><td style="border: 1px solid black"><b>{salesperson_totals[s.name]["net_total"]}</b></td><td colspan=1 style="border: 1px solid black"><b>{salesperson_totals[s.name]["collected_total"]}</b></td><td colspan=1 style="border: 1px solid black"><b>{salesperson_totals[s.name]["balance"]}</b></td></tr>'
  
#     elif not doc.company_multiselect and sales_person_user:
#         company = frappe.get_all("Company", fields=["name"])
#         data += f'<tr><td colspan="12" style="border: 1px solid black; background-color:#EBF4FA"><b style="color:black;font-size:8px">{sales_person_user}</b></td></tr>'
#         all_companies_total = {
#             "total": 0,
#             "discount_amount": 0,
#             "convert_float": 0,
#             "net_total": 0,
#             "collected_total": 0,
#             "balance": 0
#         }

#         for c in company:
#             sales_invoice = frappe.get_all("Sales Invoice", {"sales_person_user": sales_person_user,'posting_date': ('between', (from_date, to_date)), 'company': c.name,'docstatus':("not in",[0,2])}, ["name", "customer", "total", "discount_amount", "posting_date", "outstanding_amount", "po_no"],order_by="posting_date")
#             if sales_invoice:
#                 data += f'<tr><td colspan="12" style="border: 1px solid black; background-color:#EBF4FA"><b style="color:black;font-size:8px" height:25%>{c.name}</b></td></tr>'
#                 j = 1
#                 company_totals = {}
				
#                 for i in sales_invoice:
#                     net = (i.total - i.discount_amount)
#                     net_int = int(net)
#                     net_amount = (i.total - net)
#                     convert_float = int(net_amount)
#                     collected = (net - i.outstanding_amount)
#                     collected_int = int(collected)
#                     data += f'<tr style = font-size:8px><td colspan="1" style="border: 1px solid black;" height:25%>{j}</td>'
#                     data += f'<td colspan="2" style="border: 1px solid black;" height:25% nowrap>{i.posting_date.strftime("%d-%m-%Y")}</td>'
#                     data += f'<td colspan="1" style="border: 1px solid black;" height:25%>{i.name}</td>'
#                     data += f'<td colspan="1" style="border: 1px solid black;" height:25%>{i.customer}</td>'
#                     data += f'<td colspan="1" style="border: 1px solid black;" height:25%>{i.po_no}</td>'
#                     data += f'<td colspan="1" style="border: 1px solid black;" height:25%>{i.total}</td>'
#                     data += f'<td colspan="1" style="border: 1px solid black;" height:25%>{i.discount_amount}</td>'
#                     data += f'<td colspan="1" style="border: 1px solid black;" height:25%>{convert_float}</td>'
#                     data += f'<td style="border: 1px solid black;">{net_int}</td>'
#                     data += f'<td colspan="1" style="border: 1px solid black;">{collected_int}</td>'
#                     data += f'<td colspan="1" style="border: 1px solid black;">{i.outstanding_amount}</td></tr>'
#                     if c.name not in company_totals:
#                         company_totals[c.name] = {
#                             "total": i.total,
#                             "discount_amount": i.discount_amount,
#                             "net_total": net_int,
#                             "convert_float": convert_float,
#                             "collected_total": collected_int,
#                             "balance": i.outstanding_amount
#                         }
#                     else:
#                         company_totals[c.name]["total"] += i.total
#                         company_totals[c.name]["discount_amount"] += i.discount_amount
#                         company_totals[c.name]["net_total"] += net_int
#                         company_totals[c.name]['convert_float'] += convert_float
#                         company_totals[c.name]["collected_total"] += collected_int
#                         company_totals[c.name]['balance'] += i.outstanding_amount
#                     j += 1
#                 data += f'<tr style = font-size:8px>'
#                 data += f'<td colspan="6" style="border: 1px solid black;"><b>Total for {c.name}</b></td>'
#                 data += f'<td colspan="1" style="border: 1px solid black;"><b>{round(company_totals[c.name].get("total", 0),2)}</b></td>'
#                 data += f'<td colspan="1" style="border: 1px solid black;"><b>{round(company_totals[c.name].get("discount_amount", 0),2)}</b></td>'
#                 data += f'<td colspan="1" style="border: 1px solid black;"><b>{round(company_totals[c.name].get("convert_float", 0),2)}</b></td>'
#                 data += f'<td style="border: 1px solid black;"><b>{round(company_totals[c.name].get("net_total", 0),2)}</b></td>'
#                 data += f'<td colspan="1" style="border: 1px solid black;"><b>{round(company_totals[c.name].get("collected_total", 0),2)}</b></td>'
#                 data += f'<td colspan="1" style="border: 1px solid black;"><b>{round(company_totals[c.name].get("balance", 0),2)}</b></td></tr>'

#                 all_companies_total["total"] += company_totals[c.name].get("total", 0)
#                 all_companies_total["discount_amount"] += company_totals[c.name].get("discount_amount", 0)
#                 all_companies_total["convert_float"] += company_totals[c.name].get("convert_float", 0)
#                 all_companies_total["net_total"] += company_totals[c.name].get("net_total", 0)
#                 all_companies_total["collected_total"] += company_totals[c.name].get("collected_total", 0)
#                 all_companies_total["balance"] += company_totals[c.name].get("balance", 0)
#         data += f'<tr style =font-size:8px>'
#         data += f'<td colspan="6" style="border: 1px solid black;"><b>Total for {sales_person_user}</b></td>'
#         data += f'<td colspan="1" style="border: 1px solid black;"><b>{round(all_companies_total["total"],2)}</b></td>'
#         data += f'<td colspan="1" style="border: 1px solid black;"><b>{round(all_companies_total["discount_amount"],2)}</b></td>'
#         data += f'<td colspan="1" style="border: 1px solid black;"><b>{round(all_companies_total["convert_float"],2)}</b></td>'
#         data += f'<td style="border: 1px solid black;"><b>{round(all_companies_total["net_total"],2)}</b></td>'
#         data += f'<td colspan="1" style="border: 1px solid black;"><b>{round(all_companies_total["collected_total"],2)}</b></td>'
#         data += f'<td colspan="1" style="border: 1px solid black;"><b>{round(all_companies_total["balance"],2)}</b></td>'
#         data += '</tr>'

		
#     elif doc.company_multiselect and not sales_person_user:
#         for c in doc.company_multiselect:
#             company_name_printed = False
#             company_totals = {
#                 "total": 0,
#                 "discount_amount": 0,
#                 "convert_float": 0,
#                 "net_total": 0,
#                 "collected_total": 0,
#                 "balance": 0
#             }
			
#             sales_person = frappe.get_all("Sales Person", fields=["name"])
#             for s in sales_person:
#                 sales_invoice = frappe.get_all("Sales Invoice", {"sales_person_user": s.name,'posting_date': ('between', (from_date, to_date)), 'company': c.company,'docstatus':("not in",[0,2])}, ["name", "customer", "total", "discount_amount", "posting_date", "outstanding_amount", "po_no"],order_by="posting_date")
#                 j = 1

#                 if sales_invoice:
#                     data += '<tr style =font-size:8px><td colspan="12" style="border: 1px solid black; background-color:#EBF4FA"><b style="color:black">%s</b></td></tr>' % (s.name)
#                     if not company_name_printed:
#                         data += '<tr style =font-size:8px><td colspan="12" style="border: 1px solid black; background-color:#EBF4FA"><b style="color:black">%s</b></td></tr>' % (c.company)
#                         company_name_printed = True

#                     sales_person_totals = {
#                         "total": 0,
#                         "discount_amount": 0,
#                         "convert_float": 0,
#                         "net_total": 0,
#                         "collected_total": 0,
#                         "balance": 0
#                     }

#                     for i in sales_invoice:
#                         net = (i.total - i.discount_amount)
#                         net_int = int(net)
#                         net_amount = (i.total - net)
#                         convert_float = int(net_amount)
#                         collected = (net - i.outstanding_amount)
#                         collected_int = int(collected)
#                         data += '<tr style =font-size:8px><td colspan="1" style="border: 1px solid black">%s</td><td colspan="2" style="border: 1px solid black" nowrap>%s</td><td colspan="1" style="border: 1px solid black">%s</td><td colspan="1" style="border: 1px solid black">%s</td><td colspan="1" style="border: 1px solid black">%s</td><td colspan="1" style="border: 1px solid black">%s</td><td colspan="1" style="border: 1px solid black">%s</td><td colspan="1" style="border: 1px solid black">%s</td><td style="border: 1px solid black">%s</td><td colspan="1" style="border: 1px solid black">%s</td><td colspan="1" style="border: 1px solid black">%s</td></tr>' % (j, i.posting_date.strftime("%d-%m-%Y"), i.name, i.customer, i.po_no, i.total, i.discount_amount, convert_float, net_int, collected_int, i.outstanding_amount)
#                         j += 1
#                         sales_person_totals["total"] += i.total
#                         sales_person_totals["discount_amount"] += i.discount_amount
#                         sales_person_totals["convert_float"] += convert_float
#                         sales_person_totals["net_total"] += net_int
#                         sales_person_totals["collected_total"] += collected_int
#                         sales_person_totals["balance"] += i.outstanding_amount
						
#                         company_totals["total"] += i.total
#                         company_totals["discount_amount"] += i.discount_amount
#                         company_totals["convert_float"] += convert_float
#                         company_totals["net_total"] += net_int
#                         company_totals["collected_total"] += collected_int
#                         company_totals["balance"] += i.outstanding_amount

#                     data += '<tr style =font-size:8px><td colspan="6" style="border: 1px solid black"><b>Total for %s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td><td style="border: 1px solid black"><b>%s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td></tr>' % (s.name, round(sales_person_totals["total"],2), round(sales_person_totals["discount_amount"],2), round(sales_person_totals["convert_float"],2), round(sales_person_totals["net_total"],2), round(sales_person_totals["collected_total"],2), round(sales_person_totals["balance"],2))
		
#             data += '<tr style =font-size:8px><td colspan="6" style="border: 1px solid black"><b>Total for %s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td><td style="border: 1px solid black"><b>%s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td></tr>' % (c.company, round(company_totals["total"],2), round(company_totals["discount_amount"],2), round(company_totals["convert_float"],2), round(company_totals["net_total"],2), round(company_totals["collected_total"],2), round(company_totals["balance"],2))

#     else:
#         salesperson_totals = {
#             "total": 0,
#             "discount_amount": 0,
#             "convert_float": 0,
#             "net_total": 0,
#             "collected_total": 0,
#             "balance": 0
#         }
		
#         for c in doc.company_multiselect:
#             company_name_printed = False
#             company_totals = {
#                 "total": 0,
#                 "discount_amount": 0,
#                 "convert_float": 0,
#                 "net_total": 0,
#                 "collected_total": 0,
#                 "balance": 0
#             }
			
#             sales_invoice = frappe.get_all("Sales Invoice", {"sales_person_user": sales_person_user, 'posting_date': ('between', (from_date, to_date)), 'company': c.company,'docstatus':("not in",[0,2])}, ["name", "customer", "total", "discount_amount", "posting_date", "outstanding_amount", "po_no"],order_by="posting_date")
#             j = 1
			
#             if sales_invoice:
#                 if not company_name_printed:
#                     data += '<tr style =font-size:8px><td colspan="12" style="border: 1px solid black; background-color:#EBF4FA"><b style="color:black">%s</b></td></tr>' % (c.company)
#                     company_name_printed = True
				
#                 data += '<tr style =font-size:8px><td colspan="12" style="border: 1px solid black; background-color:#EBF4FA"><b style="color:black">%s</b></td></tr>' % (sales_person_user)
				
#                 for i in sales_invoice:
#                     net = (i.total - i.discount_amount)
#                     net_int = int(net)
#                     net_amount = (i.total - net)
#                     convert_float = int(net_amount)
#                     collected = (net - i.outstanding_amount)
#                     collected_int = int(collected)
#                     data += '<tr style =font-size:8px><td colspan="1" style="border: 1px solid black">%s</td><td colspan="2" style="border: 1px solid black" nowrap>%s</td><td colspan="1" style="border: 1px solid black" nowrap>%s</td><td colspan="1" style="border: 1px solid black">%s</td><td colspan="1" style="border: 1px solid black">%s</td><td colspan="1" style="border: 1px solid black">%s</td><td colspan="1" style="border: 1px solid black">%s</td><td colspan="1" style="border: 1px solid black">%s</td><td style="border: 1px solid black">%s</td><td colspan="1" style="border: 1px solid black">%s</td><td colspan="1" style="border: 1px solid black">%s</td></tr>' % (j, i.posting_date.strftime("%d-%m-%Y"), i.name, i.customer, i.po_no, i.total, i.discount_amount, convert_float, net_int, collected_int, i.outstanding_amount)
#                     j += 1

#                     company_totals["total"] += i.total
#                     company_totals["discount_amount"] += i.discount_amount
#                     company_totals["convert_float"] += convert_float
#                     company_totals["net_total"] += net_int
#                     company_totals["collected_total"] += collected_int
#                     company_totals["balance"] += i.outstanding_amount
				
#                 data += '<tr style =font-size:8px><td colspan="6" style="border: 1px solid black"><b>Total for %s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td><td style="border: 1px solid black"><b>%s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td></tr>' % (c.company, round(company_totals["total"],2), round(company_totals["discount_amount"],2), round(company_totals["convert_float"],2), round(company_totals["net_total"],2), round(company_totals["collected_total"],2), round(company_totals["balance"],2))
				
#                 salesperson_totals["total"] += company_totals["total"]
#                 salesperson_totals["discount_amount"] += company_totals["discount_amount"]
#                 salesperson_totals["convert_float"] += company_totals["convert_float"]
#                 salesperson_totals["net_total"] += company_totals["net_total"]
#                 salesperson_totals["collected_total"] += company_totals["collected_total"]
#                 salesperson_totals["balance"] += company_totals["balance"]
		
#         data += '<tr style =font-size:8px><td colspan="6" style="border: 1px solid black"><b>Total for %s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td><td style="border: 1px solid black"><b>%s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td></tr>' % (sales_person_user, round(salesperson_totals["total"],2), round(salesperson_totals["discount_amount"],2), round(salesperson_totals["convert_float"],2), round(salesperson_totals["net_total"],2), round(salesperson_totals["collected_total"],2), round(salesperson_totals["balance"],2))
		
#     data += '</table>'
#     return data

#The below method is using print format in Sales Invoice
@frappe.whitelist()
def get_items(doc):
	data = "<table  width = 100% border= 1px solid black><tr style= background-color:#e35310; ><td colspan = 1  width = 10 ><b style = color:white ; text-align:center;>S.No</b></td><td colspan = 1 width=10 style = color:white><b style = color:white ; text-align:center;>Item Code</b></td><td colspan = 1 width=10 style = color:white ><b style = color:white ; text-align:center;>Description</b></td><td colspan = 1 width=10 style = color:white><b style = color:white ; text-align:center;>Quantity</b></td><td colspan = 1 width= 10 style = color:white><b style = color:white ; text-align:center;>Unit</b><td colspan = 1 width=10 style = color:white><b style = color:white ; text-align:center;>Rate</b></td><td colspan = 1 width=10 style = color:white><b style = color:white ; text-align:center;>Amount</b></td></tr>"
	j=1
	for i in doc.items:
		list_item = []
		designation = frappe.db.get_list("Designation",['name'])
		for k in designation:
			list_item.append(k.name)
		if i.item_code not in list_item:
		# for j in designation:
			# if i.item_code != j.name:
			data += '<tr><td colspan = 1 style="border: 1px solid black">%s</td><td colspan = 1 style="border: 1px solid black">%s</td><td colspan = 1 style="border: 1px solid black">%s</td><td colspan = 1 style="border: 1px solid black">%s</td><td colspan=1 style="border:1px solid black">%s</td><td colspan = 1 style="border: 1px solid black">%s</td><td colspan = 1 style="border: 1px solid black">%s</td></tr>'%(j,i.item_code,i.description,i.qty,i.uom,i.rate,i.amount)
			j+=1
	data += '</table>'
	return data


@frappe.whitelist()
def get_items_list():
	list_item = []
	designation = frappe.db.get_list("Designation",['name'])
	for i in designation:
		list_item.append(i.name)
	print(list_item)


#While enter the no_of_days_absent_in_previous_month in additional salary to calculate the below condition and return the amount
@frappe.whitelist()
def calculate_abs_detec(days,pd,gross):
	import datetime
	specific_date = datetime.datetime.strptime(pd, "%Y-%m-%d").date()
	# first_day_current_month = datetime.date(specific_date.year, specific_date.month, 1)
	# first_day_previous_month = first_day_current_month - datetime.timedelta(days=1)
	# previous_month_num_days = first_day_previous_month.day
	per_day_amount = int(gross) / 30
	tot_amount = float(per_day_amount) * float(days)
	return int(tot_amount)

#While enter the no_of_days_absent_in_previous_month in additional salary to calculate the below condition and return the amount
@frappe.whitelist()
def calculate_add(days,pd,gross):
	import datetime
	import calendar
	# specific_date = datetime.datetime.strptime(pd, "%Y-%m-%d").date()
	# first_day_current_month = datetime.date(specific_date.year, specific_date.month, 1)
	# last_day = calendar.monthrange(first_day_current_month.year, first_day_current_month.month)[1]
	per_day_amount = int(gross) / int(30)
	tot_amount = float(per_day_amount) * float(days)
	return int(tot_amount)

#While enter the Hot Hour in additional salary to calculate the below condition and return the amount
@frappe.whitelist()
def calculate_hot(hours,gross):
	amount = (int(gross)/(240)) * int(hours) * 1.5
	return round(amount)

#While enter the Not Hour in additional salary to calculate the below condition and return the amount
@frappe.whitelist()
def calculate_not(hours,gross):
	amount = (int(gross)/(240)) * float(hours) * 1.25
	return round(amount)

@frappe.whitelist()
def take_po():
	po = frappe.db.get_list("Purchase Order",{"supplier":"AFFIX SCAFFOLDING WLL.","docstatus":1},['name'])
	for i in po:
		pi_item = frappe.get_value("Purchase Invoice Item",{"purchase_order":i.name},['parent'])
		if not pi_item:
			print(i.name)

import frappe
from datetime import datetime
from dateutil import relativedelta

#Check the below condition is validation of loan application
@frappe.whitelist()
def check_loan_amount(doc,method):
	# Get the employee's basic amount
	basic_amount = frappe.db.get_value('Employee', doc.applicant, 'basic')
	print(basic_amount)
	employees = frappe.get_all('Employee', {'status': 'Active'}, ['name', 'employee_name', 'date_of_joining'])
	for emp in employees:
		date_2 = datetime.now()
		# Get the interval between two dates
		diff = relativedelta.relativedelta(date_2, emp.date_of_joining)
		yos = str(diff.years) + ' years, ' + str(diff.months) + ' months and ' + str(diff.days) + ' days'
		exp_years = diff.years
		exp_month = diff.months
		exp_days = diff.days
		total_days = (exp_years * 365) + (exp_month * 30) + exp_days
		print(total_days)
		per_day_basic = (float(basic_amount) / 30 * 21) / 365
		total_gratuity = per_day_basic * total_days
		gratuity = (total_gratuity * 80 /100)
	# Convert loan_amount and interest_rate to appropriate numeric types
	loan_amount = float(doc.loan_amount)
	interest_rate = int(doc.rate_of_interest)

	# Calculate the loan amount plus the calculated interest
	calculated_amount = loan_amount

	# Check if the calculated amount exceeds the basic amount
	
	if calculated_amount > gratuity :
		if doc.loan_product in ["Loan Against Gratuity - Marazeem","Loan Against Gratuity - AISTCC","Personals Loan - AISTCC"] :
			frappe.throw(_('Loan amount must be Greater than Employee Gratuity'))

	elif doc.loan_product in ["Loan Against Leave  Settlement - Marazeem","Loan Against Leave  Settlement - AISTCC","Loan Against Leave  Settlement - kingFisher" ]:
		gross = frappe.db.get_value('Employee', doc.applicant, 'gross_salary')
		loan_amount = float(doc.loan_amount)
		calculated_amount = loan_amount
		if calculated_amount > gross :
			frappe.throw(_('Loan amount must be Greater than Employee Gross'))   

	   


@frappe.whitelist()
def pen_consolidated_req(name):
	con = frappe.db.sql(""" select * from `tabPurchase Invoice List` where parent = '%s' """%(name),as_dict=True)
	def calculate_outstanding_payment_difference(child_table):
		amounts_by_company = defaultdict(float)
		for row in child_table:
			company = row.get("company", "")
			outstanding_amount = row.get("outstanding_amount", 0.0)
			payment_amount = row.get("payment_amount", 0.0)
			amounts_by_company[company] += (outstanding_amount - payment_amount)
		result = []
		data = '<table width="100%"><tr><td style="color:white;background-color:#e35310;padding:1px;font-weight:bold;border:1px solid black">Company</td><td style="font-weight:bold;color:white;background-color:#e35310;padding:1px;border:1px solid black">Amount</td></tr>'
		for company, difference in amounts_by_company.items():
			result.append({"company": company, "value": difference})

		for item in result:
			data += '<tr><td style="padding:1px;border:1px solid black">{}</td><td style="padding:1px;border:1px solid black">{}</td></tr>'.format(item["company"], item["value"])

		data += '</table>'

		return data

	child_table = con
	amounts_difference_by_company = calculate_outstanding_payment_difference(child_table)
	return amounts_difference_by_company


#Get the below html table and show it Consolidated payment request
@frappe.whitelist()
def pay_consolidated_req(name):
	con = frappe.db.sql("""SELECT * FROM `tabPurchase Invoice List` WHERE parent = '%s' """ % (name), as_dict=True)

	def calculate_outstanding_payment_difference(child_table):
		amounts_by_company = defaultdict(float)
		total = 0.0

		for row in child_table:
			company = row.get("company", "")
			outstanding_amount = row.get("outstanding_amount", 0.0)
			payment_amount = row.get("payment_amount", 0.0)
			amounts_by_company[company] += payment_amount
			total += payment_amount

		result = []

		for company, difference in amounts_by_company.items():
			result.append({"company": company, "value": difference})

		# Append the total row to the result
		result.append({"company": "Total", "value": total})

		data = ''
		data += '<table width="61%"><tr><td style="color:white;background-color:#e35310;padding:1px;font-weight:bold;border: 1px solid black">Company</td><td style="font-weight:bold;color:white;background-color:#e35310;padding:1px;border: 1px solid black">Amount</td></tr>'

		for i in result:
			formatted_value = '{:,.2f}'.format(i["value"])
			data += '<tr><td style="padding:1px;border: 1px solid black">%s</td><td style="padding:1px;border: 1px solid black">%s</td></tr>' % (i["company"],formatted_value)

		data += '</table>'
		return data

	child_table = con
	amounts_difference_by_company = calculate_outstanding_payment_difference(child_table)
	return amounts_difference_by_company


#Check the same id in 2 row for consolidated payment request 
@frappe.whitelist()
def check_for_duplicate_payment(name,doc_type,doc_name):
	li = []
	dat = frappe.get_value(doc_type,name,["grand_total"])
	pos = frappe.db.sql("""select sum(`tabPurchase Invoice List`.payment_amount) as payment_amount,`tabPurchase Invoice List`.parent as name from `tabConsolidated Payment Request`
	left join `tabPurchase Invoice List` on `tabConsolidated Payment Request`.name = `tabPurchase Invoice List`.parent
	where `tabPurchase Invoice List`.name1 = '%s' """%(name),as_dict=True)[0]
	po = frappe.db.sql("""select `tabPurchase Invoice List`.parent as name from `tabConsolidated Payment Request`
	left join `tabPurchase Invoice List` on `tabConsolidated Payment Request`.name = `tabPurchase Invoice List`.parent
	where `tabPurchase Invoice List`.name1 = '%s' and `tabConsolidated Payment Request`.docstatus = 0 """%(name),as_dict=True)
	for i in po:
		if i.name != doc_name:
			li.append(i.name)
	if not pos["payment_amount"]:
		pos["payment_amount"] = 0
	return pos["payment_amount"] or 0,li,dat or 0


#Get the payroll date and set it additional salary
from frappe.utils import cstr, cint, getdate,get_first_day, get_last_day, today
@frappe.whitelist()
def get_payroll_date():
	date = datetime.strptime(today(), '%Y-%m-%d').date()
	if date.day > 15:
		payroll_date = add_days(get_first_day(add_months(date,0)),0)
		return payroll_date
	else:
		payroll_date = add_days(get_first_day(add_months(date,-1)),0)
		return payroll_date


@frappe.whitelist()
def le():
	sale = frappe.db.sql("""SELECT sales_person_user,COUNT(*) AS document_count FROM (SELECT sales_person_user FROM `tabQuotation` WHERE sales_person_user = 'Rinshad Chithuparambu Mohammedunni' and status != 'Cancelled' UNION ALL SELECT sales_person_user FROM `tabSales Order` WHERE sales_person_user = 'Rinshad Chithuparambu Mohammedunni'and status != 'Cancelled') AS combined_sales GROUP BY sales_person_user;""",as_dict=True)
	print(sale)

@frappe.whitelist()
def update_ctc():
	employee_id= frappe.db.sql("""select * from `tabEmployee` where status = 'Active' """,as_dict=1)
	for t in employee_id:
		print(t.basic)
		ctc = t.basic + t.hra + t._other_allowance + t.accommodation + t.transportation + t.mobile_allowance + t.air_ticket_allowance_ + t.medical_renewal + t.visa_cost_ + t.qid_cost + t.gratuity + t.medical_allowance_ + t.leave_salary + t.compensationemployee_insurence
		print(ctc)
		print(t.name)
		t.ctc = ctc
		emp = frappe.db.sql("""update `tabEmployee`set ctc = '%s' where status = 'Active' """%(ctc),as_dict=1)


# @frappe.whitelist()
# def return_total_amt(from_date,to_date,account):
#     acct = account.split(' - ')
#     acc=''
#     if len(acct) == 2:
#         acc = acct[0]
#     if len(acct) == 3:
#         acc = f"{acct[0]} - {acct[1]}"
#     if len(acct) == 4:
#         acc = f"{acct[1]} - {acct[2]}"
#     ac = '%'+acc+'%'
#     data = '<table  border= 1px solid black width = 100%>'
#     data += '<tr style = "background-color:#D9E2ED"><td colspan =1><b></b></td><td colspan =1 style = "text-align:center"><b>Opening</b></td><td colspan =2 style = "text-align:center"><b>Movement</b></td><td colspan =1 style = "text-align:center"><b>Closing</b></td></tr>'

#     data += '<tr style = "background-color:#e35310;color:white"><td  style = "text-align:center;font-weight:bold;color:white">Company</td><td  style = "text-align:center;font-weight:bold;color:white">Balance</td><td  style = "text-align:center;font-weight:bold;color:white">Debit</td><td  style = "text-align:center;font-weight:bold;color:white">Credit</td><td  style = "text-align:center;font-weight:bold;color:white">Balance</td></tr>'
#     op_credit = 0
#     op_debit = 0
#     total_op_debit = 0
#     total_op_credit = 0
#     t_c_credit = 0
#     t_p_credit = 0
#     t_c_debit = 0
#     t_p_debit = 0
	
#     li = []
#     company = frappe.db.sql(""" select name from `tabCompany` where is_group = 1""",as_dict=1)
#     for com in company:
#         li.append(com.name)
#         comp = frappe.db.get_list("Company",{"parent_company":com.name},['name'])
#         for j in comp:
#             li.append(j.name)
#     for c in li:
#         gle = frappe.db.sql("""select account, sum(debit) as opening_debit, sum(credit) as opening_credit from `tabGL Entry` where company = '%s'	and (posting_date < '%s' or (ifnull(is_opening, 'No') = 'Yes' and posting_date > '%s')) and account like '%s' and is_cancelled = 0  """%(c,from_date,to_date,ac),as_dict=True)
#         for g in gle:
#             if not g.opening_debit:
#                 g.opening_debit = 0
#             if not g.opening_credit:
#                 g.opening_credit = 0
#             t_p_debit += g.opening_debit
#             t_p_credit += g.opening_credit
#             balance_op = t_p_debit - t_p_credit
#             data += '<tr><td>%s</td><td style = text-align:right >%s</td>'%(c,fmt_money(g.opening_debit - g.opening_credit))
#             sq = frappe.db.sql(""" select company,sum(debit_in_account_currency) as debit,sum(credit_in_account_currency) as credit from `tabGL Entry` where company = '%s' and account like '%s' and posting_date between '%s' and '%s' and is_opening = 'No' and is_cancelled = 0 """%(c,ac,from_date,to_date),as_dict=True)
#             for i in sq:
#                 if not i.credit:
#                     i.credit = 0
#                 if not i.debit:
#                     i.debit = 0
#                 op_credit = g.opening_credit + i.credit
#                 op_debit = g.opening_debit + i.debit
#                 total_op_debit += i.debit
#                 total_op_credit += i.credit
#                 t_c_credit += op_credit
#                 t_c_debit += op_debit
#                 balance_cl = t_c_debit - t_c_credit
#                 balance_move=total_op_debit-total_op_credit
#                 data += '<td style = text-align:right >%s</td><td style = text-align:right >%s</td><td style = text-align:right >%s</td></tr>'%(fmt_money(i.debit),fmt_money(i.credit),fmt_money(op_debit - op_credit))
#     data += '<tr style = "text-align:right;font-weight:bold"><td style = "text-align:center;font-weight:bold">Total</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>'%(fmt_money(balance_op),fmt_money(total_op_debit),fmt_money(total_op_credit),fmt_money(balance_cl))
#     # data += '<tr style = "text-align:right;font-weight:bold"><td style = "text-align:center;font-weight:bold">Balance</td><td colspan =3>%s</td><td colspan =3></td><td colspan=3>%s</td></tr>'%(fmt_money(balance_op),fmt_money(balance_cl))
#     data += '</table>'
#     return data

# @frappe.whitelist()
# def return_total_amt_consolidate(from_date,to_date,account):
#     acct = account.split(' - ')
#     acc=''
#     if len(acct) == 2:
#         acc = acct[0]
#     if len(acct) == 3:
#         acc = f"{acct[0]} - {acct[1]}"
#     if len(acct) == 4:
#         acc = f"{acct[1]} - {acct[2]}"
#     ac = '%'+acc+'%'
#     data = '<table  border= 1px solid black width = 100%>'
#     # data += '<tr style = "background-color:#D9E2ED"><td colspan =1><b></b></td><td colspan =1 style = "text-align:center"><b>Opening</b></td><td colspan =2 style = "text-align:center"><b>Movement</b></td><td colspan =1 style = "text-align:center"><b>Closing</b></td></tr>'

#     data += '<tr style = "background-color:#e35310;color:white"><td  style = "text-align:center;font-weight:bold;color:white">Company</td><td  style = "text-align:center;font-weight:bold;color:white">Opening Debit</td><td  style = "text-align:center;font-weight:bold;color:white">Opening Credit</td><td  style = "text-align:center;font-weight:bold;color:white">Movement Debit</td><td  style = "text-align:center;font-weight:bold;color:white">Movement Credit</td><td  style = "text-align:center;font-weight:bold;color:white">Closing Debit</td><td  style = "text-align:center;font-weight:bold;color:white">Closing Credit</td></tr>'
#     op_credit = 0
#     op_debit = 0
#     total_op_debit = 0
#     total_op_credit = 0
#     t_c_credit = 0
#     t_p_credit = 0
#     t_c_debit = 0
#     t_p_debit = 0
	
#     li = []
#     company = frappe.db.sql(""" select name from `tabCompany` where is_group = 1""",as_dict=1)
#     for com in company:
#         li.append(com.name)
#         comp = frappe.db.get_list("Company",{"parent_company":com.name},['name'])
#         for j in comp:
#             li.append(j.name)
#     for c in li:
#         gle = frappe.db.sql("""select account, sum(debit) as opening_debit, sum(credit) as opening_credit from `tabGL Entry` where company = '%s'	and (posting_date < '%s' or (ifnull(is_opening, 'No') = 'Yes' and posting_date > '%s')) and account like '%s' and is_cancelled = 0  """%(c,from_date,to_date,ac),as_dict=True)
#         for g in gle:
#             if not g.opening_debit:
#                 g.opening_debit = 0
#             if not g.opening_credit:
#                 g.opening_credit = 0
#             t_p_debit += g.opening_debit
#             t_p_credit += g.opening_credit
#             balance_op = t_p_debit - t_p_credit
#             data += '<tr><td>%s</td><td style = text-align:right >%s</td><td style = text-align:right >%s</td>'%(c,fmt_money(g.opening_debit) ,fmt_money(g.opening_credit))
#             sq = frappe.db.sql(""" select company,sum(debit_in_account_currency) as debit,sum(credit_in_account_currency) as credit from `tabGL Entry` where company = '%s' and account like '%s' and posting_date between '%s' and '%s' and is_opening = 'No' and is_cancelled = 0 """%(c,ac,from_date,to_date),as_dict=True)
#             for i in sq:
#                 if not i.credit:
#                     i.credit = 0
#                 if not i.debit:
#                     i.debit = 0
#                 op_credit = g.opening_credit + i.credit
#                 op_debit = g.opening_debit + i.debit
#                 total_op_debit += i.debit
#                 total_op_credit += i.credit
#                 t_c_credit += op_credit
#                 t_c_debit += op_debit
#                 balance_cl = t_c_debit - t_c_credit
#                 balance_move=total_op_debit-total_op_credit
#                 data += '<td style = text-align:right >%s</td><td style = text-align:right >%s</td><td style = text-align:right >%s</td><td style = text-align:right >%s</td></tr>'%(fmt_money(i.debit),fmt_money(i.credit),fmt_money(op_debit),fmt_money(op_credit))
#     data += '<tr style = "text-align:right;font-weight:bold"><td style = "text-align:center;font-weight:bold">Total</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>'%(fmt_money(t_p_debit),fmt_money(t_p_credit),fmt_money(total_op_debit),fmt_money(total_op_credit),fmt_money(t_c_debit),fmt_money(t_c_credit))
#     # data += '<tr style = "text-align:right;font-weight:bold"><td style = "text-align:center;font-weight:bold">Balance</td><td colspan =3>%s</td><td colspan =3></td><td colspan=3>%s</td></tr>'%(fmt_money(balance_op),fmt_money(balance_cl))
#     data += '</table>'
#     return data


# @frappe.whitelist()
# def return_account_total(from_date,to_date,account):
#     data = '<table border="1" style="border-collapse: collapse; width: 100%;">'
#     data += '<tr style="background-color: #D9E2ED;"><td colspan="1"><b></b></td><td colspan="1"><b></b></td><td colspan="3" style="text-align:center;"><b>Opening</b></td><td colspan="3" style="text-align:center;"><b>Movement</b></td><td colspan="3" style="text-align:center;"><b>Closing</b></td></tr>'
#     data += '<tr style="background-color: #e35310; color: white;"><td style="text-align:center; font-weight:bold; color:white;">Party</td><td style="text-align:center; font-weight:bold; color:white;">Party Name</td><td style="text-align:center; font-weight:bold; color:white;">Debit</td><td style="text-align:center; font-weight:bold; color:white;">Credit</td><td style="text-align:center; font-weight:bold; color:white;">Balance</td><td style="text-align:center; font-weight:bold; color:white;">Debit</td><td style="text-align:center; font-weight:bold; color:white;">Credit</td><td style="text-align:center; font-weight:bold; color:white;">Balance</td><td style="text-align:center; font-weight:bold; color:white;">Debit</td><td style="text-align:center; font-weight:bold; color:white;">Credit</td><td style="text-align:center; font-weight:bold; color:white;">Balance</td></tr>'
#     employee = frappe.get_all("Employee",["name","employee_name"])
#     op_credit = 0
#     op_debit = 0
#     total_op_debit = 0
#     total_op_credit = 0
#     t_c_credit = 0
#     t_p_credit = 0
#     t_c_debit = 0
#     t_p_debit = 0
#     for j in employee:
#         gle = frappe.db.sql("""
#         SELECT name, party, sum(debit) as debit_amount, sum(credit) as credit_amount
#         FROM `tabGL Entry` 
#         WHERE account = %s and posting_date < %s and is_opening = 'No'
#         and party = %s and party_type ='Employee' and is_cancelled = 0
#         """, (account,from_date,j.name), as_dict=True)
#         for g in gle:
#             if not g.debit_amount:
#                 g.debit_amount = 0
#             if not g.credit_amount:
#                 g.credit_amount = 0
#             t_p_debit += g.debit_amount
#             t_p_credit += g.credit_amount
#             balance_op = t_p_debit - t_p_credit
#             sq = frappe.db.sql(""" select name,party,sum(debit_in_account_currency) as debit,sum(credit_in_account_currency) as credit from `tabGL Entry` where account = '%s' and party = '%s' and party_type = 'Employee' and posting_date between '%s' and '%s' and is_opening = 'No' and is_cancelled = 0 """%(account,j.name,from_date,to_date),as_dict=True)
#             for i in sq:
#                 if not i.credit:
#                     i.credit = 0
#                 if not i.debit:
#                     i.debit = 0
#                 op_credit = g.credit_amount + i.credit
#                 op_debit = g.debit_amount + i.debit
#                 total_op_debit += i.debit
#                 total_op_credit += i.credit
#                 mo_balance = total_op_debit - total_op_credit
#                 t_c_credit += op_credit
#                 t_c_debit += op_debit
#                 balance_cl = t_c_debit - t_c_credit
#                 if g.debit_amount or g.credit_amount or i.credit or i.debit:
#                     data += '<tr><td >%s</td><td>%s</td><td style = text-align:right>%s</td><td style = text-align:right>%s</td><td style = text-align:right>%s</td>'%(j.name,j.employee_name,fmt_money(g.debit_amount),fmt_money(g.credit_amount),fmt_money(g.debit_amount - g.credit_amount ))
#                     data += '<td style = text-align:right>%s</td><td style = text-align:right>%s</td><td style = text-align:right>%s</td><td style = text-align:right>%s</td><td style = text-align:right>%s</td><td style = text-align:right>%s</td></tr>'%(fmt_money(i.debit),fmt_money(i.credit),fmt_money(i.debit - i.credit),fmt_money(op_debit),fmt_money(op_credit),fmt_money(op_debit-op_credit))	
#     data += '<tr style="text-align:right; font-weight:bold;"><td colspan = 2 style="text-align:center; font-weight:bold;">Total</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (fmt_money(t_p_debit), fmt_money(t_p_credit),fmt_money(balance_op),fmt_money(total_op_debit), fmt_money(total_op_credit), fmt_money(mo_balance),fmt_money(t_c_debit), fmt_money(t_c_credit),fmt_money(balance_cl))
#     data += '</table>'
#     return data


@frappe.whitelist()
def pb_update_status():
	pb = frappe.db.sql("""update `tabProject Budget` set docstatus = 0 where name = 'INT-PB-2023-00006-2'  """,as_dict=1)


@frappe.whitelist()
def comp():
	acc = ''
	account = "FUEL EXPENSE - ASTCC"
	acct = account.split(' - ')
	if len(acct) == 2:
		acc = acct[0]
	if len(acct) == 3:
		acc = acct[1]
	print(acc)

@frappe.whitelist()
def get_name_pb(so):
	sale = frappe.get_doc("Sales Order",so)
	return sale.items


# @frappe.whitelist()
# def get_st_name():
# 	name = "SHBW-STI-2023-00037"
# 	si = frappe.get_doc("Sales Invoice",name)
# 	for i in si.items:
# 		count = i.idx
# 		# print(count)
# 		vs = name.split('-')[0]
# 		filt = '%'+vs+'%'
# 		st = frappe.db.get_all("Stock Transfer",{'name': ['like', filt]},['name'])
# 		for j in st:
# 			sto = frappe.get_doc("Stock Transfer",j.name)
# 			for k in sto.items:
# 				cnt = k.idx
# 			if cnt == count:
# 				pass
# 				# print(j.name)
# 			for a in sto.items:
# 				if a.item_code == i.item_code:
# 					print(j.name)
# 	# print(st)



@frappe.whitelist()
def set_stock_qty():
	stock = frappe.db.sql(""" select sales_person_user,sum(grand_total) as total from `tabSales Invoice` where posting_date between '%s' and '%s'  group by sales_person_user """%("2023-01-01","2023-12-31"),as_dict=1)
	print(stock)



@frappe.whitelist()
def get_st_qty():
	item = "122-31P090BL"
	# count = frappe.db.sql(""" select sum(actual_qty) as sum from `tabStock Ledger Entry` where item_code = '%s' and company = 'ELECTRA - BARWA SHOWROOM' and posting_date between '%s' and '%s' group by item_code """%(item,"2023-01-01","2023-07-01"),as_dict=1)[0]
	coun = frappe.db.sql(""" select sum(actual_qty) as sum from `tabStock Ledger Entry` where item_code = '%s' and company = 'ELECTRA - BARWA SHOWROOM' group by item_code """%(item),as_dict=1)[0]
	# print(count)
	print(coun['sum'])
	# print(count['sum'])
	# print(coun['sum'] - count['sum'])


#While enter the employee in leave extension get the below values and set it
@frappe.whitelist()
def get_leave_details_le(employee):
	leave_application = frappe.get_list(
		"Leave Application",
		filters={
			"employee": employee,
			"status": "Approved",
			"leave_type":("In",("Annual Leave","Leave Without Pay"))
		},
		fields=["name", "from_date", "to_date"],
		order_by="creation desc",
		limit=1
	)
	if leave_application:
		leave_details = leave_application[0]
		return {
			"leave_application": leave_details.name,
			"from_date": leave_details.from_date,
			"to_date": leave_details.to_date
		}

	else:
		frappe.throw(_("No approved annual leave application found for the employee."))


@frappe.whitelist()
def updat_series():
	from electra.utils import get_series
	ab = get_series("TRADING DIVISION - ELECTRA","Purchase Invoice")
	bc = ab.split('-')
	print(bc[0]+'-'+bc[1]+'-'+'STC-'+bc[2]+'-')
 
#Run the below method for every half and hour
@frappe.whitelist()    
def mark_absent():
	from_date = add_days(today(),-1)  
	to_date = today()
	# from_date = "2023-10-02"
	# to_date = "2023-10-03"
	dates = [from_date,to_date]
	for date in dates:
		employee = frappe.db.get_all('Employee',{'status':'Active','date_of_joining':['<=',from_date]},['*'])
		for emp in employee:
			# hh = check_holiday(date,emp.name)
			# if not hh:
			if not frappe.db.exists('Attendance',{'attendance_date':date,'employee':emp.name,'docstatus':('!=','2')}):
				att = frappe.new_doc('Attendance')
				att.employee = emp.name
				att.status = 'Absent'
				att.attendance_date = date
				att.company = emp.company
				att.save(ignore_permissions=True)
				frappe.db.commit() 
				print(date)  
	return "ok"

from datetime import date, timedelta

def get_dates_between(start_date, end_date):
	"""
	Get a list of dates between start_date and end_date (inclusive).

	Args:
	start_date (datetime.date): The start date.
	end_date (datetime.date): The end date.

	Returns:
	list: A list of datetime.date objects representing the dates.
	"""
	dates = []
	current_date = start_date
	while current_date <= end_date:
		dates.append(current_date)
		current_date += timedelta(days=1)
	return dates

#check the holiday for given employee
def check_holiday(date,emp):
	holiday_list = frappe.db.get_value('Employee',emp,'holiday_list')
	holiday = frappe.db.sql("""select `tabHoliday`.holiday_date,`tabHoliday`.weekly_off from `tabHoliday List` 
	left join `tabHoliday` on `tabHoliday`.parent = `tabHoliday List`.name where `tabHoliday List`.name = '%s' and holiday_date = '%s' """%(holiday_list,date),as_dict=True)
	if holiday:
		if holiday[0].weekly_off == 1:
			return "WW"
		else:
			return "HH"


  

#The below code is showing html format for norden site product search
@frappe.whitelist()
def get_norden_details(**args):
	data = ''
	data1 = ''
	i = 0
	aa = args['item']
	item = frappe.get_value('Item',{'item_code':args['item']},'item_code')
	if item:
		rate = frappe.get_value('Item',{'item_code':args['item']},'valuation_rate')
		group = frappe.get_value('Item',{'item_code':args['item']},'item_group')
		des = frappe.get_value('Item',{'item_code':args['item']},'description')
		price = frappe.get_value('Item Price',{'item_code':args['item'],'price_list':'Cost'},'price_list_rate')
		c_s_p = frappe.get_value('Item Price',{'item_code':args['item'],'price_list':'Standard Selling'},'price_list_rate') or 0
		spn = frappe.db.sql(""" select supplier_part_no from `tabItem` left join `tabItem Supplier` on `tabItem`.name = `tabItem Supplier`.parent 
		where `tabItem`.item_code = '%s' """%(args['item']),as_dict=True)[0]
		csp = 'Current Selling Price'
		cpp = 'Current Purchase Price'
		cost = 'COST'
		pso = 'Pending Sales order'
		po ='Total Purchase Order'
		ppo = 'Pending Purchase order'
		cspp_rate = 0
		cppp_rate = 0
		del_total = 0
		psoc = 0
		ppoc = 0
		ppoc_total = 0
		cspp_query = frappe.db.sql("""select `tabSales Order Item`.rate as rate, `tabSales Order`.creation from `tabSales Order`
		left join `tabSales Order Item` on `tabSales Order`.name = `tabSales Order Item`.parent
		where `tabSales Order Item`.item_code = '%s' order by `tabSales Order`.creation """ % (item), as_dict=True)
		if cspp_query:
				cspp = cspp_query[-1]
				cspp_rate = cspp['rate']

		cppp_query = frappe.db.sql("""select `tabPurchase Order Item`.rate as rate, `tabPurchase Order`.creation  from `tabPurchase Order`
		left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
		where `tabPurchase Order Item`.item_code = '%s' order by `tabPurchase Order`.creation """ % (item), as_dict=True)
		if cppp_query:
				cppp = cppp_query[-1]
				cppp_rate = cppp['rate']

		new_so = frappe.db.sql("""select sum(`tabSales Order Item`.qty) as qty,sum(`tabSales Order Item`.delivered_qty) as d_qty from `tabSales Order`
		left join `tabSales Order Item` on `tabSales Order`.name = `tabSales Order Item`.parent
		where `tabSales Order Item`.item_code = '%s' and `tabSales Order`.docstatus = 1 and `tabSales Order`.status != "Closed"  """ % (item), as_dict=True)[0]
		if not new_so['qty']:
			new_so['qty'] = 0
		if not new_so['d_qty']:
			new_so['d_qty'] = 0
		del_total = new_so['qty'] - new_so['d_qty']		
				
		new_po = frappe.db.sql("""select sum(`tabPurchase Order Item`.qty) as qty,sum(`tabPurchase Order Item`.received_qty) as d_qty from `tabPurchase Order`
		left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
		where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus = 1 and `tabPurchase Order`.status != 'Closed'  """ % (item), as_dict=True)[0]
		if not new_po['qty']:
			new_po['qty'] = 0
		if not new_po['d_qty']:
			new_po['d_qty'] = 0
		ppoc_total = new_po['qty'] - new_po['d_qty']	



		a = frappe.db.sql("""select `tabPurchase Receipt Item`.purchase_order as name from `tabPurchase Receipt`
				left join `tabPurchase Receipt Item` on `tabPurchase Receipt`.name = `tabPurchase Receipt Item`.parent
				where `tabPurchase Receipt Item`.item_code = '%s' """ % (aa), as_dict=True) or "-"

		b = frappe.db.sql("""select `tabPurchase Receipt`.name as name from `tabPurchase Receipt`
				left join `tabPurchase Receipt Item` on `tabPurchase Receipt`.name = `tabPurchase Receipt Item`.parent
				where `tabPurchase Receipt Item`.item_code = '%s' """ % (aa), as_dict=True) or "-"
		
		data = ''
		stocks_query = frappe.db.sql("""select actual_qty,warehouse,valuation_rate,stock_uom,stock_value from tabBin
				where item_code = '%s' """%(item),as_dict=True)
		if stocks_query:
				stocks = stocks_query
		data += '<table class="table table-bordered" style="width:70%"><tr><th style="padding:1px;border: 1px solid black;background-color:#fe3f0c;color:white" colspan=5><center> ELECTRA PRODUCT SEARCH</center></th></tr>'
		data += '<tr><td colspan = 2 style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;text-align: left"><b>Item Code</b></td><td colspan = 3 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(item)
		data += '<tr><td colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;text-align: left"><b>Item Name</b></td><td colspan = 3 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(frappe.db.get_value('Item',item,'item_name'))
		data += '<tr><td colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;text-align: left"><b>Item Group</b></td><td colspan = 3 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(group)
		data += '<tr><td colspan = 2 style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;text-align: left"><b>%s</b></td><td colspan = 3 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(pso,del_total)
		if ppoc_total != 0.0:
				data += '<tr><td colspan = 2 style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;text-align: left"><b>%s</b></td><td colspan = 3 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(ppo,ppoc_total)
		else:
				data += '<tr><td colspan = 2 style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;text-align: left"><b>%s</b></td><td colspan = 3 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(ppo,ppoc_total) 
		if ppoc_total > 0 :
				ppoc_date_query = frappe.db.sql("""select `tabPurchase Order Item`.schedule_date  as date from `tabPurchase Order`
				left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
				where `tabPurchase Order Item`.item_code = '%s' """ % (item), as_dict=True)[0]
				if ppoc_date_query:
					po_date = ppoc_date_query['date']
			
		data += '<tr><td colspan = 2 style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;text-align: left"><b>%s</b></td><td colspan = 3 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(csp,(c_s_p))
		data += '<tr><td colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Warehouse</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>QTY</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>UOM</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>COST</b></center></td></tr>'		
		i = 0
		cou = 0
		tot = 'Total'
		uom = 'Nos'
		for stock in stocks_query:
			if stock.actual_qty >= 0:       
				comp = frappe.db.get_value("Employee",{'user_id':frappe.session.user},'company')
				w_house = frappe.db.get_value("Warehouse",{'company':comp,'default_for_stock_transfer':1},['name'])
				if w_house == stock.warehouse:
					data += '<tr><td colspan = 2 style="padding:1px;border: 1px solid black">%s</td><td colspan = 1 style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td></tr>'%(stock.warehouse,int(stock.actual_qty) or '-',stock.stock_uom or '-',(round(float(stock.valuation_rate) , 2)) or '-')
					i += 1
					cou += stock.actual_qty  

		for stock in stocks_query:
			if stock.actual_qty >= 0:       
				comp = frappe.db.get_value("Employee",{'user_id':frappe.session.user},'company')
				w_house = frappe.db.get_value("Warehouse",{'company':comp,'default_for_stock_transfer':1},['name'])
				if w_house != stock.warehouse:
					data += '<tr><td colspan = 2 style="padding:1px;border: 1px solid black">%s</td><td colspan = 1 style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td></tr>'%(stock.warehouse,int(stock.actual_qty) or '-',stock.stock_uom or '-',(round(float(stock.valuation_rate) , 2)) or '-')
					i += 1
					cou += stock.actual_qty
		data += '<tr><td align="right" colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><b>%s</b></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>%s</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>%s</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>-</b></center></td></tr>'%(tot,int(cou),uom)
		data += '</table>'
	else:
		i += 1
		data1 += '<table width = "100%"><tr><td align="center" colspan = 10  style="padding:1px;border: 1px solid black;background-color:#fe3f0c;color:white;"><b>No Stock Available</b></td></tr>'
		data1 += '</table>'
		data += data1
	if i > 0:
		return data
	
#The below code is showing html format for norden site product search
@frappe.whitelist()
def norden_details(**args):
	data = ''
	data1 = ''
	i = 0
	aa = args['item']
	item = frappe.get_value('Item',{'item_code':args['item']},'item_code')
	if item:
		rate = frappe.get_value('Item',{'item_code':args['item']},'valuation_rate')
		group = frappe.get_value('Item',{'item_code':args['item']},'item_group')
		des = frappe.get_value('Item',{'item_code':args['item']},'description')
		price = frappe.get_value('Item Price',{'item_code':args['item'],'price_list':'Cost'},'price_list_rate')
		c_s_p = frappe.get_value('Item Price',{'item_code':args['item'],'price_list':'Standard Selling'},'price_list_rate') or 0
		spn = frappe.db.sql(""" select supplier_part_no from `tabItem` left join `tabItem Supplier` on `tabItem`.name = `tabItem Supplier`.parent 
		where `tabItem`.item_code = '%s' """%(args['item']),as_dict=True)[0]
		csp = 'Current Selling Price'
		cpp = 'Current Purchase Price'
		cost = 'COST'
		pso = 'Pending Sales order'
		po ='Total Purchase Order'
		ppo = 'Pending Purchase order'
		cspp_rate = 0
		cppp_rate = 0
		psoc = 0
		del_total = 0
		ppoc = 0
		ppoc_total = 0
		cspp_query = frappe.db.sql("""select `tabSales Order Item`.rate as rate, `tabSales Order`.creation from `tabSales Order`
		left join `tabSales Order Item` on `tabSales Order`.name = `tabSales Order Item`.parent
		where `tabSales Order Item`.item_code = '%s' order by `tabSales Order`.creation """ % (item), as_dict=True)
		if cspp_query:
			cspp = cspp_query[-1]
			cspp_rate = cspp['rate']

		cppp_query = frappe.db.sql("""select `tabPurchase Order Item`.rate as rate, `tabPurchase Order`.creation  from `tabPurchase Order`
		left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
		where `tabPurchase Order Item`.item_code = '%s' order by `tabPurchase Order`.creation """ % (item), as_dict=True)
		if cppp_query:
			cppp = cppp_query[-1]
			cppp_rate = cppp['rate']


		new_so = frappe.db.sql("""select sum(`tabSales Order Item`.qty) as qty,sum(`tabSales Order Item`.delivered_qty) as d_qty from `tabSales Order`
		left join `tabSales Order Item` on `tabSales Order`.name = `tabSales Order Item`.parent
		where `tabSales Order Item`.item_code = '%s' and `tabSales Order`.docstatus = 1 and `tabSales Order`.status != "Closed"  """ % (item), as_dict=True)[0]
		if not new_so['qty']:
			new_so['qty'] = 0
		if not new_so['d_qty']:
			new_so['d_qty'] = 0
		del_total = new_so['qty'] - new_so['d_qty']		

		new_po = frappe.db.sql("""select sum(`tabPurchase Order Item`.qty) as qty,sum(`tabPurchase Order Item`.received_qty) as d_qty from `tabPurchase Order`
		left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
		where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus = 1 and `tabPurchase Order`.status != 'Closed'  """ % (item), as_dict=True)[0]
		if not new_po['qty']:
			new_po['qty'] = 0
		if not new_po['d_qty']:
			new_po['d_qty'] = 0
		ppoc_total = new_po['qty'] - new_po['d_qty']	


		a = frappe.db.sql("""select `tabPurchase Receipt Item`.purchase_order as name from `tabPurchase Receipt`
			left join `tabPurchase Receipt Item` on `tabPurchase Receipt`.name = `tabPurchase Receipt Item`.parent
			where `tabPurchase Receipt Item`.item_code = '%s' """ % (aa), as_dict=True) or "-"
		b = frappe.db.sql("""select `tabPurchase Receipt`.name as name from `tabPurchase Receipt`
			left join `tabPurchase Receipt Item` on `tabPurchase Receipt`.name = `tabPurchase Receipt Item`.parent
			where `tabPurchase Receipt Item`.item_code = '%s' """ % (aa), as_dict=True) or "-"
		data = ''
		stocks_query = frappe.db.sql("""select actual_qty,warehouse,stock_uom,stock_value from tabBin
			where item_code = '%s' """%(item),as_dict=True)
		if stocks_query:
			stocks = stocks_query
		data += '<table class="table table-bordered" style="width:50%"><tr><th style="padding:1px;border: 1px solid black;background-color:#fe3f0c;" colspan=4><center>ELECTRA PRODUCT SEARCH</center></th></tr>'
		data += '<tr><td colspan = 2 style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;text-align: left"><b>Item Code</b></td><td colspan = 2 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(item)
		data += '<tr><td colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;text-align: left"><b>Item Name</b></td><td colspan = 2 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(frappe.db.get_value('Item',item,'item_name'))
		data += '<tr><td colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;text-align: left"><b>Item Group</b></td><td colspan = 2 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(group)
		data += '<tr><td colspan = 2 style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;text-align: left"><b>%s</b></td><td colspan = 2 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(pso,del_total)
		if ppoc_total != 0.0:
			data += '<tr><td colspan = 2 style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;text-align: left"><b>%s</b></td><td colspan = 2 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(ppo,ppoc_total)
		else:
			data += '<tr><td colspan = 2 style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;text-align: left"><b>%s</b></td><td colspan = 2 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(ppo,ppoc_total) 
		if ppoc_total > 0 :
			ppoc_date_query = frappe.db.sql("""select `tabPurchase Order Item`.schedule_date  as date from `tabPurchase Order`
			left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
			where `tabPurchase Order Item`.item_code = '%s' """ % (item), as_dict=True)[0]
			if ppoc_date_query:
				po_date = ppoc_date_query['date']
			
		data += '<tr><td colspan = 2 style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;text-align: left"><b>%s</b></td><td colspan = 2 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(csp,(c_s_p))
		data += '<tr><td colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Warehouse</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>QTY</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>UOM</b></center></tr>'
		i = 0
		cou = 0
		tot = 'Total'
		uom = 'Nos'
  
		for stock in stocks_query:
			if int(stock.actual_qty) >= 0:
				comp = frappe.db.get_value("Employee",{'user_id':frappe.session.user},'company')
				w_house = frappe.db.get_value("Warehouse",{'company':comp,'default_for_stock_transfer':1},['name'])
				if w_house == stock.warehouse:        
					data += '<tr><td colspan = 2 style="padding:1px;border: 1px solid black">%s</td><td colspan = 1 style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td></tr>'%(stock.warehouse,int(stock.actual_qty) or '-',stock.stock_uom or '-')
					i += 1
					cou += stock.actual_qty

		for stock in stocks_query:
			if int(stock.actual_qty) >= 0:
				comp = frappe.db.get_value("Employee",{'user_id':frappe.session.user},'company')
				w_house = frappe.db.get_value("Warehouse",{'company':comp,'default_for_stock_transfer':1},['name'])
				if w_house != stock.warehouse:        
					data += '<tr><td colspan = 2 style="padding:1px;border: 1px solid black">%s</td><td colspan = 1 style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td></tr>'%(stock.warehouse,int(stock.actual_qty) or '-',stock.stock_uom or '-')
					i += 1
					cou += stock.actual_qty
		data += '<tr><td align="right" colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><b>%s</b></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>%s</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>%s</b></center></td></tr>'%(tot,int(cou),uom)
		data += '</table>'
	else:
		i += 1
		data1 += '<table width = "100%"><tr><td align="center" colspan = 10  style="padding:1px;border: 1px solid black;background-color:#fe3f0c;color:white;"><b>No Stock Available</b></td></tr>'
		data1 += '</table>'
		data += data1
	if i > 0:
		return data

	

@frappe.whitelist()
def val_rate():
	valuation_rate = 0
	source_warehouse = frappe.db.get_value('Warehouse', {'default_for_stock_transfer':1,'company': "TRADING DIVISION - ELECTRA" }, ["name"])
	latest_vr = frappe.db.sql("""select valuation_rate as vr from tabBin
			where item_code = '%s' and warehouse = '%s' """%("AF21036",source_warehouse),as_dict=True)

	if len(latest_vr) > 0:
		valuation_rate = latest_vr[0]["vr"]
	# else:
	#     val_rate = []
	#     l_vr = frappe.db.sql("""select valuation_rate as vr from tabBin
	#             where item_code = '%s' """%("AF21036"),as_dict=True)
	#     for item in l_vr: 
	#         if item not in val_rate: 
	#             val_rate.append(item.vr)
	#     if len(val_rate) > 1 :
	#         valuation_rate = max(val_rate)
	print(valuation_rate)


@frappe.whitelist()
def get_project_name():
	pro = frappe.db.sql("""SELECT
		`tabProject Budget`.name AS project_budget_name
	FROM
		`tabProject`
	JOIN
		`tabProject Budget` ON `tabProject`.`budgeting` = `tabProject Budget`.name and `tabProject`.name = 'ENG-PRO-2023-00014';
	""")
	
	# project = frappe.db.sql(""" select `tabProject Budget`.name from `tabProject Budget` left join `tabProject` where `tabProject Budget`.name = `tabProject`.budgeting   """)
	print(pro)

@frappe.whitelist()
def update_workflow_state():
	stock = frappe.db.get_list("Stock Request",{"docstatus":2},["workflow_state","name"])
	for i in stock:
		if i.workflow_state == "Transfer Requested":
			frappe.db.set_value("Stock Request",i.name,"workflow_state","Cancelled")
			print(i.name)

from frappe.utils.file_manager import get_file
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe import _, get_file_items
@frappe.whitelist()
def update_att(import_file):
	filepath = get_file(import_file)
	pps = read_csv_content(filepath[1])
	for pp in pps:
		print(pp[0])
		if not frappe.db.exists('Attendance',{'attendance_date':pp[1],'employee':pp[0],'docstatus':('!=','2')}):
			att = frappe.new_doc('Attendance')
			att.employee = pp[0]
			att.status = pp[4]
			att.attendance_date = pp[1]
			att.leave_type = pp[2]
			att.leave_application = pp[3]
			att.save(ignore_permissions=True)
			frappe.db.commit() 
		else:
			att = frappe.get_doc('Attendance',{'attendance_date':pp[1],'employee':pp[0],'docstatus':('!=','2')})
			att.employee = pp[0]
			att.status = pp[4]
			att.attendance_date = pp[1]
			att.leave_type = pp[2]
			att.leave_application = pp[3]
			att.save(ignore_permissions=True)
			frappe.db.commit() 

	return 'ok'
	

#check the below condition is on submission of additional salary
@frappe.whitelist()
def salary_payroll(doc,method):
	salary = frappe.db.exists("Additional Salary",{"payroll_date":doc.payroll_date,"salary_component":doc.salary_component,"employee":doc.employee,'docstatus':('!=','2'),'name':('!=',doc.name)})
	if salary:
		print("HI")
		frappe.throw(_('This employee already has a Additional Salary against same component & same payroll period'))
	

@frappe.whitelist()
def update_workflow_state_po():
	po = frappe.db.get_list("Purchase Order",{"docstatus":2},["workflow_state","name"])
	for i in po:
		if i.workflow_state == "Submitted":
			frappe.db.set_value("Purchase Order",i.name,"workflow_state","Cancelled")
			print(i.name)

@frappe.whitelist()
def get_date():
	doc =frappe.db.get_list("Delivery Note",["name","date","posting_date"])
	for i in doc:
		frappe.db.set_value("Delivery Note",i.name,"date",i.posting_date)

#Get the below value while enter the company in day schedule form
@frappe.whitelist()
def get_project(company):
	value = frappe.db.get_list("Project",{"company":company,"status":"Open"},["name","customer","project_name"])
	return value

		
#Append the below values in Project Timesheet
@frappe.whitelist()
def get_day_plan_details(day_plan):
	doc = frappe.get_doc("Single Project Day Plan",day_plan)
	day_plan_details = []
	for i in doc.worker_multiselect:
		day_plan_details.append(frappe._dict({
			'project': doc.project,
			'project_name': doc.project_name,
			'employee': i.employee,
			'employee_name' : i.non_staff,
			'designation' : i.designation,
		}))
	for s in doc.staff_multiselect:
		day_plan_details.append(frappe._dict({
			'project': doc.project,
			'project_name': doc.project_name,
			'employee': s.employee,
			'employee_name' : s.staff,
			'designation' : s.designation,
		}))
	for p in doc.supervisor_multiselect:
		day_plan_details.append(frappe._dict({
			'project': doc.project,
			'project_name': doc.project_name,
			'employee': p.employee,
			'employee_name' : p.supervisor,
			'designation' : p.designation,
		}))
	return day_plan_details

#Below values using in print format for Cost Estimation , SO , SI , QN and PB
@frappe.whitelist()
def scope_of_work(doc):
	list=[]
	for i in doc.scope_of_work:
		if(i.delivery):
			list.append(i.delivery)
	return len(list)

@frappe.whitelist()
def list_ce():
	val = frappe.db.get_value('Quotation', { 'cost_estimation': "INT-CE-2023-00145-1"}, 'name',order_by='transaction_date asc',)
	print(val)


# @frappe.whitelist()
# def is_cost_estimation_submitted(cost_estimation):
#     cost_estimation_doc = frappe.get_doc("Cost Estimation", cost_estimation)
	
#     if cost_estimation_doc.docstatus == 1:
#         return True 
#     else:
#         return False 

# @frappe.whitelist()
# def is_cost_estimation_submitted(cost_estimation):
# 	cost_estimation_doc = frappe.get_doc("Cost Estimation", cost_estimation)

# 	if cost_estimation_doc.docstatus == 1:
# 		frappe.permissions.update_permission_property(
# 			"CE SOW", cost_estimation_doc.name, "total_overhead", "read_only", 1
# 		)
# 	else:
# 		frappe.permissions.update_permission_property(
# 			"CE SOW", cost_estimation_doc.name, "total_overhead", "read_only", 0
# 		)

#is_consumable check is enable or disable based on below condition in DN
@frappe.whitelist()
def consumables_table_dn(delivery_note):
	item_list = []
	tot = []
	total = 0
	dn = frappe.get_doc("Delivery Note",delivery_note)
	for i in dn.items:
		if i.item_code not in  item_list:
			item_list.append(i.item_code)
		if i.item_code == "CONS":
			tot.append(i.amount)
	if "CONS" in item_list:
		for ele in range(0, len(tot)):
			total = total + tot[ele]
		return 1,total
	else:
		return 0,total

#is_consumable check is enable or disable based on below condition in DN WIP	
@frappe.whitelist()
def consumables_table_dn_wip(delivery_note):
	item_list = []
	tot = []
	total = 0
	dn = frappe.get_doc("Delivery Note WIP",delivery_note)
	for i in dn.items:
		if i.item_code not in  item_list:
			item_list.append(i.item_code)
		if i.item_code == "CONS":
			tot.append(i.amount)
	if "CONS" in item_list:
		for ele in range(0, len(tot)):
			total = total + tot[ele]
		return 1,total
	else:
		return 0,total

@frappe.whitelist()
def return_consumables_sales_price(item_code):
	price = frappe.get_value('Item Price',{'item_code':item_code,'price_list':'Standard Selling'},'price_list_rate') or 0
	return price

#Create new stock entry on submission of DN
@frappe.whitelist()
def create_stock_out(doc,method):
	if doc.cons_items:
		se = frappe.new_doc("Stock Entry")
		se.stock_entry_type = "Material Issue" 
		se.company = doc.company
		se.reference_number = doc.name
		for i in doc.cons_items:
			se.append("items",{
				's_warehouse':doc.set_warehouse,
				'item_code':i.item_code,
				'qty':i.qty,
				'basic_rate':i.rate
			})
		se.save(ignore_permissions=True)
		se.submit()

#Cancel the linked DN cancel of this dn
@frappe.whitelist()
def on_cancel_dn(doc,method):
	if frappe.db.exists("Stock Entry",{"reference_number":doc.name}):
		se = frappe.db.get_value("Stock Entry",{"reference_number":doc.name},['name'])
		if se:
			se_cancel = frappe.get_doc("Stock Entry",se)
			se_cancel.cancel()
			se_cancel.delete()

@frappe.whitelist()
def get_company():
	sa=frappe.db.sql("""select name from `tabCompany`""",as_dict=True)
	return sa

@frappe.whitelist()
def get_sales_person_name():
	sa=frappe.db.sql("""select name from `tabSales Person`""",as_dict=True)
	return sa

#Cancel the linked DN cancel of this SO , id dn status is 0 then throw the below message
@frappe.whitelist()
def sales_order_cancel(doc,method):
	dn_link = frappe.db.sql("""SELECT `tabDelivery Note`.name , `tabDelivery Note`.docstatus FROM `tabDelivery Note` LEFT JOIN `tabDelivery Note Item` ON `tabDelivery Note`.name = `tabDelivery Note Item`.parent WHERE
				`tabDelivery Note Item`.against_sales_order = '%s'
		""" % (doc.name), as_dict=True)
	if dn_link:
		for i in dn_link:
			if i.docstatus == 0:
				frappe.throw(_("Sales Order is linked to a Delivery Note "+i.name+" . Cannot cancel."))

@frappe.whitelist()
def salary_employee():
	value=frappe.db.get_all("Employee",{"status":"Active"},["status","name","company"])
	for i in value:
		# print(i.name)
		if frappe.db.exists('Salary Structure Assignment',{'employee':i.name,'docstatus':('!=','2')}):
			ad = frappe.db.sql("""select * from `tabSalary Structure Assignment` where employee = '%s' and docstatus = 1 ORDER BY from_date DESC LIMIT 1  """%(i.name),as_dict=True)[0]
			if i.company != ad.company:
				print(i.name)
			# print(ad.company)
   
# @frappe.whitelist()
# def customer_credit_note(customer):
# 	cust = frappe.get_doc("Customer",customer)
# 	return cust.credit_limits

# @frappe.whitelist()
# def validate_cl(doc,method):
# 	customer_type = frappe.db.get_value("Customer",doc.customer,['customer_type'])
# 	customer_list = get_details(doc.company,doc.customer)
# 	if not customer_list and customer_type == "Company":
# 		frappe.throw(_("Kindly Set the Credit Limit in Customer"))

def get_details(company,customer):
	sql_query = """SELECT
						c.name, c.customer_name,
						ccl.bypass_credit_limit_check,
						c.is_frozen, c.disabled
					FROM    `tabCustomer` c, `tabCustomer Credit Limit` ccl
					WHERE
						c.name = ccl.parent
						AND ccl.company = '%s' AND c.name = '%s' """ % (company,customer)

	return frappe.db.sql(sql_query, as_dict=1)

# @frappe.whitelist()
# def invoice(doc,method):
# 	total = 0
# 	for i in doc.sow:
# 		for j in doc.items:
# 			if j.msow==i.msow:
# 				total += j.amount
# 		i.invoice_amount = total
# 		i.invoice_percent = (i.invoice_amount/i.total_bidding_price)*100
# 		total = 0
# 		so = frappe.get_doc("Sales Order",doc.sales_order)
# 		for k in so.scope_of_work:
# 			k.invoice_amount += i.invoice_amount
# 			k.invoice_percent += k.invoice_percent
			

@frappe.whitelist()
def invoice(doc, method):
	if doc.order_type == "Project":
		total = 0
		for i in doc.sow:
			for j in doc.items:
				if j.msow == i.msow:
					total += j.amount
			i.invoice_amount = total
			i.invoice_percent = (i.invoice_amount / i.total_bidding_price) * 100
			total = 0
		if doc.sales_order:
			so = frappe.get_doc("Sales Order", doc.sales_order)			
			for i in doc.sow:
				for k in so.scope_of_work:
					if i.msow == k.msow:
						k.invoice_amount += i.invoice_amount
						k.invoice_percent += i.invoice_percent
			so.save()

#The below values has been revert on cancel of Invoice
@frappe.whitelist()
def invoice_cancel(doc, method):
	if doc.order_type == "Project":
		if doc.sow:
			total = 0
			for i in doc.sow:
				for j in doc.items:
					if j.msow == i.msow:
						total += j.amount
				i.invoice_amount = total
				i.invoice_percent = (i.invoice_amount / i.total_bidding_price) * 100
				total = 0
			if doc.so_no:
				so = frappe.get_doc("Sales Order", doc.so_no)
				
				for i in doc.sow:
					for k in so.scope_of_work:
						if i.msow == k.msow and k.invoice_amount != 0:
							k.invoice_amount -= i.invoice_amount
							k.invoice_percent -= i.invoice_percent

				so.save()

	
#Update the below values of sales order is on update of after submit
@frappe.whitelist()
def update_bidding_price(doc,method):
	total_bidding_price = 0
	for i in doc.scope_of_work:
		total_bidding_price += i.total_bidding_price
	doc.total_bidding_price = total_bidding_price
	doc.total_bidding_price = total_bidding_price - doc.project_discount_amt

@frappe.whitelist()
def update_jv():
	jv = frappe.db.get_list("Journal Entry",{"total_amount":0,"docstatus":1},['name','total_credit'])
	for i in jv:
		frappe.db.set_value("Journal Entry",i.name,"total_amount",i.total_credit)
		print(i.name)
		print(i.total_credit)

#Below values are showing html format on while opening the Consolidated receivable
@frappe.whitelist()
def sal_consolidated_req(name):
	con = frappe.db.sql("""SELECT * FROM `tabOrder Invoice List` WHERE parent = %s""", name, as_dict=True)
	child_table = con

	def calculate_outstanding_payment_difference(child_table):
		amounts_by_company = defaultdict(float)
		for row in child_table:
			company = row.get("company", "")
			outstanding_amount = row.get("outstanding_amount", 0.0)
			payment_amount = row.get("receivable_amount", 0.0)
			amounts_by_company[company] += (outstanding_amount - payment_amount)
		
		result = []
		for company, difference in amounts_by_company.items():
			result.append({"company": company, "value": difference})
		# Build the HTML table
		data = '<table width="100%"><tr><td style="color:white;background-color:#e35310;padding:1px;font-weight:bold;border:1px solid black">Company</td><td style="font-weight:bold;color:white;background-color:#e35310;padding:1px;border:1px solid black">Amount</td></tr>'
		for i in result:
			data += '<tr><td style="padding:1px;border:1px solid black">%s</td><td style="padding:1px;border:1px solid black">%s</td></tr>' % (i["company"], i["value"])
		data += '</table>'
		return data

	amounts_difference_by_company = calculate_outstanding_payment_difference(child_table)
	return amounts_difference_by_company

@frappe.whitelist()
def update_pr():
	frappe.db.set_value("Purchase Invoice Item",{"parent":"TRD-PI-2023-00362"},'purchase_receipt',"TRD-PR-2023-00184-1",update_modified=False)

#Get the dn values and set to sales invoice
@frappe.whitelist()
def get_dn_no(sale_ord):
	dn = []
	sales_order = frappe.db.sql("""select parent as dn from `tabDelivery Note Item` where against_sales_order = '%s' """ % sale_ord,as_dict=1)
	if sales_order:
		for item in sales_order: 
			if item.dn not in dn:
				if item.dn:
					dn.append(item.dn)
	if len(dn) > 0:
		return str(dn)
	else:
		return ''
	

	
# import requests

# @frappe.whitelist()
# def fetch_location_name(pickup_point):
#     api_key = '583681db58de9d7'
#     geocoding_url = f'https://maps.googleapis.com/maps/api/geocode/json?key={api_key}&address={pickup_point}'

#     try:
#         response = requests.get(geocoding_url)
#         data = response.json()

#         if data.get('status') == 'OK' and data.get('results'):
#             location_name = data['results'][0]['formatted_address']
#             return {"location_name": location_name}
#         else:
#             return {"location_name": "Location Not Found"}
#     except Exception as e:
#         frappe.log_error(f'Error fetching location name: {str(e)}')
#         return {"location_name": "Error"}




@frappe.whitelist()
def lcv_value():
	lc = frappe.db.sql("""select sum(`tabLanded Cost Item`.applicable_charges) as pd from `tabLanded Cost Voucher`
			left join `tabLanded Cost Item` on `tabLanded Cost Voucher`.name = `tabLanded Cost Item`.parent
			where `tabLanded Cost Item`.receipt_document = '%s' and `tabLanded Cost Voucher`.docstatus =1 """%("INT-PR-2023-00008"),as_dict=True)
	
	print(lc)

#The below mail sent every day (Expired Legal Compliance Monitor)
@frappe.whitelist()    
def monthly_expiry_doc():
	current_date = nowdate()
	next_date = add_days(current_date , 30)
	doc=frappe.get_all("Legal Compliance Monitor",{'next_due': ('between', (current_date, next_date))},['name','days_left','next_due'])
	documents = ''
	documents += '<table class = table table - bordered style=border-width:2px><tr><td colspan = 6><b>Legal Compliance Monitor List</b></td></tr>'
	documents += '<tr><td colspan=2>Document Name</td><td colspan=2>Expiry Date</td><td colspan=2>Days Left</td>'
	for i in doc:
		documents += '<tr><td colspan = 2>%s</td><td colspan=2>%s</td><td colspan=2>%s</td></tr>'%(i.name,formatdate(i.next_due),i.days_left)
	documents += '</table>' 
	frappe.sendmail(
			   recipients=['suvarna@electraqatar.com'],
			 cc = ['yahia@electraqatar.com'],
			 subject=('Legal Compliance Monitor'),
			 message="""
					Dear Sir/Mam,<br>
					<p>Kindly find the attached Legal Compliance Monitor List for Expiry Soon</p>
					%s
					""" % (documents)
		) 
	
	return True

@frappe.whitelist()
def return_valu():
	doc = frappe.get_all("Sales Order Item",{'parent':"ENG-SO-2023-00185"},['name'])
	print(doc)

@frappe.whitelist()
def update_vmc():
	sep = frappe.db.sql("""update `tabVehicle Maintenance Check List` set docstatus = 0 where name = "ELE/HRA/05/00466" """,as_dict = True)
	print(sep)
 
 
#Get the below values and set it so
@frappe.whitelist()
def get_so_details(sales):
	dict_list = []
	so = frappe.get_doc("Sales Order",sales)
	for i in so.items:
		rem_qty=i.qty-i.custom_mr_qty
		dict_list.append(frappe._dict({"name":i.name,"item_code":i.item_code,"pending_qty":rem_qty,"bom":i.bom_no,"description": i.description,"warehouse":i.warehouse,"rate":i.rate,"amount":i.amount}))
		# dict_list.append(frappe._dict({"name":i.name,"item_code":i.item_code,"pending_qty":i.qty,"bom":i.bom_no,"description": i.description,"warehouse":i.warehouse,"rate":i.rate,"amount":i.amount}))
	return dict_list

@frappe.whitelist()
def update_stock_qty(doc,method):
	for i in doc.items:
		if doc.sales_order:
			sales_order = frappe.get_doc("Sales Order", doc.sales_order)
			for so_item in sales_order.items:
				if i.item_code == so_item.item_code:
					so_item.custom_sr_qty += i.qty
			sales_order.flags.ignore_validate_update_after_submit = True
			sales_order.save(ignore_permissions=True)

@frappe.whitelist()
def update_stock_qty_cancel(doc,method):
	for i in doc.items:
		if doc.sales_order:
			sales_order = frappe.get_doc("Sales Order", doc.sales_order)
			for so_item in sales_order.items:
				if i.item_code == so_item.item_code:
					so_item.custom_sr_qty -= i.qty
			sales_order.flags.ignore_validate_update_after_submit = True
			sales_order.save(ignore_permissions=True)

#Get the below values and set it so
@frappe.whitelist()
def make_material_request(items, prepared,sales_order, company, project=None):
	"""Make Work Orders against the given Sales Order for the given `items`"""
	items = json.loads(items).get("items")    
	material_req = frappe.get_doc({
		"doctype": "Material Request",
		"company": company,
		"sales_order": sales_order,
		"project": project,
		"prepared_by":prepared,
	})    
	for i in items:
		if i["bom"]:
			bom_list = frappe.get_all("BOM", {"name": i["bom"]},["*"])
			for bom in bom_list:
				items_list = frappe.get_all("BOM Item", {"parent": bom.name}, ["*"])
				for b in items_list:
					material_req.append('items',{
						"item_code":b.item_code,
						"rate":b.rate,
						"amount":b.amount * i["pending_qty"],
						"item_name":b.item_name,
						"qty":b.qty * i["pending_qty"],
						"schedule_date":today(),
						"sales_order":sales_order
					})
		else:
			material_req.append('items', {
				"item_code": i['item_code'],
				"qty": i['pending_qty'],
				"rate":i["rate"],
				"amount":i["amount"] ,
				"schedule_date": today(),
				"sales_order":sales_order

			})
	material_req.save()
	return material_req

#Check the relieving date there in active employee , if yes throww the below message
@frappe.whitelist()
def inactive_employee(doc,method):
	if doc.status=="Active":
		if doc.relieving_date:
			throw(_("Please remove the relieving date for the Active Employee."))

# @frappe.whitelist()
# def update_amount(employee,ot,wh):
# 	basic = frappe.db.get_value("Employee",{'employee_number':employee},[['basic']])
# 	one_day_amount = basic / 26
# 	wh = one_day_amount * float(wh)
# 	ot = wh * ot
# 	return ot

#Get these below value and set to Multi Project day plan
@frappe.whitelist()
def update_emp_value(employee):
	value=frappe.db.get_value("Employee",{'employee_number':employee},['employee_name'])
	designation = frappe.db.get_value("Employee",{'employee_number':employee},['designation'])
	grade = frappe.db.get_value("Employee",{'employee_number':employee},['grade'])
	return value,designation,grade

# @frappe.whitelist()
# def create_entry():
# 	employee = '629'
# 	leave_type = 'Annual Leave'
# 	leaves = -33
# 	from_date = '2023-01-15'
# 	to_date = '2023-02-16'
# 	transaction_type = 'Leave Application'
# 	transaction_name = 'HR-LAP-2023-00009'
# 	entry = frappe.new_doc('Leave Ledger Entry')
# 	entry.employee = employee
# 	entry.leave_type = leave_type
# 	entry.leaves = leaves
# 	entry.from_date = from_date
# 	entry.to_date = to_date
# 	entry.leave_type = leave_type
# 	entry.transaction_type = transaction_type
# 	entry.transaction_name = transaction_name
# 	entry.insert()
# 	return _("Leave Ledger Entry created successfully.")


@frappe.whitelist()
def update_lareei():
	frappe.db.sql("""update `tabAttendance` set docstatus = 0 where attendance_date between "2023-11-01" and "2023-11-30" and company = "Al - Shaghairi Trading and Contracting Company W.L.L (ELECTRA)" """,as_dict = True)
	
	# frappe.db.sql("""delete from `tabLeave Ledger Entry` where name = '2a43efd06b' """,as_dict = True)
	# frappe.db.sql("""delete from `tabRejoining Form` where name = 'ELE/HRA/04/0093' """,as_dict = True)
				
#Update the employee number While click the update employee number button in employee
@frappe.whitelist()
def update_employee_no(name,employee_number):
	emp = frappe.get_doc("Employee",name)
	emps=frappe.get_all("Employee",{"status":"Active"},['*'])
	for i in emps:
		if emp.employee_number == employee_number:
			pass
		elif i.employee_number == employee_number:
			frappe.throw(f"Employee Number already exists for {i.name}")
		else:
			frappe.db.set_value("Employee",name,"employee_number",employee_number)
			frappe.rename_doc("Employee", name, employee_number, force=1)
			return employee_number


#Update the work end date while enter the value of expected_tenure sales order
@frappe.whitelist()
def get_expected_dates(expect):
	exp_date = ''
	if expect == "In 15 Days":
		exp_date = add_days(nowdate(), 15)
		
	elif expect == "In 30 Days":
		exp_date = add_days(nowdate(), 30)
	
	elif expect == "In 45 Days":
		exp_date = add_days(nowdate(), 45)
	
	elif expect == "In 60 Days":
		exp_date = add_days(nowdate(), 60)
	
	elif expect == "End of next Month":
		exp_date = get_last_day(add_months(nowdate(), 1))

	return exp_date

#Update the due date while enter the value of payment terms in Fuel Maintenance
@frappe.whitelist()
def get_due_dates(payment,bill):
	due_date = ''
	if payment == "Due on receipt":
		due_date = bill

	elif payment == "Due end of next month":
		due_date = get_last_day(add_months(nowdate(), 1))

	elif payment == "Net 15":
		due_date = add_days(nowdate(), 15)

	elif payment == "Net 30":
		due_date = add_days(nowdate(), 30)

	elif payment == "Net 60":
		due_date = add_days(nowdate(), 60)

	elif payment == "Due end of the month":
		due_date = get_last_day(nowdate())

	return due_date

#Get the below values and set to SI
@frappe.whitelist()
def return_item_values(name,item_code):
	item_value=frappe.get_all("Sales Invoice Item",{'parent':name,'item_code':item_code},['qty','amount','base_amount','rate','base_rate','net_rate','base_net_rate','net_amount','base_net_amount'])
	for i in item_value:
		return i.qty,i.amount,i.base_amount,i.rate,i.base_rate,i.net_rate,i.base_net_rate,i.net_amount,i.base_net_amount
	
		
#Create new document while click the material request button for SO
@frappe.whitelist()
def get_single_trip(sale):
	order = frappe.get_doc("Sales Order",sale)
	
	if order.work_end_date:
		new_doc = frappe.new_doc("Trip")
		new_doc.customer = order.customer
		new_doc.work_order_no = sale
		new_doc.distance_map = order.distance_map
		new_doc.billing = order.billing
		new_doc.company = order.company
		new_doc.rate_per_trip = order.rate_per_trip
		new_doc.total_trip1 = order.total_trip
		new_doc.total_amount = order.total_amount
		new_doc.from_location = order.destination[0].location_place
		new_doc.to_location = order.destination[-1].location_place
		for i in order.destination:
			new_doc.append('destination', {
				'location_place': i.location_place,
				'latitude_longitude':i.latitude_longitude,
				'point':i.point
			})

		for j in order.items:
			new_doc.append('trip_item', {
				'item_code': j.item_code,
				'customer_item_code':j.customer_item_code,
				'item_name':j.item_name,
				'delivery_date':j.delivery_date,
				'item_group':j.item_group,
				'description':j.description,
				'qty':j.qty,
				'uom':j.uom,
				'rate':j.rate,
				'amount':j.amount,
				'base_rate':j.base_rate,
				'base_amount':j.base_amount,
				'net_rate':j.net_rate,
				'net_amount':j.net_amount,
				'base_net_rate':j.base_net_rate,
				'base_net_amount':j.base_net_amount,
				'valuation_rate':j.valuation_rate,
				'warehouse':j.warehouse,
				'conversion_factor':j.conversion_factor,
				'prevdoc_docname':j.prevdoc_docname,
	
			})
			
		new_doc.save(ignore_permissions=True)
	return "ok"

#Create new document while click the Multi Trip button for SO
@frappe.whitelist()
def create_multi(items, sale):
	order = frappe.get_doc("Sales Order", sale)
	item = int(items)
	for k in range(item):
		if order.work_end_date:
			new_doc = frappe.new_doc("Trip")
			new_doc.customer = order.customer
			new_doc.work_order_no = sale
			new_doc.distance_map = order.distance_map
			new_doc.billing = order.billing
			new_doc.company = order.company
			new_doc.rate_per_trip = order.rate_per_trip
			new_doc.total_trip1 = order.total_trip
			new_doc.total_amount = order.total_amount
			new_doc.from_location = order.destination[0].location_place
			new_doc.to_location = order.destination[-1].location_place
			for i in order.destination:
				new_doc.append('destination', {
					'location_place': i.location_place,
					'latitude_longitude': i.latitude_longitude,
					'point': i.point
				})

			for j in order.items:
				new_doc.append('trip_item', {
					'item_code': j.item_code,
					'customer_item_code':j.customer_item_code,
					'item_name':j.item_name,
					'item_group':j.item_group,
					'description':j.description,
					'uom':j.uom,
					'rate':j.rate,
					'base_net_rate':j.base_net_rate,
					'valuation_rate':j.valuation_rate,
					'warehouse':j.warehouse,
					'conversion_factor':j.conversion_factor,
					'prevdoc_docname':j.prevdoc_docname,
		
				})
			
			new_doc.insert(ignore_permissions=True)
	return "ok"


from electra.utils import get_series

@frappe.whitelist()
def get_si_doc(sales):
	trip = frappe.get_doc("Trip", sales)
	if trip.trip_item:
		si = frappe.new_doc("Sales Invoice")
		si.customer = trip.customer
		si.company = "KINGFISHER - TRANSPORTATION"
		si.naming_series = get_series("KINGFISHER - TRANSPORTATION", "Sales Invoice")
		for j in trip.trip_item:
			si.append('items', {
				'item_code': j.item_code,
				'customer_item_code': j.customer_item_code,
				'item_name': j.item_name,
				'delivery_date': j.delivery_date,
				'item_group': j.item_group,
				'description': j.description,
				'qty': j.qty,
				'uom': j.uom,
				'rate': j.rate,
				'amount': j.amount,
				'base_rate': j.base_rate,
				'base_amount': j.base_amount,
				'net_rate': j.net_rate,
				'net_amount': j.net_amount,
				'base_net_rate': j.base_net_rate,
				'base_net_amount': j.base_net_amount,
				'valuation_rate': j.valuation_rate,
				'warehouse': j.warehouse,
				'conversion_factor': j.conversion_factor,
				'prevdoc_docname': j.prevdoc_docname,
			})
		si.save(ignore_permissions=True)
		frappe.db.commit()
	return "ok"

#Get the below values and set to Sales Invoice Items
@frappe.whitelist()
def get_child_si(trip):
	dict_list = []
	trip_new = frappe.get_doc("Trip",trip)
	for i in trip_new.trip_item:
		dict_list.append(frappe._dict({"item_code":i.item_code,"qty":i.qty,"item_name":i.item_name,"item_group":i.item_group,"description": i.description,"uom":i.uom,"rate":i.rate,"amount":i.amount,"valuation_rate":i.valuation_rate,"base_rate":i.base_rate,"base_amount":i.base_amount,"net_rate":i.net_rate,"net_amount":i.net_amount,"base_net_rate":i.base_net_rate,"base_net_amount":i.base_net_amount,"warehouse":i.warehouse}))
	return dict_list


#Create new journal entry
@frappe.whitelist()
def create_journal_for_retention(doc,method):
	if doc.ret_amount > 0:
		jv = frappe.new_doc("Journal Entry")
		jv.voucher_type = "Journal Entry"
		jv.company = doc.company
		jv.posting_date = nowdate()
		jv.append("accounts", {
			"account": frappe.db.get_value("Account",filters={'name': ['like', '%Retention -%'],'company':doc.company}),
			"party_type": "Customer",
			"party": doc.customer,
			"debit_in_account_currency": doc.ret_amount,
			"cost_center": erpnext.get_default_cost_center(doc.company),
		})

		jv.append("accounts", {
			"account": frappe.db.get_value("Account",filters={'name': ['like', '%4110 - Sales -%'],'company':doc.company}),
			"credit_in_account_currency": doc.ret_amount,
		})
		jv.save(ignore_permissions=True)


# @frappe.whitelist()
# def send_email_on_pending_hod(name):
	hod_users = frappe.get_all(
		'User',
		filters={'enabled': 1},
		fields=['name', 'full_name', 'email'],
	)
	mr_list = frappe.get_all("Material Request", {'name': name}, ['company'])
	
	for user in hod_users:
		if 'HOD' in frappe.get_roles(user['name']):
			companies = frappe.get_all(
				"User Permission", {'user': user['email'], 'allow': 'Company'}, ['for_value'])
			unique_documents = set()
			for company in companies:
				if mr_list and mr_list[0].get('company') == company.get('for_value'):
					unique_documents.add(name)

			if unique_documents:
				data = ''
				data += f'Dear {user["full_name"]},<br><br>Kindly Find the below Document Moved To Pending for HOD<br>'
				data += '<table class="table table-bordered"><tr><th>Document Name</th></tr>'

				for document_name in unique_documents:
					document_link = f'https://erp.electraqatar.com/app/material-request/{document_name}'
					data += '<tr><td><a href="{0}">{1}</a></td></tr>'.format(document_link, document_name)

				data += '</table>'

				frappe.sendmail(
					recipients=user['email'],
					subject=('Documents Moved to Pending for HOD'),
					message=data
				)


@frappe.whitelist()
def update_company():
	company = frappe.get_all("Stock Confirmation",['name','source_company'])
	for i in company:
		frappe.db.set_value("Stock Confirmation",i.name,"from__company",i.source_company)

@frappe.whitelist()
def update_workflow():
	value = frappe.get_all("Leave Application",{"status":"Cancelled"},["name"])
	for i in value:
		frappe.db.set_value("Leave Application",i.name,"workflow_state","Cancelled")

#Update the below value for on submission of SI
@frappe.whitelist()
def update_return_doc(doc, method):
	if doc.return_against:      
	   frappe.db.set_value("Sales Invoice",doc.return_against,"return_document",doc.name)

@frappe.whitelist()
def update_return_documents():
	document = frappe.get_all("Sales Invoice",['return_against','name'])
	for i in document:
		if i.return_against:   
			frappe.db.set_value("Sales Invoice",i.return_against,"return_document",i.name)
	
#Update the below value is on update of SI
@frappe.whitelist()
def update_ret_out(doc,method):
	if doc.order_type == "Project":
		outstanding_without_retention = doc.outstanding_amount -doc.ret_amount
		doc.outstanding_amount_without_retention = outstanding_without_retention


# calculate the total amount for trip 
@frappe.whitelist(allow_guest=True)
def total_wh_hrs(in_time,out_time):
	inti = frappe.utils.get_datetime(in_time)
	out_ti = frappe.utils.get_datetime(out_time)
	if inti and out_ti:
		wh = time_diff_in_hours(out_ti,inti)
		return wh

#Run the below method for cron create_leave_allocation
@frappe.whitelist()
def update_leave_policy():
	current_date = datetime.now()
	start_of_year = date(current_date.year, 1, 1)
	end_of_year = date(current_date.year, 12, 31)
	leave = frappe.get_all("Leave Policy Detail", ["leave_type", "annual_allocation"])
	for i in leave:
		if i.leave_type =="Medical Leave" or i.leave_type =="Annual Leave":
			employees = frappe.get_all("Employee",{"status": "Active"},["employee_number","company"])
			for emp in employees:
				allocation = frappe.new_doc("Leave Allocation")
				allocation.employee = emp.employee_number
				allocation.leave_type = i.leave_type
				allocation.new_leaves_allocated = i.annual_allocation
				allocation.total_leaves_allocated = i.annual_allocation
				allocation.from_date=start_of_year
				allocation.to_date = end_of_year
				allocation.company = emp.company
				allocation.save(ignore_permissions=True)
				allocation.submit()
	frappe.db.commit()


@frappe.whitelist()
def update_si_details():
	value = frappe.get_all("Purchase Invoice",["name","stock_confirmation"])
	for i in value:
		if not i.stock_confirmation:
			frappe.db.set_value("Purchase Invoice",i.name,"stock_confirmation","Stock Transfer")

# @frappe.whitelist()
# def receivable_report(doc):
#     if doc.project:
#         data = "<table width=100% border=1px solid black><tr style=background-color:#e35310;font-size:8px><td colspan=1 ><b style=color:white; text-align:center;width:320px>Date</b></td><td colspan=2  style=color:white><b style=color:white; text-align:center;>Referance No</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Voucher Type</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Particulars</b></td><td colspan=1 style=color:white><b style=color:white; text-align:center;>Invoice Amount</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Advance</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Retention</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Net Amount</b></td><td  style=color:white><b style=color:white; text-align:center;>Received</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Balance Due</b></td></tr>"
#         pay = frappe.get_all("Payment Entry Reference",{"reference_name":doc.so_no},['parent'])
#         for i in pay:
#             pay_entry = frappe.get_all("Payment Entry",{"name":i.parent},["posting_date","paid_amount","remarks","status"],order_by="posting_date")
#             for j in pay_entry:
#                 if j.status =="Submitted":
#                     data += f'<tr><td colspan=1 style="border: 1px solid black; font-size:8px" nowrap>{j.posting_date.strftime("%d-%m-%Y")}</td><td colspan=2 style="border: 1px solid black; font-size:8px" nowrap>{i.parent}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >Payment Entry</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{j.remarks}</td><td colspan=1 style="border: 1px solid black; font-size:8px">-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{j.paid_amount}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td></tr>'
#         si  = frappe.get_all("Sales Invoice",{"sales_order":doc.so_no},['name','posting_date','total','adv_amount','ret_amount','net_total_project'],order_by="posting_date")
#         for k in si:
#             data += f'<tr><td colspan=1 style="border: 1px solid black; font-size:8px" nowrap>{k.posting_date.strftime("%d-%m-%Y")}</td><td colspan=2 style="border: 1px solid black; font-size:8px" nowrap>{k.name}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >Sales Invoice</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px">{k.total}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{k.adv_amount}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{k.ret_amount}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{k.net_total_project}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td></tr>'
#             si_pay= frappe.get_all("Payment Entry Reference",{"reference_name":k.name},['parent'])
#             for s in si_pay:
#                 si_pay_entry = frappe.get_all("Payment Entry",{"name":s.parent},["posting_date","paid_amount","remarks","status"],order_by="posting_date")
#                 for v in si_pay_entry:
#                     if v.status =="Submitted":
#                         data += f'<tr><td colspan=1 style="border: 1px solid black; font-size:8px" nowrap>{v.posting_date.strftime("%d-%m-%Y")}</td><td colspan=2 style="border: 1px solid black; font-size:8px" nowrap>{s.parent}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >Payment Entry</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{v.remarks}</td><td colspan=1 style="border: 1px solid black; font-size:8px">-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{v.paid_amount}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td></tr>'
#     else:
#         data = "<table width=100% border=1px solid black><tr style=background-color:#e35310;font-size:10px;text-align:center><td colspan=1 ><b style=color:white; text-align:center;width:320px>Sales Order Documents Not Avaliable</b></td></tr>"
#     data += '</table>'
#     return data


@frappe.whitelist()
def update_invoice():
	dn = frappe.get_all("Sales Invoice Item",["delivery_note","parent"])
	for j in dn:
		frappe.db.set_value("Delivery Note",j.delivery_note,"custom_invoice_number",j.parent)

#Update the Sales invoice number in DN
@frappe.whitelist()
def update_invoice_number(doc,method):
	for j in doc.items:
		frappe.db.set_value("Delivery Note",j.delivery_note,"custom_invoice_number",doc.name)

@frappe.whitelist()
def update_terms():
	quot = frappe.get_all("Quotation",{"tc_name":"Trading Div Sales",'docstatus':('!=','2')},["name"])
	for i in quot:
		frappe.db.set_value("Quotation",i.name,"tc_name","Trading Div Sales")



@frappe.whitelist()
def list_workspace():
	list = frappe.get_doc("Workspace","Accounting")
	for i in list.shortcuts:
		if i.link_to == "General Ledger":
			print(i.name)
			# # i.link_to = "General Ledger Report"
			# # i.save()
			frappe.db.set_value("Workspace Shortcut",'73c7d12c58','link_to',"General Ledger Report",update_modified=False)
			

# @frappe.whitelist()
# def create_employee():
# 	print("HI")
# 	emp = frappe.get_all("Employee Grade")
# 	print(emp)

#Update the value based on below condition while enter the payment terms
@frappe.whitelist()
def update_isadvance(payment_terms,name):
	payment_terms=frappe.db.get_value("Payment Schedule Sales Order",{"payment_terms":['like',"%Advance%"],'parent':name},'name')
	if payment_terms:
		frappe.db.set_value("Payment Schedule Sales Order",
						{"parent": name, "payment_terms": ["like", "%Advance%"]},
						"is_advance", 1)
		return 'OK'

#Get the company while enter the applicant name of the loan application
@frappe.whitelist()
def get_employee_company(employee):
	company = frappe.db.get_value("Employee",{"name":employee},["company"])
	return company

#Sent mail for the below details
@frappe.whitelist()
def day_book():
	table_html = ''
	yesterday_date = (datetime.now() - timedelta(days=1)).date()
	si_list = frappe.db.sql("""SELECT name, posting_date, paid_amount FROM `tabSales Invoice` WHERE posting_date = %s AND docstatus != 2 AND invoice_type = 'Cash' AND is_paid = 1""", (yesterday_date,), as_dict=True)
	if si_list:
		header = f"""<p>Dear Sir/Mam, <br> Please find the below list of Day Book As On {yesterday_date}</p><table class='table table-bordered'>"""
		regards = "Thanks & Regards"
		table_html += '<table class="table table-bordered" style="width:100%;">'
		table_html += '<tr>'
		table_html += '<th>Date</th>'
		table_html += '<th>Particulars</th>'
		table_html += '<th>Debit</th>'
		table_html += '<th>Credit</th>'
		table_html += '</tr>'
		total=0
		for si in si_list:
			total += si.paid_amount
			table_html += '<tr>'
			table_html += '<td>{}</td>'.format(si.posting_date)
			table_html += '<td>{}</td>'.format(si.name)
			table_html += '<td>{}</td>'.format(si.paid_amount)
			table_html += '<td></td>'
			table_html += '</tr>'
		table_html += '<tr>'
		table_html += '<td></td>'
		table_html += '<td><b>Total</b></td>'
		table_html += '<td>{}</td>'.format(total)
		table_html += '<td></td>'
		table_html += '</tr>'
		table_html += '</table><br>'
		subject = 'Reg. DAYBOOK AS ON {}'.format(yesterday_date)
		frappe.sendmail(
			recipients='',
			subject=subject,
			message=header + table_html + regards
		)


# @frappe.whitelist()
# def sales_order_duplicate(sales_order):
# 	childtab = frappe.db.sql(""" select `tabSales Order Item`.item_code,`tabSales Order Item`.is_free,`tabSales Order Item`.warehouse
# 	`tabSales Order Item`.item_name,`tabSales Order Item`.description,`tabSales Order Item`.msow,
# 	`tabSales Order Item`.prevdoc_docname,
# 	`tabSales Order Item`.base_rate,`tabSales Order Item`.conversion_factor,
# 	sum(`tabSales Order Item`.base_amount) as base_amount,
# 	`tabSales Order Item`.net_rate,
# 	`tabSales Order Item`.base_net_rate,
# 	sum(`tabSales Order Item`.net_amount) as net_amount,
# 	sum(`tabSales Order Item`.base_net_amount) as base_net_amount,
# 	`tabSales Order Item`.valuation_rate,
# 	`tabSales Order Item`.delivery_date,sum(`tabSales Order Item`.qty) as qty,
# 	`tabSales Order Item`.uom,`tabSales Order Item`.rate,`tabSales Order Item`.warehouse,
# 	`tabSales Order Item`.bom_no,sum(`tabSales Order Item`.amount) as amount from `tabSales Order` 
# 	left join `tabSales Order Item` on `tabSales Order`.name = `tabSales Order Item`.parent where `tabSales Order`.name = '%s' group by `tabSales Order Item`.item_code,`tabSales Order Item`.msow order by `tabSales Order Item`.idx """%(sales_order),as_dict = 1)
# 	return childtab

#Check the duplicate item in item table
@frappe.whitelist()
def sales_order_duplicate(sales_order):
	childtab = frappe.db.sql("""
		SELECT
			`tabSales Order Item`.item_code,
			`tabSales Order Item`.warehouse,
			`tabSales Order Item`.item_name,
			`tabSales Order Item`.description,
			`tabSales Order Item`.msow,
			`tabSales Order Item`.prevdoc_docname,
			`tabSales Order Item`.base_rate,
			`tabSales Order Item`.conversion_factor,
			SUM(`tabSales Order Item`.base_amount) as base_amount,
			`tabSales Order Item`.net_rate,
			`tabSales Order Item`.base_net_rate,
			SUM(`tabSales Order Item`.net_amount) as net_amount,
			SUM(`tabSales Order Item`.base_net_amount) as base_net_amount,
			`tabSales Order Item`.valuation_rate,
			`tabSales Order Item`.delivery_date,
			SUM(`tabSales Order Item`.qty) as qty,
			`tabSales Order Item`.uom,
			`tabSales Order Item`.rate,
			`tabSales Order Item`.warehouse,
			`tabSales Order Item`.bom_no,
			SUM(`tabSales Order Item`.amount) as amount
		FROM
			`tabSales Order`
		LEFT JOIN
			`tabSales Order Item` ON `tabSales Order`.name = `tabSales Order Item`.parent
		WHERE
			`tabSales Order`.name = '%s'
		GROUP BY
			`tabSales Order Item`.item_code, `tabSales Order Item`.msow
		ORDER BY
			`tabSales Order Item`.idx
	""" % (sales_order), as_dict=1)
	return childtab


#While submission of sales order check the linked project budget status , if not approved the PB throw the below message
@frappe.whitelist()
def validate_project_budget(doc,method):
	if doc.order_type == 'Project':
		workflow_state = frappe.db.get_value("Project Budget",doc.project_budget,['workflow_state'])
		if workflow_state != 'Approved':
			frappe.throw("Project Budget has to be Approved to submit Sales Order")
			
# @frappe.whitelist()
# def statement_of_account(company, from_date, to_date, customer):
#     data = ''

#     data += "<table border='1px solid black' width='100%' style='margin-left:2px;margin-right:2px;'><tr style='font-size:10px;background-color:#D3D3D3'><td width=10%><b>Date</b></td><td width=10%><b style='text-align:center;'>Voucher Type</b></td><td width=10%><b style='text-align:center'>Voucher No</b></td><td width=25%><b style='text-align:center'>Remarks</b></td><td width=10%><b style='text-align:center'>Debit(QAR)</b></td><td width=10%><b style='text-align:center'>Credit(QAR)</b></td><td width=10%><b style='text-align:center'>Balance(QAR)</b></td></tr>"
#     if customer:
#         gl_entry = frappe.db.sql("""select voucher_type,voucher_no,posting_date,sum(debit) as debit,sum(credit) as credit from `tabGL Entry` where company = %s and posting_date between %s and %s and is_cancelled = 0 and party = %s  and party_type ='Customer' group by voucher_no order by posting_date""", (company, from_date, to_date, customer), as_dict=True)
#         gle = frappe.db.sql("""select sum(debit) as opening_debit, sum(credit) as opening_credit from `tabGL Entry` where company = '%s'	and (posting_date < '%s' or (ifnull(is_opening, 'No') = 'Yes' and posting_date >= '%s')) and party = '%s' and is_cancelled = 0 and party_type ='Customer' """%(company,from_date,to_date,customer),as_dict=True)
#         frappe.log_error(f'Value: {gle}')
#     else:
#         gl_entry = frappe.db.sql("""select voucher_type,voucher_no,posting_date,sum(debit) as debit,sum(credit) as credit from `tabGL Entry` where company = %s and posting_date between %s and %s and is_cancelled = 0 group by voucher_no order by posting_date""", (company, from_date, to_date), as_dict=True)
#         gle = frappe.db.sql("""select sum(debit) as opening_debit, sum(credit) as opening_credit from `tabGL Entry` where company = '%s'	and (posting_date < '%s' or (ifnull(is_opening, 'No') = 'Yes' and posting_date >= '%s')) and is_cancelled = 0  """%(company,from_date,to_date),as_dict=True)

#     opening_balance = 0
#     t_p_debit = 0
#     t_p_credit = 0
	
#     for g in gle:
#         if not g.opening_debit:
#             g.opening_debit = 0
#         if not g.opening_credit:
#             g.opening_credit = 0
#         t_p_debit += g.opening_debit
#         t_p_credit += g.opening_credit
#         opening_balance = t_p_debit - t_p_credit
#     data += f'<tr style="font-size:10px"><td colspan =6 style="text-align:right" width=85%><b>Opening Balance</b></td></td><td style="text-align:right" width=10%><b>{fmt_money(round(opening_balance,2))}</b></td></tr>'
#     balance=opening_balance
#     for i in gl_entry:
#         balance += (i.debit -i.credit)
#         if i.voucher_type == "Payment Entry":
#             ref_no = frappe.db.get_value("Payment Entry",{"name":i.voucher_no},['reference_no'])
#             if ref_no:
#                 check_no = ref_no
#             else:
#                 check_no = ''
#         else:
#             check_no = ''
#         if i.voucher_type == "Sales Invoice":
#             remarks = ''
#             remark = frappe.db.get_all("Sales Invoice", filters={"name": i.voucher_no}, fields=['*'])

#             if remark:
#                 remark_data = remark[0]
#                 dn = remark_data.get('delivery_note', '')
#                 po = remark_data.get('po_no', '')

#                 if dn and po:
#                     remarks = f"DN No.{dn} & LPO No.{po}"
#                 elif dn:
#                     remarks = f"DN No.{dn}"
#                 elif po:
#                     remarks =f"LPO No.{po}"
#         else:
#             remarks = ''

#         data += f'<tr style="font-size:10px"><td width=10% nowrap>{i.posting_date.strftime("%d-%m-%Y")}</td><td width=10%>{i.voucher_type}</td><td width=10%>{i.voucher_no}</td><td width=25%>{remarks}{check_no}</td><td width=10% style="text-align:right">{fmt_money(round(i.debit,2)) or "-"}</td><td width=10% style="text-align:right">{fmt_money(round(i.credit,2)) or "-"}</td><td style="text-align:right" width=10%>{fmt_money(round(balance,2))}</td></tr>'
#     tp_credit=0
#     tp_debit=0
#     bal=0
#     for i in gl_entry:
#         tp_credit += i.credit 
#         t_p_credit += i.credit
#         tp_debit += i.debit
#         t_p_debit += i.debit
#         bal = tp_debit-tp_credit

#     data += f'<tr style="font-size:10px"><td colspan=4 style="text-align:right"><b>Total</b></td><td style="text-align:right"><b>{fmt_money(round(tp_debit,2))}</b></td><td style="text-align:right"><b>{fmt_money(round(tp_credit,2))}</b></td><td style="text-align:right"><b></b></td></tr>'
#     data += f'<tr style="font-size:10px"><td colspan =6 style="text-align:right" width=85%><b>Closing Balance</b></td></td><td style="text-align:right" width=10%><b>{fmt_money(round(balance,2))}</b></td></tr>'
#     data += '</table>'
#     return data


# @frappe.whitelist()
# def check_gle():
# 	acc ='COMMERCIAL BANK QATAR -4020 4361 81001'

# 	ac = '%'+acc+'%'
# 	sq = frappe.db.sql(""" select name,company,sum(debit_in_account_currency) as debit,sum(credit_in_account_currency) as credit from `tabGL Entry` where company = '%s' and account like '%s' """%("KINGFISHER - SHOWROOM","COMMERCIAL BANK QATAR -4020 4361 81001"),as_dict=True)


# @frappe.whitelist()
# def ageing_report(doc):
#     in_amount = 0
#     paid_amount = 0
#     credit_note = 0
#     out_amount = 0
#     age_0_30 = 0
#     age_31_60 = 0
#     age_61_90 = 0
#     age_91_120 = 0
#     age_above_121 = 0
#     paid = 0
#     data = "<table border='1px solid black' width='100%'><tr style='font-size:12px'><td width=10%><b>Posting Date</b></td><td width=10%><b style='text-align:center;'>Voucher No</b></td><td width=10%><b style='text-align:center'>Customer LPO</b></td><td width=10%><b style='text-align:center'>Invoiced Amount</b></td><td width=10%><b style='text-align:center'>Paid Amount</b></td><td width=10%><b style='text-align:center'>Credit Note</b></td><td width=10%><b style='text-align:center'>Outstanding Amount</b></td><td width=5%><b style='text-align:center'>Age (Days)</b></td><td width=5%><b style='text-align:center'>0- 30</b></td><td width=5%><b style='text-align:center'>31-  60</b></td><td width=5%><b style='text-align:center'>61-  90</b></td><td width=5%><b style='text-align:center'>91-120</b></td><td width=5%><b style='text-align:center'>Above 121</b></td></tr>"
#     for c in doc.company_multiselect:
#         if doc.customer:
#             si_list = frappe.db.sql(
#                 """SELECT * FROM `tabSales Invoice` WHERE company = %s and customer = %s and docstatus = 1 and is_return = 0 and invoice_type='Credit' and status !='Paid'  order by posting_date  ASC""",
#                 (c.company, doc.customer),
#                 as_dict=True
#             )
#             for i in si_list:
#                 result= frappe.db.sql("""
#                     SELECT sum(grand_total) as total 
#                     FROM `tabSales Invoice` 
#                     WHERE company = %s AND return_against = %s AND docstatus = 1
#                 """, (c.company, i.name))
#                 return_amount = result[0][0] if result and result[0][0] else 0
#                 # return_amount = frappe.db.get_value("Sales Invoice", {"name": i.return_document,"docstatus":("!=",2)}, ["grand_total"])
#                 pay = frappe.db.sql("""
#                     SELECT per.allocated_amount 
#                     FROM `tabPayment Entry Reference` AS per
#                     LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
#                     WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
#                 """, (c.company, i.name), as_dict=True)

#                 value = sum(j.allocated_amount for j in pay)

#                 jv = frappe.db.sql("""
#                     SELECT credit_in_account_currency 
#                     FROM `tabJournal Entry Account` AS per
#                     LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
#                     WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
#                 """, (c.company, i.name), as_dict=True)
#                 for k in jv:
#                     value += k.credit_in_account_currency
#                 if value and return_amount:
#                     outstanding = i.grand_total - value + return_amount
#                 elif value:
#                     outstanding = i.grand_total - value
#                 elif return_amount:
#                     outstanding = i.grand_total + return_amount
#                 else:
#                     outstanding = i.grand_total
				
#                 out_amount += outstanding
#                 age = date_diff(today(), i.posting_date) if i.posting_date else 0

#                 if round(outstanding) != 0:
#                     if value:
#                         paid_amount += value
#                     if return_amount:
#                         credit_note += return_amount
#                     in_amount += i.grand_total
#                     data += f"<tr style=font-size:10px><td width=10%>{formatdate(i.posting_date, 'dd-MM-yyyy')}</td><td width=10%>{i.name}</td><td width=10%>{i.po_no if i.po_no else '-'}</td><td width=10%>{fmt_money(round(i.grand_total, 2))}</td><td width=10%>{fmt_money(round(value, 2)) if value else '-'}</td><td width=10%>{fmt_money(round(return_amount, 2)) if return_amount else '-'}</td><td width=10%>{fmt_money(round(outstanding, 2)) if outstanding else '-'}</td><td width=5%>{age}</td>"

#                     if 0 <= age <= 30:
#                         age_0_30 += outstanding
#                     elif 31 <= age <= 60:
#                         age_31_60 += outstanding
#                     elif 61 <= age <= 90:
#                         age_61_90 += outstanding
#                     elif 91 <= age <= 120:
#                         age_91_120 += outstanding
#                     else:
#                         age_above_121 += outstanding

#                     data += f"<td width=5%>{fmt_money(round(outstanding, 2)) if 0 <= age <= 30 else '-'}</td><td width=5%>{fmt_money(round(outstanding, 2)) if 31 <= age <= 60 else '-'}</td><td width=5%>{fmt_money(round(outstanding, 2)) if 61 <= age <= 90 else '-'}</td><td width=5%>{fmt_money(round(outstanding, 2)) if 91 <= age <= 120 else '-'}</td><td width=5%>{fmt_money(round(outstanding, 2)) if age > 120 else '-'}</td></tr>"

#             sales = frappe.db.sql(
#                 """SELECT * FROM `tabSales Invoice` WHERE company = %s and customer = %s and docstatus = 1 and is_return = 1 and invoice_type='Credit' and status !='Paid' AND (return_against IS NULL OR return_against = '')  order by posting_date  ASC""",
#                 (c.company, doc.customer),
#                 as_dict=True
#             )
#             for i in sales:
#                 pay = frappe.db.sql(""" SELECT per.allocated_amount FROM `tabPayment Entry Reference` AS per
#                 LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
#                 WHERE per.reference_name = '%s' AND pe.docstatus = 1 and pe.company = '%s' """ % (i.name, c.company), as_dict=True)
#                 value = sum(j.allocated_amount for j in pay)

#                 jv = frappe.db.sql(""" SELECT credit_in_account_currency FROM `tabJournal Entry Account` AS per
#                 LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
#                 WHERE per.reference_name = '%s' AND pe.docstatus = 1 and pe.company = '%s'""" % (i.name, c.company), as_dict=True)
#                 value += sum(k.credit_in_account_currency for k in jv)

#                 outstanding = i.grand_total - value if value else i.grand_total
#                 out_amount += outstanding
#                 age = date_diff(today(), i.posting_date) if i.posting_date else 0

#                 if round(outstanding) != 0:
#                     if value:
#                         paid_amount += value
#                     if i.grand_total:
#                         credit_note += i.grand_total
#                     in_amount += i.grand_total
#                     data += f"<tr style=font-size:10px><td width=10%>{formatdate(i.posting_date, 'dd-MM-yyyy')}</td><td width=10%>{i.name}</td><td width=10%>{i.po_no if i.po_no else '-'}</td><td width=10%>{fmt_money(round(i.grand_total, 2))}</td><td width=10%>{fmt_money(round(value, 2)) if value else '-'}</td><td width=10%>{fmt_money(round(i.grand_total, 2)) if i.grand_total else '-'}</td><td width=10%>{fmt_money(round(outstanding, 2)) if outstanding else '-'}</td><td width=5%>{age}</td>"

#                     if 0 <= age <= 30:
#                         age_0_30 += outstanding
#                     elif 31 <= age <= 60:
#                         age_31_60 += outstanding
#                     elif 61 <= age <= 90:
#                         age_61_90 += outstanding
#                     elif 91 <= age <= 120:
#                         age_91_120 += outstanding
#                     else:
#                         age_above_121 += outstanding

#                     data += f"<td width=5%>{fmt_money(round(outstanding, 2)) if 0 <= age <= 30 else '-'}</td><td width=5%>{fmt_money(round(outstanding, 2)) if 31 <= age <= 60 else '-'}</td><td width=5%>{fmt_money(round(outstanding, 2)) if 61 <= age <= 90 else '-'}</td><td width=5%>{fmt_money(round(outstanding, 2)) if 91 <= age <= 120 else '-'}</td><td width=5%>{fmt_money(round(outstanding, 2)) if age > 120 else '-'}</td></tr>"

#             payment = frappe.db.sql("""
#                 SELECT * FROM `tabPayment Entry` 
#                 WHERE company = %s AND party = %s AND docstatus = 1 
#                 AND payment_type = 'Receive' 
#                 ORDER BY posting_date ASC
#             """, (c.company, doc.customer), as_dict=True)

#             for v in payment:
#                 unallocated_amount = v.unallocated_amount
#                 paid_amount += unallocated_amount
#                 out_amount -= unallocated_amount
#                 age = date_diff(today(), v.posting_date)
#                 if unallocated_amount != 0:
#                     data += f"<tr style=font-size:10px><td width=10%>{formatdate(v.posting_date, 'dd-MM-yyyy')}</td><td width=10%>{v.name}</td><td width=10%>{v.reference_no if v.reference_no else '-'}</td><td width=10%>-</td><td width=10%>{fmt_money(round(unallocated_amount, 2),)}</td><td width=10%>-</td><td width=10%>{fmt_money(round(-unallocated_amount, 2),)}</td><td width=5%>{age}</td>"
					
#                     if 0 <= age <= 30:
#                         age_0_30 -= unallocated_amount
#                     elif 31 <= age <= 60:
#                         age_31_60 -= unallocated_amount
#                     elif 61 <= age <= 90:
#                         age_61_90 -= unallocated_amount
#                     elif 91 <= age <= 120:
#                         age_91_120 -= unallocated_amount
#                     else:
#                         age_above_121 -= unallocated_amount

#                     data += f"<td width=5%>{fmt_money(round(-unallocated_amount, 2),) if 0 <= age <= 30 else '-'}</td><td width=5%>{fmt_money(round(-unallocated_amount, 2),) if 31 <= age <= 60 else '-'}</td><td width=5%>{fmt_money(round(-unallocated_amount, 2),) if 61 <= age <= 90 else '-'}</td><td width=5%>{fmt_money(round(-unallocated_amount, 2),) if 91 <= age <= 120 else '-'}</td><td width=5%>{fmt_money(round(-unallocated_amount, 2),) if age > 120 else '-'}</td></tr>"
			
#             journal = frappe.db.sql("""
#                 SELECT per.credit_in_account_currency, per.debit_in_account_currency, pe.name, pe.posting_date 
#                 FROM `tabJournal Entry Account` AS per
#                 LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
#                 WHERE pe.company = %s AND per.account LIKE %s AND pe.docstatus = 1 
#                 AND party_type = 'Customer' AND party = %s AND per.reference_name IS NULL
#             """, (c.company, '%Debtors -%', doc.customer), as_dict=True)
#             for jour in journal:
#                 if jour.credit_in_account_currency:
#                     journ_amount_credit = jour.credit_in_account_currency
#                     paid_amount += journ_amount_credit
#                     in_amount -= journ_amount_credit
#                     out_amount -= journ_amount_credit
#                     age = date_diff(today(), jour.posting_date)

#                     if 0 <= age <= 30:
#                         age_0_30 -= jour.credit_in_account_currency
#                     elif 31 <= age <= 60:
#                         age_31_60 -= jour.credit_in_account_currency
#                     elif 61 <= age <= 90:
#                         age_61_90 -= jour.credit_in_account_currency
#                     elif 91 <= age <= 120:
#                         age_91_120 -= jour.credit_in_account_currency
#                     else:
#                         age_above_121 -= jour.credit_in_account_currency
#                     data += f"<tr style='font-size:10px'><td width=10%>{formatdate(jour.posting_date, 'dd-MM-yyyy')}</td><td width=10%>{jour.name}</td><td width=10%></td><td width=10%>{fmt_money(round(-jour.credit_in_account_currency, 2))}</td><td width=10%></td><td width =10%></td><td width=10%>{fmt_money(round(-jour.credit_in_account_currency, 2))}</td><td width=5%>{age}</td><td width=5%>{fmt_money(round(-jour.credit_in_account_currency, 2) if 0 <= age <= 30 else '-')}</td><td width=5%>{fmt_money(round(-jour.credit_in_account_currency, 2) if 31 <= age <= 60 else '-')}</td><td width=5%>{fmt_money(round(-jour.credit_in_account_currency, 2) if 61 <= age <= 90 else '-')}</td><td width=5%>{fmt_money(round(-jour.credit_in_account_currency, 2) if 91 <= age <= 120 else '-')}</td><td width=5%>{fmt_money(round(-jour.credit_in_account_currency, 2) if age > 120 else '-')}</td></tr>"
	
#                 elif jour.debit_in_account_currency:
#                     journ_amount_debit = jour.debit_in_account_currency
#                     in_amount += journ_amount_debit
#                     out_amount += journ_amount_debit
#                     age = date_diff(today(), jour.posting_date)

#                     if 0 <= age <= 30:
#                         age_0_30 += jour.debit_in_account_currency
#                     elif 31 <= age <= 60:
#                         age_31_60 += jour.debit_in_account_currency
#                     elif 61 <= age <= 90:
#                         age_61_90 += jour.debit_in_account_currency
#                     elif 91 <= age <= 120:
#                         age_91_120 += jour.debit_in_account_currency
#                     else:
#                         age_above_121 += jour.debit_in_account_currency
	
#                     data += f"<tr style='font-size:10px'><td width=10%>{formatdate(jour.posting_date, 'dd-MM-yyyy')}</td><td width=10%>{jour.name}</td><td width=10%></td><td width=10%>{fmt_money(round(jour.credit_in_account_currency, 2))}</td><td width=10%></td><td width =10%></td><td width=10%>{fmt_money(round(jour.credit_in_account_currency, 2))}</td><td width=5%>{age}</td><td width=5%>{fmt_money(round(jour.credit_in_account_currency, 2) if 0 <= age <= 30 else '-')}</td><td width=5%>{fmt_money(round(jour.credit_in_account_currency, 2) if 31 <= age <= 60 else '-')}</td><td width=5%>{fmt_money(round(jour.credit_in_account_currency, 2) if 61 <= age <= 90 else '-')}</td><td width=5%>{fmt_money(round(jour.credit_in_account_currency, 2) if 91 <= age <= 120 else '-')}</td><td width=5%>{fmt_money(round(jour.credit_in_account_currency, 2) if age > 120 else '-')}</td></tr>"
#     data += f"<tr style='font-size:10px'><td width=10%></td><td width=10%></td><td width=10%><b>Total</b></td><td width=10%><b>{fmt_money(round(in_amount, 2))}</b></td><td width=10%><b>{fmt_money(round(paid_amount, 2))}</b></td><td width=10%><b>{fmt_money(round(credit_note, 2))}</b></td><td width=10%><b>{fmt_money(round(out_amount, 2))}</b></td><td width=5%></td><td width=5%><b>{fmt_money(round(age_0_30, 2))}</b></td><td width=5%><b>{fmt_money(round(age_31_60, 2))}</b></td><td width=5%><b>{fmt_money(round(age_61_90, 2))}</b></td><td width=5%><b>{fmt_money(round(age_91_120, 2))}</b></td><td width=5%><b>{fmt_money(round(age_above_121, 2))}</b></td></tr>"
#     data += "</table>"
#     return data


# @frappe.whitelist()
# def unallocated_amount(doc):
# 	payment = frappe.get_all("Payment Entry",{"party":doc.customer,"docstatus":("!=",2)},["unallocated_amount","name","posting_date"])
# 	for i in payment:
# 		if i.unallocated_amount > 0:
# 			data =''
# 			data += "<table border='1px solid black' width='100%'><tr style='font-size:12px'><td width=40%><b>Reference Name</b></td><td width=30%><b style='text-align:center;'>Poting Date</b></td><td width=30%><b style='text-align:center'>Amount</b></td></tr>"
# 			data += f"<tr style=font-size:10px><td width=40%>{i.name or ''}</b></td><td width=30%>{formatdate(i.posting_date, 'dd-MM-yyyy')}</td><td width=30%>{i.unallocated_amount or ''}</td></tr>"
# 			data += '</table>'
# 			return data

#Calculate the payment days of leave salary while enter the vacation start date		
@frappe.whitelist()
def pay_days(start_date):
	date_object = datetime.strptime(start_date, "%Y-%m-%d")
	month_value = date_object.month
	date = f"2024-{month_value:02d}-01"
	diff = date_diff(start_date,date)
	return diff

#Calculate the payment days of leave salary while enter the vacation start date		
@frappe.whitelist()
def leave_salary_days_calculation(start_date,last_date,leave_days):
	total_days = date_diff( start_date,last_date)
	payment_days = int(total_days)
	return payment_days


@frappe.whitelist()
def get_employee_details(employee):
	values = frappe.get_all("Employee",{"name":employee},["*"])
	for i in values:
		return i.basic,i.hra,i._other_allowance

# @frappe.whitelist()
# def receipt_report(doc):
#     data = "<table border='1px solid black' width='100%'><tr><td style='text-align:center;'width='10%'><b>Posting Date</b></td><td style='text-align:center;'width='10%'><b>Voucher No</b></td><td style='text-align:center;'width='20%'><b>Party Name</b></td><td style='text-align:center;'width='20%'><b>Received Amount</b></td><td style='text-align:center;overflow-wrap: break-word;'width='10%'><b>Sales Person</b></td><td style='text-align:center;overflow-wrap: break-word;'width='30%'><b>Remarks</b></td></tr>"
#     sales_person = []
#     total = 0
#     ind = 0
#     sa = frappe.db.sql("""
#         SELECT * 
#         FROM `tabPayment Entry`
#         WHERE company = %s AND posting_date BETWEEN %s AND %s AND payment_type ='Receive' AND party_type = 'Customer' AND docstatus = 1 order by posting_date  ASC
#     """, (doc.company, doc.from_date, doc.to_date), as_dict=True)
	
#     journal = frappe.db.sql("""
#         SELECT * 
#         FROM `tabJournal Entry`
#         WHERE company = %s AND posting_date BETWEEN %s AND %s AND docstatus = 1 order by posting_date  ASC
#     """, (doc.company, doc.from_date, doc.to_date), as_dict=True)
	
#     for journ in journal:
#         doc_journal = frappe.get_all("Journal Entry Account", {"parent": journ.name,"party_type":"customer"}, ["party", "credit_in_account_currency"])
#         for c in doc_journal:
#             if c.credit_in_account_currency>0:
#                 ind += 1
#                 total += c.credit_in_account_currency
#                 data += f"<tr><td width='10%' nowrap>{formatdate(journ.posting_date, 'dd-MM-yyyy')}</td><td width='10%'>{journ.name}</td><td width='20%'>{c.party}</td><td style='text-align:right;'width='20%'>{fmt_money(round(c.credit_in_account_currency, 2))}</td><td style='overflow-wrap: break-word;'width='10%'></td><td style='text-align:left;overflow-wrap: break-word;'width='30%'>{journ.remarks or ''}</td></tr>"

#     for i in sa:
#         document = frappe.get_all("Payment Entry Reference", {"parent": i.name}, ["reference_doctype", "reference_name"])
#         if document:
#             for j in document:
#                 if j.reference_doctype == "Sales Order":
#                     sales_person = frappe.db.get_value("Sales Order", {"name": j.reference_name}, ["sales_person_user"])
#                 elif j.reference_doctype == "Sales Invoice":
#                     sales_person = frappe.db.get_value("Sales Invoice", {"name": j.reference_name}, ["sales_person_user"])
#         else:
#             sales_person =''
#         ind += 1
#         total += i.received_amount
#         data += f"<tr><td width='10%' nowrap>{formatdate(i.posting_date, 'dd-MM-yyyy')}</td><td width='10%'>{i.name}</td><td width='20%'>{i.party_name}</td><td style='text-align:right;'width='20%'>{fmt_money(round(i.received_amount, 2))}</td><td  style='text-align:left;overflow-wrap: break-word'width='10%'>{sales_person}</td><td style='text-align:left;overflow-wrap: break-word;'width='30%'>{i.remarks or ''}</td></tr>"
#     data += f"<tr><td width='10%'> </td><td width='10%'> </td><td width='20%'>Total</td><td style='text-align:right;'width='20%'>{fmt_money(round(total, 2))}</td><td style='overflow-wrap: break-word;'width='10%'></td><td width='30%'></td></tr>"
#     data += '</table>'
#     return data

from six import BytesIO, string_types

#Download the excel sheet for Purchase order details
@frappe.whitelist()
def make_po():
	args = frappe.local.form_dict
	filename = args.name
	test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
	ws = wb.create_sheet(sheet_name, 0)
	ws.column_dimensions['A'].width = 20
	ws.column_dimensions['B'].width = 20
	ws.column_dimensions['C'].width = 20 
	ws.column_dimensions['D'].width = 20
	ws.column_dimensions['E'].width = 20
	ws.column_dimensions['F'].width = 20 
	ws.column_dimensions['G'].width = 15
	ws.column_dimensions['H'].width = 15
	ws.column_dimensions['I'].width = 15
	ws.column_dimensions['J'].width = 15
	ws.column_dimensions['K'].width = 15
	doc = frappe.get_doc("Purchase Order",args.name)
	if doc:
		ws.append(["Purchase Order Number","Date","Material Request Number","Supplier Name","Supplier Address","Item Code","Description","Qty","UOM","Rate","Amount"])
		for i in doc.items:
			ws.append([doc.name, doc.transaction_date.strftime("%d-%m-%Y"), doc.material_request_number, doc.supplier, doc.address_display, i.item_code, i.description, i.qty, i.uom, i.base_rate, i.base_amount])

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'


# @frappe.whitelist()
# def add_salary(name, remark):
# 	additional = frappe.get_doc("Additional Salary", name)
# 	loan_repayment_doc = frappe.get_doc("Loan Repayment Schedule", additional.custom_loan_repayment)
# 	last_row = loan_repayment_doc.repayment_schedule[-1]
# 	next_month = frappe.utils.add_months(last_row.payment_date, 1)
# 	payroll_date = frappe.utils.get_first_day(next_month)  
# 	additional_salary_doc = frappe.new_doc("Additional Salary")
# 	additional_salary_doc.employee = additional.employee
# 	additional_salary_doc.salary_component = "Loan Deduction"
# 	additional_salary_doc.amount = additional.amount
# 	additional_salary_doc.custom_loan_repayment = additional.custom_loan_repayment
# 	additional_salary_doc.payroll_date = payroll_date
# 	additional_salary_doc.insert(ignore_permissions=True)
# 	additional_salary_doc.submit()
# 	frappe.db.sql("""DELETE FROM `tabRepayment Schedule` WHERE parent = %s""", (loan_repayment_doc.name,))
# 	additional_salaries = frappe.get_all("Additional Salary", {"custom_loan_repayment": additional.custom_loan_repayment, "docstatus":1})
# 	for salary in additional_salaries:
# 		repayment = frappe.get_doc("Loan Repayment Schedule", salary.custom_loan_repayment)
# 		repayment.append("repayment_schedule", {
# 			"payment_date": salary.posting_date
# 		})
# 		repayment.save(ignore_permissions=True)
# 		frappe.db.commit()

#Create additional salary for on submission of Loan Repayment Schedule
@frappe.whitelist()
def create_additional_salary(doc,method):
	loan_repayment_doc = frappe.get_doc("Loan Repayment Schedule",doc.name)
	table = loan_repayment_doc.repayment_schedule
	applicant = frappe.db.get_value("Loan",{"name":doc.loan},["applicant"])
	if doc.custom_custom_additional_created == 0 :
		for i in doc.repayment_schedule:
			add= frappe.new_doc("Additional Salary")
			add.employee = applicant
			add.salary_component = "Loan Deduction"
			add.payroll_date = i.payment_date
			add.amount = i.total_payment
			add.overwrite_salary_structure_amount=0
			add.custom_loan_repayment = doc.name
			add.custom_principal_amount = i.principal_amount
			add.custom_balance_total_amount = i.balance_loan_amount
			add.save(ignore_permissions=True)
			add.submit()
			i.additional_salary = add.name
	doc.save(ignore_permissions = True)
	frappe.db.set_value("Loan Repayment Schedule",doc.name,'custom_custom_additional_created',1)

@frappe.whitelist()
def update_loan():
	loan = frappe.db.sql("""update `tabLoan Repayment Schedule` set docstatus = 0 where name = "LN-RS-2024-00001" """,as_dict = True)
	print(loan)
	loan = frappe.db.sql("""update `tabLoan Repayment Schedule` set custom_custom_additional_created = 0 where name = "LN-RS-2024-00001" """,as_dict = True)
	print(loan)

#Create new additional salary and update the details for loan repayment schedule for on cancel of additional salary
@frappe.whitelist()
def add_salary(name, remark):
	additional = frappe.get_doc("Additional Salary", name)
	doc = frappe.get_doc("Loan Repayment Schedule", additional.custom_loan_repayment)
	last_row = doc.repayment_schedule[-1]
	next_month = frappe.utils.add_months(last_row.payment_date, 1)
	payroll_date = frappe.utils.get_first_day(next_month)  
	additional_salary_doc = frappe.new_doc("Additional Salary")
	additional_salary_doc.employee = additional.employee
	additional_salary_doc.salary_component = "Loan Deduction"
	additional_salary_doc.amount = additional.amount
	additional_salary_doc.custom_loan_repayment = additional.custom_loan_repayment
	additional_salary_doc.payroll_date = payroll_date
	additional_salary_doc.custom_principal_amount =additional.custom_principal_amount
	additional_salary_doc.custom_balance_total_amount =additional.custom_balance_total_amount
	additional_salary_doc.insert(ignore_permissions=True)
	additional_salary_doc.submit()
	item = frappe.db.sql(""" delete from `tabRepayment Schedule` where parent = '%s'  """%(doc.name))
	if doc.custom_custom_additional_created == 1:
		add=frappe.get_all("Additional Salary",{"custom_loan_repayment":additional_salary_doc.custom_loan_repayment,"docstatus":("!=",2)},['*'],order_by='payroll_date ASC')
		for j in add:
			repayment=frappe.get_doc("Loan Repayment Schedule",j.custom_loan_repayment)
			repayment.append("repayment_schedule",{

				"payment_date":j.payroll_date,
				"principal_amount":j.custom_principal_amount,
				"total_payment":j.amount,
				"balance_loan_amount":j.custom_balance_total_amount
			})
			repayment.save(ignore_permissions=True)
			frappe.db.commit()
			balance = sorted([entry.balance_loan_amount for entry in repayment.repayment_schedule], reverse=True)
			for i, entry in enumerate(repayment.repayment_schedule):
				entry.balance_loan_amount = balance[i]
			repayment.save(ignore_permissions=True)
			frappe.db.commit()
	

#The below is using print format
@frappe.whitelist()
def accounts_ledger_table(doc):
	total_amount = 0
	total_paid = 0
	total_credit_note = 0
	total_outstanding = 0
	total_0_30 = 0
	total_31_60 = 0
	total_61_90 = 0
	total_91_above = 0
	sales_invoices = frappe.get_all("Sales Invoice", {'company': doc.company, 'customer': doc.customer}, ['posting_date', 'name', 'total', 'outstanding_amount'])
	if sales_invoices:
		data = "<table width=100% border=1px solid black><tr style=background-color:#e35310;font-size:8px><td colspan=1 ><b style=color:white; text-align:center;width:320px>Invoice No</b></td><td colspan=2  style=color:white><b style=color:white; text-align:center;>Date</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Age</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Amount QR</b></td><td colspan=1 style=color:white><b style=color:white; text-align:center;>Paid QR</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Credit Note QR</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Outstanding QR</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>0-30</b></td><td  style=color:white><b style=color:white; text-align:center;>31-60</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>61-90</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>91-Above</b></td></tr>"
		for i in sales_invoices:
			days = date_diff(doc.from_date, i.posting_date)
			if 0 <= days <= 120:
				total_amount += i.total
				total_outstanding += i.outstanding_amount
				data += f'<tr><td colspan=1 style="border: 1px solid black; font-size:8px" >{i.name}</td><td colspan=2 style="border: 1px solid black; font-size:8px" >{i.posting_date}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{days}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{i.total}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{i.outstanding_amount}</td>'
				if 0 <= days <= 30:
					total_0_30 += i.total
					data += f'<td colspan=1 style="border: 1px solid black; font-size:8px" >{i.total}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td>'
				if 31 <= days <= 60:
					total_31_60 += i.total
					data += f'<td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{i.total}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td>'
				if 61 <= days <= 90:
					total_61_90 += i.total
					data += f'<td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{i.total}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td>'
				if 91 <= days <= 120:
					total_91_above += i.total
					data += f'<td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{i.total}</td>'					
		data += '</tr>'
		data += f'<tr><td colspan=4 style="border: 1px solid black; font-size:8px" >Total</td><td style="border: 1px solid black; font-size:8px" >{total_amount}</td><td style="border: 1px solid black; font-size:8px" >-</td><td style="border: 1px solid black; font-size:8px" >-</td><td style="border: 1px solid black; font-size:8px" >{total_outstanding}</td><td style="border: 1px solid black; font-size:8px" >{total_0_30}</td><td style="border: 1px solid black; font-size:8px" >{total_31_60}</td><td style="border: 1px solid black; font-size:8px" >{total_61_90}</td><td style="border: 1px solid black; font-size:8px" >{total_91_above}</td></tr>'
		data += '</table>'
	else:
		data = "<table width=100% border=1px solid black><tr style=background-color:#e35310;font-size:10px;text-align:center><td colspan=1 ><b style=color:white; text-align:center;width:320px>Sales Invoice Documents Not Avaliable</b></td></tr>"

	return data



#Going to connected document While using the duplicate ce sow button in Cost estimation
@frappe.whitelist()
def duplicate_ce(name,doc_name):
	# ce = frappe.db.get_value("Cost Estimation",{'quotation':"INT-QTN-2023-00098-3"},['name'], order_by='name desc')
	sn = frappe.db.get_list("CE SOW",{'cost_estimation':name},['name'])
	for i in sn:
		doc = frappe.get_doc("CE SOW",i.name)
		dn = frappe.copy_doc(doc)
		dn.cost_estimation = doc_name
		dn.insert()

#Get the total value to match with sales order
@frappe.whitelist()
def less_previous_payment(name):
	si=frappe.get_all("Sales Invoice",{"sales_order":name,'docstatus':("!=",2)},['grand_total'])
	total=0
	for i in si:
		total+=i.grand_total
	return total

#Show the below details in Quotation
@frappe.whitelist()
def margin_tool(item_details):
	item_details = json.loads(item_details)
	data = ''
	data+= '<br><table ><style>td { text-align:left } table,tr,td { padding:5px;border: 1px solid black; font-size:11px;} </style>'
	data+='<tr><th  colspan=13 style="padding:1px;font-size:14px;background-color:#fe3f0c;color:white;"><center><b>ITEM DETAILS</b></center></th></tr>'
	data+='<tr style="background-color:lightgrey"><td width="150px"><b>ITEM</b></td><td width="400px;"><b>ITEM NAME</b></td><td><b>QTY</b></td><td><b>UOM</b></td><td><b>RATE</b></td><td><b>AMOUNT</b></td>'

	for i in item_details:
		data+='<tr><td  align = "right" >%s</td><td  align = "right" >%s</td><td  align = "right" >%s</td><td  align = "right" >%s</td><td  align = "right" >%s</td><td  align = "right" >%s</td>'%(i["item_code"],i["item_name"],i["qty"],i["uom"],i["rate"],i["amount"])
	data+='</table>'
	return data

#Show the below details in SO 
@frappe.whitelist()
def margin_tool_so(item_details):
	item_details = json.loads(item_details)
	data = ''
	data+= '<br><table ><style>td { text-align:left } table,tr,td { padding:5px;border: 1px solid black; font-size:11px;} </style>'
	data+='<tr><th  colspan=13 style="padding:1px;font-size:14px;background-color:#fe3f0c;color:white;"><center><b>ITEM DETAILS</b></center></th></tr>'
	data+='<tr style="background-color:lightgrey"><td width="150px"><b>ITEM</b></td><td width="400px;"><b>ITEM NAME</b></td><td><b>QTY</b></td><td><b>UOM</b></td><td><b>RATE</b></td><td><b>AMOUNT</b></td><td><b>SO Qty</b></td>'

	for i in item_details:
		data+='<tr><td  align = "right" >%s</td><td  align = "right" >%s</td><td  align = "right" >%s</td><td  align = "right" >%s</td><td  align = "right" >%s</td><td  align = "right" >%s</td><td  align = "right" >%s</td>'%(i["item_code"],i["item_name"],i["qty"],i["uom"],i["rate"],i["amount"],i["so_qty"])
	data+='</table>'
	return data

#Show the below details in  SI
@frappe.whitelist()
def margin_tool_si(item_details):
	item_details = json.loads(item_details)
	data = ''
	data+= '<br><table ><style>td { text-align:left } table,tr,td { padding:5px;border: 1px solid black; font-size:11px;} </style>'
	data+='<tr><th  colspan=13 style="padding:1px;font-size:14px;background-color:#fe3f0c;color:white;"><center><b>ITEM DETAILS</b></center></th></tr>'
	data+='<tr style="background-color:lightgrey"><td width="150px"><b>ITEM</b></td><td width="400px;"><b>ITEM NAME</b></td><td><b>QTY</b></td><td><b>UOM</b></td><td><b>RATE</b></td><td><b>AMOUNT</b></td><td><b>Ordered Qty</b></td>'

	for i in item_details:
		data+='<tr><td  align = "right" >%s</td><td  align = "right" >%s</td><td  align = "right" >%s</td><td  align = "right" >%s</td><td  align = "right" >%s</td><td  align = "right" >%s</td><td  align = "right" >%s</td>'%(i["item_code"],i["item_name"],i["qty"],i["uom"],i["rate"],i["amount"],i["ordered_qty"])
	data+='</table>'
	return data


@frappe.whitelist()
def updatecustomer():
	list = frappe.db.get_list("Customer Credit Limit")
	for i in list:
		frappe.db.set_value("Customer Credit Limit",i.name,"bypass_credit_limit_check",1)



# @frappe.whitelist()
# def reservation_entry():
#     print("hi")
#     creation = frappe.get_all("Stock Reservation Entry",{"docstatus":("!=",2)},["creation","name"])
#     print("hi")
#     for i in creation:
#         date = date_diff(nowdate(),i.creation.date())
#         print(date)
#         if date==10:
#             frappe.sendmail(
#                 recipients= ['rifdy@electraqatar.com','aneesh@electraqatar.com','kapil@electraqatar.com'],
#                 subject=('Stock Reservation Entry'),
#                 message="""
#                     Dear Sir,<br><br>
#                     Kindly find the document mentioned below, which was created 10 days ago. ,<br><br>
#                     and this document going to Unreserve in 5 days,<brr><br>
#                     %s
#                     """%(i.name)
#                 )

#Create new Journal Entry On submission of Sales Invoice
@frappe.whitelist()
def create_ret_je(doc,method):
	from electra.utils import get_series
	if doc.ret_amount and doc.ret_amount>0:
		jv = frappe.new_doc("Journal Entry")
		jv.naming_series = get_series(doc.company,"Journal Entry")
		jv.voucher_type = "Journal Entry"
		jv.company = doc.company
		jv.posting_date = nowdate()
		# jv.bill_no = tn.bill_no
		jv.bill_date = nowdate()
		acc = frappe.db.get_value("Account",{"account_type":"Receivable","account_number":"1310","company":doc.company})
		jv.append("accounts", {
			"account": acc,
			"party_type": "Customer",
			"party":doc.customer,
			"debit": doc.ret_amount,
			"reference_type": "Sales Invoice",
			"reference_name": doc.name,
			"cost_center": erpnext.get_default_cost_center(doc.company),
			"debit_in_account_currency": doc.ret_amount
		})
		je_acc=frappe.db.get_value("Account",{"name":['like','RETENTION RECEIVABLE%'],"company":doc.company},["name"])
		jv.append("accounts", {
			"account": je_acc,			
			"cost_center": erpnext.get_default_cost_center(doc.company),
			"credit": doc.ret_amount,
			"credit_in_account_currency": doc.ret_amount
		})
		jv.insert()
		jv.submit()

#Cancel Journal entry with linked sales invoice
@frappe.whitelist()
def cancel_ret_je(doc,method):
	if doc.order_type == "Project":
		parent = frappe.db.get_value("Journal Entry Account",{"reference_type": "Sales Invoice","reference_name": doc.name},['parent'])
		if parent:
			je = frappe.get_doc("Journal Entry",parent)
			je.cancel()

#Calculate the below amount set the sales invoice
@frappe.whitelist()
def calculate_amount(so):
	amount = frappe.db.get_value("Sales Order", {"name": so}, ["grand_total"])
	if amount:
		value = 0	
		query = """
			SELECT SUM(per.base_net_amount) AS total_amount
			FROM `tabSales Invoice Item` AS per
			LEFT JOIN `tabSales Invoice` AS pe ON per.parent = pe.name
			WHERE pe.docstatus != 2 and per.sales_order = %s
			"""
		pay = frappe.db.sql(query, (so,), as_dict=True)
		for i in pay:
			if i.get('total_amount') is not None: 
				value = i.get('total_amount')
			if amount <= value:
				return "Hii"

from frappe.query_builder.functions import Abs, Sum	

#Update the percetage in Sales order only the project type
@frappe.whitelist()
def update_advance_percentage(doc,method):
	list = []
	for reference in doc.references:
		if reference.reference_doctype == "Sales Order":
			sql = frappe.db.sql(""" select parent from `tabPayment Entry Reference` where reference_name = '%s' and docstatus != 0 """%(reference.reference_name),as_dict=1)
			for s in sql:
				list.append(s.parent)
			ple = frappe.qb.DocType("Payment Ledger Entry")
			reference = frappe.qb.DocType("Payment Entry Reference")
			advance = (
				frappe.qb.from_(ple)
				.inner_join(reference)
				.on(reference.parent == ple.against_voucher_no)
				.select(reference.reference_name,reference.reference_doctype,ple.account_currency, Abs(Sum(ple.amount_in_account_currency)).as_("amount"))
				.where(
					(ple.against_voucher_type == doc.doctype)
					& (ple.against_voucher_no).isin(list)
					& (ple.docstatus == 1)
					& (ple.company == doc.company)
				)
				.run(as_dict=True)
			)
			if advance:
				advance = advance[0]
				if advance.reference_doctype == "Sales Order":
					grand_total = frappe.db.get_value("Sales Order",advance.reference_name,"grand_total")
					order_type = frappe.db.get_value("Sales Order",advance.reference_name,"order_type")
					if order_type == "Project":
						advance_paid = flt(advance.amount)
						adv_percentage = (advance_paid/grand_total) * 100
						frappe.db.set_value(advance.reference_doctype, advance.reference_name, "advance_paid", advance_paid)
						frappe.db.set_value(advance.reference_doctype, advance.reference_name, "custom_adv_paid_percentage", adv_percentage)


@frappe.whitelist()
def list_aging():	
	journal = frappe.db.sql("""SELECT per.credit_in_account_currency,per.debit_in_account_currency, pe.name, pe.posting_date FROM `tabJournal Entry Account` AS per
			LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
			WHERE per.account like '%s' AND pe.docstatus = 1 AND party_type = 'Customer' AND party ='%s'AND per.reference_name IS NULL """ % ('%Debtors -%',"TECHNOLOGY PROFESSIONALS TRADING"), as_dict=True)
	print(journal)

#Calculate the below amount and check to delivery note
@frappe.whitelist()
def calculate_so_amount(so):
	amount = frappe.db.get_value("Sales Order", {"name": so}, ["grand_total"])
	value = 0
	query = """
		SELECT SUM(per.base_net_amount) AS total_amount
		FROM `tabDelivery Note Item` AS per
		LEFT JOIN `tabDelivery Note` AS pe ON per.parent = pe.name
		WHERE pe.docstatus != 2 and per.against_sales_order = %s
		"""
	pay = frappe.db.sql(query, (so,), as_dict=True)
	
	for i in pay:
		if i.get('total_amount') is not None: 
			value = i.get('total_amount')
		if not (amount == 0 and value ==0):
			if amount <= value:
				return "Hii"

#Set the series in lead
@frappe.whitelist()
def get_lead_series(company,sales_person):
	company_series = frappe.db.get_value("Company Series",{'company':company,'document_type':'Lead'},'series')
	name = sales_person.split(' ')
	return f"{company_series}-{name[0]}-"

#update date different in vacation final exit form
@frappe.whitelist()
def update_date_different(join_date,last_date):
	from_date = datetime.strptime(join_date, '%Y-%m-%d')
	to_date = datetime.strptime(last_date, '%Y-%m-%d')
	difference = to_date - from_date
	years = difference.days // 365
	months = (difference.days % 365) // 30
	days = (difference.days % 365) % 30
	return f"{years} years, {months} months, {days} days"


@frappe.whitelist()
def calculate_deduction_amount(loan,visa,air_ticket,other):
	total = loan if loan else 0 + visa if visa else 0 + air_ticket if air_ticket else 0 +other if other else 0
	return total

@frappe.whitelist()
def updatece():
	doc = frappe.db.sql(""" select qty,vr,name,item,vr_with_tr_percent,parent,cost,transfer_cost_amount from `tabCE MEP Materials` """,as_dict = 1)
	for i in doc:
		if i.vr > 0:
			estimation = frappe.db.get_value("CE SOW",i.parent,"cost_estimation")
			est = frappe.get_doc("Cost Estimation",estimation)
			if est.docstatus != 2:
				est.material_cost_price = sum(i.amount for i in est.materials)
				est.flags.ignore_validate_update_after_submit = True
				est.save(ignore_permissions=True)

@frappe.whitelist()
def update_cus_so():
	from erpnext.selling.doctype.customer.customer import get_customer_outstanding, get_credit_limit
	customer ="harvest"
	company =  "INTERIOR DIVISION - ELECTRA"
	# Outstanding based on GL Entries
	outstanding_based_on_gle = frappe.db.sql(
		"""
		select sum(debit) - sum(credit)
		from `tabGL Entry` where party_type = 'Customer'
		and is_cancelled = 0 and party = '%s'
		and company= '%s' """%(customer, company),
	)
	outstanding_based_on_gle = flt(outstanding_based_on_gle[0][0]) if outstanding_based_on_gle else 0

	# Outstanding based on Sales Order
	outstanding_based_on_so = 0

	outstanding_based_on_so = frappe.db.sql(
		"""
		select sum(base_grand_total*(100 - per_billed)/100)
		from `tabSales Order`
		where customer=%s and docstatus = 1 and company=%s
		and per_billed < 100 and status != 'Closed'""",
		(customer, company),
	)

	outstanding_based_on_so = flt(outstanding_based_on_so[0][0]) if outstanding_based_on_so else 0

	# Outstanding based on Delivery Note, which are not created against Sales Order
	outstanding_based_on_dn = 0

	unmarked_delivery_note_items = frappe.db.sql(
		"""select
			dn_item.name, dn_item.amount, dn.base_net_total, dn.base_grand_total
		from `tabDelivery Note` dn, `tabDelivery Note Item` dn_item
		where
			dn.name = dn_item.parent
			and dn.customer=%s and dn.company=%s
			and dn.docstatus = 0 and dn.status not in ('Closed', 'Stopped')
			and ifnull(dn_item.against_sales_order, '') = ''
			and ifnull(dn_item.against_sales_invoice, '') = ''
		""",
		(customer, company),
		as_dict=True,
	)

	if not unmarked_delivery_note_items:
		return outstanding_based_on_gle + outstanding_based_on_so

	si_amounts = frappe.db.sql(
		"""
		SELECT
			dn_detail, sum(amount) from `tabSales Invoice Item`
		WHERE
			docstatus = 1
			and dn_detail in ({})
		GROUP BY dn_detail""".format(
			", ".join(frappe.db.escape(dn_item.name) for dn_item in unmarked_delivery_note_items)
		)
	)

	si_amounts = {si_item[0]: si_item[1] for si_item in si_amounts}

	for dn_item in unmarked_delivery_note_items:
		dn_amount = flt(dn_item.amount)
		si_amount = flt(si_amounts.get(dn_item.name))

		if dn_amount > si_amount and dn_item.base_net_total:
			outstanding_based_on_dn += (
				(dn_amount - si_amount) / dn_item.base_net_total
			) * dn_item.base_grand_total

	print(outstanding_based_on_gle + outstanding_based_on_so + outstanding_based_on_dn)
	
@frappe.whitelist()
def update_dn_items(so_no,item):
	query = """
		SELECT soi.item_code, 
			   soi.qty - SUM(per.qty) AS qty_difference
		FROM `tabSales Order Item` AS soi
		LEFT JOIN `tabDelivery Note Item` AS per 
			ON soi.parent = per.against_sales_order
		LEFT JOIN `tabDelivery Note` AS pe 
			ON per.parent = pe.name and soi.item_code = per.item_code
		WHERE pe.docstatus != 2 
			AND soi.parent = %s
		GROUP BY soi.item_code
	"""
	dn_value = frappe.db.sql(query, (so_no,), as_dict=True)
	return dn_value

#Get the resevation entry document and sent mail for every day
@frappe.whitelist()
def reservation_entry():
	sre = []
	entries = frappe.get_all("Stock Reservation Entry", {"docstatus": 1}, ["creation", "name"])
	for entry in entries:
		creation_date = entry.creation.date()
		days_diff = (datetime.now().date() - creation_date).days
		sre.append(days_diff)
		if days_diff == 14:
			
			sre = frappe.get_value("Stock Reservation Entry",entry.name)
			frappe.get_doc("Stock Reservation Entry", entry.name).cancel()




@frappe.whitelist()
def find_role():
	value = frappe.db.get_value("Has Role",{"parent":"kapil@electraqatar.com","role":"Purchase Manager"},["role"])
	print(value)
	if value:
		print(value)

# Sent mail for expire passport details
@frappe.whitelist()
def passport_expire():
	visa = frappe.db.sql(""" select name,custom_valid_upto,custom_visa_valid_upto from `tabEmployee`""",as_dict = 1)
	data = ''
	data += '<table class = table table - bordered><tr><td colspan = 5>Expired Passport</td><td colspan = 5>Visa Passport</td></tr>'
	for date in visa:
		Visa = date.custom_visa_valid_upto
		passport = date.custom_valid_upto
		if Visa:
			str_date = datetime.strptime(str(today()),'%Y-%m-%d').date()
			expiry = add_days(today(),30)
			expiry_date = datetime.strptime(str(expiry),'%Y-%m-%d').date()
			if Visa < expiry_date:
				print(Visa)
				data += '<tr><td>%s</td><td>%s</td>'%(Visa,date.name)
		if passport:
			str_date = datetime.strptime(str(today()),'%Y-%m-%d').date()
			Pass = add_days(today(),30)
			Pass_date = datetime.strptime(str(Pass),'%Y-%m-%d').date()
			if passport < Pass_date:
				print(Visa)
				data += '<td>%s</td><td>%s</td>'%(passport,date.name)
	data+='</tr>'
	data += '</table>'
	frappe.sendmail(
			recipients=[''],
			subject=('Expiry Details'),
			header=('Visa,Passport and Job Loss Insurance Expiry List'),
			message="""
					Dear Mam,<br><br>
					%s
					""" % (data)
		)

#Create new payment entry	
@frappe.whitelist()
def create_single_pe():
	from frappe.utils import now
	self = frappe.get_doc("Consolidated Payment Request","ACC-CPRQ-2023-00182")
	pay = frappe.new_doc("Payment Entry")
	pay.payment_type = "Pay"
	pay.posting_date = now()
	pay.company = self.purchase_invoice_list[0].company
	pay.mode_of_payment = self.mode_of_payment
	pay.party = self.purchase_invoice_list[0].supplier
	pay.party_name = self.purchase_invoice_list[0].supplier
	pay.party_type = "Supplier"
	for i in self.purchase_invoice_list:
		pay.append("references",{
			"reference_doctype":i.type,
			"reference_name":i.name1,
			"allocated_amount":i.payment_amount,
			"total_amount": i.grand_total,
			"allocated_amount": i.payment_amount,
			"outstanding_amount": i.outstanding_amount,
		})
		
	acc = frappe.db.get_list("Account",{"company":self.purchase_invoice_list[0].company,"account_number":"1211"},["name","account_currency"])
	for ac in acc:
		pay.paid_from = ac.name
		pay.paid_from_account_currency = ac.account_currency
	pay_acc = frappe.db.get_list("Account",{"company":self.purchase_invoice_list[0].company,"account_number":"2110"},["name","account_currency"])
	for ac_pay in pay_acc:
		pay.paid_to = ac_pay.name
		pay.paid_to_account_currency = ac_pay.account_currency
	# pay.paid_amount = i.payment_amount
	# pay.received_amount = i.payment_amount
	# pay.reference_no = i.reference_no
	pay.consolidated_payment_request = self.name
	pay.reference_date = self.purchase_invoice_list[0].reference_date
	pay.flags.ignore_mandatory = True
	pay.flags.ignore_validate = True
	pay.save(ignore_permissions = True)



@frappe.whitelist()
def payconsolidatedreq():
	from collections import defaultdict
	from frappe.utils import now
	name = "ACC-CPRQ-2023-00182"
	con = frappe.db.sql("""SELECT * FROM `tabPurchase Invoice List` WHERE parent = '%s' """ % (name), as_dict=True)

	def calculate_outstanding_payment_difference(child_table):
		amounts_by_company = defaultdict(float)
		total = 0.0

		for row in child_table:
			company = row.get("company", "")
			outstanding_amount = row.get("outstanding_amount", 0.0)
			payment_amount = row.get("payment_amount", 0.0)
			amounts_by_company[company] += payment_amount
			total += payment_amount

		result = []

		for company, difference in amounts_by_company.items():
			result.append({"company": company, "value": difference})

		for k in result:
			self = frappe.get_doc("Consolidated Payment Request",name)
			pay = frappe.new_doc("Payment Entry")
			pay.payment_type = "Pay"
			for i in self.purchase_invoice_list:
				if i.company == k['company']:
					pay.append("references",{
						"reference_doctype":i.type,
						"reference_name":i.name1,
						"allocated_amount":i.payment_amount,
						"total_amount": i.grand_total,
						"allocated_amount": i.payment_amount,
						"outstanding_amount": i.outstanding_amount,
					})
			pay.posting_date = now()
			pay.company = k['company']
			pay.mode_of_payment = self.mode_of_payment
			pay.party = self.supplier
			pay.party_name = self.supplier
			pay.party_type = "Supplier"
			acc = frappe.db.get_list("Account",{"company":k['company'],"account_number":"1211"},["name","account_currency"])
			for ac in acc:
				pay.paid_from = ac.name
				pay.paid_from_account_currency = ac.account_currency

			co = [{"company":"KINGFISHER TRADING AND CONTRACTING COMPANY","name":"Creditors - KTCC"},
					{"company":"KINGFISHER - TRANSPORTATION","name":"Creditors - KT"},
					{"company":"KINGFISHER - SHOWROOM","name":"Creditors - KS"}]
			for company_info in co:
				if k['company'] == company_info['company']:
					print(company_info['name'])
					pay_acc = frappe.db.get_list("Account",{"name":company_info['name']},["name","account_currency"])
				else:
					pay_acc = frappe.db.get_list("Account",{"company":k['company'],"account_number":"2110"},["name","account_currency"])
			for ac_pay in pay_acc:
				pay.paid_to = ac_pay.name
				pay.paid_to_account_currency = ac_pay.account_currency
			pay.paid_amount = k['value']
			pay.received_amount = k['value']
			# pay.reference_no = i.reference_no
			pay.consolidated_payment_request = self.name
			# pay.reference_date = i.reference_date
			pay.flags.ignore_mandatory = True
			pay.flags.ignore_validate = True
			pay.save(ignore_permissions = True)

	child_table = con
	amounts_difference_by_company = calculate_outstanding_payment_difference(child_table)
	return amounts_difference_by_company

@frappe.whitelist()
def update_entry():
	loan = frappe.db.sql("""update `tabStock Reservation Entry` set status = 'Cancelled' where name = "MAT-SRE-2024-00044" """,as_dict = True)

@frappe.whitelist()
def lcv_supplier():
	value = frappe.get_all("Landed Cost Voucher",["name"])
	for i in value:
		doc = frappe.get_doc("Landed Cost Voucher",i.name)
		for j in doc.purchase_receipts:
			frappe.db.set_value("Landed Cost Voucher",i.name,"custom_supplier",j.supplier)


#Show all linked documents in SI
@frappe.whitelist()
def update_link_field(name):
	has_data = 0
	data = ''
	dn_list = []
	doc = frappe.get_doc("Sales Invoice", name)
	data += '<table class="table table-bordered" border=1><tr><td style="background-color:#fe3f0c;text-align:center" colspan=2>Link Documents</td></tr>'
	for i in doc.items:
		if i.delivery_note and i.delivery_note not in dn_list:
			has_data = 1
			dn_list.append(i.delivery_note)
			link = f'https://erp.electraqatar.com/app/delivery-note/{i.delivery_note}'
			data += '<tr style=text-align:center><td><a href="{}" target="_blank">{}</a></td><td>Delivery Note</td></tr>'.format(link, i.delivery_note)
	if doc.so_no:
		has_data = 1
		link = f'https://erp.electraqatar.com/app/sales-order/{doc.so_no}'
		data += '<tr style=text-align:center><td ><a href="{}" target="_blank">{}</a></td><td>Sales Order</td></tr>'.format(link, doc.so_no)
	if doc.return_document:
		has_data = 1
		link = f'https://erp.electraqatar.com/app/sales-invoice/{doc.return_document}'
		data += '<tr style=text-align:center><td><a href="{}" target="_blank">{}</a></td><td>Return Document</td></tr>'.format(link, doc.return_document)
		
	# if doc.return_against:
	# 	has_data = 1
	# 	link = f'https://erp.electraqatar.com/app/sales-invoice/{doc.return_against}'
	# 	data += '<tr style=text-align:center><td <a href="{}" target="_blank">{}</a></td><td>Return Against</td></tr>'.format(link, doc.return_against)

	data += '</table>'
	if has_data:
		return data

#Show all linked documents in Qn
@frappe.whitelist()
def update_qn_field(name):
	data = ''
	doc = frappe.get_doc("Quotation", name)
	data += '<table class="table table-bordered" border=1><tr><td style="background-color:#fe3f0c;text-align:center" colspan=2>Link Documents</td></tr>'
	if doc.opportunity:
		link = f'https://erp.electraqatar.com/app/opportunity/{doc.opportunity}'
		data += '<tr style=text-align:center><td><a href="{}" target="_blank">{}</a></td><td>Opportunity</td></tr>'.format(link, doc.opportunity)
	if doc.cost_estimation:
		link = f'https://erp.electraqatar.com/app/cost-estimation/{doc.cost_estimation}'
		data += '<tr style=text-align:center><td ><a href="{}" target="_blank">{}</a></td><td>Cost Estimation</td></tr>'.format(link, doc.cost_estimation)
	data += '</table>'
	if doc.opportunity or doc.cost_estimation:
		return data

#Show all linked documents in SO
@frappe.whitelist()
def update_so_field(name):
	data = ''
	doc = frappe.get_doc("Sales Order", name)
	data += '<table class="table table-bordered" border=1><tr><td style="background-color:#fe3f0c;text-align:center" colspan=2>Link Documents</td></tr>'
	if doc.quotation:
		link = f'https://erp.electraqatar.com/app/quotation/{doc.quotation}'
		data += '<tr style=text-align:center><td><a href="{}" target="_blank">{}</a></td><td>Quotation</td></tr>'.format(link, doc.quotation)
	if doc.cost_estimation:
		link = f'https://erp.electraqatar.com/app/cost-estimation/{doc.cost_estimation}'
		data += '<tr style=text-align:center><td ><a href="{}" target="_blank">{}</a></td><td>Cost Estimation</td></tr>'.format(link, doc.cost_estimation)
	if doc.project_budget:
		link = f'https://erp.electraqatar.com/app/project-budget/{doc.project_budget}'
		data += '<tr style=text-align:center><td ><a href="{}" target="_blank">{}</a></td><td>Project Budget</td></tr>'.format(link, doc.project_budget)
	data += '</table>'
	if doc.quotation or doc.cost_estimation or doc.project_budget:
		return data

#Show all linked documents in DN
@frappe.whitelist()
def update_dn_field(name):
	data = ''
	doc = frappe.get_doc("Delivery Note", name)
	data += '<table class="table table-bordered" border=1><tr><td style="background-color:#fe3f0c;text-align:center" colspan=2>Link Documents</td></tr>'
	for i in doc.items:
		if i.against_sales_order:
			link =f'https://erp.electraqatar.com/app/sales-order/{i.against_sales_order}'
			data += '<tr style=text-align:center><td><a href="{}" target="_blank">{}</a></td><td>Sales Order</td></tr>'.format(link, i.against_sales_order)
	if doc.custom_invoice_number:
		link =f'https://erp.electraqatar.com/app/sales-invoice/{doc.custom_invoice_number}'
		data += '<tr style=text-align:center><td <a href="{}" target="_blank">{}</a></td><td>Sales Invoice</td></tr>'.format(link, doc.custom_invoice_number)
	data += '</table>'
	if i.against_sales_order or doc.custom_invoice_number:
		return data


#Set the all values in advance invoice table from the below so
@frappe.whitelist()
def update_child_values(so):
	sales=frappe.get_doc("Sales Order",so)
	return sales.items , sales.scope_of_work , sales.so_work_title_item,sales.payment_schedule_2s,sales.retention_table

#Set the all values in advance invoice from the below so
@frappe.whitelist()
def update_normal_values(so):
	sales=frappe.get_doc("Sales Order",so)
	return sales.total_bidding_price , sales.project_discount_per , sales.project_discount_amt,sales.discount_tolerance_amount,sales.net_bidding_price,sales.discount_upto

#Set the all values in advance invoice from the below so
@frappe.whitelist()
def update_item_values(so):
	sales=frappe.get_doc("Sales Order",so)
	return sales.total_qty , sales.base_total , sales.base_net_total,sales.total,sales.net_total,sales.base_total_taxes_and_charges,sales.total_taxes_and_charges,sales.base_grand_total,sales.base_rounding_adjustment,sales.base_rounded_total,sales.grand_total,sales.rounding_adjustment,sales.rounded_total,sales.additional_discount_percentage,sales.discount,sales.discount_amount,sales.percent_of_discounted_amount

#Set the advance amount in advance invoice
@frappe.whitelist()
def update_advance_amount(doc,method):
	frappe.db.set_value("Advance Invoice",doc.name,"advance_amount",doc.advance_amount1)

#Set the advance amount in advance invoice
@frappe.whitelist()
def update_retention_amount(doc,method):
	frappe.db.set_value("Retention Invoice",doc.name,"advance_amount",doc.advance_amount1)

# @frappe.whitelist()
# def update_amount_advance(doc,method):
# 	if doc.custom_is_advance==1:
# 		jv = frappe.new_doc("Journal Entry")
# 		jv.voucher_type = "Journal Entry"
# 		jv.company = doc.company
# 		jv.posting_date = today()
# 		jv.user_remark = "On Submission of Advance Invoice - Cr. on Advance to be Received / Dt. on Debtor / Party - Customer"
# 		jv.cheque_no=doc.name
# 		jv.custom_advance_invoice=doc.custom_advance_invoice
# 		jv.cheque_date= doc.posting_date
# 		accounts = [
			
# 			{
# 				'account': frappe.db.get_value("Company",{"name":doc.company},["custom_default_advance_account"]),
# 				'debit_in_account_currency': doc.paid_amount,
				
# 			},{
# 				'account': frappe.db.get_value("Company",{"name":doc.company},["default_receivable_account"]),
# 				'credit_in_account_currency': doc.paid_amount,
# 				'party_type': 'Customer',
# 				'party': doc.party_name
# 			},
# 		]
# 		for account in accounts:
# 			jv.append('accounts', account)
# 		jv.save(ignore_permissions=True)
# 		jv.submit()
# 		frappe.db.set_value("Advance Invoice",doc.custom_advance_invoice,"advance_amount",doc.custom_outstanding_amount-doc.paid_amount)
# 		custom_invoice = jv.custom_advance_invoice
# 		existing_docs = frappe.db.sql("""
# 			SELECT name FROM `tabJournal Entry`
# 			WHERE name LIKE %s
# 			ORDER BY name DESC
# 		""", (custom_invoice + '%'), as_dict=True)
# 		if not existing_docs:
# 			new_name = f"{custom_invoice}-01"
# 		else:
# 			last_doc = existing_docs[0]
# 			last_number = int(last_doc.name.split('-')[-1])
# 			new_sequence = last_number + 1
# 			new_name = f"{custom_invoice}-{new_sequence}"
# 		frappe.rename_doc("Journal Entry", jv.name, new_name, force=1)

# @frappe.whitelist()
# def update_pay_name(doc,method):
# 	if doc.custom_is_advance==1:
# 		frappe.db.set_value("Advance Invoice",doc.custom_advance_invoice,"	",doc.name)

#Set the all values in advance invoice from the below so
@frappe.whitelist()
def update_ad_values(so,advance):
	sales=frappe.get_doc("Sales Order",so)
	value = frappe.db.get_value("Advance Invoice",{"name":advance},["advance_amount"])
	return sales.name,sales.total,value,frappe.db.get_value("Company",{"name":sales.company},["default_cash_account"]),frappe.db.get_value("Company",{"name":sales.company},["default_receivable_account"])


# to create Item while submitting the Cost Estimation
@frappe.whitelist()
def create_item_on_ce_submit(doc,method):
	if doc.master_scope_of_work:
		for i in doc.master_scope_of_work:
			exists = frappe.db.exists("Item",i.msow)
			if not exists:
				item = frappe.new_doc("Item")
				item.item_code = i.msow
				item.item_name = i.msow_desc
				item.description = i.msow_desc
				item.is_stock_item = 0
				item.stock_uom = i.unit
				item.item_group = "Projects"
				item.save(ignore_permissions = True)

#Get the below value and check with dn wip
@frappe.whitelist()
def check_child_item(against_pb_item):
	quantity = frappe.db.sql(""" select item,(qty - delivered_qty) as qty from `tabProject Budget Items` where name = '%s' """%(against_pb_item),as_dict = 1)
	return quantity

@frappe.whitelist()
def check_child_item_mr(against_pb_item):
	quantity = frappe.db.sql(""" select item,qty from `tabProject Budget Items` where name = '%s' """%(against_pb_item),as_dict = 1)
	return quantity

#Get the items details for main document and set to return document
@frappe.whitelist()
def get_main_dn(name):
	doc = frappe.get_doc("Delivery Note WIP",name)
	return doc.items

#Create stock entry while submission of dn wip
@frappe.whitelist()
def project_delivery(doc,method):
	project=frappe.get_value("Sales Order",{'name':doc.sales_order},['project'])
	if doc.is_return == 0:
		so = frappe.get_doc("Project Budget",doc.project_budget)
		if so.docstatus == 1:
			se = frappe.new_doc("Stock Entry")
			se.stock_entry_type = "Material Transfer" 
			se.posting_date=doc.posting_date
			se.set_posting_time=1
			se.posting_time=doc.posting_time
			se.company = doc.company
			se.reference_number = doc.name
			se.custom_reference_type = "Delivery Note WIP"
			for i in doc.items:
				is_stock_item = frappe.db.get_value("Item",i.item_code,'is_stock_item')
				if is_stock_item:
					for s in so.item_table:
						if s.docname == i.custom_against_pbsow:
							if s.delivered_qty < s.qty:
								s.delivered_qty +=i.qty
							else:
								frappe.throw("Row " + i.idx +' : Item '+i.item_code+' already delivered fully')
					do = frappe.get_doc(i.custom_against_pbsow_doctype,i.custom_against_pbsow)
					do.delivered_qty += i.qty
					do.save(ignore_permissions=True)	
					se.append("items",{
						's_warehouse':i.warehouse,
						't_warehouse':i.target_warehouse,
						'item_code':i.item_code,
						'qty':i.qty,
						'basic_rate':i.rate,
						'project':project
					})
				else:
					for cn in doc.cons_items:		
						se.append("items",{
							's_warehouse':doc.set_warehouse,
							't_warehouse':doc.set_target_warehouse,
							'item_code':cn.item,
							'qty':cn.qty,
							'basic_rate':cn.rate
						})	
			# for i in doc.items:
			# 	do = frappe.get_doc(i.custom_against_pbsow_doctype,i.custom_against_pbsow)
			# 	do.delivered_qty += i.qty
			# 	do.save(ignore_permissions=True)

			# total_qty = sum(i.delivered_qty for i in so.items)
			# so.per_delivered = (total_qty / so.total_qty)*100

			so.flags.ignore_validate_update_after_submit = True
			so.save(ignore_permissions=True)
			se.save(ignore_permissions=True)
			se.submit()
		else:
			frappe.throw("Cannot update against cancel document - <b>"+doc.project_budget+"</b>")
	else:
		so = frappe.get_doc("Project Budget",doc.project_budget)
		if so.docstatus == 1:
			if not frappe.db.exists("Sales Invoice",{'sales_order':doc.sales_order,'docstatus':('!=',2)}):
				se = frappe.new_doc("Stock Entry")
				se.stock_entry_type = "Material Transfer" 
				se.company = doc.company
				se.reference_number = doc.name
				se.custom_reference_type = "Delivery Note WIP"
				for i in doc.items:
					for s in so.item_table:
						if s.docname == i.custom_against_pbsow:
							if s.delivered_qty > 0:
								s.delivered_qty -=i.qty
							else:
								frappe.throw("Row " + i.idx +' : Item '+i.item_code+' already returned fully')	
					se.append("items",{
						's_warehouse':i.warehouse,
						't_warehouse':i.target_warehouse,
						'item_code':i.item_code,
						'qty':i.qty,
						'basic_rate':i.rate,
						'project':project
					})			
				for i in doc.items:
					do = frappe.get_doc(i.custom_against_pbsow_doctype,i.custom_against_pbsow)
					do.delivered_qty -= i.qty
					do.save(ignore_permissions=True)	

				# total_qty = sum(i.delivered_qty for i in so.items)
				# so.per_delivered = (total_qty / so.total_qty)*100

				so.flags.ignore_validate_update_after_submit = True
				so.save(ignore_permissions=True)
				se.save(ignore_permissions=True)
				se.submit()
			else:
				flag=False
				sales_invoices=frappe.get_all("Sales Invoice",{'sales_order':doc.sales_order,'docstatus':('!=',2)},['*'])
				for si in sales_invoices:
					s = si.delivery_note_list
					stat = s.strip("[']'")
					dn_lst = stat.split(",")
					for s in dn_lst:
						if doc.return_against==s.strip().lstrip("'").rstrip("'"):
							frappe.throw("This Delivery Note - <b>"+doc.project_budget+"</b> has been already invoiced. Kindly use the Sales Invoice Return")
							flag=False
							break
						else:
							flag=True
				if flag:
					se = frappe.new_doc("Stock Entry")
					se.stock_entry_type = "Material Transfer" 
					se.company = doc.company
					se.reference_number = doc.name
					se.custom_reference_type = "Delivery Note WIP"
					for i in doc.items:
						for s in so.item_table:
							if s.docname == i.custom_against_pbsow:
								s.delivered_qty -=i.qty	
						se.append("items",{
							's_warehouse':i.warehouse,
							't_warehouse':i.target_warehouse,
							'item_code':i.item_code,
							'qty':i.qty,
							'basic_rate':i.rate,
							'project':project
						})			
					for i in doc.items:
						do = frappe.get_doc(i.custom_against_pbsow_doctype,i.custom_against_pbsow)
						do.delivered_qty -= i.qty
						do.save(ignore_permissions=True)	

					# total_qty = sum(i.delivered_qty for i in so.items)
					# so.per_delivered = (total_qty / so.total_qty)*100

					so.flags.ignore_validate_update_after_submit = True
					so.save(ignore_permissions=True)
					se.save(ignore_permissions=True)
					se.submit()
		else:
			frappe.throw("Cannot update against cancel document - <b>"+doc.project_budget+"</b>")

#Revert the below values while cancel of dn wip
@frappe.whitelist()
def update_so_delivered_qty(doc,method):
	# to update the delivered qty in Project Budget
	if doc.is_return == 0:
		so = frappe.get_doc("Project Budget",doc.project_budget)	
		for i in doc.items:
			for s in so.item_table:
				if s.docname == i.custom_against_pbsow:
					s.delivered_qty -=i.qty
		for i in doc.items:
			do = frappe.get_doc(i.custom_against_pbsow_doctype,i.custom_against_pbsow)
			do.delivered_qty -= i.qty
			do.save(ignore_permissions=True)
		# total_qty = sum(i.delivered_qty for i in so.items)
		# so.per_delivered = (total_qty / so.total_qty)*100
		so.flags.ignore_validate_update_after_submit = True
		so.save(ignore_permissions=True)
	else:
		so = frappe.get_doc("Project Budget",doc.project_budget)
		for i in doc.items:
			for s in so.item_table:
				if s.docname == i.custom_against_pbsow:
					s.delivered_qty +=i.qty
		for i in doc.items:
			do = frappe.get_doc(i.custom_against_pbsow_doctype,i.custom_against_pbsow)
			do.delivered_qty += i.qty
			do.save(ignore_permissions=True)	

		# total_qty = sum(i.delivered_qty for i in so.items)
		# so.per_delivered = (total_qty / so.total_qty)*100

		so.flags.ignore_validate_update_after_submit = True
		so.save(ignore_permissions=True)


#Create stock entry for submission of SI
@frappe.whitelist()
def update_stock_on_si(doc,method):
	se_items = []
	project=frappe.get_value("Sales Order",{'name':doc.so_no},['project'])
	if doc.order_type =="Project":
		if doc.is_return==0:
			project_budget = frappe.db.get_value("Sales Order",doc.so_no,'project_budget')
			pb = frappe.get_doc("Project Budget",project_budget)
			se = frappe.new_doc("Stock Entry")
			se.stock_entry_type = "Material Issue" 
			se.set_posting_time = 1
			se.posting_date=doc.posting_date
			se.posting_time=doc.posting_time
			se.company = doc.company
			se.reference_number = doc.name
			se.custom_reference_type = "Sales Invoice"
			set_warehouse = frappe.db.get_value("Warehouse",{'warehouse_name': ['like', '%Work In Progress%'],'company':doc.company},'name')
			for i in pb.item_table:
				if i.delivered_qty:
					to_bill_qty = i.delivered_qty - i.billed_qty
					if to_bill_qty > 0:
						if frappe.db.exists("Item", {'name': i.item, 'is_stock_item': 1}):
							se.append("items",{
								's_warehouse':set_warehouse,
								'item_code':i.item,
								'transfer_qty':to_bill_qty,
								'uom':i.unit,
								'stock_uom':frappe.get_value('Item',i.item,'stock_uom'),
								'conversion_factor': 1,
								'qty':to_bill_qty,
								'basic_rate':i.rate_with_overheads,
								'custom_against_pb':i.docname,
								'project':project
							})
						i.billed_qty += to_bill_qty
						do = frappe.get_doc(i.pb_doctype,i.docname)
						do.billed_qty += to_bill_qty
						do.save(ignore_permissions=True)
			if len(se.items) > 0:
				se.save(ignore_permissions=True)
				se.submit()
			pb.flags.ignore_validate_update_after_submit = True
			pb.save(ignore_permissions=True)
		else:
			project_budget = frappe.db.get_value("Sales Order",doc.so_no,'project_budget')
			so = frappe.get_doc("Project Budget",project_budget)
			se = frappe.new_doc("Stock Entry")
			se.stock_entry_type = "Material Receipt" 
			se.company = doc.company
			se.reference_number = doc.name
			se.custom_reference_type = "Sales Invoice"
			set_warehouse = frappe.db.get_value("Warehouse",{'warehouse_name': ['like', '%Work In Progress%'],'company':doc.company},'name')
			for i in doc.custom_project_materials:
				if i.quantity:
					# project_name = frappe.get_value("Project",{'name':i.project},['project_name'])
					# pro_name = i.project +' - '+project_name
					# set_warehouse = frappe.db.get_value("Warehouse",{'warehouse_name': ['like', i.project]},'name')
					se.append("items",{
						't_warehouse':set_warehouse,
						'item_code':i.item_code,
						'qty':-(i.quantity),
						'basic_rate':0,
						'project':i.project
					})
					for j in so.item_table:
						if i.item_code==j.item:
							j.billed_qty += i.quantity
							do = frappe.get_doc(i.pb_doctype,i.docname)
							do.billed_qty += i.quantity
							do.save(ignore_permissions=True)
			so.flags.ignore_validate_update_after_submit = True
			so.save(ignore_permissions=True)
			se.save(ignore_permissions=True)
			se.submit()



#cancel the stock entry - on cancel of dn
@frappe.whitelist()
def on_cancel_dn(doc,method):
	if frappe.db.exists("Stock Entry",{"reference_number":doc.name}):
		se = frappe.db.get_value("Stock Entry",{"reference_number":doc.name},['name'])
		if se:
			se_cancel = frappe.get_doc("Stock Entry",se)
			if se_cancel.docstatus == 1:
				se_cancel.cancel()

#Revert the below process while cancel of sales invoice
@frappe.whitelist()
def cancel_stock_on_si(doc,method):
	if doc.is_return==0:
		if frappe.db.exists("Stock Entry",{"reference_number":doc.name,'docstatus':1}):
			se = frappe.db.get_value("Stock Entry",{"reference_number":doc.name},['name'])
			if se:
				se_cancel = frappe.get_doc("Stock Entry",se)
				se_cancel.cancel()
				for i in se_cancel.items:
					project_budet=frappe.db.get_value("Sales Order",{'name':doc.so_no},['project_budget'])
					do = frappe.get_doc("Project Budget Items",{"docname": i.custom_against_pb,"parent":project_budet})
					do.billed_qty -= i.qty
					pb = frappe.get_doc(do.pb_doctype,do.docname)
					pb.billed_qty -= i.qty
					pb.save(ignore_permissions=True)
					if do.docstatus != 2:
						do.save(ignore_permissions=True)
				
	else:
		se = frappe.db.get_value("Stock Entry",{"reference_number":doc.name,'docstatus':1},['name'])
		if se:
			se_cancel = frappe.get_doc("Stock Entry",se)
			se_cancel.cancel()
		for i in doc.custom_project_materials:
			j = frappe.get_doc("Project Budget Items",{'docname':i.docname})
			j.billed_qty -= (float(i.quantity))
			do = frappe.get_doc(i.pb_doctype,i.docname)
			do.billed_qty -= (float(i.quantity))
			do.save(ignore_permissions=True)
			j.flags.ignore_validate_update_after_submit = True
			j.save(ignore_permissions=True)
	
# @frappe.whitelist()
# def get_so_child_values(sales_order):
# 	document = frappe.get_doc("Sales Order",sales_order)
# 	return document.items


#Update the money in words for advance invoice
@frappe.whitelist()
def update_advance_money(advance):
	words=frappe.utils.money_in_words(advance)
	return words

@frappe.whitelist()
def update_outstanding_amount(doc,method):
	value=0
	for i in doc.scope_of_work:
		value+=i.invoice_amount
	return value

#Create new journal entry while submission of advance invoice
@frappe.whitelist()
def create_new_journal_entry(doc,method):
	document = frappe.get_doc("Advance Invoice", doc.name)
	sales=frappe.get_doc("Sales Order",doc.sales_order)
	jv = frappe.new_doc("Journal Entry")
	jv.voucher_type = "Journal Entry"
	jv.company = document.company
	jv.posting_date = doc.transaction_date
	jv.user_remark = f"Customer Name : {sales.customer} Order No : {doc.sales_order} Order Date : {sales.transaction_date} Project : {sales.project}"
	jv.cheque_no = doc.name
	jv.custom_advance_invoice=doc.name
	jv.cheque_date = document.transaction_date
	advance_account = frappe.db.get_value("Company", {"name": document.company}, "custom_default_advance_account")
	receivable_account = frappe.db.get_value("Company", {"name": document.company}, "default_receivable_account")
	accounts = [
		{
			'account': advance_account,
			'credit_in_account_currency': document.advance_amount
		},
		{
			'account': receivable_account,
			'debit_in_account_currency': document.advance_amount,
			'party_type': 'Customer',
			'party': document.customer
		}
	]

	for account in accounts:
		jv.append('accounts', account)

	jv.save(ignore_permissions=True)
	jv.submit()
	custom_invoice = jv.custom_advance_invoice
	existing_docs = frappe.db.sql("""
		SELECT name FROM `tabJournal Entry`
		WHERE name LIKE %s
		ORDER BY name DESC
	""", (custom_invoice + '%'), as_dict=True)
	if not existing_docs:
		new_name = f"{custom_invoice}-01"
	else:
		last_doc = existing_docs[0]
		last_number = int(last_doc.name.split('-')[-1])
		new_sequence = last_number + 1
		new_name = f"{custom_invoice}-{new_sequence}"
	frappe.rename_doc("Journal Entry", jv.name, new_name, force=1)


#Create new journal entry while submission of advance invoice
@frappe.whitelist()
def create_new_journal_entry_retention(doc,method):
	document = frappe.get_doc("Retention Invoice", doc.name)
	sales=frappe.get_doc("Sales Order",doc.sales_order)
	jv = frappe.new_doc("Journal Entry")
	jv.voucher_type = "Journal Entry"
	jv.company = document.company
	jv.posting_date = doc.transaction_date
	jv.user_remark = f"Customer Name : {sales.customer} Order No : {doc.sales_order} Order Date : {sales.transaction_date} Project : {sales.project}"
	jv.cheque_no = doc.name
	jv.custom_advance_invoice=doc.name
	jv.cheque_date = document.transaction_date
	advance_account = frappe.db.get_value("Company", {"name": document.company}, "custom_default_retention_account")
	receivable_account = frappe.db.get_value("Company", {"name": document.company}, "default_receivable_account")
	accounts = [
		{
			'account': advance_account,
			'credit_in_account_currency': document.advance_amount
		},
		{
			'account': receivable_account,
			'debit_in_account_currency': document.advance_amount,
			'party_type': 'Customer',
			'party': document.customer
		}
	]

	for account in accounts:
		jv.append('accounts', account)

	jv.save(ignore_permissions=True)
	jv.submit()
	custom_invoice = jv.custom_advance_invoice
	existing_docs = frappe.db.sql("""
		SELECT name FROM `tabJournal Entry`
		WHERE name LIKE %s
		ORDER BY name DESC
	""", (custom_invoice + '%'), as_dict=True)
	if not existing_docs:
		new_name = f"{custom_invoice}-01"
	else:
		last_doc = existing_docs[0]
		last_number = int(last_doc.name.split('-')[-1])
		new_sequence = last_number + 1
		new_name = f"{custom_invoice}-{new_sequence}"
	frappe.rename_doc("Journal Entry", jv.name, new_name, force=1)

#Check the below condition is validate of advance invoice
@frappe.whitelist()
def calculate_adv_amount(so,advance):
	invoice=frappe.db.get_value("Advance Invoice",{"sales_order":so,"docstatus":("!=",2)},["advance_amount"])
	sales_invoices = frappe.get_all("Sales Invoice", {"sales_order": so, "docstatus":1}, ["adv_amount"])
	total_adv_amount = sum(flt(invoice['adv_amount']) for invoice in sales_invoices)
	if invoice:
		if flt(invoice) >= (flt(advance) + flt(total_adv_amount)):
			return "True"
		else:
			frappe.throw("Not allowed to advance amount is greater than advance invoice amount")

@frappe.whitelist()
def delete_leave_application_test():
	leave_application=frappe.get_all("Leave Revocation",{'employee':'345'},['*'])
	i=0
	for la in leave_application:
		print(la.name)
		frappe.db.sql("delete from `tabLeave Revocation` where employee=%s"%(la.employee))
		i+=1
	print(i)

#Update the project budget linked in MR While submission of MR
@frappe.whitelist()	
def mr_qty(doc,method):
	if doc.project_budget:
		s=frappe.get_doc("Project Budget",doc.project_budget)
		if s.docstatus == 1:
			for i in doc.items:
				for j in s.item_table:
					if i.sales_order==s.sales_order and i.item_code == j.item:
						if i.custom_project_budget_item:
							if j.name == i.custom_project_budget_item:
								j.custom_mr_qty +=i.qty
						else:
							j.custom_mr_qty +=i.qty
				s.save(ignore_permissions=True)
				s.flags.ignore_validate_update_after_submit = True


#Update the below values while open the dn
@frappe.whitelist()
def update_dnwip_value(items, budget):
	project = frappe.get_doc("Project Budget", budget)
	for item_code in items:
		for j in project.item_table:
			if item_code == j.item:
				index = items.index(item_code)
				items[index] = {
					'item_code': item_code,
					'rate': j.rate_with_overheads
				}
	return items


#Check the below condition while submission of stock entry
@frappe.whitelist()
def update_req_qty(doc,method):
	budget = frappe.db.get_value("Sales Order", {"name":doc.sales_order, "docstatus": ("!=", 2)}, "project_budget")
	if budget:
		project = frappe.get_doc("Project Budget", budget)
		for item in doc.items:
			for j in project.item_table:
				if item.item_code==j.item and (item.docname == j.docname and item.msow == j.msow):
					if item.qty > j.qty:
						frappe.throw(f"Stock Request qty for item {item.item_code} exceeds the project Budget qty.")

#Cancel the journal entry to match with cancel advance invoice
@frappe.whitelist()
def cancel_journal_entry(doc,method):
	jv = frappe.db.get_value("Journal Entry",{"cheque_no":doc.name},["name"])
	if jv:
		jv_cancel=frappe.get_doc("Journal Entry",jv)
		if jv_cancel.docstatus == 1:
			jv_cancel.cancel()

#Cancel the journal entry to match with cancel advance invoice
@frappe.whitelist()
def cancel_journal_entry_retention(doc,method):
	jv = frappe.db.get_value("Journal Entry",{"cheque_no":doc.name},["name"])
	if jv:
		jv_cancel=frappe.get_doc("Journal Entry",jv)
		if jv_cancel.docstatus == 1:
			jv_cancel.cancel()

#Check the approved by finance button is enable or disable , if disable throw the below message in sales invoice
@frappe.whitelist()
def check_approved_by_finance(doc,method):
	if doc.custom_approved_by_finance==0 and doc.order_type=="Project" and doc.is_return==0:
		frappe.throw("This document can be submitted only after approval from the Finance Department")


#Get the reservation entry list and sent to mail for below receiptant
@frappe.whitelist()
def reservation_entrylist():
	create = frappe.db.sql("""
	SELECT creation,name,item_code, voucher_type, status
	FROM `tabStock Reservation Entry`
	WHERE status IN ('Partially Reserved','Reserved', 'Partially Delivered') AND docstatus!=2
	""", as_dict=True)
	s_no=0
	data = f"""<table class='table table-bordered' style='border-collapse: collapse; width: 100%;'><tr style='border: 1px solid black; background-color: #0f1568; color: white;'><th>S No</th><th>ID</th><th>Voucher Type</th><th>Status</th></tr>"""

	for i in create:
		s_no+=1
		d=date_diff(nowdate(),i.creation.date())
		# print(d)
		if d>10:
			data+= f"""
					<tr style='border: 1px solid black;'>
						<td>{s_no}</td><td>{i.name}</td><td>{i.voucher_type}</td><td>{i.status}</td>
					</tr>"""
	data+=f"""</table>"""
	subject="Stock Reservation Entry"
	message=f"""
	Dear Sir,<br><br>
	Kindly find the document mentioned below, which was created 15 days ago.<br><br>
	{data}"""
	frappe.sendmail(
		recipients=['rifdy@electraqatar.com','aneesh@electraqatar.com','kapil@electraqatar.com'],
		subject=subject,
		message=message,
	)
					


#Get the previous document purchase rate and set the current document previous purchase rate field in landed cost voucher
@frappe.whitelist()
def get_previous_purchase_rate(item_code):
	pr = frappe.db.sql("""
		SELECT 
			lci.current_rate_after_lcv
		FROM 
			`tabLanded Cost Item` lci
		LEFT JOIN 
			`tabLanded Cost Voucher` lcv
		ON 
			lci.parent = lcv.name
		JOIN 
			(SELECT 
				item_code, MAX(creation) as latest_creation
			FROM 
				`tabLanded Cost Item`
			WHERE 
				current_rate_after_lcv IS NOT NULL
			GROUP BY 
				item_code) latest_lci
		ON 
			lci.item_code= latest_lci.item_code
			AND lci.creation = latest_lci.latest_creation
		WHERE 
			lci.item_code = %s and lcv.docstatus !=2
	""", item_code, as_dict=True)

	if pr:
		return pr[0].get('current_rate_after_lcv')
	else:
		return '0.0'

#Update the below details based on next due in legal compliance monitor
@frappe.whitelist()
def update_days_left():
	value = frappe.get_all("Legal Compliance Monitor",["next_due","name"])
	for i in value:
		if i.next_due:
			diff = date_diff(i.next_due, today())
			frappe.db.set_value("Legal Compliance Monitor",i.name,"days_left",diff)
			if 0 <= diff < 31:
				frappe.db.set_value("Legal Compliance Monitor",i.name,"status","Expiring Soon")
			elif diff <0:
				frappe.db.set_value("Legal Compliance Monitor",i.name,"status","Expired")
			else:
				frappe.db.set_value("Legal Compliance Monitor",i.name,"status","Valid")


#Check the supplier is approved or not while save the PO , if not throw the below error
@frappe.whitelist()
def check_approved_supplier(doc,method):
	value = frappe.db.get_value("Supplier",{"name":doc.supplier},["approved_supplier"])
	if value==0:
		frappe.throw("Not an approved supplier")

#Linked dn has been cancelled while cancel of si
@frappe.whitelist()
def cancel_dn_on_si_cancel(doc,method):
	if doc.custom_delivery_note_return:
		dn_return_doc=frappe.get_doc('Delivery Note',doc.custom_delivery_note_return)
		dn_return_doc.cancel()

#Set the warehouse name as like project name in material request
@frappe.whitelist()
def set_project_warehouse(project):
	value=frappe.db.sql("""SELECT name FROM `tabWarehouse` WHERE name LIKE %s""",(f"%{project}%"),as_dict=True)
	return value[0].name

@frappe.whitelist()
def update_ctc():
	value = frappe.get_all("Employee",{'status':'Active'},['*'])
	for i in value:
		ctc = i.gross_salary + i.gratuity + i.leave_salary + i.air_ticket_allowance_ + i.qid_cost + i.medical_renewal + i.medical_allowance_ + i.visa_cost_ + i.compensationemployee_insurence + i.custom_transporttion_expense
		frappe.db.set_value("Employee",i.name,'ctc',ctc)
				
#set the below values are submission of SI
@frappe.whitelist()
def update_gl_internal_customer(doc,method):
	if doc.is_internal_customer and doc.stock_transfer_numner:
		stock_in_hand = frappe.get_all('GL Entry',{'voucher_no':doc.name,'account':['like', '%Stock In Hand%']},['name','credit','credit_in_account_currency'])
		cogs = frappe.get_all('GL Entry',{'voucher_no':doc.name,'account':['like', '%Cost of Goods Sold%']},['name','debit','debit_in_account_currency'])
		frappe.db.set_value('GL Entry',stock_in_hand[0],'credit',doc.grand_total)
		frappe.db.set_value('GL Entry',stock_in_hand[0],'credit_in_account_currency',doc.grand_total)
		frappe.db.set_value('GL Entry',cogs[0],'debit',doc.grand_total)
		frappe.db.set_value('GL Entry',cogs[0],'debit_in_account_currency',doc.grand_total)



# @frappe.whitelist()
# def remove_credit_limit():
#     rows_to_delete = frappe.get_all(
#         "Customer Credit Limit",
#         filters={"credit_limit": 100000},
#         fields=["name"]
#     )
#     for row in rows_to_delete:
#         frappe.db.delete("Customer Credit Limit", {"name": row["name"]})
	
#     remaining_rows = frappe.get_all(
#         "Customer Credit Limit",
#         fields=["name", "parent", "idx"],
#         order_by="parent, idx"
#     )
	
#     grouped_rows = {}
#     for row in remaining_rows:
#         if row["parent"] not in grouped_rows:
#             grouped_rows[row["parent"]] = []
#         grouped_rows[row["parent"]].append(row)

#     for parent, rows in grouped_rows.items():
#         for i, row in enumerate(rows):
#             frappe.db.set_value("Customer Credit Limit", row["name"], "idx", i + 1)

@frappe.whitelist()
def update_confirm_date():
	item = frappe.db.sql(""" update `tabSales Invoice` set return_document = ' ' where name = 'SHBO-CRD-2024-00073'  """)

@frappe.whitelist()
def update_query():
	# updated=frappe.db.sql(""" update `tabCost Estimation` set lead_customer="LAYALI RESTAURANT",lead_customer_name="LAYALI RESTAURANT"  where name ="MEP-CE-2024-00138" """ )
	frappe.db.sql("""UPDATE `tabSales Invoice` SET po_no="4496002353-1" WHERE name="MEP-CRD-2024-00372" """)
	# project=frappe.db.sql("""  update `tabProject Budget set lead_customer_name="LAYALI RESTAURANT" where name="MEP-PB-2024-00104"  """)

@frappe.whitelist()
def custom_submitted_date():
	value = frappe.get_all("Purchase Invoice",{"stock_confirmation":"Stock Confirmation","docstatus":("!=",2)},["posting_date","confirmation_number"])
	for i in value:
		frappe.db.set_value("Stock Confirmation",i.confirmation_number,"custom_submitted_date",i.posting_date)

@frappe.whitelist()
def custom_submitted_date_si():
	value = frappe.get_all("Sales Invoice",{"stock_transfer":"Stock Transfer","docstatus":("!=",2)},["posting_date","stock_transfer_numner"])
	for i in value:
		frappe.db.set_value("Stock Transfer",i.stock_transfer_numner,"custom_submitted_date",i.posting_date)

#Set the submission date stock transfer
@frappe.whitelist()
def set_submitted_date(doc,method):
	frappe.db.sql(""" update `tabStock Transfer` set custom_submitted_date = '%s' where name = '%s' """,(nowdate(),doc.name))

#Set the submission date of stock confirmation
@frappe.whitelist()
def set_submitted_date_con(doc,method):
	frappe.db.sql(""" update `tabStock Confirmation` set custom_submitted_date = '%s' where name = '%s' """,(nowdate(),doc.name))

#Check the items table are empty , if empty throw the message while save the cesow
@frappe.whitelist()
def validate_item(doc,method):
	if doc.a_design and not doc.design:
		return frappe.throw(_("Item not added in Table: Design"))
	if doc.b_materials:
		if doc.company == 'MEP DIVISION - ELECTRA':
			if not doc.mep_materials:
				return frappe.throw(_("Item not added in Table: Supply Materials"))
		if doc.company != 'MEP DIVISION - ELECTRA':
			if not doc.materials:
				return frappe.throw(_("Item not added in Table: Supply Materials"))
	if doc.c_finishing_work and not doc.finishing_work:
		return frappe.throw(_("Item not added in Table: Finishing Work"))
	if doc.d_accessories and not doc.bolts_accessories:
		return frappe.throw(_("Item not added in Table: Accessories"))
	if doc.e_installation and not doc.installation:
		return frappe.throw(_("Item not added in Table: Installation"))
	if doc.f_manpower and not doc.manpower:
		return frappe.throw(_("Item not added in Table: Manpower"))
	if doc.g_tools__equipment__transport and not doc.heavy_equipments:
		return frappe.throw(_("Item not added in Table: Heavy Equipments"))
	if doc.h_others and not doc.others:
		return frappe.throw(_("Item not added in Table: Subcontract"))
	if doc.j_raw and not doc.raw_materials:
		return frappe.throw(_("Item not added in Table: Raw Materials"))

#Check the item there in item list , if not throw the below message while save the PB SOW
@frappe.whitelist()
def validate_item_pb(doc,method):
	if doc.design:
		for i in doc.design:
			if not frappe.db.exists("Item",{'name':i.item}):
				return frappe.throw(_("Item: {0} in Design table not found in Item Master").format(i.item))
	if doc.mep_materials:
		for i in doc.mep_materials:
			if not frappe.db.exists("Item",{'name':i.item}):
				return frappe.throw(_("Item: {0} in Materials table not found in Item Master").format(i.item))
	if doc.materials:
		for i in doc.materials:
			if not frappe.db.exists("Item",{'name':i.item}):
				return frappe.throw(_("Item: {0} in Materials table not found in Item Master").format(i.item))
	if doc.finishing_work:
		for i in doc.finishing_work:
			if not frappe.db.exists("Item",{'name':i.item}):
				return frappe.throw(_("Item: {0} in Finishing Work table not found in Item Master").format(i.item))
	if doc.bolts_accessories:
		for i in doc.bolts_accessories:
			if not frappe.db.exists("Item",{'name':i.item}):
				return frappe.throw(_("Item: {0} in Accessories table not found in Item Master").format(i.item))
	if doc.installation:
		for i in doc.installation:
			if not frappe.db.exists("Item",{'name':i.item}):
				return frappe.throw(_("Item: {0} in Installation table not found in Item Master").format(i.item))
	if doc.heavy_equipments:
		for i in doc.heavy_equipments:
			if not frappe.db.exists("Item",{'name':i.item}):
				return frappe.throw(_("Item: {0} in Heavy Equipments table not found in Item Master").format(i.item))
	if doc.raw_materials:
		for i in doc.raw_materials:
			if not frappe.db.exists("Item",{'name':i.item}):
				return frappe.throw(_("Item: {0} in Raw Materials table not found in Item Master").format(i.item))

#Update the dn wip child table from the project budget
@frappe.whitelist()
def update_wip_child(budget):
	document = frappe.get_doc("Project Budget",budget)
	return document.finished_goods

#update stock entry table
@frappe.whitelist()
def update_raw_material(project):
	pro=frappe.db.get_value("Project",{"name":project},["budgeting"])
	parent_doc=frappe.get_doc("Project Budget",pro)
	warehouse=frappe.db.get_value('Warehouse',{'warehouse_name': ['like', '%'+project+'%']},['name'])
	return parent_doc.raw_materials,warehouse

#Update the date while open the leave salary
@frappe.whitelist()
def update_start_date(employee):
	value = frappe.db.get_value("Rejoining Form",{"emp_no":employee,"workflow_state":"Approved"},["re_join"])
	return value

from dateutil.relativedelta import relativedelta
#Update the service between the given 2 date in leave salary
@frappe.whitelist()
def update_service_period(start_date,end_date):

	start_date = datetime.strptime(start_date, '%Y-%m-%d')
	end_date = datetime.strptime(end_date, '%Y-%m-%d')
	difference = relativedelta(end_date, start_date)
	years = difference.years
	months = difference.months
	days = difference.days +1
	if days < 0:
		months -= 1
		prev_month = end_date - relativedelta(months=1)
		days_in_prev_month = (prev_month + relativedelta(months=1)).replace(day=1) - prev_month.replace(day=1)
		days += days_in_prev_month.days
	return f"{years} Years, {months} Months, {days} Days"

#Update the leave count between the given 2 date in leave salary
@frappe.whitelist()
def leave_count(start_date,end_date,employee):
	count = 0
	start_date = str(start_date)
	end_date = str(end_date)
	attendance = frappe.get_all(
	"Attendance",
		filters={
			"employee": employee,
			"attendance_date": ["between", (start_date, end_date)],
			"docstatus": ["!=", 2],
			"leave_type": ["in", ["Annual Leave", "Leave Without Pay"]]
		},
		fields=["status"]
	)

	for i in attendance:
		if i.status =="On Leave":
			count +=1
		if i.status =="Half Day":
			count+=0.5
	return count

#Avoid duplication
@frappe.whitelist()
def validate_item_price(doc,method):
	if doc.is_new():
		value = frappe.db.exists("Item Price",{"item_code":doc.item_code,"price_list":doc.price_list})
		if value:
			frappe.throw("Item Price Already there in system")

#Update the below values while submission of sales invoice
@frappe.whitelist()
def update_net_total(doc, method):
	if doc.is_return:
		if doc.adv_amount and doc.ret_amount:
			final_net_total=-(-doc.net_total_project+doc.adv_amount+doc.ret_amount)
			frappe.db.set_value("Sales Invoice",doc.name,"net_total_project",final_net_total)
		elif doc.adv_amount and not doc.ret_amount:
			final_net_total=-(-doc.net_total_project+doc.adv_amount)
			frappe.db.set_value("Sales Invoice",doc.name,"net_total_project",final_net_total)
		elif doc.ret_amount and not doc.adv_amount:
			final_net_total=-(-doc.net_total_project+doc.ret_amount)
			frappe.db.set_value("Sales Invoice",doc.name,"net_total_project",final_net_total)
	else:
		if doc.adv_amount and doc.ret_amount:
			final_net_total=doc.net_total_project+doc.adv_amount+doc.ret_amount
			frappe.db.set_value("Sales Invoice",doc.name,"net_total_project",final_net_total)
		elif doc.adv_amount and not doc.ret_amount:
			final_net_total=doc.net_total_project+doc.adv_amount
			frappe.db.set_value("Sales Invoice",doc.name,"net_total_project",final_net_total)
		elif doc.ret_amount and not doc.adv_amount:
			final_net_total=doc.net_total_project+doc.ret_amount
			frappe.db.set_value("Sales Invoice",doc.name,"net_total_project",final_net_total)
  
@frappe.whitelist()
def update_net_total_cl(name):
	doc = frappe.get_doc("Sales Invoice", name)
	if doc.adv_amount and doc.ret_amount:
		
		final_net_total=doc.net_total_project+doc.adv_amount+doc.ret_amount
		frappe.db.set_value("Sales Invoice",doc.name,"net_total_project",final_net_total)
	elif doc.adv_amount and not doc.ret_amount:
		final_net_total=doc.net_total_project+doc.adv_amount
		frappe.db.set_value("Sales Invoice",doc.name,"net_total_project",final_net_total)
	elif doc.ret_amount and not doc.adv_amount:
		final_net_total=doc.net_total_project+doc.ret_amount
		frappe.db.set_value("Sales Invoice",doc.name,"net_total_project",final_net_total)

import math
#Update the per hour cost and gratuity for every month day first
@frappe.whitelist()
def update_per_hour_cost():
	employee = frappe.get_all("Employee",{"status":"Active"},["*"])
	for i in employee:
		doj = frappe.db.get_value('Employee',{"name":i.name},['date_of_joining'])
		current_yoe_days = date_diff(nowdate(),doj)
		current_yoe = round((current_yoe_days / 365),3)
		if current_yoe < 5:
			gratuity_per_year = (int(i.basic)/30) * 12
			total_gratuity = math.ceil(gratuity_per_year/365*current_yoe_days)
		if current_yoe >= 5:
			gratuity_per_year = (int(i.basic)/21) * 12
			total_gratuity = math.ceil(gratuity_per_year/365*current_yoe_days)
		frappe.db.set_value("Employee",i.name,"gratuity",total_gratuity)	
		value = i.basic + i._other_allowance + i.custom_accomodation_ctc + i.transportation + (i.air_ticket_allowance_/12)+(gratuity_per_year/12)+(i.medical_renewal/12)+(i.visa_cost_/12)+(i.leave_salary/12)+(i.compensationemployee_insurence/12)
		# new_per_cost_hour = (value/30/8)
		new_per_cost_hour = i.basic/30/8
		frappe.db.set_value("Employee",i.name,"per_hour_cost",new_per_cost_hour)
		ctc = i.gross_salary + i.gratuity + i.leave_salary + i.air_ticket_allowance_ + i.qid_cost + i.medical_renewal + i.medical_allowance_ + i.visa_cost_ + i.compensationemployee_insurence
		frappe.db.set_value("Employee",i.name,"ctc",ctc)
# @frappe.whitelist()
# def create_per_hour_cost():
# 	job = frappe.db.exists('Scheduled Job Type', 'electra.custom.update_per_hour_cost')
# 	if not job:
# 		att = frappe.new_doc("Scheduled Job Type")
# 		att.update({
# 			"method": 'electra.custom.update_per_hour_cost',
# 			"frequency": 'Cron',
# 			"cron_format": '0 0 1 * *'
# 		})
# 		att.save(ignore_permissions=True)
		
#Cance PI while cancel of stock confirmation
@frappe.whitelist()
def cancel_purchase_invoice(doc,method):
	pi = frappe.db.get_value("Purchase Invoice",{"stock_confirmation":"Stock Confirmation","confirmation_number":doc.name,"docstatus":1},["name"])
	if pi:
		document = frappe.get_doc("Purchase Invoice",pi)
		document.cancel()

#Cance SI while cancel of stock Transfer
@frappe.whitelist()
def cancel_sales_invoice(doc,method):
	pi = frappe.db.get_value("Sales Invoice",{"stock_transfer":"Stock Transfer","stock_transfer_numner":doc.name,"docstatus":1},["name"])
	if pi:
		document = frappe.get_doc("Sales Invoice",pi)
		document.cancel()
	sc = frappe.db.get_value("Stock Confirmation",{"ic_material_transfer_confirmation":doc.name},["name"])
	if sc:
		value = frappe.get_doc("Stock Confirmation",sc)
		if value.docstatus == 0:
			value.delete()

# def itemcode_del():
#     filename='Cutt Off(File) - Sheet1 (1).csv'
#     from frappe.utils.file_manager import get_file
#     filepath = get_file(filename)
#     pps = read_csv_content(filepath[1])
#     ind=0
#     for pp in pps:
		
#         ip=frappe.db.exists("Item Price",{"item_code":pp[0],"price_list":"Cut Off Price"},["name"])
#         if ip:
#             print(pp[0])
#             ind+=1
#             frappe.delete_doc("Item Price", ip)

# @frappe.whitelist()
# def get_timesheet(from_date,to_date):
#     pos = frappe.db.sql("""
#         SELECT
#             SUM(`tabTimesheet Detail`.billing_hours) AS hours,
#             SUM(`tabTimesheet Detail`.costing_amount) AS cost_amount,
#             `tabTimesheet Detail`.project,
#             `tabTimesheet Detail`.project_name,
#             `tabTimesheet`.employee,
#             `tabTimesheet`.employee_name
#         FROM
#             `tabTimesheet`
#         LEFT JOIN
#             `tabTimesheet Detail`
#         ON
#             `tabTimesheet`.name = `tabTimesheet Detail`.parent
#         WHERE
#             `tabTimesheet`.docstatus != 2 AND
#             `tabTimesheet`.start_date BETWEEN %s AND %s AND
#         GROUP BY
#             `tabTimesheet Detail`.project,
#             `tabTimesheet`.employee
#         ORDER BY
#             `tabTimesheet`.employee
#         """, (from_date, to_date), as_dict=True)

	
#     data = "<table class='table table-bordered=1'>"
#     data += """<tr><td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Sl No.</td>
#             <td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Project</td>
#             <td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Project Name</td>
#             <td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Total Hours</td>
#             <td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Total Cost</td></tr>"""
	
	# current_employee = None
	# serial_number = 1
	# total_hours = 0
	# total_cost = 0
	# total_hours_all = 0
	# total_cost_all = 0
	# total_hours_all = sum(row.hours for row in pos)
	# total_cost_all = sum(row.cost_amount for row in pos)
	# for row in pos:
	#     if row.employee != current_employee:
	#         if current_employee is not None:
	#             data += """<tr><td colspan=3 style='border: 1px solid black;text-align:right'><b>Total</b></td>
	#                         <td style='border: 1px solid black;text-align:right'>%s</td>
	#                         <td style='border: 1px solid black;text-align:right'>%s</td></tr>""" % (total_hours, round(total_cost,2))
	#             total_hours = 0
	#             total_cost = 0
				
	#         data += """<tr><td colspan=5 style='background-color:#D5D8DC;border: 1px solid black'><b>%s</b></td></tr>""" % (row.employee +" - "+ row.employee_name)
	#         current_employee = row.employee
	#         serial_number = 1
		
	#     total_hours += row.hours
	#     total_cost += row.cost_amount
		
	#     data += """<tr><td style='border: 1px solid black;text-align:right'>%s</td>
	#                 <td style='border: 1px solid black'>%s</td>
	#                 <td style='border: 1px solid black'>%s</td>
	#                 <td style='border: 1px solid black;text-align:right'>%s</td>
	#                 <td style='border: 1px solid black;text-align:right'>%s</td></tr>""" % (serial_number, row.project, row.project_name, row.hours, round(row.cost_amount,2))
		
	#     serial_number += 1

	# if current_employee is not None:
	#     data += """<tr><td colspan=3 style='border: 1px solid black;text-align:right'><b>Total</b></td>
	#                 <td style='border: 1px solid black;text-align:right'>%s</td>
	#                 <td style='border: 1px solid black;text-align:right'>%s</td></tr>""" % (total_hours, round(total_cost,2))
				
	# data += """<tr><td colspan=3 style='background-color:#D5D8DC;border: 1px solid black;text-align:right'><b>Overall Total</b></td>
	#             <td style='background-color:#D5D8DC;border: 1px solid black;text-align:right'><b>%s</b></td>
	#             <td style='background-color:#D5D8DC;border: 1px solid black;text-align:right'><b>%s</b></td></tr>""" % (round(total_hours_all,2), round(total_cost_all,2))

	# data += "</table>"
	# return data

#Update the stock value return of the sales invoice submission
@frappe.whitelist()
def update_stock_after_return(doc,method):
	if doc.is_return == 1:
		if doc.update_stock == 0:
			if doc.delivery_note:
				s = doc.delivery_note_list
				stat = s.strip("[']'")
				dn_lst = stat.split(",")
				for s in dn_lst:
					sn = frappe.get_doc("Delivery Note",s.strip().lstrip("'").rstrip("'"))
					dn = frappe.copy_doc(sn)
					dn.is_return = 1
					dn.delivery_return = doc.delivery_note_list
					dn.posting_date = doc.posting_date
					dn.naming_series = get_dn_return_series(doc.company,"Delivery Note")
					# dn.set("items",[])
					dn.items=[]
					dn.posting_time = doc.posting_time
					dn.return_against=s.strip().lstrip("'").rstrip("'")
					
					for i in doc.items:
						dn_returns=frappe.get_all("Delivery Note",{'return_against':s.strip().lstrip("'").rstrip("'"),'docstatus':('!=',2)},['name'])
						total_return_qty = -i.qty
						for dnote in dn_returns:
							dn_ret = frappe.get_doc("Delivery Note", dnote.name)
							for dn_r in dn_ret.items:
								if dn_r.item_code == i.item_code and dn_r.name == i.dn_detail:
									total_return_qty += -(dn_r.qty)
						dn_original = frappe.get_doc("Delivery Note", s.strip().lstrip("'").rstrip("'"))
						original_qty = 0
						for item in dn_original.items:
							if item.item_code == i.item_code and item.name == i.dn_detail:
								original_qty = item.qty
								break
						if total_return_qty > original_qty:
							print("hi")
							# frappe.throw(_("Return quantity for item {0} exceeds the original quantity. And already {1} qty have been returned against the DN: {2}").format(i.item_code,(total_return_qty-(-i.qty)),s.strip().lstrip("'").rstrip("'")))
						else:
							
							dn.append("items", {
								"item_code": i.item_code,
								"item_name": i.item_name,
								"description": i.description,
								"qty": i.qty,
								"uom": i.uom,
								"stock_uom": i.uom,
								"rate": i.rate,
								"conversion_factor": 1,
								"base_rate": i.base_rate,
								"amount": i.amount,
								"warehouse": i.warehouse,
								"base_amount": i.base_amount,
								'dn_detail':i.dn_detail,
								# "against_sales_order":i.sales_order,
								# "so_detail":i.so_detail
							})
							
					if dn.items:
						dn.save(ignore_permissions = True)
						dn.submit()
						frappe.db.set_value("Sales Invoice",doc.name,'custom_delivery_note_return',dn.name)
					sn.reload()
					for i in doc.items:
						for si in sn.items:
							if si.item_code == i.item_code and si.name == i.dn_detail:
								si.returned_qty += -i.qty
					ret_qty = sum(i.returned_qty for i in sn.items)
					sn.per_returned = (ret_qty / sn.total_qty)*100
					sn.flags.ignore_validate_update_after_submit = True
					sn.save(ignore_permissions = True)
					
					sn.update_prevdoc_status()
					sn.update_billing_status()
					# dn_item=frappe.get_doc("Delivery Note",dn.name)
					# for dnit in dn_item.items:
					# 	frappe.db.set_value("Delivery Note Item",dnit.dn_detail,'returned_qty',(dnit.dn_detail+(-(dnit.qty))))
				

#Get the dn value and set the sales invoice
@frappe.whitelist()
def update_dn_wip_list(sales_order):
	dn_wip_list=[]
	delivery_note_wip=frappe.get_all("Delivery Note WIP",{'sales_order':sales_order,'docstatus':1},['name'])
	for dn_wip in delivery_note_wip:
		dn_wip_list.append(dn_wip.name)
	return str(dn_wip_list)

#Avoid the duplication entry of Customer
@frappe.whitelist()
def avoid_duplicate_name(doc, method):
	if doc.is_new():
		first_word = doc.customer_name.split()[0]
		
		similar_names = frappe.db.sql("""
			SELECT name 
			FROM `tabCustomer` 
			WHERE customer_name LIKE %s
		""", (f"{first_word}%",))

		if similar_names:
			frappe.msgprint(f"A customer with a similar name exists: {', '.join(row[0] for row in similar_names)}")


#Check Selling price and cut off Price
@frappe.whitelist()
def check_cut_off_amount(doc,method):
	if doc.is_new():
		if doc.quotation_to == 'Customer':
			customer = frappe.db.get_value("Customer",{"name":doc.customer},["custom_allow_bypass_cut_office_price"])
			if customer == 0:
				for i in doc.items:
					if frappe.db.exists("Item Price",{"item_code":i.item_code,"price_list":"Cut Off Price"}):
						rate = frappe.db.get_value("Item Price",{"item_code":i.item_code,"price_list":"Cut Off Price"},["price_list_rate"])
						if rate>i.rate:
							frappe.throw(f"Cut off Rate is Greater than the Selling Rate for this Item - <b>{i.item_code}</b>.The cut-off rate for this item is <b>{rate}</b>,The selling rate is <b>{i.rate}</b>")


@frappe.whitelist()
def count_transfer():
	count = 0
	count1=0
	value1 = frappe.get_all("Stock Transfer", filters={"docstatus":1,"transferred_date":("between",("2024-01-01","2024-09-30")),"source_company":"ELECTRA - BINOMRAN SHOWROOM"}, fields=["name"])	

	for i in value1:
		count +=1
	value = frappe.get_all("Sales Invoice", filters={ "docstatus": ("!=", 2),"posting_date":("between",("2024-01-01","2024-09-30")),"company":"ELECTRA - BINOMRAN SHOWROOM"}, fields=["name","stock_transfer_numner"])	
	for j in value:
		if i.name != j.stock_transfer_numner:
			print(i.name)

	print(count)
	print(count1)


# @frappe.whitelist()
# def last_date(employee):
# 	value = frappe.db.get_value("Leave Salary and Final Exit Request",{"employee_number":employee},["request_date"])
# 	frappe.errprint(value)
# 	return value

@frappe.whitelist()
def get_project_items():
	frappe.db.sql("""update `tabStock Entry Detail` set custom_against_pb = '339db269b8' where parent = 'MAT-STE-2024-01370' and item_code = 'FEVST15K' """,as_dict=1)

	# item_table =  frappe.get_all("Project Budget Items",{'parent':"INT-PB-2024-00020-10",'docname':'3d22399814'},['*'])
	# print(item_table)

from num2words import num2words
@frappe.whitelist()
def amount_in_words_ar(grand_total):
	amount_in_words=num2words(grand_total, lang='ar')
	return amount_in_words

@frappe.whitelist()
def amount_in_words(grand_total):
	amount_in_words=num2words(grand_total, lang='en')
	return amount_in_words

@frappe.whitelist()
def dn_html(si_name):
	data = ''
	data+= '<br><table width=100% border=1px solid black>'
	data+='<tr style="background-color:lightgrey;text-align:center"><td width="full"><b>DN Return</b></td>'

	dn=frappe.db.get_all("Delivery Note",{"custom_invoice_number":si_name,"is_return":1},["*"])
	for i in dn:
		js_route = (
			f"frappe.route_options = {{"
			f"voucher_no: '{i.name}',"
			f"company: '{i.company}',"
			f"from_date: '{i.posting_date}',"
			f"to_date: '{i.posting_date}',"
			f"ignore_prepared_report: true"
			f"}}; "
			f"frappe.set_route('query-report', 'Stock Ledger');"
		)
		data += f"<tr><td align='center'><a href='#' onclick=\"{js_route}\">{i['name']}</a></td></tr>"
	data+='</table>'
	return data

# @frappe.whitelist()
# def update_cust_po_num():
# 	frappe.db.set_value("Sales Order","ENG-SO-2024-00206",'po_no','P-1248-5777-NFS-SCA-0095')


#Get the items values from Product bundle , if enter the value in Material Transfer dialog box
@frappe.whitelist()
def product_bundle_items(bundle):
	bundle_doc = frappe.get_doc("Product Bundle",bundle)
	return bundle_doc.items

#Deleting the CESOW on deleting Cost Estimation
@frappe.whitelist()
def linked_cesow_delete(doc, method):
	cesow_list = frappe.get_all("CE SOW", filters={"cost_estimation": doc.name}, fields=["name"])
	for cesow in cesow_list:
		frappe.db.sql("""delete from `tabCE SOW` where name = '%s' """%(cesow["name"]))
		

#Gratuity Calculation - Leave Salary document
@frappe.whitelist()
def gratuity_calculation(doc,method):
	if doc.is_new():
		count = 0
		start_date = str(doc.joining_or_last_rejoining_date)
		end_date = str(doc.date_of_service)
		basic = frappe.db.get_value("Employee",{"name":doc.employee_number},["basic"])
		hra = frappe.db.get_value("Employee",{"name":doc.employee_number},["hra"])
		allowance = frappe.db.get_value("Employee",{"name":doc.employee_number},["_other_allowance"])

		attendance = frappe.get_all(
		"Attendance",
			filters={
				"employee": doc.employee_number,
				"attendance_date": ["between", (start_date, end_date)],
				"docstatus": ["!=", 2],
				"leave_type": ["in", ["Annual Leave", "Leave Without Pay"]]
			},
			fields=["status"]
		)

		for i in attendance:
			if i.status =="On Leave":
				count +=1
			if i.status =="Half Day":
				count+=0.5

		doc.total_leave = count
		
		
		total_days = date_diff( doc.date_of_service,doc.joining_date)
		payment_days = int(total_days)  + 1 

		doc.leave_salary_days = payment_days

		
		if doc.grade !="STAFF":
			gratuity =(basic /30 * 21)/365*payment_days

		if doc.grade =="STAFF":
			gratuity = (basic /30 * 21)/365* payment_days

		doc.gratuity = gratuity
		
		
		date_object = datetime.strptime(doc.date_of_service, "%Y-%m-%d")
		month_value = date_object.month
		date_of_service = datetime.strptime(doc.date_of_service, "%Y-%m-%d")
		year = date_of_service.year
		date = f"{year}-{month_value:02d}-01"
		diff = date_diff(doc.date_of_service,date)

		doc.salary_payable_days = diff + 1
		

		salary = (basic + allowance) /30 * (diff + 1)

		doc.salary = salary
		

		total = date_diff( end_date,start_date)
		salary_days = int(total) - int(count) + 1
		doc.leave_salary_days_salary = salary_days

		

		if doc.grade !="STAFF":
			leave_salary = (basic /30 * 30)/365* (salary_days)
		
		if doc.grade =="STAFF":
			leave_salary =(basic /30 * 30)/335* salary_days
		
		doc.leave_salary = leave_salary
		

		if doc.type == "End of Service":
			total_earning = (
				(doc.hot_amount or 0) + 
				(doc.salary or 0) + 
				(doc.ticket_allowance or 0) + 
				(doc.leave_salary or 0) + 
				(doc.not_amount or 0) + 
				(doc.gratuity or 0) + 
				(doc.hra or 0) + 
				(doc.other_benefits or 0)
			)

		if doc.type != "End of Service":
			total_earning = (
				(doc.hot_amount or 0) + 
				(doc.leave_salary or 0) + 
				(doc.other_benefits or 0) + 
				(doc.hra or 0) + 
				(doc.ticket_allowance or 0)
			)

		
		doc.total_earnings = total_earning
		
		doc.total_deduction = (
			(doc.loan or 0) + 
			(doc.advance or 0) + 
			(doc.others or 0) + 
			(doc.immigration_cost or 0) + 
			(doc.visa or 0) + 
			(doc.air_ticket_expense or 0) + 
			(doc.mess_expense or 0)
		)

		
		doc.net_pay = doc.total_earnings or 0 -  doc.total_deduction or 0

@frappe.whitelist()
def update_against_pb():
	frappe.db.sql("""update `tabStock Entry Detail` set custom_against_pb = '41afee58f4' where custom_against_pb = '734df2028c'""")
	frappe.db.sql("""update `tabStock Entry Detail` set custom_against_pb = '40d0b746eb' where custom_against_pb = 'ebcfa4307f'""")
	frappe.db.sql("""update `tabStock Entry Detail` set custom_against_pb = '3fd63bfa1b' where custom_against_pb = 'd29356bf05'""")

@frappe.whitelist()
def create_project_wo_print(doc):
	so_name = frappe.db.get_value("Sales Order", {"project": doc.get("name")}, "name")
	if not so_name:
		return "<p>No Sales Order found linked to this project.</p>"
	so = frappe.get_doc("Sales Order", so_name)
	
	data = f"""
	<table width="100%" style="border-collapse: collapse; font-size:10px">
		<tr>
			<th colspan="3" style="text-align: center; font-size: 18px; padding: 10px;">Work Order</th>
		</tr>
		<br><br><br>
		<tr>
			<td width="50%">{so.company or ''}</td>
			<td width="25%"><b>W.O # :</b></td>
			<td width="25%">{so.name or ''}</td>
		</tr>
		<tr>
			<td width="50%">{so.address_display or ''}</td>
			<td width="25%"><b>W.O Date :</b></td>
			<td width="25%">{frappe.utils.formatdate(so.transaction_date) or ''}</td>
		</tr>
		<tr>
			<td width="50%"><b>Attention :</b></td>
			<td width="25%"><b>Quotation :</b></td>
			<td width="25%">{so.quotation or ''}</td>
		</tr>
		<tr>
			<td width="50%"></td>
			<td width="25%"><b>P.O :</b></td>
			<td width="25%">{so.po_no or ''}</td>
		</tr>
	</table><br><br>
	"""
	data += f"""
	<table width="100%" style =font-size:10px>
		<tr>
			<td width="10%"><b>PROJECT:</b></td>
			<td width="90%">{doc.get('project_name') or ''}</td>
		</tr>
	</table><br><br><br>
	"""
	data += """
	<table width="100%" border="1" style="border-collapse: collapse; text-align: left;font-size:10px">
		<thead>
			<tr style="background-color: #f2f2f2;">
				<th width="25%" style = text-align:center>Scope of Work</th>
				<th width="10%" style = text-align:center>Qty</th>
				<th width="15%" style = text-align:center>Unit</th>
				<th width="15%" style = text-align:center>Unit Price</th>
				<th width="15%" style = text-align:center>Discount</th>
				<th width="15%" style = text-align:center>Total SO</th>
			</tr>
		</thead>
		<tbody>
	"""
	for i in so.scope_of_work:
		data += f"""
			<tr>
				<td>{i.msow_desc or ''}</td>
				<td style = text-align:right>{i.qty or 0}</td>
				<td>{i.unit or ''}</td>
				<td style = text-align:right>{i.unit_price or 0.0}</td>
				<td style = text-align:right>{i.discount or 0.0}</td>
				<td style = text-align:right>{i.total_bidding_price or 0.0}</td>
			</tr>
		"""
	data += """
		</tbody>
	</table>
	"""
	data += f"""
	<table width="100%" style =font-size:12px>
		<tr>
			<td width="100%"><b>Other Comments or Special Instructions :</b></td>
		
		</tr>
	</table><br><br><br>
	"""
	return data


@frappe.whitelist()
def create_project_material_print(doc):
	so_name = frappe.db.get_value("Sales Order", {"project": doc.get("name")}, "name")
	if not so_name:
		return "<p>No Sales Order found linked to this project.</p>"
	
	so = frappe.get_doc("Sales Order", so_name)

	data = ""

	def generate_table(items):
		j = 1
		table = f"""
		<h5 style="text-align:center"><b>Project Materials</b></h5>
		<table width=100% border=1>
			<tr>
				<td style='background-color:#f2f2f2;text-align:center;color:black;width:6%'><b>S.No</b></td>
				<td style='background-color:#f2f2f2;text-align:center;color:black;width:20%'><b>Item</b></td>
				<td style='background-color:#f2f2f2;text-align:center;color:black;width:7%'><b>Budget Qty</b></td>
				<td style='background-color:#f2f2f2;text-align:center;color:black;width:20%'><b>Project Warehouse</b></td>
				<td style='background-color:#f2f2f2;text-align:center;color:black;width:20%'><b>Default Warehouse</b></td>
				<td style='background-color:#f2f2f2;text-align:center;color:black;width:7%'><b>Excess Qty</b></td>
				<td style='background-color:#f2f2f2;text-align:center;color:black;width:20%'><b>Remarks</b></td>
			</tr>"""
		
		project_warehouse = frappe.db.get_value("Warehouse", {"warehouse_name": doc.name}, "name")
		default_warehouse = frappe.db.get_value("Warehouse", {"company": so.company, "default_for_stock_transfer": 1}, "name")
		for i in items:
			qty_value = 0
			default_qty_value = 0
			project_warehouse_qty = frappe.get_all("Bin", {"item_code": i.item, "warehouse": project_warehouse}, ["actual_qty"])
			if project_warehouse_qty:
				qty_value = sum(k.actual_qty for k in project_warehouse_qty)
			else:
				qty_value =0
			default_warehouse_qty = frappe.get_all("Bin", {"item_code": i.item, "warehouse": default_warehouse}, ["actual_qty"])
			if default_warehouse_qty:
				default_qty_value = sum(k.actual_qty for k in default_warehouse_qty)
			else:
				default_qty_value =0
			table += f"""
				<tr>
					<td>{j}</td>
					<td>{i.item}</td>
					<td style='text-align:right'>{"%.2f" % i.qty}</td>
					<td style='text-align:right'>{qty_value}</td>
					<td style='text-align:right'>{default_qty_value}</td>
					<td></td>
					<td></td>
				</tr>"""
			j += 1
		table += "</table><br>"
		return table



	all_items = []
	if so.custom_raw_materials:
		all_items.extend(so.custom_raw_materials)
	if so.design:
		all_items.extend(so.design)
	if so.materials:
		all_items.extend(so.materials)
	if so.finishing_work:
		all_items.extend(so.finishing_work)
	if so.accessories:
		all_items.extend(so.accessories)
	if so.installation:
		all_items.extend(so.installation)
	if so.finished_goods:
		all_items.extend(so.finished_goods)
	if so.heavy_equipments:
		all_items.extend(so.heavy_equipments)
	if so.others:
		all_items.extend(so.others)

	if not all_items:
		return "<p>No materials available in this Sales Order.</p>"

	
	data += generate_table(all_items)

	
	return data if data else "<p>No data available for this project.</p>"

@frappe.whitelist()
def sales_order_items_with_bundle(sales_order_name):
	sales_order = frappe.get_doc("Sales Order", sales_order_name)
	all_items = []

	for item in sales_order.items:
		if frappe.db.exists("Product Bundle", item.item_code):
			product_bundle = frappe.get_doc("Product Bundle", item.item_code)
			if product_bundle and product_bundle.items:
				for bundle_item in product_bundle.items:
					all_items.append({
						'item_code': bundle_item.item_code,
						'item_name': bundle_item.description or bundle_item.item_code,
						'uom': bundle_item.uom,
						'qty': bundle_item.qty * item.qty
					})
		else:
			if frappe.db.exists("Item", {"is_stock_item": 1, "item_code": item.item_code}):
				all_items.append({
					'item_code': item.item_code,
					'item_name': item.item_name,
					'uom': item.uom,
					'qty': item.qty
				})

	return all_items

	
@frappe.whitelist()
def update_pb_total_cost():
	project_budgets = frappe.get_all("Project Budget",{"docstatus": 1, "company": "ENGINEERING DIVISION - ELECTRA"},["name", "total_cost_of_the_project"],)
	for pb in project_budgets:
		pb_doc = frappe.get_doc("Project Budget", pb["name"])
		sow = frappe.db.sql("""SELECT name, msow, parent, total_cost FROM `tabBudget Scope of Work` WHERE parent = %s""",(pb_doc.name,),as_dict=True,)
		tot = sum(i["total_cost"] for i in sow)
		if tot != pb_doc.total_cost_of_the_project:
			print([pb_doc.name, pb_doc.total_cost_of_the_project, tot])
			pb_doc.total_cost_of_the_project = tot
			pb_doc.save(ignore_permissions=True)


@frappe.whitelist()
def get_ot_amount(date,employee,ot):
	if date and employee and ot:
		basic = frappe.db.get_value("Employee",{"name":employee},["basic"])
		holiday = frappe.db.get_value("Employee",{"name":employee},["holiday_list"])
		ot_value = ot.split(":")[0]
		ot_hour = float(ot_value)
		if frappe.db.get_value("Holiday",{"parent":holiday,"holiday_date":date}):
			return (basic/30/8 * ot_hour * 1.5)
		else:
			return (basic/30/8 * ot_hour * 1.25)


@frappe.whitelist()
def get_data_for_pb(tc, ec, cp, gpp, tcc, tec, cpc, gppc, tcp, tpb, cmp, netp, neta, contract_value):
	gppc = contract_value - tcp
	gpp = (gppc/contract_value) * 100
	if cmp == "MEP DIVISION - ELECTRA":
		tot = round(float(tec), 2) + round(float(cpc), 2) + round(float(tcc), 2)
		data = f"""
		<table style='width:100%'>
		<tr><td colspan='20' style='text-align:center;border:1px solid black;background-color:#e35310;color:white;font-weight:bold;'>TOTAL COST</td></tr>
		<tr style='background-color:#878f99;color:white'>
			<td colspan='2' style='text-align:center;border:1px solid black;color:white;font-weight:bold;'>S No</td>
			<td colspan='6' style='text-align:center;border:1px solid black;color:white;font-weight:bold;'>Title</td>
			<td colspan='6' style='text-align:center;border:1px solid black;color:white;font-weight:bold;'>Percentage(%)</td>
			<td colspan='6' style='text-align:center;border:1px solid black;color:white;font-weight:bold;'>Amount(QAR)</td>
		</tr>
		<tr>
			<td colspan='2' style='text-align:center;border:1px solid black'><b>A</b></td>
			<td colspan='12' style='text-align:left;border:1px solid black'><b>Total Cost of the Project</b></td>
			<td colspan='6' style='text-align:center;border:1px solid black'>QAR {round(float(tcp), 2)}</td>
		</tr>
		<tr>
			<td colspan='2' style='text-align:center;border:1px solid black'><b>B</b></td>
			<td colspan='6' style='text-align:left;border:1px solid black'><b>Gross Profit</b></td>
			<td colspan='6' style='text-align:center;border:1px solid black'>{round(float(gpp), 2)}</td>
			<td colspan='6' style='text-align:center;border:1px solid black'>QAR {round(float(gppc), 2)}</td>
		</tr>
		<tr>
			<td colspan='2' style='text-align:center;border:1px solid black'><b>C</b></td>
			<td colspan='12' style='text-align:left;border:1px solid black'><b>Total Business Promotion</b></td>
			<td colspan='6' style='text-align:center;border:1px solid black'>QAR {round(float(tpb), 2)}</td>
		</tr>
		</table>
		"""
	else:
		tot = round(float(tec), 2) + round(float(cpc), 2) + round(float(tcc), 2)
		eng = round(float(tcp), 2) + round(float(gppc), 2)
		data = f"""
		<table style='width:100%'>
		<tr><td colspan='20' style='text-align:center;border:1px solid black;background-color:#e35310;color:white;font-weight:bold;'>TOTAL COST</td></tr>
		<tr style='background-color:#878f99;color:white'>
			<td colspan='2' style='text-align:center;border:1px solid black;color:white;font-weight:bold;'>S No</td>
			<td colspan='6' style='text-align:center;border:1px solid black;color:white;font-weight:bold;'>Title</td>
			<td colspan='6' style='text-align:center;border:1px solid black;color:white;font-weight:bold;'>Percentage(%)</td>
			<td colspan='6' style='text-align:center;border:1px solid black;color:white;font-weight:bold;'>Amount(QAR)</td>
		</tr>
		<tr>
			<td colspan='2' style='text-align:center;border:1px solid black'><b>A</b></td>
			<td colspan='12' style='text-align:left;border:1px solid black'><b>Total Cost of the Project</b></td>
			<td colspan='6' style='text-align:center;border:1px solid black'>QAR {round(float(tcp), 2)}</td>
		</tr>
		<tr>
			<td colspan='2' style='text-align:center;border:1px solid black'><b>B</b></td>
			<td colspan='6' style='text-align:left;border:1px solid black'><b>Overhead</b></td>
			<td colspan='6' style='text-align:center;border:1px solid black'>{round(float(tc), 2)}</td>
			<td colspan='6' style='text-align:center;border:1px solid black'>QAR {round(float(tcc), 2)}</td>
		</tr>
		<tr>
			<td colspan='2' style='text-align:center;border:1px solid black'><b>C</b></td>
			<td colspan='6' style='text-align:left;border:1px solid black'><b>Engineering Overhead</b></td>
			<td colspan='6' style='text-align:center;border:1px solid black'>{round(float(ec), 2)}</td>
			<td colspan='6' style='text-align:center;border:1px solid black'>QAR {round(float(tec), 2)}</td>
		</tr>
		<tr>
			<td colspan='2' style='text-align:center;border:1px solid black'><b>D</b></td>
			<td colspan='6' style='text-align:left;border:1px solid black'><b>Contingency</b></td>
			<td colspan='6' style='text-align:center;border:1px solid black'>{round(float(cp), 2)}</td>
			<td colspan='6' style='text-align:center;border:1px solid black'>QAR {round(float(cpc), 2)}</td>
		</tr>
		<tr>
			<td colspan='2' style='text-align:center;border:1px solid black'><b>E</b></td>
			<td colspan='12' style='text-align:left;border:1px solid black'><b>Total Overhead (B+C+D)</b></td>
			<td colspan='6' style='text-align:center;border:1px solid black'>QAR {round(float(tot), 2)}</td>
		</tr>
		<tr>
			<td colspan='2' style='text-align:center;border:1px solid black'><b>F</b></td>
			<td colspan='6' style='text-align:left;border:1px solid black'><b>Gross Profit</b></td>
			<td colspan='6' style='text-align:center;border:1px solid black'>{round(float(gpp), 2)}</td>
			<td colspan='6' style='text-align:center;border:1px solid black'>QAR {round(float(gppc), 2)}</td>
		</tr>
		
		<tr>
			<td colspan='2' style='text-align:center;border:1px solid black'><b>G</b></td>
			<td colspan='6' style='text-align:left;border:1px solid black'><b>Net Profit</b></td>
			<td colspan='6' style='text-align:center;border:1px solid black'>{round(float(netp), 2)}</td>
			<td colspan='6' style='text-align:center;border:1px solid black'>QAR {round(float(neta), 2)}</td>
		</tr>
		<tr>
			<td colspan='2' style='text-align:center;border:1px solid black'><b>H</b></td>
			<td colspan='12' style='text-align:left;border:1px solid black'><b>Total Business Promotion</b></td>
			<td colspan='6' style='text-align:center;border:1px solid black'>QAR {round(float(tpb), 2)}</td>
		</tr>
		<tr>
			<td colspan='14' style='text-align:center;border:1px solid black;background-color:#878f99;color:white;font-weight:bold;'>Total Bidding Price</td>
			<td colspan='6' style='text-align:center;border:1px solid black;background-color:#878f99;color:white;font-weight:bold;'>QAR {round(float(eng), 2)}</td>
		</tr>
		</table>
		"""
	return data

#Get the documents from Consolidated Payment Request and append the Payment Entry child table
@frappe.whitelist()
def get_payment_request(company=None, supplier=None, from_date=None, to_date=None):
	

	payment_requests = frappe.get_all(
		"Consolidated Payment Request",
		filters={"supplier": supplier, "transaction_date": ("between", (from_date, to_date)),"docstatus":0},
		fields=["name"]
	)

	consolidated_data = []

	for request in payment_requests:
		payment = frappe.get_doc("Consolidated Payment Request", request.name)
		
		for entry in payment.purchase_invoice_list:
			po_status = None
			pi_status = None
			pi_status1 = None
			
			if entry.type == "Purchase Order":
				po_status = frappe.db.get_value(
					"Purchase Order", 
					{"status": "Completed", "per_billed": 100, "name": entry.name1}, 
					"name"
				)
			
			if entry.type == "Purchase Invoice":
				pi_status = frappe.db.get_value(
					"Purchase Invoice", 
					{"name": entry.name1,"status":("!=","Paid"),"docstatus":("!=",2)}, 
					"outstanding_amount"
				)
				if pi_status:
					pi_status =pi_status
				else:
					pi_status =0

			if entry.type == "Purchase Invoice":
				pi_status1 = frappe.db.get_value(
					"Purchase Invoice", 
					{"name": entry.name1,"docstatus":("!=",2)}, 
					"status"
				)
			outstanding_amount = float(entry.outstanding_amount or 0.0)
		
			if entry.company == company:
				if not po_status and  not pi_status1 == "Paid":
					
					consolidated_data.append({
						"reference_doctype": entry.type,
						"reference_name": entry.name1,
						"outstanding_amount": outstanding_amount,
						"name":request.name
					})

	return consolidated_data


@frappe.whitelist()
def update_warehouse():
	frappe.db.sql("""update `tabSales Order Item` set warehouse = 'Electra Interior Warehouse - INE' where parent = 'INT-SO-2024-00152'""")

# @frappe.whitelist()
# def set_so_value():
#     material_requests = frappe.get_all("Material Request",{"project": ["is", "set"]}, ["name", "project"])
#     for material_request in material_requests:
#         request = frappe.get_doc("Material Request", material_request.name)
#         sales_order = frappe.db.get_value(
#             "Project", 
#             {"name": material_request.project}, 
#             "sales_order"
#         )
#         for item in request.items:
#             if not item.sales_order:
#                 frappe.db.sql(
#                     """
#                     UPDATE `tabMaterial Request Item`
#                     SET sales_order = %s
#                     WHERE parent = %s
#                     """,
#                     (sales_order, material_request.name)
#                 )

def check_current_user_role(doc,method):
	if doc.custom_approved_by_finance==1:
		current_user = frappe.session.user
		user_roles = frappe.get_roles(current_user)
		if 'Projects Manager' in user_roles:
			frappe.throw("Cannot not cancel this document as it has been already approved by Finance")

@frappe.whitelist()
def update_project_in_payment(so_name):
	project=frappe.db.get_value("Sales Order",{"name":so_name},["project"])
	
	return project

@frappe.whitelist()
def update_project_in_payement_entry(so_name):
	project=frappe.db.get_value("Sales Order",{"name":so_name},["project"])
	
	return project

@frappe.whitelist()
def sales_order_item_name():
	# sales_order=frappe.get_all("Sales Order Item",{'parent':"INT-SO-2024-00212"},['*'])
	# for s in sales_order:
	# 	print(s.name)
	sales_order=frappe.get_all("Sales Invoice Item",{'parent':"INT-CRD-2024-00466"},['*'])
	for s in sales_order:
		print(s.so_detail)

@frappe.whitelist()
def title_of_project():
	frappe.db.set_value("Sales Order","INT-SO-2025-00042",'title_of_project','INT-PRO-2025-00042- The Line Fitout Interiors & Contracting- RAISED FLOOR ')
	frappe.db.set_value("Sales Invoice","INT-CRD-2025-00097",'title_of_project','INT-PRO-2025-00042- The Line Fitout Interiors & Contracting- RAISED FLOOR ')
	# frappe.db.set_value("Sales Invoice","INT-CRD-2025-00077",'title_of_project','INT-PRO-2025-00042- The Line Fitout Interiors & Contracting- RAISED FLOOR ')
	# frappe.db.set_value("Delivery Note WIP","INT-PDN-2025-00151",'title_of_project','INT-PRO-2025-00042- The Line Fitout Interiors & Contracting- RAISED FLOOR ')
	
	# frappe.db.set_value("Delivery Note WIP","INT-PDN-2025-00153",'title_of_project','INT-PRO-2025-00042- The Line Fitout Interiors & Contracting- RAISED FLOOR ')
	# frappe.db.set_value("Advance Invoice","INT-ADV-2024-00097",'title_of_project','INT-PRO-2025-00042- The Line Fitout Interiors & Contracting- RAISED FLOOR ')
	frappe.db.set_value("Quotation","INT-QTN-2025-00042",'title_of_project','INT-PRO-2025-00042- The Line Fitout Interiors & Contracting- RAISED FLOOR ')
	frappe.db.set_value("Cost Estimation","INT-CE-2025-00043",'project_name','INT-PRO-2025-00042- The Line Fitout Interiors & Contracting- RAISED FLOOR ')


@frappe.whitelist()
def create_job_fail():
	job = frappe.db.exists('Scheduled Job Type', 'cron_failed')
	if not job:
		emc = frappe.new_doc("Scheduled Job Type")
		emc.update({
			"method": 'electra.custom.cron_failed_method',
			"frequency": 'Cron',
			"cron_format": '*/5 * * * *'
		})
		emc.save(ignore_permissions=True)


@frappe.whitelist()
def cron_failed_method():
	cutoff_time = datetime.now() - timedelta(minutes=5)
	failed_jobs = frappe.get_all(
		"Scheduled Job Log",
		filters={
			"status": "Failed",
			"creation": [">=", cutoff_time]
		},
		fields=["scheduled_job_type"]
	)
	unique_job_types = set()
	for job in failed_jobs:
		unique_job_types.add(job['scheduled_job_type'])

	for job_type in unique_job_types:
		frappe.sendmail(
			recipients = ["erp@groupteampro.com","jenisha.p@groupteampro.com","pavithra.s@groupteampro.com","gifty.p@groupteampro.com"],
			subject = 'Failed Cron List - Electra',
			message = 'Dear Sir / Mam <br> Kindly find the below failed Scheduled Job  %s'%(job_type)
		)

@frappe.whitelist()
def get_series_value():
	print(frappe.db.get_value("Purchase Order",{"name":"ELV-PO-2025-00002"},["naming_series"]))

@frappe.whitelist()
def cal_dn_amount():
	dn_amount=frappe.db.sql("""SELECT per.parent,sum(per.base_net_amount) AS total_amount, per.item_code
		FROM `tabDelivery Note Item` AS per
		LEFT JOIN `tabDelivery Note` AS pe ON per.parent = pe.name
		WHERE pe.docstatus != 2 and per.against_sales_order = 'MEP-SO-2024-00255' group by per.parent""",as_dict=True)
	print(dn_amount)

# Manually update the status
@frappe.whitelist()
def attendance_correction():
	checkin = frappe.db.sql("""
	UPDATE `tabStock Confirmation`
	SET workflow_state = 'Material Recieved' , docstatus=1
	WHERE name = 'TRD-STC-2024-00152-1'
""", as_dict=True)
 
@frappe.whitelist()
def find_dn_qty_for_item():
	sales_order = "ENG-SO-2024-00186"
	item = "MSCHS16876M"
	dn_wip_list = frappe.get_all("Delivery Note WIP", {"sales_order": sales_order, "is_return": 0})
	item_qty = 0
	for dn_wip in dn_wip_list:
		stock_entry_exist = frappe.db.exists("Stock Entry", {"stock_entry_type": "Material Transfer", "reference_number": dn_wip.name, "docstatus": 1})
		if stock_entry_exist:
			se = frappe.get_doc("Stock Entry", {"stock_entry_type": "Material Transfer", "reference_number": dn_wip.name, "docstatus": 1})
			for row in se.items:
				if row.item_code == item:
					item_qty += row.qty
	return item_qty

@frappe.whitelist()
def get_sales_order(mr):
	
	return frappe.db.get_value("Material Request Item",{"parent":mr},["sales_order"])


@frappe.whitelist()
def cancel_dn_on_si_cancel_test():
	
	dn_return_doc=frappe.get_doc('Delivery Note',"TRD-DNR-2025-00031")
	for dnit in dn_return_doc.items:
		dn_items = frappe.get_doc("Delivery Note Item",dnit.dn_detail)
		print(dnit.qty)
		print(dn_items.returned_qty)
		frappe.db.set_value("Delivery Note Item",dnit.dn_detail,'returned_qty',(dn_items.returned_qty+(dnit.qty)))
  
@frappe.whitelist()
def check_dup_mr():
	project_budget = frappe.get_all("Project Budget", {"docstatus": 1}, "name")
	for pbs in project_budget:
		if pbs.name not in ("MEP-PB-2023-00020-2"):
			pb = frappe.get_doc("Project Budget", pbs.name)
			for row in pb.item_table:
				if row.custom_mr_qty > row.qty:
					print([pbs.name, row.item])

@frappe.whitelist()
def check_so_total():
	so = frappe.get_doc("Sales Order", "INT-SO-2024-00222")
	tot = 0
	for row in so.materials:
		tot += row.amount_with_overheads
	for row in so.accessories:
		tot += row.amount_with_overheads
	for row in so.manpower:
		tot += row.amount_with_overheads
	print(tot)


@frappe.whitelist()
def change_wip_status():
	frappe.db.sql("""update `tabSales Invoice` set docstatus = 0 where name ='ELV-CRD-2025-00047' """)

@frappe.whitelist()
def update_per_delivered_in_so(doc,method):
	project_budget = frappe.get_doc("Project Budget", {'sales_order': doc.sales_order, 'docstatus': 1})
	if project_budget:
		tot_qty = sum(pb.qty for pb in project_budget.item_table)
		del_qty = sum(pb.delivered_qty for pb in project_budget.item_table)

		per_delivered = (del_qty / tot_qty) * 100 if tot_qty else 0
		print(per_delivered)
		frappe.db.set_value("Sales Order",doc.sales_order,'per_delivered',per_delivered)

def test_check():
	doc = frappe.get_doc("Sales Invoice", "ENG-CRD-2025-00275")
	update_stock_on_si_new(doc,None)
@frappe.whitelist()
def update_stock_on_si_new(doc,method):

	se_items = []
	project=frappe.get_value("Sales Order",{'name':doc.so_no},['project'])
	if doc.order_type =="Project":
		if doc.is_return==0:
			project_budget = frappe.db.get_value("Sales Order",doc.so_no,'project_budget')
			pb = frappe.get_doc("Project Budget",project_budget)
			se = frappe.new_doc("Stock Entry")
			se.stock_entry_type = "Material Issue" 
			se.set_posting_time = 1
			se.posting_date=doc.posting_date
			se.posting_time=doc.posting_time
			se.company = doc.company
			se.reference_number = doc.name
			se.custom_reference_type = "Sales Invoice"
			posting_time = frappe.utils.format_time(doc.posting_time)
			set_warehouse = frappe.db.get_value("Warehouse",{'warehouse_name': ['like', '%Work In Progress%'],'company':doc.company},'name')
			for i in pb.item_table:
				if i.delivered_qty:
					to_bill_qty = i.delivered_qty - i.billed_qty
					if to_bill_qty > 0:
						tot_bill_qty = 0
						if frappe.db.exists("Item", {'name': i.item, 'is_stock_item': 1}):
							dn_wip_qty = frappe.db.sql(
								"""SELECT SUM(dni.qty) AS billable_qty
								FROM `tabDelivery Note Item` dni
								INNER JOIN `tabDelivery Note WIP` dnw ON dni.parent = dnw.name
								WHERE dni.item_code = %s AND dni.custom_against_pbsow = %s AND dnw.sales_order = %s AND dnw.docstatus = 1 AND TIMESTAMP(dnw.posting_date, dnw.posting_time) <= TIMESTAMP(%s, %s)""",
								(i.item,i.docname, doc.so_no,doc.posting_date,posting_time),
								as_dict=True
							)
							dn_qty = frappe.db.sql(
								"""SELECT SUM(dni.qty) AS billable_qty
								FROM `tabDelivery Note Item` dni
								INNER JOIN `tabDelivery Note` dnw ON dni.parent = dnw.name
								WHERE dni.item_code = %s AND dnw.sales_order = %s AND dnw.docstatus = 1 AND TIMESTAMP(dnw.posting_date, dnw.posting_time) <= TIMESTAMP(%s, %s)""",
								(i.item, doc.so_no,doc.posting_date,posting_time),
								as_dict=True
							)
							print([dn_wip_qty, dn_qty])
							billable_qty = dn_wip_qty[0].billable_qty if dn_wip_qty and dn_wip_qty[0].billable_qty else 0
							billable_qty_dn = dn_qty[0].billable_qty if dn_qty and dn_qty[0].billable_qty else 0
							tot_bill_qty += (billable_qty + billable_qty_dn)
							if (tot_bill_qty - i.billed_qty) > 0:
								se.append("items",{
									's_warehouse':set_warehouse,
									'item_code':i.item,
									'transfer_qty':(tot_bill_qty - i.billed_qty),
									'uom':i.unit,
									'stock_uom':frappe.get_value('Item',i.item,'stock_uom'),
									'conversion_factor': 1,
									'qty':(tot_bill_qty - i.billed_qty),
									'basic_rate':i.rate_with_overheads,
									'custom_against_pb':i.docname,
									'project':project
								})
						if (tot_bill_qty - i.billed_qty)>0:
							do = frappe.get_doc(i.pb_doctype,i.docname)
							do.billed_qty += (tot_bill_qty - i.billed_qty)
							do.save(ignore_permissions=True)
							i.billed_qty += (tot_bill_qty - i.billed_qty)
							
						else:
							do = frappe.get_doc(i.pb_doctype,i.docname)
							do.billed_qty += to_bill_qty
							do.save(ignore_permissions=True)
							i.billed_qty += to_bill_qty
								
			if len(se.items) > 0:
				se.save(ignore_permissions=True)
				se.submit()
			pb.flags.ignore_validate_update_after_submit = True
			pb.save(ignore_permissions=True)
		else:
			project_budget = frappe.db.get_value("Sales Order",doc.so_no,'project_budget')
			so = frappe.get_doc("Project Budget",project_budget)
			se = frappe.new_doc("Stock Entry")
			se.stock_entry_type = "Material Receipt" 
			se.company = doc.company
			se.reference_number = doc.name
			se.custom_reference_type = "Sales Invoice"
			set_warehouse = frappe.db.get_value("Warehouse",{'warehouse_name': ['like', '%Work In Progress%'],'company':doc.company},'name')
			for i in doc.custom_project_materials:
				if i.quantity:
					# project_name = frappe.get_value("Project",{'name':i.project},['project_name'])
					# pro_name = i.project +' - '+project_name
					# set_warehouse = frappe.db.get_value("Warehouse",{'warehouse_name': ['like', i.project]},'name')
					se.append("items",{
						't_warehouse':set_warehouse,
						'item_code':i.item_code,
						'qty':-(i.quantity),
						'basic_rate':0,
						'project':i.project
					})
					for j in so.item_table:
						if i.item_code==j.item:
							j.billed_qty += i.quantity
							do = frappe.get_doc(i.pb_doctype,i.docname)
							do.billed_qty += i.quantity
							do.save(ignore_permissions=True)
			so.flags.ignore_validate_update_after_submit = True
			so.save(ignore_permissions=True)
			se.save(ignore_permissions=True)
			se.submit()

@frappe.whitelist()
def check_dnwip():
	doc=frappe.get_doc("Sales Invoice",'ENG-CRD-2025-00022')
	se_items = []
	project=frappe.get_value("Sales Order",{'name':doc.so_no},['project'])
	if doc.order_type =="Project":
		if doc.is_return==0:
			project_budget = frappe.db.get_value("Sales Order",doc.so_no,'project_budget')
			pb = frappe.get_doc("Project Budget",project_budget)
			se = frappe.new_doc("Stock Entry")
			se.stock_entry_type = "Material Issue" 
			se.set_posting_time = 1
			se.posting_date=doc.posting_date
			se.posting_time=doc.posting_time
			se.company = doc.company
			se.reference_number = doc.name
			se.custom_reference_type = "Sales Invoice"
			set_warehouse = frappe.db.get_value("Warehouse",{'warehouse_name': ['like', '%Work In Progress%'],'company':doc.company},'name')
			for i in pb.item_table:
				if i.delivered_qty:
					to_bill_qty = i.delivered_qty - i.billed_qty
					if to_bill_qty > 0:
						tot_bill_qty = 0
						if frappe.db.exists("Item", {'name': i.item, 'is_stock_item': 1}):
							dn_wip_qty = frappe.db.sql(
								"""SELECT SUM(dni.qty) AS billable_qty
								FROM `tabDelivery Note Item` dni
								INNER JOIN `tabDelivery Note WIP` dnw ON dni.parent = dnw.name
								WHERE dni.item_code = %s AND dni.custom_against_pbsow = %s AND dnw.sales_order = %s AND dnw.posting_date < %s AND dnw.docstatus = 1""",
								(i.item,i.docname, doc.so_no,doc.posting_date),
								as_dict=True
							)
							dn_qty = frappe.db.sql(
								"""SELECT SUM(dni.qty) AS billable_qty
								FROM `tabDelivery Note Item` dni
								INNER JOIN `tabDelivery Note` dnw ON dni.parent = dnw.name
								WHERE dni.item_code = %s AND dnw.sales_order = %s AND dnw.posting_date < %s AND dnw.docstatus = 1""",
								(i.item, doc.so_no,doc.posting_date),
								as_dict=True
							)

							billable_qty = dn_wip_qty[0].billable_qty if dn_wip_qty and dn_wip_qty[0].billable_qty else 0
							billable_qty_dn = dn_qty[0].billable_qty if dn_qty and dn_qty[0].billable_qty else 0
							tot_bill_qty += (billable_qty + billable_qty_dn)
							print(i.billed_qty)
							print(i.item)
							print(tot_bill_qty)
	

@frappe.whitelist()
def return_total_amt1(from_date, to_date, account):

	base_account = account.rsplit(" - ", 1)[0]

	data = """
	<table border="1" width="100%">
	<tr style="background-color:#e35310;color:white;font-weight:bold;text-align:center">
		<td><b>Company</b></td>
		<td align="center"><b>Opening</b></td>
		<td align="center"><b>Debit</b></td>
		<td align="center"><b>Credit</b></td>
		<td align="center"><b>Closing</b></td>
	</tr>
	"""

	total_open = total_debit = total_credit = total_close = 0

	companies = frappe.get_all(
		"Company",
		filters={"is_group": 0},
		fields=["name", "abbr"]
	)

	for c in companies:

		account_name = f"{base_account} - {c.abbr}"

		# Opening Balance
		opening = frappe.db.sql("""
			SELECT
				IFNULL(SUM(debit_in_account_currency),0) AS debit,
				IFNULL(SUM(credit_in_account_currency),0) AS credit
			FROM `tabGL Entry`
			WHERE
				company=%s
				AND account=%s
				AND is_cancelled=0
				AND (
					posting_date < %s
					OR (ifnull(is_opening,'No')='Yes' AND posting_date > %s)
				)
		""", (c.name, account_name, from_date, to_date), as_dict=1)[0]

		open_balance = opening.debit - opening.credit

		# Period Movement
		move = frappe.db.sql("""
			SELECT
				IFNULL(SUM(debit_in_account_currency),0) AS debit,
				IFNULL(SUM(credit_in_account_currency),0) AS credit
			FROM `tabGL Entry`
			WHERE
				company=%s
				AND account=%s
				AND posting_date BETWEEN %s AND %s
				AND is_opening='No'
				AND is_cancelled=0
		""", (c.name, account_name, from_date, to_date), as_dict=1)[0]

		close_balance = open_balance + move.debit - move.credit

		data += f"""
		<tr>
			<td>{c.name}</td>
			<td align="right">{fmt_money(open_balance,2)}</td>
			<td align="right">{fmt_money(move.debit,2)}</td>
			<td align="right">{fmt_money(move.credit,2)}</td>
			<td align="right">{fmt_money(close_balance,2)}</td>
		</tr>
		"""

		total_open += open_balance
		total_debit += move.debit
		total_credit += move.credit
		total_close += close_balance

	data += f"""
	<tr style="font-weight:bold">
		<td align="center">Total</td>
		<td align="right">{fmt_money(total_open,2)}</td>
		<td align="right">{fmt_money(total_debit,2)}</td>
		<td align="right">{fmt_money(total_credit,2)}</td>
		<td align="right">{fmt_money(total_close,2)}</td>
	</tr>
	</table>
	"""

	return data


@frappe.whitelist()
def validation_on_submission(doc, methods):
	if doc.order_type == "Project":
		current_user = frappe.session.user
		user_roles = frappe.get_roles(current_user)
		if "Accounts User" not in user_roles:
			frappe.throw(
				"Only Accounts User can submit the invoice",
				title="Permission Error",
			)

@frappe.whitelist()
def validation_on_cancellation(doc, methods):
	if doc.order_type == "Project" and getdate(doc.posting_date) < getdate("2025-01-01"):
		current_user = frappe.session.user
		user_roles = frappe.get_roles(current_user)
		if "Accounts User" not in user_roles:
			frappe.throw(
				"Only Accounts User can cancel the previous year invoices",
				title="Permission Error",
			)
		
@frappe.whitelist()
def cancel_dn_on_si_cancel_manuel():
	dn_return_doc=frappe.get_doc('Delivery Note',"ENG-DNR-2025-00005")
	# dn_return_doc.cancel()
	dn = frappe.get_doc("Delivery Note",dn_return_doc.return_against)
	dn.per_returned =0
	dn.flags.ignore_validate_update_after_submit = True
	dn.save(ignore_permissions = True)
	print(((dn.per_returned-(dn_return_doc.total_qty) / dn.total_qty)*100))
	
	# for dnit in dn_return_doc.items:
	# 	dn_items = frappe.get_doc("Delivery Note Item",dnit.dn_detail)
	# 	frappe.db.set_value("Delivery Note Item",dnit.dn_detail,'returned_qty',0)
 
@frappe.whitelist()
def get_pb_materials(doc):
	project_budget_name = frappe.db.get_value("Project Budget", {"Sales_order": doc.so_no}, "name")
	pb = frappe.get_doc("Project Budget", project_budget_name)
	idx = 1
	tot_rate = 0
	tot_amount = 0
	data = ''' '''
	if pb.materials:
		data = '''
		<p style="font-weight: bold;">Materials</p>
		<table width = '100%' border = 1 style="font-size: 10px;">
		<tr>
			<td width='6%' style ='background-color:#e35310;text-align:center;color:white'>S.No.</td>
			<td width='18%' style ='background-color:#e35310;text-align:center;color:white'>Item Code</td>
			<td width='42%' style ='background-color:#e35310;text-align:center;color:white'>Item Name</td>
			<td width='10%' style ='background-color:#e35310;text-align:center;color:white'>Qty</td>
			<td width='12%' style ='background-color:#e35310;text-align:center;color:white'>Rate</td>
			<td width='12%' style ='background-color:#e35310;text-align:center;color:white'>Amount</td>
		</tr>
	  '''
		for row in pb.materials:
			data += f'''
					<tr>
						<td style = 'text-align:right'> { idx }</td>
						<td>{ row.item  or ""} </td>
						<td>{ row.item_name  or ""} </td>
						<td style = 'text-align:right'> { fmt_money(round(row.qty, 2), 2)} </td>
						<td style = 'text-align:right'>{ fmt_money(round(row.rate_with_overheads, 2), 2) or "" }  </td>
						<td style = 'text-align:right'>{ fmt_money(round(row.amount_with_overheads, 2), 2) or "" } </td>
					 </tr>
					   '''
			idx += 1
			tot_rate += row.rate_with_overheads
			tot_amount += row.amount_with_overheads
	data += f'''<tr>
					<td colspan=5 style = 'text-align:right; font-weight: bold;'> Total</td>
					<td style = 'text-align:right; font-weight: bold;'>{ fmt_money(round(tot_amount, 2), 2) or "" } </td>
				</tr>
			'''
	data += "</table>"
	return data

@frappe.whitelist()
def materials_test():
	doc = frappe.get_doc("Sales Invoice", "ENG-CRD-2025-00073")
	project_budget_name = frappe.db.get_value("Project Budget", {"Sales_order": doc.so_no}, "name")
	pb = frappe.get_doc("Project Budget", project_budget_name)
	idx = 1
	tot_rate = 0
	tot_amount = 0
	data = ''' '''
	if pb.materials:
		data = '''
		<p style="font-weight: bold;">Materials</p>
		<table width = '100%' border = 1>
		<tr>
			<td width='6%' style ='background-color:#e35310;text-align:center;color:white'>S.No.</td>
			<td width='18%' style ='background-color:#e35310;text-align:center;color:white'>Item Code</td>
			<td width='37%' style ='background-color:#e35310;text-align:center;color:white'>Item Name</td>
			<td width='12%' style ='background-color:#e35310;text-align:center;color:white'>Qty</td>
			<td width='12%' style ='background-color:#e35310;text-align:center;color:white'>Rate</td>
			<td width='15%' style ='background-color:#e35310;text-align:center;color:white'>Amount</td>
		</tr>
	  '''
		for row in pb.materials:
			data += f'''
					<tr>
						<td style = 'text-align:right'> { idx }</td>
						<td>{ row.item  or ""} </td>
						<td>{ row.item_name  or ""} </td>
						<td style = 'text-align:right'> { fmt_money(round(row.qty, 2), 2)} </td>
						<td style = 'text-align:right'>{ fmt_money(round(row.rate_with_overheads, 2), 2) or "" }  </td>
						<td style = 'text-align:right'>{ fmt_money(round(row.amount_with_overheads, 2), 2) or "" } </td>
					 </tr>
					   '''
			idx += 1
			tot_rate += row.rate_with_overheads
			tot_amount += row.amount_with_overheads
	data += f'''<tr>
					<td colspan=3 style = 'text-align:right; font-weight: bold;'> Total</td>
					<td style = 'text-align:right; font-weight: bold;'>{ fmt_money(round(tot_rate, 2), 2) or "" }  </td>
					<td style = 'text-align:right; font-weight: bold;'>{ fmt_money(round(tot_amount, 2), 2) or "" } </td>
				</tr>
			'''
	data += "</table>"
	return data

@frappe.whitelist()
def validate_posting_date_for_pe(doc,method):
	if doc.posting_date > today():
		frappe.throw("Not Allowed to create Payment Entry for future date")

@frappe.whitelist()
def cancel_delete_timesheets():
	proj_time=frappe.get_doc("Projects Timesheet",'ENG-DPT-250409-001')
	i=0
	for tlog in proj_time.project_day_plan_employee:
		query = """
			SELECT DISTINCT ts.name
			FROM `tabTimesheet` ts
			JOIN `tabTimesheet Detail` tsd ON ts.name = tsd.parent
			WHERE '%s' BETWEEN ts.start_date AND ts.end_date
			AND ts.employee = '%s'
			AND tsd.project = '%s'
		""" % (proj_time.plan_date, tlog.employee, tlog.project)
		# query = """
		# 	select * from `tabTimesheet` ts where '%s' between ts.start_date and ts.end_date and employee = '%s' 
		# 				""" % (self.plan_date,tlog.employee)
		timesheet_id = frappe.db.sql(query,as_dict=True)
		
		if timesheet_id:
			i+=1
			timesheet = frappe.get_doc("Timesheet",timesheet_id)
			timesheet.cancel()
			# timesheet.delete()
	print(i)



@frappe.whitelist()
def return_total_amt2():
	from_date = '2025-02-01'
	to_date = '2025-02-28'
	account = '1102 - CASH FROM SALES - ASTCC'
	acct = account.split(' - ')
	acc=''
	if len(acct) == 2:
		acc = acct[0]
	if len(acct) == 3:
		acc = f"{acct[0]} - {acct[1]}"
	if len(acct) == 4:
		acc = f"{acct[1]} - {acct[2]}"
	ac = '%'+acc+'%'
	data = '<table  border= 1px solid black width = 100%>'
	data += '<tr style = "background-color:#D9E2ED"><td colspan =1><b></b></td><td colspan =1 style = "text-align:center"><b>Opening</b></td><td colspan =2 style = "text-align:center"><b>Movement</b></td><td colspan =1 style = "text-align:center"><b>Closing</b></td></tr>'

	data += '<tr style = "background-color:#e35310;color:white"><td  style = "text-align:center;font-weight:bold;color:white">Company</td><td  style = "text-align:center;font-weight:bold;color:white">Balance</td><td  style = "text-align:center;font-weight:bold;color:white">Debit</td><td  style = "text-align:center;font-weight:bold;color:white">Credit</td><td  style = "text-align:center;font-weight:bold;color:white">Balance</td></tr>'
	op_credit = 0
	op_debit = 0
	total_op_debit = 0
	total_op_credit = 0
	t_c_credit = 0
	t_p_credit = 0
	t_c_debit = 0
	t_p_debit = 0
	
	li = []
	company = frappe.db.sql(""" select name from `tabCompany` where is_group = 1""",as_dict=1)
	for com in company:
		li.append(com.name)
		comp = frappe.db.get_list("Company",{"parent_company":com.name},['name'])
		for j in comp:
			li.append(j.name)
	for c in li:
		gle = frappe.db.sql("""select account, sum(debit_in_account_currency) as opening_debit, sum(credit_in_account_currency) as opening_credit from `tabGL Entry` where company = '%s'	and (posting_date < '%s' or (ifnull(is_opening, 'No') = 'Yes' and posting_date > '%s')) and account like '%s' and is_cancelled = 0  """%(c,from_date,to_date,ac),as_dict=True)
		for g in gle:
			if not g.opening_debit:
				g.opening_debit = 0
			if not g.opening_credit:
				g.opening_credit = 0
			t_p_debit += g.opening_debit
			t_p_credit += g.opening_credit
			balance_op = t_p_debit - t_p_credit
			data += '<tr><td>%s</td><td style = text-align:right >%s</td>'%(c,fmt_money(g.opening_debit - g.opening_credit,2))
			sq = frappe.db.sql(""" select company,sum(debit_in_account_currency) as debit,sum(credit_in_account_currency) as credit from `tabGL Entry` where company = '%s' and account like '%s' and posting_date between '%s' and '%s' and is_opening = 'No' and is_cancelled = 0 """%(c,ac,from_date,to_date),as_dict=True)
			for i in sq:
				if not i.credit:
					i.credit = 0
				if not i.debit:
					i.debit = 0
				op_credit = g.opening_credit + i.credit
				op_debit = g.opening_debit + i.debit
				total_op_debit += i.debit
				total_op_credit += i.credit
				t_c_credit += op_credit
				t_c_debit += op_debit
				balance_cl = t_c_debit - t_c_credit
				balance_move=total_op_debit-total_op_credit
	# 			data += '<td style = text-align:right >%s</td><td style = text-align:right >%s</td><td style = text-align:right >%s</td></tr>'%(fmt_money(i.debit,2),fmt_money(i.credit,2),fmt_money(op_debit - op_credit,2))
	# data += '<tr style = "text-align:right;font-weight:bold"><td style = "text-align:center;font-weight:bold">Total</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>'%(fmt_money(balance_op,2),fmt_money(total_op_debit,2),fmt_money(total_op_credit,2),fmt_money(balance_cl,2))
	# # data += '<tr style = "text-align:right;font-weight:bold"><td style = "text-align:center;font-weight:bold">Balance</td><td colspan =3>%s</td><td colspan =3></td><td colspan=3>%s</td></tr>'%(fmt_money(balance_op),fmt_money(balance_cl))
	# data += '</table>'
				print(c)
				print(round(op_debit,2) - round(op_credit,2))

@frappe.whitelist()
def money_in_words(
	number: str | float | int,
	main_currency: str | None = None,
	fraction_currency: str | None = None,
):
	"""
	Returns string in words with currency and fraction currency.
	"""
	from frappe.utils import get_defaults

	_ = frappe._
	number = 3650.616
	try:
		# note: `flt` returns 0 for invalid input and we don't want that
		number = float(number)
	except ValueError:
		return ""

	number = flt(number)
	if number < 0:
		return ""

	d = get_defaults()
	if not main_currency:
		main_currency = d.get("currency", "INR")
	if not fraction_currency:
		fraction_currency = frappe.db.get_value("Currency", main_currency, "fraction", cache=True) or _(
			"Cent"
		)

	number_format = (
		frappe.db.get_value("Currency", main_currency, "number_format", cache=True)
		or frappe.db.get_default("number_format")
		or "#,###.##"
	)

	fraction_length = 3
	n = f"%.{fraction_length}f" % number

	numbers = n.split(".")
	main, fraction = numbers if len(numbers) > 1 else [n, "00"]

	if len(fraction) < fraction_length:
		zeros = "0" * (fraction_length - len(fraction))
		fraction += zeros

	in_million = True
	if number_format == "#,##,###.##":
		in_million = False

	# 0.00
	if main == "0" and fraction in ["00", "000"]:
		out = _(main_currency, context="Currency") + " " + _("Zero")
	# 0.XX
	elif main == "0":
		out = _(in_words(fraction, in_million).title()) + " " + fraction_currency
	else:
		out = _(main_currency, context="Currency") + " " + _(in_words(main, in_million).title())
		if cint(fraction):
			out = (
				out
				+ " "
				+ _("and")
				+ " "
				+ _(in_words(fraction, in_million).title())
				+ " "
				+ fraction_currency
			)

	return out + " " + _("only.")

number_format_info = {
	"#,###.##": (".", ",", 2),
	"#.###,##": (",", ".", 2),
	"# ###.##": (".", " ", 2),
	"# ###,##": (",", " ", 2),
	"#'###.##": (".", "'", 2),
	"#, ###.##": (".", ", ", 2),
	"#,##,###.##": (".", ",", 2),
	"#,###.###": (".", ",", 3),
	"#.###": ("", ".", 0),
	"#,###": ("", ",", 0),
	"#.########": (".", "", 8),
}

def get_number_format_info(format: str) -> tuple[str, str, int]:
	return number_format_info.get(format) or (".", ",", 2)

def in_words(integer: int, in_million=True) -> str:
	"""
	Returns string in words for the given integer.
	"""
	from num2words import num2words

	locale = "en_IN" if not in_million else frappe.local.lang
	integer = int(integer)
	try:
		ret = num2words(integer, lang=locale)
	except NotImplementedError:
		ret = num2words(integer, lang="en")
	except OverflowError:
		ret = num2words(integer, lang="en")
	return ret.replace("-", " ")

@frappe.whitelist()
def update_stock_on_si_new_manual():
	se_items = []
	doc = frappe.get_doc("Sales Invoice","INT-CRD-2024-00390")
	project=frappe.get_value("Sales Order",{'name':doc.so_no},['project'])
	if doc.order_type =="Project":
		if doc.is_return==0:
			project_budget = frappe.db.get_value("Sales Order",doc.so_no,'project_budget')
			pb = frappe.get_doc("Project Budget",project_budget)
			se = frappe.new_doc("Stock Entry")
			se.stock_entry_type = "Material Issue" 
			se.set_posting_time = 1
			se.posting_date=doc.posting_date
			se.posting_time=doc.posting_time
			se.company = doc.company
			se.reference_number = doc.name
			se.custom_reference_type = "Sales Invoice"
			# posting_time = doc.posting_time[:8]
			set_warehouse = frappe.db.get_value("Warehouse",{'warehouse_name': ['like', '%Work In Progress%'],'company':doc.company},'name')
			for i in pb.item_table:
				if i.delivered_qty:
					to_bill_qty = i.delivered_qty - i.billed_qty
					if to_bill_qty > 0:
						tot_bill_qty = 0
						if frappe.db.exists("Item", {'name': i.item, 'is_stock_item': 1}):
							dn_wip_qty = frappe.db.sql(
								"""SELECT SUM(dni.qty) AS billable_qty
								FROM `tabDelivery Note Item` dni
								INNER JOIN `tabDelivery Note WIP` dnw ON dni.parent = dnw.name
								WHERE dni.item_code = %s AND dni.custom_against_pbsow = %s AND dnw.sales_order = %s AND dnw.posting_date < %s AND dnw.docstatus = 1""",
								(i.item,i.docname, doc.so_no,doc.posting_date),
								as_dict=True
							)
							dn_qty = frappe.db.sql(
								"""SELECT SUM(dni.qty) AS billable_qty
								FROM `tabDelivery Note Item` dni
								INNER JOIN `tabDelivery Note` dnw ON dni.parent = dnw.name
								WHERE dni.item_code = %s AND dnw.sales_order = %s AND dnw.posting_date < %s AND dnw.docstatus = 1""",
								(i.item, doc.so_no,doc.posting_date),
								as_dict=True
							)

							billable_qty = dn_wip_qty[0].billable_qty if dn_wip_qty and dn_wip_qty[0].billable_qty else 0
							billable_qty_dn = dn_qty[0].billable_qty if dn_qty and dn_qty[0].billable_qty else 0
							tot_bill_qty += (billable_qty + billable_qty_dn)
							if (tot_bill_qty - i.billed_qty) > 0:
								se.append("items",{
									's_warehouse':set_warehouse,
									'item_code':i.item,
									'transfer_qty':(tot_bill_qty - i.billed_qty),
									'uom':i.unit,
									'stock_uom':frappe.get_value('Item',i.item,'stock_uom'),
									'conversion_factor': 1,
									'qty':(tot_bill_qty - i.billed_qty),
									'basic_rate':i.rate_with_overheads,
									'custom_against_pb':i.docname,
									'project':project
								})
						if float(tot_bill_qty - i.billed_qty)>0:
							do = frappe.get_doc(i.pb_doctype,i.docname)
							print(do.billed_qty)
							do.billed_qty += (tot_bill_qty - i.billed_qty)
							print([i.item,i.pb_doctype,i.docname,do.billed_qty])
							
							i.billed_qty += (tot_bill_qty - i.billed_qty)
							print([i.item,i.billed_qty])
							
							
							# do.save(ignore_permissions=True)
						else:
							do = frappe.get_doc(i.pb_doctype,i.docname)
							do.billed_qty += to_bill_qty
							print(tot_bill_qty - i.billed_qty)
							i.billed_qty += to_bill_qty
							
							# do.save(ignore_permissions=True)
			if len(se.items) > 0:
				se.save(ignore_permissions=True)
				se.submit()
			pb.flags.ignore_validate_update_after_submit = True
			pb.save(ignore_permissions=True)

@frappe.whitelist()
def delete_warehouse_from_so(doc, method):
	if frappe.db.exists("Warehouse", {"warehouse_name": doc.project, "company": doc.company}):
		warehouse = frappe.get_doc("Warehouse", {"warehouse_name": doc.project, "company": doc.company}, ignore_permissions=True)
		warehouse.delete()

@frappe.whitelist()
def add_description_in_appraisal(doc, method):
	for row in doc.goals:
		description = frappe.db.get_value("KRA", row.kra, "description")
		row.custom_description = description




@frappe.whitelist()
def update_dnwip_field(name):
	data = ''
	doc = frappe.get_doc("Delivery Note WIP", name)
	data += '<table class="table table-bordered" border=1><tr><td style="background-color:#fe3f0c;text-align:center, " colspan=2><p style="color:white; text-align:center;">Link Documents</p></td></tr>'
	if doc.project:
		link = f'https://erp.electraqatar.com/app/project/{doc.project}'
		data += '<tr style=text-align:center><td><a href="{}" target="_blank">{}</a></td><td>Project</td></tr>'.format(link, doc.project)
	if doc.sales_order:
		link = f'https://erp.electraqatar.com/app/sales-order/{doc.sales_order}'
		data += '<tr style=text-align:center><td ><a href="{}" target="_blank">{}</a></td><td>Sales Order</td></tr>'.format(link, doc.sales_order)
	
	data += '</table>'
	if doc.project or doc.sales_order:
		return data

@frappe.whitelist()
def sales_order_item_name():
	print(frappe.db.get_value("Sales Order Item",{'parent':'INT-SO-2024-00254','item_code':'REMOVAL AND RE-INSTALLATION WORKS'},['name']))

# from frappe.model.rename_doc import rename_doc
# @frappe.whitelist()
# def update_sales_order_item_table():
	# child = frappe.new_doc("Sales Order Item")
	# child.update({
	# "actual_qty": 2343,
	# "amount": 150,
	# "base_amount": 150,
	# "base_net_amount": 147.48,
	# "base_net_rate": 4.92,
	# "base_price_list_rate": 6.5,
	# "base_rate": 5,
	# "brand": "Altair",
	# "conversion_factor": 1,
	# "creation": "2025-07-20 11:56:11.374849",
	# "custom_row_id": 26,
	# "custom_so_qty": 30,
	# "delivered_qty": 30,
	# "delivery_date": "2025-07-20",
	# "delivery_terms": "Ex-Stock",
	# "description": "Metal Clad Box 3X6 (141X80X33X 0.9MM Thickness)- Altair",
	# "discount_amount": 1.5,
	# "discount_percentage": 23.077,
	# "docstatus": 1,
	# "doctype": "Sales Order Item",
	# "grant_commission": 1,
	# "gross_profit": 38.4,
	# "idx": 26,
	# "item_code": "AMCB36",
	# "item_group": "Altair-GI",
	# "item_name": "Metal Clad Box 3X6 (141X80X33X 0.9MM Thickness)- Altair",
	# "modified": "2025-07-27 08:25:48.026677",
	# "modified_by": "sales.electrical@electraqatar.com",
	# "name": "1db6a4a6e1",
	# "net_amount": 147.48,
	# "net_rate": 4.92,
	# "owner": "sales.electrical@electraqatar.com",
	# "parent": "ELE-SO-2025-00280",
	# "parentfield": "items",
	# "parenttype": "Sales Order",
	# "prevdoc_docname": "ELE-QTN-2025-00683-1",
	# "projected_qty": 992,
	# "qty": 30,
	# "quotation_item": "cf1538f4a2",
	# "rate": 5,
	# "reserve_stock": 1,
	# "stock_qty": 30,
	# "stock_uom": "Nos",
	# "stock_uom_rate": 5,
	# "transaction_date": "2025-07-20",
	# "uom": "Nos",
	# "valuation_rate": 3.72,
	# "warehouse": "Electra Electrical Warehouse - EDE"
	# 						})
	# child.insert()
	# frappe.db.commit()
	# child = frappe.get_doc("Sales Order Item",{"item_code":'ATMFPVC-25','custom_row_id':20,'parent':'ELE-SO-2025-00280'})
	# child.idx = 20
	# child.save()

	# frappe.db.commit()
	# print(frappe.db.get_value("Sales Order Item",{"item_code":'ATRC25-4','parent':'ELE-SO-2025-00280'},['name']))
	# doc_name = frappe.db.get_value("Sales Order Item",{"item_code": "ATS163-25", "parent": "ELE-SO-2025-00280"},"name")
	# if doc_name:
	# 	new_name = 'd3094f47ac'
		# print(frappe.db.get_value('Sales Order Item',{'name':doc_name},['parent']))
		# rename_doc("Sales Order Item", doc_name, new_name)
		# frappe.db.commit()

@frappe.whitelist()
def create_annual_leave_allocation():
	current_date = datetime.now().date()
	employee_grades = frappe.db.get_all('Employee Grade',{'custom_monthly_annual_leave_allocation': ['!=', 0],'custom_completed_months_of_service': ['!=', 0]},['name', 'custom_completed_months_of_service', 'custom_monthly_annual_leave_allocation', 'custom_new_leaves'])
	
	for employee_grade in employee_grades:
		employees = frappe.get_all('Employee',{'status': 'Active','custom_employee_grade': employee_grade.name},['name', 'company', 'date_of_joining'],order_by ='name')
		for emp in employees:
			if not emp.date_of_joining:
				continue
			eligibility_date = emp.date_of_joining
			if getdate(current_date) >= eligibility_date:
				# expire_date = add_years(eligibility_date, 100)
				if frappe.get_all("Leave Allocation",filters={"docstatus":1,"employee": emp.name, "leave_type": "Annual Leave"},fields=["to_date"],order_by="to_date desc",limit=1): 
					# to_date = frappe.get_all("Leave Allocation",filters={"docstatus":1,"employee": emp.name, "leave_type": "Annual Leave"},fields=["to_date"],order_by="to_date desc",limit=1)[0].to_date
					leave_allocation_name = frappe.db.get_all("Leave Allocation",filters={"docstatus":1,"employee": emp.name, "leave_type": "Annual Leave"},fields=["name"],order_by="to_date desc",limit=1)[0].name
				else:
					# to_date = None
					leave_allocation_name = None
				# if to_date and to_date < current_date:
				#     eligibility_date =add_days(to_date,1)
				# if to_date and to_date > current_date:
				#     if to_date.year == current_date.year and to_date.month >= current_date.month and to_date.day >= current_date.day:
				#         eligibility_date =add_days(to_date,1)
				#         continue
				# leave_allocation_name = frappe.db.get_value("Leave Allocation",{"docstatus":1,"employee": emp.name,"leave_type": "Annual Leave","from_date": [">=",eligibility_date],"to_date": ["<=",expire_date]},"name")
				# if not leave_allocation_name:
				#     if employee_grade.name =='NON STAFF':
				#         today = date.today()
				#         join_date = emp.date_of_joining
				#         months_diff = (today.year - join_date.year) * 12 + (today.month - join_date.month)
				#         if months_diff == 24:
				#             leaves = employee_grade.custom_new_leaves
				#         elif months_diff < 23:
				#             leaves = 2.5 * months_diff   
				#         else:
				#             months = today.month
				#             leaves = 2.5 * months  
				#     elif employee_grade.name =='STAFF':
				#         join_date = emp.date_of_joining
				#         today = date.today()
				#         months_diff = (today.year - join_date.year) * 12 + (today.month - join_date.month)
				#         if  months_diff == 11:
				#             leaves = employee_grade.custom_new_leaves
				#         elif  months_diff < 11:
				#             leaves = 2.73 * months_diff       
				#         else:
				#             months = today.month
				#             leaves = 2.73 * months 
					# allocation = frappe.new_doc("Leave Allocation")
					# allocation.employee = emp.name
					# allocation.leave_type = "Annual Leave"
					# allocation.new_leaves_allocated = leaves
					# allocation.total_leaves_allocated = leaves
					# allocation.from_date = eligibility_date
					# allocation.to_date = expire_date
					# allocation.company = emp.company
					# allocation.save(ignore_permissions=True)
					# allocation.submit()
					# frappe.db.commit()
				if leave_allocation_name:
					# print('Existing Allocation')
					leave_doc = frappe.get_doc("Leave Allocation", leave_allocation_name)
					leave_doc.new_leaves_allocated += employee_grade.custom_monthly_annual_leave_allocation
					leave_doc.total_leaves_allocated += employee_grade.custom_monthly_annual_leave_allocation
					leave_doc.save(ignore_permissions=True)
					frappe.db.commit()


import frappe
from frappe.utils import getdate

@frappe.whitelist()
def update_cost_rate():
	timesheets = frappe.get_all(
		"Timesheet",
		filters={
			'status': "Submitted",
			'start_date': ['between', ['2024-01-01', '2025-12-31']]
		},
		fields=["name", "employee", "start_date"]
	)

	for t in timesheets:
		emp = frappe.db.get_value("Employee", t.employee, ["basic", "holiday_list"], as_dict=True)
		basic_salary = emp.basic or 0
		holiday_list = emp.holiday_list
		cost_rate = basic_salary / 30 / 8  

		if holiday_list:
			start_date = getdate(t.start_date)
			is_holiday = frappe.db.exists(
				"Holiday",
				{"parent": holiday_list, "holiday_date": start_date}
			)
			if is_holiday:
				cost_rate = (basic_salary / 30 / 8) * 1.5
		print(start_date)

		child_rows = frappe.get_all(
			"Timesheet Detail",
			filters={"parent": t.name},
			fields=["name", "billing_hours"]
		)
		cost_rate = round(cost_rate, 2)

		for row in child_rows:
			cost_amount = round(cost_rate * row.billing_hours, 2)
			frappe.db.set_value("Timesheet Detail", row.name, "costing_rate", cost_rate)
			frappe.db.set_value("Timesheet Detail", row.name, "costing_amount", cost_amount)




@frappe.whitelist()
def update_stock_on_si_new_create():
	project=frappe.get_value("Sales Order",{'name':"INT-SO-2024-00199"},['project'])
	project_budget = frappe.db.get_value("Sales Order","INT-SO-2024-00199",'project_budget')
	pb = frappe.get_doc("Project Budget",project_budget)
	# se = frappe.new_doc("Stock Entry")
	# se.stock_entry_type = "Material Issue" 
	# se.set_posting_time = 1
	# se.posting_date=doc.posting_date
	# se.posting_time=doc.posting_time
	# se.company = doc.company
	# se.reference_number = doc.name
	# se.custom_reference_type = "Sales Invoice"
	# set_warehouse = frappe.db.get_value("Warehouse",{'warehouse_name': ['like', '%Work In Progress%'],'company':doc.company},'name')
	for i in pb.item_table:
		if i.delivered_qty:
			issue_qty = frappe.db.sql(
				"""SELECT SUM(dni.qty) AS billable_qty
				FROM `tabStock Entry Detail` dni
				INNER JOIN `tabStock Entry` dnw ON dni.parent = dnw.name
				WHERE dni.item_code = %s AND dnw.stock_entry_type = "Material Issue" AND dnw.docstatus = 1 AND dni.custom_against_pb = %s """,
				(i.item,i.docname),
				as_dict=True
			)
			i.billed_qty = issue_qty[0].billable_qty
			pb.flags.ignore_validate_update_after_submit = True
			pb.save(ignore_permissions=True)
			print(i.item)
			print(issue_qty[0].billable_qty)    
	# pb.flags.ignore_validate_update_after_submit = True
	# pb.save(ignore_permissions=True)


@frappe.whitelist()
def create_annual_leave_allocation_for_new_joinees(doc,method):
	if doc.date_of_joining and doc.name:
		if doc.custom_employee_grade and frappe.db.exists('Employee Grade',{'custom_monthly_annual_leave_allocation': ['!=', 0],'name': doc.custom_employee_grade}):   
			to_date = add_years(getdate(doc.date_of_joining), 100)
			allocation = frappe.new_doc("Leave Allocation")
			allocation.employee = doc.name
			allocation.leave_type = "Annual Leave"
			allocation.new_leaves_allocated = 0
			allocation.total_leaves_allocated = 0
			allocation.from_date = getdate(doc.date_of_joining)
			allocation.to_date = to_date
			allocation.company = doc.company
			allocation.save(ignore_permissions=True)
			allocation.submit()
			frappe.db.commit()


@frappe.whitelist()
def delete_employee_leave_allocation(doc, method):
	allocations = frappe.get_all('Leave Allocation',{'employee': doc.name, 'leave_type': 'Annual Leave'},['name'])
	for alloc in allocations:
		leave_doc = frappe.get_doc('Leave Allocation', alloc.name)
		if leave_doc.docstatus == 1:
			leave_doc.cancel()
			frappe.delete_doc('Leave Allocation', alloc.name, ignore_permissions=True)
		elif leave_doc.docstatus in [0, 2]:
			frappe.delete_doc('Leave Allocation', alloc.name, ignore_permissions=True)



@frappe.whitelist()
def set_uom():
	doc = frappe.get_doc("Sales Order","MEP-SO-2024-00175")
	for i in doc.items:
		uom = frappe.db.get_value("Item",i.item_code,["stock_uom"])
		frappe.db.set_value("Sales Order Item",i.name,"uom",uom)


@frappe.whitelist()
def change_row_name(so,item):
	sales = frappe.get_doc("Sales Order",so)
	for i in sales.items:
		if i.item_code == item:
			return i.name

@frappe.whitelist()
def set_sales_order():
	frappe.db.set_value("Delivery Note WIP","INT-PDN-2025-00599","sales_order","INT-SO-2025-00135")

# @frappe.whitelist()
# def get_company_data():
#     doc = frappe.get_doc("Company", "ELECTRA - UPS DIVISION")

#     from erpnext.accounts.doctype.account.chart_of_accounts.chart_of_accounts import create_charts
#     from frappe import DuplicateEntryError

#     frappe.local.flags.ignore_root_company_validation = True
#     try:
		
#         create_charts(
#             doc.name,
#             doc.chart_of_accounts,
#             doc.existing_company
#         )
#     except DuplicateEntryError as e:
#         frappe.log_error(
#             message=str(e),
#             title="Duplicate Account Skipped during create_charts"
#         )

#     if not doc.default_receivable_account:
#         doc.db_set(
#             "default_receivable_account",
#             frappe.db.get_value(
#                 "Account",
#                 {
#                     "company": doc.name,
#                     "account_type": "Receivable",
#                     "is_group": 0
#                 }
#             ),
#         )

#     if not doc.default_payable_account:
#         doc.db_set(
#             "default_payable_account",
#             frappe.db.get_value(
#                 "Account",
#                 {
#                     "company": doc.name,
#                     "account_type": "Payable",
#                     "is_group": 0
#                 }
#             ),
#         )

#     return "Ok"

