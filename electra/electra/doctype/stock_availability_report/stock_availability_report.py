# Copyright (c) 2022, Abdulla and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file
from frappe.utils.file_manager import get_file

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
from frappe.utils import (
    flt,
    cint,
    cstr,
    get_html_format,
    get_url_to_form,
    gzip_decompress,
    format_duration,
)

class StockAvailabilityReport(Document):
      @frappe.whitelist()
      def get_data(self):
            item = frappe.get_value('Item',{'item_code':self.item_code},'item_code')
            rate = frappe.get_value('Item',{'item_code':self.item_code},'valuation_rate')
            group = frappe.get_value('Item',{'item_code':self.item_code},'item_group')
            des = frappe.get_value('Item',{'item_code':self.item_code},'description')
            price = frappe.get_value('Item Price',{'item_code':self.item_code},'price_list_rate')
            spn = frappe.db.sql(""" select supplier_part_no from `tabItem` left join `tabItem Supplier` on `tabItem`.name = `tabItem Supplier`.parent 
            where `tabItem`.item_code = '%s' """%(self.item_code),as_dict=True)[0]
            csp = 'Current Sales Price'
            cpp = 'Current Purchase Price'
            cos = 'COS'
            pso = 'Pending Sales order'
            ppo = 'Pending Purchase order'
            cspp_rate = 0
            cppp_rate = 0
            psoc = 0
            ppoc = 0
            cspp_query = frappe.db.sql("""select `tabSales Order Item`.rate as rate, `tabSales Order`.creation from `tabSales Order`
            left join `tabSales Order Item` on `tabSales Order`.name = `tabSales Order Item`.parent
            where `tabSales Order Item`.item_code = '%s' order by `tabSales Order`.creation """ % (item), as_dict=True)
            # frappe.errprint(cspp_query)
            if cspp_query:
                  cspp = cspp_query[-1]
                  cspp_rate = cspp['rate']
            cppp_query = frappe.db.sql("""select `tabPurchase Order Item`.rate as rate, `tabPurchase Order`.creation  from `tabPurchase Order`
            left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
            where `tabPurchase Order Item`.item_code = '%s' order by `tabPurchase Order`.creation """ % (item), as_dict=True)
            # frappe.errprint(cppp_query)
            if cppp_query:
                  cppp = cppp_query[-1]
                  cppp_rate = cppp['rate']
            psoc_query = frappe.db.sql("""select sum(`tabSales Order Item`.qty) as qty from `tabSales Order`
            left join `tabSales Order Item` on `tabSales Order`.name = `tabSales Order Item`.parent
            where `tabSales Order Item`.item_code = '%s' and `tabSales Order`.docstatus = 0 """ % (item), as_dict=True)[0]
            if psoc_query:
                  psoc = psoc_query['qty']
            ppoc_query = frappe.db.sql("""select sum(`tabPurchase Order Item`.qty) as qty from `tabPurchase Order`
            left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
            where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus = 0 """ % (item), as_dict=True)[0]
            if ppoc_query:
                  ppoc = ppoc_query['qty']
            data = ''
            stocks_query = frappe.db.sql("""select actual_qty,warehouse,stock_uom,stock_value from tabBin
                  where item_code = '%s' """%(item),as_dict=True)
            if stocks_query:
                  stocks = stocks_query
            data += '<table class="table table-bordered"><tr><th style="padding:1px;border: 1px solid black;background-color:#fe3f0c;" colspan=6><center>Stock Item Availability</center></th></tr>'
            data += '<tr><td style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;"><center><b>Item Code</b></center></td><td style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Item Name</b></center></td><td style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Item Group</b></center></td><td style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Item Price</b></center></td><td style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Supplier Part No</b></center></td><td style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Description</b></center></td></tr>'
            data += '<tr><td style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td><td style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td><td style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td><td style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td><td style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td><td style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td></tr>'%(item,frappe.db.get_value('Item',item,'item_name'),group,price,spn["supplier_part_no"],des)
            data += '<tr><td style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;"><center><b>%s</b></center></td><td style="padding:1px;border: 1px solid black;"><center><b>%s</b></center></td><td style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;"><center><b>%s</b></center></td><td style="padding:1px;border: 1px solid black;"><center><b>%s</b></center></td><td style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;"><center><b>%s</b></center></td><td style="padding:1px;border: 1px solid black;"><center><b>%s</b></center></td></tr>'%(csp,(cspp_rate),cpp,(cppp_rate),cos,rate or '')
            data += '<tr><td style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;"><center><b>%s</b></center></td><td style="padding:1px;border: 1px solid black;"><center><b>%s</b></center></td><td style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;"><center><b>%s</b></center></td><td colspan = 3 style="padding:1px;border: 1px solid black;"><center><b>%s</b></center></td></tr>'%(pso,psoc,ppo,ppoc)
            data += '<tr><td colspan = 3 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Warehouse</b></center></td><td colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>QTY</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>UOM</b></center></td></tr>'
            i = 0
            cou = 0
            tot = 'Total'
            uom = 'Nos'
            for stock in stocks:
                  if stock.actual_qty >= 0:
                        data += '<tr><td colspan = 3 style="padding:1px;border: 1px solid black">%s</td><td colspan = 2 style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td></tr>'%(stock.warehouse,int(stock.actual_qty) or '-',stock.stock_uom or '-')
                        i += 1
                        frappe.errprint(type(stock.actual_qty))
                        cou += stock.actual_qty  
            data += '<tr><td align="right" colspan = 3 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><b>%s</b></td><td colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>%s</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>%s</b></center></td></tr>'%(tot,int(cou),uom)
            data += '</table>'
            if i > 0:
                  return data
           
