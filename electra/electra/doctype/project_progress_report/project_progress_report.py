import calendar
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from io import BytesIO
import frappe
from frappe.utils import nowdate
from openpyxl.utils import get_column_letter
from frappe.utils import getdate
from frappe.utils import fmt_money
from frappe.model.document import Document
from frappe.utils import flt,time_diff_in_hours,to_timedelta
import datetime

class ProjectProgressReport(Document):
	pass
@frappe.whitelist()
def download(posting_date=None):
	# if not posting_date:
	posting_date = nowdate()  # Default to current date
	filename = f"Project Progress Report - {posting_date}"
	build_xlsx_response_clsr(filename)

@frappe.whitelist()
def build_xlsx_response_clsr(filename):
	xlsx_file = make_xlsx_clsr(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.errprint("Test2")
	frappe.response['type'] = 'binary'

def make_xlsx_clsr(sheet_name="Project Progress Report", posting_date=None, wb=None):
	if wb is None:
		wb = Workbook()
	args = frappe.local.form_dict
	ws = wb.active
	ws.title = sheet_name
	if args.project:
		project = args.project
		sales_order = frappe.db.get_value("Project", project, "sales_order")
		project_budget = frappe.db.get_value("Project", project, "budgeting")
	if args.sales_order:
		sales_order = args.sales_order
		project = frappe.db.get_value("Sales Order", sales_order, "project")
		project_budget = frappe.db.get_value("Sales Order", sales_order, "project_budget")

	
	blue_color = PatternFill(start_color="ebedef", end_color="ebedef", fill_type="solid")
	red_color = PatternFill(start_color="ff5050", end_color="ff5050", fill_type="solid")
	orange_color = PatternFill(start_color="e26b0a", end_color="e26b0a", fill_type="solid")
	teal_color = PatternFill(start_color="d8e4bc", end_color="d8e4bc", fill_type="solid")
	green_color = PatternFill(start_color="92d050", end_color="92d050", fill_type="solid")
	yellow_color = PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid")
	light_orange = PatternFill(start_color="ffeb9c", end_color="ffeb9c", fill_type="solid")
	sunday_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
	orange_color = PatternFill(start_color="ffc000", end_color="ffc000", fill_type="solid")
	sky_blue = PatternFill(start_color="92cddc", end_color="92cddc", fill_type="solid")
	light_blue = PatternFill(start_color="b7dee8", end_color="b7dee8", fill_type="solid")
	purple = PatternFill(start_color="ccc0da", end_color="ccc0da", fill_type="solid")
	light_green = PatternFill(start_color="57ff57", end_color="57ff57", fill_type="solid")
	dark_green = PatternFill(start_color="00b050", end_color="00b050", fill_type="solid")
	tomato = PatternFill(start_color="ff6d4b", end_color="ff6d4b", fill_type="solid")
	bright_blue = PatternFill(start_color="00b0f0", end_color="00b0f0", fill_type="solid")
	skin = PatternFill(start_color="fcd5b4", end_color="fcd5b4", fill_type="solid")
	actual_green = PatternFill(start_color="99ff99", end_color="99ff99", fill_type="solid")
	light_pink = PatternFill(start_color="fde9d9", end_color="fde9d9", fill_type="solid")
	week_off_pink = PatternFill(start_color="ffc7ce", end_color="ffc7ce", fill_type="solid")
	schedule_green = PatternFill(start_color="7fb957", end_color="7fb957", fill_type="solid")
	actual_header_green = PatternFill(start_color="99ff66", end_color="99ff66", fill_type="solid")
	continue_green = PatternFill(start_color="e26b0a", end_color="e26b0a", fill_type="solid")
	onleave_color = PatternFill(start_color="ffc000", end_color="ffc000", fill_type="solid")
	holiday_color = PatternFill(start_color="8db4e2", end_color="8db4e2", fill_type="solid")
 
	red_font = Font(color="FF0000")
	dark_red = Font(color="9c0000")
 
	font = Font(bold=True, color="000000")
	align_center_center = Alignment(vertical="center", horizontal="center", wrap_text=True)
	alignment = Alignment(horizontal="center")
	alignment_right = Alignment(horizontal="right")
	alignment_left = Alignment(horizontal="left")
	wrap_text = Alignment(wrap_text=True)
	title_font = Font(bold=True, size=14)
	thin_border = Border(
		left=Side(style="thin"),
		right=Side(style="thin"),
		top=Side(style="thin"),
		bottom=Side(style="thin")
	)
	
	ws.column_dimensions["A"].width = 5
	ws.column_dimensions["B"].width = 20
	ws.column_dimensions["C"].width = 40
	ws.column_dimensions["D"].width = 18
	ws.column_dimensions["E"].width = 15
	ws.column_dimensions["F"].width = 10
	ws.column_dimensions["G"].width = 10
	ws.column_dimensions["H"].width = 10
	ws.column_dimensions["I"].width = 10
 
	ws["A3"].value = f"PROJECT PROGRESS REPORT"
	ws["B5"].value = f"Project: {project}"
	ws["B6"].value = f"Sales Order: {sales_order}"
	ws["B7"].value = f"Project Budget: {project_budget}"
	
 
	ws["A3"].font = font
	ws["C3"].font = font
	ws["B5"].font = font
	ws["B6"].font = font
	ws["B7"].font = font
 
	ws["A3"].alignment = align_center_center
 
	ws.merge_cells("A3:I3")
	ws.merge_cells("B5:I5")
	ws.merge_cells("B6:I6")
	ws.merge_cells("B7:I7")
 
	ws.append([""])
	ws.append(["Delivery"])
	ws.cell(row=ws.max_row, column=1).font = font
 
	ws.append([""])
	rows = delivery_cost(project)
	for index, row in enumerate(rows):
		ws.append(row)
		last_row = ws.max_row
		for col in range(1, len(row) + 1):
			ws.cell(row=last_row, column=col).border = thin_border
		if index == 0 or index == len(rows) - 1:
			for col in range(1, len(row) + 1):
				ws.cell(row=last_row, column=col).font = font
					
	
	ws.append([""])
	ws.append(["Timesheet"])
	ws.cell(row=ws.max_row, column=1).font = font
  
	ws.append([""])
	rows = timesheet(project)
	for index, row in enumerate(rows):
		ws.append(row)
		last_row = ws.max_row
		for col in range(1, len(row) + 1):
			ws.cell(row=last_row, column=col).border = thin_border
		if index == 0 or index == len(rows) - 1:
			for col in range(1, len(row) + 1):
				ws.cell(row=last_row, column=col).font = font
	
	ws.append([""])
	ws.append(["Petty Cash"])
	ws.cell(row=ws.max_row, column=1).font = font
	
	ws.append([""])
	rows = petty_cash(project)
	for index, row in enumerate(rows):
		ws.append(row)
		last_row = ws.max_row
		for col in range(1, len(row) + 1):
			ws.cell(row=last_row, column=col).border = thin_border
		if index == 0 or index == len(rows) - 1:
			for col in range(1, len(row) + 1):
				ws.cell(row=last_row, column=col).font = font


	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	xlsx_file.seek(0)
	return xlsx_file

@frappe.whitelist()
def delivery_cost(project):
	data = [
		["SI.No", "Item Code", "Item Name", "UOM", "Qty", "Rate", "Amount"]
	]
	idx = 0
	tot_rate = 0
	tot_amount = 0
	tot_qty = 0
	sales_order = frappe.db.get_value("Sales Order", {"project": project}, "name")
	pb_name = frappe.db.get_value("Project", project, "budgeting")
	pb = frappe.get_doc("Project Budget", pb_name)
 
	# Delivery Note WIP
	dn_wip_list = frappe.get_all("Delivery Note WIP", {"sales_order": sales_order,"docstatus": 1})
	for child in pb.item_table:
		item_qty = 0
		rate = 0
		for dn_wip in dn_wip_list:
			del_wip = frappe.get_doc("Delivery Note WIP",dn_wip.name)
			stock_entry_exist = frappe.db.exists("Stock Entry", {"stock_entry_type": "Material Transfer", "reference_number": dn_wip.name, "docstatus": 1,})
			if stock_entry_exist:
				se = frappe.get_doc("Stock Entry", {"stock_entry_type": "Material Transfer", "reference_number": dn_wip.name, "docstatus": 1})
				for row in se.items:
					if row.item_code == child.item:
						rate = row.basic_rate
			for i in del_wip.items:
				if i.item_code == child.item and i.custom_against_pbsow == child.docname:
					amount = i.qty * rate
					idx += 1
					if del_wip.is_return == 1:
						tot_qty -= i.qty
						tot_rate -= rate
						tot_amount -= amount
						data.append([idx, i.item_code, i.item_name, i.uom, -(i.qty), fmt_money(-(round(rate, 2)), 2), fmt_money(-(round(amount, 2)), 2)])
					else:
						tot_qty += i.qty
						tot_rate += rate
						tot_amount += amount
						data.append([idx, i.item_code, i.item_name, i.uom, i.qty, fmt_money(round(rate, 2), 2), fmt_money(round(amount, 2), 2)])
	# Delivery Note
	dn_list = frappe.db.sql("""
		SELECT 
			c.item_code, c.item_name, c.qty, c.uom, c.rate, c.amount, p.is_return
		FROM 
			`tabDelivery Note` p
		INNER JOIN 
			`tabDelivery Note Item` c ON c.parent = p.name
		WHERE 
			p.project = %s AND p.docstatus = 1
	""", (project,), as_dict=True)
	if dn_list:
		for item in dn_list:
			idx += 1
			if item.is_return == 1:
				data.append([idx, item.item_code, item.item_name, item.uom, item.qty, fmt_money(-(round(item.rate, 2)), 2), fmt_money(round(item.amount, 2), 2)])
				tot_qty += item.qty
				tot_rate -= item.rate
				tot_amount += item.amount
			else:
				data.append([idx, item.item_code, item.item_name, item.uom, item.qty, fmt_money(round(item.rate, 2), 2), fmt_money(round(item.amount, 2), 2)])
				tot_qty += item.qty
				tot_rate += item.rate
				tot_amount += item.amount
	if dn_wip_list or dn_list:
		data.append(["", "", "Total", "", tot_qty, "", tot_amount])
	else:
		data.append(["", "-", "-", "-", "-", "-", "-"])
	return data
	
@frappe.whitelist()
def petty_cash(project):
	idx = 0
	data = [
		["SI.No", "Projet Name", "Voucher Name", "Voucher Type", "Expense For", "Amount"]
	]
	journal_entry = frappe.db.sql("""
		SELECT 
  			c.debit_in_account_currency, c.account, p.name, sum(c.debit_in_account_currency) as tot_amt
		FROM 
  			`tabJournal Entry` p
		INNER JOIN 
  			`tabJournal Entry Account` c ON c.parent = p.name 
		WHERE c.project = %s AND p.docstatus = 1 
		GROUP BY c.project
	""", (project,), as_dict=True)
	if journal_entry:
		for jl in journal_entry:
			idx += 1
			data.append([idx, project, jl.name, "Journal Entry", jl.account, fmt_money(round(jl.debit_in_account_currency, 2), 2)])
		data.append(["",  "", "Total", "", "", fmt_money(round(jl.tot_amt, 2), 2)])
	else:
		data.append(["", "-", "-", "-", "-", "-"])
	return data

@frappe.whitelist()
def timesheet(project):
	idx = 0
	data = [
		["SI.No", "Employee", "Employee Name", "Activity", "Activity Type", "Hours", "Qty", "Rate", "Total Cost"]
	]
	timesheet = frappe.db.sql("""
							SELECT 
								p.employee, 
								p.employee_name AS employee_name,
								c.activity_type,
								(c.task) as task,
								SUM(c.hours) AS total_hours,
								SUM(c.costing_amount) AS total_costing_amount,
								(c.costing_rate) AS total_costing_rate
							FROM `tabTimesheet` p
							INNER JOIN `tabTimesheet Detail` c 
								ON c.parent = p.name
							WHERE c.project = %s 
							AND p.docstatus = 1
							GROUP BY p.employee, c.activity_type, c.task
							order by p.employee, c.activity_type desc
						""", (project,), as_dict=1)


	if timesheet:
		tot_hrs = 0
		tot_qty = 0
		tot_rate = 0
		tot_cost = 0
		for t in timesheet:
			idx += 1
			tot_hrs += t.total_hours
			tot_qty += get_project_timesheet_qty(project, t.employee, t.task, t.activity_type)
			tot_rate += t.total_costing_rate
			tot_cost += t.total_costing_amount
			data.append([idx, t.employee, t.employee_name, t.task, t.activity_type, t.total_hours, get_project_timesheet_qty(project, t.employee, t.task, t.activity_type), fmt_money(round(t.total_costing_rate, 2), 2), fmt_money(round(t.total_costing_amount, 2), 2)])
		data.append(["", "", "Total", "", "", tot_hrs, tot_qty, "", fmt_money(round(tot_cost, 2), 2)])
	else:
		data.append(["", "-", "-", "-", "-", "-", "-", "-"])
	return data

@frappe.whitelist()
def get_task(task_id):
	subject = frappe.db.get_value("Task", task_id, "subject")
	return subject

@frappe.whitelist()
def get_project_timesheet_qty(project, employee, task, activity_type):
	if activity_type == "Regular Work":
		projects_timesheet = frappe.db.sql("""
											select sum(c.qty) as total_qty
											from `tabProjects Timesheet` p
											inner join `tabProject Day Plan Employee` c on c.parent = p.name
											where p.docstatus = 1 and c.employee = '%s' and c.activity = '%s' and c.project = '%s' 
											group by c.employee, c.activity
											""" %(employee, task, project), as_dict=1)
		if projects_timesheet:
			total_qty = projects_timesheet[0]['total_qty']
		else:
			total_qty = 0
	else: 
		total_qty = 0
	return total_qty
	
	