# Copyright (c) 2024, Abdulla and contributors
# For license information, please see license.txt

# import frappe
from frappe.model.document import Document


class SupplierStatementofAccount(Document):
	pass


import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from frappe import _
import frappe
from frappe.model.document import Document
from datetime import date, timedelta, datetime,time
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
	nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
import math
from frappe.utils import add_months, cint, flt, getdate, time_diff_in_hours,time_diff_in_seconds
import locale


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import openpyxl.styles as styles
from frappe.utils import flt, fmt_money


@frappe.whitelist()
def download():
    filename = 'Supplier Statement of Account'
    build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
		 
    ws = wb.create_sheet(sheet_name, 0)
    ws.column_dimensions['A'].width = 13
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20 
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15 
    ws.column_dimensions['G'].width = 15 

    company = frappe.db.get_value('Company', {'name': args.company_multiselect.split(',')[0]}, 'parent_company')
    if company:
        ws.append([company, " ", " ", " ", " ", " ", " ", "", "", "", ""])
        ws.append([args.company, " ", " ", " ", " ", " ", " ", "", "", "", ""])
    else:
        ws.append([args.company, " ", " ", " ", " ", " ", " ", "", "", "", ""]) 
        ws.append(["", " ", " ", " ", " ", " ", " ", "", "", "", ""])
    ws.append(["Supplier Statement of Account", " ", " ", " ", " ", " ", " "])

    from_date_str = args.from_date
    from_date_obj = datetime.strptime(from_date_str, "%Y-%m-%d")
    formatted_from_date = from_date_obj.strftime("%d-%m-%Y")

    to_date_str = args.to_date
    to_date_obj = datetime.strptime(to_date_str, "%Y-%m-%d")
    formatted_to_date = to_date_obj.strftime("%d-%m-%Y")

    header3 = ["For the Period: " + formatted_from_date + " to " + formatted_to_date]
    ws.append(header3 + [""] * 7)
    supplier = frappe.db.get_value('Supplier Statement of Account', {"name": args.name}, ['supplier'])
    header4 = ['Supplier: ' + supplier]
    ws.append(header4 + [""] * 7)
    ws.append([''])
    ws.append(['Date', 'Voucher Type', 'Voucher No', 'Remarks', 'Debit(QAR)', 'Credit(QAR)', 'Balance(QAR)'])

    data1 = get_data(args)
    for row in data1:
        ws.append(row)
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='bottom')
    bold_font = Font(bold=True)

    for header in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=7):
        for cell in header:
            cell.font = bold_font
            cell.alignment = align_center

    for header in ws.iter_rows(min_row=2, max_row=5, min_col=1, max_col=7):
        for cell in header:
            cell.font = bold_font
            cell.alignment = align_center

    for header in ws.iter_rows(min_row=8, max_row=8, min_col=1, max_col=7):
        for cell in header:
            cell.font = bold_font
            cell.alignment = align_right

    for header in ws.iter_rows(min_row=len(data1) + 6, max_row=len(data1) + 6, min_col=1, max_col=6):
        for cell in header:
            cell.font = bold_font
            cell.alignment = align_right

    for header in ws.iter_rows(min_row=len(data1) + 7, max_row=len(data1) + 7, min_col=1, max_col=7):
        for cell in header:
            cell.font = bold_font
            cell.alignment = align_right

    for header in ws.iter_rows(min_row=7, max_row=7, min_col=1, max_col=7):
        for cell in header:
            cell.font = bold_font
            cell.fill = PatternFill(fgColor='D3D3D3', fill_type="solid")

    for row in ws.iter_rows(min_row=9, max_row=len(data1) + 7):
        if 'Opening Balance' in row[0].value or 'Closing Balance' in row[0].value or 'Total' in row[0].value:
            for cell in row:
                cell.font = bold_font

    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'))

    header_range = ws['A1':ws.cell(row=len(data1) + 7, column=7).coordinate]
    for row in header_range:
        for cell in row:
            cell.border = border_thin

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=7)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=7)
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=7)
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=7)
    ws.merge_cells(start_row=8, start_column=1, end_row=8, end_column=6)
    ws.merge_cells(start_row=len(data1) + 6, start_column=1, end_row=len(data1) + 6, end_column=4)
    ws.merge_cells(start_row=len(data1) + 7, start_column=1, end_row=len(data1) + 7, end_column=6)

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def get_data(args):
    data = []
    supplier = frappe.db.get_value('Supplier Statement of Account', {"name": args.name}, ['supplier'])

    for company in args.company_multiselect.split(','):
        if supplier:
            gl_entry = frappe.db.sql("""
                SELECT voucher_type, voucher_no, posting_date, SUM(debit) AS debit, SUM(credit) AS credit
                FROM `tabGL Entry`
                WHERE company = %s AND posting_date BETWEEN %s AND %s AND is_cancelled = 0 AND party = %s
                GROUP BY voucher_no
                ORDER BY posting_date
                """, (company, args.from_date, args.to_date, supplier), as_dict=True)

            gle = frappe.db.sql("""
                SELECT SUM(debit) AS opening_debit, SUM(credit) AS opening_credit
                FROM `tabGL Entry`
                WHERE company = %s AND (posting_date < %s OR (IFNULL(is_opening, 'No') = 'Yes' AND posting_date >= %s)) AND party = %s AND is_cancelled = 0 and party_type='Supplier'                """, (company, args.from_date, args.to_date,supplier), as_dict=True)
        else:
            gl_entry = frappe.db.sql("""
                SELECT voucher_type, voucher_no, posting_date, SUM(debit) AS debit, SUM(credit) AS credit
                FROM `tabGL Entry`
                WHERE company = %s AND posting_date BETWEEN %s AND %s AND is_cancelled = 0
                GROUP BY voucher_no
                ORDER BY posting_date
                """, (company, args.from_date, args.to_date), as_dict=True)

            gle = frappe.db.sql("""
                SELECT SUM(debit) AS opening_debit, SUM(credit) AS opening_credit
                FROM `tabGL Entry`
                WHERE company = %s AND (posting_date < %s OR (IFNULL(is_opening, 'No') = 'Yes' AND posting_date >= %s)) AND is_cancelled = 0 and party_type='Supplier'
                """, (company, args.from_date, args.to_date), as_dict=True)

        opening_balance = 0
        t_p_debit = 0
        t_p_credit = 0
        for g in gle:
            if not g.opening_debit:
                g.opening_debit = 0
            if not g.opening_credit:
                g.opening_credit = 0
            t_p_debit += g.opening_debit
            t_p_credit += g.opening_credit
            opening_balance = t_p_credit -  t_p_debit 
            ob = float(fmt_money(round(opening_balance, 2)).replace(',', ''))

        data.append(['Company: ' + company])
        data.append(['Opening Balance', "", "", "", "", "", ob])

        balance = opening_balance
        for i in gl_entry:
            row = []
            balance += (i.credit - i.debit)
            if i.voucher_type == "Payment Entry":
                ref_no = frappe.db.get_value("Payment Entry", {"name": i.voucher_no}, ['reference_no'])
                check_no = ref_no if ref_no else ''
            else:
                check_no = ''

            if i.voucher_type == "Payment Entry":
                ref_no = frappe.db.get_value("Payment Entry",{"name":i.voucher_no},['reference_no'])
                if ref_no:
                    check_no = ref_no
                else:
                    check_no = ''
            else:
                check_no = ''
            if i.voucher_type == "Purchase Invoice":
                remarks = ''
                invoice = frappe.get_doc("Purchase Invoice",i.voucher_no)
                for k in invoice.items:
                    dn = frappe.db.get_value("Purchase Receipt",{"name":k.purchase_receipt},["supplier_delivery_note"])
                    po = k.purchase_order

                    if dn and po:
                        remarks = f"DN No.{dn} & PO No.{po}"
                    elif dn:
                        remarks = f"DN No.{dn}"
                    elif po:
                        remarks =f"PO No.{po}"
            elif i.voucher_type == "Journal Entry":
                remarks = ''
                remark = frappe.db.get_value("Journal Entry", {"name": i.voucher_no},['user_remark'])
                cheque_no = frappe.db.get_value("Journal Entry", {"name": i.voucher_no},['cheque_no'])
                if remark:
                    remarks = remark
                elif cheque_no:
                    check_no = cheque_no
            else:
                remarks = ''

            row.append(i.posting_date.strftime("%d-%m-%Y"))
            row.append(i.voucher_type)
            row.append(i.voucher_no)
            row.append(f"{remarks}{check_no}")
            row.append(float(fmt_money(round(i.debit, 2)).replace(',', '')) or "-")
            row.append(float(fmt_money(round(i.credit, 2)).replace(',', '')) or "-")
            row.append(float(fmt_money(round(balance, 2)).replace(',', '')) or "-")
            data.append(row)

        tp_credit = 0
        tp_debit = 0
        for i in gl_entry:
            tp_credit += i.credit
            t_p_credit += i.credit
            tp_debit += i.debit
            t_p_debit += i.debit

        data.append(["Total", "", "", "", float(fmt_money(round(tp_debit, 2)).replace(',', '')), float(fmt_money(round(tp_credit, 2)).replace(',', '')), ''])
        data.append(['Closing Balance', "", "", "", "", "", float(fmt_money(round(balance, 2)).replace(',', ''))])
        data.append([''])

    return data
