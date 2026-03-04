# Copyright (c) 2024, Abdulla and contributors
# For license information, please see license.txt

from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from frappe import _
import frappe
from frappe.model.document import Document
from datetime import date, timedelta, datetime,time
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
	nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
import math
from frappe.utils import add_months, cint, flt, getdate, time_diff_in_hours,time_diff_in_seconds
import locale


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import openpyxl.styles as styles
from frappe.utils import flt, fmt_money
import datetime

@frappe.whitelist()
def download():
	filename = 'Monthly MIS Report'
	build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
	current_date = datetime.date.today()
	current_year = current_date.strftime("%Y")
	
	previous_year = str(int(current_year) - 1)

	ws = wb.create_sheet(sheet_name, 0)
	ws.column_dimensions['A'].width = 20

	ws.append(['','Jan','','Feb','','Mar','','Apr','','May','','Jun','','Jul','','Aug','','Sep','','Oct','','Nov','','Dec','','Total'])
	ws.append(['Particulars',current_year + ' actual',current_year + ' budgeted',current_year + ' actual',current_year + ' budgeted',current_year + ' actual',current_year + ' budgeted',current_year + ' actual',current_year + ' budgeted',current_year + ' actual',current_year + ' budgeted',current_year + ' actual',current_year + ' budgeted',current_year + ' actual',current_year + ' budgeted',current_year + ' actual',current_year + ' budgeted',current_year + ' actual',current_year + ' budgeted',current_year + ' actual',current_year + ' budgeted',current_year + ' actual',current_year + ' budgeted',current_year + ' actual',current_year + ' budgeted',current_year + ' actual',current_year + ' budgeted'])
	data=get_data(args)
	for i in data:
		ws.append(i)
	data3=get_data_direct(args)
	for i in data3:
		ws.append(i)
	data2=get_data2(args)
	if data2:
		ws.append(['Other Direct Expenses'])
	
	for i in data2:
		ws.append(i)
	gross = ['Total']
	for col in ws.iter_cols(min_row=4, min_col=2):
		column_total = sum(cell.value for cell in col if cell.value is not None)
		gross.append(column_total)
	ws.append(gross)
	gross = ['Gross Profit']
	gp=[]
	start_row = len(get_data2(args)) + 5
	for cell, denom_cell in zip( ws[start_row][1:],ws[3][1:]):
		net_amt = (denom_cell.value - cell.value) 
		gp.append(net_amt)
	ws.append(gross + [p for p in gp])
	start_row = len(get_data2(args)) + 6
	per = ['Gross Profit %']
	gross_tot_per = []
	for cell, denom_cell in zip(ws[start_row][1:], ws[3][1:]):
		if denom_cell.value != 0:
			net_amt_per = (cell.value / denom_cell.value)
			gross_tot_per.append(net_amt_per)
		else:
			gross_tot_per.append(0)
	ws.append(per + [f"{p:.4f}%" for p in gross_tot_per])
	
	data=get_indirect_expenses(args)
	
	ws.append(['Other Expenses'])
	for i in data:
		ws.append(i)
	
	totals = ['Total Indirect Expenses']
	start_row = len(get_data2(args)) + 8
	for col in ws.iter_cols(min_row=start_row, min_col=2):
		column_total = sum(cell.value for cell in col if cell.value is not None)
		totals.append(column_total)
	ws.append(totals)
	totals_per = ['Total Indirect Expenses %'] 
	percentages = []
	start_row = len(get_indirect_expenses(args)) + 8 +len(get_data2(args))
	for cell, denom_cell in zip(ws[start_row][1:], ws[3][1:]):
		if cell.value and denom_cell.value is not None and denom_cell.value != 0:
			percentage = (cell.value / denom_cell.value) 
			percentages.append(percentage)
		else:
			percentages.append(0)
	ws.append(totals_per + [f"{p:.4f}%" for p in percentages])
	net_total =['Net Total']
	net_tot = []
	start_row = len(get_indirect_expenses(args)) + 9 + len(get_data2(args))
	h8=len(get_data2(args)) + 6
	for cell, denom_cell in zip(ws[start_row][1:], ws[h8][1:]):
		if cell.value and denom_cell.value is not None and denom_cell.value != 0:
			net_amt = (denom_cell.value - cell.value) 
			net_tot.append(net_amt)
		else:
			net_tot.append(0)
	ws.append(net_total + [p for p in net_tot])
	net_total_per =['Net Total %']
	net_tot_per = []
	start_row = len(get_indirect_expenses(args)) + 11 +len(get_data2(args))
	for cell, denom_cell in zip(ws[start_row][1:], ws[3][1:]):
		if cell.value and denom_cell.value is not None and denom_cell.value != 0:
			net_amt_per = (cell.value / denom_cell.value)
			net_tot_per.append(net_amt_per)
		else:
			net_tot_per.append(0)
	ws.append(net_total_per + [f"{p:.4f}%" for p in net_tot_per])



	align_center = Alignment(horizontal='center',vertical='center')
	for header in ws.iter_rows(min_row=1 , max_row=3, min_col=1, max_col=53):
		for cell in header:
			cell.font = Font(bold=True)
	for header in ws.iter_rows(min_row=5 , max_row=5, min_col=1, max_col=1):
		for cell in header:
			cell.font = Font(bold=True)
	# for header in ws.iter_rows(min_row=3 , max_row=len(get_indirect_expenses(args))+11, min_col=2, max_col=53):
	# 	for cell in header:
	# 		cell.alignment = align_center
	for header in ws.iter_rows(min_row=5 , max_row=7, min_col=1, max_col=53):
		for cell in header:
			cell.font = Font(bold=True)
	# for header in ws.iter_rows(min_row=len(get_indirect_expenses(args))+8 , max_row=ws.max_row, min_col=1, max_col=53):
	# 	for cell in header:
	# 		cell.font = Font(bold=True)

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'


def get_data(args):
	data = []
	
	current_date = datetime.date.today()
	current_year = current_date.strftime("%Y")
	data.append(['Sales'])
	company_sc=frappe.get_value("Company",{'name':args.company},['abbr'])
	# cost_center=frappe.get_value("Cost Center",{'cost_center_name':args.territory},['name'])
	sales_invoice = frappe.db.sql("""
		SELECT 
			MONTH(posting_date) AS month,
			SUM(base_grand_total) AS grand_total
		FROM 
			`tabSales Invoice`
		WHERE 
			YEAR(posting_date) = %s 
			AND company = %s
			AND docstatus = 1
		GROUP BY 
			MONTH(posting_date)
	""", (current_year,args.company), as_dict=True)
	total1 = 0
	budget1 = 0
	for order in sales_invoice:
		total1 += order.grand_total
		data[0].append((order.grand_total))
		if frappe.db.exists("Budget",{'company':args.company,'fiscal_year':('like',current_year)}):
			budget=frappe.get_doc("Budget",{'company':args.company,'fiscal_year':('like',current_year)})
			for i in budget.accounts:
				if i.account=="Sales - "+company_sc:
					budget1 += i.budget_amount/12			
					data[0].append(i.budget_amount/12)
		else:
			data[0].append(0)
	for month in range(1, 13):
		if not any(order['month'] == month for order in sales_invoice):
			data[0].append(0)
			if frappe.db.exists("Budget",{'company':args.company,'fiscal_year':('like',current_year)}):
				budget=frappe.get_doc("Budget",{'company':args.company,'fiscal_year':('like',current_year)})
				for i in budget.accounts:
					if i.account=="Sales - "+company_sc:
						budget1 += i.budget_amount/12			
						data[0].append(i.budget_amount/12)
			else:
				data[0].append(0)

	data[0].append((total1))
	data[0].append(budget1)
	return data


def get_data_direct(args):
	data = []
	current_date = datetime.date.today()
	current_year_ = current_date.strftime("%Y")
	current_year = str(int(current_year_))
	company_sc=frappe.get_value("Company",{'name':args.company},['abbr'])
	d_exp='Cost of Goods Sold - ' + company_sc
	
	should_append = False 
	gl_entry = frappe.db.sql("""
		SELECT 
			MONTH(posting_date) AS month,
			SUM(debit_in_transaction_currency) AS debit,
			SUM(credit_in_transaction_currency) AS credit
		FROM 
			`tabGL Entry`
		WHERE 
			YEAR(posting_date) = %s 
			AND account like %s
			AND company = %s
			AND is_cancelled=0 
		GROUP BY 
			MONTH(posting_date)
	""", (current_year,'%' + d_exp, args.company), as_dict=True)
	total_for_expense = 0
	budget1 =0
	for month in range(1, 13):
		tot=sum(entry['debit'] + entry['credit'] for entry in gl_entry if entry['month'] == month)
		total_for_expense += tot
	if total_for_expense != 0: 
		should_append = True
	# if should_append:
	data.append(['Total Material Cost'])
	for month in range(1, 13):
		tot=sum(entry['debit'] + entry['credit'] for entry in gl_entry if entry['month'] == month)

		data[-1].append(tot)
		if frappe.db.exists("Budget",{'company':args.company,'fiscal_year':('like',current_year)}):
			budget=frappe.get_doc("Budget",{'company':args.company,'fiscal_year':('like',current_year)})

			for i in budget.accounts:
				if i.account==d_exp:
					budget1 += i.budget_amount/12			
					data[-1].append(i.budget_amount/12)
		else:
			data[-1].append(0)
	data[-1].append(total_for_expense)
	data[-1].append(budget1)
		

	return data

def get_data2(args):
	data = []
	current_date = datetime.date.today()
	current_year_ = current_date.strftime("%Y")
	current_year = str(int(current_year_))
	company_sc=frappe.get_value("Company",{'name':args.company},['abbr'])
	d_exp='Direct Expenses - ' + company_sc
	direct_expense_account=frappe.get_value("Account",{'company':args.company,'name':d_exp},['name'])
	direct_expenses=frappe.db.sql("""select * from `tabAccount` where parent_account = %s""",(direct_expense_account),as_dict=True)
	for i in direct_expenses:
		should_append = False 
		gl_entry = frappe.db.sql("""
			SELECT 
				MONTH(posting_date) AS month,
				SUM(debit_in_transaction_currency) AS debit,
				SUM(credit_in_transaction_currency) AS credit
			FROM 
				`tabGL Entry`
			WHERE 
				YEAR(posting_date) = %s 
				AND account = %s
			GROUP BY 
				MONTH(posting_date)
		""", (current_year, i.name), as_dict=True)
		total_for_expense = 0
		budget1=0
		for month in range(1, 13):
			tot=sum(entry['debit'] + entry['credit'] for entry in gl_entry if entry['month'] == month)
			total_for_expense += tot
		if total_for_expense != 0: 
			should_append = True
		if should_append:
			data.append([i.name])
			for month in range(1, 13):
				tot=sum(entry['debit'] + entry['credit'] for entry in gl_entry if entry['month'] == month)

				data[-1].append(tot)
				if frappe.db.exists("Budget",{'company':args.company,'fiscal_year':('like',current_year)}):
					budget=frappe.get_doc("Budget",{'company':args.company,'fiscal_year':('like',current_year)})
					for i in budget.accounts:
						if i.account==i.name:
							budget1 += i.budget_amount/12			
							data[0].append(i.budget_amount/12)
				else:
					data[0].append(0)
			data[-1].append(total_for_expense)
			data[-1].append(budget1)
		if i.is_group==1:
			direct_expenses=frappe.db.sql("""select * from `tabAccount` where parent_account = %s""",(i.name),as_dict=True)
			for j in direct_expenses:
				should_append = False 
				gl_entry = frappe.db.sql("""
					SELECT 
						MONTH(posting_date) AS month,
						SUM(debit_in_transaction_currency) AS debit,
						SUM(credit_in_transaction_currency) AS credit
					FROM 
						`tabGL Entry`
					WHERE 
						YEAR(posting_date) = %s 
						AND account = %s
					GROUP BY 
						MONTH(posting_date)
				""", (current_year, j.name), as_dict=True)
				total_for_expense = 0
				budget1=0
				for month in range(1, 13):
					tot=sum(entry['debit'] + entry['credit'] for entry in gl_entry if entry['month'] == month)
					total_for_expense += tot
				if total_for_expense != 0: 
					should_append = True
				if should_append:
					if j.name!='%Cost of Goods Sold - '+company_sc:
						data.append([j.name])
						for month in range(1, 13):
							tot=sum(entry['debit'] + entry['credit'] for entry in gl_entry if entry['month'] == month)

							data[-1].append(tot)
							if frappe.db.exists("Budget",{'company':args.company,'fiscal_year':('like',current_year)}):
								budget=frappe.get_doc("Budget",{'company':args.company,'fiscal_year':('like',current_year)})
								for i in budget.accounts:
									if i.account==j.name:
										budget1 += i.budget_amount/12			
										data[0].append(i.budget_amount/12)
							else:
								data[0].append(0)
						data[-1].append(total_for_expense)
						data[-1].append(budget1)

	return data

def get_indirect_expenses(args):
	data = []
	current_date = datetime.date.today()
	current_year_ = current_date.strftime("%Y")
	current_year = str(int(current_year_))
	company_sc=frappe.get_value("Company",{'name':args.company},['abbr'])
	ind_exp='%Indirect Expenses - ' + company_sc
	indirect_expense_account=frappe.get_value("Account",{'company':args.company,'name':("like",ind_exp)},['name'])
	indirect_expenses=frappe.db.sql("""select name from `tabAccount` where parent_account = %s""",(indirect_expense_account),as_dict=True)

	
	for i in indirect_expenses:
		should_append = False 
		gl_entry = frappe.db.sql("""
			SELECT 
				MONTH(posting_date) AS month,
				SUM(debit_in_transaction_currency) AS debit,
				SUM(credit_in_transaction_currency) AS credit
			FROM 
				`tabGL Entry`
			WHERE 
				YEAR(posting_date) = %s 
				AND account = %s
			GROUP BY 
				MONTH(posting_date)
		""", (current_year, i.name), as_dict=True)
		total_for_expense = 0
		budget1=0
		for month in range(1, 13):
			tot=sum(entry['debit'] + entry['credit'] for entry in gl_entry if entry['month'] == month)
			total_for_expense += tot
		if frappe.db.exists("Budget",{'company':args.company,'fiscal_year':('like',current_year)}):
			budget=frappe.get_doc("Budget",{'company':args.company,'fiscal_year':('like',current_year)})
			for j in budget.accounts:
				if j.account==i.name:
					if j.budget_amount:
						data.append([i.name])	
						for month in range(1, 13):
							tot=sum(entry['debit'] + entry['credit'] for entry in gl_entry if entry['month'] == month)

							data[-1].append(tot)
							budget1 += j.budget_amount/12			
							data[-1].append(j.budget_amount/12)
						data[-1].append(total_for_expense)
						data[-1].append(budget1)
		else:
			if total_for_expense!=0:
				data.append([i.name])	
				for month in range(1, 13):
					tot=sum(entry['debit'] + entry['credit'] for entry in gl_entry if entry['month'] == month)
					data[-1].append(tot)		
					data[-1].append(0)
				data[-1].append(total_for_expense)
				data[-1].append(budget1)
	return data


class MonthlyMISReport(Document):
	pass
