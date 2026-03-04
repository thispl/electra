from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from frappe import _
import frappe
from frappe.model.document import Document
from datetime import date, timedelta, datetime,time
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
	nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
import math
from frappe.utils import add_months, cint, flt, getdate, time_diff_in_hours,time_diff_in_seconds
import locale


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import openpyxl.styles as styles
from frappe.utils import flt, fmt_money


@frappe.whitelist()
def download():
    filename = 'Statement of Account - Project'
    build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
		 
    ws = wb.create_sheet(sheet_name, 0)
    ws.column_dimensions['A'].width = 13
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20 
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15 
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15 

    company = frappe.db.get_value('Company', {'name': args.company_multiselect.split(',')[0]}, 'parent_company')
    if company:
        ws.append([company, " ", " ", " ", " ", " ", " ", "", "", "", ""])
        ws.append([args.company, " ", " ", " ", " ", " ", " ", "", "", "", ""])
    else:
        ws.append([args.company, " ", " ", " ", " ", " ", " ", "", "", "", ""]) 
        ws.append(["", " ", " ", " ", " ", " ", " ", "", "", "", ""])
    ws.append(["Statement of Account - Project", " ", " ", " ", " ", " ", " "])

    from_date_str = args.from_date
    from_date_obj = datetime.strptime(from_date_str, "%Y-%m-%d")
    formatted_from_date = from_date_obj.strftime("%d-%m-%Y")

    to_date_str = args.to_date
    to_date_obj = datetime.strptime(to_date_str, "%Y-%m-%d")
    formatted_to_date = to_date_obj.strftime("%d-%m-%Y")

    header3 = ["For the Period: " + formatted_from_date + " to " + formatted_to_date]
    ws.append(header3 + [""] * 10)
    customer = frappe.db.get_value('Statement of Accounts Report - Project', {"name": args.name}, ['customer'])
    project = frappe.db.get_value('Statement of Accounts Report - Project', {"name": args.name}, ['project'])
    project_name = frappe.db.get_value('Project', {"name": project}, ['project_name'])
    lpo = frappe.db.get_value('Sales Order', {"project": project}, ['po_no'])
    total_billed_amount = frappe.db.get_value('Sales Order', {"project": project}, ['net_bidding_price'])

    header4 = ['Customer: ' + customer]
    ws.append(header4 + [""] * 10)
    ws.append([''])
    header5 = ['Project Name: ' + project_name]
    header6 = ['Reference: ' + lpo]
    header7 = ['Project: ' + project]
    header8 = ['PO value: QAR ' + str(total_billed_amount)]
    ws.append(header5 + [""] * 10)
    ws.append(header6 + [""] * 10)
    ws.append(header7 + [""] * 10)
    ws.append(header8 + [""] * 10)
    ws.append(['Date', 'Voucher No', 'Voucher Type', 'Remarks', 'Invoice Amount', 'Advance', 'Retention' , 'Net Amount' , 'Received' , 'Balance Due'])

    data1 = get_data(args)
    for row in data1:
        ws.append(row)
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='bottom')
    bold_font = Font(bold=True)

    for header in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=10):
        for cell in header:
            cell.font = bold_font
            cell.alignment = align_center

    for header in ws.iter_rows(min_row=5, max_row=10, min_col=1, max_col=10):
        for cell in header:
            cell.font = bold_font
    

    for header in ws.iter_rows(min_row=11, max_row=11, min_col=1, max_col=10):
        for cell in header:
            cell.font = bold_font
            cell.fill = PatternFill(fgColor='D3D3D3', fill_type="solid")

    
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'))
    
    header_range = ws['A1':ws.cell(row=len(data1) + 11, column=10).coordinate]
    for row in header_range:
        for cell in row:
            cell.border = border_thin

    for row in ws.iter_rows(min_row=9, max_row=len(data1) + 11):
        if 'Opening Balance' in row[0].value or 'Closing Balance' in row[0].value or 'Total' in row[0].value:
            for cell in row:
                cell.font = bold_font
                
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=10)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=10)
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=10)
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=10)
    ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=10)
    ws.merge_cells(start_row=8, start_column=1, end_row=8, end_column=10)
    ws.merge_cells(start_row=9, start_column=1, end_row=9, end_column=10)
    ws.merge_cells(start_row=10, start_column=1, end_row=10, end_column=10)

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def get_data(args):
    data = []
    customer = frappe.db.get_value('Statement of Accounts Report - Project', {"name": args.name}, ['customer'])
    project = frappe.db.get_value('Statement of Accounts Report - Project', {"name": args.name}, ['project'])
    for company in args.company_multiselect.split(','):
        if customer and project:
            adv_invoice = frappe.db.sql("""select * from `tabAdvance Invoice` where company = %s  and transaction_date between %s and %s and docstatus = 1 and customer = %s and project = %s  order by transaction_date""", (company, args.from_date, args.to_date, customer,project), as_dict=True)
            sales_invoice = frappe.db.sql("""select * from `tabSales Invoice` where company = %s  and posting_date between %s and %s and docstatus = 1 and customer = %s and project = %s  order by posting_date""", (company, args.from_date, args.to_date, customer,project), as_dict=True)
            ret_invoice = frappe.db.sql("""select * from `tabRetention Invoice` where company = %s  and transaction_date between %s and %s and docstatus = 1 and customer = %s and project = %s  order by transaction_date""", (company, args.from_date, args.to_date, customer,project), as_dict=True)
        elif customer and not project:
            adv_invoice = frappe.db.sql("""select * from `tabAdvance Invoice` where company = %s  and transaction_date between %s and %s and docstatus = 1 and customer = %s order by transaction_date""", (company, args.from_date, args.to_date, customer), as_dict=True)
            sales_invoice = frappe.db.sql("""select * from `tabsales Invoice` where company = %s  and posting_date between %s and %s and docstatus = 1 and customer = %s order by posting_date""", (company, args.from_date, args.to_date, customer), as_dict=True)
            ret_invoice = frappe.db.sql("""select * from `tabRetention Invoice` where company = %s  and transaction_date between %s and %s and docstatus = 1 and customer = %s order by transaction_date""", (company, args.from_date, args.to_date, customer), as_dict=True)
        elif project and not customer:
            adv_invoice = frappe.db.sql("""select * from `tabAdvance Invoice` where company = %s  and transaction_date between %s and %s and docstatus = 1 and project = %s  order by transaction_date""", (company, args.from_date, args.to_date,project), as_dict=True)
            sales_invoice = frappe.db.sql("""select * from `tabSales Invoice` where company = %s  and posting_date between %s and %s and docstatus = 1 and project = %s  order by posting_date""", (company, args.from_date, args.to_date,project), as_dict=True)
            ret_invoice = frappe.db.sql("""select * from `tabRetention Invoice` where company = %s  and transaction_date between %s and %s and docstatus = 1 and project = %s  order by transaction_date""", (company, args.from_date, args.to_date,project), as_dict=True)
        else:
            adv_invoice = frappe.db.sql("""select * from `tabAdvance Invoice` where company = %s  and transaction_date between %s and %s and docstatus = 1 order by transaction_date""", (company, args.from_date, args.to_date), as_dict=True)
            sales_invoice = frappe.db.sql("""select * from `tabSales Invoice` where company = %s  and posting_date between %s and %s and docstatus = 1 order by posting_date""", (company, args.from_date, args.to_date), as_dict=True)
            ret_invoice = frappe.db.sql("""select * from `tabRetention Invoice` where company = %s  and transaction_date between %s and %s and docstatus = 1 order by transaction_date""", (company, args.from_date, args.to_date), as_dict=True)
        combined_data = []
        # Process Advance Invoice data
        for i in adv_invoice:
            journal = frappe.db.get_value("Journal Entry", {"cheque_no": i.name,"docstatus":1}, "name")
            query = """
            SELECT 
                SUM(`per`.`allocated_amount`) AS total_allocated_amount
            FROM 
                `tabPayment Entry Reference` `per`
            JOIN 
                `tabPayment Entry` `pe` ON `per`.`parent` = `pe`.`name`
            WHERE 
                `pe`.`docstatus` = 1 AND `per`.`reference_name` = %s
        """

            # Execute the query with the 'journal' parameter
            result = frappe.db.sql(query, (journal,), as_dict=True)

            # Check and fetch the result
            if result:
                received = result[0].get('total_allocated_amount') if result[0].get('total_allocated_amount') else 0
            else:
                received = 0

            combined_data.append({
                'posting_date': i.transaction_date,
                'name': i.name,
                'type': "Advance Invoice",
                'remarks': f"Customer: {i.customer}, Order No: {i.sales_order}, Project: {i.project}",
                'invoice_amount': 0,
                'advance_amount': i.get('advance_amount1', 0),
                'retention': 0,
                'net_amount': 0,
                'received': received,
                'balance': i.get('advance_amount1', 0) - received
            })

        # Process Sales Invoice data
        for j in sales_invoice:
            remarks = []
            if j.delivery_note:
                remarks.append(f"DN No: {j.delivery_note}")
            if j.po_no:
                remarks.append(f"LPO No: {j.po_no}")
            
            query = """
            SELECT 
                SUM(`per`.`allocated_amount`) AS total_allocated_amount
            FROM 
                `tabPayment Entry Reference` `per`
            JOIN 
                `tabPayment Entry` `pe` ON `per`.`parent` = `pe`.`name`
            WHERE 
                `pe`.`docstatus` = 1 AND `per`.`reference_name` = %s
        """

            # Execute the query with the 'journal' parameter
            result = frappe.db.sql(query, (j.name,), as_dict=True)

            # Check and fetch the result
            if result:
                received = result[0].get('total_allocated_amount') if result[0].get('total_allocated_amount') else 0
            else:
                received = 0

            combined_data.append({
                'posting_date': j.posting_date,
                'name': j.name,
                'type': "Sales Invoice",
                'remarks': ", ".join(remarks),
                'invoice_amount': j.get('custom_total_invoice_amount', 0),
                'advance_amount': -j.get('adv_amount', 0),
                'retention': -j.get('ret_amount', 0),
                'net_amount': j.get('net_total_project', 0),
                'received': received,
                'balance': j.get('net_total_project', 0) - received
            })
        for a in ret_invoice:
            journal = frappe.db.get_value("Journal Entry", {"cheque_no": a.name,"docstatus":1}, "name")
            query = """
                SELECT 
                    SUM(`per`.`allocated_amount`) AS total_allocated_amount
                FROM 
                    `tabPayment Entry Reference` `per`
                JOIN 
                    `tabPayment Entry` `pe` ON `per`.`parent` = `pe`.`name`
                WHERE 
                    `pe`.`docstatus` = 1 AND `per`.`reference_name` = %s
            """

            # Execute the query with the 'journal' parameter
            result = frappe.db.sql(query, (journal,), as_dict=True)

            # Check and fetch the result
            if result:
                received = result[0].get('total_allocated_amount') if result[0].get('total_allocated_amount') else 0
            else:
                received = 0
            combined_data.append({
                'posting_date': a.transaction_date,
                'name': a.name,
                'type': "Retention Invoice",
                'remarks':a.po_no or '',
                'invoice_amount': 0,
                'advance_amount': 0,
                'retention':a.get('advance_amount1', 0),
                'net_amount': 0,
                'received': received,
                'balance': a.get('advance_amount1', 0) - received
            })

    # Sort combined data by posting date
    combined_data.sort(key=lambda x: x['posting_date'])

    # Build table rows
    total_invoice = total_advance = total_retention = total_net = total_received = total_balance = 0

    for entry in combined_data:
        row = []
        row.append(entry["posting_date"].strftime("%d-%m-%Y"))
        row.append(entry["name"])
        row.append(entry["type"])
        row.append(entry["remarks"])
        row.append(float(fmt_money(round(entry["invoice_amount"], 2)).replace(',', '')) or "-")
        row.append(float(fmt_money(round(entry["advance_amount"], 2)).replace(',', '')) or "-")
        row.append(float(fmt_money(round(entry["retention"], 2)).replace(',', '')) or "-")
        row.append(float(fmt_money(round(entry["net_amount"], 2)).replace(',', '')) or "-")
        row.append(float(fmt_money(round(entry["received"], 2)).replace(',', '')) or "-")
        row.append(float(fmt_money(round(entry["balance"], 2)).replace(',', '')) or "-")

        total_invoice += entry['invoice_amount']
        total_advance += entry['advance_amount']
        total_retention += entry['retention']
        total_net += entry['net_amount']
        total_received += entry['received']
        total_balance += entry['balance']
            
        data.append(row)
        
    # Append totals to the data
    data.append([
        "Total", "", "", "",  
        float(fmt_money(round(total_invoice, 2)).replace(',', '')), 
        float(fmt_money(round(total_advance, 2)).replace(',', '')), 
        float(fmt_money(round(total_retention, 2)).replace(',', '')), 
        float(fmt_money(round(total_net, 2)).replace(',', '')), 
        float(fmt_money(round(total_received, 2)).replace(',', '')), 
        float(fmt_money(round(total_balance, 2)).replace(',', ''))
    ])

    return data
