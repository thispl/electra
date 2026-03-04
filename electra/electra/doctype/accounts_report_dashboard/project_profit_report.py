from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from frappe import _
import frappe
from frappe.model.document import Document
from datetime import date, timedelta, datetime,time
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
	nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
import math
from frappe.utils import add_months, cint, flt, getdate, time_diff_in_hours,time_diff_in_seconds
import locale


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import openpyxl.styles as styles
from frappe.utils import flt, fmt_money
from electra.project_details import cost_till_date,project_cost_till_date,total_jl_cost


@frappe.whitelist()
def download():
	filename = 'Project Profit Report'
	build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
		 
	ws = wb.create_sheet(sheet_name, 0)

	ws.column_dimensions['A'].width = 20
	ws.column_dimensions['B'].width = 20
	ws.column_dimensions['C'].width = 20 
	ws.column_dimensions['D'].width = 18
	ws.column_dimensions['E'].width = 20
	ws.column_dimensions['F'].width = 18 
	ws.column_dimensions['G'].width = 18
	ws.column_dimensions['H'].width = 18 
	ws.column_dimensions['I'].width = 18 
	ws.column_dimensions['J'].width = 18 
	ws.column_dimensions['K'].width = 18 
	ws.column_dimensions['L'].width = 18 
	ws.column_dimensions['M'].width = 15 
	

	from_date_str = today()
	from_date_obj = datetime.strptime(from_date_str, "%Y-%m-%d")
	formatted_from_date = from_date_obj.strftime("%d-%b-%Y")
	ws.append(["PROJECT PROFIT REPORT AS ON " +formatted_from_date ])
	ws.append([''])
	ws.append(['Customer','Order Ref.','Project Ref.','Order Value','Invoice Value till Date','Material Cost','Operational Cost','Cost Till Date','Gross Profit','Gross Profit %','Balance To Invoice','Total Collection','O/S Receipts'])
	data1 = get_data(args)
	for row in data1:
		ws.append(row)


	ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=13)
	ws.merge_cells(start_row=len(data1)+3, start_column=1, end_row=len(data1)+3, end_column=3)
	align_center = Alignment(horizontal='center', vertical='center')
	bold_font = Font(bold=True)

	for header in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=13):
		for cell in header:
			cell.font = bold_font
			cell.alignment = align_center
	for header in ws.iter_rows(min_row=len(data1)+3, max_row=len(data1)+3, min_col=1, max_col=13):
		for cell in header:
			cell.font = bold_font
			cell.alignment = align_center

	border_thin = Border(
		left=Side(style='thin'),
		right=Side(style='thin'),
		top=Side(style='thin'),
		bottom=Side(style='thin'))

	header_range = ws['A1':ws.cell(row=len(data1) + 3, column=13).coordinate]
	for row in header_range:
		for cell in row:
			cell.border = border_thin
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'

def get_data(args):
	data = []
	if not frappe.db.get_value("Project Profit Report",{"name":args.name},["project"]):
		total_grand = 0
		total_sales_in = 0
		total_mat_cost1 = 0
		total_op_cost1 = 0
		total_proj_cost = 0
		total_gr_pr = 0
		total_per_gp = 0
		total_bal = 0
		total_revenue = 0
		total_out = 0
		projects=frappe.get_all("Project",{'company':args.company,'creation':('between',(args.from_date,args.to_date))},['name'],order_by='name ASC')
					
		for proj in projects:
			name=proj.name
			pro_info = frappe.db.get_value(
				"Project", {'name': name}, [
					'customer', 'sales_order', 'custom_order_value', 'custom_invoice_value_till_date',
					'custom_balance_to_invoice', 'custom_gross_profit', 'custom_total_collection',
					'custom_material_cost', 'custom_operational_cost', 'custom_cost_till_date',
					'custom_gross_profit_per', 'custom_os_receipts'
				]
			)
			mr_cost = frappe.db.get_value("Overall Summary",{"parent":name,"item":"Supply Materials"},["actual"])
			data.append([pro_info[0],pro_info[1],name,float(fmt_money(pro_info[2], 2).replace(',', '')), float(fmt_money(pro_info[3], 2).replace(',', '')), mr_cost if mr_cost else 0,float(fmt_money(pro_info[8], 2).replace(',', '')),float(fmt_money(pro_info[9], 2).replace(',', '')),
				float(fmt_money(round(pro_info[5], 2), 2).replace(',', '')), float(fmt_money(round(pro_info[10], 2), 2).replace(',', '')), float(fmt_money(round(pro_info[4], 2), 2).replace(',', '')),
					float(fmt_money(round(pro_info[6], 2) or 0, 2).replace(',', '')), float(fmt_money(round(pro_info[11], 2), 2).replace(',', ''))])
			if mr_cost:
				mr_cost = mr_cost
			else:
				mr_cost = 0
			total_grand += pro_info[2]
			total_sales_in += pro_info[3] or 0
			total_mat_cost1 += mr_cost
			total_op_cost1 += pro_info[8]
			total_proj_cost += pro_info[9]
			total_gr_pr += pro_info[5]
			total_bal += pro_info[4]
			total_revenue += pro_info[6]
			total_out += pro_info[11]

		# Calculate average gross profit percentage
		if total_sales_in > 0:
			total_per_gp = (total_gr_pr / total_sales_in) * 100
		else:
			total_per_gp = 0

		# Append totals as the last row
		data.append([
			"Total", "-", "-",  # Placeholder for non-numeric columns
			float(fmt_money(total_grand, 2).replace(',', '')),
			float(fmt_money(total_sales_in, 2).replace(',', '')),
			float(fmt_money(total_mat_cost1, 2).replace(',', '')),
			float(fmt_money(total_op_cost1, 2).replace(',', '')),
			float(fmt_money(total_proj_cost, 2).replace(',', '')),
			float(fmt_money(total_gr_pr, 2).replace(',', '')),
			float(fmt_money(total_per_gp, 2).replace(',', '')),
			float(fmt_money(total_bal, 2).replace(',', '')),
			float(fmt_money(total_revenue, 2).replace(',', '')),
			float(fmt_money(total_out, 2).replace(',', ''))
		])

	else:
		name = frappe.db.get_value("Project Profit Report",{"name":args.name},["project"])
		pro_info = frappe.db.get_value(
			"Project", {'name': name}, [
				'customer', 'sales_order', 'custom_order_value', 'custom_invoice_value_till_date',
				'custom_balance_to_invoice', 'custom_gross_profit', 'custom_total_collection',
				'custom_material_cost', 'custom_operational_cost', 'custom_cost_till_date',
				'custom_gross_profit_per', 'custom_os_receipts'
			]
		)
		data.append([pro_info[0],pro_info[1],name,float(fmt_money(pro_info[2], 2).replace(',', '')), float(fmt_money(pro_info[3], 2).replace(',', '')), float(fmt_money(pro_info[7], 2).replace(',', '')),float(fmt_money(pro_info[8], 2).replace(',', '')),float(fmt_money(pro_info[9], 2).replace(',', '')),
			float(fmt_money(round(pro_info[5], 2), 2).replace(',', '')), float(fmt_money(round(pro_info[10], 2), 2).replace(',', '')), float(fmt_money(round(pro_info[4], 2), 2).replace(',', '')),
			float(fmt_money(round(pro_info[6], 2) or 0, 2).replace(',', '')), float(fmt_money(round(pro_info[11], 2), 2).replace(',', ''))])
		data.append(['Total','','',float(fmt_money(pro_info[2], 2).replace(',', '')), float(fmt_money(pro_info[3], 2).replace(',', '')), float(fmt_money(pro_info[7], 2).replace(',', '')),float(fmt_money(pro_info[8], 2).replace(',', '')),float(fmt_money(pro_info[9], 2).replace(',', '')),
			float(fmt_money(round(pro_info[5], 2), 2).replace(',', '')), float(fmt_money(round(pro_info[10], 2), 2).replace(',', '')), float(fmt_money(round(pro_info[4], 2), 2).replace(',', '')),
			float(fmt_money(round(pro_info[6], 2) or 0, 2).replace(',', '')), float(fmt_money(round(pro_info[11], 2), 2).replace(',', ''))])

				
	return data

