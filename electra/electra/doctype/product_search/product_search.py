# Copyright (c) 2022, Abdulla and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file
from frappe.utils.file_manager import get_file
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
from frappe.utils import (
	flt,
	cint,
	cstr,
	get_html_format,
	get_url_to_form,
	gzip_decompress,
	format_duration,
)

class ProductSearch(Document):
	@frappe.whitelist()
	def get_data(self):
		maintain_stock = frappe.db.get_value("Item",{"name":self.item_code},["is_stock_item"])
		if maintain_stock==1:
			company = frappe.db.get_value("Employee",{'user_id':frappe.session.user},'company')
			w_house = frappe.db.get_value("Warehouse",{'company':company,'default_for_stock_transfer':1},['name'])
			item = frappe.get_value('Item',{'item_code':self.item_code},'item_code')
			group = frappe.get_value('Item',{'item_code':self.item_code},'item_group')
			uom = frappe.get_value('Item',{'item_code':self.item_code},'stock_uom')
			# valuation = frappe.get_value('Item',{'item_code':self.item_code},'valuation_rate')
			des = frappe.get_value('Item',{'item_code':self.item_code},'description')
			csp = 'Current Selling Price'
			c_s_p = frappe.get_value('Item Price',{'item_code':self.item_code,'price_list':'Standard Selling'},'price_list_rate') or 0
			cpp = 'Current Purchase Price'
			cost = 'COST'
			pso = 'Pending Sales order'
			po ='Total Purchase Order'
			ppo = 'Pending Purchase order'
			cspp_rate = 0
			cppp_rate = 0
			del_total = 0
			psoc = 0
			ppoc = 0
			ppoc_total = 0
			total_qty=0
			total_res=0
			data = ''
			pos = frappe.db.sql("""select `tabPurchase Order Item`.item_code as item_code,`tabPurchase Order Item`.item_name as item_name,`tabPurchase Order`.supplier as supplier,`tabPurchase Order Item`.qty as qty,`tabPurchase Order Item`.rate as rate,`tabPurchase Order`.transaction_date as date,`tabPurchase Order`.name as po from `tabPurchase Order`
			left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
			where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus != 2 """ % (self.item_code), as_dict=True)
			new_so = frappe.db.sql("""select sum(`tabSales Order Item`.qty) as qty,sum(`tabSales Order Item`.delivered_qty) as d_qty from `tabSales Order`
			left join `tabSales Order Item` on `tabSales Order`.name = `tabSales Order Item`.parent
			where `tabSales Order Item`.item_code = '%s' and `tabSales Order`.docstatus = 1 and `tabSales Order`.status != 'Closed' """ % (item), as_dict=True)[0]
			if not new_so['qty']:
				new_so['qty'] = 0
			if not new_so['d_qty']:
				new_so['d_qty'] = 0
			del_total = new_so['qty'] - new_so['d_qty']		
					
			new_po = frappe.db.sql("""select sum(`tabPurchase Order Item`.qty) as qty,sum(`tabPurchase Order Item`.received_qty) as d_qty from `tabPurchase Order`
			left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
			where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus = 1 and `tabPurchase Order`.status != 'Closed' """ % (item), as_dict=True)[0]
			if not new_po['qty']:
				new_po['qty'] = 0
			if not new_po['d_qty']:
				new_po['d_qty'] = 0
			ppoc_total = new_po['qty'] - new_po['d_qty']	
			data += '<table class="table table-bordered" style="width:50%"><tr><th style="padding:1px;border: 1px solid black;background-color:#fe3f0c;" colspan=11><center>PRODUCT SEARCH</center></th></tr>'
			data += '<tr><td colspan = 3 style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;text-align: left"><b>Item Code</b></td><td colspan = 8 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(item)
			# data += '<tr><td colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;text-align: left"><b>Item Name</b></td><td colspan = 3 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(frappe.db.get_value('Item',item,'item_name'))
			data += '<tr><td colspan = 3 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;text-align: left"><b>Description</b></td><td colspan = 8 style="padding:1px;border: 1px solid black;text-align: left" nowrap><b>%s</b></td></tr>'%(des)
			data += '<tr><td colspan = 3 style="padding:1px;border: 1px solid black;text-align: left;background-color:#FFD580;color:black"><b>%s</b></td><td colspan = 8 style="padding:1px;border: 1px solid black;text-align: left;background-color:#FFD580;color:black"><b>%s</b></td></tr>'%(csp,f"{c_s_p:.2f}")
			data += '<tr><td colspan = 3 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;text-align: left"><b>Item Group</b></td><td colspan = 8 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(group)
			
			data += '<tr><td colspan = 3 style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;text-align: left"><b>%s</b></td><td colspan = 8 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(pso,(del_total))
			if ppoc_total != 0.0:
				data += '<tr><td colspan = 3 style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;text-align: left"><b>%s</b></td><td colspan = 8 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(ppo,(ppoc_total))
			else:
				data += '<tr><td colspan = 3 style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;text-align: left"><b>%s</b></td><td colspan = 8 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(ppo,(ppoc_total)) 
			
			if not frappe.session.user == "anoop@marazeemqatar.com":
				data += '<tr><td colspan = 3 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Company</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>QTY</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Res.qty</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>UOM</b></center></td></tr>'
			
			if frappe.session.user == "anoop@marazeemqatar.com":
			
				if frappe.session.user == "anoop@marazeemqatar.com" and group == "Eyenor" or group == "NVS" or group == "Secnor":
					data += '<tr><td colspan = 3 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Company</b></center></td>td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>QTY</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Res.qty</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>UOM</b></center></td></tr>'
				else:
					data += '<tr><td colspan = 3 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Company</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>QTY</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Res.qty</b></center></td><td colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>UOM</b></center></td></tr>'
			stocks = frappe.get_all(
				"Bin",
				filters={
					"item_code": item,
					"warehouse": ("not like", "%Work In Progress%"),
				},
				fields=["actual_qty", "warehouse", "stock_uom", "stock_value", "reserved_qty", "valuation_rate"]
			)
			warehouse_data = {}
			i = 0
			cou = 0
			reser =0
			tot = 'Total'
			if group=='OptiNor':
				uom='Meter'
			# else:
			# 	uom = 'Nos'
			for stock in stocks:
				if stock.warehouse == w_house:
					
					comp = frappe.get_value("Warehouse", stock.warehouse, 'company')
					valuation = frappe.db.get_value("Bin",{"item_code": item,"warehouse": ("not like", "%Work In Progress%"),"valuation_rate":("!=",0)},["valuation_rate"]) or 0
					entries = frappe.db.get_all("Stock Reservation Entry", 
												filters={"item_code": item, "warehouse": stock.warehouse,"status":("Not In",("Cancelled","Delivered")),"docstatus":("!=",2)}, 
												fields=["reserved_qty"])
					total_reserved_qty = sum(flt(entry["reserved_qty"]) for entry in entries)
					available_qty = stock.actual_qty
					if comp in warehouse_data:
						warehouse_data[comp]['available_qty'] += available_qty
						warehouse_data[comp]['ppoc_total'] += total_reserved_qty
						warehouse_data[comp]['del_total'] += del_total
						warehouse_data[comp]['valuation'] =valuation
					else:
						warehouse_data[comp] = {
							'available_qty': available_qty,
							'ppoc_total': total_reserved_qty,
							'del_total': del_total,
							'valuation':valuation
						}
			for comp, values in warehouse_data.items():
				valuation = values['valuation']
				if not frappe.session.user == "anoop@marazeemqatar.com":
					if group=='OptiNor':
						data += '''
							<tr>
								<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=3 nowrap>{}</td>
								<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
								<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
								<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>Meter</td>
							</tr>
						'''.format(comp, round(values['available_qty'],2)-(values['ppoc_total']), (values['ppoc_total']))	
			
					# else:
					# 	data += '''
					# 		<tr>
					# 			<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1 nowrap>{}</td>
					# 			<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
					# 			<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
					# 			<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
					# 		</tr>
					# 	'''.format(comp, int(values['available_qty']), int(values['ppoc_total']),uom)	
		
			i = 0
			for stock in stocks:
				if stock.warehouse != w_house:
					valuation = frappe.db.get_value("Bin",{"item_code": item,"warehouse": ("not like", "%Work In Progress%"),"valuation_rate":("!=",0)},["valuation_rate"]) or 0
					comp = frappe.get_value("Warehouse", stock.warehouse, 'company')
					entries = frappe.db.get_all("Stock Reservation Entry", 
												filters={"item_code": item, "warehouse": stock.warehouse,"status":("Not In",("Cancelled","Delivered")),"docstatus":("!=",2)}, 
												fields=["reserved_qty"])
					total_reserved_qty = sum(flt(entry["reserved_qty"]) for entry in entries)
					available_qty = stock.actual_qty
					if comp in warehouse_data:
						warehouse_data[comp]['available_qty'] += available_qty
						warehouse_data[comp]['ppoc_total'] += total_reserved_qty
						warehouse_data[comp]['del_total'] += del_total
						warehouse_data[comp]['valuation'] = valuation
					else:
						warehouse_data[comp] = {
							'available_qty': available_qty,
							'ppoc_total': total_reserved_qty,
							'del_total': del_total,
							'valuation':valuation
						}
			for comp, values in warehouse_data.items():
				valuation=values['valuation']
				if not frappe.session.user == "anoop@marazeemqatar.com":
					if group=='OptiNor':
						data += '''
							<tr>
								<td style="padding:1px;border: 1px solid black;color:black;font-weight:bold" colspan=3 nowrap>{}</td>
								<td style="padding:1px;border: 1px solid black;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
								<td style="padding:1px;border: 1px solid black;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
								<td style="padding:1px;border: 1px solid black;color:black;font-weight:bold;text-align:center" colspan=1>Meter</td>
							</tr>
						'''.format(comp, round(values['available_qty'],2)-(values['ppoc_total']), (values['ppoc_total']))	
					else:
						data += '''
							<tr>
								<td style="padding:1px;border: 1px solid black;color:black;font-weight:bold" colspan=3 nowrap>{}</td>
								<td style="padding:1px;border: 1px solid black;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
								<td style="padding:1px;border: 1px solid black;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
								<td style="padding:1px;border: 1px solid black;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
							</tr>
						'''.format(comp, round(values['available_qty'],2)-(values['ppoc_total']), (values['ppoc_total']),uom)	
			for stock in stocks:
				
				if stock.warehouse == w_house:
					comp = frappe.get_value("Warehouse", stock.warehouse, 'company')
					entries = frappe.db.get_all("Stock Reservation Entry", 
												filters={"item_code": item, "warehouse": stock.warehouse,"status":("Not In",("Cancelled","Delivered")),"docstatus":("!=",2)}, 
												fields=["reserved_qty"])
					total_reserved_qty = sum(flt(entry["reserved_qty"]) for entry in entries)
					available_qty = stock.actual_qty
					
					total_res
					if comp in warehouse_data:
						warehouse_data[comp]['available_qty'] = available_qty
						warehouse_data[comp]['ppoc_total'] = total_reserved_qty
						warehouse_data[comp]['del_total'] = del_total
					else:
						warehouse_data[comp] = {
							'available_qty': available_qty,
							'ppoc_total': total_reserved_qty,
							'del_total': del_total
						}
			for comp, values in warehouse_data.items():
				total_qty +=values['available_qty']
				total_res +=values['ppoc_total']
			if not frappe.session.user == "anoop@marazeemqatar.com":
				data += '''
					<tr>
						<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=3>{}</td>
						<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
						<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
						<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
					</tr>
				'''.format("Total",round(total_qty,2)-(total_res),(total_res),uom)

			if frappe.session.user == "anoop@marazeemqatar.com":
				if group=='OptiNor':
					data += '''
						<tr>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=3>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
						</tr>
					'''.format("Total",round(total_qty,2)-(total_res),(total_res),uom)	

				elif frappe.session.user == "anoop@marazeemqatar.com" and group == "Eyenor" or group == "NVS" or group == "Secnor":
					data += '''
						<tr>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=3>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
						</tr>
					'''.format("Total",round(total_qty,2)-(total_res),(total_res),uom)	

				
				else:
					data += '''
						<tr>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=3>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
						</tr>
					'''.format("Total",round(total_qty,2)-(total_res),(total_res),uom)	
			
			data += '</table>'

			return data

		else:
			company = frappe.db.get_value("Employee",{'user_id':frappe.session.user},'company')
			w_house = frappe.db.get_value("Warehouse",{'company':company,'default_for_stock_transfer':1},['name'])
			item = frappe.get_value('Item',{'item_code':self.item_code},'item_code')
			group = frappe.get_value('Item',{'item_code':self.item_code},'item_group')
			uom = frappe.get_value('Item',{'item_code':self.item_code},'stock_uom')
			# valuation = frappe.get_value('Item',{'item_code':self.item_code},'valuation_rate')
			des = frappe.get_value('Item',{'item_code':self.item_code},'description')
			csp = 'Current Selling Price'
			c_s_p = frappe.get_value('Item Price',{'item_code':self.item_code,'price_list':'Standard Selling'},'price_list_rate') or 0
			cpp = 'Current Purchase Price'
			cost = 'COST'
			pso = 'Pending Sales order'
			po ='Total Purchase Order'
			ppo = 'Pending Purchase order'
			cspp_rate = 0
			cppp_rate = 0
			del_total = 0
			psoc = 0
			ppoc = 0
			ppoc_total = 0
			total_qty=0
			total_res=0
			data = ''
			pos = frappe.db.sql("""select `tabPurchase Order Item`.item_code as item_code,`tabPurchase Order Item`.item_name as item_name,`tabPurchase Order`.supplier as supplier,`tabPurchase Order Item`.qty as qty,`tabPurchase Order Item`.rate as rate,`tabPurchase Order`.transaction_date as date,`tabPurchase Order`.name as po from `tabPurchase Order`
			left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
			where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus != 2 """ % (self.item_code), as_dict=True)
			new_so = frappe.db.sql("""select sum(`tabSales Order Item`.qty) as qty,sum(`tabSales Order Item`.delivered_qty) as d_qty from `tabSales Order`
			left join `tabSales Order Item` on `tabSales Order`.name = `tabSales Order Item`.parent
			where `tabSales Order Item`.item_code = '%s' and `tabSales Order`.docstatus = 1 and `tabSales Order`.status != 'Closed' """ % (item), as_dict=True)[0]
			if not new_so['qty']:
				new_so['qty'] = 0
			if not new_so['d_qty']:
				new_so['d_qty'] = 0
			del_total = new_so['qty'] - new_so['d_qty']		
					
			new_po = frappe.db.sql("""select sum(`tabPurchase Order Item`.qty) as qty,sum(`tabPurchase Order Item`.received_qty) as d_qty from `tabPurchase Order`
			left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
			where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus = 1 and `tabPurchase Order`.status != 'Closed' """ % (item), as_dict=True)[0]
			if not new_po['qty']:
				new_po['qty'] = 0
			if not new_po['d_qty']:
				new_po['d_qty'] = 0
			ppoc_total = new_po['qty'] - new_po['d_qty']	
			data += '<table class="table table-bordered" style="width:75%"><tr><th style="padding:1px;border: 1px solid black;background-color:#fe3f0c;" colspan=12><center>PRODUCT SEARCH</center></th></tr>'
			data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;text-align: left"><b>Item Code</b></td><td colspan = 11 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(item)
			# data += '<tr><td colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;text-align: left"><b>Item Name</b></td><td colspan = 3 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(frappe.db.get_value('Item',item,'item_name'))
			data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;text-align: left"><b>Description</b></td><td colspan = 11 style="padding:1px;border: 1px solid black;text-align: left" nowrap><b>%s</b></td></tr>'%(des)
			data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;text-align: left;background-color:#FFD580;color:black"><b>%s</b></td><td colspan = 11 style="padding:1px;border: 1px solid black;text-align: left;background-color:#FFD580;color:black"><b>%s</b></td></tr>'%(csp,f"{c_s_p:.2f}")
			data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;text-align: left"><b>Item Group</b></td><td colspan = 11 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(group)
			
			data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;text-align: left"><b>%s</b></td><td colspan = 11 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(pso,(del_total))
			if ppoc_total != 0.0:
				data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;text-align: left"><b>%s</b></td><td colspan = 11 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(ppo,(ppoc_total))
			else:
				data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;text-align: left"><b>%s</b></td><td colspan = 11 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(ppo,(ppoc_total)) 
			
			if not frappe.session.user == "anoop@marazeemqatar.com":
				data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Company</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Item Code</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>QTY</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Res.qty</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>UOM</b></center></td></tr>'
			
			if frappe.session.user == "anoop@marazeemqatar.com":
			
				if frappe.session.user == "anoop@marazeemqatar.com" and group == "Eyenor" or group == "NVS" or group == "Secnor":
					data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Company</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Item Code</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>QTY</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Res.qty</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>UOM</b></center></td></tr>'
				else:
					data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Company</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Item Code</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>QTY</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Res.qty</b></center></td><td colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>UOM</b></center></td></tr>'

			product_bundle = frappe.get_doc("Product Bundle",self.item_code)
			for bundle in product_bundle.items:
				frappe.errprint(bundle.item_code)
				stocks = frappe.get_all(
				"Bin",
				filters={
					"item_code": bundle.item_code,
					"warehouse": ("not like", "%Work In Progress%"),
				},
				fields=["actual_qty", "warehouse", "stock_uom", "stock_value", "reserved_qty", "valuation_rate"] 
			)

				warehouse_data = {}
				i = 0
				cou = 0
				reser =0
				tot = 'Total'
				if group=='OptiNor':
					uom='Meter'
				# else:
				# 	uom = 'Nos'
				for stock in stocks:
					# frappe.errprint(stock.warehouse)
					valuation = frappe.db.get_value("Bin",{"item_code": item,"warehouse": ("not like", "%Work In Progress%"),"valuation_rate":("!=",0)},["valuation_rate"]) or 0
					comp = frappe.get_value("Warehouse", stock.warehouse, 'company')
					entries = frappe.db.get_all("Stock Reservation Entry", 
												filters={"item_code": bundle.item_code, "warehouse": stock.warehouse,"status":("Not In",("Cancelled","Delivered")),"docstatus":("!=",2)}, 
												fields=["reserved_qty"])
					total_reserved_qty = sum(flt(entry["reserved_qty"]) for entry in entries)
					available_qty = stock.actual_qty
					if comp in warehouse_data:
						warehouse_data[comp]['available_qty'] += available_qty
						warehouse_data[comp]['ppoc_total'] += total_reserved_qty
						warehouse_data[comp]['del_total'] += del_total
						warehouse_data[comp]['valuation'] =valuation
					else:
						warehouse_data[comp] = {
							'available_qty': available_qty,
							'ppoc_total': total_reserved_qty,
							'del_total': del_total,
							'valuation':valuation
						}
				for comp, values in warehouse_data.items():
					valuation = values['valuation']
					if not frappe.session.user == "anoop@marazeemqatar.com":
						if (values['available_qty'])-(values['ppoc_total'])>0:
							data += '''
								<tr>
									<td style="padding:1px;border: 1px solid;color:black;font-weight:bold" colspan=1 nowrap>{}</td>
									<td style="padding:1px;border: 1px solid;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
									<td style="padding:1px;border: 1px solid;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
									<td style="padding:1px;border: 1px solid;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
									<td style="padding:1px;border: 1px solid;color:black;font-weight:bold;text-align:center" colspan=1>Meter</td>
								</tr>
							'''.format(comp,bundle.item_code, round(values['available_qty'],2)-(values['ppoc_total']), (values['ppoc_total']))	

							total_qty +=values['available_qty']
							total_res +=values['ppoc_total']
			if not frappe.session.user == "anoop@marazeemqatar.com":
				data += '''
					<tr>
						<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
						<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
						<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
						<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
						<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
					</tr>
				'''.format("Total"," " ,round(total_qty,2)-(total_res),(total_res),uom)

			if frappe.session.user == "anoop@marazeemqatar.com":
				if group=='OptiNor':
					data += '''
						<tr>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
						</tr>
					'''.format("Total"," ",round(total_qty,2)-(total_res),(total_res),uom)	

				elif frappe.session.user == "anoop@marazeemqatar.com" and group == "Eyenor" or group == "NVS" or group == "Secnor":
					data += '''
						<tr>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
						</tr>
					'''.format("Total"," " ,round(total_qty,2)-(total_res),(total_res),uom)	

				
				else:
					data += '''
						<tr>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
						</tr>
					'''.format("Total"," " ,round(total_qty,2)-(total_res),(total_res),uom)	
			data += '</table>'
			return data

	@frappe.whitelist()
	def get_pod(self):
		maintain_stock = frappe.db.get_value("Item",{"name":self.item_code},["is_stock_item"])
		if maintain_stock==1:	
			item = frappe.get_value('Item',{'item_code':self.item_code},'item_code')
			new_po = frappe.db.sql("""select sum(`tabPurchase Order Item`.qty) as qty,sum(`tabPurchase Order Item`.received_qty) as d_qty from `tabPurchase Order`
			left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
			where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus = 1  and `tabPurchase Order`.status != 'Closed' order by `tabPurchase Order Item`.schedule_date""" % (item), as_dict=True)[0]
			if not new_po['qty']:
				new_po['qty'] = 0
			if not new_po['d_qty']:
				new_po['d_qty'] = 0
			ppoc_total = new_po['qty'] - new_po['d_qty']
			if ppoc_total == 0.0 :
				frappe.msgprint("No Pending Purchase Receipt")
				data = ''
			else:
				data = ''
				data += '<table class="table table-bordered" style="width:75%"><tr><td style="padding:1px;border: 1px solid black;background-color:#fe3f0c;width:75%" colspan=3><center>Pending PO Details</center></td></tr>'
				data += '<tr><td style="padding:1px;width:25%;border: 1px solid black;background-color:#6f6f6f;color:white;text-align: center"><b>Purchase Order</b></td><td style="padding:1px;width:25%;border: 1px solid black;background-color:#6f6f6f;color:white;text-align: center"><b>Expected Date of Arrival</b></td><td style="padding:1px;width:25%;border: 1px solid black;background-color:#6f6f6f;color:white;text-align: center"><b>Pending Qty</b></td></tr>'
				sa = frappe.db.sql(""" select parent,qty,schedule_date,received_qty,rate from `tabPurchase Order Item` where `tabPurchase Order Item`.item_code = '%s' order by `tabPurchase Order Item`.schedule_date """ %(item),as_dict=True)
				for i in sa:
					if i.qty == i.received_qty:
						pass
					else:
						sb = frappe.get_doc("Purchase Order", i.parent)
						if sb.docstatus == 1 and sb.status!= "Closed":
							pending_qty = i.qty - i.received_qty
							data += '<tr><td style="border: 1px solid black;text-align: center"><b>%s</b></td><td style="border: 1px solid black;text-align: center"><b>%s</b></td><td style="border: 1px solid black;text-align: center"><b>%s</b></td></tr>' %(i.parent,format_date(i.schedule_date),"{:.2f}".format(pending_qty))

				print(data)
			data += '</table>'
			return data
		else:
			bundle_item = frappe.get_doc("Product Bundle",self.item_code)
			for i in bundle_item.items:
				new_po = frappe.db.sql("""select sum(`tabPurchase Order Item`.qty) as qty,sum(`tabPurchase Order Item`.received_qty) as d_qty from `tabPurchase Order`
				left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
				where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus = 1  and `tabPurchase Order`.status != 'Closed' order by `tabPurchase Order Item`.schedule_date""" % (i.item_code), as_dict=True)[0]
				if not new_po['qty']:
					new_po['qty'] = 0
				if not new_po['d_qty']:
					new_po['d_qty'] = 0
				ppoc_total = new_po['qty'] - new_po['d_qty']
				if ppoc_total == 0.0 :
					frappe.msgprint("No Pending Purchase Receipt")
					data = ''
				else:
					data = ''
					data += '<table class="table table-bordered" style="width:75%"><tr><td style="padding:1px;border: 1px solid black;background-color:#fe3f0c;width:75%" colspan=3><center>Pending PO Details</center></td></tr>'
					data += '<tr><td style="padding:1px;width:25%;border: 1px solid black;background-color:#6f6f6f;color:white;text-align: center"><b>Purchase Order</b></td><td style="padding:1px;width:25%;border: 1px solid black;background-color:#6f6f6f;color:white;text-align: center"><b>Expected Date of Arrival</b></td><td style="padding:1px;width:25%;border: 1px solid black;background-color:#6f6f6f;color:white;text-align: center"><b>Pending Qty</b></td></tr>'
					sa = frappe.db.sql(""" select parent,qty,schedule_date,received_qty,rate from `tabPurchase Order Item` where `tabPurchase Order Item`.item_code = '%s' order by `tabPurchase Order Item`.schedule_date """ %(i.item_code),as_dict=True)
					for i in sa:
						if i.qty == i.received_qty:
							pass
						else:
							sb = frappe.get_doc("Purchase Order", i.parent)
							if sb.docstatus == 1 and sb.status!= "Closed":
								pending_qty = i.qty - i.received_qty
								data += '<tr><td style="border: 1px solid black;text-align: center"><b>%s</b></td><td style="border: 1px solid black;text-align: center"><b>%s</b></td><td style="border: 1px solid black;text-align: center"><b>%s</b></td></tr>' %(i.parent,format_date(i.schedule_date),"{:.2f}".format(pending_qty))

					print(data)
				data += '</table>'
				return data


	@frappe.whitelist()
	def get_data_perm(self):
		maintain_stock = frappe.db.get_value("Item",{"name":self.item_code},["is_stock_item"])
		if maintain_stock==1:
			company = frappe.db.get_value("Employee",{'user_id':frappe.session.user},'company')
			w_house = frappe.db.get_value("Warehouse",{'company':company,'default_for_stock_transfer':1},['name'])
			item = frappe.get_value('Item',{'item_code':self.item_code},'item_code')
			group = frappe.get_value('Item',{'item_code':self.item_code},'item_group')
			uom = frappe.get_value('Item',{'item_code':self.item_code},'stock_uom')
			# valuation = frappe.get_value('Item',{'item_code':self.item_code},'valuation_rate')
			des = frappe.get_value('Item',{'item_code':self.item_code},'description')
			csp = 'Current Selling Price'
			c_s_p = frappe.get_value('Item Price',{'item_code':self.item_code,'price_list':'Standard Selling'},'price_list_rate') or 0
			cpp = 'Current Purchase Price'
			cost = 'COST'
			pso = 'Pending Sales order'
			po ='Total Purchase Order'
			ppo = 'Pending Purchase order'
			cspp_rate = 0
			cppp_rate = 0
			del_total = 0
			psoc = 0
			ppoc = 0
			ppoc_total = 0
			total_qty=0
			total_res=0
			data = ''
			pos = frappe.db.sql("""select `tabPurchase Order Item`.item_code as item_code,`tabPurchase Order Item`.item_name as item_name,`tabPurchase Order`.supplier as supplier,`tabPurchase Order Item`.qty as qty,`tabPurchase Order Item`.rate as rate,`tabPurchase Order`.transaction_date as date,`tabPurchase Order`.name as po from `tabPurchase Order`
			left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
			where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus != 2 """ % (self.item_code), as_dict=True)
			new_so = frappe.db.sql("""select sum(`tabSales Order Item`.qty) as qty,sum(`tabSales Order Item`.delivered_qty) as d_qty from `tabSales Order`
			left join `tabSales Order Item` on `tabSales Order`.name = `tabSales Order Item`.parent
			where `tabSales Order Item`.item_code = '%s' and `tabSales Order`.docstatus = 1 and `tabSales Order`.status != 'Closed' """ % (item), as_dict=True)[0]
			if not new_so['qty']:
				new_so['qty'] = 0
			if not new_so['d_qty']:
				new_so['d_qty'] = 0
			del_total = new_so['qty'] - new_so['d_qty']		
					
			new_po = frappe.db.sql("""select sum(`tabPurchase Order Item`.qty) as qty,sum(`tabPurchase Order Item`.received_qty) as d_qty from `tabPurchase Order`
			left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
			where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus = 1 and `tabPurchase Order`.status != 'Closed' """ % (item), as_dict=True)[0]
			if not new_po['qty']:
				new_po['qty'] = 0
			if not new_po['d_qty']:
				new_po['d_qty'] = 0
			ppoc_total = new_po['qty'] - new_po['d_qty']	
			data += '<table class="table table-bordered" style="width:75%"><tr><th style="padding:1px;border: 1px solid black;background-color:#fe3f0c;" colspan=12><center>PRODUCT SEARCH</center></th></tr>'
			data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;text-align: left"><b>Item Code</b></td><td colspan = 11 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(item)
			# data += '<tr><td colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;text-align: left"><b>Item Name</b></td><td colspan = 3 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(frappe.db.get_value('Item',item,'item_name'))
			data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;text-align: left"><b>Description</b></td><td colspan = 11 style="padding:1px;border: 1px solid black;text-align: left" nowrap><b>%s</b></td></tr>'%(des)
			data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;text-align: left;background-color:#FFD580;color:black"><b>%s</b></td><td colspan = 11 style="padding:1px;border: 1px solid black;text-align: left;background-color:#FFD580;color:black"><b>%s</b></td></tr>'%(csp,f"{c_s_p:.2f}")
			data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;text-align: left"><b>Item Group</b></td><td colspan = 11 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(group)
			
			data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;text-align: left"><b>%s</b></td><td colspan = 11 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(pso,(del_total))
			if ppoc_total != 0.0:
				data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;text-align: left"><b>%s</b></td><td colspan = 11 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(ppo,(ppoc_total))
			else:
				data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;text-align: left"><b>%s</b></td><td colspan = 11 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(ppo,(ppoc_total)) 
			
			if not frappe.session.user == "anoop@marazeemqatar.com":
				data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Company</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>QTY</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Res.qty</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>UOM</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>COST</b></center></td></tr>'
			
			if frappe.session.user == "anoop@marazeemqatar.com":
			
				if frappe.session.user == "anoop@marazeemqatar.com" and group == "Eyenor" or group == "NVS" or group == "Secnor":
					data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Company</b></center></td>td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>QTY</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Res.qty</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>UOM</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>COST</b></center></td></tr>'
				else:
					data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Company</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>QTY</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Res.qty</b></center></td><td colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>UOM</b></center></td></tr>'
			stocks = frappe.get_all(
				"Bin",
				filters={
					"item_code": item,
					"warehouse": ("not like", "%Work In Progress%"),
				},
				fields=["actual_qty", "warehouse", "stock_uom", "stock_value", "reserved_qty", "valuation_rate"]
			)

			warehouse_data = {}
			i = 0
			cou = 0
			reser =0
			tot = 'Total'
			if group=='OptiNor':
				uom='Meter'
			# else:
			# 	uom = 'Nos'
			for stock in stocks:
				if stock.warehouse == w_house:
					frappe.errprint(stock.warehouse)
					
					comp = frappe.get_value("Warehouse", stock.warehouse, 'company')
					valuation = frappe.db.get_value("Bin",{"item_code": item,"warehouse": ("not like", "%Work In Progress%"),"valuation_rate":("!=",0)},["valuation_rate"]) or 0
					entries = frappe.db.get_all("Stock Reservation Entry", 
												filters={"item_code": item, "warehouse": stock.warehouse,"status":("Not In",("Cancelled","Delivered")),"docstatus":("!=",2)}, 
												fields=["reserved_qty"])
					total_reserved_qty = sum(flt(entry["reserved_qty"]) for entry in entries)
					available_qty = stock.actual_qty
					if comp in warehouse_data:
						warehouse_data[comp]['available_qty'] += available_qty
						warehouse_data[comp]['ppoc_total'] += total_reserved_qty
						warehouse_data[comp]['del_total'] += del_total
						warehouse_data[comp]['valuation'] =valuation
					else:
						warehouse_data[comp] = {
							'available_qty': available_qty,
							'ppoc_total': total_reserved_qty,
							'del_total': del_total,
							'valuation':valuation
						}
			for comp, values in warehouse_data.items():
				valuation = values['valuation']
				if not frappe.session.user == "anoop@marazeemqatar.com":
					if group=='OptiNor':
						data += '''
							<tr>
								<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1 nowrap>{}</td>
								<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
								<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
								<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>Meter</td>
								<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:right" colspan=1>{}</td>
							</tr>
						'''.format(comp, round(values['available_qty'],2)-(values['ppoc_total']), (values['ppoc_total']), "{:.2f}".format(valuation))	
				
			i = 0
			for stock in stocks:
				if stock.warehouse != w_house:					
					comp = frappe.get_value("Warehouse", stock.warehouse, 'company')
					valuation = frappe.db.get_value("Bin",{"item_code": item,"warehouse": ("not like", "%Work In Progress%"),"valuation_rate":("!=",0)},["valuation_rate"]) or 0
					entries = frappe.db.get_all("Stock Reservation Entry", 
												filters={"item_code": item, "warehouse": stock.warehouse,"status":("Not In",("Cancelled","Delivered")),"docstatus":("!=",2)}, 
												fields=["reserved_qty"])
					total_reserved_qty = sum(flt(entry["reserved_qty"]) for entry in entries)
					available_qty = stock.actual_qty
					if comp in warehouse_data:
						warehouse_data[comp]['available_qty'] += available_qty
						warehouse_data[comp]['ppoc_total'] += total_reserved_qty
						warehouse_data[comp]['del_total'] += del_total
						warehouse_data[comp]['valuation'] = valuation
					else:
						warehouse_data[comp] = {
							'available_qty': available_qty,
							'ppoc_total': total_reserved_qty,
							'del_total': del_total,
							'valuation':valuation
						}
			for comp, values in warehouse_data.items():
				valuation=values['valuation']
				if not frappe.session.user == "anoop@marazeemqatar.com":
					if group=='OptiNor':
						data += '''
							<tr>
								<td style="padding:1px;border: 1px solid black;color:black;font-weight:bold" colspan=1 nowrap>{}</td>
								<td style="padding:1px;border: 1px solid black;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
								<td style="padding:1px;border: 1px solid black;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
								<td style="padding:1px;border: 1px solid black;color:black;font-weight:bold;text-align:center" colspan=1>Meter</td>
								<td style="padding:1px;border: 1px solid black;color:black;font-weight:bold;text-align:right" colspan=1>{}</td>
							</tr>
						'''.format(comp, round(values['available_qty'],2)-(values['ppoc_total']), (values['ppoc_total']), "{:.2f}".format(valuation))	
					else:
						data += '''
							<tr>
								<td style="padding:1px;border: 1px solid black;color:black;font-weight:bold" colspan=1 nowrap>{}</td>
								<td style="padding:1px;border: 1px solid black;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
								<td style="padding:1px;border: 1px solid black;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
								<td style="padding:1px;border: 1px solid black;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
								<td style="padding:1px;border: 1px solid black;color:black;font-weight:bold;text-align:right" colspan=1>{}</td>
							</tr>
						'''.format(comp, round(values['available_qty'],2)-(values['ppoc_total']), (values['ppoc_total']),uom,"{:.2f}".format(valuation))	
			for stock in stocks:
				
				if stock.warehouse == w_house:
					comp = frappe.get_value("Warehouse", stock.warehouse, 'company')
					entries = frappe.db.get_all("Stock Reservation Entry", 
												filters={"item_code": item, "warehouse": stock.warehouse,"status":("Not In",("Cancelled","Delivered")),"docstatus":("!=",2)}, 
												fields=["reserved_qty"])
					total_reserved_qty = sum(flt(entry["reserved_qty"]) for entry in entries)
					if frappe.session.user == "rifdy@electraqatar.com":
						available_qty = stock.actual_qty
					else:
						available_qty = stock.actual_qty - total_reserved_qty
					total_res
					if comp in warehouse_data:
						warehouse_data[comp]['available_qty'] = available_qty
						warehouse_data[comp]['ppoc_total'] = total_reserved_qty
						warehouse_data[comp]['del_total'] = del_total
					else:
						warehouse_data[comp] = {
							'available_qty': available_qty,
							'ppoc_total': total_reserved_qty,
							'del_total': del_total
						}
			for comp, values in warehouse_data.items():
				total_qty +=values['available_qty']
				total_res +=values['ppoc_total']
			if not frappe.session.user == "anoop@marazeemqatar.com":
				data += '''
					<tr>
						<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
						<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
						<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
						<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
						<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:right" colspan=1>{}</td>
					</tr>
				'''.format("Total",round(total_qty,2)-(total_res),(total_res),uom,"-")

			if frappe.session.user == "anoop@marazeemqatar.com":
				if group=='OptiNor':
					data += '''
						<tr>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:right" colspan=1>{}</td>
						</tr>
					'''.format("Total",round(total_qty,2)-(total_res),(total_res),uom,"-")	

				elif frappe.session.user == "anoop@marazeemqatar.com" and group == "Eyenor" or group == "NVS" or group == "Secnor":
					data += '''
						<tr>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
						</tr>
					'''.format("Total",round(total_qty,2)-(total_res),(total_res),uom,"-")	

				
				else:
					data += '''
						<tr>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
						</tr>
					'''.format("Total",round(total_qty,2)-(total_res),(total_res),uom,"-")	
			
			data += '</table>'

			return data
		else:
			company = frappe.db.get_value("Employee",{'user_id':frappe.session.user},'company')
			w_house = frappe.db.get_value("Warehouse",{'company':company,'default_for_stock_transfer':1},['name'])
			item = frappe.get_value('Item',{'item_code':self.item_code},'item_code')
			group = frappe.get_value('Item',{'item_code':self.item_code},'item_group')
			uom = frappe.get_value('Item',{'item_code':self.item_code},'stock_uom')
			# valuation = frappe.get_value('Item',{'item_code':self.item_code},'valuation_rate')
			des = frappe.get_value('Item',{'item_code':self.item_code},'description')
			csp = 'Current Selling Price'
			c_s_p = frappe.get_value('Item Price',{'item_code':self.item_code,'price_list':'Standard Selling'},'price_list_rate') or 0
			cpp = 'Current Purchase Price'
			cost = 'COST'
			pso = 'Pending Sales order'
			po ='Total Purchase Order'
			ppo = 'Pending Purchase order'
			cspp_rate = 0
			cppp_rate = 0
			del_total = 0
			psoc = 0
			ppoc = 0
			ppoc_total = 0
			total_qty=0
			total_res=0
			data = ''
			pos = frappe.db.sql("""select `tabPurchase Order Item`.item_code as item_code,`tabPurchase Order Item`.item_name as item_name,`tabPurchase Order`.supplier as supplier,`tabPurchase Order Item`.qty as qty,`tabPurchase Order Item`.rate as rate,`tabPurchase Order`.transaction_date as date,`tabPurchase Order`.name as po from `tabPurchase Order`
			left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
			where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus != 2 """ % (self.item_code), as_dict=True)
			new_so = frappe.db.sql("""select sum(`tabSales Order Item`.qty) as qty,sum(`tabSales Order Item`.delivered_qty) as d_qty from `tabSales Order`
			left join `tabSales Order Item` on `tabSales Order`.name = `tabSales Order Item`.parent
			where `tabSales Order Item`.item_code = '%s' and `tabSales Order`.docstatus = 1 and `tabSales Order`.status != 'Closed' """ % (item), as_dict=True)[0]
			if not new_so['qty']:
				new_so['qty'] = 0
			if not new_so['d_qty']:
				new_so['d_qty'] = 0
			del_total = new_so['qty'] - new_so['d_qty']		
					
			new_po = frappe.db.sql("""select sum(`tabPurchase Order Item`.qty) as qty,sum(`tabPurchase Order Item`.received_qty) as d_qty from `tabPurchase Order`
			left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
			where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus = 1 and `tabPurchase Order`.status != 'Closed' """ % (item), as_dict=True)[0]
			if not new_po['qty']:
				new_po['qty'] = 0
			if not new_po['d_qty']:
				new_po['d_qty'] = 0
			ppoc_total = new_po['qty'] - new_po['d_qty']	
			data += '<table class="table table-bordered" style="width:75%"><tr><th style="padding:1px;border: 1px solid black;background-color:#fe3f0c;" colspan=12><center>PRODUCT SEARCH</center></th></tr>'
			data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;text-align: left"><b>Item Code</b></td><td colspan = 11 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(item)
			# data += '<tr><td colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;text-align: left"><b>Item Name</b></td><td colspan = 3 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(frappe.db.get_value('Item',item,'item_name'))
			data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;text-align: left"><b>Description</b></td><td colspan = 11 style="padding:1px;border: 1px solid black;text-align: left" nowrap><b>%s</b></td></tr>'%(des)
			data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;text-align: left;background-color:#FFD580;color:black"><b>%s</b></td><td colspan = 11 style="padding:1px;border: 1px solid black;text-align: left;background-color:#FFD580;color:black"><b>%s</b></td></tr>'%(csp,f"{c_s_p:.2f}")
			data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;text-align: left"><b>Item Group</b></td><td colspan = 11 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(group)
			
			data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;text-align: left"><b>%s</b></td><td colspan = 11 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(pso,(del_total))
			if ppoc_total != 0.0:
				data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;text-align: left"><b>%s</b></td><td colspan = 11 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(ppo,(ppoc_total))
			else:
				data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;text-align: left"><b>%s</b></td><td colspan = 11 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(ppo,(ppoc_total)) 
			
			if not frappe.session.user == "anoop@marazeemqatar.com":
				data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Company</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Item Code</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>QTY</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Res.qty</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>UOM</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>COST</b></center></td></tr>'
			
			if frappe.session.user == "anoop@marazeemqatar.com":
			
				if frappe.session.user == "anoop@marazeemqatar.com" and group == "Eyenor" or group == "NVS" or group == "Secnor":
					data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Company</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Item Code</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>QTY</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Res.qty</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>UOM</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>COST</b></center></td></tr>'
				else:
					data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Company</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Item Code</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>QTY</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Res.qty</b></center></td><td colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>UOM</b></center></td></tr>'

			product_bundle = frappe.get_doc("Product Bundle",self.item_code)
			for bundle in product_bundle.items:
				frappe.errprint(bundle.item_code)
				stocks = frappe.get_all(
				"Bin",
				filters={
					"item_code": bundle.item_code,
					"warehouse": ("not like", "%Work In Progress%"),
				},
				fields=["actual_qty", "warehouse", "stock_uom", "stock_value", "reserved_qty", "valuation_rate"]
			)

				warehouse_data = {}
				i = 0
				cou = 0
				reser =0
				tot = 'Total'
				if group=='OptiNor':
					uom='Meter'
				# else:
				# 	uom = 'Nos'
				for stock in stocks:
					# frappe.errprint(stock.warehouse)
					valuation = round(stock['valuation_rate'],2)
					comp = frappe.get_value("Warehouse", stock.warehouse, 'company')
					entries = frappe.db.get_all("Stock Reservation Entry", 
												filters={"item_code": bundle.item_code, "warehouse": stock.warehouse,"status":("Not In",("Cancelled","Delivered")),"docstatus":("!=",2)}, 
												fields=["reserved_qty"])
					total_reserved_qty = sum(flt(entry["reserved_qty"]) for entry in entries)
					available_qty = stock.actual_qty
					if comp in warehouse_data:
						warehouse_data[comp]['available_qty'] += available_qty
						warehouse_data[comp]['ppoc_total'] += total_reserved_qty
						warehouse_data[comp]['del_total'] += del_total
						warehouse_data[comp]['valuation'] =valuation
					else:
						warehouse_data[comp] = {
							'available_qty': available_qty,
							'ppoc_total': total_reserved_qty,
							'del_total': del_total,
							'valuation':valuation
						}
				for comp, values in warehouse_data.items():
					valuation = values['valuation']
					if not frappe.session.user == "anoop@marazeemqatar.com":
						if (values['available_qty'])-(values['ppoc_total'])>0:
							data += '''
								<tr>
									<td style="padding:1px;border: 1px solid;color:black;font-weight:bold" colspan=1 nowrap>{}</td>
									<td style="padding:1px;border: 1px solid;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
									<td style="padding:1px;border: 1px solid;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
									<td style="padding:1px;border: 1px solid;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
									<td style="padding:1px;border: 1px solid;color:black;font-weight:bold;text-align:center" colspan=1>Meter</td>
									<td style="padding:1px;border: 1px solid;color:black;font-weight:bold;text-align:right" colspan=1>{}</td>
								</tr>
							'''.format(comp,bundle.item_code, round(values['available_qty'],2)-(values['ppoc_total']), (values['ppoc_total']), "{:.2f}".format(valuation))	

							total_qty +=values['available_qty']
							total_res +=values['ppoc_total']
			if not frappe.session.user == "anoop@marazeemqatar.com":
				data += '''
					<tr>
						<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
						<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
						<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
						<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
						<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
						<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:right" colspan=1>{}</td>
					</tr>
				'''.format("Total"," " ,round(total_qty,2)-(total_res),(total_res),uom,"-")

			if frappe.session.user == "anoop@marazeemqatar.com":
				if group=='OptiNor':
					data += '''
						<tr>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:right" colspan=1>{}</td>
						</tr>
					'''.format("Total"," ",round(total_qty,2)-(total_res),(total_res),uom,"-")	

				elif frappe.session.user == "anoop@marazeemqatar.com" and group == "Eyenor" or group == "NVS" or group == "Secnor":
					data += '''
						<tr>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold;text-align:center" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
						</tr>
					'''.format("Total"," " ,round(total_qty,2)-(total_res),(total_res),uom,"-")	

				
				else:
					data += '''
						<tr>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
							<td style="padding:1px;border: 1px solid black;background-color:#ffe9ad;color:black;font-weight:bold" colspan=1>{}</td>
						</tr>
					'''.format("Total"," " ,round(total_qty,2)-(total_res),(total_res),uom,"-")	
			data += '</table>'
			return data



