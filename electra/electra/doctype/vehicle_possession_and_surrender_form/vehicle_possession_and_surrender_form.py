# Copyright (c) 2023, Abdulla and contributors
# For license information, please see license.txt

import frappe
from datetime import date

today = date.today()
from frappe.model.document import Document
import datetime 
import frappe,erpnext
from frappe.utils import cint
import json
from frappe.utils import date_diff, add_months,today,add_days,add_years,nowdate,flt
from frappe.model.mapper import get_mapped_doc
from frappe.utils.file_manager import get_file
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
import datetime
from datetime import date,datetime,timedelta
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
import pandas as pd
from frappe.utils import formatdate
from frappe.utils import now
from erpnext.setup.utils import get_exchange_rate
from datetime import date


class VehiclePossessionandSurrenderForm(Document):
	pass

	@frappe.whitelist()
	def on_submit(self):
		vehicle = frappe.get_doc("Vehicle",{'name':self.plate or self.plate_no2})
		vehicle.append('record',{
			"type":self.type_of_file,
			"employee":self.emp_no,
			"date":self.possession_date or self.surrender_date,
			"remarks":self.department_manager,
			"document":self.name
		})
		vehicle.save(ignore_permissions=True)
		frappe.db.commit()
		
