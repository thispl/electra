from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from frappe import _
import frappe
from frappe.model.document import Document
from datetime import date, timedelta, datetime,time
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
    nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
import math
from frappe.utils import add_months, cint, flt, getdate, time_diff_in_hours,time_diff_in_seconds
import locale


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import openpyxl.styles as styles
from frappe.utils import flt, fmt_money


@frappe.whitelist()
def download():
    filename = 'Sales Person Wise Income Report'
    test = build_xlsx_response(filename)
    
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
        
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
         
    ws = wb.create_sheet(sheet_name, 0)
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15 
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15 
    ws.column_dimensions['G'].width = 15 
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    company = frappe.db.get_value('Company',{'name':args.company_multiselect},'parent_company')
    if(company):
        ws.append([company," "," "," "," "," "," ","","","",""])
        ws.append([args.company," "," "," "," "," "," ","","","",""])
    else:
        ws.append([args.company," "," "," "," "," "," ","","","",""]) 
        ws.append([""," "," "," "," "," "," ","","","",""])
    ws.append(["Sales Person Wise Income Report"," "," "," "," "," "," ","","","",""])
    ws.append(['','','','','','','',"","",""])
    ws.append(["SL No",'Date','Invoice Number','Customer Name','LPO No','Gross Amount','Discount','Ret Amount','Net','Collected','Balance'])
    
    data1= get_data(args)
    for row in data1:
        ws.append(row)
    align_center = Alignment(horizontal='center',vertical='center')
    align_right = Alignment(horizontal='right',vertical='bottom')
    for header in ws.iter_rows(min_row=1 , max_row=1, min_col=1, max_col=11):
        for cell in header:
            cell.font = Font(bold=True)
            cell.alignment = align_center
            # cell.fill = PatternFill(fgColor='fcff42', fill_type = "solid")
    for header in ws.iter_rows(min_row=2 , max_row=3, min_col=1, max_col=11):
        for cell in header:
            cell.font = Font(bold=True)
            cell.alignment = align_center
    for header in ws.iter_rows(min_row=5 , max_row=5, min_col=1, max_col=11):
        for cell in header:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fgColor='D3D3D3', fill_type = "solid")
    border_thin = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'))
    header_range = ws['A1':ws.cell(row=len(get_data(args))+5, column=11).coordinate]
    for row in header_range:
        for cell in row:
            cell.border = border_thin
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11 )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11 )
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=11 )
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=11 )

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def get_data(args):
    data = []
    if not args.company and not args.sales_person_user:
        sales_person = frappe.get_all("Sales Person", fields=["name"])
        j = 1
        for s in sales_person:
            company_set = set()
            prev_company_name = None
            sales = frappe.get_all("Sales Invoice", {'posting_date': ('between', (args.from_date, args.to_date)), 'sales_person_user': s.name,'docstatus':1}, ["name", "customer", "total", "discount_amount", "posting_date", "outstanding_amount", "po_no", "company"],order_by="posting_date")
            company_totals = {}
            salesperson_totals = {}
            
            if sales:
                data.append([s.name])

            for i in sales:
                company_name = i["company"]
                if company_name not in company_set:
                    if prev_company_name:
                        data.append(['Total for ' + prev_company_name, "", "", "", "", round(company_totals[prev_company_name]["total"],2), round(company_totals[prev_company_name]["discount_amount"],2), round(company_totals[prev_company_name]["convert_float"],2), round(company_totals[prev_company_name]["net_total"],2), round(company_totals[prev_company_name]["collected_total"],2), round(company_totals[prev_company_name]["balance"],2)])

                    data.append([company_name])
                    prev_company_name = company_name
                    company_set.add(company_name)

                net = (i.total - i.discount_amount)
                net_int = int(net)
                net_amount = (i.total - net)
                convert_float = int(net_amount)
                collected = (net - i.outstanding_amount)
                collected_int = int(collected)

                data.append([j, i.posting_date.strftime("%d-%m-%Y"), i.name, i.customer, i.po_no, i.total, i.discount_amount, convert_float, net_int, collected_int, i.outstanding_amount])

                if company_name not in company_totals:
                    company_totals[company_name] = {
                        "total": i.total,
                        "discount_amount": i.discount_amount,
                        "net_total": net_int,
                        "convert_float": convert_float,
                        "collected_total": collected_int,
                        "balance": i.outstanding_amount
                    }
                else:
                    company_totals[company_name]["total"] += i.total
                    company_totals[company_name]["discount_amount"] += i.discount_amount
                    company_totals[company_name]["net_total"] += net_int
                    company_totals[company_name]["convert_float"] += convert_float
                    company_totals[company_name]["collected_total"] += collected_int
                    company_totals[company_name]["balance"] += i.outstanding_amount

                if s.name not in salesperson_totals:
                    salesperson_totals[s.name] = {
                        "total": i.total,
                        "discount_amount": i.discount_amount,
                        "net_total": net_int,
                        "convert_float": convert_float,
                        "collected_total": collected_int,
                        "balance": i.outstanding_amount
                    }
                else:
                    salesperson_totals[s.name]["total"] += i.total
                    salesperson_totals[s.name]["discount_amount"] += i.discount_amount
                    salesperson_totals[s.name]["net_total"] += net_int
                    salesperson_totals[s.name]["convert_float"] += convert_float
                    salesperson_totals[s.name]["collected_total"] += collected_int
                    salesperson_totals[s.name]["balance"] += i.outstanding_amount

                j += 1

            if prev_company_name:
                data.append(['Total for ' + prev_company_name, "", "", "", "", round(company_totals[prev_company_name]["total"],2), round(company_totals[prev_company_name]["discount_amount"],2), round(company_totals[prev_company_name]["convert_float"],2), round(company_totals[prev_company_name]["net_total"],2), round(company_totals[prev_company_name]["collected_total"],2), round(company_totals[prev_company_name]["balance"],2)])
            prev_company_name = None

            if s.name in salesperson_totals:
                data.append(['Total for ' + s.name, "", "", "", "", round(salesperson_totals[s.name]["total"],2), round(salesperson_totals[s.name]["discount_amount"],2), round(salesperson_totals[s.name]["convert_float"],2), round(salesperson_totals[s.name]["net_total"],2), round(salesperson_totals[s.name]["collected_total"],2), round(salesperson_totals[s.name]["balance"],2)])
   
    elif not args.company and args.sales_person_user:
        company = frappe.get_all("Company", fields=["name"])
        data.append([args.sales_person_user])
        all_companies_total = {
            "total": 0,
            "discount_amount": 0,
            "convert_float": 0,
            "net_total": 0,
            "collected_total": 0,
            "balance": 0
        }

        for c in company:
            sales_invoice = frappe.get_all("Sales Invoice", {"sales_person_user": args.sales_person_user, 'posting_date': ('between', (args.from_date,args.to_date)), 'company': c.name}, ["name", "customer", "total", "discount_amount", "posting_date", "outstanding_amount", "po_no"],order_by="posting_date")
            if sales_invoice:
                data.append([c.name])
                j = 1
                company_totals = {}

                for i in sales_invoice:
                    net = (i.total - i.discount_amount)
                    net_int = int(net)
                    net_amount = (i.total - net)
                    convert_float = int(net_amount)
                    collected = (net - i.outstanding_amount)
                    collected_int = int(collected)

                    data.append([j, i.posting_date.strftime("%d-%m-%Y"), i.name, i.customer, i.po_no, i.total, i.discount_amount, convert_float, net_int, collected_int, i.outstanding_amount])

                    if c.name not in company_totals:
                        company_totals[c.name] = {
                            "total": i.total,
                            "discount_amount": i.discount_amount,
                            "net_total": net_int,
                            "convert_float": convert_float,
                            "collected_total": collected_int,
                            "balance": i.outstanding_amount
                        }
                    else:
                        company_totals[c.name]["total"] += i.total
                        company_totals[c.name]["discount_amount"] += i.discount_amount
                        company_totals[c.name]["net_total"] += net_int
                        company_totals[c.name]['convert_float'] += convert_float
                        company_totals[c.name]["collected_total"] += collected_int
                        company_totals[c.name]['balance'] += i.outstanding_amount
                    j += 1

                data.append([c.name, "", "", "", "", company_totals[c.name].get("total", 0), company_totals[c.name].get("discount_amount", 0), company_totals[c.name].get("convert_float", 0), company_totals[c.name].get("net_total", 0), company_totals[c.name].get("collected_total", 0), company_totals[c.name].get("balance", 0)])
                
                all_companies_total["total"] += company_totals[c.name].get("total", 0)
                all_companies_total["discount_amount"] += company_totals[c.name].get("discount_amount", 0)
                all_companies_total["convert_float"] += company_totals[c.name].get("convert_float", 0)
                all_companies_total["net_total"] += company_totals[c.name].get("net_total", 0)
                all_companies_total["collected_total"] += company_totals[c.name].get("collected_total", 0)
                all_companies_total["balance"] += company_totals[c.name].get("balance", 0)

        data.append(['Total for ' + args.sales_person_user, "", "", "", "", round(all_companies_total["total"],2), round(all_companies_total["discount_amount"],2), round(all_companies_total["convert_float"],2),round(all_companies_total["net_total"],2), round(all_companies_total["collected_total"],2), round(all_companies_total["balance"],2)])
    elif args.company and not args.sales_person_user:
        company = args.company
        company_name_printed = False
        company_totals = {
            "total": 0,
            "discount_amount": 0,
            "convert_float": 0,
            "net_total": 0,
            "collected_total": 0,
            "balance": 0
        }
        
        sales_person = frappe.get_all("Sales Person", fields=["name"])
        for s in sales_person:
            sales_invoice = frappe.get_all("Sales Invoice", {"sales_person_user": s.name, 'posting_date': ('between', (args.from_date, args.to_date)), 'company':company}, ["name", "customer", "total", "discount_amount", "posting_date", "outstanding_amount", "po_no"],order_by="posting_date")
            j = 1
            if sales_invoice:
                data.append([s.name])
                if not company_name_printed:
                    data.append([company])
                    company_name_printed = True

                sales_person_totals = {
                    "total": 0,
                    "discount_amount": 0,
                    "convert_float": 0,
                    "net_total": 0,
                    "collected_total": 0,
                    "balance": 0
                }

                for i in sales_invoice:
                    net = (i.total - i.discount_amount)
                    net_int = int(net)
                    net_amount = (i.total - net)
                    convert_float = int(net_amount)
                    collected = (net - i.outstanding_amount)
                    collected_int = int(collected)
                    data.append([j, i.posting_date.strftime("%d-%m-%Y"), i.name, i.customer, i.po_no, i.total, i.discount_amount, convert_float, net_int, collected_int, i.outstanding_amount])
                    j += 1
                    sales_person_totals["total"] += i.total
                    sales_person_totals["discount_amount"] += i.discount_amount
                    sales_person_totals["convert_float"] += convert_float
                    sales_person_totals["net_total"] += net_int
                    sales_person_totals["collected_total"] += collected_int
                    sales_person_totals["balance"] += i.outstanding_amount
                    
                    company_totals["total"] += i.total
                    company_totals["discount_amount"] += i.discount_amount
                    company_totals["convert_float"] += convert_float
                    company_totals["net_total"] += net_int
                    company_totals["collected_total"] += collected_int
                    company_totals["balance"] += i.outstanding_amount
                data.append(["Total for"+s.name,"","","","",round(sales_person_totals["total"],2), round(sales_person_totals["discount_amount"],2), round(sales_person_totals["convert_float"],2), round(sales_person_totals["net_total"],2), round(sales_person_totals["collected_total"],2), round(sales_person_totals["balance"],2)])
            data.append(["Total for"+company,"","","","",round(company_totals["total"],2), round(company_totals["discount_amount"],2), round(company_totals["convert_float"],2), round(company_totals["net_total"],2), round(company_totals["collected_total"],2), round(company_totals["balance"],2)])
    else:
        salesperson_totals = {
            "total": 0,
            "discount_amount": 0,
            "convert_float": 0,
            "net_total": 0,
            "collected_total": 0,
            "balance": 0
        }

        data.append([args.sales_person_user]) 
        
        for company in args.company_multiselect.split(','):
            data.append([company])
            company_totals = {
                "total": 0,
                "discount_amount": 0,
                "convert_float": 0,
                "net_total": 0,
                "collected_total": 0,
                "balance": 0
            }

            sales_invoice = frappe.get_all("Sales Invoice", {"sales_person_user": args.sales_person_user, 'posting_date': ('between', (args.from_date, args.to_date)), 'company':company}, ["name", "customer", "total", "discount_amount", "posting_date", "outstanding_amount", "po_no"],order_by="posting_date")
            j = 1
            if sales_invoice:
                
                data.append([args.sales_person_user])
                for i in sales_invoice:
                    net = (i.total - i.discount_amount)
                    net_int = int(net)
                    net_amount = (i.total - net)
                    convert_float = int(net_amount)
                    collected = (net - i.outstanding_amount)
                    collected_int = int(collected)
                    data.append([j, i.posting_date.strftime("%d-%m-%Y"), i.name, i.customer, i.po_no, i.total, i.discount_amount, convert_float, net_int, collected_int, i.outstanding_amount])
                    j += 1
                    company_totals["total"] += i.total
                    company_totals["discount_amount"] += i.discount_amount
                    company_totals["convert_float"] += convert_float
                    company_totals["net_total"] += net_int
                    company_totals["collected_total"] += collected_int
                    company_totals["balance"] += i.outstanding_amount

                data.append(["Total for"+company, "", "", "", "", round(company_totals["total"],2), round(company_totals["discount_amount"],2), round(company_totals["convert_float"],2), round(company_totals["net_total"],2), round(company_totals["collected_total"],2), round(company_totals["balance"],2)])
                salesperson_totals["total"] += company_totals["total"]
                salesperson_totals["discount_amount"] += company_totals["discount_amount"]
                salesperson_totals["convert_float"] += company_totals["convert_float"]
                salesperson_totals["net_total"] += company_totals["net_total"]
                salesperson_totals["collected_total"] += company_totals["collected_total"]
                salesperson_totals["balance"] += company_totals["balance"]
        data.append(["Total for"+args.sales_person_user, "", "", "", "", round(salesperson_totals["total"],2), round(salesperson_totals["discount_amount"],2), round(salesperson_totals["convert_float"],2), round(salesperson_totals["net_total"],2), round(salesperson_totals["collected_total"],2), round(salesperson_totals["balance"],2)])
    return data


def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'
