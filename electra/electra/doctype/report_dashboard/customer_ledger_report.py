from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from frappe import _
import frappe
from frappe.model.document import Document
from datetime import date, timedelta, datetime,time
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
    nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
import math
from frappe.utils import add_months, cint, flt, getdate, time_diff_in_hours,time_diff_in_seconds
import locale


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import openpyxl.styles as styles
from frappe.utils import flt, fmt_money


@frappe.whitelist()
def download():
    filename = 'Customer Ledger Summary'
    test = build_xlsx_response(filename)
    
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
        
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
         
    ws = wb.create_sheet(sheet_name, 0)
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20 
    ws.column_dimensions['D'].width = 50
    company = frappe.db.get_value('Company',{'name':args.company},'parent_company')
    if(company):
        ws.append([company," "," "," "," "," "," ","","","",""])
        ws.append([args.company," "," "," "," "," "," ","","","",""])
    else:
        ws.append([args.company," "," "," "," "," "," ","","","",""]) 
        ws.append([""," "," "," "," "," "," ","","","",""])
    ws.append(["Customer Ledger Summary Report"," "," "," ","",""])
    from_date_str = args.from_date
    from_date_obj = datetime.strptime(from_date_str, "%Y-%m-%d")
    formatted_from_date = from_date_obj.strftime("%d-%m-%Y")
    to_date_str = args.to_date
    to_date_obj = datetime.strptime(to_date_str, "%Y-%m-%d")
    formatted_to_date = to_date_obj.strftime("%d-%m-%Y")
    header3=["For the Period From: "+ formatted_from_date +" "+"to"+" "+formatted_to_date]
    ws.append(header3 + [""] * 5)
    ws.append([''])
    ws.append(["Customer Name",args.customer,"","Currency",args.currency,""])
    ws.append(['Date','Voucher No','Voucher Type','Narration',"Debit","Credit"])
    ws.append(['','','','','',''])
    data1= get_data(args)
    for row in data1:
        ws.append(row)
    align_center = Alignment(horizontal='center',vertical='center')
    align_right = Alignment(horizontal='right',vertical='bottom')
    for header in ws.iter_rows(min_row=1 , max_row=1, min_col=1, max_col=4):
        for cell in header:
            cell.font = Font(bold=True)
            cell.alignment = align_center
            # cell.fill = PatternFill(fgColor='fcff42', fill_type = "solid")
    for header in ws.iter_rows(min_row=2 , max_row=6, min_col=1, max_col=5):
        for cell in header:
            cell.font = Font(bold=True)
            cell.alignment = align_center
    for header in ws.iter_rows(min_row=8 , max_row=8, min_col=1, max_col=6):
        for cell in header:
            cell.font = Font(bold=True)
            cell.alignment = align_right
    for header in ws.iter_rows(min_row=len(get_data(args))+7 , max_row=len(get_data(args))+7, min_col=1, max_col=6):
        for cell in header:
            cell.font = Font(bold=True)
            cell.alignment = align_center
    for header in ws.iter_rows(min_row=len(get_data(args))+8 , max_row=len(get_data(args))+8, min_col=1, max_col=6):
        for cell in header:
            cell.font = Font(bold=True)
            cell.alignment = align_right
    for header in ws.iter_rows(min_row=7 , max_row=7, min_col=1, max_col=6):
        for cell in header:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fgColor='D3D3D3', fill_type = "solid")
    border_thin = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'))
    header_range = ws['A1':ws.cell(row=len(get_data(args))+8, column=6).coordinate]
    for row in header_range:
        for cell in row:
            cell.border = border_thin
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6 )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6 )
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6 )
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=6 )
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=6 )
    ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=3 )
    ws.merge_cells(start_row=6, start_column=5, end_row=6, end_column=6 )
    ws.merge_cells(start_row=8, start_column=1, end_row=8, end_column=6 )
    ws.merge_cells(start_row=len(get_data(args))+7, start_column=2, end_row=len(get_data(args))+7, end_column=6)
    ws.merge_cells(start_row=len(get_data(args))+8, start_column=1, end_row=len(get_data(args))+8, end_column=4 )
    

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def get_data(args):
    row2=[]
    row3=[]
    data=[]
    credit=0
    debit=0
    currency = frappe.db.get_all("GL Entry", {'against': args.customer}, ['name', 'posting_date', 'account_currency', 'voucher_type', 'remarks', 'credit_in_account_currency', 'debit_in_account_currency'], order_by="posting_date")
    for cur in currency:
        row1=[]
        credit +=cur.credit_in_account_currency
        debit +=cur.debit_in_account_currency
        row1+=[cur.posting_date.strftime("%d-%m-%Y"),cur.name,cur.voucher_type,cur.remarks,float(fmt_money(cur.debit_in_account_currency).replace(',', '')),float(fmt_money(cur.credit_in_account_currency).replace(',', ''))]
        data.append(row1)
    count = frappe.db.count("GL Entry",{'against':args.customer})
    row2+=["No. of Entries:",count,"","","",""]
    row3+=["Total","","","",debit,credit]
    data.append(row2)
    data.append(row3)
    return data
    
    

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'
