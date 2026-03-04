# Copyright (c) 2024, Abdulla and contributors
# For license information, please see license.txt

# import frappe
from frappe.model.document import Document


class AccountsReceivableAgeing(Document):
    pass

import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model import document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from frappe import _
import frappe
from frappe.model import document
from datetime import date, timedelta, datetime,time
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
    nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
import math
from frappe.utils import add_months, cint, flt, getdate, time_diff_in_hours,time_diff_in_seconds
import locale


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import openpyxl.styles as styles
from frappe.utils import flt, fmt_money

@frappe.whitelist()
def download():
    filename = 'Accounts Receivable Aging Report'
    test = build_xlsx_response(filename)
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)
    ws.column_dimensions['A'].width = 13
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20 
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15 
    ws.column_dimensions['G'].width = 20 
    ws.column_dimensions['H'].width = 10 
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 10 
    ws.column_dimensions['L'].width = 15
    # ws.column_dimensions['M'].width = 15
    company = frappe.db.get_value('Company',{'name':args.company_multiselect.split(',')[0]},'parent_company')
    if company:
        # ws.append([company," "," "," "," "," "," ","","","",""])
        # ws.append([args.company," "," "," "," "," "," ","","","",""])
        ws.append([company," "," "," "," "," "," ","","",""])
        ws.append([args.company," "," "," "," "," "," ","","",""])
    else:
        # ws.append([args.company," "," "," "," "," "," ","","","",""]) 
        # ws.append([""," "," "," "," "," "," ","","","",""])
        ws.append([args.company," "," "," "," "," "," ","","",""]) 
        ws.append([""," "," "," "," "," "," ","","",""])
    # ws.append(["Accounts Receivable Aging Report"," "," "," "," "," "," "," "])
    ws.append(["Accounts Receivable Aging Report"," "," "," "," "," "," "])

    date = datetime.now()
    header2=["Report Date: "+date.strftime("%d-%m-%Y")]
    # ws.append(header2 + [""] * 7)
    ws.append(header2 + [""] * 6)

    customer=frappe.db.get_value('Accounts Receivable Aging Report',{"name":args.name},['customer'])
    header3=['Party: '+customer]
    # ws.append(header3 + [""] * 7)
    ws.append(header3 + [""] * 6)

    ws.append([''])
    # ws.append(["Posting Date","Voucher Number","Customer LPO","Invoiced Amount","Paid Amount","Credit Note","Outstanding Amount","Age(Days)","0-30","31-60","61-90","91-120","Above 121"])
    ws.append(["Posting Date","Voucher Number","Customer LPO","Invoiced Amount","Paid Amount","Credit Note","Outstanding Amount","Age(Days)","0-30","31-60","61-90","Above 90"])
    
    data1= accounts_aging(args)
    for row in data1:
        ws.append(row)
    align_center = Alignment(horizontal='center',vertical='center')
    align_right = Alignment(horizontal='right',vertical='bottom')
    for header in ws.iter_rows(min_row=1 , max_row=1, min_col=1, max_col=12):
        for cell in header:
            cell.font = Font(bold=True)
            cell.alignment = align_center
            # cell.fill = PatternFill(fgColor='fcff42', fill_type = "solid")
    for header in ws.iter_rows(min_row=2 , max_row=5, min_col=1, max_col=12):
        for cell in header:
            cell.font = Font(bold=True)
            cell.alignment = align_center
    for header in ws.iter_rows(min_row=7 , max_row=7, min_col=1, max_col=12):
        for cell in header:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fgColor='D3D3D3', fill_type = "solid")
            cell.alignment = align_center
    for header in ws.iter_rows(min_row=8 , max_row=len(accounts_aging(args)) +302 , min_col=1, max_col=12):
        for cell in header:
            cell.alignment = align_center
    for header in ws.iter_rows(min_row=len(accounts_aging(args)) + 7 , max_row=len(accounts_aging(args)) + 7 , min_col=1, max_col=12):
        for cell in header:
            cell.font = Font(bold=True)
            cell.alignment = align_right
    border_thin = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'))
    header_range = ws['A1':ws.cell(row=len(accounts_aging(args))+7, column=12).coordinate]
    for row in header_range:
        for cell in row:
            cell.border = border_thin
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12 )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12 )
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=12 )
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=12 )
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=12 )
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=12 )
    ws.merge_cells(start_row=len((accounts_aging(args))) + 7, start_column=1, end_row=len((accounts_aging(args))) + 7, end_column=3 )
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file
def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'


def accounts_aging(args):
    data = []
    in_amount = 0
    paid_amount = 0
    credit_note = 0
    out_amount = 0
    age_0_30 = 0
    age_31_60 = 0
    age_61_90 = 0
    # age_91_120 = 0
    # age_above_121 = 0
    age_above_90 = 0
    combined_data = []
    customer=frappe.db.get_value('Accounts Receivable Aging Report',{"name":args.name},['customer'])
    if customer:
        for company in args.company_multiselect.split(','):
            account_name = frappe.db.get_value(
                "Company",
                company,
                "default_receivable_account"
            )

            if not account_name:
                continue
            
            si_list = frappe.db.sql("""
                SELECT * FROM `tabSales Invoice` 
                WHERE company = %s AND customer = %s AND docstatus = 1 
                AND is_return = 0 AND invoice_type='Credit' 
                AND status != 'Paid' 
                ORDER BY posting_date ASC
            """, (company, customer), as_dict=True)

            for i in si_list:
                result= frappe.db.sql("""
                    SELECT sum(grand_total) as total 
                    FROM `tabSales Invoice` 
                    WHERE company = %s AND return_against = %s AND docstatus = 1
                """, (company, i.name))
                return_amount = result[0][0] if result and result[0][0] else 0
                
                result_doc = frappe.db.sql("""
                    SELECT name
                    FROM `tabSales Invoice` 
                    WHERE company = %s AND return_against = %s AND docstatus = 1
                """, (company, i.name), as_dict=True)
                pay_doc = []
                if result_doc:
                    pay_doc = frappe.db.sql("""
                        SELECT per.allocated_amount 
                        FROM `tabPayment Entry Reference` AS per
                        LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
                        WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
                    """, (company, result_doc[0]["name"]), as_dict=True)
                pay = frappe.db.sql("""
                    SELECT per.allocated_amount 
                    FROM `tabPayment Entry Reference` AS per
                    LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
                    WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
                """, (company, i.name), as_dict=True)
                value = sum(j["allocated_amount"] for j in pay) + sum(r["allocated_amount"] for r in pay_doc)

                jv = frappe.db.sql("""
                    SELECT credit_in_account_currency 
                    FROM `tabJournal Entry Account` AS per
                    LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
                    WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
                """, (company, i.name), as_dict=True)

                for k in jv:
                    value += k.credit_in_account_currency

                if value and return_amount:
                    outstanding = i.grand_total - value + return_amount
                elif value:
                    outstanding = i.grand_total - value
                elif return_amount:
                    outstanding = i.grand_total + return_amount
                else:
                    outstanding = i.grand_total

                out_amount += outstanding
                age = date_diff(today(), i.posting_date) if i.posting_date else 0

                if round(outstanding) != 0:
                    if value:
                        paid_amount += value
                    if return_amount:
                        credit_note += return_amount
                    in_amount += i.grand_total

                    if 0 <= age <= 30:
                        age_0_30 += outstanding
                    elif 31 <= age <= 60:
                        age_31_60 += outstanding
                    elif 61 <= age <= 90:
                        age_61_90 += outstanding
                    else:
                        age_above_90 += outstanding
                    # elif 91 <= age <= 120:
                    #     age_91_120 += outstanding
                    # else:
                    #     age_above_121 += outstanding

                    combined_data.append({
                        'posting_date': i.posting_date,
                        'name': i.name,
                        'po_no': i.po_no if i.po_no else '-',
                        'grand_total': i.grand_total,
                        'paid_amount': value if value else 0,
                        'credit_note': return_amount if return_amount else 0,
                        'outstanding': outstanding if outstanding else 0,
                        'age': age,
                        'oustanding_0_30':outstanding if 0 <= age <= 30 else 0,
                        'oustanding_31_60':outstanding if 31 <= age <= 60 else 0,
                        'oustanding_61_90':outstanding if 61 <= age <= 90 else 0,
                        'oustanding_above_90':outstanding if age > 90 else 0,
                        # 'oustanding_91_120':outstanding if 91 <= age <= 120 else 0,
                        # 'oustanding_above_121':outstanding if age > 120 else 0,
                    })

            # Fetch Sales Invoices that are returns
            sales = frappe.db.sql("""
                SELECT * FROM `tabSales Invoice` 
                WHERE company = %s AND customer = %s AND docstatus = 1 
                AND is_return = 1 AND invoice_type='Credit' 
                AND status != 'Paid' 
                AND (return_against IS NULL OR return_against = '')  
                ORDER BY posting_date ASC
            """, (company, customer), as_dict=True)

            for i in sales:
                pay = frappe.db.sql("""
                    SELECT per.allocated_amount 
                    FROM `tabPayment Entry Reference` AS per
                    LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
                    WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
                """, (company, i.name), as_dict=True)

                value = sum(j.allocated_amount for j in pay)
                jv = frappe.db.sql("""
                    SELECT credit_in_account_currency 
                    FROM `tabJournal Entry Account` AS per
                    LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
                    WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
                """, (company, i.name), as_dict=True)

                for k in jv:
                    value += k.credit_in_account_currency

                if value:
                    outstanding = i.grand_total - value
                else:
                    outstanding = i.grand_total

                out_amount += outstanding
                age = date_diff(today(), i.posting_date) if i.posting_date else 0

                if round(outstanding) != 0:
                    if value:
                        paid_amount += value
                    if i.grand_total:
                        credit_note += i.grand_total
                    in_amount += i.grand_total

                    if 0 <= age <= 30:
                        age_0_30 += outstanding
                    elif 31 <= age <= 60:
                        age_31_60 += outstanding
                    elif 61 <= age <= 90:
                        age_61_90 += outstanding
                    else:
                        age_above_90 += outstanding
                    # elif 91 <= age <= 120:
                    #     age_91_120 += outstanding
                    # else:
                    #     age_above_121 += outstanding

                    combined_data.append({
                        'posting_date': i.posting_date,
                        'name': i.name,
                        'po_no': i.po_no if i.po_no else '-',
                        'grand_total': i.grand_total,
                        'paid_amount': value if value else 0,
                        'credit_note': i.grand_total if i.grand_total else 0,
                        'outstanding': outstanding if outstanding else 0,
                        'age': age,
                        'oustanding_0_30':outstanding if 0 <= age <= 30 else 0,
                        'oustanding_31_60':outstanding if 31 <= age <= 60 else 0,
                        'oustanding_61_90':outstanding if 61 <= age <= 90 else 0,
                        'oustanding_above_90':outstanding if age > 90 else 0,
                        # 'oustanding_91_120':outstanding if 91 <= age <= 120 else 0,
                        # 'oustanding_above_121':outstanding if age > 120 else 0,
                    })

            # Fetch Payment Entries
            payment = frappe.db.sql("""
                SELECT * FROM `tabPayment Entry` 
                WHERE company = %s AND party = %s AND docstatus = 1 
                AND payment_type = 'Receive' 
                ORDER BY posting_date ASC
            """, (company, customer), as_dict=True)

            for v in payment:
                unallocated_amount = v.unallocated_amount
                paid_amount += unallocated_amount
                out_amount -= unallocated_amount
                age = date_diff(today(), v.posting_date)

                if 0 <= age <= 30:
                    age_0_30 -= unallocated_amount
                elif 31 <= age <= 60:
                    age_31_60 -= unallocated_amount
                elif 61 <= age <= 90:
                    age_61_90 -= unallocated_amount
                else:
                    age_above_90 -= unallocated_amount
                # elif 91 <= age <= 120:
                #     age_91_120 -= unallocated_amount
                # else:
                #     age_above_121 -= unallocated_amount

                if unallocated_amount != 0:
                    combined_data.append({
                        'posting_date': v.posting_date,
                        'name': v.name,
                        'po_no': v.reference_no if v.reference_no else '-',
                        'grand_total': 0,
                        'paid_amount': unallocated_amount if unallocated_amount else 0,
                        'credit_note': 0,
                        'outstanding': -unallocated_amount if unallocated_amount else 0,
                        'age': age,
                        'oustanding_0_30':-unallocated_amount if 0 <= age <= 30 else 0,
                        'oustanding_31_60':-unallocated_amount if 31 <= age <= 60 else 0,
                        'oustanding_61_90':-unallocated_amount if 61 <= age <= 90 else 0,
                        'oustanding_above_90':-unallocated_amount if age > 90 else 0,
                        # 'oustanding_91_120':-unallocated_amount if 91 <= age <= 120 else 0,
                        # 'oustanding_above_121':-unallocated_amount if age > 120 else 0,
                    })

            # Fetch Journal Entries
            # journal = frappe.db.sql("""
            #     SELECT per.credit_in_account_currency, per.debit_in_account_currency, pe.name, pe.posting_date,pe.user_remark,pe.cheque_no
            #     FROM `tabJournal Entry Account` AS per
            #     LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
            #     WHERE pe.company = %s AND per.account LIKE %s AND pe.docstatus = 1 
            #     AND party_type = 'Customer' AND party = %s AND per.reference_name IS NULL
            # """, (company, '%Debtors -%', customer), as_dict=True)
            journal = frappe.db.sql("""
                SELECT per.credit_in_account_currency, per.debit_in_account_currency, pe.name, pe.posting_date,pe.user_remark,pe.cheque_no
                FROM `tabJournal Entry Account` AS per
                LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
                WHERE pe.company = %s AND per.account = %s AND pe.docstatus = 1 
                AND party_type = 'Customer' AND party = %s AND per.reference_name IS NULL
            """, (company, account_name, customer), as_dict=True)

            for jour in journal:
                pay_journ = frappe.db.sql("""
                    SELECT per.allocated_amount 
                    FROM `tabPayment Entry Reference` AS per
                    LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
                    WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
                """, (company, jour.name), as_dict=True)
                value_journ = sum(jo["allocated_amount"] for jo in pay_journ)
                if value_journ:
                    value_journ = value_journ
                else:
                    value_journ = 0
                cheque_no = frappe.db.get_value("Retention Invoice",{"name":jour.cheque_no},["po_no"]) or frappe.db.get_value("Advance Invoice",{"name":jour.cheque_no},["po_no"])
                if cheque_no:
                    remark = cheque_no
                else:
                    remark = " "
                if jour.credit_in_account_currency:
                    journ_amount_credit = jour.credit_in_account_currency
                    paid_amount += journ_amount_credit - value_journ
                    in_amount -= journ_amount_credit - value_journ
                    out_amount -= journ_amount_credit - value_journ
                    age = date_diff(today(), jour.posting_date)
                    if 0 <= age <= 30:
                        age_0_30 -= (jour.credit_in_account_currency -value_journ)
                    elif 31 <= age <= 60:
                        age_31_60 -= (jour.credit_in_account_currency -value_journ)
                    elif 61 <= age <= 90:
                        age_61_90 -= (jour.credit_in_account_currency -value_journ)
                    else:
                        age_above_90 -= (jour.credit_in_account_currency -value_journ)
                    # elif 91 <= age <= 120:
                    #     age_91_120 -= (jour.credit_in_account_currency -value_journ)
                    # else:
                    #     age_above_121 -= (jour.credit_in_account_currency -value_journ)
                    combined_data.append({
                        'posting_date': jour.posting_date,
                        'name': jour.name,
                        'po_no':remark,
                        'grand_total': -jour.credit_in_account_currency,
                        'paid_amount': 0,
                        'credit_note': 0,
                        'outstanding': -(jour.credit_in_account_currency-value_journ) or 0,
                        'age': age,
                        'oustanding_0_30':-(jour.credit_in_account_currency-value_journ) if 0 <= age <= 30 else 0,
                        'oustanding_31_60':-(jour.credit_in_account_currency-value_journ) if 31 <= age <= 60 else 0,
                        'oustanding_61_90':-(jour.credit_in_account_currency-value_journ) if 61 <= age <= 90 else 0,
                        'oustanding_above_90':-(jour.credit_in_account_currency-value_journ) if age > 90 else 0,
                        # 'oustanding_91_120':-(jour.credit_in_account_currency-value_journ) if 91 <= age <= 120 else 0,
                        # 'oustanding_above_121':-(jour.credit_in_account_currency-value_journ) if age > 120 else 0,
                    })
    
                elif jour.debit_in_account_currency:
                    journ_amount_debit = jour.debit_in_account_currency
                    in_amount += journ_amount_debit - value_journ
                    out_amount += journ_amount_debit - value_journ
                    age = date_diff(today(), jour.posting_date)

                    if 0 <= age <= 30:
                        age_0_30 += (jour.debit_in_account_currency -value_journ)
                    elif 31 <= age <= 60:
                        age_31_60 += (jour.debit_in_account_currency -value_journ)
                    elif 61 <= age <= 90:
                        age_61_90 += (jour.debit_in_account_currency -value_journ)
                    else:
                        age_above_90 += (jour.debit_in_account_currency -value_journ)
                    # elif 91 <= age <= 120:
                    #     age_91_120 += (jour.debit_in_account_currency -value_journ)
                    # else:
                    #     age_above_121 += (jour.debit_in_account_currency -value_journ)
    
                    combined_data.append({
                        'posting_date': jour.posting_date,
                        'name': jour.name,
                        'po_no':remark,
                        'grand_total': jour.debit_in_account_currency,
                        'paid_amount':value_journ or 0,
                        'credit_note': 0,
                        'outstanding': (jour.debit_in_account_currency -value_journ) or 0,
                        'age': age,
                        'oustanding_0_30':jour.debit_in_account_currency -value_journ if 0 <= age <= 30 else 0,
                        'oustanding_31_60':jour.debit_in_account_currency -value_journ if 31 <= age <= 60 else 0,
                        'oustanding_61_90':jour.debit_in_account_currency -value_journ if 61 <= age <= 90 else 0,
                        'oustanding_above_90':jour.debit_in_account_currency -value_journ if age > 90 else 0,
                        # 'oustanding_91_120':jour.debit_in_account_currency -value_journ if 91 <= age <= 120 else 0,
                        # 'oustanding_above_121':jour.debit_in_account_currency -value_journ if age > 120 else 0,
                    })
    combined_data = sorted(combined_data, key=lambda x: x['posting_date'])
    for entry in combined_data:
        if entry['outstanding'] != 0:
            data.append([
                entry['posting_date'].strftime("%d-%m-%Y"),entry['name'],
                entry['po_no'],(round(entry['grand_total'], 2)),
                (round(entry['paid_amount'], 2)) if entry['paid_amount']  else '-',
                (round(entry['credit_note'],2)) if entry['credit_note'] else '-',
                (round(entry['outstanding'], 2)) if entry['outstanding'] else '-',
                entry['age'],round(entry['oustanding_0_30'],2),
                round(entry['oustanding_31_60'],2),
                round(entry['oustanding_61_90'],2),
                round(entry['oustanding_above_90'],2)
                # round(entry['oustanding_91_120'],2),
                # round(entry['oustanding_above_121'],2)
            ])
    data.append([
        'Total', '', '', round(in_amount, 2),
        round(paid_amount, 2) if paid_amount else '-',
        round(credit_note, 2) if credit_note else '-',
        round(out_amount, 2), '',
        round(age_0_30, 2),
        round(age_31_60, 2),
        round(age_61_90, 2),
        round(age_above_90, 2)
        # round(age_91_120, 2),
        # round(age_above_121, 2)
    ])

    return data