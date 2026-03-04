import frappe
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from frappe.utils import fmt_money
from frappe.utils.file_manager import save_file
from io import BytesIO
from datetime import datetime

@frappe.whitelist()
def download():
    filename = 'Group Summary Report'
    build_xlsx_response(filename)

def make_xlsx(sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    
    ws = wb.active
    ws.title = sheet_name or "Group Summary"
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15 
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15 
    ws.column_dimensions['G'].width = 15 
    ws.column_dimensions['H'].width = 15
    
    com = args.get("company")
    company = frappe.db.get_value('Company', {'name': com}, 'parent_company')
    if company:
        ws.append([company, " "])
        ws.append([com, " "])
    else:
        ws.append([com, " "])
        ws.append(["", " "])
    
    ws.append(["Group Summary Report", " "])
    to_date_str = args.get("to_date")
    to_date_obj = datetime.strptime(to_date_str, "%Y-%m-%d")
    formatted_to_date = to_date_obj.strftime("%d-%m-%Y")
    ws.append(["As On: " + formatted_to_date])
    ws.append([''])
    ws.append(["Group Name", args.get("account"), "", "", "Currency", "", args.get("currency")])
    ws.append(["Account Number", "Account Name", "Opening", "", "Movement", "", "Closing", ""])
    ws.append(["", "", "Debit", "Credit", "Debit", "Credit", "Debit", "Credit"])
    
    ws.merge_cells('A1:H1')
    ws.merge_cells('A2:H2')
    ws.merge_cells('A3:H3')
    ws.merge_cells('A4:H4')
    ws.merge_cells('C7:D7')
    ws.merge_cells('E7:F7')
    ws.merge_cells('G7:H7')
    
    data = get_data(args)
    for row in data:
        ws.append(row)
    
    align_center = Alignment(horizontal='center', vertical='center')
    for row in range(1, 5):
        for cell in ws[row]:
            cell.alignment = align_center
    
    for cell in ws[6]:
        cell.font = Font(bold=True)
        cell.alignment = align_center
    for header in ws[1]:
        header.font = Font(bold=True)
        header.alignment = align_center
    
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border_thin
    
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def get_data(args):
    data = []
    t_p_debit = 0
    t_p_credit = 0
    t_c_debit = 0
    t_c_credit = 0
    total_op_debit = 0
    total_op_credit = 0
    main_account = frappe.db.get_value("Account", args.get("account"), ["name", "account_number"], as_dict=True)
    def get_all_accounts(account_code):
        accounts = frappe.db.get_all("Account", filters={"parent_account": account_code, "disabled": 0}, fields=["name", "account_number"])
        all_accounts = accounts[:]
        for acc in accounts:
            child_accounts = get_all_accounts(acc.name)  # Recursively fetch child accounts
            all_accounts.extend(child_accounts)
        return all_accounts

    accounts = [main_account] + get_all_accounts(args.get("account"))
    # accounts = frappe.db.get_all("Account", filters={"parent_account": args.get("account"), "disabled": 0}, fields=["name", "account_number"])
    
    for j in accounts:
        gle = frappe.db.sql("""
        SELECT sum(debit) as debit_amount, sum(credit) as credit_amount
        FROM `tabGL Entry` 
        WHERE account = %s and posting_date < %s and is_opening = 'No' and is_cancelled = 0
        """, (j.name, args.get("from_date")), as_dict=True)
        
        for g in gle:
            g.debit_amount = g.debit_amount or 0
            g.credit_amount = g.credit_amount or 0
            t_p_debit += g.debit_amount
            t_p_credit += g.credit_amount
            
            sq = frappe.db.sql("""
            SELECT sum(debit_in_account_currency) as debit, sum(credit_in_account_currency) as credit 
            FROM `tabGL Entry` 
            WHERE account = %s AND posting_date BETWEEN %s AND %s AND is_opening = 'No' AND is_cancelled = 0
            """, (j.name, args.get("from_date"), args.get("to_date")), as_dict=True)
            
            for i in sq:
                i.credit = i.credit or 0
                i.debit = i.debit or 0
                op_credit = g.credit_amount + i.credit
                op_debit = g.debit_amount + i.debit
                total_op_debit += i.debit
                total_op_credit += i.credit
                t_c_credit += op_credit
                t_c_debit += op_debit
                
                if g.debit_amount or g.credit_amount or i.credit or i.debit:
                    data.append([j.account_number, j.name, g.debit_amount, g.credit_amount, i.debit, i.credit, op_debit, op_credit])
    
    data.append(["Total", '', t_p_debit, t_p_credit, total_op_debit, total_op_credit, t_c_debit, t_c_credit])
    return data

def build_xlsx_response(filename):
    xlsx_file = make_xlsx()
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'
