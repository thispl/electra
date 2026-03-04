from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from frappe import _
import frappe
from frappe.model.document import Document
from datetime import date, timedelta, datetime,time
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
    nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
import math
from frappe.utils import add_months, cint, flt, getdate, time_diff_in_hours,time_diff_in_seconds
import locale


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import openpyxl.styles as styles
from frappe.utils import flt, fmt_money


@frappe.whitelist()
def download():
    filename = 'Group Summary Report'
    test = build_xlsx_response(filename)
    
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
        
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
         
    ws = wb.create_sheet(sheet_name, 0)
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15 
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15 
    ws.column_dimensions['G'].width = 15 
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15 
    ws.column_dimensions['J'].width = 15 
    ws.column_dimensions['K'].width = 15 
    com = args.company
    company = frappe.db.get_value('Company',{'name':com},'parent_company')
    if(company):
        ws.append([company," "," "," "," "," "," ","","","",""])
        ws.append([com," "," "," "," "," "," ","","","",""])
    else:
        ws.append([com," "," "," "," "," "," ","","","",""]) 
        ws.append([""," "," "," "," "," "," ","","","",""])
    
    ws.append(["Group Summary Report"," "," "," "," "," "," ","","","",""])
    to_date_str = args.to_date
    to_date_obj = datetime.strptime(to_date_str, "%Y-%m-%d")
    formatted_to_date = to_date_obj.strftime("%d-%m-%Y")
    header3=["As On: "+ formatted_to_date]
    ws.append(header3 + [""] * 11)
    ws.append([''])
    ws.append(["Group Name",args.account,"","","","Currency","",args.currency,"","",""])
    ws.append([' ','','Opening',"","",'Movement',"","",'Closing',"",""])
    ws.append(['Party','Party Name','Debit','Credit','Balance','Debit','Credit','Balance','Debit','Credit','Balance'])
    ws.append(['','','','','','','','',"","",""])
    data1= get_data(args)
    for row in data1:
        ws.append(row)
    align_center = Alignment(horizontal='center',vertical='center')
    align_right = Alignment(horizontal='right',vertical='bottom')
    for header in ws.iter_rows(min_row=1 , max_row=1, min_col=1, max_col=11):
        for cell in header:
            cell.font = Font(bold=True)
            cell.alignment = align_center
            # cell.fill = PatternFill(fgColor='fcff42', fill_type = "solid")
    for header in ws.iter_rows(min_row=2 , max_row=6, min_col=1, max_col=11):
        for cell in header:
            cell.font = Font(bold=True)
            cell.alignment = align_center
    # for header in ws.iter_rows(min_row=7 , max_row=7, min_col=1, max_col=11):
    # 	for cell in header:
    # 		cell.font = Font(bold=True)
    # 		cell.alignment = align_right
    for header in ws.iter_rows(min_row=len(get_data(args))+8 , max_row=len(get_data(args))+8, min_col=1, max_col=11):
        for cell in header:
            cell.font = Font(bold=True)
            cell.alignment = align_right
    # for header in ws.iter_rows(min_row=len(get_data(args))+7 , max_row=len(get_data(args))+7, min_col=1, max_col=11):
    # 	for cell in header:
    # 		cell.font = Font(bold=True)
    # 		cell.alignment = align_right
    for header in ws.iter_rows(min_row=8 , max_row=8, min_col=1, max_col=11):
        for cell in header:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fgColor='D3D3D3', fill_type = "solid")
    for header in ws.iter_rows(min_row=7, max_row=7, min_col=1, max_col=11):
        for cell in header:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fgColor='D3D3D3', fill_type = "solid")
    border_thin = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'))
    header_range = ws['A1':ws.cell(row=len(get_data(args))+8, column=11).coordinate]
    for row in header_range:
        for cell in row:
            cell.border = border_thin
    
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11 )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11 )
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=11 )
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=11 )
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=11 )
    ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=5 )
    ws.merge_cells(start_row=6, start_column=6, end_row=6, end_column=7 )
    ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=2 )
    ws.merge_cells(start_row=7, start_column=3, end_row=7, end_column=5 )
    ws.merge_cells(start_row=7, start_column=6, end_row=7, end_column=8 )
    ws.merge_cells(start_row=7, start_column=9, end_row=7, end_column=11 )
    # ws.merge_cells(start_row=8, start_column=1, end_row=8, end_column=11 )
    

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def get_data(args):
    data = []
    row2=[]
    row3=[]
    employee = frappe.get_all("Employee",["name","employee_name"])
    op_credit = 0
    op_debit = 0
    total_op_debit = 0
    total_op_credit = 0
    t_c_credit = 0
    t_p_credit = 0
    t_c_debit = 0
    t_p_debit = 0
    t_op_debit =0
    t_op_credit = 0
    t_op_balance = 0
    t_mov_debit = 0
    t_mov_credit = 0
    t_mov_balance = 0
    t_clo_debit = 0
    t_clo_credit = 0
    t_clo_balance = 0
    for j in employee:
        gle = frappe.db.sql("""
        SELECT name, party, sum(debit) as opening_debit, sum(credit) as opening_credit
        FROM `tabGL Entry` 
        WHERE account = %s and posting_date < %s and is_opening = 'No'
        and party = %s and party_type ='Employee' and is_cancelled = 0
        """, (args.account,args.from_date,j.name), as_dict=True)
        for g in gle:
            row1=[]
            if not g.opening_debit:
                g.opening_debit = 0
            if not g.opening_credit:
                g.opening_credit = 0
            t_p_debit += g.opening_debit
            t_p_credit += g.opening_credit
            balance_op = t_p_debit - t_p_credit
            row1+=[j.name,j.employee_name,float(fmt_money(g.opening_debit).replace(',', '')),float(fmt_money(g.opening_credit).replace(',', '')),(float(fmt_money(g.opening_debit).replace(',', '')) - float(fmt_money(g.opening_credit).replace(',', '')))]
            sq = frappe.db.sql(""" select name,party,sum(debit_in_account_currency) as debit,sum(credit_in_account_currency) as credit from `tabGL Entry` where account = '%s' and party = '%s' and party_type = 'Employee' and posting_date between '%s' and '%s' and is_opening = 'No' and is_cancelled = 0 """%(args.account,j.name,args.from_date,args.to_date),as_dict=True)
            for i in sq:
                if not i.credit:
                    i.credit = 0
                if not i.debit:
                    i.debit = 0
                op_credit = g.opening_credit + i.credit
                op_debit = g.opening_debit + i.debit
                total_op_debit += i.debit
                total_op_credit += i.credit
                balance_mo = total_op_debit - total_op_credit
                t_c_credit += op_credit
                t_c_debit += op_debit
                balance_cl = t_c_debit - t_c_credit
                row1+=[float(fmt_money(i.debit).replace(',', '')),float(fmt_money(i.credit).replace(',', '')),(float(fmt_money(i.debit).replace(',', ''))-float(fmt_money(i.credit).replace(',', ''))),float(fmt_money(op_debit).replace(',', '')),float(fmt_money(op_credit).replace(',', '')),(float(fmt_money(op_debit).replace(',', ''))-float(fmt_money(op_credit).replace(',', '')))]
                if g.opening_debit or g.opening_credit or i.debit or i.credit:
                    if op_debit-op_credit:
                        data.append(row1)
        if op_debit-op_credit:
            t_op_debit += g.opening_debit
            t_op_credit += g.opening_credit
            t_op_balance = t_op_debit - t_op_credit
            t_mov_debit += i.debit
            t_mov_credit += i.credit
            t_mov_balance = t_mov_debit - t_mov_credit
            t_clo_debit += op_debit
            t_clo_credit += op_credit
            t_clo_balance =  t_clo_debit - t_clo_credit
    row2+=["Total",'',float(fmt_money(t_op_debit).replace(',', '')), float(fmt_money(t_op_credit).replace(',', '')),float(fmt_money(t_op_balance).replace(',', '')), float(fmt_money(t_mov_debit).replace(',', '')), float(fmt_money(t_mov_credit).replace(',', '')), float(fmt_money(t_mov_balance).replace(',', '')),float(fmt_money(t_clo_debit).replace(',', '')), float(fmt_money(t_clo_credit).replace(',', '')),float(fmt_money(t_clo_balance).replace(',', ''))]
    data.append(row2)
    data.append(row3)
    return data

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'
