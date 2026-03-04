from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from frappe import _
import frappe
from frappe.model.document import Document
from datetime import date, timedelta, datetime,time
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
	nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
import math
from frappe.utils import add_months, cint, flt, getdate, time_diff_in_hours,time_diff_in_seconds
import locale


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import openpyxl.styles as styles
from frappe.utils import flt, fmt_money


@frappe.whitelist()
def download():
	filename = 'Consolidated Ledger Report'
	test = build_xlsx_response(filename)
	
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
		
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
		 
	ws = wb.create_sheet(sheet_name, 0)
	ws.column_dimensions['A'].width = 50
	ws.column_dimensions['B'].width = 30
	ws.column_dimensions['C'].width = 30 
	ws.column_dimensions['D'].width = 30
	ws.column_dimensions['E'].width = 30
	ws.column_dimensions['F'].width = 30
	ws.column_dimensions['G'].width = 30
	company = frappe.db.get_value('Company',{'name':args.company},'parent_company')
	# if(company):
	# 	ws.append([company," "," "," "," "," "," ","","","",""])
	# 	ws.append([args.company," "," "," "," "," "," ","","","",""])
	# else:
	# 	ws.append([args.company," "," "," "," "," "," ","","","",""]) 
	# 	ws.append([""," "," "," "," "," "," ","","","",""])
	ws.append(["Consolidated Ledger Report"," "," "," "," "," "," ","","",""])
	ws.append(['','','','','','','','','',''])
	from_date_str = args.from_date
	from_date_obj = datetime.strptime(from_date_str, "%Y-%m-%d")
	formatted_from_date = from_date_obj.strftime("%d-%m-%Y")
	to_date_str = args.to_date
	to_date_obj = datetime.strptime(to_date_str, "%Y-%m-%d")
	formatted_to_date = to_date_obj.strftime("%d-%m-%Y")
	ws.append(["Account",args.account," ","From Date",formatted_from_date,"To Date",formatted_to_date])
	ws.append(['Company','Opening Debit','Opening Credit','Movement Debit','Movement Credit','Closing Debit','Closing Credit'])
	ws.append(['','','','','','',''])
	data1= get_data(args)
	for row in data1:
		ws.append(row)
	align_center = Alignment(horizontal='center',vertical='center')
	align_right = Alignment(horizontal='right',vertical='bottom')
	for header in ws.iter_rows(min_row=1 , max_row=1, min_col=1, max_col=10):
		for cell in header:
			cell.font = Font(bold=True)
			cell.alignment = align_center
			# cell.fill = PatternFill(fgColor='fcff42', fill_type = "solid")
	for header in ws.iter_rows(min_row=2 , max_row=5, min_col=1, max_col=10):
		for cell in header:
			cell.font = Font(bold=True)
			cell.alignment = align_center
	# for header in ws.iter_rows(min_row=len(get_data(args))+5 , max_row=len(get_data(args))+5, min_col=1, max_col=10):
	#     for cell in header:
	#         cell.font = Font(bold=True)
	#         cell.alignment = align_right
	for header in ws.iter_rows(min_row=24 , max_row=24, min_col=2, max_col=7):
		for cell in header:
			cell.font = Font(bold=True)
			cell.alignment = align_right
	for header in ws.iter_rows(min_row=24 , max_row=24, min_col=1, max_col=1):
		for cell in header:
			cell.font = Font(bold=True)
	for header in ws.iter_rows(min_row=3 , max_row=3, min_col=1, max_col=7):
		for cell in header:
			cell.font = Font(bold=True)
			cell.fill = PatternFill(fgColor='D3D3D3', fill_type = "solid")
	for header in ws.iter_rows(min_row=4 , max_row=4, min_col=1, max_col=7):
		for cell in header:
			cell.font = Font(bold=True)
			cell.fill = PatternFill(fgColor='D3D3D3', fill_type = "solid")
	border_thin = Border(
	left=Side(style='thin'),
	right=Side(style='thin'),
	top=Side(style='thin'),
	bottom=Side(style='thin'))
	header_range = ws['A1':ws.cell(row=len(get_data(args))+5, column=7).coordinate]
	for row in header_range:
		for cell in row:
			cell.border = border_thin
	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7 )
	ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7 )
	ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=1 )
	ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=3 )
	# ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=5 )
	ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=7 )
	# ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=2 )
	# ws.merge_cells(start_row=6, start_column=3, end_row=6, end_column=4 )
	# ws.merge_cells(start_row=6, start_column=5, end_row=6, end_column=5 )
	# ws.merge_cells(start_row=8, start_column=1, end_row=8, end_column=5 )
	

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def get_data(args):
	data=[]
	row2=[]
	row3=[]
	acc=''
	acct = args.account.split(' - ')
	base_account = args.account.rsplit(" - ", 1)[0]
	if len(acct) == 2:
		acc = acct[0]
	if len(acct) == 3:
		acc = f"{acct[0]} - {acct[1]}"
	if len(acct) == 4:
		acc = f"{acct[1]} - {acct[2]}"
	ac = '%'+acc+'%'
	op_credit = 0
	op_debit = 0
	total_op_debit = 0
	total_op_credit = 0
	t_c_credit = 0
	t_p_credit = 0
	t_c_debit = 0
	t_p_debit = 0
	li = []
	company = frappe.db.sql(""" select name from `tabCompany` where is_group = 1""",as_dict=1)
	for com in company:
		li.append(com.name)
		comp = frappe.db.get_list("Company",{"parent_company":com.name},['name'])
		for j in comp:
			li.append(j.name)
	for c in li:
		company_name = c
		abbr = frappe.db.get_value("Company", company_name, "abbr")
		account_name = f"{base_account} - {abbr}"
		# gle = frappe.db.sql("""select account, sum(debit) as opening_debit, sum(credit) as opening_credit from `tabGL Entry` where company = '%s'	and (posting_date < '%s' or (ifnull(is_opening, 'No') = 'Yes' and posting_date >= '%s')) and account like '%s' and is_cancelled = 0  """%(c,args.from_date,args.to_date,ac),as_dict=True)
		gle = frappe.db.sql("""select account, sum(debit) as opening_debit, sum(credit) as opening_credit from `tabGL Entry` where company = '%s'	and (posting_date < '%s' or (ifnull(is_opening, 'No') = 'Yes' and posting_date >= '%s')) and account = '%s' and is_cancelled = 0  """%(c,args.from_date,args.to_date,account_name),as_dict=True)
		for g in gle:
			row1=[]
			if not g.opening_debit:
				g.opening_debit = 0
			if not g.opening_credit:
				g.opening_credit = 0
			t_p_debit += g.opening_debit
			t_p_credit += g.opening_credit
			balance_op = t_p_debit - t_p_credit
			row1+=[c,float(fmt_money(g.opening_debit).replace(',', '')),float(fmt_money(g.opening_credit).replace(',', ''))]
			sq = frappe.db.sql(""" select company,sum(debit_in_account_currency) as debit,sum(credit_in_account_currency) as credit from `tabGL Entry` where company = '%s' and account like '%s' and posting_date between '%s' and '%s' and is_opening = 'No' and is_cancelled = 0 """%(c,ac,args.from_date,args.to_date),as_dict=True)
			for i in sq:
				if not i.credit:
					i.credit = 0
				if not i.debit:
					i.debit = 0
				op_credit = g.opening_credit + i.credit
				op_debit = g.opening_debit + i.debit
				total_op_debit += i.debit
				total_op_credit += i.credit
				t_c_credit += op_credit
				t_c_debit += op_debit
				balance_cl = t_c_debit - t_c_credit
				balance_move=total_op_debit-total_op_credit
				row1+=[float(fmt_money(i.debit).replace(',', '')),float(fmt_money(i.credit).replace(',', '')),float(fmt_money(op_debit).replace(',', '')),float(fmt_money(op_credit).replace(',', ''))]
			data.append(row1)
	row2 += ["Total",float(fmt_money(t_p_debit).replace(',', '')),float(fmt_money(t_p_credit).replace(',', '')),float(fmt_money(total_op_debit).replace(',', '')),float(fmt_money(total_op_credit).replace(',', '')),float(fmt_money(t_c_debit).replace(',', '')),float(fmt_money(t_c_credit).replace(',', ''))]
	# row3 += ["Balance",'','',fmt_money(balance_op),'','','','','',fmt_money(balance_cl)]
	data.append(row2)
	# data.append(row3)
	return data

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'
