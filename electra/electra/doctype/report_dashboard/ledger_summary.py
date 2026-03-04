from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from frappe import _
import frappe
from frappe.model.document import Document
from datetime import date, timedelta, datetime,time
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
	nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
import math
from frappe.utils import add_months, cint, flt, getdate, time_diff_in_hours,time_diff_in_seconds
import locale


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import openpyxl.styles as styles
from frappe.utils import flt, fmt_money


@frappe.whitelist()
def download():
	filename = 'Ledger Summary Report'
	test = build_xlsx_response(filename)
	
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
		
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
		 
	ws = wb.create_sheet(sheet_name, 0)
	ws.column_dimensions['A'].width = 50
	ws.column_dimensions['B'].width = 30
	ws.column_dimensions['C'].width = 30 
	ws.column_dimensions['D'].width = 30
	ws.column_dimensions['E'].width = 30
	company = frappe.db.get_value('Company',{'name':args.company},'parent_company')
	if(company):
		ws.append([company," "," "," "," "," "," ","","","",""])
		ws.append([args.company," "," "," "," "," "," ","","","",""])
	else:
		ws.append([args.company," "," "," "," "," "," ","","","",""]) 
		ws.append([""," "," "," "," "," "," ","","","",""])
	ws.append(["Ledger Summary Report"," "," "," "," "," "," ","","",""])
	ws.append(['','','','','','','','','',''])
	from_date_str = args.from_date
	from_date_obj = datetime.strptime(from_date_str, "%Y-%m-%d")
	formatted_from_date = from_date_obj.strftime("%d-%m-%Y")
	to_date_str = args.to_date
	to_date_obj = datetime.strptime(to_date_str, "%Y-%m-%d")
	formatted_to_date = to_date_obj.strftime("%d-%m-%Y")
	ws.append(["Account"+':'+args.account,"From Date",formatted_from_date,"To Date",formatted_to_date])
	ws.append(['','Opening',"Movement",'',"Closing"])
	ws.append(['Company','Balance','Debit','Credit','Balance'])
	ws.append(['','','','',''])
	data1= get_data(args)
	for row in data1:
		ws.append(row)
	align_center = Alignment(horizontal='center',vertical='center')
	align_right = Alignment(horizontal='right',vertical='bottom')
	for header in ws.iter_rows(min_row=1 , max_row=1, min_col=1, max_col=10):
		for cell in header:
			cell.font = Font(bold=True)
			cell.alignment = align_center
			# cell.fill = PatternFill(fgColor='fcff42', fill_type = "solid")
	for header in ws.iter_rows(min_row=2 , max_row=6, min_col=1, max_col=10):
		for cell in header:
			cell.font = Font(bold=True)
			cell.alignment = align_center
	# for header in ws.iter_rows(min_row=len(get_data(args))+5 , max_row=len(get_data(args))+5, min_col=1, max_col=10):
	#     for cell in header:
	#         cell.font = Font(bold=True)
	#         cell.alignment = align_right
	for header in ws.iter_rows(min_row=len(get_data(args))+8 , max_row=len(get_data(args))+8, min_col=1, max_col=5):
		for cell in header:
			cell.font = Font(bold=True)
			cell.alignment = align_right
	for header in ws.iter_rows(min_row=6 , max_row=6, min_col=1, max_col=5):
		for cell in header:
			cell.font = Font(bold=True)
			cell.fill = PatternFill(fgColor='D3D3D3', fill_type = "solid")
	for header in ws.iter_rows(min_row=7 , max_row=7, min_col=1, max_col=5):
		for cell in header:
			cell.font = Font(bold=True)
			cell.fill = PatternFill(fgColor='D3D3D3', fill_type = "solid")
	border_thin = Border(
	left=Side(style='thin'),
	right=Side(style='thin'),
	top=Side(style='thin'),
	bottom=Side(style='thin'))
	header_range = ws['A1':ws.cell(row=len(get_data(args))+8, column=5).coordinate]
	for row in header_range:
		for cell in row:
			cell.border = border_thin
	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5 )
	ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5 )
	ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=5 )
	ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=5 )
	ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=2 )
	ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=2 )
	ws.merge_cells(start_row=6, start_column=3, end_row=6, end_column=4 )
	ws.merge_cells(start_row=6, start_column=5, end_row=6, end_column=5 )
	ws.merge_cells(start_row=8, start_column=1, end_row=8, end_column=5 )
	

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def get_data(args):

	data = []

	base_account = args.account.rsplit(" - ", 1)[0]

	total_open = 0
	total_debit = 0
	total_credit = 0
	total_close = 0

	companies = frappe.get_all(
		"Company",
		filters={"is_group": 0},
		fields=["name", "abbr"]
	)

	for c in companies:

		account_name = f"{base_account} - {c.abbr}"

		# ---------- OPENING ----------
		opening = frappe.db.sql("""
			SELECT
				IFNULL(SUM(debit),0) AS debit,
				IFNULL(SUM(credit),0) AS credit
			FROM `tabGL Entry`
			WHERE
				company=%s
				AND account=%s
				AND is_cancelled=0
				AND (
					posting_date < %s
					OR (IFNULL(is_opening,'No')='Yes' AND posting_date > %s)
				)
		""", (c.name, account_name, args.from_date, args.to_date), as_dict=1)[0]

		open_balance = opening.debit - opening.credit

		# ---------- MOVEMENT ----------
		move = frappe.db.sql("""
			SELECT
				IFNULL(SUM(debit),0) AS debit,
				IFNULL(SUM(credit),0) AS credit
			FROM `tabGL Entry`
			WHERE
				company=%s
				AND account=%s
				AND posting_date BETWEEN %s AND %s
				AND is_opening='No'
				AND is_cancelled=0
		""", (c.name, account_name, args.from_date, args.to_date), as_dict=1)[0]

		close_balance = open_balance + move.debit - move.credit

		data.append([
			c.name,
			open_balance,
			move.debit,
			move.credit,
			close_balance
		])

		total_open += open_balance
		total_debit += move.debit
		total_credit += move.credit
		total_close += close_balance

	# ---------- TOTAL ROW ----------
	data.append([
		"Total",
		total_open,
		total_debit,
		total_credit,
		total_close
	])

	return data

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'
