from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model import document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from frappe import _
import frappe
from frappe.model import document
from datetime import date, timedelta, datetime,time
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
	nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
import math
from frappe.utils import add_months, cint, flt, getdate, time_diff_in_hours,time_diff_in_seconds
import locale


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import openpyxl.styles as styles
from frappe.utils import flt, fmt_money

@frappe.whitelist()
def download():
	filename = 'Accounts Ledger Summary Report'
	limit = frappe.db
	test = build_xlsx_response(filename)
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
	ws = wb.create_sheet(sheet_name, 0)
	ws.column_dimensions['A'].width = 20
	ws.column_dimensions['B'].width = 15
	ws.column_dimensions['C'].width = 10
	ws.column_dimensions['D'].width = 20
	ws.column_dimensions['E'].width = 20
	ws.column_dimensions['F'].width = 15 
	ws.column_dimensions['G'].width = 20 
	ws.column_dimensions['H'].width = 10 
	ws.column_dimensions['I'].width = 10
	ws.column_dimensions['J'].width = 10
	ws.column_dimensions['K'].width = 10 
	data2=for_limit(args)
	company = frappe.db.get_value('Company',{'name':args.company},'parent_company')
	if(company):
		ws.append([company," "," "," "," "," "," ","","","",""])
		ws.append([args.company," "," "," "," "," "," ","","","",""])
	else:
		ws.append([args.company," "," "," "," "," "," ","","","",""]) 
		ws.append([""," "," "," "," "," "," ","","","",""])
	ws.append(['','SOA :',args.customer,'','','Credit Limit :',str(data2)])
	date = datetime.now()
	ws.append(['','Division :',args.company,'','','Posting Date :',date.strftime("%d-%m-%Y")])
	from_date_str = args.from_date
	from_date_obj = datetime.strptime(from_date_str, "%Y-%m-%d")
	formatted_from_date = from_date_obj.strftime("%d-%m-%Y")
	ws.append(['','Due Date Untill :',formatted_from_date])
	ws.append([''])
	ws.append(['Invoice No','Date','Age','Account QR','Paid QR','Credit Note QR','Outstanding QR','0-30','31-60','61-90','91-Above'])
	data1= accounts_ledger(args)
	for row in data1:
		ws.append(row)
	align_center = Alignment(horizontal='center',vertical='center')
	align_right = Alignment(horizontal='right',vertical='bottom')
	for header in ws.iter_rows(min_row=1 , max_row=1, min_col=1, max_col=11):
		for cell in header:
			cell.font = Font(bold=True)
			cell.alignment = align_center
			# cell.fill = PatternFill(fgColor='fcff42', fill_type = "solid")
	for header in ws.iter_rows(min_row=2 , max_row=2, min_col=1, max_col=11):
		for cell in header:
			cell.font = Font(bold=True)
			cell.alignment = align_center
	for header in ws.iter_rows(min_row=3 , max_row=6, min_col=1, max_col=11):
		for cell in header:
			cell.font = Font(bold=True)
	for header in ws.iter_rows(min_row=7 , max_row=7, min_col=1, max_col=11):
		for cell in header:
			cell.font = Font(bold=True)
			cell.fill = PatternFill(fgColor='D3D3D3', fill_type = "solid")
			cell.alignment = align_center
	for header in ws.iter_rows(min_row=8 , max_row=len(accounts_ledger(args)) +302 , min_col=1, max_col=11):
		for cell in header:
			cell.alignment = align_center
	for header in ws.iter_rows(min_row=len(accounts_ledger(args)) + 7 , max_row=len(accounts_ledger(args)) + 7 , min_col=1, max_col=11):
		for cell in header:
			cell.font = Font(bold=True)
			cell.alignment = align_right
	border_thin = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'))
	header_range = ws['A1':ws.cell(row=len(accounts_ledger(args))+7, column=11).coordinate]
	for row in header_range:
		for cell in row:
			cell.border = border_thin
	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11 )
	ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11 )
	ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=4 )
	ws.merge_cells(start_row=3, start_column=7, end_row=3, end_column=8 )
	ws.merge_cells(start_row=4, start_column=3, end_row=4, end_column=4 )
	ws.merge_cells(start_row=4, start_column=7, end_row=4, end_column=8 )
	ws.merge_cells(start_row=5, start_column=3, end_row=5, end_column=4 )
	ws.merge_cells(start_row=5, start_column=7, end_row=5, end_column=8 )
	ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=11 )
	ws.merge_cells(start_row=len((accounts_ledger(args))) + 7, start_column=1, end_row=len((accounts_ledger(args))) + 7, end_column=3 )
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file
def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'

def for_limit(args):
		for_limit= frappe.db.get_value("Customer Credit Limit", {"parent": args.customer},["credit_limit"])
		return for_limit


	
def accounts_ledger(args):
	data=[]
	total_amount = 0
	total_paid = 0
	total_credit_note = 0
	total_outstanding = 0
	total_0_30 = 0
	total_31_60 = 0
	total_61_90 = 0
	total_91_above = 0
	sales_invoices = frappe.get_all("Sales Invoice", {'company': args.company, 'customer': args.customer}, ['posting_date', 'name', 'total', 'outstanding_amount'],order_by="posting_date")
	if sales_invoices:
		for i in sales_invoices:
			days = date_diff(args.from_date, i.posting_date)
			if 0 <= days <= 120:
				total_amount += i.total
				total_outstanding += i.outstanding_amount
				row = [
					  i.name,i.posting_date.strftime("%d-%m-%Y"),days,float(fmt_money(round(i.total, 2)).replace(',', '')),'-','-',float(fmt_money(round(i.outstanding_amount, 2)).replace(',', '')),
					  float(fmt_money(round(i.total, 2)).replace(',', '')) if 0 <= days <= 30 else '-',
					  float(fmt_money(round(i.total, 2)).replace(',', '')) if 31 <= days <= 60 else '-',
					  float(fmt_money(round(i.total, 2)).replace(',', '')) if 61 <= days <= 90 else '-',
					  float(fmt_money(round(i.total, 2)).replace(',', '')) if 91 <= days <= 120 else '-'
				]
				data.append(row)
				if 0 <= days <= 30:
					total_0_30 += i.total
				if 31 <= days <= 60:
					total_31_60 += i.total
				if 61 <= days <= 90:
					total_61_90 += i.total
				if 91 <= days <= 120:
					total_91_above += i.total
		data.append([
			'Total','','',float(fmt_money(round(total_amount, 2)).replace(',', '')),'-','-',
			 float(fmt_money(round(total_outstanding, 2)).replace(',', '')),
			 float(fmt_money(round(total_0_30, 2)).replace(',', '')),
			 float(fmt_money(round(total_31_60, 2)).replace(',', '')),
			 float(fmt_money(round(total_61_90, 2)).replace(',', '')),
			 float(fmt_money(round(total_91_above, 2)).replace(',', ''))
		])		
		
	return data