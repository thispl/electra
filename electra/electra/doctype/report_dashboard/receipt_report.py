from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from frappe import _
import frappe
from frappe.model.document import Document
from datetime import date, timedelta, datetime,time
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
	nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
import math
from frappe.utils import add_months, cint, flt, getdate, time_diff_in_hours,time_diff_in_seconds
import locale


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import openpyxl.styles as styles
from frappe.utils import flt, fmt_money

@frappe.whitelist()
def download():
	filename = 'Receipt Report'
	test = build_xlsx_response(filename)
	
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()

	ws = wb.create_sheet(sheet_name, 0)
	ws.column_dimensions['A'].width = 13
	ws.column_dimensions['B'].width = 15
	ws.column_dimensions['C'].width = 20 
	ws.column_dimensions['D'].width = 20
	ws.column_dimensions['E'].width = 20
	ws.column_dimensions['F'].width = 15
	ws.column_dimensions['G'].width = 50
	company = frappe.db.get_value('Company',{'name':args.company},'parent_company')
	if(company):
		ws.append([company," "," "," "," "," "," ","","","",""])
		ws.append([args.company," "," "," "," "," "," ","","","",""])
	else:
		ws.append([args.company," "," "," "," "," "," ","","","",""]) 
		ws.append([""," "," "," "," "," "," ","","","",""])
	from_date_str = args.from_date
	from_date_obj = datetime.strptime(from_date_str, "%Y-%m-%d")
	formatted_from_date = from_date_obj.strftime("%d-%m-%Y")
	to_date_str = args.to_date
	to_date_obj = datetime.strptime(to_date_str, "%Y-%m-%d")
	formatted_to_date = to_date_obj.strftime("%d-%m-%Y")
	header2=["Report Period: "+formatted_from_date + " " +" to "+ " "+formatted_to_date]
	ws.append( header2 + [""] * 5)
	ws.append([''])
	ws.append(["SI NO","Posting Date","Voucher Number","Party Name","Received Amount","Sales Person","Remarks"])
	data1= receipt_report(args)
	for row in data1:
		ws.append(row)
	align_center = Alignment(horizontal='center',vertical='center')
	align_right = Alignment(horizontal='right',vertical='bottom')
	for header in ws.iter_rows(min_row=1 , max_row=1, min_col=1, max_col=7):
		for cell in header:
			cell.font = Font(bold=True)
			cell.alignment = align_center
			# cell.fill = PatternFill(fgColor='fcff42', fill_type = "solid")

	for header in ws.iter_rows(min_row=2 , max_row=3, min_col=1, max_col=7):
		for cell in header:
			cell.font = Font(bold=True)
			cell.alignment = align_center
	for header in ws.iter_rows(min_row=5 , max_row=5, min_col=1, max_col=7):
		for cell in header:
			cell.font = Font(bold=True)
			cell.alignment = align_center
			cell.fill = PatternFill(fgColor='D3D3D3', fill_type = "solid")
	for header in ws.iter_rows(min_row=len(receipt_report(args)) + 5 , max_row=len(receipt_report(args)) + 5 , min_col=1, max_col=5):
		for cell in header:
			cell.font = Font(bold=True)
			cell.alignment = align_right
	for header in ws.iter_rows(min_row=6 , max_row=len(receipt_report(args)) +301 , min_col=1, max_col=7):
		for cell in header:
			cell.alignment = align_center
	for header in ws.iter_rows(min_row=len(receipt_report(args)) + 5 , max_row=len(receipt_report(args)) + 5 , min_col=1, max_col=7):
		for cell in header:
			cell.font = Font(bold=True)
			cell.alignment = align_right
	border_thin = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'))
	header_range = ws['A1':ws.cell(row=len(receipt_report(args))+5, column=7).coordinate]
	for row in header_range:
		for cell in row:
			cell.border = border_thin
			

	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7 )
	ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7 )
	ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=7 )
	ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=7 )
	ws.merge_cells(start_row=len((receipt_report(args))) + 5, start_column=1, end_row=len((receipt_report(args))) + 5, end_column=4 )
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'


def receipt_report(args):
	data =[]
	total = 0
	ind = 0
	sa = frappe.db.sql("""
		SELECT * 
		FROM `tabPayment Entry`
		WHERE company = %s AND posting_date BETWEEN %s AND %s AND payment_type ='Receive' AND party_type = 'Customer' AND docstatus = 1 order by posting_date  ASC
	""", (args.company, args.from_date, args.to_date), as_dict=True)
	
	journal = frappe.db.sql("""
        SELECT * 
        FROM `tabJournal Entry`
        WHERE 
            company = %s 
            AND posting_date BETWEEN %s AND %s 
            AND docstatus = 1 
            AND voucher_type IN ("Bank Entry", "Cash Entry")
        ORDER BY posting_date ASC
    """, (args.company, args.from_date, args.to_date), as_dict=True)	
	
	for journ in journal:
		doc_journal = frappe.get_all("Journal Entry Account", {"parent": journ.name,"party_type":"customer"}, ["party", "credit_in_account_currency"])
		for c in doc_journal:
			if c.credit_in_account_currency>0:
				ind += 1
				total += c.credit_in_account_currency
				row=[len(data) + 1,journ.posting_date.strftime("%d-%m-%Y"), journ.name, c.party, float(fmt_money(round(c.credit_in_account_currency, 2)).replace(',', '')), '',journ.remarks]
				data.append(row)

	for i in sa:
		document = frappe.get_all("Payment Entry Reference", {"parent": i.name}, ["reference_doctype", "reference_name"])
		sales_person = ''
		if document:
			for j in document:
				if j.reference_doctype == "Sales Order":
					sales_person = frappe.db.get_value("Sales Order", {"name": j.reference_name}, ["sales_person_user"])
				elif j.reference_doctype == "Sales Invoice":
					sales_person = frappe.db.get_value("Sales Invoice", {"name": j.reference_name}, ["sales_person_user"])	
		ind += 1
		total += i.received_amount
		row = [len(data) + 1, i.posting_date.strftime("%d-%m-%Y"), i.name, i.party_name, float(fmt_money(round(i.received_amount, 2)).replace(',', '')), sales_person,i.remarks]
		data.append(row)

	data.append(["Total", "", "", "", float(fmt_money(round(total, 2)).replace(',', '')), '',''])
	return data
	# 	total += i.received_amount
	# row1 =[ind,i.posting_date,i.name,i.party_name,fmt_money(round(i.received_amount, 2)),sales_person]
	# row2 =["Total","","","",fmt_money(round(total, 2)),'']
	# data.append(row1)
	# data.append(row2)
   
