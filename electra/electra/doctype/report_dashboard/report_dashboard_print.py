from __future__ import unicode_literals
import frappe,erpnext
from frappe.utils import cint
from frappe.utils import flt, fmt_money
import json
from frappe.utils import date_diff, add_months,today,add_days,add_years,nowdate,flt
from frappe.utils import cstr, cint, getdate,get_first_day, get_last_day, today, time_diff_in_hours

from frappe.model.mapper import get_mapped_doc
from frappe.utils.file_manager import get_file
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
import datetime
from datetime import date,datetime,timedelta
from electra.utils import get_dn_return_series
import openpyxl
from collections import defaultdict
from openpyxl import Workbook
import openpyxl
import xlrd
import re
from frappe.utils import date_diff,today,cstr
from datetime import datetime
from dateutil import relativedelta
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
import pandas as pd
from frappe.utils import formatdate


from erpnext.setup.utils import get_exchange_rate
from frappe import throw,_, db, get_doc, throw, whitelist


from frappe.utils import (
    add_days,
    cint,
    cstr,
    date_diff,
    flt,
    formatdate,
    get_first_day,
    get_link_to_form,
    getdate,
    money_in_words,
    rounded,
)


#Group summary report dashboard
@frappe.whitelist()
def return_account_total(from_date,to_date,account):
    
    data = '<table border="1" style="border-collapse: collapse; width: 100%;">'
    data += '<tr style="background-color: #D9E2ED;"><td colspan="1"><b></b></td><td colspan="1"><b></b></td><td colspan="3" style="text-align:center;"><b>Opening</b></td><td colspan="3" style="text-align:center;"><b>Movement</b></td><td colspan="3" style="text-align:center;"><b>Closing</b></td></tr>'
    data += '<tr style="background-color: #e35310; color: white;"><td style="text-align:center; font-weight:bold; color:white;">Party</td><td style="text-align:center; font-weight:bold; color:white;">Party Name</td><td style="text-align:center; font-weight:bold; color:white;">Debit</td><td style="text-align:center; font-weight:bold; color:white;">Credit</td><td style="text-align:center; font-weight:bold; color:white;">Balance</td><td style="text-align:center; font-weight:bold; color:white;">Debit</td><td style="text-align:center; font-weight:bold; color:white;">Credit</td><td style="text-align:center; font-weight:bold; color:white;">Balance</td><td style="text-align:center; font-weight:bold; color:white;">Debit</td><td style="text-align:center; font-weight:bold; color:white;">Credit</td><td style="text-align:center; font-weight:bold; color:white;">Balance</td></tr>'
    employee = frappe.get_all("Employee",["name","employee_name"])
    op_credit = 0
    op_debit = 0
    total_op_debit = 0
    total_op_credit = 0
    t_c_credit = 0
    t_p_credit = 0
    t_c_debit = 0
    t_p_debit = 0
    t_op_debit =0
    t_op_credit = 0
    t_op_balance = 0
    t_mov_debit = 0
    t_mov_credit = 0
    t_mov_balance = 0
    t_clo_debit = 0
    t_clo_credit = 0
    t_clo_balance = 0
    for j in employee:
        gle = frappe.db.sql("""
        SELECT name, party, sum(debit) as debit_amount, sum(credit) as credit_amount
        FROM `tabGL Entry` 
        WHERE account = %s and posting_date < %s and is_opening = 'No'
        and party = %s and party_type ='Employee' and is_cancelled = 0
        """, (account,from_date,j.name), as_dict=True)
        for g in gle:
            if not g.debit_amount:
                g.debit_amount = 0
            if not g.credit_amount:
                g.credit_amount = 0
            t_p_debit += g.debit_amount
            t_p_credit += g.credit_amount
            balance_op = t_p_debit - t_p_credit
            sq = frappe.db.sql(""" select name,party,sum(debit_in_account_currency) as debit,sum(credit_in_account_currency) as credit from `tabGL Entry` where account = '%s' and party = '%s' and party_type = 'Employee' and posting_date between '%s' and '%s' and is_opening = 'No' and is_cancelled = 0 """%(account,j.name,from_date,to_date),as_dict=True)
            for i in sq:
                if not i.credit:
                    i.credit = 0
                if not i.debit:
                    i.debit = 0
                op_credit = g.credit_amount + i.credit
                op_debit = g.debit_amount + i.debit
                total_op_debit += i.debit
                total_op_credit += i.credit
                mo_balance = total_op_debit - total_op_credit
                t_c_credit += op_credit
                t_c_debit += op_debit
                balance_cl = t_c_debit - t_c_credit
                if g.debit_amount or g.credit_amount or i.credit or i.debit:
                    if op_debit-op_credit:
                        data += '<tr><td >%s</td><td>%s</td><td style = text-align:right>%s</td><td style = text-align:right>%s</td><td style = text-align:right>%s</td>'%(j.name,j.employee_name,fmt_money(g.debit_amount),fmt_money(g.credit_amount),fmt_money(g.debit_amount - g.credit_amount ))
                        data += '<td style = text-align:right>%s</td><td style = text-align:right>%s</td><td style = text-align:right>%s</td><td style = text-align:right>%s</td><td style = text-align:right>%s</td><td style = text-align:right>%s</td></tr>'%(fmt_money(i.debit),fmt_money(i.credit),fmt_money(i.debit - i.credit),fmt_money(op_debit),fmt_money(op_credit),fmt_money(op_debit-op_credit))	
            if op_debit-op_credit:
                t_op_debit += g.debit_amount
                t_op_credit += g.credit_amount
                t_op_balance = t_op_debit - t_op_credit
                t_mov_debit += i.debit
                t_mov_credit += i.credit
                t_mov_balance = t_mov_debit - t_mov_credit
                t_clo_debit += op_debit
                t_clo_credit += op_credit
                t_clo_balance =  t_clo_debit - t_clo_credit
    data += '<tr style="text-align:right; font-weight:bold;"><td colspan = 2 style="text-align:center; font-weight:bold;">Total</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (fmt_money(t_op_debit), fmt_money(t_op_credit),fmt_money(t_op_balance),fmt_money(t_mov_debit), fmt_money(t_mov_credit), fmt_money(t_mov_balance),fmt_money(t_clo_debit), fmt_money(t_clo_credit),fmt_money(t_clo_balance))
    data += '</table>'
    return data

@frappe.whitelist()
def return_account_summary_total(from_date,to_date,account):
    data = '<table border="1" style="border-collapse: collapse; width: 100%;">'
    data += '<tr style="background-color: #D9E2ED;"><td colspan="1"><b></b></td><td colspan="1"><b></b></td>'
    data += '<td colspan="2" style="text-align:center;"><b>Opening</b></td>'
    data += '<td colspan="2" style="text-align:center;"><b>Movement</b></td>'
    data += '<td colspan="2" style="text-align:center;"><b>Closing</b></td></tr>'
    
    data += '<tr style="background-color: #e35310; color: white;">'
    data += '<td style="text-align:center; font-weight:bold; color:white;">Account Code</td>'
    data += '<td style="text-align:center; font-weight:bold; color:white;">Account Name</td>'
    data += '<td style="text-align:center; font-weight:bold; color:white;">Debit</td>'
    data += '<td style="text-align:center; font-weight:bold; color:white;">Credit</td>'
    data += '<td style="text-align:center; font-weight:bold; color:white;">Debit</td>'
    data += '<td style="text-align:center; font-weight:bold; color:white;">Credit</td>'
    data += '<td style="text-align:center; font-weight:bold; color:white;">Debit</td>'
    data += '<td style="text-align:center; font-weight:bold; color:white;">Credit</td></tr>'
    main_account = frappe.db.get_value("Account", account, ["name", "account_number"], as_dict=True)
    if not main_account:
        frappe.throw(f"Account '{account}' not found")
    def get_all_accounts(account_code):
        accounts = frappe.db.get_all("Account", filters={"parent_account": account_code, "disabled": 0}, fields=["name", "account_number"])
        all_accounts = accounts[:]
        for acc in accounts:
            child_accounts = get_all_accounts(acc.name)  # Recursively fetch child accounts
            all_accounts.extend(child_accounts)
        return all_accounts

    accounts = [main_account] + get_all_accounts(account)
    # accounts = frappe.db.get_all("Account", filters={"parent_account": account, "disabled": 0}, fields=["name","account_number"])
    total_op_debit = total_op_credit = t_c_credit = t_p_credit = t_c_debit = t_p_debit = 0
    for acc in accounts:
        gle = frappe.db.sql("""
        SELECT sum(debit) as debit_amount, sum(credit) as credit_amount
        FROM `tabGL Entry` 
        WHERE account = %s and posting_date < %s and is_opening = 'No'
        and is_cancelled = 0
        """, (acc.name,from_date), as_dict=True)
        for g in gle:
            g.debit_amount = g.debit_amount or 0
            g.credit_amount = g.credit_amount or 0
            t_p_debit += g.debit_amount
            t_p_credit += g.credit_amount
            
            sq = frappe.db.sql("""
            SELECT sum(debit_in_account_currency) as debit, sum(credit_in_account_currency) as credit 
            FROM `tabGL Entry` 
            WHERE account = %s AND posting_date BETWEEN %s AND %s AND is_opening = 'No' AND is_cancelled = 0
            """, (acc.name, from_date, to_date), as_dict=True)
            
            for i in sq:
                i.credit = i.credit or 0
                i.debit = i.debit or 0
                op_credit = g.credit_amount + i.credit
                op_debit = g.debit_amount + i.debit
                total_op_debit += i.debit
                total_op_credit += i.credit
                t_c_credit += op_credit
                t_c_debit += op_debit
                
                if g.debit_amount or g.credit_amount or i.credit or i.debit:
                    data += '<tr><td>%s</td><td>%s</td><td style="text-align:right">%s</td><td style="text-align:right">%s</td>' % (
                        acc.account_number or '-', acc.name, fmt_money(g.debit_amount), fmt_money(g.credit_amount))
                    data += '<td style="text-align:right">%s</td><td style="text-align:right">%s</td>' % (
                        fmt_money(i.debit), fmt_money(i.credit))
                    data += '<td style="text-align:right">%s</td><td style="text-align:right">%s</td></tr>' % (
                        fmt_money(op_debit), fmt_money(op_credit))
    
    data += '<tr style="text-align:right; font-weight:bold;"><td colspan = 2 style="text-align:center; font-weight:bold;">Total</td>'
    data += '<td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (
        fmt_money(t_p_debit), fmt_money(t_p_credit),
        fmt_money(total_op_debit), fmt_money(total_op_credit),
        fmt_money(t_c_debit), fmt_money(t_c_credit))
    data += '</table>'
    return data
#Consolidated ledger report
@frappe.whitelist()
def return_total_amt(from_date,to_date,account):
    acct = account.split(' - ')
    acc=''
    if len(acct) == 2:
        acc = acct[0]
    if len(acct) == 3:
        acc = f"{acct[0]} - {acct[1]}"
    if len(acct) == 4:
        acc = f"{acct[1]} - {acct[2]}"
    ac = '%'+acc+'%'
    data = '<table  border= 1px solid black width = 100%>'
    data += '<tr style = "background-color:#D9E2ED"><td colspan =1><b></b></td><td colspan =1 style = "text-align:center"><b>Opening</b></td><td colspan =2 style = "text-align:center"><b>Movement</b></td><td colspan =1 style = "text-align:center"><b>Closing</b></td></tr>'

    data += '<tr style = "background-color:#e35310;color:white"><td  style = "text-align:center;font-weight:bold;color:white">Company</td><td  style = "text-align:center;font-weight:bold;color:white">Balance</td><td  style = "text-align:center;font-weight:bold;color:white">Debit</td><td  style = "text-align:center;font-weight:bold;color:white">Credit</td><td  style = "text-align:center;font-weight:bold;color:white">Balance</td></tr>'
    op_credit = 0
    op_debit = 0
    total_op_debit = 0
    total_op_credit = 0
    t_c_credit = 0
    t_p_credit = 0
    t_c_debit = 0
    t_p_debit = 0
    
    li = []
    company = frappe.db.sql(""" select name from `tabCompany` where is_group = 1""",as_dict=1)
    for com in company:
        li.append(com.name)
        comp = frappe.db.get_list("Company",{"parent_company":com.name},['name'])
        for j in comp:
            li.append(j.name)
    for c in li:
        gle = frappe.db.sql("""select account, sum(debit_in_account_currency) as opening_debit, sum(credit_in_account_currency) as opening_credit from `tabGL Entry` where company = '%s'	and (posting_date < '%s' or (ifnull(is_opening, 'No') = 'Yes' and posting_date > '%s')) and account like '%s' and is_cancelled = 0  """%(c,from_date,to_date,ac),as_dict=True)
        for g in gle:
            if not g.opening_debit:
                g.opening_debit = 0
            if not g.opening_credit:
                g.opening_credit = 0
            t_p_debit += g.opening_debit
            t_p_credit += g.opening_credit
            balance_op = t_p_debit - t_p_credit
            data += '<tr><td>%s</td><td style = text-align:right >%s</td>'%(c,fmt_money(g.opening_debit - g.opening_credit,2))
            sq = frappe.db.sql(""" select company,sum(debit_in_account_currency) as debit,sum(credit_in_account_currency) as credit from `tabGL Entry` where company = '%s' and account like '%s' and posting_date between '%s' and '%s' and is_opening = 'No' and is_cancelled = 0 """%(c,ac,from_date,to_date),as_dict=True)
            for i in sq:
                if not i.credit:
                    i.credit = 0
                if not i.debit:
                    i.debit = 0
                op_credit = g.opening_credit + i.credit
                op_debit = g.opening_debit + i.debit
                total_op_debit += i.debit
                total_op_credit += i.credit
                t_c_credit += op_credit
                t_c_debit += op_debit
                balance_cl = t_c_debit - t_c_credit
                balance_move=total_op_debit-total_op_credit
                data += '<td style = text-align:right >%s</td><td style = text-align:right >%s</td><td style = text-align:right >%s</td></tr>'%(fmt_money(i.debit,2),fmt_money(i.credit,2),fmt_money(op_debit - op_credit,2))
    data += '<tr style = "text-align:right;font-weight:bold"><td style = "text-align:center;font-weight:bold">Total</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>'%(fmt_money(balance_op,2),fmt_money(total_op_debit,2),fmt_money(total_op_credit,2),fmt_money(balance_cl,2))
    # data += '<tr style = "text-align:right;font-weight:bold"><td style = "text-align:center;font-weight:bold">Balance</td><td colspan =3>%s</td><td colspan =3></td><td colspan=3>%s</td></tr>'%(fmt_money(balance_op),fmt_money(balance_cl))
    data += '</table>'
    return data


#Sales Person wise income report
@frappe.whitelist()
def get_sales_person(doc, name, from_date, to_date, company, sales_person_user):
    data = "<table width=100% border=1px solid black><tr style=background-color:#4682B4;font-size:8px><td colspan=1 ><b style=color:white; text-align:center;width:320px>SL.NO</b></td><td colspan=2  style=color:white><b style=color:white; text-align:center;>Date</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Invoice Number</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Customer Name</b></td><td colspan=1 style=color:white><b style=color:white; text-align:center;>LPO NO</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Gross Amount</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Discount</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Ret. Amount</b></td><td  style=color:white><b style=color:white; text-align:center;>Net</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Collected</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Balance</b></td></tr>"

    if not doc.company_multiselect and not sales_person_user:
        sales_person = frappe.get_all("Sales Person", fields=["name"])
        j = 1
        for s in sales_person:
            company_set = set()
            prev_company_name = None
            sales = frappe.get_all("Sales Invoice", {'posting_date': ('between', (from_date, to_date)), 'sales_person_user': s.name,'docstatus':("not in",[0,2])}, ["name", "customer", "total", "discount_amount", "posting_date", "outstanding_amount", "po_no", "company"],order_by="posting_date")
            company_totals = {}  
            salesperson_totals = {} 

            if sales:
                data += '<tr><td colspan=12 style="border: 1px solid black; background-color:#EBF4FA; font:size:8px" ><b style=color:"red">%s</b></td></tr>' % (s.name)

            for i in sales:
                company_name = i["company"]
                if company_name not in company_set:
                    if prev_company_name:
                        data += f'<tr style = font-size:8px><td colspan=6 style="border: 1px solid black"><b>Total for {prev_company_name}</b></td><td colspan=1 style="border: 1px solid black"><b>{round(company_totals[prev_company_name]["total"],2)}</b></td><td colspan=1 style="border: 1px solid black"><b>{round(company_totals[prev_company_name]["discount_amount"],2)}</b></td><td colspan=1 style="border: 1px solid black"><b>{round(company_totals[prev_company_name]["convert_float"],2)}</b></td><td style="border: 1px solid black"><b>{round(company_totals[prev_company_name]["net_total"],2)}</b></td><td colspan=1 style="border: 1px solid black"><b>{round(company_totals[prev_company_name]["collected_total"],2)}</b></td><td colspan=1 style="border: 1px solid black"><b>{round(company_totals[prev_company_name]["balance"],2)}</b></td></tr>'
                    data += f'<tr><td colspan=12 style="border: 1px solid black; background-color:#EBF4FA;font-size:8px" ><b style=color:"red">{company_name}</b></td></tr>'
                    prev_company_name = company_name
                    company_set.add(company_name)

                net = (i.total - i.discount_amount)
                net_int = int(net)
                net_amount = (i.total - net)
                convert_float = int(net_amount)
                collected = (net - i.outstanding_amount)
                collected_int = int(collected)

                data += f'<tr style=font-size:8px width:100 height:50><td colspan=1 style="border: 1px solid black">{j}</td>'
                data += f'<td colspan=2 style="border: 1px solid black" nowrap>{i.posting_date.strftime("%d-%m-%Y")}</td>'
                data += f'<td colspan=1 style="border: 1px solid black" nowrap>{i.name}</td>'
                data += f'<td colspan=1 style="border: 1px solid black">{i.customer}</td>'
                data += f'<td colspan=1 style="border: 1px solid black">{i.po_no}</td>'
                data += f'<td colspan=1 style="border: 1px solid black">{i.total}</td>'
                data += f'<td colspan=1 style="border: 1px solid black">{i.discount_amount}</td>'
                data += f'<td colspan=1 style="border: 1px solid black">{convert_float}</td>'
                data += f'<td style="border: 1px solid black">{net_int}</td>'
                data += f'<td colspan=1 style="border: 1px solid black">{collected_int}</td>'
                data += f'<td colspan=1 style="border: 1px solid black">{i.outstanding_amount}</td></tr>'

    
                if company_name not in company_totals:
                    company_totals[company_name] = {
                        "total": i.total,
                        "discount_amount": i.discount_amount,
                        "net_total": net_int,
                        "convert_float": convert_float,
                        "collected_total": collected_int,
                        "balance": i.outstanding_amount
                    }
                else:
                    company_totals[company_name]["total"] += i.total
                    company_totals[company_name]["discount_amount"] += i.discount_amount
                    company_totals[company_name]["net_total"] += net_int
                    company_totals[company_name]["convert_float"] += convert_float
                    company_totals[company_name]["collected_total"] += collected_int
                    company_totals[company_name]["balance"] += i.outstanding_amount

                
                if s.name not in salesperson_totals:
                    salesperson_totals[s.name] = {
                        "total": i.total,
                        "discount_amount": i.discount_amount,
                        "net_total": net_int,
                        "convert_float": convert_float,
                        "collected_total": collected_int,
                        "balance": i.outstanding_amount
                    }
                else:
                    salesperson_totals[s.name]["total"] += i.total
                    salesperson_totals[s.name]["discount_amount"] += i.discount_amount
                    salesperson_totals[s.name]["net_total"] += net_int
                    salesperson_totals[s.name]["convert_float"] += convert_float
                    salesperson_totals[s.name]["collected_total"] += collected_int
                    salesperson_totals[s.name]["balance"] += i.outstanding_amount

                j += 1

            if prev_company_name:
                
                data += f'<tr style = font-size:8px><td colspan=6 style="border: 1px solid black"><b>Total for {prev_company_name}</b></td><td colspan=1 style="border: 1px solid black"><b>{round(company_totals[prev_company_name]["total"],2)}</b></td><td colspan=1 style="border: 1px solid black"><b>{round(company_totals[prev_company_name]["discount_amount"],2)}</b></td><td colspan=1 style="border: 1px solid black"><b>{round(company_totals[prev_company_name]["convert_float"],2)}</b></td><td style="border: 1px solid black"><b>{round(company_totals[prev_company_name]["net_total"],2)}</b></td><td colspan=1 style="border: 1px solid black"><b>{round(company_totals[prev_company_name]["collected_total"],2)}</b></td><td colspan=1 style="border: 1px solid black"><b>{round(company_totals[prev_company_name]["balance"],2)}</b></td></tr>'

            prev_company_name = None

            
            if s.name in salesperson_totals:
                data += f'<tr style= font-size:8px><td colspan=6 style="border: 1px solid black"><b>Total for {s.name}</b></td><td colspan=1 style="border: 1px solid black"><b>{salesperson_totals[s.name]["total"]}</b></td><td colspan=1 style="border: 1px solid black"><b>{salesperson_totals[s.name]["discount_amount"]}</b></td><td colspan=1 style="border: 1px solid black"><b>{salesperson_totals[s.name]["convert_float"]}</b></td><td style="border: 1px solid black"><b>{salesperson_totals[s.name]["net_total"]}</b></td><td colspan=1 style="border: 1px solid black"><b>{salesperson_totals[s.name]["collected_total"]}</b></td><td colspan=1 style="border: 1px solid black"><b>{salesperson_totals[s.name]["balance"]}</b></td></tr>'
  
    elif not doc.company_multiselect and sales_person_user:
        company = frappe.get_all("Company", fields=["name"])
        data += f'<tr><td colspan="12" style="border: 1px solid black; background-color:#EBF4FA"><b style="color:black;font-size:8px">{sales_person_user}</b></td></tr>'
        all_companies_total = {
            "total": 0,
            "discount_amount": 0,
            "convert_float": 0,
            "net_total": 0,
            "collected_total": 0,
            "balance": 0
        }

        for c in company:
            sales_invoice = frappe.get_all("Sales Invoice", {"sales_person_user": sales_person_user,'posting_date': ('between', (from_date, to_date)), 'company': c.name,'docstatus':("not in",[0,2])}, ["name", "customer", "total", "discount_amount", "posting_date", "outstanding_amount", "po_no"],order_by="posting_date")
            if sales_invoice:
                data += f'<tr><td colspan="12" style="border: 1px solid black; background-color:#EBF4FA"><b style="color:black;font-size:8px" height:25%>{c.name}</b></td></tr>'
                j = 1
                company_totals = {}
                
                for i in sales_invoice:
                    net = (i.total - i.discount_amount)
                    net_int = int(net)
                    net_amount = (i.total - net)
                    convert_float = int(net_amount)
                    collected = (net - i.outstanding_amount)
                    collected_int = int(collected)
                    data += f'<tr style = font-size:8px><td colspan="1" style="border: 1px solid black;" height:25%>{j}</td>'
                    data += f'<td colspan="2" style="border: 1px solid black;" height:25% nowrap>{i.posting_date.strftime("%d-%m-%Y")}</td>'
                    data += f'<td colspan="1" style="border: 1px solid black;" height:25%>{i.name}</td>'
                    data += f'<td colspan="1" style="border: 1px solid black;" height:25%>{i.customer}</td>'
                    data += f'<td colspan="1" style="border: 1px solid black;" height:25%>{i.po_no}</td>'
                    data += f'<td colspan="1" style="border: 1px solid black;" height:25%>{i.total}</td>'
                    data += f'<td colspan="1" style="border: 1px solid black;" height:25%>{i.discount_amount}</td>'
                    data += f'<td colspan="1" style="border: 1px solid black;" height:25%>{convert_float}</td>'
                    data += f'<td style="border: 1px solid black;">{net_int}</td>'
                    data += f'<td colspan="1" style="border: 1px solid black;">{collected_int}</td>'
                    data += f'<td colspan="1" style="border: 1px solid black;">{i.outstanding_amount}</td></tr>'
                    if c.name not in company_totals:
                        company_totals[c.name] = {
                            "total": i.total,
                            "discount_amount": i.discount_amount,
                            "net_total": net_int,
                            "convert_float": convert_float,
                            "collected_total": collected_int,
                            "balance": i.outstanding_amount
                        }
                    else:
                        company_totals[c.name]["total"] += i.total
                        company_totals[c.name]["discount_amount"] += i.discount_amount
                        company_totals[c.name]["net_total"] += net_int
                        company_totals[c.name]['convert_float'] += convert_float
                        company_totals[c.name]["collected_total"] += collected_int
                        company_totals[c.name]['balance'] += i.outstanding_amount
                    j += 1
                data += f'<tr style = font-size:8px>'
                data += f'<td colspan="6" style="border: 1px solid black;"><b>Total for {c.name}</b></td>'
                data += f'<td colspan="1" style="border: 1px solid black;"><b>{round(company_totals[c.name].get("total", 0),2)}</b></td>'
                data += f'<td colspan="1" style="border: 1px solid black;"><b>{round(company_totals[c.name].get("discount_amount", 0),2)}</b></td>'
                data += f'<td colspan="1" style="border: 1px solid black;"><b>{round(company_totals[c.name].get("convert_float", 0),2)}</b></td>'
                data += f'<td style="border: 1px solid black;"><b>{round(company_totals[c.name].get("net_total", 0),2)}</b></td>'
                data += f'<td colspan="1" style="border: 1px solid black;"><b>{round(company_totals[c.name].get("collected_total", 0),2)}</b></td>'
                data += f'<td colspan="1" style="border: 1px solid black;"><b>{round(company_totals[c.name].get("balance", 0),2)}</b></td></tr>'

                all_companies_total["total"] += company_totals[c.name].get("total", 0)
                all_companies_total["discount_amount"] += company_totals[c.name].get("discount_amount", 0)
                all_companies_total["convert_float"] += company_totals[c.name].get("convert_float", 0)
                all_companies_total["net_total"] += company_totals[c.name].get("net_total", 0)
                all_companies_total["collected_total"] += company_totals[c.name].get("collected_total", 0)
                all_companies_total["balance"] += company_totals[c.name].get("balance", 0)
        data += f'<tr style =font-size:8px>'
        data += f'<td colspan="6" style="border: 1px solid black;"><b>Total for {sales_person_user}</b></td>'
        data += f'<td colspan="1" style="border: 1px solid black;"><b>{round(all_companies_total["total"],2)}</b></td>'
        data += f'<td colspan="1" style="border: 1px solid black;"><b>{round(all_companies_total["discount_amount"],2)}</b></td>'
        data += f'<td colspan="1" style="border: 1px solid black;"><b>{round(all_companies_total["convert_float"],2)}</b></td>'
        data += f'<td style="border: 1px solid black;"><b>{round(all_companies_total["net_total"],2)}</b></td>'
        data += f'<td colspan="1" style="border: 1px solid black;"><b>{round(all_companies_total["collected_total"],2)}</b></td>'
        data += f'<td colspan="1" style="border: 1px solid black;"><b>{round(all_companies_total["balance"],2)}</b></td>'
        data += '</tr>'

        
    elif doc.company_multiselect and not sales_person_user:
        for c in doc.company_multiselect:
            company_name_printed = False
            company_totals = {
                "total": 0,
                "discount_amount": 0,
                "convert_float": 0,
                "net_total": 0,
                "collected_total": 0,
                "balance": 0
            }
            
            sales_person = frappe.get_all("Sales Person", fields=["name"])
            for s in sales_person:
                sales_invoice = frappe.get_all("Sales Invoice", {"sales_person_user": s.name,'posting_date': ('between', (from_date, to_date)), 'company': c.company,'docstatus':("not in",[0,2])}, ["name", "customer", "total", "discount_amount", "posting_date", "outstanding_amount", "po_no"],order_by="posting_date")
                j = 1

                if sales_invoice:
                    data += '<tr style =font-size:8px><td colspan="12" style="border: 1px solid black; background-color:#EBF4FA"><b style="color:black">%s</b></td></tr>' % (s.name)
                    if not company_name_printed:
                        data += '<tr style =font-size:8px><td colspan="12" style="border: 1px solid black; background-color:#EBF4FA"><b style="color:black">%s</b></td></tr>' % (c.company)
                        company_name_printed = True

                    sales_person_totals = {
                        "total": 0,
                        "discount_amount": 0,
                        "convert_float": 0,
                        "net_total": 0,
                        "collected_total": 0,
                        "balance": 0
                    }

                    for i in sales_invoice:
                        net = (i.total - i.discount_amount)
                        net_int = int(net)
                        net_amount = (i.total - net)
                        convert_float = int(net_amount)
                        collected = (net - i.outstanding_amount)
                        collected_int = int(collected)
                        data += '<tr style =font-size:8px><td colspan="1" style="border: 1px solid black">%s</td><td colspan="2" style="border: 1px solid black" nowrap>%s</td><td colspan="1" style="border: 1px solid black">%s</td><td colspan="1" style="border: 1px solid black">%s</td><td colspan="1" style="border: 1px solid black">%s</td><td colspan="1" style="border: 1px solid black">%s</td><td colspan="1" style="border: 1px solid black">%s</td><td colspan="1" style="border: 1px solid black">%s</td><td style="border: 1px solid black">%s</td><td colspan="1" style="border: 1px solid black">%s</td><td colspan="1" style="border: 1px solid black">%s</td></tr>' % (j, i.posting_date.strftime("%d-%m-%Y"), i.name, i.customer, i.po_no, i.total, i.discount_amount, convert_float, net_int, collected_int, i.outstanding_amount)
                        j += 1
                        sales_person_totals["total"] += i.total
                        sales_person_totals["discount_amount"] += i.discount_amount
                        sales_person_totals["convert_float"] += convert_float
                        sales_person_totals["net_total"] += net_int
                        sales_person_totals["collected_total"] += collected_int
                        sales_person_totals["balance"] += i.outstanding_amount
                        
                        company_totals["total"] += i.total
                        company_totals["discount_amount"] += i.discount_amount
                        company_totals["convert_float"] += convert_float
                        company_totals["net_total"] += net_int
                        company_totals["collected_total"] += collected_int
                        company_totals["balance"] += i.outstanding_amount

                    data += '<tr style =font-size:8px><td colspan="6" style="border: 1px solid black"><b>Total for %s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td><td style="border: 1px solid black"><b>%s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td></tr>' % (s.name, round(sales_person_totals["total"],2), round(sales_person_totals["discount_amount"],2), round(sales_person_totals["convert_float"],2), round(sales_person_totals["net_total"],2), round(sales_person_totals["collected_total"],2), round(sales_person_totals["balance"],2))
        
            data += '<tr style =font-size:8px><td colspan="6" style="border: 1px solid black"><b>Total for %s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td><td style="border: 1px solid black"><b>%s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td></tr>' % (c.company, round(company_totals["total"],2), round(company_totals["discount_amount"],2), round(company_totals["convert_float"],2), round(company_totals["net_total"],2), round(company_totals["collected_total"],2), round(company_totals["balance"],2))

    else:
        salesperson_totals = {
            "total": 0,
            "discount_amount": 0,
            "convert_float": 0,
            "net_total": 0,
            "collected_total": 0,
            "balance": 0
        }
        
        for c in doc.company_multiselect:
            company_name_printed = False
            company_totals = {
                "total": 0,
                "discount_amount": 0,
                "convert_float": 0,
                "net_total": 0,
                "collected_total": 0,
                "balance": 0
            }
            
            sales_invoice = frappe.get_all("Sales Invoice", {"sales_person_user": sales_person_user, 'posting_date': ('between', (from_date, to_date)), 'company': c.company,'docstatus':("not in",[0,2])}, ["name", "customer", "total", "discount_amount", "posting_date", "outstanding_amount", "po_no"],order_by="posting_date")
            j = 1
            
            if sales_invoice:
                if not company_name_printed:
                    data += '<tr style =font-size:8px><td colspan="12" style="border: 1px solid black; background-color:#EBF4FA"><b style="color:black">%s</b></td></tr>' % (c.company)
                    company_name_printed = True
                
                data += '<tr style =font-size:8px><td colspan="12" style="border: 1px solid black; background-color:#EBF4FA"><b style="color:black">%s</b></td></tr>' % (sales_person_user)
                
                for i in sales_invoice:
                    net = (i.total - i.discount_amount)
                    net_int = int(net)
                    net_amount = (i.total - net)
                    convert_float = int(net_amount)
                    collected = (net - i.outstanding_amount)
                    collected_int = int(collected)
                    data += '<tr style =font-size:8px><td colspan="1" style="border: 1px solid black">%s</td><td colspan="2" style="border: 1px solid black" nowrap>%s</td><td colspan="1" style="border: 1px solid black" nowrap>%s</td><td colspan="1" style="border: 1px solid black">%s</td><td colspan="1" style="border: 1px solid black">%s</td><td colspan="1" style="border: 1px solid black">%s</td><td colspan="1" style="border: 1px solid black">%s</td><td colspan="1" style="border: 1px solid black">%s</td><td style="border: 1px solid black">%s</td><td colspan="1" style="border: 1px solid black">%s</td><td colspan="1" style="border: 1px solid black">%s</td></tr>' % (j, i.posting_date.strftime("%d-%m-%Y"), i.name, i.customer, i.po_no, i.total, i.discount_amount, convert_float, net_int, collected_int, i.outstanding_amount)
                    j += 1

                    company_totals["total"] += i.total
                    company_totals["discount_amount"] += i.discount_amount
                    company_totals["convert_float"] += convert_float
                    company_totals["net_total"] += net_int
                    company_totals["collected_total"] += collected_int
                    company_totals["balance"] += i.outstanding_amount
                
                data += '<tr style =font-size:8px><td colspan="6" style="border: 1px solid black"><b>Total for %s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td><td style="border: 1px solid black"><b>%s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td></tr>' % (c.company, round(company_totals["total"],2), round(company_totals["discount_amount"],2), round(company_totals["convert_float"],2), round(company_totals["net_total"],2), round(company_totals["collected_total"],2), round(company_totals["balance"],2))
                
                salesperson_totals["total"] += company_totals["total"]
                salesperson_totals["discount_amount"] += company_totals["discount_amount"]
                salesperson_totals["convert_float"] += company_totals["convert_float"]
                salesperson_totals["net_total"] += company_totals["net_total"]
                salesperson_totals["collected_total"] += company_totals["collected_total"]
                salesperson_totals["balance"] += company_totals["balance"]
        
        data += '<tr style =font-size:8px><td colspan="6" style="border: 1px solid black"><b>Total for %s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td><td style="border: 1px solid black"><b>%s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td><td colspan="1" style="border: 1px solid black"><b>%s</b></td></tr>' % (sales_person_user, round(salesperson_totals["total"],2), round(salesperson_totals["discount_amount"],2), round(salesperson_totals["convert_float"],2), round(salesperson_totals["net_total"],2), round(salesperson_totals["collected_total"],2), round(salesperson_totals["balance"],2))
        
    data += '</table>'
    return data

#Accounts receivable report dashboard
@frappe.whitelist()
def receivable_report(doc):
    if doc.project:
        data = "<table width=100% border=1px solid black><tr style=background-color:#e35310;font-size:8px><td colspan=1 ><b style=color:white; text-align:center;width:320px>Date</b></td><td colspan=2  style=color:white><b style=color:white; text-align:center;>Referance No</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Voucher Type</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Particulars</b></td><td colspan=1 style=color:white><b style=color:white; text-align:center;>Invoice Amount</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Advance</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Retention</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Net Amount</b></td><td  style=color:white><b style=color:white; text-align:center;>Received</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Balance Due</b></td></tr>"
        pay = frappe.get_all("Payment Entry Reference",{"reference_name":doc.so_no},['parent'])
        for i in pay:
            pay_entry = frappe.get_all("Payment Entry",{"name":i.parent},["posting_date","paid_amount","remarks","status"],order_by="posting_date")
            for j in pay_entry:
                if j.status =="Submitted":
                    data += f'<tr><td colspan=1 style="border: 1px solid black; font-size:8px" nowrap>{j.posting_date.strftime("%d-%m-%Y")}</td><td colspan=2 style="border: 1px solid black; font-size:8px" nowrap>{i.parent}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >Payment Entry</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{j.remarks}</td><td colspan=1 style="border: 1px solid black; font-size:8px">-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{j.paid_amount}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td></tr>'
        si  = frappe.get_all("Sales Invoice",{"sales_order":doc.so_no},['name','posting_date','total','adv_amount','ret_amount','net_total_project'],order_by="posting_date")
        for k in si:
            data += f'<tr><td colspan=1 style="border: 1px solid black; font-size:8px" nowrap>{k.posting_date.strftime("%d-%m-%Y")}</td><td colspan=2 style="border: 1px solid black; font-size:8px" nowrap>{k.name}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >Sales Invoice</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px">{k.total}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{k.adv_amount}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{k.ret_amount}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{k.net_total_project}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td></tr>'
            si_pay= frappe.get_all("Payment Entry Reference",{"reference_name":k.name},['parent'])
            for s in si_pay:
                si_pay_entry = frappe.get_all("Payment Entry",{"name":s.parent},["posting_date","paid_amount","remarks","status"],order_by="posting_date")
                for v in si_pay_entry:
                    if v.status =="Submitted":
                        data += f'<tr><td colspan=1 style="border: 1px solid black; font-size:8px" nowrap>{v.posting_date.strftime("%d-%m-%Y")}</td><td colspan=2 style="border: 1px solid black; font-size:8px" nowrap>{s.parent}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >Payment Entry</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{v.remarks}</td><td colspan=1 style="border: 1px solid black; font-size:8px">-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{v.paid_amount}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td></tr>'
    else:
        data = "<table width=100% border=1px solid black><tr style=background-color:#e35310;font-size:10px;text-align:center><td colspan=1 ><b style=color:white; text-align:center;width:320px>Sales Order Documents Not Avaliable</b></td></tr>"
    data += '</table>'
    return data

#Accounts ledger summary
@frappe.whitelist()
def get_accounts_ledger(doc):
    total_amount = 0
    total_paid = 0
    total_credit_note = 0
    total_outstanding = 0
    total_0_30 = 0
    total_31_60 = 0
    total_61_90 = 0
    total_91_above = 0
    sales_invoices = frappe.get_all("Sales Invoice", {'company': doc.company, 'customer': doc.customer}, ['posting_date', 'name', 'total', 'outstanding_amount'],order_by="posting_date")
    if sales_invoices:
        data = "<table width=100% border=1px solid black><tr style=background-color:#e35310;font-size:8px><td colspan=1 ><b style=color:white; text-align:center;width:320px>Invoice No</b></td><td colspan=2  style=color:white><b style=color:white; text-align:center;>Date</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Age</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Amount QR</b></td><td colspan=1 style=color:white><b style=color:white; text-align:center;>Paid QR</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Credit Note QR</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Outstanding QR</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>0-30</b></td><td  style=color:white><b style=color:white; text-align:center;>31-60</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>61-90</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>91-Above</b></td></tr>"
        for i in sales_invoices:
            days = date_diff(doc.from_date, i.posting_date)
            if 0 <= days <= 120:
                total_amount += i.total
                total_outstanding += i.outstanding_amount
                data += f'<tr><td colspan=1 style="border: 1px solid black; font-size:8px" >{i.name}</td><td colspan=2 style="border: 1px solid black; font-size:8px" >{i.posting_date.strftime("%d-%m-%Y")}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{days}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{i.total}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{i.outstanding_amount}</td>'
                if 0 <= days <= 30:
                    total_0_30 += i.total
                    data += f'<td colspan=1 style="border: 1px solid black; font-size:8px" >{i.total}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td>'
                if 31 <= days <= 60:
                    total_31_60 += i.total
                    data += f'<td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{i.total}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td>'
                if 61 <= days <= 90:
                    total_61_90 += i.total
                    data += f'<td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{i.total}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td>'
                if 91 <= days <= 120:
                    total_91_above += i.total
                    data += f'<td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{i.total}</td>'					
        data += '</tr>'
        data += f'<tr><td colspan=4 style="border: 1px solid black; font-size:8px" >Total</td><td style="border: 1px solid black; font-size:8px" >{total_amount}</td><td style="border: 1px solid black; font-size:8px" >-</td><td style="border: 1px solid black; font-size:8px" >-</td><td style="border: 1px solid black; font-size:8px" >{total_outstanding}</td><td style="border: 1px solid black; font-size:8px" >{total_0_30}</td><td style="border: 1px solid black; font-size:8px" >{total_31_60}</td><td style="border: 1px solid black; font-size:8px" >{total_61_90}</td><td style="border: 1px solid black; font-size:8px" >{total_91_above}</td></tr>'
        data += '</table>'
    else:
        data = "<table width=100% border=1px solid black><tr style=background-color:#e35310;font-size:10px;text-align:center><td colspan=1 ><b style=color:white; text-align:center;width:320px>Sales Invoice Documents Not Avaliable</b></td></tr>"

    return data


#Statement of account
@frappe.whitelist()
def statement_of_account(company, from_date, to_date, customer):
    data = ''

    data += "<table border='1px solid black' width='100%'><tr style='font-size:10px;background-color:#D3D3D3'><td width=10%><b>Date</b></td><td width=10%><b style='text-align:center;'>Voucher Type</b></td><td width=10%><b style='text-align:center'>Voucher No</b></td><td width=25%><b style='text-align:center'>Remarks</b></td><td width=10%><b style='text-align:center' width=10%>Debit(QAR)</b></td><td width=10%><b style='text-align:center'>Credit(QAR)</b></td><td width=10%><b style='text-align:center'>Balance(QAR)</b></td></tr>"
    if customer:
        gl_entry = frappe.db.sql("""select voucher_type,voucher_no,posting_date,sum(debit) as debit,sum(credit) as credit from `tabGL Entry` where company = %s and posting_date between %s and %s and is_cancelled = 0 and party = %s  and party_type ='Customer' group by voucher_no order by posting_date""", (company, from_date, to_date, customer), as_dict=True)
        gle = frappe.db.sql("""select sum(debit) as opening_debit, sum(credit) as opening_credit from `tabGL Entry` where company = '%s'	and (posting_date < '%s' or (ifnull(is_opening, 'No') = 'Yes' and posting_date >= '%s')) and party = '%s' and is_cancelled = 0 and party_type ='Customer' """%(company,from_date,to_date,customer),as_dict=True)
    else:
        gl_entry = frappe.db.sql("""select voucher_type,voucher_no,posting_date,sum(debit) as debit,sum(credit) as credit from `tabGL Entry` where company = %s and posting_date between %s and %s and is_cancelled = 0 group by voucher_no order by posting_date""", (company, from_date, to_date), as_dict=True)
        gle = frappe.db.sql("""select sum(debit) as opening_debit, sum(credit) as opening_credit from `tabGL Entry` where company = '%s'	and (posting_date < '%s' or (ifnull(is_opening, 'No') = 'Yes' and posting_date >= '%s')) and is_cancelled = 0  """%(company,from_date,to_date),as_dict=True)

    opening_balance = 0
    t_p_debit = 0
    t_p_credit = 0
    
    for g in gle:
        if not g.opening_debit:
            g.opening_debit = 0
        if not g.opening_credit:
            g.opening_credit = 0
        t_p_debit += g.opening_debit
        t_p_credit += g.opening_credit
        opening_balance = t_p_debit - t_p_credit
    data += f'<tr style="font-size:9px"><td colspan =6 style="text-align:right" width=85%><b>Opening Balance</b></td></td><td style="text-align:right" width=10%><b>{fmt_money(round(opening_balance,2))}</b></td></tr>'
    balance=opening_balance
    for i in gl_entry:
        balance += (i.debit -i.credit)
        if i.voucher_type == "Payment Entry":
            ref_no = frappe.db.get_value("Payment Entry",{"name":i.voucher_no},['reference_no'])
            if ref_no:
                check_no = ref_no
            else:
                check_no = ''
        else:
            check_no = ''
        if i.voucher_type == "Sales Invoice":
            remarks = ''
            remark = frappe.db.get_all("Sales Invoice", filters={"name": i.voucher_no}, fields=['*'])

            if remark:
                remark_data = remark[0]
                dn = remark_data.get('delivery_note', '')
                po = remark_data.get('po_no', '')

                if dn and po:
                    remarks = f"DN No.{dn} & LPO No.{po}"
                elif dn:
                    remarks = f"DN No.{dn}"
                elif po:
                    remarks =f"LPO No.{po}"
        elif i.voucher_type == "Journal Entry":
            remarks = ''
            
            remark = frappe.db.get_value("Journal Entry", {"name": i.voucher_no}, "user_remark")
            cheque_no = frappe.db.get_value("Journal Entry", {"name": i.voucher_no}, "cheque_no")
            
            if cheque_no:
                po_no = frappe.db.get_value("Retention Invoice", {"name": cheque_no}, "po_no") or frappe.db.get_value("Advance Invoice", {"name": cheque_no}, "po_no")
            else:
                po_no = None
            
            if remark and cheque_no:
                remarks = f"LPO No.{po_no} {remark}" if po_no else ''
                check_no = cheque_no
            elif remark:
                remarks = remark
            elif cheque_no:
                remarks = f"LPO No.{po_no}" if po_no else ''
                check_no = cheque_no

        else:
            remarks = ''


        data += f'''
            <tr style="font-size:9px">
                <td width=10% nowrap>{i.posting_date.strftime("%d-%m-%Y")}</td>
                <td width=10%>{i.voucher_type}</td>
                <td width=10% nowrap>{i.voucher_no}</td>
                <td style="max-width: 150px; word-wrap: break-word; overflow-wrap: break-word; white-space: normal;">
    { remarks }{ check_no }"
</td>
                <td width=10% style="text-align:right">{fmt_money(round(i.debit, 2)) or "-"}</td>
                <td width=10% style="text-align:right">{fmt_money(round(i.credit, 2)) or "-"}</td>
                <td style="text-align:right" width=10%>{fmt_money(round(balance, 2))}</td>
            </tr>
            '''

    tp_credit=0
    tp_debit=0
    bal=0
    for i in gl_entry:
        tp_credit += i.credit 
        t_p_credit += i.credit
        tp_debit += i.debit
        t_p_debit += i.debit
        bal = tp_debit-tp_credit

    data += f'<tr style="font-size:9px"><td colspan=4 style="text-align:right"><b>Total</b></td><td style="text-align:right"><b>{fmt_money(round(tp_debit,2))}</b></td><td style="text-align:right"><b>{fmt_money(round(tp_credit,2))}</b></td><td style="text-align:right"><b></b></td></tr>'
    data += f'<tr style="font-size:9px"><td colspan =6 style="text-align:right" width=85%><b>Closing Balance</b></td></td><td style="text-align:right" width=10%><b>{fmt_money(round(balance,2))}</b></td></tr>'
    data += '</table>'
    return data


#Statement of account
@frappe.whitelist()
def supplier_statement_of_account(company, from_date, to_date, supplier):
    data = ''

    data += "<table border='1px solid black' width='100%'><tr style='font-size:10px;background-color:#D3D3D3'><td width=10%><b>Date</b></td><td width=10%><b style='text-align:center;'>Voucher Type</b></td><td width=10%><b style='text-align:center'>Voucher No</b></td><td width=25%><b style='text-align:center'>Remarks</b></td><td width=10%><b style='text-align:center' width=10%>Debit(QAR)</b></td><td width=10%><b style='text-align:center'>Credit(QAR)</b></td><td width=10%><b style='text-align:center'>Balance(QAR)</b></td></tr>"
    if supplier:
        gl_entry = frappe.db.sql("""select voucher_type,voucher_no,posting_date,sum(debit) as debit,sum(credit) as credit from `tabGL Entry` where company = %s and posting_date between %s and %s and is_cancelled = 0 and party = %s  and party_type ='Supplier' group by voucher_no order by posting_date""", (company, from_date, to_date, supplier), as_dict=True)
        gle = frappe.db.sql("""select sum(debit) as opening_debit, sum(credit) as opening_credit from `tabGL Entry` where company = '%s'	and (posting_date < '%s' or (ifnull(is_opening, 'No') = 'Yes' and posting_date >= '%s')) and party = '%s' and is_cancelled = 0 and party_type ='Supplier' """%(company,from_date,to_date,supplier),as_dict=True)
    else:
        gl_entry = frappe.db.sql("""select voucher_type,voucher_no,posting_date,sum(debit) as debit,sum(credit) as credit from `tabGL Entry` where company = %s and posting_date between %s and %s and is_cancelled = 0 group by voucher_no order by posting_date""", (company, from_date, to_date), as_dict=True)
        gle = frappe.db.sql("""select sum(debit) as opening_debit, sum(credit) as opening_credit from `tabGL Entry` where company = '%s'	and (posting_date < '%s' or (ifnull(is_opening, 'No') = 'Yes' and posting_date >= '%s')) and is_cancelled = 0  """%(company,from_date,to_date),as_dict=True)

    opening_balance = 0
    t_p_debit = 0
    t_p_credit = 0
    
    for g in gle:
        if not g.opening_debit:
            g.opening_debit = 0
        if not g.opening_credit:
            g.opening_credit = 0
        t_p_debit += g.opening_debit
        t_p_credit += g.opening_credit
        opening_balance = t_p_credit - t_p_debit
    data += f'<tr style="font-size:9px"><td colspan =6 style="text-align:right" width=85%><b>Opening Balance</b></td></td><td style="text-align:right" width=10%><b>{fmt_money(round(opening_balance,2))}</b></td></tr>'
    balance=opening_balance
    for i in gl_entry:
        balance += (i.credit-i.debit)
        if i.voucher_type == "Payment Entry":
            ref_no = frappe.db.get_value("Payment Entry",{"name":i.voucher_no},['reference_no'])
            if ref_no:
                check_no = ref_no
            else:
                check_no = ''
        else:
            check_no = ''
        if i.voucher_type == "Purchase Invoice":
            remarks = ''
            invoice = frappe.get_doc("Purchase Invoice",i.voucher_no)
            for k in invoice.items:
                dn = frappe.db.get_value("Purchase Receipt",{"name":k.purchase_receipt},["supplier_delivery_note"])
                po = k.purchase_order

                if dn and po:
                    remarks = f"DN No.{dn} & PO No.{po}"
                elif dn:
                    remarks = f"DN No.{dn}"
                elif po:
                    remarks =f"PO No.{po}"
                else:
                    remarks =""
        elif i.voucher_type == "Journal Entry":
            remarks = ''
            remark = frappe.db.get_value("Journal Entry", {"name": i.voucher_no},['user_remark'])
            cheque_no = frappe.db.get_value("Journal Entry", {"name": i.voucher_no},['cheque_no'])
            if remark:
                remarks = remark
            elif cheque_no:
                check_no = cheque_no
            else:
                remarks = ""
                check_no = ""
        else:
            remarks = ''

        data += f'''
            <tr style="font-size:9px">
                <td width=10% nowrap>{i.posting_date.strftime("%d-%m-%Y")}</td>
                <td width=10%>{i.voucher_type}</td>
                <td width=10% nowrap>{i.voucher_no}</td>
                <td style="max-width: 150px; word-wrap: break-word; overflow-wrap: break-word; white-space: normal;">
    { remarks }{ check_no }
</td>
                <td width=10% style="text-align:right">{fmt_money(round(i.debit, 2)) or "-"}</td>
                <td width=10% style="text-align:right">{fmt_money(round(i.credit, 2)) or "-"}</td>
                <td style="text-align:right" width=10%>{fmt_money(round(balance, 2))}</td>
            </tr>
            '''

    tp_credit=0
    tp_debit=0
    bal=0
    for i in gl_entry:
        tp_credit += i.credit 
        t_p_credit += i.credit
        tp_debit += i.debit
        t_p_debit += i.debit
        bal = tp_debit-tp_credit

    data += f'<tr style="font-size:9px"><td colspan=4 style="text-align:right"><b>Total</b></td><td style="text-align:right"><b>{fmt_money(round(tp_debit,2))}</b></td><td style="text-align:right"><b>{fmt_money(round(tp_credit,2))}</b></td><td style="text-align:right"><b></b></td></tr>'
    data += f'<tr style="font-size:9px"><td colspan =6 style="text-align:right" width=85%><b>Closing Balance</b></td></td><td style="text-align:right" width=10%><b>{fmt_money(round(balance,2))}</b></td></tr>'
    data += '</table>'
    return data


#Accounts receivable ageing report
@frappe.whitelist()
def ageing_report(doc):
    in_amount = 0
    paid_amount = 0
    credit_note = 0
    out_amount = 0
    age_0_30 = 0
    age_31_60 = 0
    age_61_90 = 0
    age_91_120 = 0
    age_above_121 = 0
    paid = 0
    combined_data =[]
    data = "<table border='1px solid black' width='100%'><tr style='font-size:12px'><td width=10%><b>Posting Date</b></td><td width=10%><b style='text-align:center;'>Voucher No</b></td><td width=10%><b style='text-align:center'>Customer LPO</b></td><td width=10%><b style='text-align:center'>Invoiced Amount</b></td><td width=10%><b style='text-align:center'>Paid Amount</b></td><td width=10%><b style='text-align:center'>Credit Note</b></td><td width=10%><b style='text-align:center'>Outstanding Amount</b></td><td width=5%><b style='text-align:center'>Age (Days)</b></td><td width=5%><b style='text-align:center'>0- 30</b></td><td width=5%><b style='text-align:center'>31-  60</b></td><td width=5%><b style='text-align:center'>61-  90</b></td><td width=5%><b style='text-align:center'>91-120</b></td><td width=5%><b style='text-align:center'>Above 121</b></td></tr>"
    for c in doc.company_multiselect:
        if doc.customer:
            si_list = frappe.db.sql(
                """SELECT * FROM `tabSales Invoice` WHERE company = %s and customer = %s and docstatus = 1 and is_return = 0 and invoice_type='Credit' and status !='Paid'  order by posting_date  ASC""",
                (c.company, doc.customer),
                as_dict=True
            )
            for i in si_list:
                result= frappe.db.sql("""
                    SELECT sum(grand_total) as total
                    FROM `tabSales Invoice` 
                    WHERE company = %s AND return_against = %s AND docstatus = 1
                """, (c.company, i.name))
                return_amount = result[0][0] if result and result[0][0] else 0
                
                result_doc = frappe.db.sql("""
                    SELECT name
                    FROM `tabSales Invoice` 
                    WHERE company = %s AND return_against = %s AND docstatus = 1
                """, (c.company, i.name), as_dict=True)
                
                pay_doc = []
                if result_doc:
                    pay_doc = frappe.db.sql("""
                        SELECT per.allocated_amount 
                        FROM `tabPayment Entry Reference` AS per
                        LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
                        WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
                    """, (c.company, result_doc[0]["name"]), as_dict=True)
                pay = frappe.db.sql("""
                    SELECT per.allocated_amount 
                    FROM `tabPayment Entry Reference` AS per
                    LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
                    WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
                """, (c.company, i.name), as_dict=True)
                value = sum(j["allocated_amount"] for j in pay) + sum(r["allocated_amount"] for r in pay_doc)

                jv = frappe.db.sql("""
                    SELECT credit_in_account_currency 
                    FROM `tabJournal Entry Account` AS per
                    LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
                    WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
                """, (c.company, i.name), as_dict=True)
                for k in jv:
                    value += k.credit_in_account_currency
                

                if value and return_amount:
                    outstanding = i.grand_total - value + return_amount
                elif value:
                    outstanding = i.grand_total - value
                elif return_amount:
                    outstanding = i.grand_total + return_amount
                else:
                    outstanding = i.grand_total
                
                out_amount += outstanding
                age = date_diff(today(), i.posting_date) if i.posting_date else 0

                if round(outstanding) != 0:
                    if value:
                        paid_amount += value
                    if return_amount:
                        credit_note += return_amount
                    in_amount += i.grand_total
                    if 0 <= age <= 30:
                        age_0_30 += outstanding
                    elif 31 <= age <= 60:
                        age_31_60 += outstanding
                    elif 61 <= age <= 90:
                        age_61_90 += outstanding
                    elif 91 <= age <= 120:
                        age_91_120 += outstanding
                    else:
                        age_above_121 += outstanding
                    combined_data.append({
                        'posting_date': i.posting_date,
                        'name': i.name,
                        'po_no': i.po_no if i.po_no else '-',
                        'grand_total': i.grand_total,
                        'paid_amount': value if value else 0,
                        'credit_note': return_amount if return_amount else 0,
                        'outstanding': outstanding if outstanding else 0,
                        'age': age,
                        'oustanding_0_30':outstanding if 0 <= age <= 30 else 0,
                        'oustanding_31_60':outstanding if 31 <= age <= 60 else 0,
                        'oustanding_61_90':outstanding if 61 <= age <= 90 else 0,
                        'oustanding_91_120':outstanding if 91 <= age <= 120 else 0,
                        'oustanding_above_121':outstanding if age > 120 else 0,
                    })

            sales = frappe.db.sql(
                """SELECT * FROM `tabSales Invoice` WHERE company = %s and customer = %s and docstatus = 1 and is_return = 1 and invoice_type='Credit' and status !='Paid' AND (return_against IS NULL OR return_against = '')  order by posting_date  ASC""",
                (c.company, doc.customer),
                as_dict=True
            )
            for a in sales:
                pay = frappe.db.sql(""" SELECT per.allocated_amount FROM `tabPayment Entry Reference` AS per
                LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
                WHERE per.reference_name = '%s' AND pe.docstatus = 1 and pe.company = '%s' """ % (a.name, c.company), as_dict=True)
                value = sum(j.allocated_amount for j in pay)

                jv = frappe.db.sql(""" SELECT credit_in_account_currency FROM `tabJournal Entry Account` AS per
                LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
                WHERE per.reference_name = '%s' AND pe.docstatus = 1 and pe.company = '%s'""" % (a.name, c.company), as_dict=True)
                value += sum(k.credit_in_account_currency for k in jv)

                outstanding = a.grand_total - value if value else a.grand_total
                out_amount += outstanding
                age = date_diff(today(), a.posting_date) if a.posting_date else 0

                if round(outstanding) != 0:
                    if value:
                        paid_amount += value
                    if a.grand_total:
                        credit_note += a.grand_total
                    in_amount += a.grand_total
                    if 0 <= age <= 30:
                        age_0_30 += outstanding
                    elif 31 <= age <= 60:
                        age_31_60 += outstanding
                    elif 61 <= age <= 90:
                        age_61_90 += outstanding
                    elif 91 <= age <= 120:
                        age_91_120 += outstanding
                    else:
                        age_above_121 += outstanding
                    combined_data.append({
                        'posting_date': a.posting_date,
                        'name': a.name,
                        'po_no':a.po_no if a.po_no else '-',
                        'grand_total': a.grand_total,
                        'paid_amount': value if value else 0,
                        'credit_note': a.grand_total if a.grand_total else 0,
                        'outstanding': outstanding if outstanding else 0,
                        'age': age,
                        'oustanding_0_30':outstanding if 0 <= age <= 30 else 0,
                        'oustanding_31_60':outstanding if 31 <= age <= 60 else 0,
                        'oustanding_61_90':outstanding if 61 <= age <= 90 else 0,
                        'oustanding_91_120':outstanding if 91 <= age <= 120 else 0,
                        'oustanding_above_121':outstanding if age > 120 else 0,
                    })

            payment = frappe.db.sql("""
                SELECT * FROM `tabPayment Entry` 
                WHERE company = %s AND party = %s AND docstatus = 1 
                AND payment_type = 'Receive' 
                ORDER BY posting_date ASC
            """, (c.company, doc.customer), as_dict=True)
            for v in payment:
                unallocated_amount = v.unallocated_amount
                paid_amount += unallocated_amount
                out_amount -= unallocated_amount
                age = date_diff(today(), v.posting_date)
                if unallocated_amount != 0:
                    if 0 <= age <= 30:
                        age_0_30 -= unallocated_amount
                    elif 31 <= age <= 60:
                        age_31_60 -= unallocated_amount
                    elif 61 <= age <= 90:
                        age_61_90 -= unallocated_amount
                    elif 91 <= age <= 120:
                        age_91_120 -= unallocated_amount
                    else:
                        age_above_121 -= unallocated_amount
                    combined_data.append({
                        'posting_date': v.posting_date,
                        'name': v.name,
                        'po_no': v.reference_no if v.reference_no else '-',
                        'grand_total': 0,
                        'paid_amount': unallocated_amount if unallocated_amount else 0,
                        'credit_note': 0,
                        'outstanding': -unallocated_amount if unallocated_amount else 0,
                        'age': age,
                        'oustanding_0_30':-unallocated_amount if 0 <= age <= 30 else 0,
                        'oustanding_31_60':-unallocated_amount if 31 <= age <= 60 else 0,
                        'oustanding_61_90':-unallocated_amount if 61 <= age <= 90 else 0,
                        'oustanding_91_120':-unallocated_amount if 91 <= age <= 120 else 0,
                        'oustanding_above_121':-unallocated_amount if age > 120 else 0,
                    })

            journal = frappe.db.sql("""
                SELECT per.credit_in_account_currency, per.debit_in_account_currency, pe.name, pe.posting_date, pe.user_remark,
                FROM `tabJournal Entry Account` AS per
                LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
                WHERE pe.company = %s AND per.account LIKE %s AND pe.docstatus = 1 
                AND party_type = 'Customer' AND party = %s AND per.reference_name IS NULL
            """, (c.company, '%Debtors -%', doc.customer), as_dict=True)
            for jour in journal:
                pay_journ = frappe.db.sql("""
                    SELECT per.allocated_amount 
                    FROM `tabPayment Entry Reference` AS per
                    LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
                    WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
                """, (c.company, jour.name), as_dict=True)
                value_journ = sum(jo["allocated_amount"] for jo in pay_journ)
                if value_journ:
                    value_journ = value_journ
                else:
                    value_journ = 0
                if jour.credit_in_account_currency:
                    journ_amount_credit = jour.credit_in_account_currency
                    paid_amount += journ_amount_credit - value_journ
                    in_amount -= journ_amount_credit - value_journ
                    out_amount -= journ_amount_credit - value_journ
                    age = date_diff(today(), jour.posting_date)
                    if 0 <= age <= 30:
                        age_0_30 -= (jour.credit_in_account_currency -value_journ)
                    elif 31 <= age <= 60:
                        age_31_60 -= (jour.credit_in_account_currency -value_journ)
                    elif 61 <= age <= 90:
                        age_61_90 -= (jour.credit_in_account_currency -value_journ)
                    elif 91 <= age <= 120:
                        age_91_120 -= (jour.credit_in_account_currency -value_journ)
                    else:
                        age_above_121 -= (jour.credit_in_account_currency -value_journ)
                    combined_data.append({
                        'posting_date': jour.posting_date,
                        'name': jour.name,
                        'po_no':jour.user_remark or '',
                        'grand_total': -jour.credit_in_account_currency,
                        'paid_amount': 0,
                        'credit_note': '-',
                        'outstanding': -(jour.credit_in_account_currency-value_journ) or 0,
                        'age': age,
                        'oustanding_0_30':-(jour.credit_in_account_currency-value_journ) if 0 <= age <= 30 else 0,
                        'oustanding_31_60':-(jour.credit_in_account_currency-value_journ) if 31 <= age <= 60 else 0,
                        'oustanding_61_90':-(jour.credit_in_account_currency-value_journ) if 61 <= age <= 90 else 0,
                        'oustanding_91_120':-(jour.credit_in_account_currency-value_journ) if 91 <= age <= 120 else 0,
                        'oustanding_above_121':-(jour.credit_in_account_currency-value_journ) if age > 120 else 0,
                    })
    
                elif jour.debit_in_account_currency:
                    journ_amount_debit = jour.debit_in_account_currency
                    in_amount += journ_amount_debit - value_journ
                    out_amount += journ_amount_debit - value_journ
                    age = date_diff(today(), jour.posting_date)

                    if 0 <= age <= 30:
                        age_0_30 += (jour.debit_in_account_currency -value_journ)
                    elif 31 <= age <= 60:
                        age_31_60 += (jour.debit_in_account_currency -value_journ)
                    elif 61 <= age <= 90:
                        age_61_90 += (jour.debit_in_account_currency -value_journ)
                    elif 91 <= age <= 120:
                        age_91_120 += (jour.debit_in_account_currency -value_journ)
                    else:
                        age_above_121 += (jour.debit_in_account_currency -value_journ)
    
                    combined_data.append({
                        'posting_date': jour.posting_date,
                        'name': jour.name,
                        'po_no':jour.user_remark or '',
                        'grand_total': jour.debit_in_account_currency,
                        'paid_amount':value_journ or 0,
                        'credit_note': 0,
                        'outstanding': (jour.debit_in_account_currency -value_journ) or 0,
                        'age': age,
                        'oustanding_0_30':jour.debit_in_account_currency -value_journ if 0 <= age <= 30 else 0,
                        'oustanding_31_60':jour.debit_in_account_currency -value_journ if 31 <= age <= 60 else 0,
                        'oustanding_61_90':jour.debit_in_account_currency -value_journ if 61 <= age <= 90 else 0,
                        'oustanding_91_120':jour.debit_in_account_currency -value_journ if 91 <= age <= 120 else 0,
                        'oustanding_above_121':jour.debit_in_account_currency -value_journ if age > 120 else 0,
                    })
    combined_data = sorted(combined_data, key=lambda x: x['posting_date'])
    for entry in combined_data:
        if entry['outstanding'] != 0:
            data += f"""<tr style='font-size:10px'>
            <td>{formatdate(entry['posting_date'],'dd-mm-yyyy')}</td>
            <td>{entry['name']}</td>
            <td>{entry['po_no']}</td>
            <td>{fmt_money(round(entry['grand_total'], 2))}</td>
            <td>{fmt_money(round(entry['paid_amount'], 2)) if entry['paid_amount']  else '-'}</td>
            <td>{fmt_money(entry['credit_note'], 2) if entry['credit_note'] else '-'}</td>
            <td>{fmt_money(round(entry['outstanding'], 2)) if entry['outstanding'] else 0}</td>
            <td>{entry['age']}</td>
            <td>{fmt_money(round(entry['oustanding_0_30'],2))}</td>
            <td>{fmt_money(round(entry['oustanding_31_60'],2))}</td>
            <td>{fmt_money(round(entry['oustanding_61_90'],2))}</td>
            <td>{fmt_money(round(entry['oustanding_91_120'],2))}</td>
            <td>{fmt_money(round(entry['oustanding_above_121'],2))}</td>
            </tr>"""
        

    data += f"<tr style='font-size:10px'><td width=10%></td><td width=10%></td><td width=10%><b>Total</b></td><td width=10%><b>{fmt_money(round(in_amount, 2))}</b></td><td width=10%><b>{fmt_money(round(paid_amount, 2))}</b></td><td width=10%><b>{fmt_money(round(credit_note, 2))}</b></td><td width=10%><b>{fmt_money(round(out_amount, 2))}</b></td><td width=5%></td><td width=5%><b>{fmt_money(round(age_0_30, 2))}</b></td><td width=5%><b>{fmt_money(round(age_31_60, 2))}</b></td><td width=5%><b>{fmt_money(round(age_61_90, 2))}</b></td><td width=5%><b>{fmt_money(round(age_91_120, 2))}</b></td><td width=5%><b>{fmt_money(round(age_above_121, 2))}</b></td></tr>"
    data += "</table>"
    return data

#Consolidated ledger report
@frappe.whitelist()
def return_total_amt_consolidate(from_date,to_date,account):
    acct = account.split(' - ')
    acc=''
    if len(acct) == 2:
        acc = acct[0]
    if len(acct) == 3:
        acc = f"{acct[0]} - {acct[1]}"
    if len(acct) == 4:
        acc = f"{acct[1]} - {acct[2]}"
    ac = '%'+acc+'%'
    data = '<table  border= 1px solid black width = 100%>'
    # data += '<tr style = "background-color:#D9E2ED"><td colspan =1><b></b></td><td colspan =1 style = "text-align:center"><b>Opening</b></td><td colspan =2 style = "text-align:center"><b>Movement</b></td><td colspan =1 style = "text-align:center"><b>Closing</b></td></tr>'
    data += '<tr style = "background-color:#D9E2ED;color:black"><td  style = "text-align:center;font-weight:bold;color:white"></td><td  colspan= 2 style = "text-align:center;font-weight:bold;color:black">Opening</td><td  colspan = 2 style = "text-align:center;font-weight:bold;color:black">Movement</td><td  colspan = 2 style = "text-align:center;font-weight:bold;color:black">Closing</td></tr>'
    data += '<tr style = "background-color:#e35310;color:white"><td  style = "text-align:center;font-weight:bold;color:white">Company</td><td  style = "text-align:center;font-weight:bold;color:white">Opening Debit</td><td  style = "text-align:center;font-weight:bold;color:white">Opening Credit</td><td  style = "text-align:center;font-weight:bold;color:white">Movement Debit</td><td  style = "text-align:center;font-weight:bold;color:white">Movement Credit</td><td  style = "text-align:center;font-weight:bold;color:white">Closing Debit</td><td  style = "text-align:center;font-weight:bold;color:white">Closing Credit</td></tr>'
    op_credit = 0
    op_debit = 0
    total_op_debit = 0
    total_op_credit = 0
    t_c_credit = 0
    t_p_credit = 0
    t_c_debit = 0
    t_p_debit = 0
    base_account =account.rsplit(" - ", 1)[0]
    li = []
    company = frappe.db.sql(""" select name from `tabCompany` where is_group = 1""",as_dict=1)
    for com in company:
        li.append(com.name)
        comp = frappe.db.get_list("Company",{"parent_company":com.name},['name'])
        for j in comp:
            li.append(j.name)
    for c in li:
        company_name = c
        abbr = frappe.db.get_value("Company", company_name, "abbr")
        account_name = f"{base_account} - {abbr}"
        gle = frappe.db.sql("""select account, sum(debit) as opening_debit, sum(credit) as opening_credit from `tabGL Entry` where company = '%s'	and (posting_date < '%s' or (ifnull(is_opening, 'No') = 'Yes' and posting_date > '%s')) and account = '%s' and is_cancelled = 0  """%(c,from_date,to_date,account_name),as_dict=True)
        # gle = frappe.db.sql("""select account, sum(debit) as opening_debit, sum(credit) as opening_credit from `tabGL Entry` where company = '%s'	and (posting_date < '%s' or (ifnull(is_opening, 'No') = 'Yes' and posting_date > '%s')) and account like '%s' and is_cancelled = 0  """%(c,from_date,to_date,ac),as_dict=True)
        for g in gle:
            if not g.opening_debit:
                g.opening_debit = 0
            if not g.opening_credit:
                g.opening_credit = 0
            t_p_debit += g.opening_debit
            t_p_credit += g.opening_credit
            balance_op = t_p_debit - t_p_credit
            data += '<tr><td>%s</td><td style = text-align:right >%s</td><td style = text-align:right >%s</td>'%(c,fmt_money(g.opening_debit) ,fmt_money(g.opening_credit))
            sq = frappe.db.sql(""" select company,sum(debit_in_account_currency) as debit,sum(credit_in_account_currency) as credit from `tabGL Entry` where company = '%s' and account like '%s' and posting_date between '%s' and '%s' and is_opening = 'No' and is_cancelled = 0 """%(c,ac,from_date,to_date),as_dict=True)
            for i in sq:
                if not i.credit:
                    i.credit = 0
                if not i.debit:
                    i.debit = 0
                op_credit = g.opening_credit + i.credit
                op_debit = g.opening_debit + i.debit
                total_op_debit += i.debit
                total_op_credit += i.credit
                t_c_credit += op_credit
                t_c_debit += op_debit
                balance_cl = t_c_debit - t_c_credit
                balance_move=total_op_debit-total_op_credit
                data += '<td style = text-align:right >%s</td><td style = text-align:right >%s</td><td style = text-align:right >%s</td><td style = text-align:right >%s</td></tr>'%(fmt_money(i.debit),fmt_money(i.credit),fmt_money(op_debit),fmt_money(op_credit))
    data += '<tr style = "text-align:right;font-weight:bold"><td style = "text-align:center;font-weight:bold">Total</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>'%(fmt_money(t_p_debit),fmt_money(t_p_credit),fmt_money(total_op_debit),fmt_money(total_op_credit),fmt_money(t_c_debit),fmt_money(t_c_credit))
    # data += '<tr style = "text-align:right;font-weight:bold"><td style = "text-align:center;font-weight:bold">Balance</td><td colspan =3>%s</td><td colspan =3></td><td colspan=3>%s</td></tr>'%(fmt_money(balance_op),fmt_money(balance_cl))
    data += '</table>'
    return data

#Receipt Report
@frappe.whitelist()
def receipt_report(doc):
    data = "<table border='1px solid black' width='100%'><tr><td style='text-align:center;'width='10%'><b>Posting Date</b></td><td style='text-align:center;'width='10%'><b>Voucher No</b></td><td style='text-align:center;'width='20%'><b>Party Name</b></td><td style='text-align:center;'width='20%'><b>Received Amount</b></td><td style='text-align:center;overflow-wrap: break-word;'width='10%'><b>Sales Person</b></td><td style='text-align:center;overflow-wrap: break-word;'width='30%'><b>Remarks</b></td></tr>"
    sales_person = []
    total = 0
    ind = 0
    sa = frappe.db.sql("""
        SELECT * 
        FROM `tabPayment Entry`
        WHERE company = %s AND posting_date BETWEEN %s AND %s AND payment_type ='Receive' AND party_type = 'Customer' AND docstatus = 1 order by posting_date  ASC
    """, (doc.company, doc.from_date, doc.to_date), as_dict=True)
    
    journal = frappe.db.sql("""
        SELECT * 
        FROM `tabJournal Entry`
        WHERE 
            company = %s 
            AND posting_date BETWEEN %s AND %s 
            AND docstatus = 1 
            AND voucher_type IN ("Bank Entry", "Cash Entry")
        ORDER BY posting_date ASC
    """, (doc.company, doc.from_date, doc.to_date), as_dict=True)
    
    for journ in journal:
        doc_journal = frappe.get_all("Journal Entry Account", {"parent": journ.name,"party_type":"customer"}, ["party", "credit_in_account_currency"])
        for c in doc_journal:
            if c.credit_in_account_currency>0:
                ind += 1
                total += c.credit_in_account_currency
                data += f"<tr><td width='10%' nowrap>{formatdate(journ.posting_date, 'dd-MM-yyyy')}</td><td width='10%'>{journ.name}</td><td width='20%'>{c.party}</td><td style='text-align:right;'width='20%'>{fmt_money(round(c.credit_in_account_currency, 2))}</td><td style='overflow-wrap: break-word;'width='10%'></td><td style='text-align:left;overflow-wrap: break-word;'width='30%'>{journ.remarks or ''}</td></tr>"

    for i in sa:
        document = frappe.get_all("Payment Entry Reference", {"parent": i.name}, ["reference_doctype", "reference_name"])
        if document:
            for j in document:
                if j.reference_doctype == "Sales Order":
                    sales_person = frappe.db.get_value("Sales Order", {"name": j.reference_name}, ["sales_person_user"])
                elif j.reference_doctype == "Sales Invoice":
                    sales_person = frappe.db.get_value("Sales Invoice", {"name": j.reference_name}, ["sales_person_user"])
        else:
            sales_person =''
        ind += 1
        total += i.received_amount
        data += f"<tr><td width='10%' nowrap>{formatdate(i.posting_date, 'dd-MM-yyyy')}</td><td width='10%'>{i.name}</td><td width='20%'>{i.party_name}</td><td style='text-align:right;'width='20%'>{fmt_money(round(i.received_amount, 2))}</td><td  style='text-align:left;overflow-wrap: break-word'width='10%'>{sales_person}</td><td style='text-align:left;overflow-wrap: break-word;'width='30%'>{i.remarks or ''}</td></tr>"
    data += f"<tr><td width='10%'> </td><td width='10%'> </td><td width='20%'>Total</td><td style='text-align:right;'width='20%'>{fmt_money(round(total, 2))}</td><td style='overflow-wrap: break-word;'width='10%'></td><td width='30%'></td></tr>"
    data += '</table>'
    return data

@frappe.whitelist()
def statement_of_account_project(company, from_date, to_date, customer=None, project=None):
    data = "<table border='1px solid black' width='100%'><tr style='font-size:10px;background-color:#D3D3D3'>"
    data += "<td width='5%'><b>Date</b></td>"
    data += "<td width='10%'><b>Voucher No</b></td>"
    data += "<td width='10%'><b>Voucher Type</b></td>"
    data += "<td width='15%'><b>Remarks</b></td>"
    data += "<td width='10%'><b>Invoice Amount</b></td>"
    data += "<td width='10%'><b>Advance</b></td>"
    data += "<td width='10%'><b>Retention</b></td>"
    data += "<td width='10%'><b>Net Amount</b></td>"
    data += "<td width='10%'><b>Received</b></td>"
    data += "<td width='10%'><b>Balance Due</b></td></tr>"

    if customer and project:
        adv_invoice = frappe.db.sql("""select * from `tabAdvance Invoice` where company = %s  and transaction_date between %s and %s and docstatus = 1 and customer = %s and project = %s  order by transaction_date""", (company, from_date, to_date, customer,project), as_dict=True)
        sales_invoice = frappe.db.sql("""select * from `tabSales Invoice` where company = %s  and posting_date between %s and %s and docstatus = 1 and customer = %s and project = %s  order by posting_date""", (company, from_date, to_date, customer,project), as_dict=True)
        payment_entry = frappe.db.sql("""select * from `tabPayment Entry` where company = %s  and posting_date between %s and %s and docstatus = 1 and party_type = 'Customer' and party = %s and project = %s  order by posting_date""", (company, from_date, to_date, customer,project), as_dict=True)
        ret_invoice = frappe.db.sql("""select * from `tabRetention Invoice` where company = %s  and transaction_date between %s and %s and docstatus = 1 and customer = %s and project = %s  order by transaction_date""", (company, from_date, to_date, customer,project), as_dict=True)
    elif customer and not project:
        adv_invoice = frappe.db.sql("""select * from `tabAdvance Invoice` where company = %s  and transaction_date between %s and %s and docstatus = 1 and customer = %s order by transaction_date""", (company, from_date, to_date, customer), as_dict=True)
        sales_invoice = frappe.db.sql("""select * from `tabsales Invoice` where company = %s  and posting_date between %s and %s and docstatus = 1 and customer = %s order by posting_date""", (company, from_date, to_date, customer), as_dict=True)
        payment_entry = frappe.db.sql("""select * from `tabPayment Entry` where company = %s  and posting_date between %s and %s and docstatus = 1 and party_type = 'Customer' and party = %s order by posting_date""", (company, from_date, to_date, customer), as_dict=True)
        ret_invoice = frappe.db.sql("""select * from `tabRetention Invoice` where company = %s  and transaction_date between %s and %s and docstatus = 1 and customer = %s order by transaction_date""", (company, from_date, to_date, customer), as_dict=True)
    elif project and not customer:
        adv_invoice = frappe.db.sql("""select * from `tabAdvance Invoice` where company = %s  and transaction_date between %s and %s and docstatus = 1 and project = %s  order by transaction_date""", (company, from_date, to_date,project), as_dict=True)
        sales_invoice = frappe.db.sql("""select * from `tabSales Invoice` where company = %s  and posting_date between %s and %s and docstatus = 1 and project = %s  order by posting_date""", (company, from_date, to_date,project), as_dict=True)
        payment_entry = frappe.db.sql("""select * from `tabPayment Entry` where company = %s  and posting_date between %s and %s and docstatus = 1 and party_type = 'Customer' and project = %s  order by posting_date""", (company, from_date, to_date,project), as_dict=True)
        ret_invoice = frappe.db.sql("""select * from `tabRetention Invoice` where company = %s  and transaction_date between %s and %s and docstatus = 1 and project = %s  order by transaction_date""", (company, from_date, to_date,project), as_dict=True)
    else:
        adv_invoice = frappe.db.sql("""select * from `tabAdvance Invoice` where company = %s  and transaction_date between %s and %s and docstatus = 1 order by transaction_date""", (company, from_date, to_date), as_dict=True)
        sales_invoice = frappe.db.sql("""select * from `tabSales Invoice` where company = %s  and posting_date between %s and %s and docstatus = 1 order by posting_date""", (company, from_date, to_date), as_dict=True)
        payment_entry = frappe.db.sql("""select * from `tabPayment Entry` where company = %s  and posting_date between %s and %s and docstatus = 1 and party_type = 'Customer' order by posting_date""", (company, from_date, to_date), as_dict=True)
        ret_invoice = frappe.db.sql("""select * from `tabRetention Invoice` where company = %s  and transaction_date between %s and %s and docstatus = 1 order by transaction_date""", (company, from_date, to_date), as_dict=True)
    combined_data = []

    # Process Advance Invoice data
    for i in adv_invoice:
        journal = frappe.db.get_value("Journal Entry", {"cheque_no": i.name,"docstatus":1}, "name")
        query = """
            SELECT 
                SUM(`per`.`allocated_amount`) AS total_allocated_amount
            FROM 
                `tabPayment Entry Reference` `per`
            JOIN 
                `tabPayment Entry` `pe` ON `per`.`parent` = `pe`.`name`
            WHERE 
                `pe`.`docstatus` = 1 AND `per`.`reference_name` = %s
        """

        # Execute the query with the 'journal' parameter
        result = frappe.db.sql(query, (journal,), as_dict=True)

        # Check and fetch the result
        if result:
            received = result[0].get('total_allocated_amount') if result[0].get('total_allocated_amount') else 0
        else:
            received = 0
        combined_data.append({
            'posting_date': i.transaction_date,
            'name': i.name,
            'type': "Advance Invoice",
            'remarks': f"Customer: {i.customer}, Order No: {i.sales_order}, Project: {i.project}",
            'invoice_amount': 0,
            'advance_amount': i.get('advance_amount1', 0),
            'retention': 0,
            'net_amount': 0,
            'received': received,
            'balance': i.get('advance_amount1', 0) - received
        })

    # Process Sales Invoice data
    for j in sales_invoice:
        remarks = []
        if j.delivery_note:
            remarks.append(f"DN No: {j.delivery_note}")
        if j.po_no:
            remarks.append(f"LPO No: {j.po_no}")
        
        query = """
            SELECT 
                SUM(`per`.`allocated_amount`) AS total_allocated_amount
            FROM 
                `tabPayment Entry Reference` `per`
            JOIN 
                `tabPayment Entry` `pe` ON `per`.`parent` = `pe`.`name`
            WHERE 
                `pe`.`docstatus` = 1 AND `per`.`reference_name` = %s
        """

        # Execute the query with the 'journal' parameter
        result = frappe.db.sql(query, (j.name,), as_dict=True)

        # Check and fetch the result
        if result:
            received = result[0].get('total_allocated_amount') if result[0].get('total_allocated_amount') else 0
        else:
            received = 0

        combined_data.append({
            'posting_date': j.posting_date,
            'name': j.name,
            'type': "Sales Invoice",
            'remarks': ", ".join(remarks),
            'invoice_amount': j.get('custom_total_invoice_amount', 0),
            'advance_amount': -j.get('adv_amount', 0),
            'retention': -j.get('ret_amount', 0),
            'net_amount': j.get('net_total_project', 0),
            'received': received,
            'balance': j.get('net_total_project', 0) - received
        })


    for a in ret_invoice:
        journal = frappe.db.get_value("Journal Entry", {"cheque_no": a.name,"docstatus":1}, "name")
        query = """
            SELECT 
                SUM(`per`.`allocated_amount`) AS total_allocated_amount
            FROM 
                `tabPayment Entry Reference` `per`
            JOIN 
                `tabPayment Entry` `pe` ON `per`.`parent` = `pe`.`name`
            WHERE 
                `pe`.`docstatus` = 1 AND `per`.`reference_name` = %s
        """

        # Execute the query with the 'journal' parameter
        result = frappe.db.sql(query, (journal,), as_dict=True)

        # Check and fetch the result
        if result:
            received = result[0].get('total_allocated_amount') if result[0].get('total_allocated_amount') else 0
        else:
            received = 0
        combined_data.append({
            'posting_date': a.transaction_date,
            'name': a.name,
            'type': "Retention Invoice",
            'remarks':a.po_no or '',
            'invoice_amount': 0,
            'advance_amount': 0,
            'retention':a.get('advance_amount1', 0),
            'net_amount': 0,
            'received': received,
            'balance': a.get('advance_amount1', 0) - received
        })

    # Sort combined data by posting date
    combined_data.sort(key=lambda x: x['posting_date'])

    # Build table rows
    total_invoice = total_advance = total_retention = total_net = total_received = total_balance = 0

    for entry in combined_data:
        data += f"<tr style='font-size:10px'>"
        data += f"<td nowrap>{formatdate(entry['posting_date'], 'dd-MM-yyyy')}</td>"
        data += f"<td nowrap>{entry['name']}</td>"
        data += f"<td nowrap>{entry['type']}</td>"
        data += f"<td>{entry['remarks']}</td>"
        data += f"<td style =text-align:right>{fmt_money(entry['invoice_amount']) if entry['invoice_amount'] else 0.00}</td>"
        data += f"<td style =text-align:right>{fmt_money(entry['advance_amount']) if entry['advance_amount'] else 0.00}</td>"
        data += f"<td style =text-align:right>{fmt_money(entry['retention']) if entry['retention'] else 0.00}</td>"
        data += f"<td style =text-align:right>{fmt_money(entry['net_amount']) if entry['net_amount'] else 0.00}</td>"
        data += f"<td style =text-align:right>{fmt_money(entry['received']) if entry['received'] else 0.00}</td>"
        data += f"<td style =text-align:right>{fmt_money(entry['balance']) if entry['balance'] else 0.00}</td>"
        data += "</tr>"

        total_invoice += entry['invoice_amount']
        total_advance += entry['advance_amount']
        total_retention += entry['retention']
        total_net += entry['net_amount']
        total_received += entry['received']
        total_balance += entry['balance']

    # Add totals row
    data += "<tr style='font-size:10px;background-color:#D3D3D3'>"
    data += f"<td colspan='4'><b>Total</b></td>"
    data += f"<td style =text-align:right>{fmt_money(total_invoice)}</td>"
    data += f"<td style =text-align:right>{fmt_money(total_advance)}</td>"
    data += f"<td style =text-align:right>{fmt_money(total_retention)}</td>"
    data += f"<td style =text-align:right>{fmt_money(total_net)}</td>"
    data += f"<td style =text-align:right>{fmt_money(total_received)}</td>"
    data += f"<td style =text-align:right>{fmt_money(total_balance)}</td>"
    data += "</tr>"

    data += "</table>"

    return data

@frappe.whitelist()
def ageing_report_test(doc):
    in_amount = 0
    paid_amount = 0
    credit_note = 0
    out_amount = 0
    age_0_30 = 0
    age_31_60 = 0
    age_61_90 = 0
    # age_91_120 = 0
    # age_above_121 = 0
    age_above_90=0
    paid = 0
    combined_data =[]
    data = "<table border='1px solid black' width='100%'><tr style='font-size:12px'><td width=10%><b>Posting Date</b></td><td width=10%><b style='text-align:center;'>Voucher No</b></td><td width=10%><b style='text-align:center'>Customer LPO</b></td><td width=10%><b style='text-align:center'>Invoiced Amount</b></td><td width=10%><b style='text-align:center'>Paid Amount</b></td><td width=10%><b style='text-align:center'>Credit Note</b></td><td width=10%><b style='text-align:center'>Outstanding Amount</b></td><td width=5%><b style='text-align:center'>Age (Days)</b></td><td width=5%><b style='text-align:center'>0- 30</b></td><td width=5%><b style='text-align:center'>31-  60</b></td><td width=5%><b style='text-align:center'>61-  90</b></td><td width=5%><b style='text-align:center'>Above 90</b></td></tr>"
    for c in doc.company_multiselect:
        if doc.customer:
            account_name = frappe.db.get_value(
                "Company",
                c.company,
                "default_receivable_account"
            )

            if not account_name:
                continue
            si_list = frappe.db.sql(
                """SELECT * FROM `tabSales Invoice` WHERE company = %s and customer = %s and docstatus = 1 and is_return = 0 and invoice_type='Credit' and status !='Paid' order by posting_date ASC""",
                (c.company, doc.customer),
                as_dict=True
            )
            for i in si_list:
                result= frappe.db.sql("""
                    SELECT sum(grand_total) as total
                    FROM `tabSales Invoice` 
                    WHERE company = %s AND return_against = %s AND docstatus = 1
                """, (c.company, i.name))
                return_amount = result[0][0] if result and result[0][0] else 0
                
                result_doc = frappe.db.sql("""
                    SELECT name
                    FROM `tabSales Invoice` 
                    WHERE company = %s AND return_against = %s AND docstatus = 1
                """, (c.company, i.name), as_dict=True)
                
                pay_doc = []
                jv_doc = []
                if result_doc:
                    pay_doc = frappe.db.sql("""
                        SELECT per.allocated_amount 
                        FROM `tabPayment Entry Reference` AS per
                        LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
                        WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
                    """, (c.company, result_doc[0]["name"]), as_dict=True)
                    jv_doc = frappe.db.sql("""
                        SELECT debit_in_account_currency 
                        FROM `tabJournal Entry Account` AS per
                        LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
                        WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
                    """, (c.company, result_doc[0]["name"]), as_dict=True)
                pay = frappe.db.sql("""
                    SELECT per.allocated_amount 
                    FROM `tabPayment Entry Reference` AS per
                    LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
                    WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
                """, (c.company, i.name), as_dict=True)
                value = sum(j["allocated_amount"] for j in pay) + sum(r["allocated_amount"] for r in pay_doc)
                # value = sum(j["allocated_amount"] for j in pay) + sum(r["allocated_amount"] for r in pay_doc) - sum(k["debit_in_account_currency"] for k in jv_doc)

                jv = frappe.db.sql("""
                    SELECT credit_in_account_currency 
                    FROM `tabJournal Entry Account` AS per
                    LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
                    WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
                """, (c.company, i.name), as_dict=True)
                for k in jv:
                    value += k.credit_in_account_currency
                

                if value and return_amount:
                    outstanding = i.grand_total + i.rounding_adjustment - value + return_amount + sum(k["debit_in_account_currency"] for k in jv_doc)
                elif value:
                    outstanding = i.grand_total + i.rounding_adjustment - value
                elif return_amount:
                    outstanding = i.grand_total + i.rounding_adjustment + return_amount
                else:
                    outstanding = i.grand_total + i.rounding_adjustment
                
                out_amount += outstanding
                age = date_diff(today(), i.posting_date) if i.posting_date else 0

                if round(outstanding) != 0:
                    if value:
                        paid_amount += value
                    if return_amount:
                        credit_note += return_amount
                    in_amount += i.grand_total + i.rounding_adjustment
                    if 0 <= age <= 30:
                        age_0_30 += outstanding
                    elif 31 <= age <= 60:
                        age_31_60 += outstanding
                    elif 61 <= age <= 90:
                        age_61_90 += outstanding
                    else:
                        age_above_90 += outstanding
                    # elif 91 <= age <= 120:
                    #     age_91_120 += outstanding
                    # else:
                    #     age_above_121 += outstanding
                    combined_data.append({
                        'posting_date': i.posting_date,
                        'name': i.name,
                        'po_no': i.po_no if i.po_no else '-',
                        'grand_total': i.grand_total + i.rounding_adjustment,
                        'paid_amount': value if value else 0,
                        'credit_note': return_amount if return_amount else 0,
                        'outstanding': outstanding if outstanding else 0,
                        'age': age,
                        'oustanding_0_30':outstanding if 0 <= age <= 30 else 0,
                        'oustanding_31_60':outstanding if 31 <= age <= 60 else 0,
                        'oustanding_61_90':outstanding if 61 <= age <= 90 else 0,
                        'oustanding_above_90':outstanding if age > 90 else 0,
                        # 'oustanding_91_120':outstanding if 91 <= age <= 120 else 0,
                        # 'oustanding_above_121':outstanding if age > 120 else 0,
                    })

            sales = frappe.db.sql(
                """SELECT * FROM `tabSales Invoice` WHERE company = %s and customer = %s and docstatus = 1 and is_return = 1 and invoice_type='Credit' and status !='Paid' AND (return_against IS NULL OR return_against = '')  order by posting_date  ASC""",
                (c.company, doc.customer),
                as_dict=True
            )
            for a in sales:
                pay = frappe.db.sql(""" SELECT per.allocated_amount FROM `tabPayment Entry Reference` AS per
                LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
                WHERE per.reference_name = '%s' AND pe.docstatus = 1 and pe.company = '%s' """ % (a.name, c.company), as_dict=True)
                value = sum(j.allocated_amount for j in pay)

                jv = frappe.db.sql(""" SELECT credit_in_account_currency FROM `tabJournal Entry Account` AS per
                LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
                WHERE per.reference_name = '%s' AND pe.docstatus = 1 and pe.company = '%s'""" % (a.name, c.company), as_dict=True)
                value += sum(k.credit_in_account_currency for k in jv)

                outstanding = a.grand_total - value if value else a.grand_total
                out_amount += outstanding
                age = date_diff(today(), a.posting_date) if a.posting_date else 0

                if round(outstanding) != 0:
                    if value:
                        paid_amount += value
                    if a.grand_total:
                        credit_note += a.grand_total
                    in_amount += a.grand_total
                    if 0 <= age <= 30:
                        age_0_30 += outstanding
                    elif 31 <= age <= 60:
                        age_31_60 += outstanding
                    elif 61 <= age <= 90:
                        age_61_90 += outstanding
                    else:
                        age_above_90 += outstanding
                    # elif 91 <= age <= 120:
                    #     age_91_120 += outstanding
                    # else:
                    #     age_above_121 += outstanding
                    combined_data.append({
                        'posting_date': a.posting_date,
                        'name': a.name,
                        'po_no':a.po_no if i.po_no else '-',
                        'grand_total': a.grand_total,
                        'paid_amount': value if value else 0,
                        'credit_note': a.grand_total if a.grand_total else 0,
                        'outstanding': outstanding if outstanding else 0,
                        'age': age,
                        'oustanding_0_30':outstanding if 0 <= age <= 30 else 0,
                        'oustanding_31_60':outstanding if 31 <= age <= 60 else 0,
                        'oustanding_61_90':outstanding if 61 <= age <= 90 else 0,
                        'oustanding_above_90':outstanding if age > 90 else 0,
                        # 'oustanding_91_120':outstanding if 91 <= age <= 120 else 0,
                        # 'oustanding_above_121':outstanding if age > 120 else 0,
                    })

            payment = frappe.db.sql("""
                SELECT * FROM `tabPayment Entry` 
                WHERE company = %s AND party = %s AND docstatus = 1 
                AND payment_type = 'Receive' 
                ORDER BY posting_date ASC
            """, (c.company, doc.customer), as_dict=True)
            for v in payment:
                unallocated_amount = v.unallocated_amount
                paid_amount += unallocated_amount
                out_amount -= unallocated_amount
                age = date_diff(today(), v.posting_date)
                if unallocated_amount != 0:
                    if 0 <= age <= 30:
                        age_0_30 -= unallocated_amount
                    elif 31 <= age <= 60:
                        age_31_60 -= unallocated_amount
                    elif 61 <= age <= 90:
                        age_61_90 -= unallocated_amount
                    else:
                        age_above_90 -= unallocated_amount
                    # elif 91 <= age <= 120:
                    #     age_91_120 -= unallocated_amount
                    # else:
                    #     age_above_121 -= unallocated_amount
                    combined_data.append({
                        'posting_date': v.posting_date,
                        'name': v.name,
                        'po_no': v.reference_no if v.reference_no else '-',
                        'grand_total': 0,
                        'paid_amount': unallocated_amount if unallocated_amount else 0,
                        'credit_note': 0,
                        'outstanding': -unallocated_amount if unallocated_amount else 0,
                        'age': age,
                        'oustanding_0_30':-unallocated_amount if 0 <= age <= 30 else 0,
                        'oustanding_31_60':-unallocated_amount if 31 <= age <= 60 else 0,
                        'oustanding_61_90':-unallocated_amount if 61 <= age <= 90 else 0,
                        'oustanding_above_90':-unallocated_amount if age > 90 else 0,
                        # 'oustanding_91_120':-unallocated_amount if 91 <= age <= 120 else 0,
                        # 'oustanding_above_121':-unallocated_amount if age > 120 else 0,
                    })

            # journal = frappe.db.sql("""
            #     SELECT per.credit_in_account_currency, per.debit_in_account_currency, pe.name, pe.posting_date ,pe.user_remark,pe.cheque_no
            #     FROM `tabJournal Entry Account` AS per
            #     LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
            #     WHERE pe.company = %s AND per.account LIKE %s AND pe.docstatus = 1
            #     AND party_type = 'Customer' AND party = %s AND per.reference_name IS NULL
            # """, (c.company, '%Debtors -%', doc.customer), as_dict=True)
            journal = frappe.db.sql("""
                SELECT per.credit_in_account_currency, per.debit_in_account_currency, pe.name, pe.posting_date ,pe.user_remark,pe.cheque_no
                FROM `tabJournal Entry Account` AS per
                LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
                WHERE pe.company = %s AND per.account = %s AND pe.docstatus = 1
                AND party_type = 'Customer' AND party = %s AND per.reference_name IS NULL
            """, (c.company,account_name, doc.customer), as_dict=True)
            for jour in journal:
                pay_journ = frappe.db.sql("""
                    SELECT per.allocated_amount 
                    FROM `tabPayment Entry Reference` AS per
                    LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
                    WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
                """, (c.company, jour.name), as_dict=True)
                value_journ = sum(jo["allocated_amount"] for jo in pay_journ)
                if value_journ:
                    value_journ = value_journ
                else:
                    value_journ = 0
                cheque_no = frappe.db.get_value("Retention Invoice",{"name":jour.cheque_no},["po_no"]) or frappe.db.get_value("Advance Invoice",{"name":jour.cheque_no},["po_no"])
                if cheque_no:
                    remark = cheque_no
                else:
                    remark = " "
                if jour.credit_in_account_currency:
                    journ_amount_credit = jour.credit_in_account_currency
                    paid_amount += journ_amount_credit - value_journ
                    in_amount -= journ_amount_credit - value_journ
                    out_amount -= journ_amount_credit - value_journ
                    age = date_diff(today(), jour.posting_date)
                    if 0 <= age <= 30:
                        age_0_30 -= (jour.credit_in_account_currency -value_journ)
                    elif 31 <= age <= 60:
                        age_31_60 -= (jour.credit_in_account_currency -value_journ)
                    elif 61 <= age <= 90:
                        age_61_90 -= (jour.credit_in_account_currency -value_journ)
                    else:
                        age_above_90 -= (jour.credit_in_account_currency -value_journ)
                    # elif 91 <= age <= 120:
                    #     age_91_120 -= (jour.credit_in_account_currency -value_journ)
                    # else:
                    #     age_above_121 -= (jour.credit_in_account_currency -value_journ)
                    combined_data.append({
                        'posting_date': jour.posting_date,
                        'name': jour.name,
                        'po_no':remark,
                        'grand_total': -jour.credit_in_account_currency,
                        'paid_amount': 0,
                        'credit_note': '-',
                        'outstanding': -(jour.credit_in_account_currency-value_journ) or 0,
                        'age': age,
                        'oustanding_0_30':-(jour.credit_in_account_currency-value_journ) if 0 <= age <= 30 else 0,
                        'oustanding_31_60':-(jour.credit_in_account_currency-value_journ) if 31 <= age <= 60 else 0,
                        'oustanding_61_90':-(jour.credit_in_account_currency-value_journ) if 61 <= age <= 90 else 0,
                        'oustanding_above_90':-(jour.credit_in_account_currency-value_journ) if age > 90 else 0,
                        # 'oustanding_91_120':-(jour.credit_in_account_currency-value_journ) if 91 <= age <= 120 else 0,
                        # 'oustanding_above_121':-(jour.credit_in_account_currency-value_journ) if age > 120 else 0,
                    })
    
                elif jour.debit_in_account_currency:
                    journ_amount_debit = jour.debit_in_account_currency
                    in_amount += journ_amount_debit - value_journ
                    out_amount += journ_amount_debit - value_journ
                    age = date_diff(today(), jour.posting_date)

                    if 0 <= age <= 30:
                        age_0_30 += (jour.debit_in_account_currency -value_journ)
                    elif 31 <= age <= 60:
                        age_31_60 += (jour.debit_in_account_currency -value_journ)
                    elif 61 <= age <= 90:
                        age_61_90 += (jour.debit_in_account_currency -value_journ)
                    else:
                        age_above_90 += (jour.debit_in_account_currency -value_journ)
                    # elif 91 <= age <= 120:
                    #     age_91_120 += (jour.debit_in_account_currency -value_journ)
                    # else:
                    #     age_above_121 += (jour.debit_in_account_currency -value_journ)
    
                    combined_data.append({
                        'posting_date': jour.posting_date,
                        'name': jour.name,
                        'po_no':remark,
                        'grand_total': jour.debit_in_account_currency,
                        'paid_amount':value_journ or 0,
                        'credit_note': 0,
                        'outstanding': (jour.debit_in_account_currency -value_journ) or 0,
                        'age': age,
                        'oustanding_0_30':jour.debit_in_account_currency -value_journ if 0 <= age <= 30 else 0,
                        'oustanding_31_60':jour.debit_in_account_currency -value_journ if 31 <= age <= 60 else 0,
                        'oustanding_61_90':jour.debit_in_account_currency -value_journ if 61 <= age <= 90 else 0,
                        'oustanding_above_90':jour.debit_in_account_currency -value_journ if age > 90 else 0,
                        # 'oustanding_91_120':jour.debit_in_account_currency -value_journ if 91 <= age <= 120 else 0,
                        # 'oustanding_above_121':jour.debit_in_account_currency -value_journ if age > 120 else 0,
                    })
    combined_data = sorted(combined_data, key=lambda x: x['posting_date'])
    for entry in combined_data:
        if entry['outstanding'] != 0:
            data += f"""<tr style='font-size:10px'>
            <td>{formatdate(entry['posting_date'],'dd-mm-yyyy')}</td>
            <td>{entry['name']}</td>
            <td>{entry['po_no']}</td>
            <td>{fmt_money(round(entry['grand_total'], 2))}</td>
            <td>{fmt_money(round(entry['paid_amount'], 2)) if entry['paid_amount']  else '-'}</td>
            <td>{fmt_money(entry['credit_note'], 2) if entry['credit_note'] else '-'}</td>
            <td>{fmt_money(round(entry['outstanding'], 2)) if entry['outstanding'] else 0}</td>
            <td>{entry['age']}</td>
            <td>{fmt_money(round(entry['oustanding_0_30'],2))}</td>
            <td>{fmt_money(round(entry['oustanding_31_60'],2))}</td>
            <td>{fmt_money(round(entry['oustanding_61_90'],2))}</td>
            <td>{fmt_money(round(entry['oustanding_above_90'],2))}</td>
            </tr>"""
            # <td>{fmt_money(round(entry['oustanding_91_120'],2))}</td>
            # <td>{fmt_money(round(entry['oustanding_above_121'],2))}</td>
            
        

    data += f"<tr style='font-size:10px'><td width=10%></td><td width=10%></td><td width=10%><b>Total</b></td><td width=10%><b>{fmt_money(round(in_amount, 2))}</b></td><td width=10%><b>{fmt_money(round(paid_amount, 2))}</b></td><td width=10%><b>{fmt_money(round(credit_note, 2))}</b></td><td width=10%><b>{fmt_money(round(out_amount, 2))}</b></td><td width=5%></td><td width=5%><b>{fmt_money(round(age_0_30, 2))}</b></td><td width=5%><b>{fmt_money(round(age_31_60, 2))}</b></td><td width=5%><b>{fmt_money(round(age_61_90, 2))}</b></td><td width=5%><b>{fmt_money(round(age_above_90, 2))}</b></td></tr>"
    data += "</table>"
    return data


@frappe.whitelist()
def ageing_report_test_test():
    doc = frappe.get_doc("Accounts Receivable Aging Report", "ARAR00865")
    in_amount = 0
    paid_amount = 0
    credit_note = 0
    out_amount = 0
    age_0_30 = 0
    age_31_60 = 0
    age_61_90 = 0
    # age_91_120 = 0
    # age_above_121 = 0
    age_above_90=0
    paid = 0
    combined_data =[]
    data = "<table border='1px solid black' width='100%'><tr style='font-size:12px'><td width=10%><b>Posting Date</b></td><td width=10%><b style='text-align:center;'>Voucher No</b></td><td width=10%><b style='text-align:center'>Customer LPO</b></td><td width=10%><b style='text-align:center'>Invoiced Amount</b></td><td width=10%><b style='text-align:center'>Paid Amount</b></td><td width=10%><b style='text-align:center'>Credit Note</b></td><td width=10%><b style='text-align:center'>Outstanding Amount</b></td><td width=5%><b style='text-align:center'>Age (Days)</b></td><td width=5%><b style='text-align:center'>0- 30</b></td><td width=5%><b style='text-align:center'>31-  60</b></td><td width=5%><b style='text-align:center'>61-  90</b></td><td width=5%><b style='text-align:center'>Above 90</b></td></tr>"
    for c in doc.company_multiselect:
        if doc.customer:
            si_list = frappe.db.sql(
                """SELECT * FROM `tabSales Invoice` WHERE company = %s and customer = %s and docstatus = 1 and is_return = 0 and invoice_type='Credit' and status !='Paid' order by posting_date ASC""",
                (c.company, doc.customer),
                as_dict=True
            )
            for i in si_list:
                result= frappe.db.sql("""
                    SELECT sum(grand_total) as total
                    FROM `tabSales Invoice` 
                    WHERE company = %s AND return_against = %s AND docstatus = 1
                """, (c.company, i.name))
                return_amount = result[0][0] if result and result[0][0] else 0
                
                result_doc = frappe.db.sql("""
                    SELECT name
                    FROM `tabSales Invoice` 
                    WHERE company = %s AND return_against = %s AND docstatus = 1
                """, (c.company, i.name), as_dict=True)
                
                pay_doc = []
                if result_doc:
                    pay_doc = frappe.db.sql("""
                        SELECT per.allocated_amount 
                        FROM `tabPayment Entry Reference` AS per
                        LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
                        WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
                    """, (c.company, result_doc[0]["name"]), as_dict=True)
                    jv_doc = frappe.db.sql("""
                        SELECT debit_in_account_currency 
                        FROM `tabJournal Entry Account` AS per
                        LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
                        WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
                    """, (c.company, result_doc[0]["name"]), as_dict=True)
                pay = frappe.db.sql("""
                    SELECT per.allocated_amount 
                    FROM `tabPayment Entry Reference` AS per
                    LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
                    WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
                """, (c.company, i.name), as_dict=True)
                value = sum(j["allocated_amount"] for j in pay) + sum(r["allocated_amount"] for r in pay_doc) - sum(k["debit_in_account_currency"] for k in jv_doc)

                jv = frappe.db.sql("""
                    SELECT credit_in_account_currency 
                    FROM `tabJournal Entry Account` AS per
                    LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
                    WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
                """, (c.company, i.name), as_dict=True)
                for k in jv:
                    value += k.credit_in_account_currency
                
                print(value)
                print(return_amount)
                if value and return_amount:
                    outstanding = i.grand_total + i.rounding_adjustment - value + return_amount
                elif value:
                    outstanding = i.grand_total + i.rounding_adjustment - value
                elif return_amount:
                    outstanding = i.grand_total + i.rounding_adjustment + return_amount
                else:
                    outstanding = i.grand_total + i.rounding_adjustment
                print(outstanding)
                out_amount += outstanding
                age = date_diff(today(), i.posting_date) if i.posting_date else 0

                if round(outstanding) != 0:
                    if value:
                        paid_amount += value
                    if return_amount:
                        credit_note += return_amount
                    in_amount += i.grand_total + i.rounding_adjustment
                    if 0 <= age <= 30:
                        age_0_30 += outstanding
                    elif 31 <= age <= 60:
                        age_31_60 += outstanding
                    elif 61 <= age <= 90:
                        age_61_90 += outstanding
                    else:
                        age_above_90 += outstanding
                    # elif 91 <= age <= 120:
                    #     age_91_120 += outstanding
                    # else:
                    #     age_above_121 += outstanding
                    combined_data.append({
                        'posting_date': i.posting_date,
                        'name': i.name,
                        'po_no': i.po_no if i.po_no else '-',
                        'grand_total': i.grand_total + i.rounding_adjustment,
                        'paid_amount': value if value else 0,
                        'credit_note': return_amount if return_amount else 0,
                        'outstanding': outstanding if outstanding else 0,
                        'age': age,
                        'oustanding_0_30':outstanding if 0 <= age <= 30 else 0,
                        'oustanding_31_60':outstanding if 31 <= age <= 60 else 0,
                        'oustanding_61_90':outstanding if 61 <= age <= 90 else 0,
                        'oustanding_above_90':outstanding if age > 90 else 0,
                        # 'oustanding_91_120':outstanding if 91 <= age <= 120 else 0,
                        # 'oustanding_above_121':outstanding if age > 120 else 0,
                    })

            # sales = frappe.db.sql(
            #     """SELECT * FROM `tabSales Invoice` WHERE company = %s and customer = %s and docstatus = 1 and is_return = 1 and invoice_type='Credit' and status !='Paid' AND (return_against IS NULL OR return_against = '')  order by posting_date  ASC""",
            #     (c.company, doc.customer),
            #     as_dict=True
            # )
            # for a in sales:
            #     pay = frappe.db.sql(""" SELECT per.allocated_amount FROM `tabPayment Entry Reference` AS per
            #     LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
            #     WHERE per.reference_name = '%s' AND pe.docstatus = 1 and pe.company = '%s' """ % (a.name, c.company), as_dict=True)
            #     value = sum(j.allocated_amount for j in pay)

            #     jv = frappe.db.sql(""" SELECT credit_in_account_currency FROM `tabJournal Entry Account` AS per
            #     LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
            #     WHERE per.reference_name = '%s' AND pe.docstatus = 1 and pe.company = '%s'""" % (a.name, c.company), as_dict=True)
            #     value += sum(k.credit_in_account_currency for k in jv)

            #     outstanding = a.grand_total - value if value else a.grand_total
            #     out_amount += outstanding
            #     age = date_diff(today(), a.posting_date) if a.posting_date else 0

            #     if round(outstanding) != 0:
            #         if value:
            #             paid_amount += value
            #         if a.grand_total:
            #             credit_note += a.grand_total
            #         in_amount += a.grand_total
            #         if 0 <= age <= 30:
            #             age_0_30 += outstanding
            #         elif 31 <= age <= 60:
            #             age_31_60 += outstanding
            #         elif 61 <= age <= 90:
            #             age_61_90 += outstanding
            #         else:
            #             age_above_90 += outstanding
            #         # elif 91 <= age <= 120:
            #         #     age_91_120 += outstanding
            #         # else:
            #         #     age_above_121 += outstanding
            #         combined_data.append({
            #             'posting_date': a.posting_date,
            #             'name': a.name,
            #             'po_no':a.po_no if i.po_no else '-',
            #             'grand_total': a.grand_total,
            #             'paid_amount': value if value else 0,
            #             'credit_note': a.grand_total if a.grand_total else 0,
            #             'outstanding': outstanding if outstanding else 0,
            #             'age': age,
            #             'oustanding_0_30':outstanding if 0 <= age <= 30 else 0,
            #             'oustanding_31_60':outstanding if 31 <= age <= 60 else 0,
            #             'oustanding_61_90':outstanding if 61 <= age <= 90 else 0,
            #             'oustanding_above_90':outstanding if age > 90 else 0,
            #             # 'oustanding_91_120':outstanding if 91 <= age <= 120 else 0,
            #             # 'oustanding_above_121':outstanding if age > 120 else 0,
            #         })

            # payment = frappe.db.sql("""
            #     SELECT * FROM `tabPayment Entry` 
            #     WHERE company = %s AND party = %s AND docstatus = 1 
            #     AND payment_type = 'Receive' 
            #     ORDER BY posting_date ASC
            # """, (c.company, doc.customer), as_dict=True)
            # for v in payment:
            #     unallocated_amount = v.unallocated_amount
            #     paid_amount += unallocated_amount
            #     out_amount -= unallocated_amount
            #     age = date_diff(today(), v.posting_date)
            #     if unallocated_amount != 0:
            #         if 0 <= age <= 30:
            #             age_0_30 -= unallocated_amount
            #         elif 31 <= age <= 60:
            #             age_31_60 -= unallocated_amount
            #         elif 61 <= age <= 90:
            #             age_61_90 -= unallocated_amount
            #         else:
            #             age_above_90 -= unallocated_amount
            #         # elif 91 <= age <= 120:
            #         #     age_91_120 -= unallocated_amount
            #         # else:
            #         #     age_above_121 -= unallocated_amount
            #         combined_data.append({
            #             'posting_date': v.posting_date,
            #             'name': v.name,
            #             'po_no': v.reference_no if v.reference_no else '-',
            #             'grand_total': 0,
            #             'paid_amount': unallocated_amount if unallocated_amount else 0,
            #             'credit_note': 0,
            #             'outstanding': -unallocated_amount if unallocated_amount else 0,
            #             'age': age,
            #             'oustanding_0_30':-unallocated_amount if 0 <= age <= 30 else 0,
            #             'oustanding_31_60':-unallocated_amount if 31 <= age <= 60 else 0,
            #             'oustanding_61_90':-unallocated_amount if 61 <= age <= 90 else 0,
            #             'oustanding_above_90':-unallocated_amount if age > 90 else 0,
            #             # 'oustanding_91_120':-unallocated_amount if 91 <= age <= 120 else 0,
            #             # 'oustanding_above_121':-unallocated_amount if age > 120 else 0,
            #         })

            # journal = frappe.db.sql("""
            #     SELECT per.credit_in_account_currency, per.debit_in_account_currency, pe.name, pe.posting_date 
            #     FROM `tabJournal Entry Account` AS per
            #     LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
            #     WHERE pe.company = %s AND per.account LIKE %s AND pe.docstatus = 1 
            #     AND party_type = 'Customer' AND party = %s AND per.reference_name IS NULL
            # """, (c.company, '%Debtors -%', doc.customer), as_dict=True)
            # for jour in journal:
            #     print(jour)
            #     pay_journ = frappe.db.sql("""
            #         SELECT per.allocated_amount 
            #         FROM `tabPayment Entry Reference` AS per
            #         LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
            #         WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
            #     """, (c.company, jour.name), as_dict=True)
            #     value_journ = sum(jo["allocated_amount"] for jo in pay_journ)
            #     if value_journ:
            #         value_journ = value_journ
            #     else:
            #         value_journ = 0
            #     if jour.credit_in_account_currency:
            #         journ_amount_credit = jour.credit_in_account_currency
            #         paid_amount += journ_amount_credit - value_journ
            #         in_amount -= journ_amount_credit - value_journ
            #         out_amount -= journ_amount_credit - value_journ
            #         age = date_diff(today(), jour.posting_date)
            #         if 0 <= age <= 30:
            #             age_0_30 -= (jour.credit_in_account_currency -value_journ)
            #         elif 31 <= age <= 60:
            #             age_31_60 -= (jour.credit_in_account_currency -value_journ)
            #         elif 61 <= age <= 90:
            #             age_61_90 -= (jour.credit_in_account_currency -value_journ)
            #         else:
            #             age_above_90 -= (jour.credit_in_account_currency -value_journ)
            #         # elif 91 <= age <= 120:
            #         #     age_91_120 -= (jour.credit_in_account_currency -value_journ)
            #         # else:
            #         #     age_above_121 -= (jour.credit_in_account_currency -value_journ)
            #         combined_data.append({
            #             'posting_date': jour.posting_date,
            #             'name': jour.name,
            #             'po_no':'-',
            #             'grand_total': -jour.credit_in_account_currency,
            #             'paid_amount': 0,
            #             'credit_note': '-',
            #             'outstanding': -(jour.credit_in_account_currency-value_journ) or 0,
            #             'age': age,
            #             'oustanding_0_30':-(jour.credit_in_account_currency-value_journ) if 0 <= age <= 30 else 0,
            #             'oustanding_31_60':-(jour.credit_in_account_currency-value_journ) if 31 <= age <= 60 else 0,
            #             'oustanding_61_90':-(jour.credit_in_account_currency-value_journ) if 61 <= age <= 90 else 0,
            #             'oustanding_above_90':-(jour.credit_in_account_currency-value_journ) if age > 90 else 0,
            #             # 'oustanding_91_120':-(jour.credit_in_account_currency-value_journ) if 91 <= age <= 120 else 0,
            #             # 'oustanding_above_121':-(jour.credit_in_account_currency-value_journ) if age > 120 else 0,
            #         })
    
            #     elif jour.debit_in_account_currency:
            #         journ_amount_debit = jour.debit_in_account_currency
            #         in_amount += journ_amount_debit - value_journ
            #         out_amount += journ_amount_debit - value_journ
            #         age = date_diff(today(), jour.posting_date)

            #         if 0 <= age <= 30:
            #             age_0_30 += (jour.debit_in_account_currency -value_journ)
            #         elif 31 <= age <= 60:
            #             age_31_60 += (jour.debit_in_account_currency -value_journ)
            #         elif 61 <= age <= 90:
            #             age_61_90 += (jour.debit_in_account_currency -value_journ)
            #         else:
            #             age_above_90 += (jour.debit_in_account_currency -value_journ)
            #         # elif 91 <= age <= 120:
            #         #     age_91_120 += (jour.debit_in_account_currency -value_journ)
            #         # else:
            #         #     age_above_121 += (jour.debit_in_account_currency -value_journ)
    
            #         combined_data.append({
            #             'posting_date': jour.posting_date,
            #             'name': jour.name,
            #             'po_no':'-',
            #             'grand_total': jour.debit_in_account_currency,
            #             'paid_amount':value_journ or 0,
            #             'credit_note': 0,
            #             'outstanding': (jour.debit_in_account_currency -value_journ) or 0,
            #             'age': age,
            #             'oustanding_0_30':jour.debit_in_account_currency -value_journ if 0 <= age <= 30 else 0,
            #             'oustanding_31_60':jour.debit_in_account_currency -value_journ if 31 <= age <= 60 else 0,
            #             'oustanding_61_90':jour.debit_in_account_currency -value_journ if 61 <= age <= 90 else 0,
            #             'oustanding_above_90':jour.debit_in_account_currency -value_journ if age > 90 else 0,
            #             # 'oustanding_91_120':jour.debit_in_account_currency -value_journ if 91 <= age <= 120 else 0,
            #             # 'oustanding_above_121':jour.debit_in_account_currency -value_journ if age > 120 else 0,
            #         })
    combined_data = sorted(combined_data, key=lambda x: x['posting_date'])
    for entry in combined_data:
        if entry['outstanding'] != 0:
            data += f"""<tr style='font-size:10px'>
            <td>{formatdate(entry['posting_date'],'dd-mm-yyyy')}</td>
            <td>{entry['name']}</td>
            <td>{entry['po_no']}</td>
            <td>{fmt_money(round(entry['grand_total'], 2))}</td>
            <td>{fmt_money(round(entry['paid_amount'], 2)) if entry['paid_amount']  else '-'}</td>
            <td>{fmt_money(entry['credit_note'], 2) if entry['credit_note'] else '-'}</td>
            <td>{fmt_money(round(entry['outstanding'], 2)) if entry['outstanding'] else 0}</td>
            <td>{entry['age']}</td>
            <td>{fmt_money(round(entry['oustanding_0_30'],2))}</td>
            <td>{fmt_money(round(entry['oustanding_31_60'],2))}</td>
            <td>{fmt_money(round(entry['oustanding_61_90'],2))}</td>
            <td>{fmt_money(round(entry['oustanding_above_90'],2))}</td>
            </tr>"""
            # <td>{fmt_money(round(entry['oustanding_91_120'],2))}</td>
            # <td>{fmt_money(round(entry['oustanding_above_121'],2))}</td>
            
        

    data += f"<tr style='font-size:10px'><td width=10%></td><td width=10%></td><td width=10%><b>Total</b></td><td width=10%><b>{fmt_money(round(in_amount, 2))}</b></td><td width=10%><b>{fmt_money(round(paid_amount, 2))}</b></td><td width=10%><b>{fmt_money(round(credit_note, 2))}</b></td><td width=10%><b>{fmt_money(round(out_amount, 2))}</b></td><td width=5%></td><td width=5%><b>{fmt_money(round(age_0_30, 2))}</b></td><td width=5%><b>{fmt_money(round(age_31_60, 2))}</b></td><td width=5%><b>{fmt_money(round(age_61_90, 2))}</b></td><td width=5%><b>{fmt_money(round(age_above_90, 2))}</b></td></tr>"
    data += "</table>"
    return "HI"

