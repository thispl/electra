from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from frappe import _
import frappe
from frappe.model.document import Document
from datetime import date, timedelta, datetime,time
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
    nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
import math
from frappe.utils import add_months, cint, flt, getdate, time_diff_in_hours,time_diff_in_seconds
import locale


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import openpyxl.styles as styles
from frappe.utils import flt, fmt_money


@frappe.whitelist()
def download():
    filename = 'Accounts Receivable Report'
    test = build_xlsx_response(filename)
    
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
        
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
         
    ws = wb.create_sheet(sheet_name, 0)
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15 
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15 
    ws.column_dimensions['G'].width = 15 
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.append(["Accounts Receivable Report"," "," "," "," "," "," ","","",""])
    header3=['Project: '+args.project]
    ws.append(header3 + [""] * 9)
    header4=['Reference:'+args.so_no]
    ws.append(header4 + [""] * 9)
    data2=so_value(args)
    header5=['SO Value:'+str(data2)]
    ws.append(header5 + [""] * 9)
    ws.append(['','','','','','','',"","",""])
    ws.append(['Date','Reference No','Voucher Type','Particulars','Invoice Amount','Advance','Retention','Net Amount','Received','Balance Due'])
    
    data1= get_data(args)
    for row in data1:
        ws.append(row)
    align_center = Alignment(horizontal='center',vertical='center')
    align_right = Alignment(horizontal='right',vertical='bottom')
    for header in ws.iter_rows(min_row=1 , max_row=4, min_col=1, max_col=10):
        for cell in header:
            cell.font = Font(bold=True)
            cell.alignment = align_center
    for header in ws.iter_rows(min_row=6 , max_row=6, min_col=1, max_col=10):
        for cell in header:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fgColor='D3D3D3', fill_type = "solid")
    border_thin = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'))
    header_range = ws['A1':ws.cell(row=len(get_data(args))+6, column=10).coordinate]
    for row in header_range:
        for cell in row:
            cell.border = border_thin
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10 )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10 )
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=10 )
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=10 )
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=10 )
    

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def get_data(args):
    data=[]
    if args.project:
        pay = frappe.get_all("Payment Entry Reference",{"reference_name":args.so_no},['parent'])
        for i in pay:
            pay_entry = frappe.get_all("Payment Entry",{"name":i.parent},["posting_date","paid_amount","remarks","status"],order_by="posting_date")
            for j in pay_entry:
                row1=[]
                if j.status =="Submitted":
                    row1+=[j.posting_date.strftime("%d-%m-%Y"),i.parent,'Payment Entry',j.remarks,'-',j.paid_amount,'-','-','-','-']               
                    data.append(row1)
        si  = frappe.get_all("Sales Invoice",{"sales_order":args.so_no},['name','posting_date','total','adv_amount','ret_amount','net_total_project'],order_by="posting_date")
        for k in si:
            row2=[]
            row2+=[k.posting_date.strftime("%d-%m-%Y"),k.name,'Sales Invoice','-',k.total,k.adv_amount,k.ret_amount,k.net_total_project,'-','-']
            data.append(row2)
            si_pay= frappe.get_all("Payment Entry Reference",{"reference_name":k.name},['parent'])
            for s in si_pay:
                si_pay_entry = frappe.get_all("Payment Entry",{"name":s.parent},["posting_date","paid_amount","remarks","status"],order_by="posting_by")
                for v in si_pay_entry:
                    row3=[]
                    if v.status =="Submitted":
                        row3+=[v.posting_date.strftime("%d-%m-%Y"),s.parent,'Payment Entry',v.remarks,'-','-','-',v.paid_amount,'-']
                        data.append(row3)
    return data

def so_value(args):
    so_value = frappe.db.get_value("Sales Order", {"name": args.so_no},["total"])
    return so_value

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'
