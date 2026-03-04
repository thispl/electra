import frappe
from frappe.model.document import Document
from frappe.utils import (cstr, add_days, date_diff, format_date, 
    nowdate, get_datetime_str, get_datetime, now_datetime, format_datetime, today, 
    time_diff_in_hours, time_diff_in_seconds, flt, fmt_money)
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.file_manager import get_file
from frappe.utils.background_jobs import enqueue

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, GradientFill, PatternFill
from openpyxl.utils import get_column_letter
from six import BytesIO, string_types
from datetime import date, timedelta, datetime, time
from frappe.utils import cstr, cint, getdate, get_last_day, get_first_day, add_days,date_diff


@frappe.whitelist()
def download():
    filename = 'Target_Vs_Achievement_Report.xlsx'
    build_xlsx_response(filename)

def make_xlsx(sheet_name=None):
    args = frappe.local.form_dict
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name if sheet_name else 'Sheet1'
    ws.append([args.get('company')])
    data = get_data(args)
    align_center = Alignment(horizontal='center', vertical='center',wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='center',wrap_text=True)
    white_font = Font(color="FFFFFF")
    border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )
    for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=8):
        for cell in rows:
            cell.font = Font(bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    
    # for row in data:
    # 	ws.append([data.index(row)])
    # 	if data.index(row)==0:
    # 		ws.merge_cells(start_row=int(data.index(row)+1), start_column=1, end_row=int(data.index(row)+2), end_column=4)

    # 	for rows in ws.iter_rows(min_row=data.index(row), max_row=data.index(row), min_col=1, max_col=8):
    # 		for cell in rows:
    # 			cell.alignment = align_center
    # 			cell.border = border
        
    # 	ws.append(row)
    current_row = 2 

    for row in data:
        ws.append(row)
        if 'MONTH:' in row:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
            ws.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=8)
            for rows in ws.iter_rows(min_row=current_row, max_row=current_row, min_col=1, max_col=8):
                for cell in rows:
                    cell.border = border
        if 'EXECUTIVE:' in row:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
            ws.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=8)
            for rows in ws.iter_rows(min_row=current_row, max_row=current_row, min_col=1, max_col=8):
                for cell in rows:
                    cell.border = border
            for rows in ws.iter_rows(min_row=current_row, max_row=current_row, min_col=7, max_col=8):
                for cell in rows:
                    cell.fill = PatternFill(fgColor='fcff42', fill_type = "solid")
        if 'Target Vs Achievement.' in row:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
            for rows in ws.iter_rows(min_row=current_row, max_row=current_row, min_col=1, max_col=8):
                for cell in rows:
                    cell.alignment = align_center
                    cell.border = border
                    cell.font = Font(bold=True,underline="single")
        if '_' in row:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
            for rows in ws.iter_rows(min_row=current_row, max_row=current_row, min_col=1, max_col=8):
                for cell in rows:
                    cell.alignment = align_center
        if 'MONTH  TO DATE   (MTD)' in row:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
            ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=8)
            for rows in ws.iter_rows(min_row=current_row, max_row=current_row, min_col=1, max_col=8):
                for cell in rows:
                    cell.alignment = align_center
                    cell.border = border
        if 'On Actual Target' in row:
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
            for rows in ws.iter_rows(min_row=current_row, max_row=current_row, min_col=1, max_col=8):
                for cell in rows:
                    cell.alignment = align_center
                    cell.border = border
        if 'Target' in row:
            for rows in ws.iter_rows(min_row=current_row, max_row=current_row, min_col=1, max_col=8):
                for cell in rows:
                    cell.alignment = align_center
                    cell.border = border
            for rows in ws.iter_rows(min_row=current_row, max_row=current_row, min_col=3, max_col=4):
                for cell in rows:
                    cell.fill = PatternFill(fgColor='00004d', fill_type = "solid")
                    cell.font = white_font
            for rows in ws.iter_rows(min_row=current_row, max_row=current_row, min_col=7, max_col=8):
                for cell in rows:
                    cell.fill = PatternFill(fgColor='00004d', fill_type = "solid")
                    cell.font = white_font
        if 'MTD (1)' in row:
            for rows in ws.iter_rows(min_row=current_row, max_row=current_row, min_col=1, max_col=8):
                for cell in rows:
                    cell.alignment = align_center
                    cell.border = border
            for rows in ws.iter_rows(min_row=current_row, max_row=current_row, min_col=3, max_col=4):
                for cell in rows:
                    cell.fill = PatternFill(fgColor='00004d', fill_type = "solid")
                    cell.font = white_font
            for rows in ws.iter_rows(min_row=current_row, max_row=current_row, min_col=7, max_col=8):
                for cell in rows:
                    cell.fill = PatternFill(fgColor='00004d', fill_type = "solid")
                    cell.font = white_font
        if '.' in row:
            for rows in ws.iter_rows(min_row=current_row, max_row=current_row, min_col=1, max_col=8):
                for cell in rows:
                    cell.alignment = align_right
                    cell.border = border
                    ws.row_dimensions[current_row].height = 50
            for rows in ws.iter_rows(min_row=current_row, max_row=current_row, min_col=2, max_col=2):
                for cell in rows:
                    cell.fill = PatternFill(fgColor='dcfeff', fill_type = "solid")	
            for rows in ws.iter_rows(min_row=current_row, max_row=current_row, min_col=6, max_col=6):
                for cell in rows:
                    cell.fill = PatternFill(fgColor='dcfeff', fill_type = "solid")
            for rows in ws.iter_rows(min_row=current_row, max_row=current_row, min_col=9, max_col=9):
                for cell in rows:
                    cell.font = white_font
        if '(3) = (2) - (1)    (4) = (3) / (1)' in row:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
            ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=8)
            for rows in ws.iter_rows(min_row=current_row, max_row=current_row, min_col=1, max_col=8):
                for cell in rows:
                    cell.alignment = align_center
                    cell.border = border

        
        current_row += 1
        
    
    

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file
def build_xlsx_response(filename):
    xlsx_file = make_xlsx(sheet_name=filename)  
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'
    
def set_column_widths(ws):
    column_widths = [10, 30] + [10] * 20  
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['J'].width = 20
        ws.column_dimensions['K'].width = 20


def get_data(args):
    
    if args.get('sales_person_name') and args.get('month') and args.get('fiscal_year') and args.get('company') and args.get('from_date') and args.get('to_date') and args.get('january'):
        data = []
        
        row1= ["", "", "", "", "", "MONTH:", f"{args.get('month')}-{args.get('fiscal_year')}"]
        row2 = ["", "", "", "", "", "EXECUTIVE:", args.get('sales_person_name')]
        row3= ["Target Vs Achievement."]
        # row4= ["",""]
        row5=["MONTH  TO DATE   (MTD)","","","","YEAR TO DATE    (YTD)"]
        row6=["","On Actual Target","","",""]
        row7=["Target","Actual","Variance","% var.","Target","Actual","Variance","% var."]
        row8=["MTD (1)","MTD  (2)", "MTD  (3)", "MTD  (4)", "YTD (5)", "YTD (6)","YTD (7)","YTD (8)"]
        
        target_query = """
        SELECT SUM(`tabTarget Detail`.target_amount) AS Target
        FROM `tabTarget Detail`
        WHERE `tabTarget Detail`.parent = %s AND `tabTarget Detail`.fiscal_year = %s
        """

        target_data = frappe.db.sql(target_query, (args.get('sales_person_name'), args.get('fiscal_year')), as_dict=True)

        sales_query = """
        SELECT SUM(`tabSales Invoice`.grand_total) AS Grand_Total
        FROM `tabSales Invoice`
        WHERE `tabSales Invoice`.company = %s
        AND `tabSales Invoice`.docstatus = 1
        AND `tabSales Invoice`.sales_person_user = %s
        AND `tabSales Invoice`.posting_date BETWEEN %s AND %s
        GROUP BY 
            `tabSales Invoice`.sales_person_user
        """
        sales_data = frappe.db.sql(sales_query, (args.get('company'), args.get('sales_person_name'), args.get('from_date'), args.get('to_date')), as_dict=True)
        sales_query1 = """
        SELECT SUM(`tabSales Invoice`.grand_total) AS Grand_Total1
        FROM `tabSales Invoice`
        WHERE `tabSales Invoice`.company = %s
        AND `tabSales Invoice`.docstatus = 1
        AND `tabSales Invoice`.sales_person_user = %s
        AND `tabSales Invoice`.posting_date BETWEEN %s AND %s
        GROUP BY 
            `tabSales Invoice`.sales_person_user
            """
        sales_data1 = frappe.db.sql(sales_query1, (args.get('company'), args.get('sales_person_name'),args.get('january'), args.get('to_date')), as_dict=True)

        if target_data and sales_data and sales_data1:
            d1=target_data[0].get('Target')/12 if target_data[0].get('Target') else 0
            d2=sales_data[0].get('Grand_Total')
            d3=d2-d1
            d4=(d3/d1)*100 if target_data[0].get('Target') else 0
            d5=target_data[0].get('Target') if target_data[0].get('Target') else 0
            d6=sales_data1[0].get('Grand_Total1')
            d7=d6-d5
            d8=(d7/d5)*100 if target_data[0].get('Target') else 0
            row9=[f"{d1:.2f}",f"{d2:.2f}",f"{d3:.2f}",f"{d4:.2f}%",d5,f"{d6:.2f}",f"{d7:.2f}",f"{d8:.2f}%","."]
        elif target_data and sales_data:
            d1=target_data[0].get('Target')/12 if target_data[0].get('Target') else 0
            d2=sales_data[0].get('Grand_Total')
            d3=d2-d1
            d4=(d3/d1)*100 if target_data[0].get('Target') else 0
            row9=[f"{d1:.2f}",f"{d2:.2f}",f"{d3:.2f}",f"{d4:.2f}%",target_data[0].get('Target'),"N/A","N/A","N/A","."]
        elif target_data:
            if target_data[0].get('Target'):
                d1=target_data[0].get('Target')/12
                row9=[f"{d1:.2f}","N/A","N/A","N/A",target_data[0].get('Target'),"N/A","N/A","N/A","."]
            else:
                row9=["N/A","N/A","N/A","N/A","N/A","N/A","N/A","N/A","."]

        row10=["(3) = (2) - (1)    (4) = (3) / (1)","","","(7) = (6) - (5)        (8) = (7) / (5)"]
        row_empty=["","_"]
        data.append(row1)
        data.append(row2)
        data.append(row3)
        # data.append(row4)
        data.append(row5)
        data.append(row6)
        data.append(row7)
        data.append(row8)
        data.append(row9)
        data.append(row10)
        data.append(row_empty)
     

    if args.get('month') and args.get('fiscal_year') and args.get('company') and args.get('from_date') and args.get('to_date') and args.get('january'):
        data = []
        sql_query = """SELECT `tabSales Invoice`.sales_person_user AS Name
                FROM `tabSales Invoice`
                LEFT JOIN `tabSales Person`
                ON `tabSales Invoice`.sales_person_user = `tabSales Person`.name
                WHERE `tabSales Person`.enabled = 1
                AND `tabSales Invoice`.company = %s
                AND `tabSales Invoice`.posting_date BETWEEN %s AND %s
                GROUP BY `tabSales Invoice`.sales_person_user
                ORDER BY `tabSales Person`.name
                """
        
        sales_person = frappe.db.sql(sql_query,(args.get('company'), args.get('from_date'), args.get('to_date')), as_dict=True)
        for i in sales_person:
            row1= ["", "", "", "", "", "MONTH:", f"{args.get('month')}-{args.get('fiscal_year')}"]
            row2 = ["", "", "", "", "", "EXECUTIVE:", i['Name']]
            row3= ["Target Vs Achievement."]
            # row4= ["","_"]
            row5=["MONTH  TO DATE   (MTD)","","","","YEAR TO DATE    (YTD)"]
            row6=["","On Actual Target","","",""]
            row7=["Target","Actual","Variance","% var.","Target","Actual","Variance","% var."]
            row8=["MTD (1)","MTD  (2)", "MTD  (3)", "MTD  (4)", "YTD (5)", "YTD (6)","YTD (7)","YTD (8)"]
            
            target_query = """
            SELECT SUM(`tabTarget Detail`.target_amount) AS Target
            FROM `tabTarget Detail`
            WHERE `tabTarget Detail`.parent = %s AND `tabTarget Detail`.fiscal_year = %s
            """

            target_data = frappe.db.sql(target_query, (i['Name'], args.get('fiscal_year')), as_dict=True)

            sales_query = """
            SELECT SUM(`tabSales Invoice`.grand_total) AS Grand_Total
            FROM `tabSales Invoice`
            WHERE `tabSales Invoice`.company = %s
            AND `tabSales Invoice`.docstatus = 1
            AND `tabSales Invoice`.sales_person_user = %s
            AND `tabSales Invoice`.posting_date BETWEEN %s AND %s
            GROUP BY 
                `tabSales Invoice`.sales_person_user
            """
            sales_data = frappe.db.sql(sales_query, (args.get('company'), i['Name'], args.get('from_date'), args.get('to_date')), as_dict=True)
            sales_query1 = """
            SELECT SUM(`tabSales Invoice`.grand_total) AS Grand_Total1
            FROM `tabSales Invoice`
            WHERE `tabSales Invoice`.company = %s
            AND `tabSales Invoice`.docstatus = 1
            AND `tabSales Invoice`.sales_person_user = %s
            AND `tabSales Invoice`.posting_date BETWEEN %s AND %s
            GROUP BY 
                `tabSales Invoice`.sales_person_user
                """
            sales_data1 = frappe.db.sql(sales_query1, (args.get('company'), i['Name'],args.get('january'), args.get('to_date')), as_dict=True)

            if target_data and sales_data and sales_data1:
                d1=target_data[0].get('Target')/12 if target_data[0].get('Target') else 0
                d2=sales_data[0].get('Grand_Total')
                d3=d2-d1
                d4=(d3/d1)*100 if target_data[0].get('Target') else 0
                d5=target_data[0].get('Target') if target_data[0].get('Target') else 0
                d6=sales_data1[0].get('Grand_Total1')
                d7=d6-d5
                d8=(d7/d5)*100 if target_data[0].get('Target') else 0
                row9=[f"{d1:.2f}",f"{d2:.2f}",f"{d3:.2f}",f"{d4:.2f}%",d5,f"{d6:.2f}",f"{d7:.2f}",f"{d8:.2f}%","."]
            elif target_data and sales_data:
                d1=target_data[0].get('Target')/12 if target_data[0].get('Target') else 0
                d2=sales_data[0].get('Grand_Total')
                d3=d2-d1
                d4=(d3/d1)*100 if target_data[0].get('Target') else 0
                row9=[f"{d1:.2f}",f"{d2:.2f}",f"{d3:.2f}",f"{d4:.2f}%",target_data[0].get('Target'),"N/A","N/A","N/A","."]
            elif target_data:
                if target_data[0].get('Target'):
                    d1=target_data[0].get('Target')/12
                    row9=[f"{d1:.2f}","N/A","N/A","N/A",target_data[0].get('Target'),"N/A","N/A","N/A","."]
                else:
                    row9=["N/A","N/A","N/A","N/A","N/A","N/A","N/A","N/A","."]

            row10=["(3) = (2) - (1)    (4) = (3) / (1)","","","(7) = (6) - (5)        (8) = (7) / (5)"]	
            row_empty=["","_"]
            data.append(row1)
            data.append(row2)
            data.append(row3)
            # data.append(row4)
            data.append(row5)
            data.append(row6)
            data.append(row7)
            data.append(row8)
            data.append(row9)
            data.append(row10)
            data.append(row_empty)
    return data

@frappe.whitelist()
def get_from_to_dates(month,fiscal_year):
    if month == 'January':
        month1 = "01"
    if month == 'February':
        month1 = "02"
    if month == 'March':
        month1 = "03"
    if month == 'April':
        month1 = "04"
    if month == 'May':
        month1 = "05"
    if month == 'June':
        month1 = "06"
    if month == 'July':
        month1 = "07"
    if month == 'Augest':
        month1 = "08"
    if month == 'September':
        month1 = "09"
    if month == 'October':
        month1 = "10"
    if month == 'November':
        month1 = "11"
    if month == 'December':
        month1 = "12"
    formatted_start_date = fiscal_year + '-' + month1 + '-01'
    formatted_end_date = get_last_day(formatted_start_date)
    return formatted_start_date,formatted_end_date
@frappe.whitelist()
def get_year_to_dates(fiscal_year):
    
    
    # frappe.errprint(fiscal_year)
    formatted_start_date2 = fiscal_year + '-01' + '-01'
    # formatted_start_date2 = '01-' + month1 + '-' + fiscal_year
    # frappe.errprint(formatted_start_date2)
    return formatted_start_date2