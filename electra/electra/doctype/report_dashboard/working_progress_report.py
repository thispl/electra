from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils import nowdate,nowtime
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types

@frappe.whitelist()
def download():
    filename = 'Working Progress Report'
    test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
         
    ws = wb.create_sheet(sheet_name, 0)
    if args.division:
        project = frappe.get_all("Project",{'company':args.division},['*'])
        if project:
            frappe.errprint(project)
            ws.append([args.division,"","","","","","","","","","","","","","","","","","",""])
            ws.append(["REVENUE FROM ONGOING CONTTRACTS TILL END OF MAY","","","","","","","","","","","","","","","","","","",""])
            ws.append(['Sl. No','Project No.','Project Reference','Project Description','Contract Value','Revised Contract Value','Estimated cost to complete','Estimated Profit Margin','ACTUAL COSTS TILL PREVIOUS YEAR','Costs incurred during the year','TOTAL COST','Total costs incurred','Percentage of Completion','Total Revenue to date','Revenue recognised upto  PY','Revenue for the Current Year - 22','Progress Bills raised till 2021','Progress Bills raised in 2022','Progress Bills raised till 2022','Short/(Excess) revenue over billings','% TO BE COMPLETED']) 
            i=1
            for proj in project:    
                name=proj.name
                ref=proj.project_name
                des=proj.customer
                cval=frappe.get_value('Sales Order',{'name':proj.sales_order},['grand_total']) or 0
                rcval=frappe.get_value('Sales Order',{'name':proj.sales_order},['grand_total']) or 0

                ws.append([i,name,ref,des,cval,rcval])
                i=i+1

    
    
  
        
    for rows in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=7):
        for cell in rows:
            cell.fill = PatternFill(fgColor="128C29", fill_type = "solid")
    for rows in ws.iter_rows(min_row=3, max_row=3, min_col=10, max_col=11):
        for cell in rows:
            cell.fill = PatternFill(fgColor="128C29", fill_type = "solid")
    for rows in ws.iter_rows(min_row=4, min_col=11, max_col=11):
        for cell in rows:
            cell.fill = PatternFill(fgColor="f7ee40", fill_type = "solid")
    for rows in ws.iter_rows(min_row=3, min_col=8, max_col=9):
        for cell in rows:
            cell.fill = PatternFill(fgColor="8a8888", fill_type = "solid")

    for rows in ws.iter_rows(min_row=4, min_col=10, max_col=10):
        for cell in rows:
            cell.fill = PatternFill(fgColor="8a8888", fill_type = "solid")
    for rows in ws.iter_rows(min_row=3, min_col=12, max_col=16):
        for cell in rows:
            cell.fill = PatternFill(fgColor="8a8888", fill_type = "solid")
    for rows in ws.iter_rows(min_row=3, min_col=17, max_col=19):
        for cell in rows:
            cell.fill = PatternFill(fgColor="85ab72", fill_type = "solid")
    for rows in ws.iter_rows(min_row=3, min_col=20, max_col=21):
        for cell in rows:
            cell.fill = PatternFill(fgColor="8a8888", fill_type = "solid")

    border = Border(left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000'))
    for rows in ws.iter_rows(min_row=3, max_row=len(project)+3, min_col=1, max_col=len(project)+14):

        for cell in rows:
            cell.border = border


    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column= 21)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column= 21)

    align_center = Alignment(horizontal='center',vertical='center')
    for cell in ws['1:1']:
        cell.alignment = align_center
        cell.font = Font(bold=True)
    for cell in ws['2:2']:
        cell.alignment = align_center
        cell.font = Font(bold=True)    

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary' 