# Copyright (c) 2021, TEAMPRO and contributors
# For license information, please see license.txt

from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils import nowdate,nowtime
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue
from openpyxl import Workbook
from openpyxl.drawing.image import Image

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import erpnext
from frappe.utils import flt

@frappe.whitelist()
def download():
	filename = 'Monthly Salary Register Print'
	test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()

	ws = wb.create_sheet(sheet_name, 0)
	if not args:
		args = {}
	currency = None
	if args.get("currency"):
		currency = args.get("currency")
	company_currency = erpnext.get_company_currency(args.get("company"))
	if args.division:
		salary_slip = frappe.get_all("Salary Slip", {'company': args.division}, ['*'])
		if salary_slip:
			header_image_path = frappe.db.get_value("Letter Header",{'name':'Electra'},['image'])
			add_header_image(ws, header_image_path)
			from_date_str = args.from_date
			from_date_obj = datetime.strptime(from_date_str, "%Y-%m-%d")
			formatted_from_date = from_date_obj.strftime("%d-%m-%Y")
			to_date_str = args.to_date
			to_date_obj = datetime.strptime(to_date_str, "%Y-%m-%d")
			formatted_to_date = to_date_obj.strftime("%d-%m-%Y")
			ws.append(['','','','','','','','','Monthly Salary Register Print'])
			ws.append(['From:', formatted_from_date])
			ws.append(['To:', formatted_to_date])
			ws.append(['Currency:', args.currency])
			ws.append(['Company:', args.division])
			ws.append(['Document Status:', args.status])
			ws.append(['Employee', 'Employee Name', 'Fixed Basic(QR)', 'Fixed HRA(QR)', 'Fixed Other Allowance(QR)',
					   'Fixed Transport Allowance(QR)', 'Fixed Total', 'Working Days', 'Absent Days', 'Not Hours',
					   'Hot Hours', 'Not Amount(QR)', 'Hot Amount(QR)', 'Previous Month Additionals(QR)',
					   'Other Additions(QR)', 'Gross Pay(QR)', 'Loan Deduction(QR)', 'Other/Advance Deduction(QR)',
					   'Mess Advance(QR)', 'Previous Month Abs Deduction', 'Net Pay(QR)', 'Remarks'])

			processed_employees = set()
			for ss in salary_slip:
				if ss.employee not in processed_employees:
					sum_emp = frappe.db.sql(
						"""select sum(basic + hra + _other_allowance + transportation) as total from `tabEmployee` where employee='%s' """ % (
							ss.employee), as_dict=1)[0]
					row = [
						ss.employee,
						ss.employee_name,
						frappe.db.get_value('Employee', ss.employee, 'basic'),
						frappe.db.get_value('Employee', ss.employee, 'hra'),
						frappe.db.get_value('Employee', ss.employee, '_other_allowance'),
						frappe.db.get_value('Employee', ss.employee, 'transportation'),
						sum_emp['total'] or "-",
						ss.total_working_days,
					]

					row += [
						ss.absent_days or 0,
						int(frappe.get_value('Additional Salary',
											 {'salary_component': "NOT Hours", 'employee': ss.employee,
											  'docstatus': 1}, ['not_hours']) or 0),
						int(frappe.get_value('Additional Salary',
											 {'salary_component': "HOT Hours", 'employee': ss.employee,
											  'docstatus': 1}, ['hot_hours']) or 0),
						int(frappe.get_value('Salary Detail', {'salary_component': "NOT Hours", 'parent': ss.name},
											 ["amount"]) or 0),
						int(frappe.get_value('Salary Detail', {'salary_component': "HOT Hours", 'parent': ss.name},
											 ["amount"]) or 0),
						int(frappe.get_value('Salary Detail',
											 {'salary_component': "Previous Month Additionals", 'parent': ss.name},
											 ["amount"]) or 0),
						int(frappe.get_value('Salary Detail',
											 {'salary_component': "Others Additions", 'parent': ss.name},
											 ["amount"]) or 0)]

					if currency == company_currency:
						row += [int(flt(ss.gross_pay)) * int(flt(ss.exchange_rate))]
					else:
						row += [int(ss.gross_pay)]

					row += [
						frappe.get_value('Salary Detail', {'salary_component': "Loan Deduction", 'parent': ss.name},
										 ["amount"]) or 0,
						frappe.get_value('Salary Detail',
										 {'salary_component': "Other/ Advance Deduction", 'parent': ss.name},
										 ["amount"]) or 0,
						frappe.get_value('Salary Detail', {'salary_component': "Mess Advance", 'parent': ss.name},
										 ["amount"]) or 0,
						frappe.get_value('Salary Detail',
										 {'salary_component': "Previous Month Abs Detection", 'parent': ss.name},
										 ["amount"]) or 0
					]

					if currency == company_currency:
						row += [
							int(flt(ss.net_pay)) * int(flt(ss.exchange_rate)),
						]
					else:
						row += [
							int(ss.net_pay) or "-",
						]

					row += [
						frappe.db.get_value('Additional Salary',
											{'employee': ss.employee, "salary_component": "Previous Month Abs Detection",
											 "payroll_date": ss.start_date}, ['Remarks']) or "-"
					]

					ws.append(row)
					processed_employees.add(ss.employee)
		



	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'

def add_header_image(ws, image_path):
	header_image_path = frappe.db.get_value("Letter Header",{'name':'Electra'},['image'])
	if header_image_path:
		img = Image(header_image_path)
		ws.add_image(img, 'A1:D5')
	else:
		print("Header image path is not available.")