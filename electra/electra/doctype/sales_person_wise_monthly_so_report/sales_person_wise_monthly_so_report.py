# Copyright (c) 2024, Abdulla and contributors
# For license information, please see license.txt

from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from frappe import _
import frappe
from frappe.model.document import Document
from datetime import date, timedelta, datetime,time
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
	nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
import math
from frappe.utils import add_months, cint, flt, getdate, time_diff_in_hours,time_diff_in_seconds
import locale


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import openpyxl.styles as styles
from frappe.utils import flt, fmt_money


@frappe.whitelist()
def download():
	filename = 'Sales Person Wise Monthly SO'
	test = build_xlsx_response(filename)
	
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
		 
	ws = wb.create_sheet(sheet_name, 0)
	ws.column_dimensions['A'].width = 20
	ws.column_dimensions['N'].width = 20
	ws.append(['Sales Person','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec','Total ' + args.year])
	data=get_data(args)
	for i in data:
		ws.append(i)
	totals = ['Total']
	for col in ws.iter_cols(min_row=2, min_col=2):
		column_total = sum(cell.value for cell in col if cell.value is not None)
		totals.append(column_total)
	ws.append(totals)
	data_rows = len(get_data(args)) + 1  # Adding 1 to account for the header row
	last_column_total = sum(cell.value for cell in list(ws.iter_cols(min_row=len(get_data(args))+2, min_col=len(ws[1]), max_col=len(ws[1])))[0] if isinstance(cell.value, (int, float)))
	per = ['Percentage']
	for col_index, col in enumerate(ws.iter_cols(min_row=len(get_data(args))+2, min_col=2, max_col=len(ws[1])-1)):  # Exclude the last column
		column_total = sum(cell.value for cell in col if cell.value is not None)
		percentages = []
		if column_total != 0:
			percentages = [(float(cell.value) / last_column_total) * 100 if cell.value is not None else 0 for cell in col]
		else:
			percentages = [0] * len(col)
		per.extend(percentages)
	ws.append(per)
	
	

	align_center = Alignment(horizontal='center',vertical='center')
	align_right = Alignment(horizontal='right',vertical='bottom')

	for header in ws.iter_rows(min_row=1 , max_row=1, min_col=1, max_col=14):
		for cell in header:
			cell.font = Font(bold=True)
			cell.alignment = align_center
	for header in ws.iter_rows(min_row=len(get_data(args))+2 , max_row=len(get_data(args))+3, min_col=1, max_col=14):
		for cell in header:
			cell.font = Font(bold=True)
			cell.alignment = align_center
	border_thin = Border(
	left=Side(style='thin'),
	right=Side(style='thin'),
	top=Side(style='thin'),
	bottom=Side(style='thin'))
	header_range = ws['A1':ws.cell(row=len(get_data(args))+3, column=14).coordinate]
	for row in header_range:
		for cell in row:
			cell.border = border_thin

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'


def get_data(args):
	data = []
	if args.sales_person:
		data.append([args.sales_person])
		sales_orders = frappe.db.sql("""
			SELECT 
				MONTH(transaction_date) AS month,
				SUM(grand_total) AS grand_total
			FROM 
				`tabSales Order`
			WHERE 
				YEAR(transaction_date) = %s 
				AND sales_person_user = %s 
				AND docstatus = 1
			GROUP BY 
				MONTH(transaction_date)
		""", (args.year, args.sales_person), as_dict=True)
		total = 0
		for order in sales_orders:
			total += order.grand_total
			data[0].append((order.grand_total))
		for month in range(1, 13):
			if not any(order['month'] == month for order in sales_orders):
				data[0].append(0)

		data[0].append((total))
		
	else:
		sales_persons = frappe.db.sql("""SELECT name FROM `tabSales Person`""", as_dict=True)
		for sp in sales_persons:
			sales_orders = frappe.db.sql("""
				SELECT 
					MONTH(transaction_date) AS month,
					SUM(grand_total) AS grand_total
				FROM 
					`tabSales Order`
				WHERE 
					YEAR(transaction_date) = %s 
					AND sales_person_user = %s
					AND docstatus = 1
				GROUP BY 
					MONTH(transaction_date)
			""", (args.year, sp.name), as_dict=True)
			total = 0
			sp_data = [sp.name]  # Create a separate list for each salesperson
			for order in sales_orders:
				
				sp_data.append(order.grand_total)
				total += order.grand_total
			for month in range(1, 13):
				if not any(order['month'] == month for order in sales_orders):
					sp_data.append(0)
			
			sp_data.append(total)
			data.append(sp_data)

	return data


class SalesPersonWiseMonthlySOReport(Document):
	pass
