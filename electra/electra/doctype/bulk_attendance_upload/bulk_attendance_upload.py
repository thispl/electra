# Copyright (c) 2021, Abdulla and contributors
# For license information, please see license.txt

# Copyright (c) 2021, TEAMPRO and contributors
# For license information, please see license.txt

from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types

class BulkAttendanceUpload(Document):
    @frappe.whitelist()
    def validate(self):
        if self.upload:
            filepath = get_file(self.upload)
            pps = read_csv_content(filepath[1])
            no_of_days = date_diff(add_days(self.to_date, 1), self.from_date)
            dates = [add_days(self.from_date, i) for i in range(0, no_of_days)]
            for pp in pps:
                if pp[1] != 'Employee':
                    frappe.errprint(pp[1])
                    if pp[1]:
                        b = 3
                        for date in dates:
                            if pp[b]:
                                doj = frappe.db.get_value('Employee',{'employee':pp[1]},['date_of_joining'])
                                company = frappe.db.get_value('Employee',{'employee':pp[1]},['company'])
                                if doj:
                                    if datetime.strptime(date,'%Y-%m-%d').date() >= doj:
                                        if not frappe.db.exists("Attendance",{'employee':pp[1],'attendance_date':date,'company':company}):
                                            doc = frappe.new_doc("Attendance")
                                            doc.employee = pp[1]
                                            doc.attendance_date = date
                                            doc.company = company
                                            frappe.errprint(company)
                                            if pp[b] == "P":
                                                doc.status = "Present"
                                            doc.save(ignore_permissions=True)
                                            doc.submit()
                                            b = b+1
                            else:
                                b = b+1
            frappe.msgprint('Attendance uploaded successfully')
                                
        
    
          

# @frappe.whitelist()
# def download():
#     filename = 'Bulk Attendance Upload'
#     test = build_xlsx_response(filename)
@frappe.whitelist()
def download():
    args = frappe.local.form_dict

    w = UnicodeWriter()
    no_of_days = date_diff(add_days(args.to_date, 1), args.from_date)
    dates = [add_days(args.from_date, i) for i in range(0, no_of_days)]
    header = ["S.no","Employee","Name","Department","Designation"]
    header1 =["DAY","","","",""]
    for date in dates:
        dt = datetime.strptime(date,'%Y-%m-%d')
        day_format = datetime.date(dt).strftime('%d/%m/%y')
        dayfor = datetime.date(dt).strftime('%A')
        header.append(day_format)
        header1.append( dayfor)
    w.writerow(header)
   
    w.writerow(header1)
    employee = frappe.db.get_all("Employee",{'status':'Active'},['*'])
    i = 1
    for emp in employee:
        emp_id = emp.employee
        name = emp.employee_name
        department = emp.department
        designation = emp.designation
        w.writerow([i,emp_id,name,department,designation])
        i = i+1

    

    # write out response as a type csv
    frappe.response['result'] = cstr(w.getvalue())
    frappe.response['type'] = 'csv'
    frappe.response['doctype'] = "Attendance"


    
    
    
    

        
        

  
    
 






    






    # xlsx_file = BytesIO()
    # w.save(xlsx_file)
    # return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.csv'
    # frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'csv'
    frappe.response['doctype'] = "Attendance"
    frappe.response['result'] = cstr(xlsx_file.getvalue())


# frappe.response['result'] = cstr(w.getvalue())
#     frappe.response['type'] = 'csv'
    # frappe.response['doctype'] = "Shift Assignment"