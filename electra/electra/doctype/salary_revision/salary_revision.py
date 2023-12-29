# Copyright (c) 2023, Abdulla and contributors
# For license information, please see license.txt

import frappe
from datetime import date

today = date.today()
from frappe.model.document import Document
import datetime 
import frappe,erpnext
from frappe.utils import cint
import json
from frappe.utils import date_diff, add_months,today,add_days,add_years,nowdate,flt
from frappe.model.mapper import get_mapped_doc
from frappe.utils.file_manager import get_file
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
import datetime
from datetime import date,datetime,timedelta
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
import pandas as pd
from frappe.utils import formatdate
from frappe.utils import now
from erpnext.setup.utils import get_exchange_rate
from datetime import date

class SalaryRevision(Document):
	pass

	@frappe.whitelist()
	def on_submit(self):
		if self.salary_revision_status == "Permanent":	
			today = date.today()
			date_str = self.effective_date 
			date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
			if date_obj < today:
				emp = frappe.get_doc("Employee",{'employee':self.employee,'status': 'Active' })
				frappe.errprint(emp.employee_name)
				emp.salary_mode = self.salary_mode
				emp.salary_currency = self.salary_currency
				emp.basic = self.basic
				emp.hra = self.hra
				emp.accommodation = self.accommodation
				emp.mobile_allowance = self.mobile_allowance
				emp.transportation = self.transport_allowance
				emp._other_allowance = self.other_allowance
				emp.medical_allowance_ = self.medical_allowance
				emp.leave_salary = self.leave_salary
				emp.air_ticket_allowance_ = self.air_ticket_allowance
				emp.qid_cost = self.qid_cost
				emp.medical_renewal = self.medical_renewal
				emp.compensationemployee_insurence = self.compensation_employee_insurence
				emp.visa_cost_ = self.visa_cost
				emp.gross_salary = self.gross_salary
				emp.gratuity = self.gratuity
				emp.ctc = self.cost_to_company
				emp.per_hour_cost = self.per_hour_cost
				emp.append('history',{
					"date":self.effective_date,
					"basic":self.basic,
					"hra":self.hra,
					"other_allowance":self.other_allowance,
					"transport_allowance":self.transport_allowance,
					"medical_allowance":self.medical_allowance,
					"air_ticket_allowance":self.air_ticket_allowance,
					"mobile_allowance":self.medical_allowance,
					"visa_cost":self.visa_cost,
					"accommodation":self.accommodation,
					"leave_salary":self.leave_salary,
					"qid_cost":self.qid_cost,
					"medical_renewal":self.medical_renewal,
					"compensation":self.compensation_employee_insurence,
					"remark":self.remarks
				})
				emp.save(ignore_permissions=True)
				frappe.db.commit()
		else:
			today = date.today()
			date_str = self.effective__from_date 
			date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
			if date_obj < today:
				frappe.errprint("HI")
				emp = frappe.get_doc("Employee",{'employee':self.employee,'status': 'Active' })
				frappe.errprint(emp.employee_name)
				emp.salary_mode = self.salary_mode1
				emp.salary_currency = self.salary_currency1
				emp.basic = self.basic2
				emp.hra = self.hra2
				emp.accommodation = self.accommodation2
				emp.mobile_allowance = self.mobile_allowance2
				emp.transportation = self.transport_allowance2
				emp._other_allowance = self.other_allowance2
				emp.medical_allowance_ = self.medical_allowance2
				emp.leave_salary = self.leave_salary2
				emp.air_ticket_allowance_ = self.air_ticket_allowance2
				emp.qid_cost = self.qid_cost2
				emp.medical_renewal = self.medical_renewal2
				emp.compensationemployee_insurence = self.compensation_2
				emp.visa_cost_ = self.visa_cost2
				emp.gross_salary = self.gross_salary2
				emp.gratuity = self.gratuity2
				emp.ctc = self.cost_to_company2
				emp.per_hour_cost = self.per_hour_cost2
				emp.save(ignore_permissions=True)
				frappe.db.commit()


