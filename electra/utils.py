from __future__ import unicode_literals
import json
import requests
from re import S
from unicodedata import name
from datetime import datetime
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from datetime import date, timedelta
import calendar
from erpnext.stock.get_item_details import get_item_price
import frappe
from frappe import _
from frappe.utils import flt
from frappe.utils import formatdate
from erpnext.selling.doctype.customer.customer import get_customer_outstanding, get_credit_limit
from erpnext.stock.get_item_details import get_valuation_rate
from erpnext.setup.utils import get_exchange_rate
from frappe.utils.background_jobs import enqueue
from frappe.utils import (
	add_days,
	add_months,
	cint,
	date_diff,
	flt,
	get_first_day,
	get_last_day,
	get_link_to_form,
	getdate,
	rounded,
	today,
)
import json
from frappe.core.doctype.session_default_settings.session_default_settings import (
	clear_session_defaults,
	set_session_default_values,
)
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types

@frappe.whitelist()
def update_customer_filter():
	ce = frappe.get_all('Cost Estimation',['name','lead_customer'])
	for c in ce:
		ce_doc = frappe.get_doc('Cost Estimation',c.name)
		ce_doc.db_set('customer',c.lead_customer)
		
#Update the below value while submission of payment entry
@frappe.whitelist()
def update_outstanding_without_retention(doc,method):
	if doc.order_type == 'Project':
		doc.outstanding_amount__without_retention = doc.outstanding_amount
		
@frappe.whitelist()
def add_company_to_session_defaults(doc):
	default_company = frappe.get_value('User Permission',{'user':frappe.session.user,'allow':'Company','is_default':1},'for_value')
	if default_company:
		frappe.defaults.set_user_default("company", default_company)

#Get the cluster value set to qn
@frappe.whitelist()
def set_cluster(sales_person):
	doc = frappe.get_value("Cluster",{"user":sales_person})
	if doc:
		return doc

#Get the below value for print
@frappe.whitelist()
def get_dns(doc):
	dn = []
	sales_order = frappe.db.sql("""select sales_order as so from `tabSales Invoice Item` where parent = '%s' """ % doc,as_dict=1)
	if sales_order:
		for item in sales_order: 
			if item.so not in dn:
				if item.so:
					dn.append(item)
	if len(dn) > 0:
		return dn
	else:
		return ''

#Get the below value for print
@frappe.whitelist()
def get_sos(doc):
	dn = []
	sales_order = frappe.db.sql("""select delivery_note as dn from `tabSales Invoice Item` where parent = '%s' """ % doc,as_dict=1)
	if sales_order:
		for item in sales_order: 
			if item.dn not in dn:
				if item.dn:
					dn.append(item.dn)
	if len(dn) > 0:
		return dn
	else:
		return ''
	
#Get the below value and set it execute method
@frappe.whitelist()
def get_so_list():
	doc = "ITD-CRD-2023-00106"
	dn = []
	sales_order = frappe.db.sql("""select delivery_note as dn from `tabSales Invoice Item` where parent = '%s' """ % doc,as_dict=1)
	if sales_order:
		for item in sales_order: 
			if item.dn not in dn:
				if item.dn:
					dn.append(item.dn)
	if len(dn) > 0:
		return dn
	else:
		return ''

	# frappe.log_error(title='dn_list',message=dn[0])
			
	# return sales_order

#Set the value to execute the method
@frappe.whitelist()
def reset_general_entry_purchase_rate():
	last_purchase_rate = frappe.get_value('Item','001','last_purchase_rate')
	if last_purchase_rate:
		frappe.db.set_value('Item','001','last_purchase_rate',0.0)
		frappe.db.commit()

#Add child values for Update selling Price
@frappe.whitelist()
def get_item_details(item_group=None,brand=None,item=None):
	if item_group and brand:
		filters = { 'item_group' : item_group,'brand' : brand }
	if brand:
		filters = { 'brand' : brand }
	if item_group:
		filters = { 'item_group' : item_group }
	if item:
		filters = {'item_code':item}  

	

	items = frappe.get_all("Item",fields=['item_code','stock_uom','item_name',],filters=filters)
	item_details = []
	for item in items:
		selling_price = 0.0
		item_price_args = {
				"item_code": item['item_code'],
				"price_list": "Standard Selling",
				"uom": item['stock_uom'],
				"batch_no": ""
				
		}
		item_code = item['item_code']
		item_price_list = frappe.db.get_value('Item Price',{'item_code':item_code},['name'])
		item_price = get_item_price(item_price_args,item_code,)
		if item_price:
			selling_price = item_price[0][1]
		item_detail = {
				"item_code": item['item_code'],
				"item_name": item['item_name'],
				"selling_price": selling_price,
				'item_marked':item_price_list
		}
		item_details.append(item_detail)
	return item_details

#Update the price while click the update button for update selling price
@frappe.whitelist()
def update_selling_price(update_selling_price,percentage):
	items = json.loads(update_selling_price)
	for i in items:
		if percentage:
			percent_calculate = i['selling_price']
			percent_value = percent_calculate * (1 +(int(percentage)/100))
			frappe.db.set_value("Item Price",i['item_marked'],'price_list_rate',percent_value)
			frappe.msgprint('Item Price Rate was updated')
		else:
			frappe.db.set_value('Item Price',i['item_marked'],'price_list_rate',i['selling_price'])
			frappe.msgprint(_('Item Price Rate was updated'))

	   
#Get the credit limit for given customer and company to match with doctype    
@frappe.whitelist()
def fetch_credit_limit(company,customer,doctype):
	result = ""
	customer_type = frappe.db.get_value("Customer",customer,['customer_type'])
	customer_list = get_details(company,customer)
	if not customer_list and customer_type == "Company" and doctype != "Quotation":
		frappe.throw(_("Kindly Set the Credit Limit in Customer"))
	bypass_credit_limit_check = is_frozen = disabled = "No"
	customer_naming_type = frappe.db.get_value("Selling Settings", None, "cust_master_name")
	for d in customer_list:
		outstanding_amt = round(get_customer_outstanding(d.name, company,
				ignore_outstanding_sales_order=d.bypass_credit_limit_check),2)
		credit_limit = round(get_credit_limit(d.name, company),2)
		bal = round(flt(credit_limit) - flt(outstanding_amt),2)
		if d.bypass_credit_limit_check:
			bypass_credit_limit_check = "Yes"
		if d.is_frozen:
			is_frozen = "Yes"
		if d.disabled:
			disabled = "Yes"
		if credit_limit and bal:
			result = '''
			<table class="table table-bordered">
			<tr>
				<td>Credit Limit</td>
				<td>Outstanding Amount</td>
				<td>Balance</td>
				<td>Bypassing Credit Limit Enabled ?</td>
				<td>Frozen ?</td>
				<td>Disabled ?</td>
			</tr>
			<tr>
				<td>%s</td>
				<td>%s</td>
				<td>%s</td>
				<td>%s</td>
				<td>%s</td>
				<td>%s</td>
			</tr>
			
			</table>
			''' %( credit_limit, outstanding_amt, bal,bypass_credit_limit_check,is_frozen,disabled )
		# else:
		#     result = "Empty"
		return result

#Get the below details to match with given company and customer
def get_details(company, customer):
	sql_query = """SELECT
						c.name, c.customer_name,
						ccl.bypass_credit_limit_check,
						c.is_frozen, c.disabled
					FROM    `tabCustomer` c
					INNER JOIN `tabCustomer Credit Limit` ccl ON c.name = ccl.parent
					WHERE
						ccl.company = %s AND c.name = %s"""

	return frappe.db.sql(sql_query, (company, customer), as_dict=True)


#Get all warehouse for to match with given company
@frappe.whitelist()
def get_wh(company):
	warehouses = frappe.get_list("Warehouse",{'company':company},['name'],ignore_permissions=True)
	wh_list = []
	for wh in warehouses:
		wh_list.append(wh['name'])
	return wh_list


#Get the below details and set to margin item table while enter the margin item calculation
@frappe.whitelist()
def get_last_valuation_rate(item_code,source_company):
	valuation_rate = 0
	source_warehouse = frappe.db.get_value('Warehouse', {'default_for_stock_transfer':1,'company': source_company }, ["name"])
	latest_vr = frappe.db.sql("""select valuation_rate as vr from tabBin
			where item_code = '%s' and warehouse = '%s' """%(item_code,source_warehouse),as_dict=True)
	# latest_vr = frappe.db.sql("""
	#             SELECT valuation_rate as vr FROM `tabStock Ledger Entry` WHERE item_code = %s AND company=%s AND valuation_rate > 0
	#             ORDER BY posting_date DESC, posting_time DESC, creation DESC LIMIT 1
	#             """, (item_code,source_warehouse),as_dict=True)
	if latest_vr:
		if latest_vr[0]["vr"] > 0:
			valuation_rate = latest_vr[0]["vr"]
		else:
			val_rate = []
			l_vr = frappe.db.sql("""select valuation_rate as vr from tabBin
					where item_code = '%s' """%(item_code),as_dict=True)
			for item in l_vr: 
				if item not in val_rate: 
					val_rate.append(item.vr)
			if len(val_rate) > 1 :
				valuation_rate = max(val_rate)
	else:
		val_rate = []
		l_vr = frappe.db.sql("""select valuation_rate as vr from tabBin
				where item_code = '%s' """%(item_code),as_dict=True)
		for item in l_vr: 
			if item not in val_rate: 
				val_rate.append(item.vr)
		if len(val_rate) > 1 :
			valuation_rate = max(val_rate)
	return valuation_rate

#Return the below values to match with coverted by value and set to sales person details in SO
@frappe.whitelist()
def make_sales_order(name):
	prepared_by,converted_by = frappe.db.get_value("Quotation",{'name':name},['user_id','converted_by']) 

	user_list=frappe._dict()
	user_list.update({
		"converted_by":converted_by,
		"prepared_by":prepared_by
	})
	return user_list


#Get the valuation rate of given below item code
@frappe.whitelist()
def get_valrate():
	item_code = '1PRTR-3SS84-3'
	source_company = 'MEP DIVISION - ELECTRA'
	valuation_rate = 0
	source_warehouse = frappe.db.get_value('Warehouse', {'default_for_stock_transfer':1,'company': source_company }, ["name"])
	latest_vr = frappe.db.sql("""select valuation_rate as vr from tabBin
			where item_code = '%s' and warehouse = '%s' """%(item_code,source_warehouse),as_dict=True)
	# latest_vr = frappe.db.sql("""
	#             SELECT valuation_rate as vr FROM `tabStock Ledger Entry` WHERE item_code = %s AND company=%s AND valuation_rate > 0
	#             ORDER BY posting_date DESC, posting_time DESC, creation DESC LIMIT 1
	#             """, (item_code,source_warehouse),as_dict=True)
	if latest_vr:
		if latest_vr[0]["vr"] > 0:
			valuation_rate = latest_vr[0]["vr"]
		else:
			val_rate = []
			l_vr = frappe.db.sql("""select valuation_rate as vr from tabBin
					where item_code = '%s' """%(item_code),as_dict=True)
			for item in l_vr: 
				if item not in val_rate: 
					val_rate.append(item.vr)
			if len(val_rate) > 1 :
				valuation_rate = max(val_rate)
	else:
		val_rate = []
		l_vr = frappe.db.sql("""select valuation_rate as vr from tabBin
				where item_code = '%s' """%(item_code),as_dict=True)
		for item in l_vr: 
			if item not in val_rate: 
				val_rate.append(item.vr)
		if len(val_rate) > 1 :
			valuation_rate = max(val_rate)
	return valuation_rate






#Return the below values to match with coverted by value and set to sales person details in DN
@frappe.whitelist()
def make_dnote(name):
	prepared_by,sales_person_user = frappe.db.get_value("Sales Order",{'name':name},['prepared_by','sales_person_user']) 

	user_list=frappe._dict()
	user_list.update({
		"sales_person_user":sales_person_user,
		"prepared_by":prepared_by
	})
	return user_list

#Set the SO name in DN 
@frappe.whitelist()
def make_dn(name):
	name =frappe.db.get_value("Sales Order",{'name':name},['name'])
	return name
# @frappe.whitelist()
# def get_dn_names(so_name):
#     name,sales_order =frappe.get_list("Delivery Invoice",{'name':name},['name','sales_order'])
#     frappe.errprint(name)
#     frappe.errprint(sales_order)
#     refer_list = frappe._dict()
#     refer_list.update({
#         'sales_invoice':name,
#         'dn':sales_order
#     })
#     return refer_list

# @frappe.whitelist()
# def make_qn(name):
#     frappe.errprint(name)
#     name = frappe.db.get_value('Opportunity',{'name':name},name)
#     return name

#Below documents has been cancelled While delete the project 
@frappe.whitelist()     
def pe_on_trash(doc,method):
	ple = frappe.qb.DocType("Payment Ledger Entry")
	frappe.qb.from_(ple).delete().where(
		(ple.voucher_type == doc.doctype) & (ple.voucher_no == doc.name)
	).run()
	frappe.db.sql(
		"delete from `tabGL Entry` where voucher_type=%s and voucher_no=%s", (doc.doctype, doc.name)
	)
	frappe.db.sql(
		"delete from `tabStock Ledger Entry` where voucher_type=%s and voucher_no=%s",
		(doc.doctype, doc.name),
	)

#Get the below details and set to User details in QN
@frappe.whitelist()
def get_user_details(user):
	employee = frappe.get_value("Sales Person",user,['user_id'])
	employees = frappe.get_list("Employee",{'user_id':employee},['user_id','designation','cell_number'],ignore_permissions=True)
	employee_list = frappe._dict()
	for e in employees:
		employee_list.update({
			'user_id': e.user_id,
			'designation': e.designation,
			'cell_number': e.cell_number
		})
	return employee_list

#Get the below details to match with given user while enter the User id in QN
@frappe.whitelist()
def get_user_detail(user):
	employees = frappe.get_list("Employee",{'user_id':user},['employee_name','designation','cell_number'],ignore_permissions=True)
	employee_list = frappe._dict()
	for e in employees:
		employee_list.update({
			'employee_name': e.employee_name,
			'designation': e.designation,
			'cell_number': e.cell_number
		})
	return employee_list

#Get the details from CECost Estimation Master Scope of Work to match with 
@frappe.whitelist()
def get_ce_msow(cost_estimation):
	return frappe.get_all("CE Master Scope of Work",{'parent':cost_estimation},['*'])


#Get the all details from Employee Property History and set to Employee Master Record
@frappe.whitelist()
def get_transfer(employee_transfer_id):
	return frappe.get_all("Employee Property History",{'parent':employee_transfer_id},['*'])

#Get the all details from External Provider Evaluation
@frappe.whitelist()
def get_evaluation_period():
	return frappe.get_all("External Provider Evaluation",['*'])

#Get the below date from given doctype and set to PO
@frappe.whitelist()
def get_evaluation_date(supplier):
	epef = frappe.get_all("External Provider Evaluation Form",{'external_provider':supplier},['max(re_evaluation_date) as re_evaluation_date'])[0]
	return epef['re_evaluation_date']


#Show the below details in HTML View while open the QN
@frappe.whitelist()
def show_valuation_rate(items,company):
	import json
	items = json.loads(items)
	data = ''
	data += """<table class="table table-bordered">
				<tr><th style="padding:1px;border: 1px solid black" colspan=6><center>Itemwise Valuation Rate</center></th></tr>
				<tr>
					<td style="padding:1px;border: 1px solid black"><b>Item Code</b></td>
					<td style="padding:1px;border: 1px solid black"><b>Item Name</b></td>
					<td style="padding:1px;border: 1px solid black"><b>Valuation Rate</b></td>
				</tr>"""
	for item in items:
		gvr = get_valuation_rate(item['item_code'], company)
		last_valuation_rate = frappe.db.sql("""
				SELECT valuation_rate FROM `tabStock Ledger Entry` WHERE item_code = %s AND valuation_rate > 0
				ORDER BY posting_date DESC, posting_time DESC, creation DESC LIMIT 1
				""", (item['item_code']))
		data += """
		<tr>
		<td style="padding:1px;border: 1px solid black">%s</td>
		<td style="padding:1px;border: 1px solid black">%s</td>
		<td style="padding:1px;border: 1px solid black">%s</td>
		</tr>""" % (item['item_code'],item['item_name'],last_valuation_rate[0][0])
	data += """</table>"""
	if items:
		return data

#Create Master Scope of work while save the SO
@frappe.whitelist()
def validate_sow(doc,method):
	sows = doc.scope_of_work
	if sows:
		for sow in sows:
			if sow.msow:
				if frappe.db.exists('Master Scope of Work',sow.msow):
					msow_id = frappe.get_doc("Master Scope of Work",sow.msow)
				else:
					msow_id = frappe.new_doc("Master Scope of Work")
				msow_id.scope_of_work = sow.msow
				msow_id.desc = sow.msow_desc
				msow_id.is_group = 1
				msow_id.save(ignore_permissions=True)
			# if sow.ssow:
			#     if frappe.db.exists('Sub Scope of Work',{"name":sow.ssow,"parent_scope_of_work":sow.msow}):
			#         ssow_id = frappe.get_doc("Sub Scope of Work",sow.ssow)
			#     else:	
			#         ssow_id = frappe.new_doc("Sub Scope of Work")
			#     ssow_id.scope_of_work = sow.ssow
			#     ssow_id.desc = sow.ssow_desc
			#     ssow_id.parent_scope_of_work = sow.msow
			#     ssow_id.save(ignore_permissions=True)

#Document series set the Given Gerenal settings
@frappe.whitelist()
def get_series(company,doctype):
	company_series = frappe.db.get_value("Company Series",{'company':company,'document_type':doctype},'series')
	return company_series

#Get the all values from SO SOW and return to PB SOW
@frappe.whitelist()
def get_so_sow(sales_order,msow):
	so_sow = frappe.get_all("SO Scope of Work",{'parent':sales_order,'msow':msow},['*'])
	return so_sow


#Get the all values from CE MSOW and return to PB SOW
@frappe.whitelist()
def get_ce_sow(cost_estimation,msow):
	ce_sow = frappe.get_all("CE Master Scope of Work",{'parent':cost_estimation,'msow':msow},['*'])
	return ce_sow

	
#Set the quotation name in Cost Estimation while save of the QN
@frappe.whitelist()
def add_quotation_ce(doc,method):
	if doc.cost_estimation:
		frappe.db.set_value('Cost Estimation',doc.cost_estimation,'quotation',doc.name)
		frappe.db.commit()

		
#Create new doc of opportunity while save of the Quotation
@frappe.whitelist()
def validate_opportunity_sow(doc,method):
	sows = doc.scope_of_work
	for sow in sows:
		if sow.msow:
			if frappe.db.exists('Master Scope of Work',sow.msow):
				msow_id = frappe.get_doc("Master Scope of Work",sow.msow)
			else:
				msow_id = frappe.new_doc("Master Scope of Work")
			msow_id.master_scope_of_work = sow.msow
			msow_id.desc = sow.msow_desc
			msow_id.save(ignore_permissions=True)


#Set the income of Sales Invoice while save and any update of the document
@frappe.whitelist()
def set_income_account(doc,method):
	if doc.invoice_type == 'Cash':
		short_code = frappe.get_value('Company',doc.company,'abbr')
		for it in doc.items:
			it.income_account = 'Sales - Cash - '+short_code

#Default item set while save of the item document
@frappe.whitelist()
def item_default_wh(doc,method):
	item_default_set = frappe.get_value('Item',doc.item_code,'item_default_set')
	if not item_default_set:
		companies = [
			{"company":"KINGFISHER - TRANSPORTATION","default_warehouse":"Kingfisher Transportation Warehouse - KT", "buying_cost_center" : "Main - KT", "selling_cost_center" : "Main - KT","expense_account": "Cost of Goods Sold - KT","income_account":"Sales - Credit - KT"},
			{"company":"STEEL DIVISION - ELECTRA","default_warehouse":"Steel Warehouse - SDE", "buying_cost_center" : "Main - SDE", "selling_cost_center" : "Main - SDE","expense_account": "5111 - Cost of Goods Sold - SDE","income_account":"Sales - Credit - SDE"},
			{"company":"MARAZEEM SECURITY SERVICES - HO","default_warehouse":"Marazeem HO Warehouse - MSSHO", "buying_cost_center" : "Main - MSSHO", "selling_cost_center" : "Main - MSSHO","expense_account": "5111 - Cost of Goods Sold - MSSHO","income_account":"Sales - Credit - MSSHO"},
			{"company":"TRADING DIVISION - ELECTRA","default_warehouse":"Electra Trading Warehouse - TDE", "buying_cost_center" : "Main - TDE", "selling_cost_center" : "Main - TDE","expense_account": "5111 - Cost of Goods Sold - TDE","income_account":"Sales - Credit - TDE"},
			{"company":"MEP DIVISION - ELECTRA","default_warehouse":"Electra MEP Warehouse - MEP", "buying_cost_center" : "Main - MEP", "selling_cost_center" : "Main - MEP","expense_account": "5111 - Cost of Goods Sold - MEP","income_account":"Sales - Credit - MEP"},
			{"company":"ELECTRA - BINOMRAN SHOWROOM","default_warehouse":"Electra Binomran Showroom Warehouse - EBO", "buying_cost_center" : "Main - EBO", "selling_cost_center" : "Main - EBO","expense_account": "5111 - Cost of Goods Sold - EBO","income_account":"Sales - Credit - EBO"},
			{"company":"KINGFISHER TRADING AND CONTRACTING COMPANY","default_warehouse":"Kingfisher Warehouse - KTCC", "buying_cost_center" : "Main - KTCC", "selling_cost_center" : "Main - KTCC","expense_account": "Cost of Goods Sold - KTCC","income_account":"Sales - Credit - KTCC"},
			{"company":"KINGFISHER - SHOWROOM","default_warehouse":"Kingfisher Showroom Warehouse - KS", "buying_cost_center" : "Main - KS", "selling_cost_center" : "Main - KS","expense_account": "Cost of Goods Sold - KS","income_account":"Sales - Credit - KS"},
			{"company":"MARAZEEM SECURITY SERVICES - SHOWROOM","default_warehouse":"Marazeem Showroom - MSSS", "buying_cost_center" : "Main - MSSS", "selling_cost_center" : "Main - MSSS","expense_account": "5111 - Cost of Goods Sold - MSSS","income_account":"Sales - Credit - MSSS"},
			{"company":"MARAZEEM SECURITY SERVICES","default_warehouse":"Marazeem Warehouse - MSS", "buying_cost_center" : "Main - MSS", "selling_cost_center" : "Main - MSS","expense_account": "5111 - Cost of Goods Sold - MSS","income_account":"Sales - Credit - MSS"},
			{"company":"ELECTRA - BARWA SHOWROOM","default_warehouse":"Barwa Showroom  - EBS", "buying_cost_center" : "Main - EBS", "selling_cost_center" : "Main - EBS","expense_account": "5111 - Cost of Goods Sold - EBS","income_account":"Sales - Credit - EBS"},
			{"company":"ELECTRA  - NAJMA SHOWROOM","default_warehouse":"Electra Najma Showroom Warehouse - ENS", "buying_cost_center" : "Main - ENS", "selling_cost_center" : "Main - ENS","expense_account": "5111 - Cost of Goods Sold - ENS","income_account":"Sales - Credit - ENS"},
			{"company":"ELECTRA - ALKHOR SHOWROOM","default_warehouse":"Alkhor Showroom Warehouse - EAS", "buying_cost_center" : "Main - EAS", "selling_cost_center" : "Main - EAS","expense_account": "5111 - Cost of Goods Sold - EAS","income_account":"Sales - Credit - EAS"},
   
			
			{"company":"ELECTRA - FIRE ALARM AND PUBLIC ADDRESS","default_warehouse":"Fire Alarm and Public Address - FAAPA", "buying_cost_center" : "Main - FAAPA", "selling_cost_center" : "Main - FAAPA","expense_account": "5111 - Cost of Goods Sold - FAAPA","income_account":"Sales - Credit - FAAPA"},
			{"company":"ELECTRA - UPS DIVISION","default_warehouse":"UPS Division - UPS", "buying_cost_center" : "Main - UPS", "selling_cost_center" : "Main - UPS","expense_account": "Cost of Goods Sold - UPS","income_account":"Sales - Credit - UPS"},
   
			{"company":"INDUSTRIAL TOOLS DIVISION","default_warehouse":"Electra Industrial Tools Warehouse - ITD", "buying_cost_center" : "Main - ITD", "selling_cost_center" : "Main - ITD","expense_account": "5111 - Cost of Goods Sold - ITD","income_account":"Sales - Credit - ITD"},
			{"company":"ELECTRICAL DIVISION - ELECTRA","default_warehouse":"Electra Electrical Warehouse - EDE", "buying_cost_center" : "Main - EDE", "selling_cost_center" : "Main - EDE","expense_account": "5111 - Cost of Goods Sold - EDE","income_account":"Sales - Credit - EDE"},
			{"company":"ENGINEERING DIVISION - ELECTRA","default_warehouse":"Electra Engineering Warehouse - EED", "buying_cost_center" : "Main - EED", "selling_cost_center" : "Main - EED","expense_account": "5111 - Cost of Goods Sold - EED","income_account":"Sales - Credit - EED"},
			{"company": "INTERIOR DIVISION - ELECTRA","default_warehouse":"Electra Interior Warehouse - INE", "buying_cost_center" : "Main - INE", "selling_cost_center" : "Main - INE","expense_account": "5111 - Cost of Goods Sold - INE","income_account":"Sales - Credit - INE"},
			{"company" :"Al - Shaghairi Trading and Contracting Company W.L.L (ELECTRA)","default_warehouse":"Electra Warehouse - ASTCC", "buying_cost_center" : "Main - ASTCC", "selling_cost_center" : "Main - ASTCC","expense_account": "5111 - Cost of Goods Sold - ASTCC","income_account":"Sales - Credit - ASTCC"}
					]
		for company in companies:
			item_default = frappe.db.exists('Item Default',{'parent':doc.item_code,'company':company['company']},'parent')
			if not item_default:
				itemid = frappe.get_doc("Item",doc.item_code)
				itemid.item_default_set = 1
				itemid.append('item_defaults',{
					'company':company['company'],
					'default_warehouse':company['default_warehouse'],
					'buying_cost_center':company['buying_cost_center'],
					'selling_cost_center':company['selling_cost_center'],
					'expense_account':company['expense_account'],
					'income_account':company['income_account'],
				})
				itemid.save(ignore_permissions=True)
			else:
				frappe.db.set_value('Item Default',{'parent':itemid.name,'company':company['company']},'default_warehouse',company['default_warehouse'])
				frappe.db.set_value('Item Default',{'parent':itemid.name,'company':company['company']},'buying_cost_center',company['buying_cost_center'])
				frappe.db.set_value('Item Default',{'parent':itemid.name,'company':company['company']},'selling_cost_center',company['selling_cost_center'])
				frappe.db.set_value('Item Default',{'parent':itemid.name,'company':company['company']},'expense_account',company['expense_account'])
				frappe.db.set_value('Item Default',{'parent':itemid.name,'company':company['company']},'income_account',company['income_account'])
		frappe.db.set_value('Item',doc.item_code,"item_default_set",1)
		
#Create RQ Job of the below mark_default_wh method
@frappe.whitelist()
def enqueue_default_wh():
	enqueue(method=mark_default_wh, queue='long', timeout=13000, is_async=True)


#Default item set while save of the item document
@frappe.whitelist()
def mark_default_wh():
	frappe.db.auto_commit_on_many_writes = 1
	# item_default = frappe.new_doc("Item Default")
	companies = [
			{"company":"KINGFISHER - TRANSPORTATION","default_warehouse":"Kingfisher Transportation Warehouse - KT", "buying_cost_center" : "Main - KT", "selling_cost_center" : "Main - KT","expense_account": "Cost of Goods Sold - KT","income_account":"Sales - Credit - KT"},
			{"company":"STEEL DIVISION - ELECTRA","default_warehouse":"Steel Warehouse - SDE", "buying_cost_center" : "Main - SDE", "selling_cost_center" : "Main - SDE","expense_account": "5111 - Cost of Goods Sold - SDE","income_account":"Sales - Credit - SDE"},
			{"company":"MARAZEEM SECURITY SERVICES - HO","default_warehouse":"Marazeem HO Warehouse - MSSHO", "buying_cost_center" : "Main - MSSHO", "selling_cost_center" : "Main - MSSHO","expense_account": "5111 - Cost of Goods Sold - MSSHO","income_account":"Sales - Credit - MSSHO"},
			{"company":"TRADING DIVISION - ELECTRA","default_warehouse":"Electra Trading Warehouse - TDE", "buying_cost_center" : "Main - TDE", "selling_cost_center" : "Main - TDE","expense_account": "5111 - Cost of Goods Sold - TDE","income_account":"Sales - Credit - TDE"},
			{"company":"MEP DIVISION - ELECTRA","default_warehouse":"Electra MEP Warehouse - MEP", "buying_cost_center" : "Main - MEP", "selling_cost_center" : "Main - MEP","expense_account": "5111 - Cost of Goods Sold - MEP","income_account":"Sales - Credit - MEP"},
			{"company":"ELECTRA - BINOMRAN SHOWROOM","default_warehouse":"Electra Binomran Showroom Warehouse - EBO", "buying_cost_center" : "Main - EBO", "selling_cost_center" : "Main - EBO","expense_account": "5111 - Cost of Goods Sold - EBO","income_account":"Sales - Credit - EBO"},
			{"company":"KINGFISHER TRADING AND CONTRACTING COMPANY","default_warehouse":"Kingfisher Warehouse - KTCC", "buying_cost_center" : "Main - KTCC", "selling_cost_center" : "Main - KTCC","expense_account": "Cost of Goods Sold - KTCC","income_account":"Sales - Credit - KTCC"},
			{"company":"KINGFISHER - SHOWROOM","default_warehouse":"Kingfisher Showroom Warehouse - KS", "buying_cost_center" : "Main - KS", "selling_cost_center" : "Main - KS","expense_account": "Cost of Goods Sold - KS","income_account":"Sales - Credit - KS"},
			{"company":"MARAZEEM SECURITY SERVICES - SHOWROOM","default_warehouse":"Marazeem Showroom - MSSS", "buying_cost_center" : "Main - MSSS", "selling_cost_center" : "Main - MSSS","expense_account": "5111 - Cost of Goods Sold - MSSS","income_account":"Sales - Credit - MSSS"},
			{"company":"MARAZEEM SECURITY SERVICES","default_warehouse":"Marazeem Warehouse - MSS", "buying_cost_center" : "Main - MSS", "selling_cost_center" : "Main - MSS","expense_account": "5111 - Cost of Goods Sold - MSS","income_account":"Sales - Credit - MSS"},
			{"company":"ELECTRA - BARWA SHOWROOM","default_warehouse":"Barwa Showroom  - EBS", "buying_cost_center" : "Main - EBS", "selling_cost_center" : "Main - EBS","expense_account": "5111 - Cost of Goods Sold - EBS","income_account":"Sales - Credit - EBS"},
			{"company":"ELECTRA  - NAJMA SHOWROOM","default_warehouse":"Electra Najma Showroom Warehouse - ENS", "buying_cost_center" : "Main - ENS", "selling_cost_center" : "Main - ENS","expense_account": "5111 - Cost of Goods Sold - ENS","income_account":"Sales - Credit - ENS"},
			{"company":"ELECTRA - ALKHOR SHOWROOM","default_warehouse":"Alkhor Showroom Warehouse - EAS", "buying_cost_center" : "Main - EAS", "selling_cost_center" : "Main - EAS","expense_account": "5111 - Cost of Goods Sold - EAS","income_account":"Sales - Credit - EAS"},
   
			{"company":"ELECTRA - FIRE ALARM AND PUBLIC ADDRESS","default_warehouse":"Fire Alarm and Public Address - FAAPA", "buying_cost_center" : "Main - FAAPA", "selling_cost_center" : "Main - FAAPA","expense_account": "5111 - Cost of Goods Sold - FAAPA","income_account":"Sales - Credit - FAAPA"},
			{"company":"ELECTRA - UPS DIVISION","default_warehouse":"UPS Division - UPS", "buying_cost_center" : "Main - UPS", "selling_cost_center" : "Main - UPS","expense_account": "Cost of Goods Sold - UPS","income_account":"Sales - Credit - UPS"},
   
			{"company":"INDUSTRIAL TOOLS DIVISION","default_warehouse":"Electra Industrial Tools Warehouse - ITD", "buying_cost_center" : "Main - ITD", "selling_cost_center" : "Main - ITD","expense_account": "5111 - Cost of Goods Sold - ITD","income_account":"Sales - Credit - ITD"},
			{"company":"ELECTRICAL DIVISION - ELECTRA","default_warehouse":"Electra Electrical Warehouse - EDE", "buying_cost_center" : "Main - EDE", "selling_cost_center" : "Main - EDE","expense_account": "5111 - Cost of Goods Sold - EDE","income_account":"Sales - Credit - EDE"},
			{"company":"ENGINEERING DIVISION - ELECTRA","default_warehouse":"Electra Engineering Warehouse - EED", "buying_cost_center" : "Main - EED", "selling_cost_center" : "Main - EED","expense_account": "5111 - Cost of Goods Sold - EED","income_account":"Sales - Credit - EED"},
			{"company": "INTERIOR DIVISION - ELECTRA","default_warehouse":"Electra Interior Warehouse - INE", "buying_cost_center" : "Main - INE", "selling_cost_center" : "Main - INE","expense_account": "5111 - Cost of Goods Sold - INE","income_account":"Sales - Credit - INE"},
			{"company" :"Al - Shaghairi Trading and Contracting Company W.L.L (ELECTRA)","default_warehouse":"Electra Warehouse - ASTCC", "buying_cost_center" : "Main - ASTCC", "selling_cost_center" : "Main - ASTCC","expense_account": "5111 - Cost of Goods Sold - ASTCC","income_account":"Sales - Credit - ASTCC"}
					]

	items = frappe.get_all('Item',limit=50)
	for item in items:
		print(item)
		for company in companies:
			item_default = frappe.db.exists('Item Default',{'parent':item.name,'company':company['company']},'parent')
			if not item_default:
				itemid = frappe.get_doc("Item",item.name)
				itemid.shelf_life_in_days = 0
				itemid.append('item_defaults',{
					'company':company['company'],
					'default_warehouse':company['default_warehouse'],
					'buying_cost_center':company['buying_cost_center'],
					'selling_cost_center':company['selling_cost_center'],
					'expense_account':company['expense_account'],
					'income_account':company['income_account']
				})
				itemid.flags.ignore_mandatory = True
				itemid.save(ignore_permissions=True)
				frappe.db.commit()
			else:
				frappe.db.set_value('Item Default',{'parent':item.name,'company':company['company']},'default_warehouse',company['default_warehouse'])
				frappe.db.set_value('Item Default',{'parent':item.name,'company':company['company']},'buying_cost_center',company['buying_cost_center'])
				frappe.db.set_value('Item Default',{'parent':item.name,'company':company['company']},'selling_cost_center',company['selling_cost_center'])
				frappe.db.set_value('Item Default',{'parent':item.name,'company':company['company']},'expense_account',company['expense_account'])
				frappe.db.set_value('Item Default',{'parent':item.name,'company':company['company']},'income_account',company['income_account'])
				frappe.db.commit()
		# frappe.db.set_value('Item',item.name,"item_default_set",1)
	frappe.db.auto_commit_on_many_writes = 0


#Get the company to permission with given user
@frappe.whitelist()
@frappe.validate_and_sanitize_search_inputs
def get_company_users(doctype, txt, searchfield, start, page_len, filters):
	cond = ''
	if filters and filters.get('company'):
		parent_company = frappe.db.get_value('Company',{'name': filters.get('company')}, 'parent_company')
	query = frappe.db.sql(
		"""select u.name from `tabUser` u inner join `tabEmployee` e on e.user_id = u.name 
			and e.company in ('%s','%s') """ %(filters.get('company'),parent_company)
					)
	return query

#Automatically create the project warehouse while save of the project
@frappe.whitelist()
def create_project_warehouse(doc,method):
	parent_warehouse = frappe.get_value("Warehouse",{"company":doc.company,"is_group": 1})
	if not frappe.db.exists('Warehouse',{"warehouse_name":doc.name}):
		wh = frappe.new_doc("Warehouse")
		wh.update({
			"warehouse_name" : doc.name,
			"project_name":doc.name,
			"company": doc.company,
			"parent_warehouse": parent_warehouse
		})
		wh.warehouse_type='Project'
		wh.save(ignore_permissions=True)

#Set the cost value in designation while save of the employee
@frappe.whitelist()
def manpower_avg_cost_calculation(doc,method):
	designation = doc.designation
	if designation:
		avg_phc = 0
		total_per_hour_cost = frappe.get_all("Employee",{'status': 'Active','designation':designation},['sum(per_hour_cost) as phc'])
		employees = frappe.db.count("Employee",{'status': 'Active','designation':designation},['name'])
		if employees:
			avg_phc = total_per_hour_cost[0]['phc'] / employees
		frappe.db.set_value("Designation",designation,"per_hour_cost",avg_phc)
		frappe.db.commit()

#Set the cost value in designation
@frappe.whitelist()
def bulk_manpower_avg_cost_calculation():
	designations = frappe.get_all('Designation')
	for designation in designations:
		designation = designation['name']
		if designation:
			avg_phc = 0
			total_per_hour_cost = frappe.get_all("Employee",{'status': 'Active','designation':designation},['sum(per_hour_cost) as phc'])
			employees = frappe.db.count("Employee",{'status': 'Active','designation':designation},['name'])
			if employees:
				avg_phc = total_per_hour_cost[0]['phc'] / employees
			frappe.db.set_value("Designation",designation,"per_hour_cost",avg_phc)
			frappe.db.commit()

#Update the below values in Project while submission of SO
@frappe.whitelist()
def create_project_from_so(doc,method):
	if doc.order_type == 'Project':
		series = get_series(doc.company,"Project")
		project = frappe.new_doc("Project")
		project.update({
			"company": doc.company,
			"naming_series": series,
			"project_name": doc.title_of_project,
			"customer":doc.customer,
			"sales_order":doc.name,
			"budgeting": frappe.get_value("Project Budget",{"sales_order":doc.name},"name")
		})
		project.save(ignore_permissions=True)
		frappe.db.commit()
		bom = frappe.db.get_list("BOM",{"project_budget":doc.project_budget},["name"])
		if bom:
			for i in bom:
				bo = frappe.get_doc("BOM",i.name)
				bo.project = project.name
				bo.save(ignore_permissions=True) 

@frappe.whitelist(allow_guest=True)
def ping():
	return 'Pong'

#Combine the all details from PB to SO while click the combine button
@frappe.whitelist()
def update_project_budget(so,name,amount):
	if so:
		
		sales = frappe.get_doc("Sales Order", so)
		if sales.custom_sow_item_table ==1:
			pro=frappe.get_doc("Project Budget",name)
			sales.items = []
			for i in pro.master_scope_of_work:
				sales.append("items", {
					'item_code': i.msow,
					'item_name': i.msow,
					'description': i.msow_desc,
					'qty': i.qty,
					'rate': i.unit_price,
					'amount': i.total_bidding_price,
					'uom': i.unit,
					'conversion_factor':1
				})
			sales.save()
			frappe.db.set_value("Sales Order",so,"grand_total",amount)
			frappe.db.set_value("Sales Order",so,"total",amount)
			
			frappe.db.commit()
			return frappe.msgprint(f"Sales Order {so} has been updated with the new items.")
	
	
#If rate is 0 throw the below message while submission of DN
@frappe.whitelist()
def submit_dummy_dn(doc,method):
	if doc.grand_total <= 0 and doc.dummy_dn == 0 and not doc.is_return:
		frappe.throw("Cannot deliver with zero rate unless it is a dummy DN")

# @frappe.whitelist()
# def stockpopup(item_code):
#     item = frappe.get_value('Item',{'item_code':item_code},'item_code')
#     stocks = frappe.db.sql("""select actual_qty,warehouse,stock_uom,stock_value from tabBin
#         where item_code = '%s' """%(item),as_dict=True)
#     frappe.errprint(stocks[8])
#     data = ''
#     data += '<table class="table table-bordered"><tr><th style="padding:1px;border:1px solid black;font-size:14px;background-color:#e35310;color:white;" colspan=8><center><b>STOCK DETAILS</b></center></th></tr>'
#     data += '<tr><td colspan=1 style="border: 1px solid black;font-size:12px;"><center><b>ITEM CODE</b><center></td></tr>'
#     data += '<tr><td colspan=1 style="border: 1px solid black;font-size:12px;"><center>%s<center><center></td></tr>'%(stocks[2])
#     data += '</table>'
#     return item

#Show the below details in HTML View while click the button of stock details in QN
@frappe.whitelist()
def get_stock_details(item_details,company):
	item_details = json.loads(item_details)
	data = ''
	data += '<table class="table table-bordered"><tr><th style="padding:1px;border:1px solid black;font-size:14px;background-color:#e35310;color:white;" colspan=8><center><b>STOCK DETAILS</b></center></th></tr>'
	data +='<tr><td colspan=1 style="border: 1px solid black;font-size:12px;"><center><b>ITEM CODE</b><center></td><td colspan=1 style="border: 1px solid black;font-size:12px;"><center><b>ITEM NAME</b><center></td><td colspan=1 style="border: 1px solid black;font-size:12px;"><center><b>IN WAREHOUSE</b><center></td><td colspan=1 style="border: 1px solid black;font-size:12px;"><center><b>IN TRANSIT</b><center></td><td colspan=1 style="border: 1px solid black;font-size:12px;"><center><b>TOTAL</b><center></td></tr>'
	
	for j in item_details:
		country = frappe.get_value("Company",{"name":company},["country"])
		warehouse_stock = frappe.db.sql("""
		select sum(b.actual_qty) as qty from `tabBin` b 
		join `tabWarehouse` wh on wh.name = b.warehouse
		join `tabCompany` c on c.name = wh.company
		where c.country = '%s' and b.item_code = '%s'
		""" % (country,j["item_code"]),as_dict=True)[0]
		if not warehouse_stock["qty"]:
			warehouse_stock["qty"] = 0
		purchase_order = frappe.db.sql("""select sum(`tabPurchase Order Item`.qty) as qty from `tabPurchase Order`
				left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
				where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus != 2 """%(j["item_code"]),as_dict=True)[0] or 0 
		if not purchase_order["qty"]:
			purchase_order["qty"] = 0
		purchase_receipt = frappe.db.sql("""select sum(`tabPurchase Receipt Item`.qty) as qty from `tabPurchase Receipt`
				left join `tabPurchase Receipt Item` on `tabPurchase Receipt`.name = `tabPurchase Receipt Item`.parent
				where `tabPurchase Receipt Item`.item_code = '%s' and `tabPurchase Receipt`.docstatus = 1 """%(j["item_code"]),as_dict=True)[0] or 0 
		if not purchase_receipt["qty"]:
			purchase_receipt["qty"] = 0
		in_transit = purchase_order["qty"] - purchase_receipt["qty"]
		total = warehouse_stock["qty"] + in_transit
		data+='<tr><td colspan=1 style="border: 1px solid black;font-size:12px;"><center>%s<center><center></td><td colspan=1 style="border: 1px solid black;font-size:12px;"><center>%s<center><center></td><td colspan=1 style="border: 1px solid black;font-size:12px;"><center>%s<center><center></td><td colspan=1 style="border: 1px solid black;font-size:12px;"><center>%s</center></td><td colspan=1 style="border: 1px solid black;font-size:12px;"><center>%s</center></td></tr>'%(j["item_code"],j["item_name"],warehouse_stock["qty"], in_transit,total)
	
	data += '</table>'
	return data

#Show the below details in HTML View while click the button of stock details in SI
@frappe.whitelist()
def getstock_detail(item_details,company):
	item_details = json.loads(item_details)
	data = ''
	data += '<h4><center><b>STOCK DETAILS</b></center></h4>'
	data += '<h6>Note:</h6>'
	data += '<table style = font-size:10px width=100% ><tr><td>Electra Interior Warehouse - <b>INE</b></td><td>Kingfisher Transportation Warehouse - <b>KT</b></td><td>Steel Warehouse - <b>SDE</b></td><td>Marazeem HO Warehouse - <b>MSSHO</b></td><td>Electra Trading Warehouse - <b>TDE</b></td></tr>'
	data += '<tr><td>Electra Warehouse - <b>ASTCC</b></td><td>Electra Binomran Showroom Warehouse - <b>EBO</b></td><td>Kingfisher Warehouse - <b>KTCC</b></td><td>Kingfisher Showroom Warehouse - <b>KS</b></td><td>Marazeem Showroom - <b>MSSS</b></td></tr>'
	data += '<tr><td>Electra Najma Showroom Warehouse - <b> ENS</b></td><td>Marazeem Warehouse - <b>MSS</b></td><td>Barwa Showroom  - <b>EBS</b></td><td>Electra Electrical Warehouse - <b>EDE</b></td><td>Electra Engineering Warehouse - <b>EED</b></td></tr>'
	data += '</table>'
	for j in item_details:
		country = frappe.get_value("Company",{"name":company},["country"])

		warehouse_stock = frappe.db.sql("""
		select sum(b.actual_qty) as qty from `tabBin` b join `tabWarehouse` wh on wh.name = b.warehouse join `tabCompany` c on c.name = wh.company where c.country = '%s' and b.item_code = '%s'
		""" % (country,j["item_code"]),as_dict=True)[0]

		if not warehouse_stock["qty"]:
			warehouse_stock["qty"] = 0
		
		
		new_po = frappe.db.sql("""select sum(`tabPurchase Order Item`.qty) as qty,sum(`tabPurchase Order Item`.received_qty) as d_qty from `tabPurchase Order` 
		left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
		where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus = 1 and `tabPurchase Order`.status != 'Closed' """ % (j["item_code"]), as_dict=True)[0]
		if not new_po['qty']:
			new_po['qty'] = 0
		if not new_po['d_qty']:
			new_po['d_qty'] = 0
		in_transit = new_po['qty'] - new_po['d_qty']


		
		total = warehouse_stock["qty"] + in_transit

		stocks = frappe.db.sql("""select actual_qty,warehouse,stock_uom,stock_value from tabBin
		where item_code = '%s' """%(j["item_code"]),as_dict=True)

		pos = frappe.db.sql("""select `tabPurchase Order Item`.item_code as item_code,`tabPurchase Order Item`.item_name as item_name,`tabPurchase Order`.supplier as supplier,sum(`tabPurchase Order Item`.qty) as qty,`tabPurchase Order Item`.rate as rate,`tabPurchase Order`.transaction_date as date,`tabPurchase Order`.name as po from `tabPurchase Order`
		left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
		where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus != 2 order by rate asc limit 1""" % (j["item_code"]), as_dict=True)
	
		new_so = frappe.db.sql("""select sum(`tabSales Order Item`.qty) as qty,sum(`tabSales Order Item`.delivered_qty) as d_qty from `tabSales Order`
		left join `tabSales Order Item` on `tabSales Order`.name = `tabSales Order Item`.parent
		where `tabSales Order Item`.item_code = '%s' and `tabSales Order`.docstatus = 1 and `tabSales Order`.status != "Closed" """ % (j["item_code"]), as_dict=True)[0]
		if not new_so['qty']:
			new_so['qty'] = 0
		if not new_so['d_qty']:
			new_so['d_qty'] = 0
		del_total = new_so['qty'] - new_so['d_qty']
			
		
		i = 0
		for po in pos:
			if pos:
				data += '<table class="table table-bordered">'
				data += '<tr>'
				data += '<td colspan=1 style="width:13%;padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e35310;color:white;"><center><b>ITEM CODE</b><center></td>'
				data += '<td colspan=1 style="width:33%;padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e35310;color:white;"><center><b>ITEM NAME</b><center></td>'
				data += '<td colspan=1 style="width:70px;padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e35310;color:white;"><center><b>STOCK</b><center></td>'

				for stock in stocks:
					if stock.actual_qty > 0:
						wh = stock.warehouse
						x = wh.split('- ')
						data += '<td colspan=1 style="width:70px;padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e35310;color:white;"><center><b>%s</b><center></td>'%(x[-1])
				data += '<td colspan=1 style="padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e35310;color:white;"><center><b>PENDING TO RECEIVE</b><center></td>'
				data += '<td colspan=1 style="padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e35310;color:white;"><center><b>PENDING TO SELL</b><center></td>'
				data += '</tr>'
				
				
				
				data +='<tr>'
				data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>'%(j["item_code"])
				data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>'%(j["item_name"])
				data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>'%(warehouse_stock['qty'] or 0)
				for stock in stocks:
					if stock.actual_qty > 0:
						data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>'%(stock.actual_qty)
				data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>'%(in_transit or 0)
				data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>'%(del_total or 0)
				data += '</tr>'
			i += 1
		data += '</table>'
			
	return data

#Show the below details in HTML View while click the button of stock details in DN,QN,SO
@frappe.whitelist()
def detail_stock(item_details, company):
	item_details = json.loads(item_details)
	data = ''
	data += '<h4><center><b>STOCK DETAILS</b></center></h4>'
	data += '<h6>Note:</h6>'
	

	data += '<table style = font-size:10px width=100% ><tr><td>Electra Warehouse - <b>ASTCC</b></td><td>Electra Trading Warehouse - <b>TDE</b></td><td>Electra Interior Warehouse - <b>INE</b></td><td>Steel Warehouse - <b>SDE</b></td><td>Electra Electrical Warehouse - <b>EDE</b></td></tr>'
	data += '<tr><td>Electra Engineering Warehouse - <b>EED</b></td><td>Barwa Showroom  - <b>EBS</b></td><td>Electra Najma Showroom Warehouse - <b> ENS</b></td><td>Electra Binomran Showroom Warehouse - <b>EBO</b></td><td>Alkhor Showroom Warehouse - <b>EAS</b></td></tr>'
	data += '<tr><td>Electra Industrial Tools Warehouse - <b> ITD</b></td><td>Kingfisher Showroom Warehouse - <b>KS</b></td><td>Kingfisher Transportation Warehouse - <b>KT</b></td><td>Kingfisher Warehouse - <b>KTCC</b></td><td>Marazeem HO Warehouse - <b>MSSHO</b></td></tr>'
	data += '<tr><td>Marazeem Warehouse - <b>MSS</b></td><td>Marazeem Showroom - <b>MSSS</b></td></tr>'


	
	data += '</table>'
	data += '<table class="table table-bordered" style = font-size:10px>'
	data += '<td colspan=1 style="width:12%;padding:1px;border:1px solid black;background-color:#e35310;color:white;"><center><b>ITEM CODE</b></center></td>'
	data += '<td colspan=1 style="width:20%;padding:1px;border:1px solid black;background-color:#e35310;color:white;"><center><b>ITEM NAME</b></center></td>'
	data += '<td colspan=1 style="width:70px;padding:1px;border:1px solid black;background-color:#e35310;color:white;"><center><b>STOCK</b></center></td>'
	comp = frappe.db.get_list("Company","name")
	for co in comp:
		st = 0
		ware = frappe.db.get_list("Warehouse",{"company":co.name,"default_for_stock_transfer":1},['name'])
		for w in ware:
			data += '<td colspan=1 style="width:70px;padding:1px;border:1px solid black;background-color:#e35310;color:white;"><center><b>%s</b></center></td>'%(w.name.split("-")[-1]) 
	data += '<td colspan=1 style="width:180px;padding:1px;border:1px solid black;background-color:#e35310;color:white;"><center><b>TO RECEIVE</b></center></td>'
	data += '<td colspan=1 style="width:180px;padding:1px;border:1px solid black;background-color:#e35310;color:white;"><center><b>TO SELL</b></center></td>'
	warehouses = []  
	for j in item_details:
		country = frappe.get_value("Company",{"name":company},["country"])

		warehouse_stock = frappe.db.sql("""
		select sum(b.actual_qty) as qty from `tabBin` b join `tabWarehouse` wh on wh.name = b.warehouse join `tabCompany` c on c.name = wh.company where c.country = '%s' and b.item_code = '%s'
		""" % (country,j["item_code"]),as_dict=True)[0]

		if not warehouse_stock["qty"]:
			warehouse_stock["qty"] = 0
		
		
		new_po = frappe.db.sql("""select sum(`tabPurchase Order Item`.qty) as qty,sum(`tabPurchase Order Item`.received_qty) as d_qty from `tabPurchase Order` 
		left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
		where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus = 1 and `tabPurchase Order`.status != 'Closed' """ % (j["item_code"]), as_dict=True)[0]
		if not new_po['qty']:
			new_po['qty'] = 0
		if not new_po['d_qty']:
			new_po['d_qty'] = 0
		in_transit = new_po['qty'] - new_po['d_qty']


		
		total = warehouse_stock["qty"] + in_transit

		stocks = frappe.db.sql("""select actual_qty,warehouse,stock_uom,stock_value from tabBin
		where item_code = '%s' """%(j["item_code"]),as_dict=True)

		pos = frappe.db.sql("""select `tabPurchase Order Item`.item_code as item_code,`tabPurchase Order Item`.item_name as item_name,`tabPurchase Order`.supplier as supplier,sum(`tabPurchase Order Item`.qty) as qty,`tabPurchase Order Item`.rate as rate,`tabPurchase Order`.transaction_date as date,`tabPurchase Order`.name as po from `tabPurchase Order`
		left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
		where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus != 2 order by rate asc limit 1""" % (j["item_code"]), as_dict=True)
	
		new_so = frappe.db.sql("""select sum(`tabSales Order Item`.qty) as qty,sum(`tabSales Order Item`.delivered_qty) as d_qty from `tabSales Order`
		left join `tabSales Order Item` on `tabSales Order`.name = `tabSales Order Item`.parent
		where `tabSales Order Item`.item_code = '%s' and `tabSales Order`.docstatus = 1 and `tabSales Order`.status != "Closed"  """ % (j["item_code"]), as_dict=True)[0]
		if not new_so['qty']:
			new_so['qty'] = 0
		if not new_so['d_qty']:
			new_so['d_qty'] = 0
		del_total = new_so['qty'] - new_so['d_qty']
		i = 0
		for po in pos:
			data += '<tr>'
			data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>' % (j["item_code"])
			data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>' % (j["item_name"])
			data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>' % (warehouse_stock['qty'] or 0)
			comp = frappe.db.get_list("Company","name")
			for co in comp:
				st = 0
				ware = frappe.db.get_list("Warehouse",{"company":co.name,"default_for_stock_transfer":1},['name'])
				for w in ware:
					sto = frappe.db.get_value("Bin",{"item_code":j["item_code"],"warehouse":w.name},['actual_qty'])
					if not sto:
						sto = 0
					st += sto
					data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>' %(st)
			data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>'%(in_transit or 0)
			data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>'%(del_total or 0)
			data += '</tr>'
		i += 1
	data += '</tr>'
	data += '</table>'
	return data



#Show the below details in HTML View while click the button oF Margin details in QN
@frappe.whitelist()
def get_item_margin(item_details,company,discount_amount):
	item_details = json.loads(item_details)
	data_4 = ''
	data_4 += '''<table class="table">
	<tr>
	<th style="padding:1px;border: 1px solid black;font-size:14px;background-color:#FF4500;color:white;" colspan =9 ><center><b>MARGIN</b></center></th>
	</tr>'''

	data_4+='''<tr>
	<td colspan=1 style="border: 1px solid black;font-size:12px;"><b>ITEM</b></td>
	<td colspan=1 style="border: 1px solid black;font-size:12px;"><b>ITEM NAME</b></td>
	<td colspan=1 style="border: 1px solid black;font-size:12px;"><b><center>QTY</center></b></td>
	<td colspan=1 style="border: 1px solid black;font-size:12px;"><b><center>Cost</center></b></td>
	<td colspan=1 style="border: 1px solid black;font-size:12px;"><b><center>Rate</center></b></td>
	<td colspan=1 style="border: 1px solid black;font-size:12px;"><b><center>Cost Amount</center></b></td>
	<td colspan=1 style="border: 1px solid black;font-size:12px;"><b><center>Selling Amount</center></b></td>
	<td colspan=1 style="border: 1px solid black;font-size:12px;"><b><center>Profit Amount</center></b></td>
	<td colspan=1 style="border: 1px solid black;font-size:12px;"><b><center>Profit %</center></b></td>
	</tr>'''
	cost_amount = 0
	total_selling_price = 0
	rate = 0
	cost = 0
	selling_amount = 0
	profit_amount_ = 0
	selling_amount_ = 0
	qty_tot = 0
	tot_per = 0
	no_item = 0
	for i in item_details:
		qty_tot += i["qty"]
		rate += i["base_net_rate"]
		selling_amount += i["base_net_amount"]
		total_selling_price = (i["base_net_rate"] * i["qty"]) + total_selling_price
		
		warehouse_stock = frappe.db.sql("""
			SELECT valuation_rate as vr 
			FROM `tabBin` b 
			JOIN `tabWarehouse` wh ON wh.name = b.warehouse
			JOIN `tabCompany` c ON c.name = wh.company
			WHERE b.item_code = %s AND c.name = %s AND wh.name = %s
			""", (i["item_code"], company, i["warehouse"]), as_dict=True)
		
		if warehouse_stock:
			warehouse_stock = warehouse_stock[0]
			if not warehouse_stock["vr"]:
				warehouse_stock["vr"] = 0
		else:
			if frappe.db.get_value("Item Price",{'item_code':i["item_code"],'price_list':'Standard Buying'}):
				std_selling=frappe.db.get_value("Item Price",{'item_code':i["item_code"],'price_list':'Standard Buying'},['price_list_rate'])
				if std_selling:
					warehouse_stock = {"vr": std_selling}
				else:
					warehouse_stock = {"vr": 0}
			else:
				warehouse_stock = {"vr": 0}

		cost += warehouse_stock["vr"]
		cost_amount = (warehouse_stock["vr"] * i["qty"]) + cost_amount
		cost_amount_ = (warehouse_stock["vr"] * i["qty"])
		
		tot = cost_amount - total_selling_price
		tot_diff = cost_amount_ - i["base_net_amount"]
		
		if cost_amount_ > 0:
			tot_ = tot_diff / cost_amount_ * 100
			tot_per += tot_
			
			data_4 += '''<tr>
				<td colspan=1 style="border: 1px solid black;font-size:12px;">{}</td>
				<td colspan=1 style="border: 1px solid black;font-size:12px;">{}</td>
				<td colspan=1 style="border: 1px solid black;font-size:12px; text-align: right">{}</td>
				<td colspan=1 style="border: 1px solid black;font-size:12px; text-align: right">{}</td>
				<td colspan=1 style="border: 1px solid black;font-size:12px; text-align: right">{}</td>
				<td colspan=1 style="border: 1px solid black;font-size:12px; text-align: right">{}</td>
				<td colspan=1 style="border: 1px solid black;font-size:12px; text-align: right">{}</td>
				<td colspan=1 style="border: 1px solid black;font-size:12px; text-align: right">{}</td>
				<td colspan=1 style="border: 1px solid black;font-size:12px; text-align: right">{}</td>
			</tr>'''.format(i["item_code"], i["description"], i["qty"], round(warehouse_stock["vr"], 2), round(i["base_net_rate"], 2), round(cost_amount_, 2), round(i["base_net_amount"], 2), round(-tot_diff, 2), round(-tot_, 2))
	
	if cost_amount_ == 0:
		total_margin_internal = (total_selling_price - cost_amount_)/100

	else:
		total_margin_internal = (-tot / cost_amount)*100

		data_4 += '''<tr>
		<td colspan=2 style="border: 1px solid black;padding:1px;font-size:14px;text-align: right"><b>Total </b></td>
		<td style="border: 1px solid black;padding:1px;font-size:12px;text-align: right"><b>%s</b></td>
		<td style="border: 1px solid black;padding:1px;font-size:12px;text-align: right"><b>%s</b></td>
		<td style="border: 1px solid black;padding:1px;font-size:12px;text-align: right"><b>%s</b></td>
		<td style="border: 1px solid black;padding:1px;font-size:12px;text-align: right"><b>%s</b></td>
		<td style="border: 1px solid black;padding:1px;font-size:12px;text-align: right"><b>%s</b></td>
		<td style="border: 1px solid black;padding:1px;font-size:12px;text-align: right"><b>%s</b></td>
		<td style="border: 1px solid black;padding:1px;font-size:12px;text-align: right"><b>%s</b></td>
		</tr>'''%(qty_tot,round(cost,2),round(rate,2),round(cost_amount,2),round(selling_amount,2),round(-tot,2),round(total_margin_internal,2))
		
		total_amount = float(discount_amount) + float(selling_amount)
		data_4+= '''<tr>
		<td colspan=2 style="border: 1px solid black;padding:1px;font-size:12px;text-align: right"><b>Before Applying Discount</b></td>
		<td style="border: 1px solid black;padding:1px;font-size:12px;text-align: right"><b>%s</b></td>
		<td colspan=2 style="border: 1px solid black;padding:1px;font-size:12px;text-align: right"><b>Discount Amount </b></td>
		<td style="border: 1px solid black;padding:1px;font-size:12px;text-align: right"><b>%s</b></td>
		<td colspan=2 style="border: 1px solid black;padding:1px;font-size:12px;text-align: right"><b>After Applying Discount</b></td>
		<td style="border: 1px solid black;padding:1px;font-size:12px;text-align: right"><b>%s</b></td>
		</tr>'''%(total_amount,discount_amount,selling_amount)
		
	
	return data_4



# @frappe.whitelist()
# def get_item_margin_percentage_calculation(item_details,company,profit):
#     profit_per = int(profit)
#     item_details = json.loads(item_details)
#     data_4 = ''
#     data_4 += '''<table class="table">
#     <tr>
#     <th style="padding:1px;border: 1px solid black;font-size:14px;background-color:#FF4500;color:white;" colspan =9 ><center><b>MARGIN</b></center></th>
#     </tr>'''

#     data_4+='''<tr>
#     <td colspan=1 style="border: 1px solid black;font-size:12px;"><b>ITEM</b></td>
#     <td colspan=1 style="border: 1px solid black;font-size:12px;"><b>ITEM NAME</b></td>
#     <td colspan=1 style="border: 1px solid black;font-size:12px;"><b><center>QTY</center></b></td>
#     <td colspan=1 style="border: 1px solid black;font-size:12px;"><b><center>Cost</center></b></td>
#     <td colspan=1 style="border: 1px solid black;font-size:12px;"><b><center>Rate</center></b></td>
#     <td colspan=1 style="border: 1px solid black;font-size:12px;"><b><center>Cost Amount</center></b></td>
#     <td colspan=1 style="border: 1px solid black;font-size:12px;"><b><center>Selling Amount</center></b></td>
#     <td colspan=1 style="border: 1px solid black;font-size:12px;"><b><center>Profit Amount</center></b></td>
#     <td colspan=1 style="border: 1px solid black;font-size:12px;"><b><center>Profit %</center></b></td>
#     </tr>'''
#     cost_amount = 0
#     total_selling_price = 0
#     rate = 0
#     cost = 0
#     selling_amount = 0
#     profit_amount_ = 0
#     selling_amount_ = 0
#     qty_tot = 0
#     for i in item_details:
#         qty_tot += i["qty"]
#         rate += i["rate"]
#         selling_amount += i["base_amount"]
#         total_selling_price =  (i["rate"] * i["qty"]) + total_selling_price
#         frappe.errprint(total_selling_price)
#         warehouse_stock = frappe.db.sql("""
#         select valuation_rate as vr from `tabBin` b 
#         join `tabWarehouse` wh on wh.name = b.warehouse
#         join `tabCompany` c on c.name = wh.company
#         where b.item_code = '%s'
#         """ % (i["item_code"]),as_dict=True)[0]
#         cost += warehouse_stock.vr 
#         if not warehouse_stock.vr :
#             warehouse_stock.vr  = 0
#         cost_amount = (warehouse_stock.vr  * i["qty"]) + cost_amount
#         cost_amount_ = (warehouse_stock.vr * i["qty"])
		
#         tot =  cost_amount - total_selling_price
#         tot_diff =  cost_amount_ - i["base_amount"]
		
#         if profit_per > 0:
#             profit_amount_calc = (cost_amount_ * profit_per)/100
#             selling_amount_calc = profit_amount_calc + cost_amount_
#             profit_amount_ += profit_amount_calc
#             selling_amount_ += selling_amount_calc
#         if cost_amount_ > 0:
#             tot_ = tot_diff/cost_amount_ *100
#         if profit_per == 0:
#             data_4+='''<tr>
#             <td colspan=1 style="border: 1px solid black;font-size:12px;">%s</td>
#             <td colspan=1 style="border: 1px solid black;font-size:12px;">%s</td>
#             <td colspan=1 style="border: 1px solid black;font-size:12px; text-align: right">%s</td>
#             <td colspan=1 style="border: 1px solid black;font-size:12px; text-align: right">%s</td>
#             <td colspan=1 style="border: 1px solid black;font-size:12px; text-align: right">%s</td>
#             <td colspan=1 style="border: 1px solid black;font-size:12px; text-align: right">%s</td>
#             <td colspan=1 style="border: 1px solid black;font-size:12px; text-align: right">%s</td>
#             <td colspan=1 style="border: 1px solid black;font-size:12px; text-align: right">%s</td>
#             <td colspan=1 style="border: 1px solid black;font-size:12px; text-align: right">%s</td>
#             </tr>'''%(i["item_code"],i["description"],i["qty"],round(warehouse_stock.vr,2) ,round(i["rate"],2),round(cost_amount_,2),round(i["base_amount"],2),round(-tot_diff,2),round(-tot_,2))
#         elif profit_per > 0:
#             data_4+='''<tr>
#             <td colspan=1 style="border: 1px solid black;font-size:12px;">%s</td>
#             <td colspan=1 style="border: 1px solid black;font-size:12px;">%s</td>
#             <td colspan=1 style="border: 1px solid black;font-size:12px; text-align: right">%s</td>
#             <td colspan=1 style="border: 1px solid black;font-size:12px; text-align: right">%s</td>
#             <td colspan=1 style="border: 1px solid black;font-size:12px; text-align: right">%s</td>
#             <td colspan=1 style="border: 1px solid black;font-size:12px; text-align: right">%s</td>
#             <td colspan=1 style="border: 1px solid black;font-size:12px; text-align: right">%s</td>
#             <td colspan=1 style="border: 1px solid black;font-size:12px; text-align: right">%s</td>
#             <td colspan=1 style="border: 1px solid black;font-size:12px; text-align: right">%s</td>
#             </tr>
#             '''%(i["item_code"],i["description"],i["qty"],round(warehouse_stock.vr,2) ,round(i["rate"],2),round(cost_amount_,2),round(selling_amount_calc,2),round(profit_amount_calc,2),round(profit_per,2))
	
	
	
#     # data_5 = ''
#     if cost_amount_ == 0:
#         total_margin_internal = (total_selling_price - cost_amount_)/100
#         frappe.errprint('total_margin_internal')

#     else:
#         total_margin_internal = (round(-tot_,2))
#         frappe.errprint(total_margin_internal)
#         frappe.errprint('oo')

		

#     if profit_per == 0:
#         data_4 += '''<tr>
#         <td colspan=2 style="border: 1px solid black;padding:1px;font-size:14px;text-align: right"><b>Total </b></td>
#         <td style="border: 1px solid black;padding:1px;font-size:12px;text-align: right"><b>%s</b></td>
#         <td style="border: 1px solid black;padding:1px;font-size:12px;text-align: right"><b>%s</b></td>
#         <td style="border: 1px solid black;padding:1px;font-size:12px;text-align: right"><b>%s</b></td>
#         <td style="border: 1px solid black;padding:1px;font-size:12px;text-align: right"><b>%s</b></td>
#         <td style="border: 1px solid black;padding:1px;font-size:12px;text-align: right"><b>%s</b></td>
#         <td style="border: 1px solid black;padding:1px;font-size:12px;text-align: right"><b>%s</b></td>
#         <td style="border: 1px solid black;padding:1px;font-size:12px;text-align: right"><b>%s</b></td>
#         </tr>'''%(qty_tot,round(cost,2),round(rate,2),round(cost_amount,2),round(selling_amount,2),round(-tot,2),round(total_margin_internal,2))
#         data_4+='''</table>'''
#     elif profit_per > 0:

#         data_4 += '''<tr>
#         <td colspan=2 style="border: 1px solid black;padding:1px;font-size:14px;text-align: right"><b>Total </b></td>
#         <td style="border: 1px solid black;padding:1px;font-size:12px;text-align: right"><b>%s</b></td>
#         <td style="border: 1px solid black;padding:1px;font-size:12px;text-align: right"><b>%s</b></td>
#         <td style="border: 1px solid black;padding:1px;font-size:12px;text-align: right"><b>%s</b></td>
#         <td style="border: 1px solid black;padding:1px;font-size:12px;text-align: right"><b>%s</b></td>
#         <td style="border: 1px solid black;padding:1px;font-size:12px;text-align: right"><b>%s</b></td>
#         <td style="border: 1px solid black;padding:1px;font-size:12px;text-align: right"><b>%s</b></td>
#         <td style="border: 1px solid black;padding:1px;font-size:12px;text-align: right"><b>%s</b></td>
		
#         </tr>'''%(qty_tot,round(cost,2),round(rate,2),round(cost_amount,2),round(selling_amount_,2),round(profit_amount_,2),round(profit_per,2))
#         data_4+='''</table>'''
# # pro/cost 100
	
#     return data_4

# @frappe.whitelist()
# def get_invoice_summary(project):
# 	# data = ''
# 	# data += '<table class="table table-bordered"><tr><th style="padding:1px;border:1px solid black;font-size:14px;background-color:#e35310;color:white;" colspan=8><center><b>INVOICE SUMMARY</b></center></th></tr>'
# 	# data +='<tr><td colspan=1 style="border: 1px solid black;font-size:12px;"><center><b>INVOICE LIST</b><center></td></tr>'
   
# 	invoice = frappe.get_all('Sales Invoice',{"project":project},["*"])
# 	for i in invoice:
	
	# data+='<tr><td colspan=1 style="border: 1px solid black;font-size:12px;"><center>%s<center><center></td></tr>'%(invoice)
	
	# data += '</table>'
	# return 

#Get the days between the given dates in Full and Final Settlement
@frappe.whitelist()
def get_dates(boarding_begins_on,last__days):
	no_of_days = date_diff(add_days(last__days, 1), boarding_begins_on)
	dates = [add_days(boarding_begins_on, i) for i in range(0, no_of_days)]
	return dates

# @frappe.whitelist()
# def calculate_attendance(employee):
#     name_reg = frappe.db.sql(
#         """select name,hods_relieving_date,actual_relieving_date from `tabResignation Form` where employee = %s """ % (employee), as_dict=True)[0]
#     current_date = name_reg.actual_relieving_date
#     first_day_of_month = current_date.replace(day=1)
#     hod_date = str(first_day_of_month)
#     app_date = str(name_reg.actual_relieving_date)
#     if name_reg:
#         att = frappe.db.sql("""select count(*) as count from `tabAttendance` where attendance_date between '%s' and '%s' and status = 'Present'  and employee = %s""" %
#                             (hod_date, app_date, employee), as_dict=True)[0]
#         cal = att
#         return cal, name_reg

# @frappe.whitelist()
# def get_reg_form(employee):
#     name_reg = frappe.db.sql(
#         """select name,hods_relieving_date,actual_relieving_date from `tabResignation Form` where employee = %s """ % (employee), as_dict=True)[0]
#     current_date = name_reg.actual_relieving_date
#     first_day_of_month = current_date.replace(day=1)
#     return name_reg.name, first_day_of_month, name_reg.actual_relieving_date


#Get the gratuity and set to Full and Final Settlement
@frappe.whitelist()
def get_gratuity(employee):
	from datetime import datetime
	from dateutil import relativedelta
	date_2 = datetime.now()
	emp = frappe.get_doc('Employee', employee)
	# Get the interval between two dates
	diff = relativedelta.relativedelta(date_2, emp.date_of_joining)

	exp_years = diff.years
	exp_month = diff.months
	exp_days = diff.days

	basic_salary = frappe.db.get_value(
		'Employee', emp.employee_number, 'basic')

	per_day_basic = basic_salary / 30

	if emp.grade == 'Office Staff':
		gratuity_per_year = per_day_basic * 30
	else:
		gratuity_per_year = per_day_basic * 21

	gratuity_per_month = gratuity_per_year / 12
	gratuity_per_day = gratuity_per_month / 30
	earned_gpy = gratuity_per_year * exp_years
	earned_gpm = gratuity_per_month * exp_month
	earned_gpd = gratuity_per_day * exp_days
	total_gratuity = earned_gpy + earned_gpm + earned_gpd

	return total_gratuity

#Update the employee status while submission of resignation
@frappe.whitelist()
def update_employee_status(doc,method):
	reg = frappe.db.sql(
		"""select * from `tabResignation Form` where docstatus = 1""", as_dict=1)
	if reg:
		for emp in reg:
			if emp.actual_relieving_date == datetime.strptime((today()), '%Y-%m-%d').date():
				emp_n = frappe.get_doc('Employee', emp.employee)
				emp_n.status = "Left"
				emp_n.relieving_date = emp.actual_relieving_date
				emp_n.save(ignore_permissions=True)


@frappe.whitelist(allow_guest=True)
def bluetooth_print_format():
	a = []
	b = {}
	b['type'] = 0
	b['content'] = 'My Title'	
	# b.bold = 1
	# b.align = 1
	b['format'] = 0; # 0 if normal, 1 if double Height, 2 if double Height + Width, 3 if double Width, 4 if small
	a.append(b)
 
	# Dictionary to JSON Object using dumps() method
	# Return JSON Object
	return json.dumps(a)

@frappe.whitelist()
# def enqueue_delete_items():
#     enqueue(method=delete_items, queue="long", timeout=6000, is_async=True)
#Delete items to match the below condition
def delete_items():
	items_list = frappe.get_all('Item',['name'])
	for it in items_list:
		item = frappe.get_doc('Item',it['name'])
		item.is_stock_item = 1
		item.flags.ignore_links = True
		item.save(ignore_permissions=True)


def delete_linked():
	linked_doctypes = [
			"Quotation Item",
			"Opportunity Item"
			
		]

	for lin in linked_doctypes:
		dle = '`tab%s`' % lin
		frappe.db.sql("""delete from %s """ %dle)

#Get the payment terms template while enter the payment_terms_temp in QN
@frappe.whitelist()
def get_pay_terms(payment):
	pay = frappe.get_doc("Payment Terms Template",payment)
	return pay.terms

#Get the exchange currency while enter the currency in consolidated payment request
@frappe.whitelist()
def get_currency_exchange(currency):
	conversion = get_exchange_rate(currency, "QAR")
	return conversion


#Get norden items details and show to product search
@frappe.whitelist(allow_guest=True)
def get_norden_item(item):
	url = "https://erp.nordencommunication.com/api/method/norden.custom.get_electra_details?item=%s" % (item)
	headers = { 'Content-Type': 'application/json','Authorization': 'token 28a1f5da5dffd46:812f4d4d2671af2'}
	# params = {"limit_start": 0,"limit_page_length": 20000}

	response = requests.request('GET',url,headers=headers)
	res = json.loads(response.text)
	return res
#Get norden items details and show to product search
@frappe.whitelist(allow_guest=True)
def get_norden_item_without_cost(item):
	url = "https://erp.nordencommunication.com/api/method/norden.custom.get_electra_details_without_cost?item=%s" % (item)
	headers = { 'Content-Type': 'application/json','Authorization': 'token 28a1f5da5dffd46:812f4d4d2671af2'}
	# params = {"limit_start": 0,"limit_page_length": 20000}

	response = requests.request('GET',url,headers=headers)
	res = json.loads(response.text)
	return res

#Get the below value while enter the item code in qn
@frappe.whitelist()
def get_stock_av(item_code,abbr):
	item = frappe.db.get_value("Item",item_code,abbr)
	return item

#Get the default letter head while enter the company
@frappe.whitelist()
def get_default_letter_head(company):
	letter_head = frappe.db.get_value("Company",company,['default_letter_head'])
	return letter_head

#Set the series of SI While save of the document
@frappe.whitelist()
def get_sales_invoice_series(company,doctype):
	company_series = frappe.db.get_value("Company Series",{'company':company,'document_type':doctype},['series'])
	series = company_series.split('-')
	ser = series[0]+"-"+ "STI" +"-"+ series[2]+ "-"
	return ser

#Get the role of the given user
@frappe.whitelist()
def get_user_role(user):
	user_roles = frappe.get_roles(user)
	return user_roles
	
#Set the series of return DN While save of the document
@frappe.whitelist()
def get_dn_return_series(company,doctype):
	company_series = frappe.db.get_value("Company Series",{'company':company,'document_type':doctype},'series')
	series = company_series.split('-')
	ser = series[0]+"-"+ "DNR" +"-"+ series[2]+ "-"
	return ser

#Check the below condition and throw message in SO
@frappe.whitelist()
def restrict_general_item_so(doc,method):
	if doc.items:
		for item in doc.items:
			if item.item_code == "001":
				frappe.throw("Not Allowed to Make Sales Order in General Entry")

#Check the below condition and throw message in SI
@frappe.whitelist()
def restrict_general_item_si(doc,method):
	if doc.items:
		for item in doc.items:
			if item.item_code == "001":
				frappe.throw("Not Allowed to Make Sales Invoice in General Entry")

#Check the below condition and throw message in DN
@frappe.whitelist()
def restrict_general_item_dn(doc,method):
	if doc.items:
		for item in doc.items:
			if item.item_code == "001":
				frappe.throw("Not Allowed to Make Delivery Note in General Entry")

# def setup_hooks():
#     from frappe.model import DocType
#     if "Sales Order" in DocType:
#         DocType["Sales Order"].before_save.append(restrict_general_item)

# # Execute the setup_hooks function
# setup_hooks()

from frappe.utils import fmt_money
from electra.project_details import cost_till_date,project_cost_till_date,total_jl_cost

#Get the below details and show to project
@frappe.whitelist()
def actual_vs_budgeted(name):
	crea = frappe.db.sql("""select creation from `tabProject` where name ='%s' """ % (name), as_dict=True)[0]
	project_budget, p_name = frappe.db.get_value("Project", {'name': name}, ['budgeting', 'project_name'])
	pb = frappe.get_doc("Project Budget", project_budget)
	data = '<table width = 100%>'
	data += '<tr><td><b>Date</b></td><td>%s</td><td><b>Refer No</b></td><td></td></tr>' % (
		format_date(crea["creation"].date()))
	data += '<tr><td><b>Project Code</b></td><td>%s</td><td><b>Project</b></td><td>%s</td></tr>' % (name, p_name)
	data += '<tr><td><b>Client</b></td><td>%s</td><td><b>Order Ref No.</b></td><td>%s</td></tr>' % (
		pb.lead_customer_name, pb.sales_order)
	data += '</table><br>'
	grand = frappe.db.get_value("Sales Order", {'project': name, 'docstatus': 1}, 'grand_total')
	sales_in = frappe.db.sql(
		""" SELECT SUM(custom_total_invoice_amount) as total FROM `tabSales Invoice` WHERE project = %s and docstatus = 1""",
		(name,),
		as_dict=True)[0]
	if not sales_in["total"]:
		sales_in["total"] = 0
	bal = grand - sales_in["total"]
	revenue = frappe.db.sql(
		""" SELECT SUM(paid_amount) as amt FROM `tabPayment Entry` WHERE project = %s""",
		(name,),
		as_dict=True)[0]
	if not revenue["amt"]:
		revenue["amt"] = 0
	out = grand - revenue["amt"]
	tot = 0
	dn_list = frappe.db.get_list("Delivery Note", {'project': name})
	if dn_list:
		for d_list in dn_list:
			dn = frappe.get_doc("Delivery Note", d_list.name)
			for d in dn.items:
				w_house = frappe.db.get_value("Warehouse",
											  {'company': dn.company, 'default_for_stock_transfer': 1},
											  ['name'])
				val = frappe.db.get_value("Bin", {"item_code": d.item_code, "warehouse": w_house}, ['valuation_rate']) or 0
				total = val * d.qty
				tot += total
	prof = grand - tot
	gp = (prof / grand) * 100
	gr_pr = sales_in["total"] - project_cost_till_date(name)
	per_gp = (gr_pr / sales_in["total"] or 0) * 100
	data += '<h3>Project Details</h3>'
	data += '<table border=1px solid black width=100%>'
	data += '<tr style="background-color:#ABB2B9;text-align:center"><td><b>Order Value</b></td><td><b>Invoice Value till Date</b></td><td><b>Cost Till Date</b></td><td><b>Gross Profit</b></td><td><b>Gross Profit %</b></td><td><b>Balance To Invoice</b></td><td><b>Total Collection</b></td><td><b>O/S Receipts</b></td></tr>'
	data += '<tr style="text-align:right"><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (
		fmt_money(grand, 2), fmt_money(sales_in["total"] or 0, 2), fmt_money(project_cost_till_date(name), 2),
		fmt_money(round(gr_pr, 2), 2), fmt_money(round(per_gp, 2), 2), fmt_money(round(bal, 2), 2),
		fmt_money(round(revenue["amt"], 2) or 0, 2), fmt_money(round(out, 2), 2))
	data += '</table><br>'			
	data += '<h3>Overall Summary</h3>'
	data += '<table border="1" style="width:100%; border-collapse:collapse;">'
	data += '<tr style="background-color:#ABB2B9; text-align:center"><td><b>S.NO</b></td><td><b>Items</b></td><td><b>Budgeted</b></td><td><b>Actual</b></td></tr>'
	
	dic = [{"name": "Supply Materials", "table": pb.materials},
		   {"name": "Accessories", "table": pb.bolts_accessories},
		   {"name": "Subcontract", "table": pb.others},
		   {"name": "Design", "table": pb.design},
		   {"name": "Finishing Work", "table": pb.finishing_work},
		   {"name": "Tools / Equipment / Transport / Others", "table": pb.heavy_equipments},
		   {"name": "Finished Goods", "table": pb.finished_goods},
		   {"name": "Manpower", "table": pb.manpower},
		   {"name": "Installation", "table": pb.installation}
		   ]
	serial_no = 1
	for j in dic:
		
		if j["table"]:
			if j["name"] == "Manpower":
					man = frappe.db.sql(
						"""select sum(`tabTimesheet`.total_costing_amount) as amt from `tabTimesheet`
						left join `tabTimesheet Detail` on `tabTimesheet`.name = `tabTimesheet Detail`.parent
						where `tabTimesheet Detail`.project = %s""",
						(name,),
						as_dict=True) or 0
					if man:
						m = man[0]['amt']
					else:
						m = 0
				
					fmt_money(round(i.amount_with_overheads, 2), 2)
					fmt_money((m or 0), 2)
			
			elif j["name"] == "Installation":
				amt = 0
				total_amt = 0
				a_amt = 0
				total_a_amt = 0
				v_amt = 0
				total_v_amt = 0

				for i in j["table"]:
					dn = frappe.db.sql("""
						SELECT 
							SUM(`tabTimesheet Detail`.costing_amount) AS total_cost,
							`tabTimesheet Detail`.task,
							`tabTask`.subject AS task_name
						FROM 
							`tabTimesheet`
						LEFT JOIN 
							`tabTimesheet Detail` ON `tabTimesheet`.name = `tabTimesheet Detail`.parent
						LEFT JOIN 
							`tabTask` ON `tabTimesheet Detail`.task = `tabTask`.name
						WHERE 
							`tabTimesheet Detail`.project = '%s' AND `tabTimesheet`.docstatus = 1 AND `tabTask`.subject = '%s'
						GROUP BY 
							`tabTimesheet Detail`.task
					""" % (name,i.description), as_dict=True)
					if dn:
						completed_qty = frappe.db.get_value("Task",dn[0]['task'],'completed_qty') or 0
						d = completed_qty
						d_amt = round(dn[0]['total_cost'],2)
					else:
						d = 0
						d_amt = 0
						
					amt += round(i.amount_with_overheads, 2)
					a_amt += round(d_amt, 2)
					v_amt += round(i.amount_with_overheads, 2) - round(d_amt, 2)
					total_amt += amt
					total_a_amt += a_amt
					total_v_amt += v_amt
			else:
				amt = 0
				total_amt = 0
				a_amt = 0
				total_a_amt = 0
				v_amt = 0
				total_v_amt = 0
				total_am = 0
				for i in j["table"]:
					dn = frappe.db.sql(
						"""select `tabDelivery Note Item`.item_code as item_code, sum(`tabDelivery Note Item`.qty) as qty, sum(`tabDelivery Note Item`.amount) as amount from `tabDelivery Note`
						left join `tabDelivery Note Item` on `tabDelivery Note`.name = `tabDelivery Note Item`.parent
						where `tabDelivery Note Item`.item_code = %s and `tabDelivery Note`.project = %s group by `tabDelivery Note Item`.item_code""",
						(i.item, name),
						as_dict=True) or 0
					dn_wip = find_dn_qty_for_item(pb.sales_order, i.item) or 0
					
					if dn and dn_wip:
						d = dn[0]['qty'] + dn_wip
						d_amt = (dn[0]['qty'] + dn_wip) * i.unit_price
					elif dn:
						d = dn[0]['qty']
						d_amt = dn[0]['qty'] * i.unit_price
					elif dn_wip:
						d = dn_wip
						d_amt = dn_wip * i.unit_price
					else:
						d = 0
						d_amt = 0
					# frappe.log_error(title="Budgeted vs Actual Report", message=[i.item, d, d_amt])
			
					# amt += round(i.amount_with_overheads, 2)
					amt += round(i.amount, 2)

					a_amt += round(d_amt, 2)
					# v_amt += round(i.amount_with_overheads, 2) - round(d_amt, 2)
					v_amt += round(i.amount, 2) - round(d_amt, 2)
					total_amt += amt
					total_a_amt += a_amt
					total_v_amt += v_amt

			
			data += '<tr style="text-align:right">'
			data += '<td>%s</td>' % serial_no
			data += '<td style="text-align:left">%s</td>' % j["name"]
			data += '<td>%s</td>' % fmt_money(round(amt, 2), 2) or 0
			data += '<td>%s</td>' %  fmt_money(round(a_amt, 2), 2) or 0
			data += '</tr>'
			
			serial_no += 1
	

	data += '</table><br>'

	data += '<table border=1px solid black width=100%>'
	data += '<tr style="background-color:#ABB2B9"><td></td><td colspan=5></td><td colspan=2 style="text-align:center"><b>Budgeted</b></td><td colspan=2 style="text-align:center"><b>Actual</b></td><td style="text-align:center" colspan=2><b>Variance</b></td><tr>'
	data += '<tr style="background-color:#D5D8DC"><td colspan=1><b>S.NO</b></td><td colspan=1><b>Budget Code</b></td><td colspan=1><b>Part Number</b></td><td colspan=1><b>Description</b></td><td colspan=1><b>Unit</b></td><td colspan=1><b>Qty</b></td><td colspan=1><b>Rate</b></td><td colspan=1><b>Amount</b></td><td colspan=1><b>Qty</b></td><td colspan=1><b>Amount</b></td><td colspan=1><b>Qty</b></td><td colspan=1><b>Amount</b></td><tr>'
	dic_items = []
	dic = [{"name": "Supply Materials", "table": pb.materials},
		   {"name": "Accessories", "table": pb.bolts_accessories},
		   {"name": "Subcontract", "table": pb.others},
		   {"name": "Design", "table": pb.design},
		   {"name": "Finishing Work", "table": pb.finishing_work},
		   {"name": "Tools / Equipment / Transport / Others", "table": pb.heavy_equipments},
		   {"name": "Finished Goods", "table": pb.finished_goods}]
		#    {"name": "Other Expense", "table": other_report}]
	serial_no = 1
	for j in dic:
		if j["table"]:
			data += '<tr><td style="background-color:#EBEDEF;text-align:center" colspan=12><b>%s</b></td><tr>' % (
				j["name"])
			amt = 0
			total_amt = 0
			a_amt = 0
			total_a_amt = 0
			v_amt = 0
			total_v_amt = 0
			for i in j["table"]:
				dn = frappe.db.sql(
					"""select `tabDelivery Note Item`.item_code as item_code, sum(`tabDelivery Note Item`.qty) as qty, sum(`tabDelivery Note Item`.amount) as amount from `tabDelivery Note`
					left join `tabDelivery Note Item` on `tabDelivery Note`.name = `tabDelivery Note Item`.parent
					where `tabDelivery Note Item`.item_code = %s and `tabDelivery Note`.project = %s group by `tabDelivery Note Item`.item_code""",
					(i.item, name),
					as_dict=True) or 0
				dn_wip = find_dn_qty_for_item(pb.sales_order, i.item) or 0
				
				if dn and dn_wip:
					d = dn[0]['qty'] + dn_wip
					d_amt = (dn[0]['qty'] + dn_wip) * i.unit_price
				elif dn:
					d = dn[0]['qty']
					d_amt = dn[0]['qty'] * i.unit_price
				elif dn_wip:
					d = dn_wip
					d_amt = dn_wip * i.unit_price
				else:
					d = 0
					d_amt = 0
				data += '<tr><td>%s</td><td colspan=1>%s</td><td colspan=1>%s</td><td colspan=1>%s</td><td colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><tr>' % (
					serial_no,project_budget, i.item, i.description, i.unit, i.qty, fmt_money(round(i.unit_price, 2), 2),
					fmt_money(round(i.amount, 2), 2), d, fmt_money(round(d_amt, 2), 2),
					fmt_money(round(i.qty - d, 2), 2), fmt_money(round(i.amount, 2) - round(d_amt, 2), 2))
				serial_no += 1
				# amt += round(i.amount_with_overheads, 2)
				amt += round(i.amount, 2)
				a_amt += round(d_amt, 2)
				v_amt += round(i.amount, 2) - round(d_amt, 2)
				# v_amt += round(i.amount_with_overheads, 2) - round(d_amt, 2)
				total_amt += amt
				total_a_amt += a_amt
				total_v_amt += v_amt
				
			data += '<tr><td></td><td colspan=6></td><td style="text-align:right" colspan=1>%s</td><td style="text-align:right" colspan=1></td><td style="text-align:right" colspan=1>%s</td><td style="text-align:right" colspan=1></td><td style="text-align:right" colspan=1>%s</td><tr>' % (
				fmt_money(round(amt, 2), 2), fmt_money(round(a_amt, 2), 2), fmt_money(round(v_amt, 2), 2))
			
			
	manpower = [{"name": "Manpower", "table": pb.manpower}]
	serial_no = 1
	for j in manpower:
		if j["table"]:
			data += '<tr><td style="background-color:#EBEDEF;text-align:center" colspan=12><b>%s</b></td><tr>' % (
				j["name"])
			for i in j["table"]:
				man = frappe.db.sql(
					"""select sum(`tabTimesheet`.total_costing_amount) as amt from `tabTimesheet`
					left join `tabTimesheet Detail` on `tabTimesheet`.name = `tabTimesheet Detail`.parent
					where `tabTimesheet Detail`.project = %s""",
					(name,),
					as_dict=True) or 0
				if man:
					m = man[0]['amt']
				else:
					m = 0
				data += '<tr><td>%s</td><td colspan=1>%s</td><td colspan=1>%s</td><td colspan=1>%s</td><td colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><tr>' % (
					serial_no,project_budget, i.worker, i.worker, "Nos", i.total_workers,
					fmt_money(round(i.rate_with_overheads, 2), 2),
					fmt_money(round(i.amount_with_overheads, 2), 2), i.total_workers,
					fmt_money((m or 0), 2), i.total_workers,
					fmt_money(((i.amount_with_overheads) - (m or 0)), 2))
				serial_no +=1
	install = [{"name": "Installation", "table": pb.installation}]
	serial_no = 1
	for j in install:
		if j["table"]:
			data += '<tr><td style="background-color:#EBEDEF;text-align:center" colspan=12><b>%s</b></td><tr>' % (
				j["name"])
			amt = 0
			total_amt = 0
			a_amt = 0
			total_a_amt = 0
			v_amt = 0
			total_v_amt = 0

			for i in j["table"]:
				dn = frappe.db.sql("""
					SELECT 
						SUM(`tabTimesheet Detail`.costing_amount) AS total_cost,
						`tabTimesheet Detail`.task,
						`tabTask`.subject AS task_name
					FROM 
						`tabTimesheet`
					LEFT JOIN 
						`tabTimesheet Detail` ON `tabTimesheet`.name = `tabTimesheet Detail`.parent
					LEFT JOIN 
						`tabTask` ON `tabTimesheet Detail`.task = `tabTask`.name
					WHERE 
						`tabTimesheet Detail`.project = '%s' AND `tabTimesheet`.docstatus = 1 AND `tabTask`.subject = '%s'
					GROUP BY 
						`tabTimesheet Detail`.task
				""" % (name,i.description), as_dict=True)
				if dn:
					completed_qty = frappe.db.get_value("Task",dn[0]['task'],'completed_qty') or 0
					d = completed_qty
					d_amt = round(dn[0]['total_cost'],2)
				else:
					d = 0
					d_amt = 0
					
				data += '<tr><td>%s</td><td colspan=1>%s</td><td colspan=1>%s</td><td colspan=1>%s</td><td colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><tr>' % (serial_no,project_budget, i.item, i.description, i.unit, i.qty,fmt_money(round(i.rate_with_overheads, 2), 2),fmt_money(round(i.amount_with_overheads, 2), 2),d,d_amt,i.qty - d, fmt_money(round(i.amount_with_overheads, 2) - round(d_amt, 2), 2))
				serial_no +=1
				amt += round(i.amount_with_overheads, 2)
				a_amt += round(d_amt, 2)
				v_amt += round(i.amount_with_overheads, 2) - round(d_amt, 2)
				total_amt += amt
				total_a_amt += a_amt
				total_v_amt += v_amt
			data += '<tr><td></td><td colspan=6></td><td style="text-align:right" colspan=1>%s</td><td style="text-align:right" colspan=1></td><td style="text-align:right" colspan=1>%s</td><td style="text-align:right" colspan=1></td><td style="text-align:right" colspan=1>%s</td><tr>' % (
				fmt_money(round(amt, 2), 2), fmt_money(round(a_amt, 2), 2), fmt_money(round(v_amt, 2), 2))
			
	# for j in other_expense:
	#     if j["table"]:
	#         data += '<tr><td style="background-color:#EBEDEF;text-align:center" colspan=11><b>%s</b></td><tr>' % (
	#             j["name"])
	#         serial_no = 1  # Initialize serial number for table rows
	#         total_amount = 0

	#         for i in j["table"]:
	#             other_items = frappe.db.sql("""
	#                 SELECT p.name AS project_name, pb.name AS budget_name, pb.voucher_type, pbi.account, pbi.debit, pbi.credit
	#                 FROM `tabProject` p
	#                 JOIN `tabJournal Entry Account` pbi ON p.name = pbi.project
	#                 JOIN `tabJournal Entry` pb ON pbi.parent = pb.name
	#                 WHERE pb.voucher_type = 'Journal Entry' AND pb.docstatus = 1
	#             """)
	#             if not other_items:
	#                 # If no data, display a message inside the table
	#                 table += '''
	#                     <tr>
	#                         <td colspan="6">No data available for the selected project.</td>
	#                     </tr>'''
	#             else:
	#                 # If data exists, iterate through the fetched data and build the table rows
	#                 for item in other_items:
	#                     # Determine whether to use debit or credit for the amount
	#                     amount = item['debit'] if item['debit'] else item['credit']
	#                     # Accumulate total amount
	#                     total_amount += amount
	#             data += '<tr><td colspan=1>%s</td><td colspan=1>%s</td><td colspan=1>%s</td><td colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><tr>' % (serial_no,{item['project_name']},{item['budget_name']},{item['voucher_type']},{item['account']},{amount})


	data += '</table>'
	data += other_report(name)
	# data += get_timesheet_project(name)
	
	
	
	return data
#Get the below details and show html format in Project document
@frappe.whitelist()
def work_order_detailed_report(name):
	crea = frappe.db.sql("""select creation from `tabProject` where name ='%s' """ % (name), as_dict=True)[0]
	project_budget,p_name = frappe.db.get_value("Project",{'name':name},['budgeting','project_name'])
	pb = frappe.get_doc("Project Budget",project_budget)
	data = '<table width = 100%>'
	data += '<tr><td><b>Date</b></td><td>%s</td><td><b>Refer No</b></td><td></td></tr>'%(format_date(crea["creation"].date()))
	data += '<tr><td><b>Project Code</b></td><td>%s</td><td><b>Project</b></td><td>%s</td></tr>'%(name,p_name)
	data += '<tr><td><b>Client</b></td><td>%s</td><td><b>Order Ref No.</b></td><td>%s</td></tr>'%(pb.lead_customer_name,pb.sales_order)
	data += '</table><br>'

	grand = frappe.db.get_value("Sales Order",{'project':name},['grand_total'])
	sales_in = frappe.db.sql(""" select sum(grand_total) as total from `tabSales Invoice` where project = '%s' """%(name),as_dict=True)[0]
	if not sales_in["total"]:
		sales_in["total"] = 0
	bal = grand - sales_in["total"]
	revenue = frappe.db.sql(""" select sum(paid_amount) as amt from `tabPayment Entry` where project = '%s' """%(name),as_dict=True)[0]
	if not revenue["amt"]:
		revenue["amt"] = 0
	out = grand - revenue["amt"]
	tot = 0
	val = 0
	dn_list = frappe.db.get_list("Delivery Note",{'project':name})
	if dn_list:
		for d_list in dn_list:
			dn = frappe.get_doc("Delivery Note",d_list.name)
			for d in dn.items:
				w_house = frappe.db.get_value("Warehouse",{'company':dn.company,'default_for_stock_transfer':1},['name'])
				val = frappe.db.get_value("Bin",{"item_code":d.item_code,"warehouse":w_house},['valuation_rate']) or 0
				if not d.qty:
					d.qty = 0
				total = val * d.qty
				tot+=total
	prof = grand - tot
	gp = (prof / grand)*100
	data += '<h3>Project Details</h3>'
	data += '<table border= 1px solid black width = 100%>'
	data += '<tr style = "background-color:#ABB2B9;text-align:center"><td><b>Order Value</b></td><td><b>Invoice Value till Date</b></td><td><b>Cost Till Date</b></td><td><b>Gross Profit</b></td><td><b>Gross Profit %</b></td><td><b>Balance To Invoice</b></td><td><b>Total Collection</b></td><td><b>O/S Receipts</b></td></tr>'
	data += '<tr style = "text-align:right"><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>'%(grand,sales_in["total"] or 0,round(tot,2),round(prof,2),round(gp,2),round(bal,2),round(revenue["amt"],2) or 0,round(out,2))
	data += '</table><br>'

	data += '<table border= 1px solid black width = 100%>'
	dic = [{"name":"Supply Materials","table":pb.materials},{"name":"Accessories","table":pb.bolts_accessories},{"name":"Installation","table":pb.installation},{"name":"Subcontract","table":pb.others},{"name":"Design","table":pb.design},{"name":"Finishing Work","table":pb.finishing_work},{"name":"Tools / Equipment / Transport / Others","table":pb.heavy_equipments}]
	for j in dic:
		if j["table"]:
			data += '<tr><td style ="background-color:#EBEDEF;text-align:center" colspan = 13><b>%s</b></td><tr>'%(j["name"])          
			data +='<table border= 1px width = 100%>'
			data += '<tr style= "background-color:#D5D8DC"><td colspan = 1><b>Budget Code</b></td><td colspan = 1><b>Part Number</b></td><td colspan = 1><b>Description</b></td><td colspan = 1><b>Unit</b></td><td colspan = 1><b>Qty</b></td><td colspan = 1><b>Rate</b></td><td colspan = 1><b>Amount</b></td><td colspan = 1><b>Stock Qty</b></td><td colspan = 1><b>MR Qty</b></td><td colspan = 1><b>MR Rate</b></td><td colspan = 1><b>MR Amount</b></td><td colspan = 1><b>DN Rate</b></td><td colspan = 1><b>DN Amount</b></td><tr>'
			
			for i in j["table"]:     
				dn = frappe.db.sql("""select `tabDelivery Note Item`.item_code as item_code, `tabDelivery Note Item`.qty as qty, `tabDelivery Note Item`.amount as amount from `tabDelivery Note`
				left join `tabDelivery Note Item` on `tabDelivery Note`.name = `tabDelivery Note Item`.parent
				where `tabDelivery Note Item`.item_code = '%s' and `tabDelivery Note`.project = '%s' """ % (i.item,name), as_dict=True) or 0
				if dn:
					d = dn[0]['qty']
					d_amt = dn[0]['amount']
				else:
					d = 0
					d_amt = 0
				mr_amount = 0
				mr_qty = 0
				mr_rate = 0
				parent = frappe.db.sql("""select `tabMaterial Request`.name as name from `tabMaterial Request`
				left join `tabMaterial Request Item` on `tabMaterial Request`.name = `tabMaterial Request Item`.parent
				where `tabMaterial Request Item`.item_code = '%s' and `tabMaterial Request`.docstatus != 2 and `tabMaterial Request Item`.project = '%s'"""%(i.item,name),as_dict=True)
				company = frappe.db.get_value("Project",{'name':name},['company'])
				w_house = frappe.db.get_value("Warehouse",{'company':company,'default_for_stock_transfer':1},['name'])
				stocks = frappe.db.sql("""select actual_qty from tabBin where item_code = '%s' and warehouse = '%s' """%(i.item,w_house),as_dict=True)
				if not stocks:
					stock = 0
				else:
					stock = stocks[0]["actual_qty"]
					wo_amt = 0
				data += '<tr><td  style = "text-align:left;width:18%% " colspan = 1>%s</td><td  style = "text-align:left;width:22%%" colspan = 1>%s</td><td  style = "text-align:left;width:28%%" colspan = 1>%s</td><td  style = "text-align:right;width:3%%" colspan = 1>%s</td><td  style = "text-align:right;width:3%%" colspan = 1>%s</td><td  style = "text-align:right;width:3%%" colspan = 1>%s</td><td  style = "text-align:right;width:3%%" colspan = 1>%s</td><td  style = "text-align:right;width:3%%" colspan = 1>%s</td><td  style = "text-align:right;width:3%%" colspan = 1>%s</td><td  style = "text-align:right;width:3%%" colspan = 1>%s</td><td  style = "text-align:right;width:3%%" colspan = 1>%s</td><td  style = "text-align:right;width:3%%" colspan = 1>%s</td><td  style = "text-align:right;width:3%%" colspan = 1>%s</td></tr>'%(project_budget,i.item,i.description,i.unit,i.qty,round(i.rate_with_overheads,2),round(i.amount_with_overheads,2),stock,mr_qty,mr_rate,mr_amount,d,round(d_amt,2))

	fg = [{"name":"Finished Goods","table":pb.finished_goods}]
	for j in fg:
		if j["table"]:
			data += '<tr><td style ="background-color:#EBEDEF;text-align:center" colspan = 13><b>%s</b></td><tr>'%(j["name"])          
			amt = 0
			total_amt = 0
			a_amt = 0 
			total_a_amt = 0 
			v_amt = 0
			total_v_amt = 0
			for i in j["table"]:     
				dn = frappe.db.sql("""select `tabDelivery Note Item`.item_code as item_code, `tabDelivery Note Item`.qty as qty, `tabDelivery Note Item`.amount as amount from `tabDelivery Note`
				left join `tabDelivery Note Item` on `tabDelivery Note`.name = `tabDelivery Note Item`.parent
				where `tabDelivery Note Item`.item_code = '%s' and `tabDelivery Note`.project = '%s' """ % (i.item,name), as_dict=True) or 0
				if dn:
					d = dn[0]['qty']
					d_amt = dn[0]['amount']
				else:
					d = 0
					d_amt = 0
				mr_amount = 0
				if i.bom:
					bom = frappe.get_doc("BOM",i.bom)
					for it in bom.items:
						parent = frappe.db.sql("""select `tabMaterial Request`.name as name from `tabMaterial Request`
						left join `tabMaterial Request Item` on `tabMaterial Request`.name = `tabMaterial Request Item`.parent
						where `tabMaterial Request Item`.item_code = '%s' and `tabMaterial Request`.docstatus != 2 and `tabMaterial Request Item`.project = '%s'"""%(it.item_code,name),as_dict=True) or ''
						
						if parent:
							
							mr = frappe.get_doc("Material Request",parent[0]["name"])
							for m in mr.items:
								if m.item_code == it.item_code:
									mr_amount += m.amount
				wo_amt = 0
				if i.bom:
					wor = frappe.db.exists("Work Order",{'bom_no':i.bom})
					if wor:
						work_order = frappe.get_doc("Work Order",{'bom_no':i.bom})
						if work_order:
							for wo in work_order.required_items:
								wo_amt += wo.amount
		 
				data +='<table border= 1px width = 100%>'
				data += '<tr style= "background-color:#D5D8DC"><td colspan = 1><b>Budget Code</b></td><td colspan = 1><b>Part Number</b></td><td colspan = 1><b>Description</b></td><td colspan = 1><b>Unit</b></td><td colspan = 1><b>Qty</b></td><td colspan = 1><b>Rate</b></td><td colspan = 1><b>Amount</b></td><td colspan = 1><b>MR Amount</b></td><td colspan = 1><b>WO Amount</b></td><tr>'
				data += '<tr><td  style = "text-align:left;width:18%% " colspan = 1>%s</td><td  style = "text-align:left;width:22%%" colspan = 1>%s</td><td  style = "text-align:left;width:28%%" colspan = 1>%s</td><td  style = "text-align:right;width:3%%" colspan = 1>%s</td><td  style = "text-align:right;width:3%%" colspan = 1>%s</td><td  style = "text-align:right;width:3%%" colspan = 1>%s</td><td  style = "text-align:right;width:3%%" colspan = 1>%s</td><td  style = "text-align:right;width:3%%" colspan = 1>%s</td><td  style = "text-align:right;width:3%%" colspan = 1>%s</td></tr>'%(project_budget,i.item,i.description,i.unit,i.qty,round(i.rate_with_overheads,2),round(i.amount_with_overheads,2),round(mr_amount,2),round(wo_amt,2))

				data += '<tr style= "background-color:#D5D8DC"><td style = "text-align:center"  colspan = 3></td><td style = "text-align:center"  colspan = 3><b>Material Request Details</b></td><td style = "text-align:center"  colspan = 3><b>Work Order Details</b></td></tr>'
				data +='<tr style ="background-color:#EBEDEF"><td style = "text-align:center;width: 25%" colspan = 1><b>Item Code</b></td><td  style = "text-align:center;width: 35%" colspan = 1><b>Item Name</b></td><td  style = "text-align:center;width: 25%" colspan = 1><b>Stock Qty</b></td><td  style = "text-align:center;width: 25%" colspan = 1><b>Qty</b></td><td  style = "text-align:center;width: 25%" colspan = 1><b>Rate</b></td><td  style = "text-align:center;width: 25%" colspan = 1><b>Amount</b></td><td  style = "text-align:center;width: 25%" colspan = 1><b>Qty</b></td><td  style = "text-align:center;width: 25%" colspan = 1><b>Rate</b></td><td  style = "text-align:center;width: 25%" colspan = 2><b>Amount</b></td><tr>'
				

				if i.bom:                    
					bom = frappe.get_doc("BOM",i.bom)
					for it in bom.items:
						company = frappe.db.get_value("Project",{'name':name},['company'])
						w_house = frappe.db.get_value("Warehouse",{'company':company,'default_for_stock_transfer':1},['name'])
						stocks = frappe.db.sql("""select actual_qty from tabBin where item_code = '%s' and warehouse = '%s' """%(it.item_code,w_house),as_dict=True)
						parent = frappe.db.sql("""select `tabMaterial Request`.name as name from `tabMaterial Request`
						left join `tabMaterial Request Item` on `tabMaterial Request`.name = `tabMaterial Request Item`.parent
						where `tabMaterial Request Item`.item_code = '%s' and `tabMaterial Request`.docstatus != 2 and `tabMaterial Request Item`.project = '%s'"""%(it.item_code,name),as_dict=True) or ''
						
						if parent:
							
							mr = frappe.get_doc("Material Request",parent[0]["name"])
							for m in mr.items:                                                            
								if m.item_code == it.item_code:    
									pos = frappe.db.sql("""select `tabWork Order Item`.rate as rate,`tabWork Order Item`.amount as amount,`tabWork Order Item`.consumed_qty as consumed_qty from `tabWork Order`
									left join `tabWork Order Item` on `tabWork Order`.name = `tabWork Order Item`.parent
									where `tabWork Order Item`.item_code = '%s' and `tabWork Order`.docstatus != 2 and `tabWork Order`.bom_no ='%s' """%(m.item_code,i.bom),as_dict=True)                                                         
									data += '<tr><td  style = "text-align:left" colspan = 1>%s</td><td  style = "text-align:left" colspan = 1>%s</td><td  style = "text-align:right" colspan = 1>%s</td><td  style = "text-align:right" colspan = 1>%s</td><td  style = "text-align:right" colspan = 1>%s</td><td  style = "text-align:right" colspan = 1>%s</td><td  style = "text-align:right" colspan = 1>%s</td><td  style = "text-align:right" colspan = 1>%s</td><td  style = "text-align:right" colspan = 1>%s</td></tr>'%(m.item_code,m.item_name or '',round(stocks[0]["actual_qty"],2) or 0,m.qty or '',m.rate or '',m.amount or '',round(pos[0]["consumed_qty"],2) or '',round(pos[0]["rate"],2) or '',round(pos[0]["amount"],2) or '')
						else:
							workorder = frappe.db.sql("""select `tabWork Order`.name as name from `tabWork Order`
							left join `tabWork Order Item` on `tabWork Order`.name = `tabWork Order Item`.parent
							where `tabWork Order Item`.item_code = '%s' and `tabWork Order`.docstatus != 2 and `tabWork Order`.bom_no ='%s' """%(it.item_code,i.bom),as_dict=True) or ''
							
							if workorder:
								work_order = frappe.get_doc("Work Order",workorder[0]["name"])
						
							pos = frappe.db.sql("""select `tabWork Order Item`.rate as rate,`tabWork Order Item`.amount as amount,`tabWork Order Item`.consumed_qty as consumed_qty,`tabWork Order Item`.required_qty as required_qty from `tabWork Order`
							left join `tabWork Order Item` on `tabWork Order`.name = `tabWork Order Item`.parent
							where `tabWork Order Item`.item_code = '%s' and `tabWork Order`.docstatus != 2 and `tabWork Order`.bom_no ='%s' """%(it.item_code,i.bom),as_dict=True)
							if pos:
								data += '<tr><td  style = "text-align:left" colspan = 1>%s</td><td  style = "text-align:left" colspan = 1>%s</td><td  style = "text-align:right" colspan = 1>%s</td><td  style = "text-align:right" colspan = 1>%s</td><td  style = "text-align:right" colspan = 1>%s</td><td  style = "text-align:right" colspan = 1>%s</td><td  style = "text-align:right" colspan = 1>%s</td><td  style = "text-align:right" colspan = 1>%s</td><td  style = "text-align:right" colspan = 2>%s</td></tr>'%(it.item_code,it.item_name,round(stocks[0]["actual_qty"],2) or 0, 0,0,0,round(pos[0]["consumed_qty"],2) or round(pos[0]["required_qty"],2),round(pos[0]["rate"],2) or '',round(pos[0]["amount"],2) or '')

				data +='</table>'
				amt += round(i.amount_with_overheads,2)
				a_amt += round(d_amt,2)
				v_amt += round(i.amount_with_overheads,2) - round(d_amt,2)
				total_amt += amt
				total_a_amt += a_amt
				total_v_amt += v_amt
		  
	return data


#Get the below details and show html format in Project document
@frappe.whitelist()
def work_order_brief_report(name):
	crea = frappe.db.sql("""select creation from `tabProject` where name ='%s' """ % (name), as_dict=True)[0]
	project_budget,p_name = frappe.db.get_value("Project",{'name':name},['budgeting','project_name'])
	pb = frappe.get_doc("Project Budget",project_budget)
	data = '<table width = 100%>'
	data += '<tr><td><b>Date</b></td><td>%s</td><td><b>Refer No</b></td><td></td></tr>'%(format_date(crea["creation"].date()))
	data += '<tr><td><b>Project Code</b></td><td>%s</td><td><b>Project</b></td><td>%s</td></tr>'%(name,p_name)
	data += '<tr><td><b>Client</b></td><td>%s</td><td><b>Order Ref No.</b></td><td>%s</td></tr>'%(pb.lead_customer_name,pb.sales_order)
	data += '</table><br>'

	grand = frappe.db.get_value("Sales Order",{'project':name},['grand_total'])
	sales_in = frappe.db.sql(""" select sum(grand_total) as total from `tabSales Invoice` where project = '%s' """%(name),as_dict=True)[0]
	if not sales_in["total"]:
		sales_in["total"] = 0
	bal = grand - sales_in["total"]
	revenue = frappe.db.sql(""" select sum(paid_amount) as amt from `tabPayment Entry` where project = '%s' """%(name),as_dict=True)[0]
	if not revenue["amt"]:
		revenue["amt"] = 0
	out = grand - revenue["amt"]
	tot = 0
	val = 0
	dn_list = frappe.db.get_list("Delivery Note",{'project':name})
	if dn_list:
		for d_list in dn_list:
			dn = frappe.get_doc("Delivery Note",d_list.name)
			for d in dn.items:
				w_house = frappe.db.get_value("Warehouse",{'company':dn.company,'default_for_stock_transfer':1},['name'])
				val = frappe.db.get_value("Bin",{"item_code":d.item_code,"warehouse":w_house},['valuation_rate']) or 0
				if not d.qty:
					d.qty = 0
				total = val * d.qty
				tot+=total
	prof = grand - tot
	gp = (prof / grand)*100
	data += '<h3>Project Details</h3>'
	data += '<table border= 1px solid black width = 100%>'
	data += '<tr style = "background-color:#ABB2B9;text-align:center"><td><b>Order Value</b></td><td><b>Invoice Value till Date</b></td><td><b>Cost Till Date</b></td><td><b>Gross Profit</b></td><td><b>Gross Profit %</b></td><td><b>Balance To Invoice</b></td><td><b>Total Collection</b></td><td><b>O/S Receipts</b></td></tr>'
	data += '<tr style = "text-align:right"><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>'%(grand,sales_in["total"] or 0,round(tot,2),round(prof,2),round(gp,2),round(bal,2),round(revenue["amt"],2) or 0,round(out,2))
	data += '</table><br>'

	data += '<table border= 1px solid black width = 100%>'
	dic = [{"name":"Supply Materials","table":pb.materials},{"name":"Accessories","table":pb.bolts_accessories},{"name":"Installation","table":pb.installation},{"name":"Subcontract","table":pb.others},{"name":"Design","table":pb.design},{"name":"Finishing Work","table":pb.finishing_work},{"name":"Tools / Equipment / Transport / Others","table":pb.heavy_equipments}]
	for j in dic:
		if j["table"]:
			data += '<tr><td style ="background-color:#EBEDEF;text-align:center" colspan = 13><b>%s</b></td><tr>'%(j["name"])          
			data +='<table border= 1px width = 100%>'
			data += '<tr style= "background-color:#D5D8DC"><td colspan = 1><b>Budget Code</b></td><td colspan = 1><b>Part Number</b></td><td colspan = 1><b>Description</b></td><td colspan = 1><b>Unit</b></td><td colspan = 1><b>Qty</b></td><td colspan = 1><b>Rate</b></td><td colspan = 1><b>Amount</b></td><td colspan = 1><b>Stock Qty</b></td><td colspan = 1><b>MR Amount</b></td><td colspan = 1><b>DN Amount</b></td><tr>'
			
			for i in j["table"]:     
				dn = frappe.db.sql("""select `tabDelivery Note Item`.item_code as item_code, `tabDelivery Note Item`.qty as qty, `tabDelivery Note Item`.amount as amount from `tabDelivery Note`
				left join `tabDelivery Note Item` on `tabDelivery Note`.name = `tabDelivery Note Item`.parent
				where `tabDelivery Note Item`.item_code = '%s' and `tabDelivery Note`.project = '%s' """ % (i.item,name), as_dict=True) or 0
				if dn:
					d = dn[0]['qty']
					d_amt = dn[0]['amount']
				else:
					d = 0
					d_amt = 0
				mr_amount = 0
				mr_qty = 0
				mr_rate = 0
				parent = frappe.db.sql("""select `tabMaterial Request`.name as name from `tabMaterial Request`
				left join `tabMaterial Request Item` on `tabMaterial Request`.name = `tabMaterial Request Item`.parent
				where `tabMaterial Request Item`.item_code = '%s' and `tabMaterial Request`.docstatus != 2 and `tabMaterial Request Item`.project = '%s'"""%(i.item,name),as_dict=True)
				company = frappe.db.get_value("Project",{'name':name},['company'])
				w_house = frappe.db.get_value("Warehouse",{'company':company,'default_for_stock_transfer':1},['name'])
				stocks = frappe.db.sql("""select actual_qty from tabBin where item_code = '%s' and warehouse = '%s' """%(i.item,w_house),as_dict=True)
				if not stocks:
					stock = 0
				else:
					stock = stocks[0]["actual_qty"]
					wo_amt = 0
				data += '<tr><td  style = "text-align:left;width:18%% " colspan = 1>%s</td><td  style = "text-align:left;width:22%%" colspan = 1>%s</td><td  style = "text-align:left;width:28%%" colspan = 1>%s</td><td  style = "text-align:right;width:3%%" colspan = 1>%s</td><td  style = "text-align:right;width:3%%" colspan = 1>%s</td><td  style = "text-align:right;width:3%%" colspan = 1>%s</td><td  style = "text-align:right;width:3%%" colspan = 1>%s</td><td  style = "text-align:right;width:3%%" colspan = 1>%s</td><td  style = "text-align:right;width:3%%" colspan = 1>%s</td><td  style = "text-align:right;width:3%%" colspan = 1>%s</td></tr>'%(project_budget,i.item,i.description,i.unit,i.qty,round(i.rate_with_overheads,2),round(i.amount_with_overheads,2),stock,mr_amount,round(d_amt,2))

	fg = [{"name":"Finished Goods","table":pb.finished_goods}]
	for j in fg:
		if j["table"]:
			data += '<tr><td style ="background-color:#EBEDEF;text-align:center" colspan = 13><b>%s</b></td><tr>'%(j["name"])
			data +='<table border= 1px width = 100%>'
			data += '<tr style= "background-color:#D5D8DC"><td colspan = 1><b>Budget Code</b></td><td colspan = 1><b>Part Number</b></td><td colspan = 1><b>Description</b></td><td colspan = 1><b>Unit</b></td><td colspan = 1><b>Qty</b></td><td colspan = 1><b>Rate</b></td><td colspan = 1><b>Amount</b></td><td colspan = 1><b>MR Amount</b></td><td colspan = 1><b>WO Amount</b></td><tr>'
					 
			amt = 0
			total_amt = 0
			a_amt = 0 
			total_a_amt = 0 
			v_amt = 0
			total_v_amt = 0
			for i in j["table"]:     
				dn = frappe.db.sql("""select `tabDelivery Note Item`.item_code as item_code, `tabDelivery Note Item`.qty as qty, `tabDelivery Note Item`.amount as amount from `tabDelivery Note`
				left join `tabDelivery Note Item` on `tabDelivery Note`.name = `tabDelivery Note Item`.parent
				where `tabDelivery Note Item`.item_code = '%s' and `tabDelivery Note`.project = '%s' """ % (i.item,name), as_dict=True) or 0
				if dn:
					d = dn[0]['qty']
					d_amt = dn[0]['amount']
				else:
					d = 0
					d_amt = 0
				mr_amount = 0
				if i.bom:
					bom = frappe.get_doc("BOM",i.bom)
					for it in bom.items:
						parent = frappe.db.sql("""select `tabMaterial Request`.name as name from `tabMaterial Request`
						left join `tabMaterial Request Item` on `tabMaterial Request`.name = `tabMaterial Request Item`.parent
						where `tabMaterial Request Item`.item_code = '%s' and `tabMaterial Request`.docstatus != 2 and `tabMaterial Request Item`.project = '%s'"""%(it.item_code,name),as_dict=True) or ''
						
						if parent:
							
							mr = frappe.get_doc("Material Request",parent[0]["name"])
							for m in mr.items:
								if m.item_code == it.item_code:
									mr_amount += m.amount
				wo_amt = 0
				if i.bom:
					wor = frappe.db.exists("Work Order",{'bom_no':i.bom})
					if wor:
						work_order = frappe.get_doc("Work Order",{'bom_no':i.bom})
						if work_order:
							for wo in work_order.required_items:
								wo_amt += wo.amount
		 
				data += '<tr><td  style = "text-align:left;width:18%% " colspan = 1>%s</td><td  style = "text-align:left;width:22%%" colspan = 1>%s</td><td  style = "text-align:left;width:28%%" colspan = 1>%s</td><td  style = "text-align:right;width:3%%" colspan = 1>%s</td><td  style = "text-align:right;width:3%%" colspan = 1>%s</td><td  style = "text-align:right;width:3%%" colspan = 1>%s</td><td  style = "text-align:right;width:3%%" colspan = 1>%s</td><td  style = "text-align:right;width:3%%" colspan = 1>%s</td><td  style = "text-align:right;width:3%%" colspan = 1>%s</td></tr>'%(project_budget,i.item,i.description,i.unit,i.qty,round(i.rate_with_overheads,2),round(i.amount_with_overheads,2),round(mr_amount,2),round(wo_amt,2))

			   
				amt += round(i.amount_with_overheads,2)
				a_amt += round(d_amt,2)
				v_amt += round(i.amount_with_overheads,2) - round(d_amt,2)
				total_amt += amt
				total_a_amt += a_amt
				total_v_amt += v_amt
		  
	return data

#Get the child table details in excel format for given documents
@frappe.whitelist()
def make_item_sheet():
	args = frappe.local.form_dict
	filename = args.name
	test = build_xlsx_response(filename)

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary' 	
	
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
	ws = wb.create_sheet(sheet_name, 0)
	if args.doctype == "Quotation":
		doc = frappe.get_doc(args.doctype,args.name)
		if doc:
			ws.append(["Item Code","Item Name","Brand","Availability","Qty","UOM","Rate","Amount"])
			for i in doc.items:
				ws.append([i.item_code,i.description,i.brand,i.delivery_terms,i.qty,i.uom,i.rate,i.amount])
	if args.doctype == "Sales Invoice":
		doc = frappe.get_doc(args.doctype,args.name)
		if doc:
			ws.append(["Item Code","Item Name","Qty","UOM","Rate","Amount"])
			for i in doc.items:
				ws.append([i.item_code,i.item_name,i.qty,i.uom,i.rate,i.amount])
	if args.doctype == "Sales Order":
		doc = frappe.get_doc(args.doctype,args.name)
		if doc:
			ws.append(["Item Code","Item Name","Delivery Date","Qty","UOM","Rate","Amount"])
			for i in doc.items:
				ws.append([i.item_code,i.item_name,i.delivery_date,i.qty,i.uom,i.rate,i.amount])
	if args.doctype == "Delivery Note":
		doc = frappe.get_doc(args.doctype,args.name)
		if doc:
			ws.append(["Item Code","Item Name","Qty","UOM","Rate","Amount"])
			for i in doc.items:
				ws.append([i.item_code,i.item_name,i.qty,i.uom,i.rate,i.amount])
	if args.doctype == "Landed Cost Voucher":
		doc = frappe.get_doc(args.doctype,args.name)
		if doc:
			ws.append(["Item Code","Description","Qty","Current Rate After LCV"])
			for i in doc.items:
				ws.append([i.item_code,i.description,i.qty,i.current_rate_after_lcv])
	if args.doctype == "Project Budget":
		doc = frappe.get_doc(args.doctype,args.name)
		if doc:
			ws.append(["Item Code","Item Name","Unit","Qty"])
			for i in doc.item_table:
				ws.append([i.item,i.item_name,i.unit,i.qty])
				
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file



#Show the below details in HTML Format while enter the stock details button
@frappe.whitelist()
def update_detail_stock(item_details, company):
	item_details = json.loads(item_details)
	data = ''
	data += '<h4><center><b>STOCK DETAILS</b></center></h4>'
	data += '<h6>Note:</h6>'
	data += '<table style = font-size:10px width=100% ><tr><td>Electra Interior Warehouse - <b>INE</b></td><td>Kingfisher Transportation Warehouse - <b>KT</b></td><td>Steel Warehouse - <b>SDE</b></td><td>Marazeem HO Warehouse - <b>MSSHO</b></td><td>Electra Trading Warehouse - <b>TDE</b></td></tr>'
	data += '<tr><td>Electra Warehouse - <b>ASTCC</b></td><td>Electra Binomran Showroom Warehouse - <b>EBO</b></td><td>Kingfisher Warehouse - <b>KTCC</b></td><td>Kingfisher Showroom Warehouse - <b>KS</b></td><td>Marazeem Showroom - <b>MSSS</b></td></tr>'
	data += '<tr><td>Electra Najma Showroom Warehouse - <b> ENS</b></td><td>Marazeem Warehouse - <b>MSS</b></td><td>Barwa Showroom  - <b>EBS</b></td><td>Electra Electrical Warehouse - <b>EDE</b></td><td>Electra Engineering Warehouse - <b>EED</b></td></tr>'
	data += '</table>'
	data += '<table class="table table-bordered" style = font-size:10px>'
	data += '<td colspan=1 style="width:12%;padding:1px;border:1px solid black;background-color:#e35310;color:white;"><center><b>ITEM CODE</b></center></td>'
	data += '<td colspan=1 style="width:20%;padding:1px;border:1px solid black;background-color:#e35310;color:white;"><center><b>ITEM NAME</b></center></td>'
	data += '<td colspan=1 style="width:70px;padding:1px;border:1px solid black;background-color:#e35310;color:white;"><center><b>STOCK</b></center></td>'
	# comp = frappe.db.get_list("Company","name")
	comp = frappe.db.sql("""select name from `tabCompany`""",as_dict=1)
	for co in comp:
		st = 0
		ware = frappe.db.sql("""select name from `tabWarehouse` where company = '%s' and default_for_stock_transfer = 1 """%(co.name),as_dict=1)
		# ware = frappe.db.get_list("Warehouse",{"company":co.name,"default_for_stock_transfer":1},['name'])
		for w in ware:

			data += '<td colspan=1 style="width:70px;padding:1px;border:1px solid black;background-color:#e35310;color:white;"><center><b>%s</b></center></td>'%(w.name.split("-")[-1]) 
	data += '<td colspan=1 style="width:180px;padding:1px;border:1px solid black;background-color:#e35310;color:white;"><center><b>TO RECEIVE</b></center></td>'
	data += '<td colspan=1 style="width:180px;padding:1px;border:1px solid black;background-color:#e35310;color:white;"><center><b>TO SELL</b></center></td>'
	warehouses = []  
	for j in item_details:
		country = frappe.get_value("Company",{"name":company},["country"])

		warehouse_stock = frappe.db.sql("""
		select sum(b.actual_qty) as qty from `tabBin` b join `tabWarehouse` wh on wh.name = b.warehouse join `tabCompany` c on c.name = wh.company where c.country = '%s' and b.item_code = '%s'
		""" % (country,j["item_code"]),as_dict=True)[0]

		if not warehouse_stock["qty"]:
			warehouse_stock["qty"] = 0
		
		
		new_po = frappe.db.sql("""select sum(`tabPurchase Order Item`.qty) as qty,sum(`tabPurchase Order Item`.received_qty) as d_qty from `tabPurchase Order` 
		left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
		where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus = 1 and `tabPurchase Order`.status != 'Closed' """ % (j["item_code"]), as_dict=True)[0]
		if not new_po['qty']:
			new_po['qty'] = 0
		if not new_po['d_qty']:
			new_po['d_qty'] = 0
		in_transit = new_po['qty'] - new_po['d_qty']


		
		total = warehouse_stock["qty"] + in_transit

		stocks = frappe.db.sql("""select actual_qty,warehouse,stock_uom,stock_value from tabBin
		where item_code = '%s' """%(j["item_code"]),as_dict=True)

		pos = frappe.db.sql("""select `tabPurchase Order Item`.item_code as item_code,`tabPurchase Order Item`.item_name as item_name,`tabPurchase Order`.supplier as supplier,sum(`tabPurchase Order Item`.qty) as qty,`tabPurchase Order Item`.rate as rate,`tabPurchase Order`.transaction_date as date,`tabPurchase Order`.name as po from `tabPurchase Order`
		left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
		where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus != 2 order by rate asc limit 1""" % (j["item_code"]), as_dict=True)
	
		new_so = frappe.db.sql("""select sum(`tabSales Order Item`.qty) as qty,sum(`tabSales Order Item`.delivered_qty) as d_qty from `tabSales Order`
		left join `tabSales Order Item` on `tabSales Order`.name = `tabSales Order Item`.parent
		where `tabSales Order Item`.item_code = '%s' and `tabSales Order`.docstatus = 1 and `tabSales Order`.status != "Closed" """ % (j["item_code"]), as_dict=True)[0]
		if not new_so['qty']:
			new_so['qty'] = 0
		if not new_so['d_qty']:
			new_so['d_qty'] = 0
		del_total = new_so['qty'] - new_so['d_qty']
		i = 0
		for po in pos:
			data += '<tr>'
			data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>' % (j["item_code"])
			data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>' % (j["item_name"])
			data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>' % (warehouse_stock['qty'] or 0				)
			comp = frappe.db.sql("""select name from `tabCompany`""",as_dict=1)
			for co in comp:
				st = 0
				ware = frappe.db.sql("""select name from `tabWarehouse` where company = '%s' and default_for_stock_transfer = 1 """%(co.name),as_dict=1)
				for w in ware:
					sto = frappe.db.get_value("Bin",{"item_code":j["item_code"],"warehouse":w.name},['actual_qty'])
					if not sto:
						sto = 0
					st += sto
					data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>' %(st)
			data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>'%(in_transit or 0)
			data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>'%(del_total or 0)
			data += '</tr>'
		i += 1
	data += '</tr>'
	data += '</table>'
	return data

# @frappe.whitelist()
# def update_child_table():
# 	emp = frappe.get_all("Employee",{"status":"Active"},["*"])
# 	for i in emp:
# 		if i.history:
# 			print("yes")
# 		else:
# 			print("no")
# 			e = frappe.get_doc("Employee",i.name)
# 			e.append("history", {
# 			'date':'2023-09-01',
# 			'basic':i.basic,
# 			'hra':i.hra,
# 			'other_allowance':i._other_allowance,
# 			'transport_allowance':i.transportation,
# 			'medical_allowance':i.medical_allowance_,
# 			'air_ticket_allowance':i.air_ticket_allowance_,
# 			'mobile_allowance':i.mobile_allowance,
# 			'visa_cost':i.visa_cost_,
# 			'accommodation':i.accommodation,
# 			'leave_salary':i.leave_salary,
# 			'qid_cost':i.qid_cost,
# 			'medical_renewal':i.medical_renewal,
# 			'compensation':i.compensationemployee_insurence,
# 			'remark':'Salary on September-01'
# 			})
# 			e.save(ignore_permissions = True)


@frappe.whitelist()
def returnall():
	k = 0
	leave = frappe.db.get_list("Leave Application",{'docstatus':1},['name','company','creation','leave_type','status'])
	for i in leave:
		doc = frappe.db.exists("Leave Ledger Entry",{'transaction_name':i.name})
		if not doc:
			# if i.leave_type == "Casual Leave" and i.status != "Rejected":
			print(i.name)
	#             k = k + 1
	# print(k)

# @frappe.whitelist()
# def return_leave():
#     doc = frappe.db.sql(""" select employee_name,sum(total_leave_days) as total from `tabLeave Application` where leave_type = 'Casual Leave' and status = 'Approved' and docstatus = 1 group by employee """,as_dict=1)
#     for i in doc:
#         print(i.employee_name)
#         print(i.total)

#Show the below value in Print format
@frappe.whitelist()
def update_mat_total(doc):
	if doc.materials:
		rate =0
		amount = 0
		for i in doc.materials:
			rate = rate + i.rate_with_overheads
			amount = amount+ i.amount_with_overheads
			rate_value = "{:.2f}".format(rate)
			amount_value = "{:.2f}".format(amount)
		data = "<table width=100% border=1px solid black>"
		data += f'<tr><td width=71% style="border: 1px solid black; font-size:12px;text-align:center" ><b>Total</b></td><td width = 14% style="border:1px solid black; font-size:12px" nowrap>{ rate_value }</td><td colspan=1 style="border: 1px solid black; font-size:12px" >{ amount_value }</td></tr>'
		data += '</table>'
		return data

#Show the below value in Print format
@frappe.whitelist()
def update_design_total(doc):
	if doc.design:
		rate =0
		amount = 0
		for i in doc.design:
			rate = rate + i.rate_with_overheads
			amount = amount+ i.amount_with_overheads
			rate_value = "{:.2f}".format(rate)
			amount_value = "{:.2f}".format(amount)
		data = "<table width=100% border=1px solid black>"
		data += f'<tr><td width=71% style="border: 1px solid black; font-size:12px;text-align:center" ><b>Total</b></td><td width = 14% style="border:1px solid black; font-size:12px" nowrap>{ rate_value }</td><td colspan=1 style="border: 1px solid black; font-size:12px" >{ amount_value }</td></tr>'
		data += '</table>'
		return data
	
#Show the below value in Print format
@frappe.whitelist()
def update_work_total(doc):
	if doc.finishing_work:
		rate =0
		amount = 0
		for i in doc.finishing_work:
			rate = rate + i.rate_with_overheads
			amount = amount+ i.amount_with_overheads
			rate_value = "{:.2f}".format(rate)
			amount_value = "{:.2f}".format(amount)
		data = "<table width=100% border=1px solid black>"
		data += f'<tr><td width=71% style="border: 1px solid black; font-size:12px;text-align:center" ><b>Total</b></td><td width = 14% style="border:1px solid black; font-size:12px" nowrap>{ rate_value }</td><td colspan=1 style="border: 1px solid black; font-size:12px" >{ amount_value }</td></tr>'
		data += '</table>'
		return data

#Show the below value in Print format
@frappe.whitelist()
def update_acc_total(doc):
	if doc.accessories:
		rate =0
		amount = 0
		for i in doc.accessories:
			rate = rate + i.rate_with_overheads
			amount = amount+ i.amount_with_overheads
			rate_value = "{:.2f}".format(rate)
			amount_value = "{:.2f}".format(amount)
		data = "<table width=100% border=1px solid black>"
		data += f'<tr><td width=71% style="border: 1px solid black; font-size:12px;text-align:center" ><b>Total</b></td><td width = 14% style="border:1px solid black; font-size:12px" nowrap>{ rate_value }</td><td colspan=1 style="border: 1px solid black; font-size:12px" >{ amount_value }</td></tr>'
		data += '</table>'
		return data

#Show the below value in Print format
@frappe.whitelist()
def update_inst_total(doc):
	if doc.installation:
		rate =0
		amount = 0
		for i in doc.installation:
			rate = rate + i.rate_with_overheads
			amount = amount+ i.amount_with_overheads
			rate_value = "{:.2f}".format(rate)
			amount_value = "{:.2f}".format(amount)
		data = "<table width=100% border=1px solid black>"
		data += f'<tr><td width=71% style="border: 1px solid black; font-size:12px;text-align:center" ><b>Total</b></td><td width = 14% style="border:1px solid black; font-size:12px" nowrap>{ rate_value }</td><td colspan=1 style="border: 1px solid black; font-size:12px" >{ amount_value }</td></tr>'
		data += '</table>'
		return data

#Show the below value in Print format
@frappe.whitelist()
def update_power_total(doc):
	if doc.manpower:
		rate =0
		amount = 0
		for i in doc.manpower:
			rate = rate + i.rate_with_overheads
			amount = amount+ i.amount_with_overheads
			rate_value = "{:.2f}".format(rate)
			amount_value = "{:.2f}".format(amount)
		data = "<table width=100% border=1px solid black>"
		data += f'<tr><td width=71% style="border: 1px solid black; font-size:12px;text-align:center" ><b>Total</b></td><td width = 14% style="border:1px solid black; font-size:12px" nowrap>{ rate_value }</td><td colspan=1 style="border: 1px solid black; font-size:12px" >{ amount_value }</td></tr>'
		data += '</table>'
		return data

#Show the below value in Print format
@frappe.whitelist()
def update_tools_total(doc):
	if doc.heavy_equipments:
		rate =0
		amount = 0
		for i in doc.heavy_equipments:
			rate = rate + i.rate_with_overheads
			amount = amount+ i.amount_with_overheads
			rate_value = "{:.2f}".format(rate)
			amount_value = "{:.2f}".format(amount)
		data = "<table width=100% border=1px solid black>"
		data += f'<tr><td width=71% style="border: 1px solid black; font-size:12px;text-align:center" ><b>Total</b></td><td width = 14% style="border:1px solid black; font-size:12px" nowrap>{ rate_value }</td><td colspan=1 style="border: 1px solid black; font-size:12px" >{ amount_value }</td></tr>'
		data += '</table>'
		return data
	
#Show the below value in Print format
@frappe.whitelist()
def update_other_total(doc):
	if doc.others:
		rate =0
		amount = 0
		for i in doc.others:
			rate = rate + i.rate_with_overheads
			amount = amount+ i.amount_with_overheads
			rate_value = "{:.2f}".format(rate)
			amount_value = "{:.2f}".format(amount)
		data = "<table width=100% border=1px solid black>"
		data += f'<tr><td width=71% style="border: 1px solid black; font-size:12px;text-align:center" ><b>Total</b></td><td width = 14% style="border:1px solid black; font-size:12px" nowrap>{ rate_value }</td><td colspan=1 style="border: 1px solid black; font-size:12px" >{ amount_value }</td></tr>'
		data += '</table>'
		return data
	
# #Check the below condition while submission of rejoining form
# @frappe.whitelist()
# def rejoining_date_validation(doc,method):
# 	rejoin = frappe.utils.add_days(doc.end,1)
# 	late_joining_days = date_diff(doc.re_join,doc.end)
# 	dates = [add_days(rejoin, i) for i in range(0, late_joining_days)]
# 	if doc.re_join != rejoin:
# 		if not frappe.db.exists("Leave Extension",{"employee":doc.emp_no,"extension_from_date":rejoin,"total_no_of_days":late_joining_days}):
# 			for date in dates:
# 				if frappe.db.exists("Attendance",{"attendance_date":date,"employee":doc.emp_no,"docstatus":("!=",2)}):
# 					att = frappe.get_doc("Attendance",{"attendance_date":date,"employee":doc.emp_no,"docstatus":("!=",2)})
# 					att.status = "Absent"
# 					att.save(ignore_permissions=True)
# 					att.submit()
# 					frappe.db.commit()
# 				else:
# 					att=frappe.new_doc("Attendance")
# 					att.employee = doc.emp_no
# 					att.status = "Absent"
# 					att.attendance_date = date
# 					att.save(ignore_permissions=True)
# 					att.submit()
# 					frappe.db.commit()
					
#Update the discount value while save of the document
@frappe.whitelist()
def update_discount(discount, total, name):
	if discount:
		discount = float(discount)
		total = float(total)
		if total != 0:
			value = (discount / total) * 100
			return value
		else:
			return 0.0


#Avoid duplicate entry in Supplier
@frappe.whitelist()
def supplier_duplicate(supplier):
	if frappe.db.exists("Supplier",supplier) :
		frappe.throw("Supplier Already Exists")
		return False
	else:
		return True


#Show the below details in HTML Format while enter the variation Details in SO
@frappe.whitelist()
def variation_report(sales_order):
	list = frappe.db.get_list("Project Budget",{"sales_order":sales_order},["name","date_of_budget","total_bidding_price","amended_from"],order_by = "name")
	data = "<table class='table table-bordered=1'>"
	data += "<tr><td style='border: 1px solid black;background-color:#e35310;color:white'><b>Sl. No.</b></td><td style='border: 1px solid black;background-color:#e35310;color:white'><b>Document Number</b></td><td style='border: 1px solid black;background-color:#e35310;color:white'><b>Date</b></td><td style='border: 1px solid black;background-color:#e35310;color:white'><b>Amount</b></td><td style='border: 1px solid black;background-color:#e35310;color:white'><b>Amended Value</b></td><td style='border: 1px solid black;background-color:#e35310;color:white'><b></b></td></tr>"
	index = 1
	for i in list:
		data += "<tr><td style='border: 1px solid black'><b>%s</b></td><td style='border: 1px solid black'><b>%s</b></td><td style='border: 1px solid black'><b>%s</b></td><td style='border: 1px solid black'><b>%s</b></td>"%(index,i.name,i.date_of_budget,i.total_bidding_price)
		amended_bp = 0
		if i.amended_from:
			amended = frappe.get_value("Project Budget",i.amended_from,['total_bidding_price'])
			amended_bp = i.total_bidding_price - amended
		if amended_bp >= 0:
			data += "<td style='border: 1px solid black;color:green'><b>%s</b></td>"%(round(amended_bp,2))
		if amended_bp < 0:
			data += "<td style='border: 1px solid black;color:red'><b>%s</b></td>"%(round(amended_bp,2))
		if i.amended_from:
			data += """<td style='border: 1px solid black;'><button onclick="showDetails('%s','%s')">Show Details</button></td></tr>""" % (i.name,i.amended_from)
		else:
			data += """<td style='border: 1px solid black;'></td></tr>"""
		index +=1
	data+= "</table>"
	return data

#Show the below details in HTML Format In SO
@frappe.whitelist()
def detailed_variation_report(name,amended_from):
	data = "<table class='table table-bordered=1'>"
	data += "<tr><td style='border: 1px solid black;background-color:#e35310;color:white'><b>Sl. No.</b></td><td style='border: 1px solid black;background-color:#e35310;color:white'><b>MSoW ID</b></td><td style='border: 1px solid black;background-color:#e35310;color:white'><b>Description</b></td><td style='border: 1px solid black;background-color:#e35310;color:white'><b>Qty</b><td style='border: 1px solid black;background-color:#e35310;color:white'><b>UOM</b><td style='border: 1px solid black;background-color:#e35310;color:white'><b>Unit Price</b></td><td style='border: 1px solid black;background-color:#e35310;color:white'><b>Total Bidding Price</b></td></tr>"
	doc = frappe.get_doc("Project Budget",name)
	amend_doc = frappe.get_doc("Project Budget",amended_from)
	index = 1
	for i in doc.master_scope_of_work:
		for d in amend_doc.master_scope_of_work:
			if i.msow == d.msow:
				data += "<tr><td style='border: 1px solid black'><b>%s</b></td><td style='border: 1px solid black'><b>%s</b></td><td style='border: 1px solid black'><b>%s</b></td><td style='border: 1px solid black'><b>%s</b></td><td style='border: 1px solid black'><b>%s</b></td>"%(index,i.msow,i.msow_desc,i.qty,i.unit)
				if i.total_bidding_price > d.total_bidding_price:            
					data += "<td style='border: 1px solid black;color:green'><b>%s</b></td><td style='border: 1px solid black;color:green'><b>%s</b></td></tr>"%(i.unit_price,i.total_bidding_price)
				if i.total_bidding_price < d.total_bidding_price:
					data += "<td style='border: 1px solid black;color:red'><b>%s</b></td><td style='border: 1px solid black;color:red'><b>%s</b></td></tr>"%(i.unit_price,i.total_bidding_price)
				if i.total_bidding_price == d.total_bidding_price:
					data += "<td style='border: 1px solid black'><b>%s</b></td><td style='border: 1px solid black'><b>%s</b></td></tr>"%(i.unit_price,i.total_bidding_price)
		index += 1
	data+= "</table>"
	return data


# @frappe.whitelist()
# def make_item_sheet():
#     args = frappe.local.form_dict
#     filename = args.name
#     test = build_xlsx_response(filename)

# def make_xlsx(data,sheet_name=None, wb=None, column_widths=None):
#     args = frappe.local.form_dict
#     company = args.company
#     column_widths = column_widths or []
#     if wb is None:
#         wb = openpyxl.Workbook()
#     ws = wb.create_sheet(sheet_name, 0)
#     warehouse = []
#     comp = frappe.db.sql("""select name from `tabCompany`""",as_dict=1)
#     for co in comp:
#         st = 0
#         ware = frappe.db.sql("""select name from `tabWarehouse` where company = '%s' and default_for_stock_transfer = 1 """%(co.name),as_dict=1)
#         for w in ware:
#             warehouse.append(w.get('name')) 
#     ws.append(['Item Code']+['Item Name']+['Stock']+warehouse+['TO RECEIVE']+['TO SELL'])
#     doc = frappe.get_doc(args.doctype,args.name)
#     for j in doc.items:
#         country = frappe.get_value("Company",{"name":company},["country"])

#         warehouse_stock = frappe.db.sql("""
#         select sum(b.actual_qty) as qty from `tabBin` b join `tabWarehouse` wh on wh.name = b.warehouse join `tabCompany` c on c.name = wh.company where c.country = '%s' and b.item_code = '%s'
#         """ % (country,j.item_code),as_dict=True)[0]

#         if not warehouse_stock["qty"]:
#             warehouse_stock["qty"] = 0
	
#         new_po = frappe.db.sql("""select sum(`tabPurchase Order Item`.qty) as qty,sum(`tabPurchase Order Item`.received_qty) as d_qty from `tabPurchase Order` 
#         left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
#         where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus = 1 and `tabPurchase Order`.status != 'Closed' """ % (j.item_code), as_dict=True)[0]
#         if not new_po['qty']:
#             new_po['qty'] = 0
#         if not new_po['d_qty']:
#             new_po['d_qty'] = 0
#         in_transit = new_po['qty'] - new_po['d_qty']
#         pos = frappe.db.sql("""select `tabPurchase Order Item`.item_code as item_code,`tabPurchase Order Item`.item_name as item_name,`tabPurchase Order`.supplier as supplier,sum(`tabPurchase Order Item`.qty) as qty,`tabPurchase Order Item`.rate as rate,`tabPurchase Order`.transaction_date as date,`tabPurchase Order`.name as po from `tabPurchase Order`
#             left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
#             where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus != 2 order by rate asc limit 1""" % (j.item_code), as_dict=True)
		
#         new_so = frappe.db.sql("""select sum(`tabSales Order Item`.qty) as qty,sum(`tabSales Order Item`.delivered_qty) as d_qty from `tabSales Order`
#             left join `tabSales Order Item` on `tabSales Order`.name = `tabSales Order Item`.parent
#             where `tabSales Order Item`.item_code = '%s' and `tabSales Order`.docstatus = 1 and `tabSales Order`.status != "Closed"  """ % (j.item_code), as_dict=True)[0]
		

#         if not new_so['qty']:
#             new_so['qty'] = 0
#         if not new_so['d_qty']:
#             new_so['d_qty'] = 0
#         del_total = new_so['qty'] - new_so['d_qty']
		
#         i= 0
#         st = 0
#         amt = []
#         comp = frappe.db.sql("""select name from `tabCompany`""",as_dict=1)
#         for co in comp:
#             ware = frappe.db.sql("""select name from `tabWarehouse` where company = '%s' and default_for_stock_transfer = 1 """%(co.name),as_dict=1)
#             for w in ware:
#                 sto = frappe.db.get_value("Bin", {"item_code": j.item_code, "warehouse": w.name}, ['actual_qty'])
#                 if not sto:
#                     sto = 0
#                 st += sto
#                 amt.append(sto)
#         ws.append([j.item_code,j.item_name,st]+amt + [int(in_transit or 0), int(del_total or 0)])
#         i += 1
#     xlsx_file = BytesIO()
#     wb.save(xlsx_file)
#     return xlsx_file

	
@frappe.whitelist()
def unallocate():
	value=0
	pay= frappe.get_all("Payment Entry Reference",{"reference_name":"TRD-CRD-2023-00935"},["allocated_amount"])
	for i in pay:
		value +=i.allocated_amount
	print(value)


# @frappe.whitelist()
# def get_timesheet(from_date,to_date,company):
#     pos = frappe.db.sql("""
#         SELECT
#             SUM(`tabTimesheet Detail`.billing_hours) AS hours,
#             SUM(`tabTimesheet Detail`.costing_amount) AS cost_amount,
#             `tabTimesheet Detail`.project,
#             `tabTimesheet Detail`.project_name,
#             `tabTimesheet`.employee,
#             `tabTimesheet`.employee_name
#         FROM
#             `tabTimesheet`
#         LEFT JOIN
#             `tabTimesheet Detail`
#         ON
#             `tabTimesheet`.name = `tabTimesheet Detail`.parent
#         WHERE
#             `tabTimesheet`.docstatus != 2 AND
#             `tabTimesheet`.start_date BETWEEN %s AND %s AND
#             `tabTimesheet`.company = %s
#         GROUP BY
#             `tabTimesheet Detail`.project,
#             `tabTimesheet`.employee
#         ORDER BY
#             `tabTimesheet`.employee
#         """, (from_date, to_date, company), as_dict=True)

	
#     data = "<table class='table table-bordered=1'>"
#     data += """<tr><td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Sl No.</td>
#             <td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Project</td>
#             <td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Project Name</td>
#             <td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Total Hours</td>
#             <td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Total Cost</td></tr>"""
	
#     current_employee = None
#     serial_number = 1
#     total_hours = 0
#     total_cost = 0
#     total_hours_all = 0
#     total_cost_all = 0
#     total_hours_all = sum(row.hours for row in pos)
#     total_cost_all = sum(row.cost_amount for row in pos)
#     for row in pos:
#         if row.employee != current_employee:
#             if current_employee is not None:
#                 data += """<tr><td colspan=3 style='border: 1px solid black;text-align:right'><b>Total</b></td>
#                             <td style='border: 1px solid black;text-align:right'>%s</td>
#                             <td style='border: 1px solid black;text-align:right'>%s</td></tr>""" % (total_hours, round(total_cost,2))
#                 total_hours = 0
#                 total_cost = 0
				
#             data += """<tr><td colspan=5 style='background-color:#D5D8DC;border: 1px solid black'><b>%s</b></td></tr>""" % (row.employee +" - "+ row.employee_name)
#             current_employee = row.employee
#             serial_number = 1
		
#         total_hours += row.hours
#         total_cost += row.cost_amount
		
#         data += """<tr><td style='border: 1px solid black;text-align:right'>%s</td>
#                     <td style='border: 1px solid black'>%s</td>
#                     <td style='border: 1px solid black'>%s</td>
#                     <td style='border: 1px solid black;text-align:right'>%s</td>
#                     <td style='border: 1px solid black;text-align:right'>%s</td></tr>""" % (serial_number, row.project, row.project_name, row.hours, round(row.cost_amount,2))
		
#         serial_number += 1

#     if current_employee is not None:
#         data += """<tr><td colspan=3 style='border: 1px solid black;text-align:right'><b>Total</b></td>
#                     <td style='border: 1px solid black;text-align:right'>%s</td>
#                     <td style='border: 1px solid black;text-align:right'>%s</td></tr>""" % (total_hours, round(total_cost,2))
				
#     data += """<tr><td colspan=3 style='background-color:#D5D8DC;border: 1px solid black;text-align:right'><b>Overall Total</b></td>
#                 <td style='background-color:#D5D8DC;border: 1px solid black;text-align:right'><b>%s</b></td>
#                 <td style='background-color:#D5D8DC;border: 1px solid black;text-align:right'><b>%s</b></td></tr>""" % (round(total_hours_all,2), round(total_cost_all,2))

#     data += "</table>"
#     return data
# @frappe.whitelist()
# def get_timesheet(from_date, to_date):
#     pos = frappe.db.sql("""
#     SELECT 
#      SUM(`tabTimesheet Detail`.billing_hours) AS hours,
#      SUM(`tabTimesheet Detail`.costing_amount) AS cost_amount,
#     `tabTimesheet Detail`.project,
#     `tabTimesheet Detail`.project_name,
#     `tabTimesheet`.employee,
#     `tabTimesheet`.employee_name
# FROM
#     `tabTimesheet`
# LEFT JOIN
#     `tabTimesheet Detail`
# ON
#     `tabTimesheet`.name = `tabTimesheet Detail`.parent
# WHERE
#     `tabTimesheet`.docstatus != 2 AND
#     `tabTimesheet`.start_date BETWEEN %s AND %s
# GROUP BY
#     `tabTimesheet Detail`.project,
#     `tabTimesheet`.employee
# ORDER BY
#     `tabTimesheet`.employee

# """, (from_date, to_date), as_dict=True)
#     frappe.msgprint(pos)

#     total_hours = 0
#     total_cost = 0
#     total_hours_all = 0
#     total_cost_all = 0
#     total_hours_all = sum(row.hours for row in pos)
#     total_cost_all = sum(row.cost_amount for row in pos)
#     data = "<table class='table table-bordered=1'>"
#     data += """<tr><td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Sl No.</td>
#             <td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Project</td>
#             <td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Project Name</td>
#             <td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Total Hours</td>
#             <td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Total Cost</td></tr>"""
#     for index, record in enumerate(pos, start=1):
#         data += """<tr><td colspan=3 style='border: 1px solid black;text-align:right'><b>Total</b></td>
#                             <td style='border: 1px solid black;text-align:right'>%s</td>
#                             <td style='border: 1px solid black;text-align:right'>%s</td></tr>""" % (total_hours, round(total_cost,2))
#         total_hours = 0
#         total_cost = 0
				
#         data += """<tr><td colspan=5 style='background-color:#D5D8DC;border: 1px solid black'><b>%s</b></td></tr>""" % (record.employee +" - "+ record.employee_name)
#         current_employee = record.employee
#         serial_number = 1
		
#         total_hours += record.hours
#         total_cost += record.cost_amount
		
#         data += """<tr><td style='border: 1px solid black;text-align:right'>%s</td>
#                     <td style='border: 1px solid black'>%s</td>
#                     <td style='border: 1px solid black'>%s</td>
#                     <td style='border: 1px solid black;text-align:right'>%s</td>
#                     <td style='border: 1px solid black;text-align:right'>%s</td></tr>""" % (index, record.project, record.project_name, record.hours, round(record.cost_amount,2))
#         data += """<tr><td colspan=3 style='border: 1px solid black;text-align:right'><b>Total</b></td>
#                     <td style='border: 1px solid black;text-align:right'>%s</td>
#                     <td style='border: 1px solid black;text-align:right'>%s</td></tr>""" % (total_hours, round(total_cost,2))
				
#     data += """<tr><td colspan=3 style='background-color:#D5D8DC;border: 1px solid black;text-align:right'><b>Overall Total</b></td>
#             <td style='background-color:#D5D8DC;border: 1px solid black;text-align:right'><b>%s</b></td>
#             <td style='background-color:#D5D8DC;border: 1px solid black;text-align:right'><b>%s</b></td></tr>""" % (round(total_hours_all,2), round(total_cost_all,2))

#     data+="</table>"
#     return data

#Show the print format
@frappe.whitelist()
def get_timesheet(from_date, to_date,company,project):
	pos = frappe.db.sql(""" 
	SELECT 
	 SUM(`tabTimesheet Detail`.billing_hours) AS hours,
	 SUM(`tabTimesheet Detail`.costing_amount) AS cost_amount,
	 `tabTimesheet Detail`.project,
	 `tabTimesheet Detail`.project_name,
	 `tabTimesheet`.employee,
	 `tabTimesheet`.employee_name
	FROM
		`tabTimesheet`
	LEFT JOIN
		`tabTimesheet Detail`
	ON
		`tabTimesheet`.name = `tabTimesheet Detail`.parent
	LEFT JOIN
		`tabProject`
	ON
		`tabTimesheet Detail`.project = `tabProject`.name
	WHERE
		`tabTimesheet`.docstatus != 2 AND
		`tabTimesheet`.start_date BETWEEN %s AND %s AND
		`tabProject`.company = %s
	GROUP BY
		`tabTimesheet Detail`.project,
		`tabTimesheet`.employee
	ORDER BY
		`tabTimesheet`.employee
	""",(from_date, to_date,company,project), as_dict=True)
	
	# Debugging output
	frappe.msgprint(str(pos))

	total_hours_all = sum(row.hours for row in pos)
	total_cost_all = sum(row.cost_amount for row in pos)

	data = "<table class='table table-bordered=1'>"
	data += """<tr><td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Sl No.</td>
			<td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Project</td>
			<td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Project Name</td>
			<td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Total Hours</td>
			<td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Total Cost</td></tr>"""

	current_employee = None
	total_hours = 0
	total_cost = 0
	serial_number = 1

	for index, record in enumerate(pos, start=1):
		if current_employee != record.employee:  # New employee
			if current_employee is not None:  # Not the first employee
				data += """<tr><td colspan=3 style='border: 1px solid black;text-align:right'><b>Total</b></td>
							<td style='border: 1px solid black;text-align:right'>%s</td>
							<td style='border: 1px solid black;text-align:right'>%s</td></tr>""" % (total_hours, round(total_cost, 2))
				total_hours = 0
				total_cost = 0
			current_employee = record.employee
			data += """<tr><td colspan=5 style='background-color:#D5D8DC;border: 1px solid black'><b>%s - %s</b></td></tr>""" % (record.employee, record.employee_name)
		
		total_hours += record.hours
		total_cost += record.cost_amount

		data += """<tr><td style='border: 1px solid black;text-align:right'>%s</td>
					<td style='border: 1px solid black'>%s</td>
					<td style='border: 1px solid black'>%s</td>
					<td style='border: 1px solid black;text-align:right'>%s</td>
					<td style='border: 1px solid black;text-align:right'>%s</td></tr>""" % (index, record.project, record.project_name, record.hours, round(record.cost_amount, 2))
	if current_employee is not None:
		data += """<tr><td colspan=3 style='border: 1px solid black;text-align:right'><b>Total</b></td>
					<td style='border: 1px solid black;text-align:right'>%s</td>
					<td style='border: 1px solid black;text-align:right'>%s</td></tr>""" % (total_hours, round(total_cost, 2))

	data += """<tr><td colspan=3 style='background-color:#D5D8DC;border: 1px solid black;text-align:right'><b>Overall Total</b></td>
				<td style='background-color:#D5D8DC;border: 1px solid black;text-align:right'><b>%s</b></td>
				<td style='background-color:#D5D8DC;border: 1px solid black;text-align:right'><b>%s</b></td></tr>""" % (round(total_hours_all, 2), round(total_cost_all, 2))

	data += "</table>"
	return data

# @frappe.whitelist()
# def get_timesheet_project(from_date, to_date,company):
#     project_pos = frappe.db.sql("""
#         SELECT
#             `tabTimesheet Detail`.billing_hours AS hours,
#             `tabTimesheet Detail`.costing_amount AS cost_amount,
#             `tabTimesheet Detail`.costing_rate AS cost_rate,
#             `tabTimesheet Detail`.project,
#             `tabTimesheet Detail`.project_name,
#             `tabTimesheet`.start_date,
#             `tabTimesheet`.employee_name
#         FROM
#             `tabTimesheet`
#         LEFT JOIN
#             `tabTimesheet Detail`
#         ON
#             `tabTimesheet`.name = `tabTimesheet Detail`.parent
#         LEFT JOIN
#             `tabProject`
#         ON
#             `tabTimesheet Detail`.project = `tabProject`.name
#         WHERE
#             `tabTimesheet`.docstatus != 2 AND
#             `tabTimesheet`.company = %s AND
#             `tabTimesheet`.start_date BETWEEN %s AND %s
#         ORDER BY
#             `tabTimesheet Detail`.project, `tabTimesheet`.employee_name, `tabTimesheet`.start_date
#         """, (company,from_date, to_date), as_dict=True)

#     data = "<table class='table table-bordered'>"
#     data += """<tr>
#             <td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Sl.No.</td>
#             <td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Date</td>
#             <td style='width:35%;text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Project Name</td>
#             <td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Hours</td>
#             <td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Cost Per Hour</td>
#             <td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Total Cost</td>
#             </tr>"""

#     serial_number = 1
#     current_project = None
#     current_employee = None
#     current_project_total_hours = 0
#     current_project_total_cost = 0
#     current_employee_total_hours = 0
#     current_employee_total_cost = 0
#     overall_total_hours = 0
#     overall_total_cost = 0

#     for index, record in enumerate(project_pos, start=1):
#         if current_employee != record.employee_name:
#             if current_employee is not None:
#                 data += "<tr>"
#                 data += "<td colspan='3' style='border: 1px solid black;text-align:right;font-weight:bold'>Total</td>"
#                 data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(current_employee_total_hours,2))
#                 data += "<td style='border: 1px solid black;text-align:right'></td>"

#                 data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(current_employee_total_cost, 2))
#                 data += "</tr>"
#                 current_employee_total_hours = 0
#                 current_employee_total_cost = 0
		
#         if current_project != record.project:
#             if current_project is not None:
#                 data += "<tr>"
#                 data += "<td colspan='3' style='border: 1px solid black;text-align:right;font-weight:bold'>Project Total</td>"
#                 data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(current_project_total_hours,2))
#                 data += "<td style='border: 1px solid black;text-align:right'></td>"
				
#                 data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(current_project_total_cost, 2))
#                 data += "</tr>"
#                 current_project_total_hours = 0
#                 current_project_total_cost = 0

#         if current_project != record.project:
#             data += "<tr>"
#             data += "<td colspan='6' style='background-color:#D5D8DC;border: 1px solid black;font-weight:bold'>{}</td>".format(record.project)
#             data += "</tr>"
#             current_project = record.project

#         if current_employee != record.employee_name:
#             data += "<tr>"
#             data += "<td colspan='6' style='background-color:#edf2f7;border: 1px solid black;font-weight:bold'>{}</td>".format(record.employee_name)
#             data += "</tr>"
#             current_employee = record.employee_name

#         data += "<tr>"
#         data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(index)
#         data += "<td style='border: 1px solid black'>{}</td>".format(format_date(record.start_date))
#         data += "<td style='border: 1px solid black'>{}</td>".format(record.project_name)
#         data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(record.hours,2))
#         data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(record.cost_rate, 2))
#         data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(record.cost_amount, 2))
#         data += "</tr>"
#         serial_number += 1


#         current_project_total_hours += round(record.hours,2)
#         current_project_total_cost += round(record.cost_amount,2)
#         current_employee_total_hours += round(record.hours,2)
#         current_employee_total_cost += round(record.cost_amount,2)
#         overall_total_hours += round(record.hours,2)
#         overall_total_cost += round(record.cost_amount,2)

#     if current_employee is not None:
#         data += "<tr>"
#         data += "<td colspan='3' style='border: 1px solid black;text-align:right;font-weight:bold'>Total</td>"
#         data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(current_employee_total_hours,2))
#         data += "<td style='border: 1px solid black;text-align:right'></td>"
#         data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(current_employee_total_cost, 2))
#         data += "</tr>"

#     if current_project is not None:
#         data += "<tr>"
#         data += "<td colspan='3' style='border: 1px solid black;text-align:right;font-weight:bold'>Project Total</td>"
#         data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(current_project_total_hours,2))
#         data += "<td style='border: 1px solid black;text-align:right'></td>"
#         data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(current_project_total_cost, 2))
#         data += "</tr>"

#     data += "<tr>"
#     data += "<td colspan='3' style='background-color:#D5D8DC;border: 1px solid black;text-align:right;font-weight:bold'>Overall Total</td>"
#     data += "<td style='background-color:#D5D8DC;border: 1px solid black;text-align:right;font-weight:bold'>{}</td>".format(round(overall_total_hours,2))
#     data += "<td style='background-color:#D5D8DC;border: 1px solid black;text-align:right'></td>"
#     data += "<td style='background-color:#D5D8DC;border: 1px solid black;text-align:right;font-weight:bold'>{}</td>".format(round(overall_total_cost, 2))
#     data += "</tr>"

#     data += "</table>"
#     return data

#Show print format
@frappe.whitelist()
def get_timesheet_project(from_date, to_date, company):
	project_pos = frappe.db.sql("""
		SELECT
			`tabTimesheet Detail`.billing_hours AS hours,
			`tabTimesheet Detail`.costing_amount AS cost_amount,
			`tabTimesheet Detail`.costing_rate AS cost_rate,
			`tabTimesheet Detail`.project,
			`tabTimesheet Detail`.project_name,
			`tabTimesheet`.start_date,
			`tabTimesheet`.employee_name,
			`tabTimesheet`.employee
		FROM
			`tabTimesheet`
		LEFT JOIN
			`tabTimesheet Detail`
		ON
			`tabTimesheet`.name = `tabTimesheet Detail`.parent
		LEFT JOIN
			`tabProject`
		ON
			`tabTimesheet Detail`.project = `tabProject`.name
		WHERE
			`tabTimesheet`.docstatus != 2
			AND `tabProject`.company = %s
			AND `tabTimesheet`.start_date BETWEEN %s AND %s
		ORDER BY
			`tabTimesheet Detail`.project, `tabTimesheet`.employee_name, `tabTimesheet`.start_date
	""", (company, from_date, to_date), as_dict=True)

	data = "<table class='table table-bordered'>"
	data += """<tr>
			<td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Sl.No.</td>
			<td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Date</td>
			<td style='width:35%;text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Project Name</td>
			<td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Hours</td>
			<td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Cost Per Hour</td>
			<td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Total Cost</td>
			</tr>"""

	# Variables to hold running totals
	serial_number = 1
	current_project = None
	current_employee = None
	current_project_total_hours = 0
	current_project_total_cost = 0
	current_employee_total_hours = 0
	current_employee_total_cost = 0
	overall_total_hours = 0
	overall_total_cost = 0

	# Loop through the fetched timesheet records
	for index, record in enumerate(project_pos, start=1):
		# Handle employee-wise totals
		if current_employee != record.employee_name:
			if current_employee is not None:
				data += "<tr>"
				data += "<td colspan='3' style='border: 1px solid black;text-align:right;font-weight:bold'>Total</td>"
				data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(current_employee_total_hours, 2))
				data += "<td style='border: 1px solid black;text-align:right'></td>"
				data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(current_employee_total_cost, 2))
				data += "</tr>"
				current_employee_total_hours = 0
				current_employee_total_cost = 0
		# Handle project-wise totals
		if current_project != record.project:
			if current_project is not None:
				data += "<tr>"
				data += "<td colspan='3' style='border: 1px solid black;text-align:right;font-weight:bold'>Project Total</td>"
				data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(current_project_total_hours, 2))
				data += "<td style='border: 1px solid black;text-align:right'></td>"
				data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(current_project_total_cost, 2))
				data += "</tr>"
				current_project_total_hours = 0
				current_project_total_cost = 0

		# New project header
		if current_project != record.project:
			data += "<tr>"
			data += "<td colspan='6' style='background-color:#D5D8DC;border: 1px solid black;font-weight:bold'>{}</td>".format(record.project)
			data += "</tr>"
			current_project = record.project

		# New employee header
		if current_employee != record.employee_name:
			data += "<tr>"
			data += "<td colspan='6' style='background-color:#edf2f7;border: 1px solid black;font-weight:bold'>{}- {}</td>".format(record.employee,record.employee_name)
			data += "</tr>"
			current_employee = record.employee_name
	   


		# Adding timesheet detail row
		data += "<tr>"
		data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(index)
		data += "<td style='border: 1px solid black'>{}</td>".format(format_date(record.start_date))
		data += "<td style='border: 1px solid black'>{}</td>".format(record.project_name)
		data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(record.hours, 2))
		data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(record.cost_rate, 2))
		data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(record.cost_amount, 2))
		data += "</tr>"

		# Update running totals
		current_project_total_hours += round(record.hours, 2)
		current_project_total_cost += round(record.cost_amount, 2)
		current_employee_total_hours += round(record.hours, 2)
		current_employee_total_cost += round(record.cost_amount, 2)
		overall_total_hours += round(record.hours, 2)
		overall_total_cost += round(record.cost_amount, 2)

	# Final employee total
	if current_employee is not None:
		data += "<tr>"
		data += "<td colspan='3' style='border: 1px solid black;text-align:right;font-weight:bold'>Total</td>"
		data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(current_employee_total_hours, 2))
		data += "<td style='border: 1px solid black;text-align:right'></td>"
		data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(current_employee_total_cost, 2))
		data += "</tr>"

	# Final project total
	if current_project is not None:
		data += "<tr>"
		data += "<td colspan='3' style='border: 1px solid black;text-align:right;font-weight:bold'>Project Total</td>"
		data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(current_project_total_hours, 2))
		data += "<td style='border: 1px solid black;text-align:right'></td>"
		data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(current_project_total_cost, 2))
		data += "</tr>"

	# Overall total row
	data += "<tr>"
	data += "<td colspan='3' style='background-color:#D5D8DC;border: 1px solid black;text-align:right;font-weight:bold'>Overall Total</td>"
	data += "<td style='background-color:#D5D8DC;border: 1px solid black;text-align:right;font-weight:bold'>{}</td>".format(round(overall_total_hours, 2))
	data += "<td style='background-color:#D5D8DC;border: 1px solid black;text-align:right'></td>"
	data += "<td style='background-color:#D5D8DC;border: 1px solid black;text-align:right;font-weight:bold'>{}</td>".format(round(overall_total_cost, 2))
	data += "</tr>"

	data += "</table>"
	return data



@frappe.whitelist()
def update_time():
	project_pos = frappe.db.sql("""
		SELECT
			`tabTimesheet`.employee_name,
			`tabTimesheet Detail`.parent,
			`tabTimesheet Detail`.project
		FROM
			`tabTimesheet`
		LEFT JOIN
			`tabTimesheet Detail`
		ON
			`tabTimesheet`.name = `tabTimesheet Detail`.parent
		WHERE
			`tabTimesheet`.docstatus != 2
		""",as_dict=True)
	for i in project_pos:
		company = frappe.db.get_value("Project",i.project,"company")
		print(i.parent)
		print(company)
		frappe.db.set_value("Timesheet",i.parent,"company",company)



from frappe.utils import fmt_money
#Show print format
@frappe.whitelist()
def project_details_report(name,company,start_date,end_date):
	if name:
		data = '<table width = 100%>'
		project_budget = frappe.db.get_value("Project", {'name': name}, ['budgeting'])
		pb = frappe.get_doc("Project Budget", project_budget)    
		grand = frappe.db.get_value("Sales Order", {'project': name}, ['grand_total'])
		sales_in = frappe.db.sql(""" SELECT SUM(grand_total) as total FROM `tabSales Invoice` WHERE project = %s and docstatus=1""",(name,),as_dict=True)[0]
		if not sales_in["total"]:
			sales_in["total"] = 0
		bal = grand - sales_in["total"]
		revenue = frappe.db.sql(""" SELECT SUM(paid_amount) as amt FROM `tabPayment Entry` WHERE project = %s and docstatus=1""",(name,),as_dict=True)[0]
		if not revenue["amt"]:
			revenue["amt"] = 0
		out = sales_in["total"] - revenue["amt"]
		tot = 0
		dn_list = frappe.db.get_list("Delivery Note", {'project': name})
		if dn_list:
			for d_list in dn_list:
				dn = frappe.get_doc("Delivery Note", d_list.name)
				for d in dn.items:
					w_house = frappe.db.get_value("Warehouse",{'company': dn.company, 'default_for_stock_transfer': 1},['name'])
					val = frappe.db.get_value("Bin", {"item_code": d.item_code, "warehouse": w_house}, ['valuation_rate']) or 0
					total = val * d.qty
					tot += total
		gross_profit_amount=sales_in["total"]-tot
		if gross_profit_amount>0:
			gross_profit_per=(gross_profit_amount/sales_in["total"])*100
		else:
			gross_profit_per=0
		# prof = grand - tot
		# gp = (prof / grand) * 100
		data += '<table border=1px solid black width=100%>'
		data += '''<tr style="font-size:12px;background-color:#ABB2B9;text-align:center">
					<td><b>Customer</b></td>
					<td><b>Order Ref.</b></td>
					<td><b>Project Ref.</b></td>
					<td><b>Order Value</b></td>
					<td><b>Invoice Value till Date</b></td>
					<td><b>Cost Till Date</b></td>
					<td><b>Gross Profit</b></td>
					<td><b>Gross Profit %</b></td>
					<td><b>Balance To Invoice</b></td>
					<td><b>Total Collection</b></td>
					<td><b>O/S Receipts</b></td></tr>'''
		data += '''<tr style="font-size:12px;text-align:right"><td>%s</td>
					<td>%s</td><td>%s</td><td>%s</td>
					<td>%s</td><td>%s</td><td>%s</td>
					<td>%s</td><td>%s</td><td>%s</td>
					<td>%s</td></tr>''' % (pb.lead_customer_name,"",name,
			fmt_money(grand, 2), fmt_money(sales_in["total"] or 0, 2), fmt_money(round(tot, 2), 2),
			fmt_money(round(gross_profit_amount, 2), 2), fmt_money(round(gross_profit_per, 2), 2), fmt_money(round(bal, 2), 2),
			fmt_money(round(revenue["amt"], 2) or 0, 2), fmt_money(round(out, 2), 2))
		data += '</table>'
	else:
		pro_list = frappe.db.sql(""" select name from `tabProject` where company = '%s' and creation between '%s' and '%s' """%(company,start_date,end_date),as_dict = 1)
		data = '<table width = 100%>'
		data += '<table border=1px solid black width=100%>'
		data += '''<tr style="font-size:12px;background-color:#ABB2B9;text-align:center">
					<td><b>Customer</b></td>
					<td><b>Order Ref.</b></td>
					<td><b>Project Ref.</b></td>
					<td><b>Order Value</b></td>
					<td><b>Invoice Value till Date</b></td>
					<td><b>Cost Till Date</b></td>
					<td><b>Gross Profit</b></td>
					<td><b>Gross Profit %</b></td>
					<td><b>Balance To Invoice</b></td>
					<td><b>Total Collection</b></td>
					<td><b>O/S Receipts</b></td></tr>'''
		total_grand = 0
		total_sales_in_total = 0
		total_cost_till_date = 0
		total_gross_profit = 0
		total_balance_to_invoice = 0
		total_total_collection = 0
		total_out = 0
		for i in pro_list:
			project_budget = frappe.db.get_value("Project", {'name': i.name}, ['budgeting'])
			if project_budget:
				pb = frappe.get_doc("Project Budget", project_budget)    
				grand = frappe.db.get_value("Sales Order", {'project': i.name}, ['grand_total'])
				sales_in = frappe.db.sql(""" SELECT SUM(grand_total) as total FROM `tabSales Invoice` WHERE project = %s and docstatus=1""",(i.name,),as_dict=True)[0]
				if not sales_in["total"]:
					sales_in["total"] = 0
				bal = grand - sales_in["total"]
				revenue = frappe.db.sql(""" SELECT SUM(paid_amount) as amt FROM `tabPayment Entry` WHERE project = %s and docstatus=1""",(i.name,),as_dict=True)[0]
				if not revenue["amt"]:
					revenue["amt"] = 0
				out = grand - revenue["amt"]
				tot = 0
				dn_list = frappe.db.get_list("Delivery Note", {'project': i.name})
				if dn_list:
					for d_list in dn_list:
						dn = frappe.get_doc("Delivery Note", d_list.name)
						for d in dn.items:
							w_house = frappe.db.get_value("Warehouse",{'company': dn.company, 'default_for_stock_transfer': 1},['name'])
							val = frappe.db.get_value("Bin", {"item_code": d.item_code, "warehouse": w_house}, ['valuation_rate']) or 0
							total = val * d.qty
							tot += total
				prof = grand - tot
				gp = (prof / grand) * 100
				total_grand += grand
				total_cost_till_date += tot
				total_gross_profit += prof
				total_balance_to_invoice += bal
				total_total_collection += revenue["amt"]
				total_out += out
				total_sales_in_total += sales_in["total"] or 0

				
				data += '''<tr style="font-size:12px;text-align:left"><td>%s</td>
							<td>%s</td><td>%s</td><td style="text-align:right">%s</td>
							<td style="text-align:right">%s</td><td style="text-align:right">%s</td><td style="text-align:right">%s</td>
							<td style="text-align:right">%s</td><td style="text-align:right">%s</td><td style="text-align:right">%s</td>
							<td style="text-align:right">%s</td></tr>''' % (pb.lead_customer_name,"",i.name,
					fmt_money(grand, 2), fmt_money(sales_in["total"] or 0, 2), fmt_money(round(tot, 2), 2),
					fmt_money(round(prof, 2), 2), fmt_money(round(gp, 2), 2), fmt_money(round(bal, 2), 2),
					fmt_money(round(revenue["amt"], 2) or 0, 2), fmt_money(round(out, 2), 2))
		data += '''<tr style="font-size:12px;text-align:right; background-color:#E5E7E9">
			<td><b>Total</b></td>
			<td></td><td></td>
			<td>{}</td><td>{}</td><td>{}</td><td>{}</td><td></td><td>{}</td><td>{}</td><td>{}</td>
			</tr>'''.format(fmt_money(total_grand, 2), fmt_money(total_sales_in_total, 2),
							fmt_money(total_cost_till_date, 2), fmt_money(total_gross_profit, 2),
							fmt_money(total_balance_to_invoice, 2), fmt_money(total_total_collection, 2),
							fmt_money(total_out, 2))
		data += '</table>'
	return data


from frappe.utils import fmt_money
#Show print format
@frappe.whitelist()
def estimated_vs_actual(name):
	crea = frappe.db.sql("""select creation from `tabProject` where name ='%s' """ % (name), as_dict=True)[0]
	project_budget, p_name = frappe.db.get_value("Project", {'name': name}, ['budgeting', 'project_name'])
	sales_order = frappe.db.get_value("Project", {'name': name}, ['sales_order'])
	cost_est = frappe.db.get_value("Sales Order",sales_order,'cost_estimation')
	pb = frappe.get_doc("Cost Estimation", cost_est)
	data = '<table width = 100%>'
	data += '<tr><td><b>Date</b></td><td>%s</td><td><b>Refer No</b></td><td></td></tr>' % (
		format_date(crea["creation"].date()))
	data += '<tr><td><b>Project Code</b></td><td>%s</td><td><b>Project</b></td><td>%s</td></tr>' % (name, p_name)
	data += '<tr><td><b>Client</b></td><td>%s</td><td><b>Order Ref No.</b></td><td>%s</td></tr>' % (
		pb.lead_customer_name, sales_order)
	data += '</table><br>'
	grand = frappe.db.get_value("Sales Order", {'project': name}, ['grand_total'])
	sales_in = frappe.db.sql(
		""" SELECT SUM(grand_total) as total FROM `tabSales Invoice` WHERE project = %s""",
		(name,),
		as_dict=True)[0]
	if not sales_in["total"]:
		sales_in["total"] = 0
	bal = grand - sales_in["total"]
	revenue = frappe.db.sql(
		""" SELECT SUM(paid_amount) as amt FROM `tabPayment Entry` WHERE project = %s""",
		(name,),
		as_dict=True)[0]
	if not revenue["amt"]:
		revenue["amt"] = 0
	out = grand - revenue["amt"]
	tot = 0
	dn_list = frappe.db.get_list("Delivery Note", {'project': name})
	if dn_list:
		for d_list in dn_list:
			dn = frappe.get_doc("Delivery Note", d_list.name)
			for d in dn.items:
				w_house = frappe.db.get_value("Warehouse",
											  {'company': dn.company, 'default_for_stock_transfer': 1},
											  ['name'])
				val = frappe.db.get_value("Bin", {"item_code": d.item_code, "warehouse": w_house}, ['valuation_rate']) or 0
				total = val * d.qty
				tot += total
	prof = grand - tot
	gp = (prof / grand) * 100
	data += '<h3>Project Details</h3>'
	data += '<table border=1px solid black width=100%>'
	data += '<tr style="background-color:#ABB2B9;text-align:center"><td><b>Order Value</b></td><td><b>Invoice Value till Date</b></td><td><b>Cost Till Date</b></td><td><b>Gross Profit</b></td><td><b>Gross Profit %</b></td><td><b>Balance To Invoice</b></td><td><b>Total Collection</b></td><td><b>O/S Receipts</b></td></tr>'
	data += '<tr style="text-align:right"><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (
		fmt_money(grand, 2), fmt_money(sales_in["total"] or 0, 2), fmt_money(round(tot, 2), 2),
		fmt_money(round(prof, 2), 2), fmt_money(round(gp, 2), 2), fmt_money(round(bal, 2), 2),
		fmt_money(round(revenue["amt"], 2) or 0, 2), fmt_money(round(out, 2), 2))
	data += '</table><br>'
	data += '<table border=1px solid black width=100%>'
	data += '<tr style="background-color:#ABB2B9"><td colspan=4></td><td colspan=3 style="text-align:center"><b>Budgeted</b></td><td colspan=2 style="text-align:center"><b>Actual</b></td><td style="text-align:center" colspan=2><b>Variance</b></td><tr>'
	data += '<tr style="background-color:#D5D8DC"><td colspan=1><b>Budget Code</b></td><td colspan=1><b>Part Number</b></td><td colspan=1><b>Description</b></td><td colspan=1><b>Unit</b></td><td colspan=1><b>Qty</b></td><td colspan=1><b>Rate</b></td><td colspan=1><b>Amount</b></td><td colspan=1><b>Qty</b></td><td colspan=1><b>Amount</b></td><td colspan=1><b>Qty</b></td><td colspan=1><b>Amount</b></td><tr>'
	dic_items = []
	dic = [{"name": "Supply Materials", "table": pb.materials},
		   {"name": "Accessories", "table": pb.bolts_accessories},
		   {"name": "Subcontract", "table": pb.manpower_subcontract},
		   {"name": "Design", "table": pb.design_calculation},
		   {"name": "Finishing Work", "table": pb.finishing_work},
		   {"name": "Tools / Equipment / Transport / Others", "table": pb.heavy_equipments},
		   {"name": "Finished Goods", "table": pb.finished_goods}]
	for j in dic:
		if j["table"]:
			data += '<tr><td style="background-color:#EBEDEF;text-align:center" colspan=11><b>%s</b></td><tr>' % (
				j["name"])
			amt = 0
			total_amt = 0
			a_amt = 0
			total_a_amt = 0
			v_amt = 0
			total_v_amt = 0
			for i in j["table"]:
				dn = frappe.db.sql(
					"""select `tabDelivery Note Item`.item_code as item_code, `tabDelivery Note Item`.qty as qty, `tabDelivery Note Item`.amount as amount from `tabDelivery Note`
					left join `tabDelivery Note Item` on `tabDelivery Note`.name = `tabDelivery Note Item`.parent
					where `tabDelivery Note Item`.item_code = %s and `tabDelivery Note`.project = %s""",
					(i.item, name),
					as_dict=True) or 0
				if dn:
					d = dn[0]['qty']
					d_amt = dn[0]['amount']
				else:
					d = 0
					d_amt = 0
				data += '<tr><td colspan=1>%s</td><td colspan=1>%s</td><td colspan=1>%s</td><td colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><tr>' % (
					cost_est, i.item, i.description, i.unit, i.qty, fmt_money(round(i.rate_with_overheads, 2), 2),
					fmt_money(round(i.amount_with_overheads, 2), 2), d, fmt_money(round(d_amt, 2), 2),
					i.qty - d, fmt_money(round(i.amount_with_overheads, 2) - round(d_amt, 2), 2))
				amt += round(i.amount_with_overheads, 2)
				a_amt += round(d_amt, 2)
				v_amt += round(i.amount_with_overheads, 2) - round(d_amt, 2)
				total_amt += amt
				total_a_amt += a_amt
				total_v_amt += v_amt
			data += '<tr><td colspan=6></td><td style="text-align:right" colspan=1>%s</td><td style="text-align:right" colspan=1></td><td style="text-align:right" colspan=1>%s</td><td style="text-align:right" colspan=1></td><td style="text-align:right" colspan=1>%s</td><tr>' % (
				fmt_money(round(amt, 2), 2), fmt_money(round(a_amt, 2), 2), fmt_money(round(v_amt, 2), 2))
	manpower = [{"name": "Manpower", "table": pb.manpower}]
	for j in manpower:
		if j["table"]:
			data += '<tr><td style="background-color:#EBEDEF;text-align:center" colspan=11><b>%s</b></td><tr>' % (
				j["name"])
			for i in j["table"]:
				man = frappe.db.sql(
					"""select sum(`tabTimesheet`.total_costing_amount) as amt from `tabTimesheet`
					left join `tabTimesheet Detail` on `tabTimesheet`.name = `tabTimesheet Detail`.parent
					where `tabTimesheet Detail`.project = %s""",
					(name,),
					as_dict=True) or 0
				if man:
					m = man[0]['amt']
				else:
					m = 0
				data += '<tr><td colspan=1>%s</td><td colspan=1>%s</td><td colspan=1>%s</td><td colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><tr>' % (
					cost_est, i.worker, i.worker, "Nos", i.total_workers,
					fmt_money(round(i.rate_with_overheads, 2), 2),
					fmt_money(round(i.amount_with_overheads, 2), 2), i.total_workers,
					fmt_money((m or 0), 2), i.total_workers,
					fmt_money(((i.amount_with_overheads) - (m or 0)), 2))
	install = [{"name": "Installation", "table": pb.installation}]
	for j in install:
		if j["table"]:
			data += '<tr><td style="background-color:#EBEDEF;text-align:center" colspan=11><b>%s</b></td><tr>' % (
				j["name"])
			amt = 0
			total_amt = 0
			a_amt = 0
			total_a_amt = 0
			v_amt = 0
			total_v_amt = 0

			for i in j["table"]:
				dn = frappe.db.sql("""
					SELECT 
						SUM(`tabTimesheet Detail`.costing_amount) AS total_cost,
						`tabTimesheet Detail`.task,
						`tabTask`.subject AS task_name
					FROM 
						`tabTimesheet`
					LEFT JOIN 
						`tabTimesheet Detail` ON `tabTimesheet`.name = `tabTimesheet Detail`.parent
					LEFT JOIN 
						`tabTask` ON `tabTimesheet Detail`.task = `tabTask`.name
					WHERE 
						`tabTimesheet Detail`.project = '%s' AND `tabTimesheet`.docstatus = 1 AND `tabTask`.subject = '%s'
					GROUP BY 
						`tabTimesheet Detail`.task
				""" % (name,i.description), as_dict=True)
				if dn:
					completed_qty = frappe.db.get_value("Task",dn[0]['task'],'completed_qty') or 0
					d = completed_qty
					d_amt = round(dn[0]['total_cost'],2)
				else:
					d = 0
					d_amt = 0
					
				data += '<tr><td colspan=1>%s</td><td colspan=1>%s</td><td colspan=1>%s</td><td colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><tr>' % (cost_est, i.item, i.description, i.unit, i.qty,fmt_money(round(i.rate_with_overheads, 2), 2),fmt_money(round(i.amount_with_overheads, 2), 2),d,d_amt,i.qty - d, fmt_money(round(i.amount_with_overheads, 2) - round(d_amt, 2), 2))
				amt += round(i.amount_with_overheads, 2)
				a_amt += round(d_amt, 2)
				v_amt += round(i.amount_with_overheads, 2) - round(d_amt, 2)
				total_amt += amt
				total_a_amt += a_amt
				total_v_amt += v_amt
			data += '<tr><td colspan=6></td><td style="text-align:right" colspan=1>%s</td><td style="text-align:right" colspan=1></td><td style="text-align:right" colspan=1>%s</td><td style="text-align:right" colspan=1></td><td style="text-align:right" colspan=1>%s</td><tr>' % (
				fmt_money(round(amt, 2), 2), fmt_money(round(a_amt, 2), 2), fmt_money(round(v_amt, 2), 2))

	data += '</table>'
	return data

#Check the below condition and throw message while save of DN and SI
@frappe.whitelist()
def update_credit_limit(doc,method):
	if doc.is_return ==0:
		is_internal_customer = frappe.db.get_value("Customer",doc.customer,['is_internal_customer'])
		if doc.doctype == "Delivery Note":
			customer_type = frappe.db.get_value("Customer",doc.customer,['customer_type'])
			customer_list = get_details(doc.company,doc.customer)
			if not customer_list and customer_type == "Company":
				frappe.throw(_(f"Credit Limit not set for this Customer-{doc.customer}"))
				doc.save(ignore_permissions=False)
		if doc.doctype == "Sales Invoice" and doc.invoice_type == "Credit" and is_internal_customer == 0:
			customer_type = frappe.db.get_value("Customer",doc.customer,['customer_type'])
			customer_list = get_details(doc.company,doc.customer)
			if not customer_list and customer_type == "Company":
				frappe.throw(_(f"Credit Limit not set for this Customer-{doc.customer}"))
				doc.save(ignore_permissions=False)

#create new document while save of supplier and customer
@frappe.whitelist()
def create_existing(doc,method):
	if doc.doctype == 'Customer':
		if not frappe.db.exists('Customer Name',doc.customer_name):
			cn = frappe.new_doc('Customer Name')
			cn.customer = doc.customer_name
			cn.insert()
			cn.save(ignore_permissions=True)
	if doc.doctype == 'Supplier':
		if not frappe.db.exists('Supplier Name',doc.supplier_name):
			sn = frappe.new_doc('Supplier Name')
			sn.supplier = doc.supplier_name
			sn.insert()
			sn.save(ignore_permissions=True)
			
#Updates the below value in item table while submission of SO
@frappe.whitelist()
def update_pb_sow(doc,method):
	if doc.order_type == "Project":
		for i in doc.items:
			if i.msow:
				list = frappe.get_doc("PB SOW",{'sales_order':doc.name,'msow':i.msow},'name')
				for d in list.design:
					if i.work_title == "DESIGN":
						if d.item == i.item_code and d.qty ==i.qty and round(d.rate_with_overheads,3) == round(i.rate,3):
							d.docname = i.name
							list.save(ignore_permissions=True)
				for d in list.materials:
					if i.work_title == "SUPPLY MATERIALS":
						if d.item == i.item_code and d.qty ==i.qty and round(d.rate_with_overheads,3) == round(i.rate,3):
							d.docname = i.name
							list.save(ignore_permissions=True)
				for d in list.mep_materials:
					if i.work_title == "SUPPLY MATERIALS":
						if d.item == i.item_code and d.qty ==i.qty and round(d.rate_with_overheads,3) == round(i.rate,3):
							d.docname = i.name
							list.save(ignore_permissions=True)
				for d in list.finishing_work:
					if i.work_title == "FINISHING WORK":
						if d.item == i.item_code and d.qty ==i.qty and round(d.rate_with_overheads,3) == round(i.rate,3):
							d.docname = i.name
							list.save(ignore_permissions=True)

				for d in list.bolts_accessories:
					if i.work_title == "ACCESSORIES":
						if d.item == i.item_code and d.qty ==i.qty and round(d.rate_with_overheads,3) == round(i.rate,3):
							d.docname = i.name
							list.save(ignore_permissions=True)
				
				for d in list.installation:
					if i.work_title == "INSTALLATION":
						if d.item == i.item_code and d.qty ==i.qty and round(d.rate_with_overheads,3) == round(i.rate,3):
							d.docname = i.name
							list.save(ignore_permissions=True)
				
				for d in list.others:
					if i.work_title == "SUBCONTRACT":
						if d.item == i.item_code and d.qty ==i.qty and round(d.rate_with_overheads,3) == round(i.rate,3):
							d.docname = i.name
							list.save(ignore_permissions=True)
				
				for d in list.finished_goods:
					if i.work_title == "FINISHED GOODS":
						if d.item == i.item_code and d.qty ==i.qty and round(d.rate_with_overheads,3) == round(i.rate,3):
							d.docname = i.name
							list.save(ignore_permissions=True)

				for d in list.heavy_equipments:
					if i.work_title == "TOOLS/EQUIPMENTS/TRANSPORT/OTHERS":
						if d.item == i.item_code and d.qty ==i.qty and round(d.rate_with_overheads,3) == round(i.rate,3):
							d.docname = i.name
							list.save(ignore_permissions=True)

				for d in list.manpower:
					if i.work_title == "MANPOWER":
						if d.worker == i.item_code and d.total_workers ==i.qty and round(d.rate_with_overheads,3) == round(i.rate,3):
							d.docname = i.name
							list.save(ignore_permissions=True)

#Updates the below value in item table while ENTER THE button of update name in SO
@frappe.whitelist()
def updatepbsow(docname):
	doc = frappe.get_doc("Sales Order",docname)
	if doc.order_type == "Project":
		for i in doc.items:

			list = frappe.get_doc("PB SOW",{'sales_order':doc.name,'msow':i.msow},'name')
			for d in list.design:
				if i.work_title == "DESIGN":
					if d.item == i.item_code and d.qty ==i.qty and round(d.rate_with_overheads,3) == round(i.rate,3):
						d.docname = i.name
						list.save(ignore_permissions=True)
			for d in list.materials:
				if i.work_title == "SUPPLY MATERIALS":
					if d.item == i.item_code and d.qty ==i.qty and round(d.rate_with_overheads,3) == round(i.rate,3):
						d.docname = i.name
						list.save(ignore_permissions=True)
			for d in list.mep_materials:
				if i.work_title == "SUPPLY MATERIALS":
					if d.item == i.item_code and d.qty ==i.qty and round(d.rate_with_overheads,3) == round(i.rate,3):
						d.docname = i.name
						list.save(ignore_permissions=True)
			for d in list.finishing_work:
				if i.work_title == "FINISHING WORK":
					if d.item == i.item_code and d.qty ==i.qty and round(d.rate_with_overheads,3) == round(i.rate,3):
						d.docname = i.name
						list.save(ignore_permissions=True)

			for d in list.bolts_accessories:
				if i.work_title == "ACCESSORIES":
					if d.item == i.item_code and d.qty ==i.qty and round(d.rate_with_overheads,3) == round(i.rate,3):
						d.docname = i.name
						list.save(ignore_permissions=True)
			
			for d in list.installation:
				if i.work_title == "INSTALLATION":
					if d.item == i.item_code and d.qty ==i.qty and round(d.rate_with_overheads,3) == round(i.rate,3):
						d.docname = i.name
						list.save(ignore_permissions=True)
			
			for d in list.others:
				if i.work_title == "SUBCONTRACT":
					if d.item == i.item_code and d.qty ==i.qty and round(d.rate_with_overheads,3) == round(i.rate,3):
						d.docname = i.name
						list.save(ignore_permissions=True)
			
			for d in list.finished_goods:
				if i.work_title == "FINISHED GOODS":
					if d.item == i.item_code and d.qty ==i.qty and round(d.rate_with_overheads,3) == round(i.rate,3):
						d.docname = i.name
						list.save(ignore_permissions=True)

			for d in list.heavy_equipments:
				if i.work_title == "TOOLS/EQUIPMENTS/TRANSPORT/OTHERS":
					if d.item == i.item_code and d.qty ==i.qty and round(d.rate_with_overheads,3) == round(i.rate,3):
						d.docname = i.name
						list.save(ignore_permissions=True)

			for d in list.manpower:
				if i.work_title == "MANPOWER":
					if d.worker == i.item_code and d.total_workers ==i.qty and round(d.rate_with_overheads,3) == round(i.rate,3):
						d.docname = i.name
						list.save(ignore_permissions=True)



def test_bin():
	# from erpnext.stock.doctype.stock_reservation_entry.stock_reservation_entry import (
	# 		get_sre_reserved_qty_for_item_and_warehouse,
	# 	)
	# item_code='114-40002104GY'
	# warehouse = 'Electra Trading Warehouse - TDE'
	# reserved_stock = get_sre_reserved_qty_for_item_and_warehouse(item_code, warehouse)
	# print(reserved_stock)
	bin = frappe.get_doc('Bin','a2d710f1a0')
	bin.db_set("reserved_stock", flt('0'), update_modified=True)


@frappe.whitelist()
def stock_request_test(req):
	at=frappe.get_doc("Sales Order",req) 
	st=frappe.get_doc("Project Budget",at.project_budget)
	if st:
		return st.item_table


#Check the avilable qty while submission of DN WIP
@frappe.whitelist()
def check_stock_quantity(item_code,warehouse,qty):
	if frappe.db.exists("Item",{'name':item_code,'is_stock_item':1}):
		if not frappe.db.exists("Bin",{'warehouse':warehouse,'item_code':item_code}):
			return "Not exists"
		else:
			bin_doc=frappe.get_doc("Bin",{'warehouse':warehouse,'item_code':item_code})
			if bin_doc.actual_qty < float(qty):
				
				return "Not exists"

#Check the avilable qty while submission of DN WIP
@frappe.whitelist()
def check_cnstock_quantity(item_code,warehouse,qty):
	if frappe.db.exists("Item",{'name':item_code,'is_stock_item':1}):
		if not frappe.db.exists("Bin",{'warehouse':warehouse,'item_code':item_code}):
			return "Not exists"
		else:
			bin_doc=frappe.get_doc("Bin",{'warehouse':warehouse,'item_code':item_code})
			if bin_doc.actual_qty < float(qty):
				
				return "Not exists"
			

@frappe.whitelist()
def update_so_from_csv(filename):
	from frappe.utils.csvutils import read_csv_content
	from frappe.utils.file_manager import get_file
	filepath = get_file(filename)
	pps = read_csv_content(filepath[1])
	i=1
	count = 1
	for pp in pps:
		so = frappe.db.exists("Sales Order",pp[0])
		if so:
			count += 1
			frappe.db.set_value('Sales Order',so,"grand_total",pp[1])
			frappe.db.set_value('Sales Order',so,"total",pp[2])
			
	
			 # frappe.db.sql("""update `tabCase` set case_status = %s where name = %s""",(pp[1],pp[0]))
			# frappe.db.sql("""delete from `tabEmployment` where name = %s""",(pp[0]),as_dict=True)

#Update values in PB SOW While submission of Stock Entry
@frappe.whitelist()
def update_finished_goods(doc,method):
	if doc.stock_entry_type =="Manufacture" and doc.company=="ENGINEERING DIVISION - ELECTRA":
		stock_entry = frappe.get_doc("Stock Entry", doc.name)
		project_name = stock_entry.project
		pro=frappe.db.get_value("Project",{"name":project_name},["budgeting"])
		pb_sow_name = frappe.db.get_value("PB SOW", {"project_budget": pro}, ["name"])
		pb_sow_doc = frappe.get_doc("PB SOW", pb_sow_name)
		for item in stock_entry.items:
			if item.is_finished_item ==1:
				pb_sow_doc.append("finished_goods", {
					"item": item.item_code,
					"qty": item.qty,
					"unit_price":item.basic_rate,
					"se_doc_name":doc.name
				})
			for i in pb_sow_doc.raw_materials:
				if i.item == item.item_code:
					
					i.delivered_qty+=item.qty
					i.billed_qty+=item.qty
		pb_sow_doc.save()
		frappe.db.commit()

#Update values in PB SOW While submission of Stock Entry
@frappe.whitelist()
def update_finished_goods_in_project_budget(doc,method):
	if doc.stock_entry_type =="Manufacture" and doc.company=="ENGINEERING DIVISION - ELECTRA":
		stock_entry = frappe.get_doc("Stock Entry", doc.name)
		project_name = stock_entry.project
		pro=frappe.db.get_value("Project",{"name":project_name},["budgeting"])
		pb_doc = frappe.get_doc("Project Budget",pro )
		for item in stock_entry.items:
			if item.is_finished_item ==1:
				
				pb_doc. append("finished_goods", {
					"item": item.item_code,
					"qty": item.qty,
					"unit_price":item.basic_rate,
					"se_doc_name":doc.name
				})
			for i in pb_doc.raw_materials:
				if i.item == item.item_code:
					i.delivered_qty+=item.qty
					i.billed_qty+=item.qty
		pb_doc.save()
		frappe.db.commit()

#Update values in PB SOW While submission of Stock Entry
@frappe.whitelist()
def update_finished_goods_in_so(doc,method):
	if doc.stock_entry_type =="Manufacture" and doc.company=="ENGINEERING DIVISION - ELECTRA":
		stock_entry = frappe.get_doc("Stock Entry", doc.name)
		project_name = stock_entry.project
		pro=frappe.db.get_value("Project",{"name":project_name},["sales_order"])
		pb_doc = frappe.get_doc("Sales Order",pro )
		for item in stock_entry.items:
			if item.is_finished_item ==1:
				
				pb_doc. append("finished_goods", {
					"item": item.item_code,
					"qty": item.qty,
					"unit_price":item.basic_rate,
					"se_doc_name":doc.name
				})
			for i in pb_doc.custom_raw_materials:
				if i.item == item.item_code:
					
					i.delivered_qty+=item.qty
					i.billed_qty+=item.qty
					
		pb_doc.save()
		frappe.db.commit()

#Update values in PB SOW While submission of Stock Entry
@frappe.whitelist()
def update_finished_goods_in_cesow(doc,method):
	if doc.stock_entry_type =="Manufacture" and doc.company=="ENGINEERING DIVISION - ELECTRA":
		stock_entry = frappe.get_doc("Stock Entry", doc.name)
		project_name = stock_entry.project
		pro=frappe.db.get_value("Project",{"name":project_name},["budgeting"])
		pb_sow_name = frappe.db.get_value("PB SOW", {"project_budget": pro}, ["cost_estimation"])
		value=frappe.db.get_value("CE SOW",{"cost_estimation":pb_sow_name},["name"])
		
		parent_doc=frappe.get_doc("CE SOW",value)
		parent_doc.i_finished_goods=1
		parent_doc.save()
		frappe.db.commit()

#Update values in PB SOW While cancel of Stock Entry
@frappe.whitelist()
def update_qty_cancel(doc,method):
	if doc.stock_entry_type =="Manufacture" and doc.company=="ENGINEERING DIVISION - ELECTRA":
		stock_entry = frappe.get_doc("Stock Entry", doc.name)
		project_name = stock_entry.project
		pro=frappe.db.get_value("Project",{"name":project_name},["budgeting"])
		pb_doc = frappe.get_doc("Project Budget",pro )
		for item in stock_entry.items:
			for i in pb_doc.raw_materials:
				if i.item == item.item_code:
					i.delivered_qty-=item.qty
					i.billed_qty-=item.qty
			pb_doc.finished_goods = [
			fg for fg in pb_doc.finished_goods if not (fg.item == item.item_code and fg.se_doc_name == doc.name)]
		for idx, fg in enumerate(pb_doc.finished_goods, start=1):
			fg.idx = idx
		pb_doc.save()
		frappe.db.commit()

#Update values in PB SOW While cancel of Stock Entry
@frappe.whitelist()
def update_qty_in_pbsow_cancel(doc,method):
	if doc.stock_entry_type =="Manufacture" and doc.company=="ENGINEERING DIVISION - ELECTRA":
		stock_entry = frappe.get_doc("Stock Entry", doc.name)
		project_name = stock_entry.project
		pro=frappe.db.get_value("Project",{"name":project_name},["budgeting"])
		pb_sow_name = frappe.db.get_value("PB SOW", {"project_budget": pro}, ["name"])
		pb_sow_doc = frappe.get_doc("PB SOW", pb_sow_name)
		for item in stock_entry.items:
			for i in pb_sow_doc.raw_materials:
				if i.item == item.item_code:
					i.delivered_qty-=item.qty
					i.billed_qty-=item.qty
			pb_sow_doc.finished_goods = [
			fg for fg in pb_sow_doc.finished_goods if not (fg.item == item.item_code and fg.se_doc_name == doc.name)]
		for idx, fg in enumerate(pb_sow_doc.finished_goods, start=1):
			fg.idx = idx
		pb_sow_doc.save()
		frappe.db.commit()

#Update values in PB SOW While cancel of Stock Entry
@frappe.whitelist()
def update_qty_in_so_cancel(doc,method):
	if doc.stock_entry_type =="Manufacture" and doc.company=="ENGINEERING DIVISION - ELECTRA":
		stock_entry = frappe.get_doc("Stock Entry", doc.name)
		project_name = stock_entry.project
		pro=frappe.db.get_value("Project",{"name":project_name},["sales_order"])
		pb_doc = frappe.get_doc("Sales Order",pro )
		for item in stock_entry.items:
			for i in pb_doc.custom_raw_materials:
				if i.item == item.item_code:
					i.delivered_qty-=item.qty
					i.billed_qty-=item.qty
			pb_doc.finished_goods = [
			fg for fg in pb_doc.finished_goods if not (fg.item == item.item_code and fg.se_doc_name == doc.name)]
		for idx, fg in enumerate(pb_doc.finished_goods, start=1):
			fg.idx = idx
		pb_doc.save()
		frappe.db.commit()

#Ageing report print format in Report dashboard
@frappe.whitelist()
def ageing_report_multicompany(doc):
	in_amount = 0
	paid_amount = 0
	credit_note = 0
	out_amount = 0
	age_0_30 = 0
	age_31_60 = 0
	age_61_90 = 0
	age_91_120 = 0
	age_above_121 = 0
	paid = 0
	combined_data =[]
	data = "<table border='1px solid black' width='100%'><tr style='font-size:12px'><td width=10%><b>Posting Date</b></td><td width=10%><b style='text-align:center;'>Voucher No</b></td><td width=10%><b style='text-align:center'>Customer LPO</b></td><td width=10%><b style='text-align:center'>Invoiced Amount</b></td><td width=10%><b style='text-align:center'>Paid Amount</b></td><td width=10%><b style='text-align:center'>Credit Note</b></td><td width=10%><b style='text-align:center'>Outstanding Amount</b></td><td width=5%><b style='text-align:center'>Age (Days)</b></td><td width=5%><b style='text-align:center'>0- 30</b></td><td width=5%><b style='text-align:center'>31-  60</b></td><td width=5%><b style='text-align:center'>61-  90</b></td><td width=5%><b style='text-align:center'>91-120</b></td><td width=5%><b style='text-align:center'>Above 121</b></td></tr>"
	for c in doc.company_multiselect:
		if doc.customer:
			si_list = frappe.db.sql(
				"""SELECT * FROM `tabSales Invoice` WHERE company = %s and customer = %s and docstatus = 1 and is_return = 0 and invoice_type='Credit' and status !='Paid'  order by posting_date  ASC""",
				(c.company, doc.customer),
				as_dict=True
			)
			for i in si_list:
				result= frappe.db.sql("""
					SELECT sum(grand_total) as total
					FROM `tabSales Invoice` 
					WHERE company = %s AND return_against = %s AND docstatus = 1
				""", (c.company, i.name))
				return_amount = result[0][0] if result and result[0][0] else 0
				
				result_doc = frappe.db.sql("""
					SELECT name
					FROM `tabSales Invoice` 
					WHERE company = %s AND return_against = %s AND docstatus = 1
				""", (c.company, i.name), as_dict=True)
				
				pay_doc = []
				if result_doc:
					pay_doc = frappe.db.sql("""
						SELECT per.allocated_amount 
						FROM `tabPayment Entry Reference` AS per
						LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
						WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
					""", (c.company, result_doc[0]["name"]), as_dict=True)
				pay = frappe.db.sql("""
					SELECT per.allocated_amount 
					FROM `tabPayment Entry Reference` AS per
					LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
					WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
				""", (c.company, i.name), as_dict=True)
				value = sum(j["allocated_amount"] for j in pay) + sum(r["allocated_amount"] for r in pay_doc)

				jv = frappe.db.sql("""
					SELECT credit_in_account_currency 
					FROM `tabJournal Entry Account` AS per
					LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
					WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
				""", (c.company, i.name), as_dict=True)
				for k in jv:
					value += k.credit_in_account_currency
				

				if value and return_amount:
					outstanding = i.grand_total - value + return_amount
				elif value:
					outstanding = i.grand_total - value
				elif return_amount:
					outstanding = i.grand_total + return_amount
				else:
					outstanding = i.grand_total
				
				out_amount += outstanding
				age = date_diff(today(), i.posting_date) if i.posting_date else 0

				if round(outstanding) != 0:
					if value:
						paid_amount += value
					if return_amount:
						credit_note += return_amount
					in_amount += i.grand_total
					if 0 <= age <= 30:
						age_0_30 += outstanding
					elif 31 <= age <= 60:
						age_31_60 += outstanding
					elif 61 <= age <= 90:
						age_61_90 += outstanding
					elif 91 <= age <= 120:
						age_91_120 += outstanding
					else:
						age_above_121 += outstanding
					combined_data.append({
						'posting_date': i.posting_date,
						'name': i.name,
						'po_no': i.po_no if i.po_no else '-',
						'grand_total': i.grand_total,
						'paid_amount': value if value else 0,
						'credit_note': return_amount if return_amount else 0,
						'outstanding': outstanding if outstanding else '-',
						'age': age,
						'oustanding_0_30':outstanding if 0 <= age <= 30 else '-',
						'oustanding_31_60':outstanding if 31 <= age <= 60 else '-',
						'oustanding_61_90':outstanding if 61 <= age <= 90 else '-',
						'oustanding_91_120':outstanding if 91 <= age <= 120 else '-',
						'oustanding_above_121':outstanding if age > 120 else '-',
					})

			sales = frappe.db.sql(
				"""SELECT * FROM `tabSales Invoice` WHERE company = %s and customer = %s and docstatus = 1 and is_return = 1 and invoice_type='Credit' and status !='Paid' AND (return_against IS NULL OR return_against = '')  order by posting_date  ASC""",
				(c.company, doc.customer),
				as_dict=True
			)
			for a in sales:
				pay = frappe.db.sql(""" SELECT per.allocated_amount FROM `tabPayment Entry Reference` AS per
				LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
				WHERE per.reference_name = '%s' AND pe.docstatus = 1 and pe.company = '%s' """ % (a.name, c.company), as_dict=True)
				value = sum(j.allocated_amount for j in pay)

				jv = frappe.db.sql(""" SELECT credit_in_account_currency FROM `tabJournal Entry Account` AS per
				LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
				WHERE per.reference_name = '%s' AND pe.docstatus = 1 and pe.company = '%s'""" % (a.name, c.company), as_dict=True)
				value += sum(k.credit_in_account_currency for k in jv)

				outstanding = a.grand_total - value if value else a.grand_total
				out_amount += outstanding
				age = date_diff(today(), a.posting_date) if a.posting_date else 0

				if round(outstanding) != 0:
					if value:
						paid_amount += value
					if a.grand_total:
						credit_note += a.grand_total
					in_amount += a.grand_total
					if 0 <= age <= 30:
						age_0_30 += outstanding
					elif 31 <= age <= 60:
						age_31_60 += outstanding
					elif 61 <= age <= 90:
						age_61_90 += outstanding
					elif 91 <= age <= 120:
						age_91_120 += outstanding
					else:
						age_above_121 += outstanding
					combined_data.append({
						'posting_date': a.posting_date,
						'name': a.name,
						'po_no':a.po_no if i.po_no else '-',
						'grand_total': a.grand_total,
						'paid_amount': value if value else 0,
						'credit_note': a.grand_total if a.grand_total else 0,
						'outstanding': outstanding if outstanding else '-',
						'age': age,
						'oustanding_0_30':outstanding if 0 <= age <= 30 else '-',
						'oustanding_31_60':outstanding if 31 <= age <= 60 else '-',
						'oustanding_61_90':outstanding if 61 <= age <= 90 else '-',
						'oustanding_91_120':outstanding if 91 <= age <= 120 else '-',
						'oustanding_above_121':outstanding if age > 120 else '-',
					})

			payment = frappe.db.sql("""
				SELECT * FROM `tabPayment Entry` 
				WHERE company = %s AND party = %s AND docstatus = 1 
				AND payment_type = 'Receive' 
				ORDER BY posting_date ASC
			""", (c.company, doc.customer), as_dict=True)
			for v in payment:
				unallocated_amount = v.unallocated_amount
				paid_amount += unallocated_amount
				out_amount -= unallocated_amount
				age = date_diff(today(), v.posting_date)
				if unallocated_amount != 0:
					if 0 <= age <= 30:
						age_0_30 -= unallocated_amount
					elif 31 <= age <= 60:
						age_31_60 -= unallocated_amount
					elif 61 <= age <= 90:
						age_61_90 -= unallocated_amount
					elif 91 <= age <= 120:
						age_91_120 -= unallocated_amount
					else:
						age_above_121 -= unallocated_amount
					combined_data.append({
						'posting_date': v.posting_date,
						'name': v.name,
						'po_no': v.reference_no if v.reference_no else '-',
						'grand_total': 0,
						'paid_amount': unallocated_amount if unallocated_amount else 0,
						'credit_note': 0,
						'outstanding': -unallocated_amount if unallocated_amount else '-',
						'age': age,
						'oustanding_0_30':-unallocated_amount if 0 <= age <= 30 else '-',
						'oustanding_31_60':-unallocated_amount if 31 <= age <= 60 else '-',
						'oustanding_61_90':-unallocated_amount if 61 <= age <= 90 else '-',
						'oustanding_91_120':-unallocated_amount if 91 <= age <= 120 else '-',
						'oustanding_above_121':-unallocated_amount if age > 120 else '-',
					})
			
			journal = frappe.db.sql("""
				SELECT per.credit_in_account_currency, per.debit_in_account_currency, pe.name, pe.posting_date 
				FROM `tabJournal Entry Account` AS per
				LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
				WHERE pe.company = %s AND per.account LIKE %s AND pe.docstatus = 1 
				AND party_type = 'Customer' AND party = %s AND per.reference_name IS NULL
			""", (c.company, '%Debtors -%', doc.customer), as_dict=True)
			for jour in journal:
				if jour.credit_in_account_currency:
					journ_amount_credit = jour.credit_in_account_currency
					paid_amount += journ_amount_credit
					in_amount -= journ_amount_credit
					out_amount -= journ_amount_credit
					age = date_diff(today(), jour.posting_date)

					if 0 <= age <= 30:
						age_0_30 -= jour.credit_in_account_currency
					elif 31 <= age <= 60:
						age_31_60 -= jour.credit_in_account_currency
					elif 61 <= age <= 90:
						age_61_90 -= jour.credit_in_account_currency
					elif 91 <= age <= 120:
						age_91_120 -= jour.credit_in_account_currency
					else:
						age_above_121 -= jour.credit_in_account_currency
					combined_data.append({
						'posting_date': jour.posting_date,
						'name': jour.name,
						'po_no':'-',
						'grand_total': -jour.credit_in_account_currency,
						'paid_amount': 0,
						'credit_note': '-',
						'outstanding': -jour.credit_in_account_currency,
						'age': age,
						'oustanding_0_30':-jour.credit_in_account_currency if 0 <= age <= 30 else '-',
						'oustanding_31_60':-jour.credit_in_account_currency if 31 <= age <= 60 else '-',
						'oustanding_61_90':-jour.credit_in_account_currency if 61 <= age <= 90 else '-',
						'oustanding_91_120':-jour.credit_in_account_currency if 91 <= age <= 120 else '-',
						'oustanding_above_121':-jour.credit_in_account_currency if age > 120 else '-',
					})
	
				elif jour.debit_in_account_currency:
					journ_amount_debit = jour.debit_in_account_currency
					in_amount += journ_amount_debit
					out_amount += journ_amount_debit
					age = date_diff(today(), jour.posting_date)

					if 0 <= age <= 30:
						age_0_30 += jour.debit_in_account_currency
					elif 31 <= age <= 60:
						age_31_60 += jour.debit_in_account_currency
					elif 61 <= age <= 90:
						age_61_90 += jour.debit_in_account_currency
					elif 91 <= age <= 120:
						age_91_120 += jour.debit_in_account_currency
					else:
						age_above_121 += jour.debit_in_account_currency
	
					combined_data.append({
						'posting_date': jour.posting_date,
						'name': jour.name,
						'po_no':'-',
						'grand_total': jour.debit_in_account_currency,
						'paid_amount':0,
						'credit_note': 0,
						'outstanding': jour.debit_in_account_currency,
						'age': age,
						'oustanding_0_30':jour.debit_in_account_currency if 0 <= age <= 30 else '-',
						'oustanding_31_60':jour.debit_in_account_currency if 31 <= age <= 60 else '-',
						'oustanding_61_90':jour.debit_in_account_currency if 61 <= age <= 90 else '-',
						'oustanding_91_120':jour.debit_in_account_currency if 91 <= age <= 120 else '-',
						'oustanding_above_121':jour.debit_in_account_currency if age > 120 else '-',
					})
	combined_data = sorted(combined_data, key=lambda x: x['posting_date'])
	for entry in combined_data:
		data += f"""<tr style='font-size:10px'>
		   <td>{formatdate(entry['posting_date'],'dd-mm-yyyy')}</td>
		   <td>{entry['name']}</td>
		   <td>{entry['po_no']}</td>
		   <td>{fmt_money(round(entry['grand_total'], 2))}</td>
		   <td>{fmt_money(round(entry['paid_amount'], 2)) if entry['paid_amount']  else '-'}</td>
		   <td>{fmt_money(round(entry['credit_note'], 2)) if entry['credit_note'] else '-'}</td>
		   <td>{fmt_money(round(entry['outstanding'], 2)) if entry['outstanding'] else '-'}</td>
		   <td>{entry['age']}</td>
		   <td>{entry['oustanding_0_30']}</td>
		   <td>{entry['oustanding_31_60']}</td>
		   <td>{entry['oustanding_61_90']}</td>
		   <td>{entry['oustanding_91_120']}</td>
		   <td>{entry['oustanding_above_121']}</td>
		   </tr>"""
		

	data += f"<tr style='font-size:10px'><td width=10%></td><td width=10%></td><td width=10%><b>Total</b></td><td width=10%><b>{fmt_money(round(in_amount, 2))}</b></td><td width=10%><b>{fmt_money(round(paid_amount, 2))}</b></td><td width=10%><b>{fmt_money(round(credit_note, 2))}</b></td><td width=10%><b>{fmt_money(round(out_amount, 2))}</b></td><td width=5%></td><td width=5%><b>{fmt_money(round(age_0_30, 2))}</b></td><td width=5%><b>{fmt_money(round(age_31_60, 2))}</b></td><td width=5%><b>{fmt_money(round(age_61_90, 2))}</b></td><td width=5%><b>{fmt_money(round(age_91_120, 2))}</b></td><td width=5%><b>{fmt_money(round(age_above_121, 2))}</b></td></tr>"
	data += "</table>"
	return data

#CESOW Added
@frappe.whitelist()
def get_cesow(cost, msid, cso):
	sow = frappe.get_value("CE Master Scope of Work", {'parent': cost, 'msow': msid}, 'name')
	if sow:
		frappe.db.set_value("CE Master Scope of Work", sow, 'cesow', cso)    


@frappe.whitelist()
def update_po(po, name):
	frappe.db.set_value("Sales Order", name, "po_no", po)
	sales_invoice = frappe.db.get_all('Sales Invoice', filters = {'sales_order': name,'docstatus':('!=',2)}, fields = ["name"])
	delivery_note = frappe.db.get_all('Delivery Note Item', filters = {'against_sales_order': name}, fields = ['parent'])
	delivery_note_wip = frappe.db.get_all("Delivery Note WIP", filters = {'sales_order': name,'docstatus':('!=',2)}, fields =  ["name"])
	for si in sales_invoice:
		frappe.db.set_value("Sales Invoice", si['name'], "po_no", po)
	for dn in delivery_note:
		del_note=frappe.get_doc("Delivery Note",dn['parent'])
		if del_note.docstatus!=2:
			frappe.db.set_value("Delivery Note", dn['parent'], "po_no", po)
	for wip in delivery_note_wip:
		frappe.db.set_value("Delivery Note WIP", wip['name'], "po_no", po)


@frappe.whitelist()
def update_project_title(po, name):
	frappe.db.set_value("Sales Order", name, "title_of_project", po)
	cost_estimation,quotation,project = frappe.db.get_value("Sales Order",{'name':name},['cost_estimation','quotation','project'])
	
	# quotation = frappe.db.get_value("Sales Order",name,quotation)
	# project = frappe.db.get_value("Sales Order",name,project)
	if cost_estimation:
		frappe.db.set_value("Cost Estimation", cost_estimation, "project_name", po)
	if project:
		frappe.db.set_value("Project", project, "project_name", po)
	if quotation:
		frappe.db.set_value("Quotation", quotation, "title_of_project", po)
	sales_invoice = frappe.db.get_all('Sales Invoice', filters = {'sales_order': name,'docstatus':('!=',2)}, fields = ["name"])
	delivery_note = frappe.db.get_all('Delivery Note Item', filters = {'against_sales_order': name}, fields = ['parent'])
	delivery_note_wip = frappe.db.get_all("Delivery Note WIP", filters = {'sales_order': name,'docstatus':('!=',2)}, fields =  ["name"])
	for si in sales_invoice:
		frappe.db.set_value("Sales Invoice", si['name'], "title_of_project", po)
	for dn in delivery_note:
		del_note=frappe.get_doc("Delivery Note",dn['parent'])
		if del_note.docstatus!=2:
			frappe.db.set_value("Delivery Note", dn['parent'], "title_of_project", po)
	for wip in delivery_note_wip:
		frappe.db.set_value("Delivery Note WIP", wip['name'], "title_of_project", po)

@frappe.whitelist()
def update_mr_qty_in_so(doc, method):
	for i in doc.items:
		if i.sales_order:
			sales_order = frappe.get_doc("Sales Order", i.sales_order)
			for so_item in sales_order.items:
				if i.item_code == so_item.item_code:
					so_item.custom_mr_qty += i.qty
			sales_order.flags.ignore_validate_update_after_submit = True
			sales_order.save(ignore_permissions=True)

@frappe.whitelist()
def update_mr_qty_in_so_cancel(doc, method):
	for i in doc.items:
		if i.sales_order:
			sales_order = frappe.get_doc("Sales Order", i.sales_order)
			for so_item in sales_order.items:
				if i.item_code == so_item.item_code:
					so_item.custom_mr_qty -= i.qty
			sales_order.flags.ignore_validate_update_after_submit = True
			sales_order.save(ignore_permissions=True)


@frappe.whitelist()
def other_report(project):
	# Prepare the initial HTML table structure with styles
	table = '''

	<table border=1px solid black width=100% style="margin-top: 20px;">
	<thead>
		<tr>
			<th style="background-color: #abb2b9 !important; color: black; border: 1px solid #808080; padding: 5px; text-align: center;" colspan="6">Other Expenses</th>
		</tr>
		<tr>
			<th style="background-color: #d5d8dc !important;">S.NO</th>
			<th style="background-color: #d5d8dc !important;">Project Name</th>
			<th style="background-color: #d5d8dc !important;">Voucher Name</th>
			<th style="background-color: #d5d8dc !important;">Voucher Type</th>
			<th style="background-color: #d5d8dc !important;">Expense For</th>
			<th style="background-color: #d5d8dc !important;">Amount</th>
		</tr>
	</thead>
		<tbody>'''

	serial_no = 1  # Initialize serial number for table rows
	total_amount = 0  # Initialize total amount accumulator

	# SQL Query to fetch project and journal entry data from the child table (Journal Entry Account)
	other_items = frappe.db.sql("""
		SELECT p.name AS project_name, pb.name AS budget_name, pb.voucher_type, pbi.account, pbi.debit, pbi.credit
		FROM `tabProject` p
		JOIN `tabJournal Entry Account` pbi ON p.name = pbi.project
		JOIN `tabJournal Entry` pb ON pbi.parent = pb.name
		WHERE pbi.project = %s
		  AND pb.voucher_type = 'Journal Entry' AND pb.docstatus = 1
	""", (project,), as_dict=True)

	# Check if no data is returned
	if not other_items:
		# If no data, display a message inside the table
		table += '''
			<tr>
				
			</tr>'''
	else:
		# If data exists, iterate through the fetched data and build the table rows
		for item in other_items:
			# Determine whether to use debit or credit for the amount
			amount = item['debit'] if item['debit'] else item['credit']
			# Accumulate total amount
			total_amount += amount
		

			# Populate table rows with data
			table += f'''
				<tr>
					<td style="border: 1px solid #808080;">{serial_no}</td>
					<td style="border: 1px solid #808080;">{item['project_name']}</td>  <!-- Project Name from Journal Entry Account -->
					<td style="border: 1px solid #808080;">{item['budget_name']}</td>   <!-- Voucher Name -->
					<td style="border: 1px solid #808080;">{item['voucher_type']}</td>  <!-- Voucher Type -->
					<td style="border: 1px solid #808080;">{item['account']}</td>       <!-- Expense For (Account) -->
					<td style="border: 1px solid #808080;">{amount}</td>                <!-- Debit or Credit -->
				</tr>'''
			serial_no += 1  # Increment serial number

		# Add a final row for the overall total
		table += f'''
			<tr>
				<td colspan="5" style="text-align: right; font-weight: bold;border: 1px solid #808080;">Total Amount</td>
				<td style="border: 1px solid #808080;">{total_amount}</td>
			</tr>'''

	# Close the table body and table tags
	table += '''</tbody>
	</table>'''

	# Return the complete HTML table
 
# 	table += '''
# 				<table border=1px solid black width=100% style="margin-top: 20px;">
# 					<tr><th colspan=6 style="text-align: center; background-color: #abb2b9 !important;">Timesheet</th></tr>
# 					<tr>
# 						<th style="background-color: #d5d8dc !important;" >SI.NO</th>
# 						<th style="background-color: #d5d8dc !important;" >Date</th>
# 						<th style="background-color: #d5d8dc !important;" >Project Name</th>
# 						<th style="background-color: #d5d8dc !important;" >Total Hours</th>
# 						<th style="background-color: #d5d8dc !important;" >Cost Per Hour</th>
# 						<th style="background-color: #d5d8dc !important;" >Total Cost</th>
# 	 				</tr> '''
# 	project_pos = frappe.db.sql("""
# 	SELECT
# 		td.billing_hours AS hours,
# 		td.costing_amount AS cost_amount,
# 		td.costing_rate AS cost_rate,
# 		td.project,
# 		td.project_name,
# 		ts.start_date,
# 		ts.employee_name,
# 		ts.employee
# 	FROM
# 		`tabTimesheet` ts
# 	LEFT JOIN
# 		`tabTimesheet Detail` td
# 	ON
# 		ts.name = td.parent
# 	WHERE
# 		ts.docstatus != 2
# 		AND td.project = %s
# 	ORDER BY
# 		td.project, ts.employee_name, ts.start_date
# """, (project,), as_dict=True)
	
# 	current_employee = None
# 	total_hours = 0
# 	cost = 0
# 	total_cost = 0
# 	ind = 1
# 	for record in project_pos:
# 		if current_employee != record.employee:
# 			table += f'''<tr style="background-color: #eeeeee !important;">
# 							<td colspan=6 style="padding-left: 10px;"><b>{record.employee}-{record.employee_name}</b></td>
# 						</tr>'''
# 			current_employee = record.employee
# 			current_date = record.from_date
# 		table += f'''<tr>
# 						<td>{ind}</td>
# 						<td>{format_date(record.start_date)}</td>
# 						<td>{record.project_name}</td>
# 						<td>{record.hours}</td>
# 						<td>{round(record.cost_rate,2)}</td>
# 						<td>{round(record.cost_amount, 2)}</td>
# 					</tr> '''
# 		ind += 1
# 		total_hours += record.hours
# 		cost += round(record.cost_rate,2)
# 		total_cost += round(record.cost_amount, 2)
# 	table += f'''<tr>
#  				<td colspan=3 style="text-align: right;"><b>Total</b></td>
# 				<td>{total_hours}</td>
# 				<td>{cost}</td>
# 				<td>{total_cost}</td>
# 	 			</tr>'''
# 	table += ''' </table> '''
	return table

@frappe.whitelist()
def get_timesheet_project(from_date, to_date, company):
	project_pos = frappe.db.sql("""
		SELECT
			`tabTimesheet Detail`.billing_hours AS hours,
			`tabTimesheet Detail`.costing_amount AS cost_amount,
			`tabTimesheet Detail`.costing_rate AS cost_rate,
			`tabTimesheet Detail`.project,
			`tabTimesheet Detail`.project_name,
			`tabTimesheet`.start_date,
			`tabTimesheet`.employee_name,
			`tabTimesheet`.employee
		FROM
			`tabTimesheet`
		LEFT JOIN
			`tabTimesheet Detail`
		ON
			`tabTimesheet`.name = `tabTimesheet Detail`.parent
		LEFT JOIN
			`tabProject`
		ON
			`tabTimesheet Detail`.project = `tabProject`.name
		WHERE
			`tabTimesheet`.docstatus != 2
			AND `tabProject`.company = %s
							 
			AND `tabTimesheet`.start_date BETWEEN %s AND %s
		ORDER BY
			`tabTimesheet Detail`.project, `tabTimesheet`.employee_name, `tabTimesheet`.start_date
	""", (company, from_date, to_date), as_dict=True)

	data = "<table class='table table-bordered'>"
	data += """<tr>
			<td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Sl.No.</td>
			<td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Date</td>
			<td style='width:35%;text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Project Name</td>
			<td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Hours</td>
			<td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Cost Per Hour</td>
			<td style='text-align:center;border: 1px solid black;background-color:#e35310;color:white;font-weight:bold'>Total Cost</td>
			</tr>"""

	# Variables to hold running totals
	serial_number = 1
	current_project = None
	current_employee = None
	current_project_total_hours = 0
	current_project_total_cost = 0
	current_employee_total_hours = 0
	current_employee_total_cost = 0
	overall_total_hours = 0
	overall_total_cost = 0

	# Loop through the fetched timesheet records
	for index, record in enumerate(project_pos, start=1):
		# Handle employee-wise totals
		if current_employee != record.employee_name:
			if current_employee is not None:
				data += "<tr>"
				data += "<td colspan='3' style='border: 1px solid black;text-align:right;font-weight:bold'>Total</td>"
				data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(current_employee_total_hours, 2))
				data += "<td style='border: 1px solid black;text-align:right'></td>"
				data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(current_employee_total_cost, 2))
				data += "</tr>"
				current_employee_total_hours = 0
				current_employee_total_cost = 0
		# Handle project-wise totals
		if current_project != record.project:
			if current_project is not None:
				data += "<tr>"
				data += "<td colspan='3' style='border: 1px solid black;text-align:right;font-weight:bold'>Project Total</td>"
				data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(current_project_total_hours, 2))
				data += "<td style='border: 1px solid black;text-align:right'></td>"
				data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(current_project_total_cost, 2))
				data += "</tr>"
				current_project_total_hours = 0
				current_project_total_cost = 0

		# New project header
		if current_project != record.project:
			data += "<tr>"
			data += "<td colspan='6' style='background-color:#D5D8DC;border: 1px solid black;font-weight:bold'>{}</td>".format(record.project)
			data += "</tr>"
			current_project = record.project

		# New employee header
		if current_employee != record.employee_name:
			data += "<tr>"
			data += "<td colspan='6' style='background-color:#edf2f7;border: 1px solid black;font-weight:bold'>{}- {}</td>".format(record.employee,record.employee_name)
			data += "</tr>"
			current_employee = record.employee_name
	   


		# Adding timesheet detail row
		data += "<tr>"
		data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(index)
		data += "<td style='border: 1px solid black'>{}</td>".format(format_date(record.start_date))
		data += "<td style='border: 1px solid black'>{}</td>".format(record.project_name)
		data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(record.hours, 2))
		data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(record.cost_rate, 2))
		data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(record.cost_amount, 2))
		data += "</tr>"

		# Update running totals
		current_project_total_hours += round(record.hours, 2)
		current_project_total_cost += round(record.cost_amount, 2)
		current_employee_total_hours += round(record.hours, 2)
		current_employee_total_cost += round(record.cost_amount, 2)
		overall_total_hours += round(record.hours, 2)
		overall_total_cost += round(record.cost_amount, 2)

	# Final employee total
	if current_employee is not None:
		data += "<tr>"
		data += "<td colspan='3' style='border: 1px solid black;text-align:right;font-weight:bold'>Total</td>"
		data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(current_employee_total_hours, 2))
		data += "<td style='border: 1px solid black;text-align:right'></td>"
		data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(current_employee_total_cost, 2))
		data += "</tr>"

	# Final project total
	if current_project is not None:
		data += "<tr>"
		data += "<td colspan='3' style='border: 1px solid black;text-align:right;font-weight:bold'>Project Total</td>"
		data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(current_project_total_hours, 2))
		data += "<td style='border: 1px solid black;text-align:right'></td>"
		data += "<td style='border: 1px solid black;text-align:right'>{}</td>".format(round(current_project_total_cost, 2))
		data += "</tr>"

	# Overall total row
	data += "<tr>"
	data += "<td colspan='3' style='background-color:#D5D8DC;border: 1px solid black;text-align:right;font-weight:bold'>Overall Total</td>"
	data += "<td style='background-color:#D5D8DC;border: 1px solid black;text-align:right;font-weight:bold'>{}</td>".format(round(overall_total_hours, 2))
	data += "<td style='background-color:#D5D8DC;border: 1px solid black;text-align:right'></td>"
	data += "<td style='background-color:#D5D8DC;border: 1px solid black;text-align:right;font-weight:bold'>{}</td>".format(round(overall_total_cost, 2))
	data += "</tr>"

	data += "</table>"
	return data


@frappe.whitelist()
def update_leave_status():
	frappe.db.sql("""update `tabLeave Application` set workflow_state = 'Cancelled' where name = 'ELE/HRA/03/0718' """)


@frappe.whitelist()
def project_project_report(name,company,from_date,to_date):
	# data = ''
	# projects=frappe.get_all("Project",{'company':company},['name'],order_by='creation ASC')
	# for proj in projects:
	# 	frappe.log_error(title='project',message=proj.name)
	# return data
	# name=''
	# company='INTERIOR DIVISION - ELECTRA'
	# from_date='2024-12-01'
	# to_date='2024-12-31'
	if name:
		crea = frappe.db.sql("""select creation from `tabProject` where name ='%s' """ % (name), as_dict=True)[0]
		project_budget, p_name = frappe.db.get_value("Project", {'name': name}, ['budgeting', 'project_name'])
		pb = frappe.get_doc("Project Budget", project_budget)
		data = ''
		grand = frappe.db.get_value("Sales Order", {'project': name, 'docstatus': 1}, 'grand_total')
		customer = frappe.db.get_value("Sales Order", {'project': name, 'docstatus': 1}, 'customer')
		so_id = frappe.db.get_value("Sales Order", {'project': name, 'docstatus': 1}, 'name')
		sales_in = frappe.db.sql(
			""" SELECT SUM(custom_total_invoice_amount) as total FROM `tabSales Invoice` WHERE project = %s and docstatus = 1""",
			(name,),
			as_dict=True)[0]
		if not sales_in["total"]:
			sales_in["total"] = 0
		bal = grand - sales_in["total"]
		revenue = frappe.db.sql(
			""" SELECT SUM(paid_amount) as amt FROM `tabPayment Entry` WHERE project = %s and docstatus=1""",
			(name,),
			as_dict=True)[0]
		if not revenue["amt"]:
			revenue["amt"] = 0
		out = grand - revenue["amt"]
		tot = 0
		dn_list = frappe.db.get_list("Delivery Note", {'project': name})
		if dn_list:
			for d_list in dn_list:
				dn = frappe.get_doc("Delivery Note", d_list.name)
				for d in dn.items:
					w_house = frappe.db.get_value("Warehouse",
												{'company': dn.company, 'default_for_stock_transfer': 1},
												['name'])
					val = frappe.db.get_value("Bin", {"item_code": d.item_code, "warehouse": w_house}, ['valuation_rate']) or 0
					total = val * d.qty
					tot += total
		prof = grand - tot
		gp = (prof / grand) * 100
		gr_pr = sales_in["total"] - project_cost_till_date(name)
		if sales_in["total"]>0:
			per_gp = (gr_pr / sales_in["total"] or 0) * 100
		else:
			per_gp=0
				
		dic = [{"name": "Supply Materials", "table": pb.materials},
			{"name": "Accessories", "table": pb.bolts_accessories},
			{"name": "Subcontract", "table": pb.others},
			{"name": "Design", "table": pb.design},
			{"name": "Finishing Work", "table": pb.finishing_work},
			{"name": "Tools / Equipment / Transport / Others", "table": pb.heavy_equipments},
			{"name": "Finished Goods", "table": pb.finished_goods},
			{"name": "Manpower", "table": pb.manpower},
			{"name": "Installation", "table": pb.installation}
			]
		serial_no = 1
		total_op_cost=0
		total_mat_cost=0
		total_mat_cost+=total_jl_cost(name)
		for j in dic:
			
			if j["table"]:
				if j["name"] == "Manpower":
					amt = 0
					total_amt = 0
					a_amt = 0
					total_a_amt = 0
					v_amt = 0
					total_v_amt = 0
					total_am = 0
					man = frappe.db.sql(
						"""select sum(`tabTimesheet`.total_costing_amount) as amt from `tabTimesheet`
						left join `tabTimesheet Detail` on `tabTimesheet`.name = `tabTimesheet Detail`.parent
						where `tabTimesheet Detail`.project = %s""",
						(name,),
						as_dict=True) or 0
					if man:
						m = man[0]['amt']
					else:
						m = 0
					for i in j["table"]:
						amt += round(i.amount_with_overheads, 2)
					if m:
						total_op_cost+=m
					else:
						total_op_cost+=0
				
				elif j["name"] == "Installation":
					amt = 0
					total_amt = 0
					a_amt = 0
					total_a_amt = 0
					v_amt = 0
					total_v_amt = 0

					for i in j["table"]:
						dn = frappe.db.sql("""
							SELECT 
								SUM(`tabTimesheet Detail`.costing_amount) AS total_cost,
								`tabTimesheet Detail`.task,
								`tabTask`.subject AS task_name
							FROM 
								`tabTimesheet`
							LEFT JOIN 
								`tabTimesheet Detail` ON `tabTimesheet`.name = `tabTimesheet Detail`.parent
							LEFT JOIN 
								`tabTask` ON `tabTimesheet Detail`.task = `tabTask`.name
							WHERE 
								`tabTimesheet Detail`.project = '%s' AND `tabTimesheet`.docstatus = 1 AND `tabTask`.subject = '%s'
							GROUP BY 
								`tabTimesheet Detail`.task
						""" % (name,i.description), as_dict=True)
						if dn:
							completed_qty = frappe.db.get_value("Task",dn[0]['task'],'completed_qty') or 0
							d = completed_qty
							d_amt = round(dn[0]['total_cost'],2)
						else:
							d = 0
							d_amt = 0
							
						amt += round(i.amount_with_overheads, 2)
						a_amt += round(d_amt, 2)
						v_amt += round(i.amount_with_overheads, 2) - round(d_amt, 2)
						total_amt += amt
						total_a_amt += a_amt
						total_v_amt += v_amt
						total_op_cost+=d_amt
				
				else:
					amt = 0
					total_amt = 0
					a_amt = 0
					total_a_amt = 0
					v_amt = 0
					total_v_amt = 0
					total_am = 0
					for i in j["table"]:
						dn = frappe.db.sql(
							"""select `tabDelivery Note Item`.item_code as item_code, `tabDelivery Note Item`.qty as qty, `tabDelivery Note Item`.amount as amount from `tabDelivery Note`
							left join `tabDelivery Note Item` on `tabDelivery Note`.name = `tabDelivery Note Item`.parent
							where `tabDelivery Note Item`.item_code = %s and `tabDelivery Note`.project = %s""",
							(i.item, name),
							as_dict=True) or 0
						dn_wip = frappe.db.sql("""select c.item_code as item, c.qty as qty, c.amount as amount from `tabStock Entry` p
											inner join `tabStock Entry Detail` c on c.parent = p.name
											where c.item_code = '%s' and c.project = '%s' and p.docstatus = 1 and p.custom_reference_type = 'Delivery Note WIP' """ 
											%(i.item, name), as_dict=1) or 0
						
						if dn and dn_wip:
							d = dn[0]['qty'] + dn_wip[0]['qty']
							d_amt = dn[0]['qty'] + dn_wip[0]['qty'] * i.unit_price
						elif dn:
							d = dn[0]['qty']
							d_amt = dn[0]['qty'] * i.unit_price
						elif dn_wip:
							d = dn_wip[0]['qty']
							d_amt = dn_wip[0]['qty'] * i.unit_price
						else:
							d = 0
							d_amt = 0
				
						amt += round(i.amount_with_overheads, 2)
						a_amt += round(d_amt, 2)
						v_amt += round(i.amount_with_overheads, 2) - round(d_amt, 2)
						total_amt += amt
						total_a_amt += a_amt
						total_v_amt += v_amt
						total_mat_cost+=d_amt

		# data += '</table><br>'
		data += '<table border=1px solid black width=100%>'
		data += '<tr style="font-size:9px;background-color:#ABB2B9;text-align:center"><td><b>Customer</b></td><td><b>Order Ref.</b></td><td><b>Project Ref.</b></td><td><b>Order Value</b></td><td><b>Invoice Value till Date</b></td><td><b>Material Cost</b></td><td><b>Operational Cost</b></td><td><b>Cost Till Date</b></td><td><b>Gross Profit</b></td><td><b>Gross Profit %</b></td><td><b>Balance To Invoice</b></td><td><b>Total Collection</b></td><td><b>O/S Receipts</b></td></tr>'
		data += '<tr style="font-size:9px;text-align:right"><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (
			customer,so_id,name,fmt_money(grand, 2), fmt_money(sales_in["total"] or 0, 2), fmt_money(total_mat_cost, 2),fmt_money(total_op_cost, 2),fmt_money(project_cost_till_date(name), 2),
			fmt_money(round(gr_pr, 2), 2), fmt_money(round(per_gp, 2), 2), fmt_money(round(bal, 2), 2),
			fmt_money(round(revenue["amt"], 2) or 0, 2), fmt_money(round(out, 2), 2))
		data += '<tr style="font-size:9px;text-align:right;font-weight:bold"><td colspan=3>Total</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (
			fmt_money(grand, 2), fmt_money(sales_in["total"] or 0, 2), fmt_money(total_mat_cost, 2),fmt_money(total_op_cost, 2),fmt_money(project_cost_till_date(name), 2),
			fmt_money(round(gr_pr, 2), 2), fmt_money(round(per_gp, 2), 2), fmt_money(round(bal, 2), 2),
			fmt_money(round(revenue["amt"], 2) or 0, 2), fmt_money(round(out, 2), 2))
		data += '</table><br>'
		return data
	else:
		total_grand = 0
		total_sales_in = 0
		total_mat_cost1 = 0
		total_op_cost1 = 0
		total_proj_cost = 0
		total_gr_pr = 0
		total_per_gp = 0
		total_bal = 0
		total_revenue = 0
		total_out = 0
		projects=frappe.get_all("Project",{'company':company,'creation':('between',(from_date,to_date))},['name'],order_by='name ASC')
		data = '<table border=1px solid black width=100%>'
		data += '<tr style="font-size:9px;background-color:#ABB2B9;text-align:center"><td><b>Customer</b></td><td><b>Order Ref.</b></td><td><b>Project Ref.</b></td><td><b>Order Value</b></td><td><b>Invoice Value till Date</b></td><td><b>Material Cost</b></td><td><b>Operational Cost</b></td><td><b>Cost Till Date</b></td><td><b>Gross Profit</b></td><td><b>Gross Profit %</b></td><td><b>Balance To Invoice</b></td><td><b>Total Collection</b></td><td><b>O/S Receipts</b></td></tr>'
			
		for proj in projects:
			name=proj.name
			
			project_budget, p_name = frappe.db.get_value("Project", {'name': name}, ['budgeting', 'project_name'])
			
			grand = frappe.db.get_value("Sales Order", {'project': name, 'docstatus': 1}, 'grand_total')
			customer = frappe.db.get_value("Sales Order", {'project': name, 'docstatus': 1}, 'customer')
			so_id = frappe.db.get_value("Sales Order", {'project': name, 'docstatus': 1}, 'name')
			sales_in = frappe.db.sql(
				""" SELECT SUM(custom_total_invoice_amount) as total FROM `tabSales Invoice` WHERE project = %s and docstatus = 1""",
				(name,),
				as_dict=True)[0]
			if not sales_in["total"]:
				sales_in["total"] = 0
			if not grand:
				grand=0
			bal = grand - sales_in["total"]
			revenue = frappe.db.sql(
				""" SELECT SUM(paid_amount) as amt FROM `tabPayment Entry` WHERE project = %s and docstatus=1""",
				(name,),
				as_dict=True)[0]
			if not revenue["amt"]:
				revenue["amt"] = 0
			out = grand - revenue["amt"]
			tot = 0
			dn_list = frappe.db.get_list("Delivery Note", {'project': name})
			if dn_list:
				for d_list in dn_list:
					dn = frappe.get_doc("Delivery Note", d_list.name)
					for d in dn.items:
						w_house = frappe.db.get_value("Warehouse",
													{'company': dn.company, 'default_for_stock_transfer': 1},
													['name'])
						val = frappe.db.get_value("Bin", {"item_code": d.item_code, "warehouse": w_house}, ['valuation_rate']) or 0
						total = val * d.qty
						tot += total
			prof = grand - tot
			if grand>0:
				gp = (prof / grand) * 100
			else:
				gp = 0
			gr_pr = sales_in["total"] - project_cost_till_date(name)
			if sales_in["total"]>0:
				per_gp = (gr_pr / sales_in["total"] or 0) * 100
			else:
				per_gp=0
			
			total_op_cost=0
			total_mat_cost=0
			total_mat_cost+=total_jl_cost(name)
			if project_budget:
				pb = frappe.get_doc("Project Budget", project_budget)		
				dic = [{"name": "Supply Materials", "table": pb.materials},
					{"name": "Accessories", "table": pb.bolts_accessories},
					{"name": "Subcontract", "table": pb.others},
					{"name": "Design", "table": pb.design},
					{"name": "Finishing Work", "table": pb.finishing_work},
					{"name": "Tools / Equipment / Transport / Others", "table": pb.heavy_equipments},
					{"name": "Finished Goods", "table": pb.finished_goods},
					{"name": "Manpower", "table": pb.manpower},
					{"name": "Installation", "table": pb.installation}
					]
				serial_no = 1
				
				for j in dic:
					
					if j["table"]:
						if j["name"] == "Manpower":
							amt = 0
							total_amt = 0
							a_amt = 0
							total_a_amt = 0
							v_amt = 0
							total_v_amt = 0
							total_am = 0
							man = frappe.db.sql(
								"""select sum(`tabTimesheet`.total_costing_amount) as amt from `tabTimesheet`
								left join `tabTimesheet Detail` on `tabTimesheet`.name = `tabTimesheet Detail`.parent
								where `tabTimesheet Detail`.project = %s""",
								(name,),
								as_dict=True) or 0
							if man:
								m = man[0]['amt']
							else:
								m = 0
							for i in j["table"]:
								amt += round(i.amount_with_overheads, 2)
							if m:
								total_op_cost+=m
							else:
								total_op_cost+=0
						
						elif j["name"] == "Installation":
							amt = 0
							total_amt = 0
							a_amt = 0
							total_a_amt = 0
							v_amt = 0
							total_v_amt = 0

							for i in j["table"]:
								dn = frappe.db.sql("""
									SELECT 
										SUM(`tabTimesheet Detail`.costing_amount) AS total_cost,
										`tabTimesheet Detail`.task,
										`tabTask`.subject AS task_name
									FROM 
										`tabTimesheet`
									LEFT JOIN 
										`tabTimesheet Detail` ON `tabTimesheet`.name = `tabTimesheet Detail`.parent
									LEFT JOIN 
										`tabTask` ON `tabTimesheet Detail`.task = `tabTask`.name
									WHERE 
										`tabTimesheet Detail`.project = '%s' AND `tabTimesheet`.docstatus = 1 AND `tabTask`.subject = '%s'
									GROUP BY 
										`tabTimesheet Detail`.task
								""" % (name,i.description), as_dict=True)
								if dn:
									completed_qty = frappe.db.get_value("Task",dn[0]['task'],'completed_qty') or 0
									d = completed_qty
									d_amt = round(dn[0]['total_cost'],2)
								else:
									d = 0
									d_amt = 0
									
								amt += round(i.amount_with_overheads, 2)
								a_amt += round(d_amt, 2)
								v_amt += round(i.amount_with_overheads, 2) - round(d_amt, 2)
								total_amt += amt
								total_a_amt += a_amt
								total_v_amt += v_amt
								total_op_cost+=d_amt
						
						else:
							amt = 0
							total_amt = 0
							a_amt = 0
							total_a_amt = 0
							v_amt = 0
							total_v_amt = 0
							total_am = 0
							for i in j["table"]:
								dn = frappe.db.sql(
									"""select `tabDelivery Note Item`.item_code as item_code, `tabDelivery Note Item`.qty as qty, `tabDelivery Note Item`.amount as amount from `tabDelivery Note`
									left join `tabDelivery Note Item` on `tabDelivery Note`.name = `tabDelivery Note Item`.parent
									where `tabDelivery Note Item`.item_code = %s and `tabDelivery Note`.project = %s""",
									(i.item, name),
									as_dict=True) or 0
								dn_wip = frappe.db.sql("""select c.item_code as item, c.qty as qty, c.amount as amount from `tabStock Entry` p
													inner join `tabStock Entry Detail` c on c.parent = p.name
													where c.item_code = '%s' and c.project = '%s' and p.docstatus = 1 and p.custom_reference_type = 'Delivery Note WIP' """ 
													%(i.item, name), as_dict=1) or 0
								
								if dn and dn_wip:
									d = dn[0]['qty'] + dn_wip[0]['qty']
									d_amt = dn[0]['qty'] + dn_wip[0]['qty'] * i.unit_price
								elif dn:
									d = dn[0]['qty']
									d_amt = dn[0]['qty'] * i.unit_price
								elif dn_wip:
									d = dn_wip[0]['qty']
									d_amt = dn_wip[0]['qty'] * i.unit_price
								else:
									d = 0
									d_amt = 0
						
								amt += round(i.amount_with_overheads, 2)
								a_amt += round(d_amt, 2)
								v_amt += round(i.amount_with_overheads, 2) - round(d_amt, 2)
								total_amt += amt
								total_a_amt += a_amt
								total_v_amt += v_amt
								total_mat_cost+=d_amt


			data += '<tr style="font-size:9px;text-align:right"><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (
				customer,so_id,name,fmt_money(grand, 2), fmt_money(sales_in["total"] or 0, 2), fmt_money(total_mat_cost, 2),fmt_money(total_op_cost, 2),fmt_money(project_cost_till_date(name), 2),
				fmt_money(round(gr_pr, 2), 2), fmt_money(round(per_gp, 2), 2), fmt_money(round(bal, 2), 2),
				fmt_money(round(revenue["amt"], 2) or 0, 2), fmt_money(round(out, 2), 2))
			total_grand += grand
			total_sales_in += sales_in["total"] or 0
			total_mat_cost1 += total_mat_cost
			total_op_cost1 += total_op_cost
			total_proj_cost += project_cost_till_date(name)
			total_gr_pr += gr_pr
			total_bal += bal
			total_revenue += revenue["amt"] or 0
			total_out += out

		# Calculate average gross profit percentage
		if total_sales_in > 0:
			total_per_gp = (total_gr_pr / total_sales_in) * 100
		else:
			total_per_gp = 0
		data += '<tr style="font-size:9px;text-align:right;font-weight:bold"><td>Total</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (
				fmt_money(total_grand, 2), fmt_money(total_sales_in, 2), fmt_money(total_mat_cost1, 2),fmt_money(total_op_cost1, 2),fmt_money(total_proj_cost, 2),
				fmt_money(round(total_gr_pr, 2), 2), fmt_money(round(total_per_gp, 2), 2), fmt_money(round(total_bal, 2), 2),
				fmt_money(round(total_revenue, 2), 2), fmt_money(round(total_out, 2), 2))
		data += '</table><br>'
		return data

@frappe.whitelist()
def actual_vs_budgeted_manuel():
	name='ENG-PRO-2024-00038'
	crea = frappe.db.sql("""select creation from `tabProject` where name ='%s' """ % (name), as_dict=True)[0]
	project_budget, p_name = frappe.db.get_value("Project", {'name': name}, ['budgeting', 'project_name'])
	pb = frappe.get_doc("Project Budget", project_budget)
	data = ''
	# data += '<tr><td><b>Date</b></td><td>%s</td><td><b>Refer No</b></td><td></td></tr>' % (
		# format_date(crea["creation"].date()))
	# data += '<tr><td><b>Project Code</b></td><td>%s</td><td><b>Project</b></td><td>%s</td></tr>' % (name, p_name)
	# data += '<tr><td><b>Client</b></td><td>%s</td><td><b>Order Ref No.</b></td><td>%s</td></tr>' % (
		# pb.lead_customer_name, pb.sales_order)
	# data += '</table><br>'
	grand = frappe.db.get_value("Sales Order", {'project': name, 'docstatus': 1}, 'grand_total')
	sales_in = frappe.db.sql(
		""" SELECT SUM(custom_total_invoice_amount) as total FROM `tabSales Invoice` WHERE project = %s and docstatus = 1""",
		(name,),
		as_dict=True)[0]
	if not sales_in["total"]:
		sales_in["total"] = 0
	bal = grand - sales_in["total"]
	revenue = frappe.db.sql(
		""" SELECT SUM(paid_amount) as amt FROM `tabPayment Entry` WHERE project = %s""",
		(name,),
		as_dict=True)[0]
	if not revenue["amt"]:
		revenue["amt"] = 0
	out = grand - revenue["amt"]
	tot = 0
	dn_list = frappe.db.get_list("Delivery Note", {'project': name})
	if dn_list:
		for d_list in dn_list:
			dn = frappe.get_doc("Delivery Note", d_list.name)
			for d in dn.items:
				w_house = frappe.db.get_value("Warehouse",
											  {'company': dn.company, 'default_for_stock_transfer': 1},
											  ['name'])
				val = frappe.db.get_value("Bin", {"item_code": d.item_code, "warehouse": w_house}, ['valuation_rate']) or 0
				total = val * d.qty
				tot += total
	prof = grand - tot
	gp = (prof / grand) * 100
	gr_pr = sales_in["total"] - project_cost_till_date(name)
	per_gp = (gr_pr / sales_in["total"] or 0) * 100
			

	

	

	data += '<table border=1px solid black width=100%>'
	data += '<tr style="background-color:#ABB2B9"><td></td><td colspan=5></td><td colspan=3 style="text-align:center"><b>Budgeted</b></td><td colspan=2 style="text-align:center"><b>Actual</b></td><td style="text-align:center" colspan=2><b>Variance</b></td><tr>'
	data += '<tr style="background-color:#D5D8DC"><td colspan=1><b>S.NO</b></td><td colspan=1><b>Budget Code</b></td><td colspan=1><b>Part Number</b></td><td colspan=1><b>Description</b></td><td colspan=1><b>Unit</b></td><td colspan=1><b>Qty</b></td><td colspan=1><b>Rate</b></td><td colspan=1><b>Amount</b></td><td colspan=1><b>Qty</b></td><td colspan=1><b>Amount</b></td><td colspan=1><b>Qty</b></td><td colspan=1><b>Amount</b></td><tr>'
	dic_items = []
	dic = [{"name": "Supply Materials", "table": pb.materials},
		   {"name": "Accessories", "table": pb.bolts_accessories},
		   {"name": "Subcontract", "table": pb.others},
		   {"name": "Design", "table": pb.design},
		   {"name": "Finishing Work", "table": pb.finishing_work},
		   {"name": "Tools / Equipment / Transport / Others", "table": pb.heavy_equipments},
		   {"name": "Finished Goods", "table": pb.finished_goods}]
		#    {"name": "Other Expense", "table": other_report}]
	serial_no = 1
	for j in dic:
		if j["table"]:
			data += '<tr><td style="background-color:#EBEDEF;text-align:center" colspan=12><b>%s</b></td><tr>' % (
				j["name"])
			amt = 0
			total_amt = 0
			a_amt = 0
			total_a_amt = 0
			v_amt = 0
			total_v_amt = 0
			for i in j["table"]:
				dn = frappe.db.sql(
					"""select `tabDelivery Note Item`.item_code as item_code, `tabDelivery Note Item`.qty as qty, `tabDelivery Note Item`.amount as amount from `tabDelivery Note`
					left join `tabDelivery Note Item` on `tabDelivery Note`.name = `tabDelivery Note Item`.parent
					where `tabDelivery Note Item`.item_code = %s and `tabDelivery Note`.project = %s""",
					(i.item, name),
					as_dict=True) or 0
				dn_wip = frappe.db.sql("""select c.item_code as item, sum(c.qty) as qty, sum(c.amount) as amount from `tabStock Entry` p
									   inner join `tabStock Entry Detail` c on c.parent = p.name
									   where c.item_code = '%s' and c.project = '%s' and p.docstatus = 1 and p.custom_reference_type = 'Delivery Note WIP' group by 'c.item_code'""" 
									   %(i.item, name), as_dict=1) or 0
				
				if dn and dn_wip:
					
					d = dn[0]['qty'] + dn_wip[0]['qty']
					d_amt = dn[0]['amount'] + dn_wip[0]['amount']
				elif dn:
					d = dn[0]['qty']
					d_amt = dn[0]['amount']
				elif dn_wip:
					d = dn_wip[0]['qty']
					d_amt = dn_wip[0]['amount']
				else:
					d = 0
					d_amt = 0
				print(d)
				# frappe.log_error(title="Budgeted vs Actual Report", message=[i.item, d, d_amt])
				data += '<tr><td>%s</td><td colspan=1>%s</td><td colspan=1>%s</td><td colspan=1>%s</td><td colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><tr>' % (
					serial_no,project_budget, i.item, i.description, i.unit, i.qty, fmt_money(round(i.unit_price, 2), 2),
					fmt_money(round(i.amount, 2), 2), d, fmt_money(round(d_amt, 2), 2),
					i.qty - d, fmt_money(round(i.amount_with_overheads, 2) - round(d_amt, 2), 2))
				serial_no += 1
				amt += round(i.amount_with_overheads, 2)
				a_amt += round(d_amt, 2)
				v_amt += round(i.amount_with_overheads, 2) - round(d_amt, 2)
				total_amt += amt
				total_a_amt += a_amt
				total_v_amt += v_amt
				
			data += '<tr><td></td><td colspan=6></td><td style="text-align:right" colspan=1>%s</td><td style="text-align:right" colspan=1></td><td style="text-align:right" colspan=1>%s</td><td style="text-align:right" colspan=1></td><td style="text-align:right" colspan=1>%s</td><tr>' % (
				fmt_money(round(amt, 2), 2), fmt_money(round(a_amt, 2), 2), fmt_money(round(v_amt, 2), 2))
			
			data += '<br>'
			
	# manpower = [{"name": "Manpower", "table": pb.manpower}]
	# serial_no = 1
	# for j in manpower:
	# 	if j["table"]:
	# 		data += '<tr><td style="background-color:#EBEDEF;text-align:center" colspan=12><b>%s</b></td><tr>' % (
	# 			j["name"])
	# 		for i in j["table"]:
	# 			man = frappe.db.sql(
	# 				"""select sum(`tabTimesheet`.total_costing_amount) as amt from `tabTimesheet`
	# 				left join `tabTimesheet Detail` on `tabTimesheet`.name = `tabTimesheet Detail`.parent
	# 				where `tabTimesheet Detail`.project = %s""",
	# 				(name,),
	# 				as_dict=True) or 0
	# 			if man:
	# 				m = man[0]['amt']
	# 			else:
	# 				m = 0
	# 			data += '<tr><td>%s</td><td colspan=1>%s</td><td colspan=1>%s</td><td colspan=1>%s</td><td colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><tr>' % (
	# 				serial_no,project_budget, i.worker, i.worker, "Nos", i.total_workers,
	# 				fmt_money(round(i.rate_with_overheads, 2), 2),
	# 				fmt_money(round(i.amount_with_overheads, 2), 2), i.total_workers,
	# 				fmt_money((m or 0), 2), i.total_workers,
	# 				fmt_money(((i.amount_with_overheads) - (m or 0)), 2))
	# 			serial_no +=1
	# install = [{"name": "Installation", "table": pb.installation}]
	# serial_no = 1
	# for j in install:
	# 	if j["table"]:
	# 		data += '<tr><td style="background-color:#EBEDEF;text-align:center" colspan=12><b>%s</b></td><tr>' % (
	# 			j["name"])
	# 		amt = 0
	# 		total_amt = 0
	# 		a_amt = 0
	# 		total_a_amt = 0
	# 		v_amt = 0
	# 		total_v_amt = 0

	# 		for i in j["table"]:
	# 			dn = frappe.db.sql("""
	# 				SELECT 
	# 					SUM(`tabTimesheet Detail`.costing_amount) AS total_cost,
	# 					`tabTimesheet Detail`.task,
	# 					`tabTask`.subject AS task_name
	# 				FROM 
	# 					`tabTimesheet`
	# 				LEFT JOIN 
	# 					`tabTimesheet Detail` ON `tabTimesheet`.name = `tabTimesheet Detail`.parent
	# 				LEFT JOIN 
	# 					`tabTask` ON `tabTimesheet Detail`.task = `tabTask`.name
	# 				WHERE 
	# 					`tabTimesheet Detail`.project = '%s' AND `tabTimesheet`.docstatus = 1 AND `tabTask`.subject = '%s'
	# 				GROUP BY 
	# 					`tabTimesheet Detail`.task
	# 			""" % (name,i.description), as_dict=True)
	# 			if dn:
	# 				completed_qty = frappe.db.get_value("Task",dn[0]['task'],'completed_qty') or 0
	# 				d = completed_qty
	# 				d_amt = round(dn[0]['total_cost'],2)
	# 			else:
	# 				d = 0
	# 				d_amt = 0
					
	# 			data += '<tr><td>%s</td><td colspan=1>%s</td><td colspan=1>%s</td><td colspan=1>%s</td><td colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><td  style="text-align:right" colspan=1>%s</td><tr>' % (serial_no,project_budget, i.item, i.description, i.unit, i.qty,fmt_money(round(i.rate_with_overheads, 2), 2),fmt_money(round(i.amount_with_overheads, 2), 2),d,d_amt,i.qty - d, fmt_money(round(i.amount_with_overheads, 2) - round(d_amt, 2), 2))
	# 			serial_no +=1
	# 			amt += round(i.amount_with_overheads, 2)
	# 			a_amt += round(d_amt, 2)
	# 			v_amt += round(i.amount_with_overheads, 2) - round(d_amt, 2)
	# 			total_amt += amt
	# 			total_a_amt += a_amt
	# 			total_v_amt += v_amt
	# 		data += '<tr><td></td><td colspan=6></td><td style="text-align:right" colspan=1>%s</td><td style="text-align:right" colspan=1></td><td style="text-align:right" colspan=1>%s</td><td style="text-align:right" colspan=1></td><td style="text-align:right" colspan=1>%s</td><tr>' % (
	# 			fmt_money(round(amt, 2), 2), fmt_money(round(a_amt, 2), 2), fmt_money(round(v_amt, 2), 2))
	
	
	
	
	return data
#Get the below details and show html format in Project document

@frappe.whitelist()
def update_project_projct_report_details(name,company,from_date,to_date):
	if name:
		crea = frappe.db.sql("""select creation from `tabProject` where name ='%s' """ % (name), as_dict=True)[0]
		project_budget, p_name = frappe.db.get_value("Project", {'name': name}, ['budgeting', 'project_name'])
		pb = frappe.get_doc("Project Budget", project_budget)
		data = ''
		grand = frappe.db.get_value("Sales Order", {'project': name, 'docstatus': 1}, 'grand_total')
		customer = frappe.db.get_value("Sales Order", {'project': name, 'docstatus': 1}, 'customer')
		so_id = frappe.db.get_value("Sales Order", {'project': name, 'docstatus': 1}, 'name')
		sales_in = frappe.db.sql(
			""" SELECT SUM(custom_total_invoice_amount) as total FROM `tabSales Invoice` WHERE project = %s and docstatus = 1""",
			(name,),
			as_dict=True)[0]
		if not sales_in["total"]:
			sales_in["total"] = 0
		bal = grand - sales_in["total"]
		revenue = frappe.db.sql(
			""" SELECT SUM(paid_amount) as amt FROM `tabPayment Entry` WHERE project = %s and docstatus=1""",
			(name,),
			as_dict=True)[0]
		if not revenue["amt"]:
			revenue["amt"] = 0
		out = grand - revenue["amt"]
		tot = 0
		dn_list = frappe.db.get_list("Delivery Note", {'project': name})
		if dn_list:
			for d_list in dn_list:
				dn = frappe.get_doc("Delivery Note", d_list.name)
				for d in dn.items:
					w_house = frappe.db.get_value("Warehouse",
												{'company': dn.company, 'default_for_stock_transfer': 1},
												['name'])
					val = frappe.db.get_value("Bin", {"item_code": d.item_code, "warehouse": w_house}, ['valuation_rate']) or 0
					total = val * d.qty
					tot += total
		prof = grand - tot
		gp = (prof / grand) * 100
		gr_pr = sales_in["total"] - project_cost_till_date(name)
		if sales_in["total"]>0:
			per_gp = (gr_pr / sales_in["total"] or 0) * 100
		else:
			per_gp=0
				
		dic = [{"name": "Supply Materials", "table": pb.materials},
			{"name": "Accessories", "table": pb.bolts_accessories},
			{"name": "Subcontract", "table": pb.others},
			{"name": "Design", "table": pb.design},
			{"name": "Finishing Work", "table": pb.finishing_work},
			{"name": "Tools / Equipment / Transport / Others", "table": pb.heavy_equipments},
			{"name": "Finished Goods", "table": pb.finished_goods},
			{"name": "Manpower", "table": pb.manpower},
			{"name": "Installation", "table": pb.installation}
			]
		serial_no = 1
		total_op_cost=0
		total_mat_cost=0
		total_mat_cost+=total_jl_cost(name)
		for j in dic:
			
			if j["table"]:
				if j["name"] == "Manpower":
					amt = 0
					total_amt = 0
					a_amt = 0
					total_a_amt = 0
					v_amt = 0
					total_v_amt = 0
					total_am = 0
					man = frappe.db.sql(
						"""select sum(`tabTimesheet`.total_costing_amount) as amt from `tabTimesheet`
						left join `tabTimesheet Detail` on `tabTimesheet`.name = `tabTimesheet Detail`.parent
						where `tabTimesheet Detail`.project = %s""",
						(name,),
						as_dict=True) or 0
					if man:
						m = man[0]['amt']
					else:
						m = 0
					for i in j["table"]:
						amt += round(i.amount_with_overheads, 2)
					if m:
						total_op_cost+=m
					else:
						total_op_cost+=0
				
				elif j["name"] == "Installation":
					amt = 0
					total_amt = 0
					a_amt = 0
					total_a_amt = 0
					v_amt = 0
					total_v_amt = 0

					for i in j["table"]:
						dn = frappe.db.sql("""
							SELECT 
								SUM(`tabTimesheet Detail`.costing_amount) AS total_cost,
								`tabTimesheet Detail`.task,
								`tabTask`.subject AS task_name
							FROM 
								`tabTimesheet`
							LEFT JOIN 
								`tabTimesheet Detail` ON `tabTimesheet`.name = `tabTimesheet Detail`.parent
							LEFT JOIN 
								`tabTask` ON `tabTimesheet Detail`.task = `tabTask`.name
							WHERE 
								`tabTimesheet Detail`.project = '%s' AND `tabTimesheet`.docstatus = 1 AND `tabTask`.subject = '%s'
							GROUP BY 
								`tabTimesheet Detail`.task
						""" % (name,i.description), as_dict=True)
						if dn:
							completed_qty = frappe.db.get_value("Task",dn[0]['task'],'completed_qty') or 0
							d = completed_qty
							d_amt = round(dn[0]['total_cost'],2)
						else:
							d = 0
							d_amt = 0
							
						amt += round(i.amount_with_overheads, 2)
						a_amt += round(d_amt, 2)
						v_amt += round(i.amount_with_overheads, 2) - round(d_amt, 2)
						total_amt += amt
						total_a_amt += a_amt
						total_v_amt += v_amt
						total_op_cost+=d_amt
				
				else:
					amt = 0
					total_amt = 0
					a_amt = 0
					total_a_amt = 0
					v_amt = 0
					total_v_amt = 0
					total_am = 0
					for i in j["table"]:
						dn = frappe.db.sql(
							"""select `tabDelivery Note Item`.item_code as item_code, `tabDelivery Note Item`.qty as qty, `tabDelivery Note Item`.amount as amount from `tabDelivery Note`
							left join `tabDelivery Note Item` on `tabDelivery Note`.name = `tabDelivery Note Item`.parent
							where `tabDelivery Note Item`.item_code = %s and `tabDelivery Note`.project = %s""",
							(i.item, name),
							as_dict=True) or 0
						dn_wip = frappe.db.sql("""select c.item_code as item, c.qty as qty, c.amount as amount from `tabStock Entry` p
											inner join `tabStock Entry Detail` c on c.parent = p.name
											where c.item_code = '%s' and c.project = '%s' and p.docstatus = 1 and p.custom_reference_type = 'Delivery Note WIP' """ 
											%(i.item, name), as_dict=1) or 0
						
						if dn and dn_wip:
							d = dn[0]['qty'] + dn_wip[0]['qty']
							d_amt = dn[0]['qty'] + dn_wip[0]['qty'] * i.unit_price
						elif dn:
							d = dn[0]['qty']
							d_amt = dn[0]['qty'] * i.unit_price
						elif dn_wip:
							d = dn_wip[0]['qty']
							d_amt = dn_wip[0]['qty'] * i.unit_price
						else:
							d = 0
							d_amt = 0
				
						amt += round(i.amount_with_overheads, 2)
						a_amt += round(d_amt, 2)
						v_amt += round(i.amount_with_overheads, 2) - round(d_amt, 2)
						total_amt += amt
						total_a_amt += a_amt
						total_v_amt += v_amt
						total_mat_cost+=d_amt

		# data += '</table><br>'
		project_doc=frappe.get_doc("Project",name)
		project_doc.custom_order_value=grand
		project_doc.custom_invoice_value_till_date=sales_in["total"]
		project_doc.custom_material_cost=total_mat_cost
		project_doc.custom_operational_cost=total_op_cost
		project_doc.custom_cost_till_date=project_cost_till_date(name)
		project_doc.custom_gross_profit=gr_pr
		project_doc.custom_gross_profit_per=per_gp
		project_doc.custom_balance_to_invoice=bal
		project_doc.custom_total_collection=revenue["amt"]
		project_doc.custom_os_receipts=out
		project_doc.save(ignore_permissions=True)
	else:
		total_grand = 0
		total_sales_in = 0
		total_mat_cost1 = 0
		total_op_cost1 = 0
		total_proj_cost = 0
		total_gr_pr = 0
		total_per_gp = 0
		total_bal = 0
		total_revenue = 0
		total_out = 0
		projects=frappe.get_all("Project",{'company':company,'creation':('between',(from_date,to_date))},['name'],order_by='name ASC')
		data = '<table border=1px solid black width=100%>'
		data += '<tr style="font-size:9px;background-color:#ABB2B9;text-align:center"><td><b>Customer</b></td><td><b>Order Ref.</b></td><td><b>Project Ref.</b></td><td><b>Order Value</b></td><td><b>Invoice Value till Date</b></td><td><b>Material Cost</b></td><td><b>Operational Cost</b></td><td><b>Cost Till Date</b></td><td><b>Gross Profit</b></td><td><b>Gross Profit %</b></td><td><b>Balance To Invoice</b></td><td><b>Total Collection</b></td><td><b>O/S Receipts</b></td></tr>'
			
		for proj in projects:
			
			name=proj.name
			if frappe.db.exists("Sales Order", {'project': name, 'docstatus': 1,'status':'Closed'}):
				continue
			else:
				project_budget, p_name = frappe.db.get_value("Project", {'name': name}, ['budgeting', 'project_name'])
				
				grand = frappe.db.get_value("Sales Order", {'project': name, 'docstatus': 1}, 'grand_total')
				customer = frappe.db.get_value("Sales Order", {'project': name, 'docstatus': 1}, 'customer')
				so_id = frappe.db.get_value("Sales Order", {'project': name, 'docstatus': 1}, 'name')
				sales_in = frappe.db.sql(
					""" SELECT SUM(custom_total_invoice_amount) as total FROM `tabSales Invoice` WHERE project = %s and docstatus = 1""",
					(name,),
					as_dict=True)[0]
				if not sales_in["total"]:
					sales_in["total"] = 0
				if not grand:
					grand=0
				bal = grand - sales_in["total"]
				revenue = frappe.db.sql(
					""" SELECT SUM(paid_amount) as amt FROM `tabPayment Entry` WHERE project = %s and docstatus=1""",
					(name,),
					as_dict=True)[0]
				if not revenue["amt"]:
					revenue["amt"] = 0
				out = grand - revenue["amt"]
				tot = 0
				dn_list = frappe.db.get_list("Delivery Note", {'project': name})
				if dn_list:
					for d_list in dn_list:
						dn = frappe.get_doc("Delivery Note", d_list.name)
						for d in dn.items:
							w_house = frappe.db.get_value("Warehouse",
														{'company': dn.company, 'default_for_stock_transfer': 1},
														['name'])
							val = frappe.db.get_value("Bin", {"item_code": d.item_code, "warehouse": w_house}, ['valuation_rate']) or 0
							total = val * d.qty
							tot += total
				prof = grand - tot
				if grand>0:
					gp = (prof / grand) * 100
				else:
					gp = 0
				gr_pr = sales_in["total"] - project_cost_till_date(name)
				if sales_in["total"]>0:
					per_gp = (gr_pr / sales_in["total"] or 0) * 100
				else:
					per_gp=0
				
				total_op_cost=0
				total_mat_cost=0
				total_mat_cost+=total_jl_cost(name)
				if project_budget:
					pb = frappe.get_doc("Project Budget", project_budget)		
					dic = [{"name": "Supply Materials", "table": pb.materials},
						{"name": "Accessories", "table": pb.bolts_accessories},
						{"name": "Subcontract", "table": pb.others},
						{"name": "Design", "table": pb.design},
						{"name": "Finishing Work", "table": pb.finishing_work},
						{"name": "Tools / Equipment / Transport / Others", "table": pb.heavy_equipments},
						{"name": "Finished Goods", "table": pb.finished_goods},
						{"name": "Manpower", "table": pb.manpower},
						{"name": "Installation", "table": pb.installation}
						]
					serial_no = 1
					
					for j in dic:
						
						if j["table"]:
							if j["name"] == "Manpower":
								amt = 0
								total_amt = 0
								a_amt = 0
								total_a_amt = 0
								v_amt = 0
								total_v_amt = 0
								total_am = 0
								man = frappe.db.sql(
									"""select sum(`tabTimesheet`.total_costing_amount) as amt from `tabTimesheet`
									left join `tabTimesheet Detail` on `tabTimesheet`.name = `tabTimesheet Detail`.parent
									where `tabTimesheet Detail`.project = %s""",
									(name,),
									as_dict=True) or 0
								if man:
									m = man[0]['amt']
								else:
									m = 0
								for i in j["table"]:
									amt += round(i.amount_with_overheads, 2)
								if m:
									total_op_cost+=m
								else:
									total_op_cost+=0
							
							elif j["name"] == "Installation":
								amt = 0
								total_amt = 0
								a_amt = 0
								total_a_amt = 0
								v_amt = 0
								total_v_amt = 0

								for i in j["table"]:
									dn = frappe.db.sql("""
										SELECT 
											SUM(`tabTimesheet Detail`.costing_amount) AS total_cost,
											`tabTimesheet Detail`.task,
											`tabTask`.subject AS task_name
										FROM 
											`tabTimesheet`
										LEFT JOIN 
											`tabTimesheet Detail` ON `tabTimesheet`.name = `tabTimesheet Detail`.parent
										LEFT JOIN 
											`tabTask` ON `tabTimesheet Detail`.task = `tabTask`.name
										WHERE 
											`tabTimesheet Detail`.project = '%s' AND `tabTimesheet`.docstatus = 1 AND `tabTask`.subject = '%s'
										GROUP BY 
											`tabTimesheet Detail`.task
									""" % (name,i.description), as_dict=True)
									if dn:
										completed_qty = frappe.db.get_value("Task",dn[0]['task'],'completed_qty') or 0
										d = completed_qty
										d_amt = round(dn[0]['total_cost'],2)
									else:
										d = 0
										d_amt = 0
										
									amt += round(i.amount_with_overheads, 2)
									a_amt += round(d_amt, 2)
									v_amt += round(i.amount_with_overheads, 2) - round(d_amt, 2)
									total_amt += amt
									total_a_amt += a_amt
									total_v_amt += v_amt
									total_op_cost+=d_amt
							
							else:
								amt = 0
								total_amt = 0
								a_amt = 0
								total_a_amt = 0
								v_amt = 0
								total_v_amt = 0
								total_am = 0
								for i in j["table"]:
									dn = frappe.db.sql(
										"""select `tabDelivery Note Item`.item_code as item_code, `tabDelivery Note Item`.qty as qty, `tabDelivery Note Item`.amount as amount from `tabDelivery Note`
										left join `tabDelivery Note Item` on `tabDelivery Note`.name = `tabDelivery Note Item`.parent
										where `tabDelivery Note Item`.item_code = %s and `tabDelivery Note`.project = %s""",
										(i.item, name),
										as_dict=True) or 0
									dn_wip = frappe.db.sql("""select c.item_code as item, c.qty as qty, c.amount as amount from `tabStock Entry` p
														inner join `tabStock Entry Detail` c on c.parent = p.name
														where c.item_code = '%s' and c.project = '%s' and p.docstatus = 1 and p.custom_reference_type = 'Delivery Note WIP' """ 
														%(i.item, name), as_dict=1) or 0
									
									if dn and dn_wip:
										d = dn[0]['qty'] + dn_wip[0]['qty']
										d_amt = dn[0]['qty'] + dn_wip[0]['qty'] * i.unit_price
									elif dn:
										d = dn[0]['qty']
										d_amt = dn[0]['qty'] * i.unit_price
									elif dn_wip:
										d = dn_wip[0]['qty']
										d_amt = dn_wip[0]['qty'] * i.unit_price
									else:
										d = 0
										d_amt = 0
							
									amt += round(i.amount_with_overheads, 2)
									a_amt += round(d_amt, 2)
									v_amt += round(i.amount_with_overheads, 2) - round(d_amt, 2)
									total_amt += amt
									total_a_amt += a_amt
									total_v_amt += v_amt
									total_mat_cost+=d_amt

				project_doc=frappe.get_doc("Project",name)
				project_doc.custom_order_value=grand
				project_doc.custom_invoice_value_till_date=sales_in["total"]
				project_doc.custom_material_cost=total_mat_cost
				project_doc.custom_operational_cost=total_op_cost
				project_doc.custom_cost_till_date=project_cost_till_date(name)
				project_doc.custom_gross_profit=gr_pr
				project_doc.custom_gross_profit_per=per_gp
				project_doc.custom_balance_to_invoice=bal
				project_doc.custom_total_collection=revenue["amt"]
				project_doc.custom_os_receipts=out
				project_doc.save(ignore_permissions=True)
				
@frappe.whitelist()
def project_profit_report_new(name,company,from_date,to_date):
	# name='ENG-PRO-2024-00002'
	# company='ENGINEERING DIVISION - ELECTRA'
	# from_date='2024-01-01'
	# to_date='2024-12-31'
	data = ''
	if name:
		pro_info = frappe.db.get_value(
			"Project", {'name': name}, [
				'customer', 'sales_order', 'custom_order_value', 'custom_invoice_value_till_date',
				'custom_balance_to_invoice', 'custom_gross_profit', 'custom_total_collection',
				'custom_material_cost', 'custom_operational_cost', 'custom_cost_till_date',
				'custom_gross_profit_per', 'custom_os_receipts'
			]
		)

		data = '<table border="1" style="border-collapse:collapse; width:100%;">'
		data += '<tr style="font-size:9px; background-color:#ABB2B9; text-align:center;">'
		data += '<td><b>Customer</b></td><td><b>Order Ref.</b></td><td><b>Project Ref.</b></td>'
		data += '<td><b>Order Value</b></td><td><b>Invoice Value till Date</b></td><td><b>Material Cost</b></td>'
		data += '<td><b>Operational Cost</b></td><td><b>Cost Till Date</b></td><td><b>Gross Profit</b></td>'
		data += '<td><b>Gross Profit %</b></td><td><b>Balance To Invoice</b></td>'
		data += '<td><b>Total Collection</b></td><td><b>O/S Receipts</b></td></tr>'

		mr_cost = frappe.db.get_value("Overall Summary",{"parent":name,"item":"Supply Materials"},["actual"])

		data += '<tr style="font-size:9px; text-align:right;">'
		data += '<td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td>'.format(
			pro_info[0] or '',  
			pro_info[1] or '',  
			name or '',         
			fmt_money(pro_info[2], 2) or '0.00',  
			fmt_money(pro_info[3], 2) or '0.00',  
			fmt_money(mr_cost, 2) or '0.00',  
			fmt_money(pro_info[8], 2) or '0.00'   
		)
		data += '<td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
			fmt_money(pro_info[9], 2) or '0.00',  
			fmt_money(pro_info[5], 2) or '0.00',  
			fmt_money(pro_info[10], 2) or '0.00', 
			fmt_money(pro_info[4], 2) or '0.00',  
			fmt_money(pro_info[6], 2) or '0.00',  
			fmt_money(pro_info[11], 2) or '0.00'  
		)

		
		data += '<tr style="font-size:9px; text-align:right; font-weight:bold;">'
		data += '<td colspan="3" style="text-align:center;">Total</td>'
		data += '<td>{}</td><td>{}</td><td>{}</td><td>{}</td>'.format(        
			fmt_money(pro_info[2], 2) or '0.00',  
			fmt_money(pro_info[3], 2) or '0.00',  
			fmt_money(pro_info[7], 2) or '0.00',  
			fmt_money(pro_info[8], 2) or '0.00'   
		)
		data += '<td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
			fmt_money(pro_info[9], 2) or '0.00',  
			fmt_money(pro_info[5], 2) or '0.00',  
			fmt_money(pro_info[10], 2) or '0.00', 
			fmt_money(pro_info[4], 2) or '0.00',  
			fmt_money(pro_info[6], 2) or '0.00',  
			fmt_money(pro_info[11], 2) or '0.00'  
		)
		data += '</table><br>'
		return data

	else:
		total_grand = 0
		total_sales_in = 0
		total_mat_cost1 = 0
		total_op_cost1 = 0
		total_proj_cost = 0
		total_gr_pr = 0
		total_per_gp = 0
		total_bal = 0
		total_revenue = 0
		total_out = 0
		projects=frappe.get_all("Project",{'company':company,'creation':('between',(from_date,to_date))},['name'],order_by='name ASC')
		data = '<table border="1" style="border-collapse:collapse; width:100%;">'
		data += '<tr style="font-size:9px; background-color:#ABB2B9; text-align:center;">'
		data += '<td><b>Customer</b></td><td><b>Order Ref.</b></td><td><b>Project Ref.</b></td>'
		data += '<td><b>Order Value</b></td><td><b>Invoice Value till Date</b></td><td><b>Material Cost</b></td>'
		data += '<td><b>Operational Cost</b></td><td><b>Cost Till Date</b></td><td><b>Gross Profit</b></td>'
		data += '<td><b>Gross Profit %</b></td><td><b>Balance To Invoice</b></td>'
		data += '<td><b>Total Collection</b></td><td><b>O/S Receipts</b></td></tr>'	
		for proj in projects:
			name=proj.name
			pro_info = frappe.db.get_value(
				"Project", {'name': name}, [
					'customer', 'sales_order', 'custom_order_value', 'custom_invoice_value_till_date',
					'custom_balance_to_invoice', 'custom_gross_profit', 'custom_total_collection',
					'custom_material_cost', 'custom_operational_cost', 'custom_cost_till_date',
					'custom_gross_profit_per', 'custom_os_receipts'
				]
			)
			

			data += '<tr style="font-size:9px; text-align:right;">'
			data += '<td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td>'.format(
				pro_info[0] or '',  
				pro_info[1] or '',  
				name or '',         
				fmt_money(pro_info[2], 2) or '0.00',  
				fmt_money(pro_info[3], 2) or '0.00',  
				fmt_money(pro_info[7], 2) or '0.00',  
				fmt_money(pro_info[8], 2) or '0.00'   
			)
			data += '<td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
				fmt_money(pro_info[9], 2) or '0.00',  
				fmt_money(pro_info[5], 2) or '0.00',  
				fmt_money(pro_info[10], 2) or '0.00', 
				fmt_money(pro_info[4], 2) or '0.00',  
				fmt_money(pro_info[6], 2) or '0.00',  
				fmt_money(pro_info[11], 2) or '0.00'  
			)
			total_grand += pro_info[2]
			total_sales_in += pro_info[3] or 0
			total_mat_cost1 += pro_info[7]
			total_op_cost1 += pro_info[8]
			total_proj_cost += pro_info[9]
			total_gr_pr += pro_info[5]
			total_bal += pro_info[4]
			total_revenue += pro_info[6]
			total_out += pro_info[11]

		if total_sales_in > 0:
			total_per_gp = (total_gr_pr / total_sales_in) * 100
		else:
			total_per_gp = 0
		data += '<tr style="font-size:9px;text-align:right;font-weight:bold"><td colspan=3>Total</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (
				fmt_money(total_grand, 2), fmt_money(total_sales_in, 2), fmt_money(total_mat_cost1, 2),fmt_money(total_op_cost1, 2),fmt_money(total_proj_cost, 2),
				fmt_money(round(total_gr_pr, 2), 2), fmt_money(round(total_per_gp, 2), 2), fmt_money(round(total_bal, 2), 2),
				fmt_money(round(total_revenue, 2), 2), fmt_money(round(total_out, 2), 2))
		data += '</table><br>'
		return data

@frappe.whitelist()
def find_dn_qty_for_item(sales_order, item):
	dn_wip_list = frappe.get_all("Delivery Note WIP", {"sales_order": sales_order, "is_return": 0})
	item_qty = 0
	for dn_wip in dn_wip_list:
		stock_entry_exist = frappe.db.exists("Stock Entry", {"stock_entry_type": "Material Transfer", "reference_number": dn_wip.name, "docstatus": 1})
		if stock_entry_exist:
			se = frappe.get_doc("Stock Entry", {"stock_entry_type": "Material Transfer", "reference_number": dn_wip.name, "docstatus": 1})
			for row in se.items:
				if row.item_code == item:
					item_qty += row.qty
	return item_qty

@frappe.whitelist()
def update_project_projct_report_details_rq_job():
	companies = ['MEP DIVISION - ELECTRA','ENGINEERING DIVISION - ELECTRA','INTERIOR DIVISION - ELECTRA']
	# companies = ['INTERIOR DIVISION - ELECTRA']

	count = 0 
	for c in companies:
		company = c
		projects=frappe.get_all("Project",{'company':company,'status':('not in',['Cancelled','Completed'])},['name'],order_by='name ASC')		
		for proj in projects:
			name=proj.name
			if frappe.db.exists("Sales Order", {'project': name, 'docstatus': 1,'status':'Closed'}):
				continue
			else:
				if frappe.db.exists("Sales Order", {'project': name, 'docstatus': 2}):
					continue
				else:
					count+=1
					crea = frappe.db.sql("""select creation from `tabProject` where name ='%s' """ % (name), as_dict=True)[0]
					project_budget, p_name = frappe.db.get_value("Project", {'name': name}, ['budgeting', 'project_name'])
					grand = frappe.db.get_value("Sales Order", {'project': name, 'docstatus': 1}, 'grand_total')
					customer = frappe.db.get_value("Sales Order", {'project': name, 'docstatus': 1}, 'customer')
					so_id = frappe.db.get_value("Sales Order", {'project': name, 'docstatus': 1}, 'name')
					sales_in = frappe.db.sql(
						""" SELECT SUM(custom_total_invoice_amount) as total FROM `tabSales Invoice` WHERE project = %s and docstatus = 1""",
						(name,),
						as_dict=True)[0]
					if not sales_in["total"]:
						sales_in["total"] = 0
					bal = grand - sales_in["total"]
					revenue = frappe.db.sql(
						""" SELECT SUM(paid_amount) as amt FROM `tabPayment Entry` WHERE project = %s and docstatus=1""",
						(name,),
						as_dict=True)[0]
					if not revenue["amt"]:
						revenue["amt"] = 0
					out = grand - revenue["amt"]
					tot = 0
					dn_list = frappe.db.get_list("Delivery Note", {'project': name})
					if dn_list:
						for d_list in dn_list:
							dn = frappe.get_doc("Delivery Note", d_list.name)
							for d in dn.items:
								w_house = frappe.db.get_value("Warehouse",
															{'company': dn.company, 'default_for_stock_transfer': 1},
															['name'])
								val = frappe.db.get_value("Bin", {"item_code": d.item_code, "warehouse": w_house}, ['valuation_rate']) or 0
								total = val * d.qty
								tot += total
					
					gr_pr = sales_in["total"] - project_cost_till_date(name)
					if sales_in["total"]>0:
						per_gp = (gr_pr / sales_in["total"] or 0) * 100
					else:
						per_gp=0
					if project_budget:
						pb = frappe.get_doc("Project Budget", project_budget)		
						dic = [{"name": "Supply Materials", "table": pb.materials},
							{"name": "Accessories", "table": pb.bolts_accessories},
							{"name": "Subcontract", "table": pb.others},
							{"name": "Design", "table": pb.design},
							{"name": "Finishing Work", "table": pb.finishing_work},
							{"name": "Tools / Equipment / Transport / Others", "table": pb.heavy_equipments},
							{"name": "Finished Goods", "table": pb.finished_goods},
							{"name": "Manpower", "table": pb.manpower},
							{"name": "Installation", "table": pb.installation}
							]
						serial_no = 1
						total_op_cost=0
						total_mat_cost=0
						total_mat_cost+=total_jl_cost(name)
						for j in dic:
							
							if j["table"]:
								if j["name"] == "Manpower":
									amt = 0
									total_amt = 0
									a_amt = 0
									total_a_amt = 0
									v_amt = 0
									total_v_amt = 0
									total_am = 0
									man = frappe.db.sql(
										"""select sum(`tabTimesheet`.total_costing_amount) as amt from `tabTimesheet`
										left join `tabTimesheet Detail` on `tabTimesheet`.name = `tabTimesheet Detail`.parent
										where `tabTimesheet Detail`.project = %s""",
										(name,),
										as_dict=True) or 0
									if man:
										m = man[0]['amt']
									else:
										m = 0
									for i in j["table"]:
										amt += round(i.amount_with_overheads, 2)
									if m:
										total_op_cost+=m
									else:
										total_op_cost+=0
								
								elif j["name"] == "Installation":
									amt = 0
									total_amt = 0
									a_amt = 0
									total_a_amt = 0
									v_amt = 0
									total_v_amt = 0

									for i in j["table"]:
										dn = frappe.db.sql("""
											SELECT 
												SUM(`tabTimesheet Detail`.costing_amount) AS total_cost,
												`tabTimesheet Detail`.task,
												`tabTask`.subject AS task_name
											FROM 
												`tabTimesheet`
											LEFT JOIN 
												`tabTimesheet Detail` ON `tabTimesheet`.name = `tabTimesheet Detail`.parent
											LEFT JOIN 
												`tabTask` ON `tabTimesheet Detail`.task = `tabTask`.name
											WHERE 
												`tabTimesheet Detail`.project = '%s' AND `tabTimesheet`.docstatus = 1 AND `tabTask`.subject = '%s'
											GROUP BY 
												`tabTimesheet Detail`.task
										""" % (name,i.description), as_dict=True)
										if dn:
											completed_qty = frappe.db.get_value("Task",dn[0]['task'],'completed_qty') or 0
											d = completed_qty
											d_amt = round(dn[0]['total_cost'],2)
										else:
											d = 0
											d_amt = 0
											
										amt += round(i.amount_with_overheads, 2)
										a_amt += round(d_amt, 2)
										v_amt += round(i.amount_with_overheads, 2) - round(d_amt, 2)
										total_amt += amt
										total_a_amt += a_amt
										total_v_amt += v_amt
										total_op_cost+=d_amt
								
								else:
									amt = 0
									total_amt = 0
									a_amt = 0
									total_a_amt = 0
									v_amt = 0
									total_v_amt = 0
									total_am = 0
									for i in j["table"]:
										dn = frappe.db.sql(
											"""select `tabDelivery Note Item`.item_code as item_code, `tabDelivery Note Item`.qty as qty, `tabDelivery Note Item`.amount as amount from `tabDelivery Note`
											left join `tabDelivery Note Item` on `tabDelivery Note`.name = `tabDelivery Note Item`.parent
											where `tabDelivery Note Item`.item_code = %s and `tabDelivery Note`.project = %s""",
											(i.item, name),
											as_dict=True) or 0
										dn_wip = frappe.db.sql("""select c.item_code as item, c.qty as qty, c.amount as amount from `tabStock Entry` p
															inner join `tabStock Entry Detail` c on c.parent = p.name
															where c.item_code = '%s' and c.project = '%s' and p.docstatus = 1 and p.custom_reference_type = 'Delivery Note WIP' """ 
															%(i.item, name), as_dict=1) or 0
										
										if dn and dn_wip:
											d = dn[0]['qty'] + dn_wip[0]['qty']
											d_amt = dn[0]['qty'] + dn_wip[0]['qty'] * i.unit_price
										elif dn:
											d = dn[0]['qty']
											d_amt = dn[0]['qty'] * i.unit_price
										elif dn_wip:
											d = dn_wip[0]['qty']
											d_amt = dn_wip[0]['qty'] * i.unit_price
										else:
											d = 0
											d_amt = 0
								
										amt += round(i.amount_with_overheads, 2)
										a_amt += round(d_amt, 2)
										v_amt += round(i.amount_with_overheads, 2) - round(d_amt, 2)
										total_amt += amt
										total_a_amt += a_amt
										total_v_amt += v_amt
										total_mat_cost+=d_amt
					print(name)
					project_doc=frappe.get_doc("Project",name)
					project_doc.custom_order_value=grand
					project_doc.custom_invoice_value_till_date=sales_in["total"]
					project_doc.custom_material_cost=total_mat_cost
					project_doc.custom_operational_cost=total_op_cost
					project_doc.custom_cost_till_date=project_cost_till_date(name)
					project_doc.custom_gross_profit=gr_pr
					project_doc.custom_gross_profit_per=per_gp
					project_doc.custom_balance_to_invoice=bal
					project_doc.custom_total_collection=revenue["amt"]
					project_doc.custom_os_receipts=out
					project_doc.save(ignore_permissions=True)
					
	frappe.log_error(title='Project Updation Count',message=count)

@frappe.whitelist()
def update_project_projct_report_details_rq_job_enqueue():
	frappe.enqueue(
		update_project_projct_report_details_rq_job, 
		queue="long",
		timeout=36000,
		is_async=True, 
		now=False, 
		job_name='Project Profit Cost In Project',
		enqueue_after_commit=False,
	)

#Get the below value for print
@frappe.whitelist()
def get_sos_so(doc):
	so = []
	sales_order = frappe.db.sql("""select sales_order as so from `tabPurchase Order Item` where parent = '%s' """ % doc,as_dict=1)
	if sales_order:
		for item in sales_order: 
			if item.so not in so:
				if item.so:
					so.append(item.so)
	if len(so) > 0:
		return so
	else:
		return ''

@frappe.whitelist()
def check_ce_sow(cost_estimation,msow):
	if not frappe.db.exists("CE SOW",{'cost_estimation':cost_estimation,'msow':msow}):
		return "hi"

@frappe.whitelist()
def project_profit_report_new1(name,company,from_date,to_date):

	data = ''
	if name:
		pro_info = frappe.db.get_value(
			"Project", {'name': name}, [
				'customer', 'sales_order', 'custom_order_value', 'custom_invoice_value_till_date',
				'custom_balance_to_invoice', 'custom_gross_profit', 'custom_total_collection',
				'custom_material_cost', 'custom_operational_cost', 'custom_cost_till_date',
				'custom_gross_profit_per', 'custom_os_receipts'
			]
		)
		sales_in = frappe.db.sql(
			""" SELECT SUM(custom_total_invoice_amount - discount_amount) as total FROM `tabSales Invoice` WHERE project = %s and docstatus = 1""",
			(name),
			as_dict=True)[0]
		if not sales_in["total"]:
			sales_in["total"] = 0
		invoiced_value = sales_in["total"]
		data = '<table border="1" style="border-collapse:collapse; width:100%;">'
		data += '<tr style="font-size:9px; background-color:#ABB2B9; text-align:center;">'
		data += '<td><b>Customer</b></td><td><b>Order Ref.</b></td><td><b>Project Ref.</b></td>'
		data += '<td><b>Order Value</b></td><td><b>Invoice Value till Date</b></td><td><b>Material Cost</b></td>'
		data += '<td><b>Operational Cost</b></td><td><b>Cost Till Date</b></td><td><b>Gross Profit</b></td>'
		data += '<td><b>Gross Profit %</b></td><td><b>Balance To Invoice</b></td>'
		data += '<td><b>Total Collection</b></td><td><b>O/S Receipts</b></td></tr>'

		
		data += '<tr style="font-size:9px; text-align:right;">'
		data += '<td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td>'.format(
			pro_info[0] or '',  
			pro_info[1] or '',  
			name or '',         
			fmt_money(pro_info[2], 2) or '0.00',  
			fmt_money(invoiced_value, 2) or '0.00',  
			fmt_money(pro_info[7], 2) or '0.00',  
			fmt_money(pro_info[8], 2) or '0.00'   
		)
		data += '<td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
			fmt_money(pro_info[9], 2) or '0.00',  
			fmt_money(pro_info[5], 2) or '0.00',  
			fmt_money(pro_info[10], 2) or '0.00', 
			fmt_money(pro_info[4], 2) or '0.00',  
			fmt_money(pro_info[6], 2) or '0.00',  
			fmt_money(pro_info[11], 2) or '0.00'  
		)

		
		data += '<tr style="font-size:9px; text-align:right; font-weight:bold;">'
		data += '<td colspan="3" style="text-align:center;">Total</td>'
		data += '<td>{}</td><td>{}</td><td>{}</td><td>{}</td>'.format(        
			fmt_money(pro_info[2], 2) or '0.00',  
			fmt_money(invoiced_value, 2) or '0.00',  
			fmt_money(pro_info[7], 2) or '0.00',  
			fmt_money(pro_info[8], 2) or '0.00'   
		)
		data += '<td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
			fmt_money(pro_info[9], 2) or '0.00',  
			fmt_money(pro_info[5], 2) or '0.00',  
			fmt_money(pro_info[10], 2) or '0.00', 
			fmt_money(pro_info[4], 2) or '0.00',  
			fmt_money(pro_info[6], 2) or '0.00',  
			fmt_money(pro_info[11], 2) or '0.00'  
		)
		data += '</table><br>'
		return data

	else:
		total_grand = 0
		total_sales_in = 0
		total_mat_cost1 = 0
		total_op_cost1 = 0
		total_proj_cost = 0
		total_gr_pr = 0
		total_per_gp = 0
		total_bal = 0
		total_revenue = 0
		total_out = 0
		projects=frappe.get_all("Project",{'company':company,'creation':('between',(from_date,to_date))},['name'],order_by='name ASC')
		data = '<table border="1" style="border-collapse:collapse; width:100%;">'
		data += '<tr style="font-size:9px; background-color:#ABB2B9; text-align:center;">'
		data += '<td><b>Customer</b></td><td><b>Order Ref.</b></td><td><b>Project Ref.</b></td>'
		data += '<td><b>Order Value</b></td><td><b>Invoice Value till Date</b></td><td><b>Material Cost</b></td>'
		data += '<td><b>Operational Cost</b></td><td><b>Cost Till Date</b></td><td><b>Gross Profit</b></td>'
		data += '<td><b>Gross Profit %</b></td><td><b>Balance To Invoice</b></td>'
		data += '<td><b>Total Collection</b></td><td><b>O/S Receipts</b></td></tr>'	
		for proj in projects:
			name=proj.name
			pro_info = frappe.db.get_value(
				"Project", {'name': name}, [
					'customer', 'sales_order', 'custom_order_value', 'custom_invoice_value_till_date',
					'custom_balance_to_invoice', 'custom_gross_profit', 'custom_total_collection',
					'custom_material_cost', 'custom_operational_cost', 'custom_cost_till_date',
					'custom_gross_profit_per', 'custom_os_receipts'
				]
			)
			sales_in = frappe.db.sql(
				""" SELECT SUM(custom_total_invoice_amount - discount_amount) as total FROM `tabSales Invoice` WHERE project = %s and docstatus = 1""",
				(name),
				as_dict=True)[0]
			if not sales_in["total"]:
				sales_in["total"] = 0
			invoiced_value = sales_in["total"]

			data += '<tr style="font-size:9px; text-align:right;">'
			data += '<td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td>'.format(
				pro_info[0] or '',  
				pro_info[1] or '',  
				name or '',         
				fmt_money(pro_info[2], 2) or '0.00',  
				fmt_money(invoiced_value, 2) or '0.00',  
				fmt_money(pro_info[7], 2) or '0.00',  
				fmt_money(pro_info[8], 2) or '0.00'   
			)
			data += '<td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
				fmt_money(pro_info[9], 2) or '0.00',  
				fmt_money(pro_info[5], 2) or '0.00',  
				fmt_money(pro_info[10], 2) or '0.00', 
				fmt_money(pro_info[4], 2) or '0.00',  
				fmt_money(pro_info[6], 2) or '0.00',  
				fmt_money(pro_info[11], 2) or '0.00'  
			)
			total_grand += pro_info[2]
			total_sales_in += invoiced_value or 0
			total_mat_cost1 += pro_info[7]
			total_op_cost1 += pro_info[8]
			total_proj_cost += pro_info[9]
			total_gr_pr += pro_info[5]
			total_bal += pro_info[4]
			total_revenue += pro_info[6]
			total_out += pro_info[11]

		if total_sales_in > 0:
			total_per_gp = (total_gr_pr / total_sales_in) * 100
		else:
			total_per_gp = 0
		data += '<tr style="font-size:9px;text-align:right;font-weight:bold"><td colspan=3>Total</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (
				fmt_money(total_grand, 2), fmt_money(total_sales_in, 2), fmt_money(total_mat_cost1, 2),fmt_money(total_op_cost1, 2),fmt_money(total_proj_cost, 2),
				fmt_money(round(total_gr_pr, 2), 2), fmt_money(round(total_per_gp, 2), 2), fmt_money(round(total_bal, 2), 2),
				fmt_money(round(total_revenue, 2), 2), fmt_money(round(total_out, 2), 2))
		data += '</table><br>'
		return data


@frappe.whitelist()
def update_so_in_project():
	frappe.db.set_value("Project","INT-PRO-2025-00127", "sales_order","INT-SO-2025-00127")
	frappe.db.set_value("Project","ELV-PRO-2025-00078", "sales_order","ELV-SO-2025-00364")
	frappe.db.set_value("Project","ENG-PRO-2025-00031", "sales_order","ENG-SO-2025-00051")
	frappe.db.set_value("Project","ENG-PRO-2025-00085", "sales_order","ENG-SO-2025-00147")

def test_check():
    new_fiscal_year_leave_allocation()

def new_fiscal_year_leave_allocation():
	today = frappe.utils.today()
	year_start_date = frappe.utils.get_first_day(today)
	employees = frappe.db.get_all(
		"Employee", 
		filters={"status": "Active", "employee": "1000","custom_employee_grade": ["in", ["NON STAFF", "STAFF"]]}, 
		fields=["name", "custom_employee_grade", "date_of_joining", "probation_end_date"]
	)
	for employee in employees:
		if employee.date_of_joining >= year_start_date:
			allocation_date = employee.date_of_joining
		else:
			allocation_date = year_start_date
		
		create_leave_allocation(employee.name, allocation_date, "Medical Leave", 15)
  
		if employee.custom_employee_grade == "STAFF":
			if employee.probation_end_date and employee.probation_end_date > year_start_date:
				allocation_date = employee.probation_end_date
			
			if not employee.probation_end_date:
				no_of_days = date_diff(today, employee.date_of_joining)
				if no_of_days >= 90:
					allocation_date = add_days(employee.date_of_joining, 90)
					if allocation_date <= year_start_date:
						allocation_date = year_start_date
					
					create_leave_allocation(employee.name, allocation_date, "Casual Leave", 3)
			else:
				if employee.probation_end_date <= getdate(today):
					create_leave_allocation(employee.name, allocation_date, "Casual Leave", 3)
	
def create_leave_allocation(employee, allocation_date, leave_type, leave_count):
	if allocation_date <= getdate(frappe.utils.today()):
		from_date = allocation_date
		to_date = date(allocation_date.year, 12, 31)
		if not frappe.db.exists(
			"Leave Allocation",
			{
				"employee": employee, "leave_type": leave_type,
				"from_date": ("between", (from_date, to_date)),
				"to_date": ("between", (from_date, to_date)),
				"docstatus": 1
			}
		):
			leave_allocation = frappe.get_doc({
				"doctype": "Leave Allocation",
				"employee": employee,
				"leave_type": leave_type,
				"from_date": from_date,
				"to_date": to_date,
				"new_leaves_allocated": leave_count,
				"total_leaves_allocated": leave_count,
			})
			leave_allocation.insert(ignore_permissions=True)
			leave_allocation.submit()